from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "mpfm_local.db"
OUTPUTS_DIR = ROOT / "data" / "outputs"


def inspect_db() -> None:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    print(f"DB: {DB_PATH}")
    print("\nFILES_IMPORTED B08 daily around month turn")
    for row in cur.execute(
        """
        SELECT filename, content_date, report_start, report_end, excel_month, processed_ok, message
        FROM files_imported
        WHERE unit_code='B08' AND ext='pdf' AND file_type='daily'
          AND content_date BETWEEN '2026-03-30' AND '2026-04-10'
        ORDER BY content_date, filename
        """
    ):
        print(dict(row))

    print("\nCURATED B08 official counts")
    for row in cur.execute(
        """
        SELECT day_ref, row_kind, COUNT(*) AS n
        FROM measurements_curated
        WHERE bank='B08'
          AND day_ref BETWEEN '2026-03-30' AND '2026-04-10'
          AND COALESCE(is_official, 1)=1
        GROUP BY day_ref, row_kind
        ORDER BY day_ref, row_kind
        """
    ):
        print(dict(row))

    conn.close()


def inspect_workbook(month_pt: str, year: int) -> None:
    path = OUTPUTS_DIR / f"MPFM_{month_pt}_{year}.xlsx"
    print(f"\nWORKBOOK {path.name} exists={path.exists()}")
    if not path.exists():
        return

    wb = load_workbook(path, read_only=True, data_only=True)
    if "BASE_UNICA_MES" not in wb.sheetnames:
        print("BASE_UNICA_MES missing")
        return

    ws = wb["BASE_UNICA_MES"]
    header = [str(value or "").strip().lower() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: header.index(name) for name in ["day_ref", "bank", "origin", "entity", "source"] if name in header}
    print("header indexes", index)

    counts: dict[tuple[str, str], int] = {}
    samples: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        bank = row[index["bank"]] if "bank" in index else None
        if bank != "B08":
            continue
        day_ref = str(row[index["day_ref"]]) if "day_ref" in index else ""
        origin = str(row[index["origin"]]) if "origin" in index else ""
        counts[(day_ref, origin)] = counts.get((day_ref, origin), 0) + 1
        if len(samples) < 12:
            samples.append(
                {
                    "day_ref": row[index["day_ref"]] if "day_ref" in index else None,
                    "origin": row[index["origin"]] if "origin" in index else None,
                    "entity": row[index["entity"]] if "entity" in index else None,
                    "source": row[index["source"]] if "source" in index else None,
                }
            )

    print("counts", counts)
    print("samples")
    for sample in samples:
        print(sample)


if __name__ == "__main__":
    inspect_db()
    inspect_workbook("MAR", 2026)
    inspect_workbook("ABR", 2026)