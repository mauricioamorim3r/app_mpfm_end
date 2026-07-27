from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "outputs" / "MPFM_ABR_2026.xlsx"


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    print(WORKBOOK)
    print(wb.sheetnames)
    ws = wb["BASE_UNICA_MES"]
    header = [str(value or "").strip() for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    lookup = {name: idx for idx, name in enumerate(header)}
    print({key: lookup.get(key) for key in ["ProductionDate", "Granularity", "Origin", "Bank", "Entity", "SourceFile"]})

    counts: dict[tuple[str, str, str], int] = defaultdict(int)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[lookup["Bank"]] != "B08":
            continue
        day = str(row[lookup["ProductionDate"]] or "")[:10]
        if day < "2026-04-01" or day > "2026-04-08":
            continue
        granularity = str(row[lookup["Granularity"]] or "")
        origin = str(row[lookup["Origin"]] or "")
        counts[(day, granularity, origin)] += 1

    for key in sorted(counts):
        print(key, counts[key])

    wb.close()


if __name__ == "__main__":
    main()