from __future__ import annotations

import json

from fastapi import HTTPException, Request
from app_config import get_inactive_tag_associados
from db_schema import rebuild_active_view


def register_admin_routes(app, ctx: dict) -> None:
    load_cadastro = ctx["load_cadastro"]
    cadastro_path = ctx["cadastro_path"]
    output_dir = ctx["output_dir"]
    load_prefs = ctx["load_prefs"]
    prefs_path = ctx["prefs_path"]
    all_metric_names = ctx["all_metric_names"]
    db_conn = ctx["db_conn"]

    @app.get("/api/cadastro")
    def api_get_cadastro():
        return load_cadastro()

    @app.post("/api/cadastro")
    async def api_save_cadastro(request: Request):
        body = await request.json()
        cadastro_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), "utf-8")
        return {"ok": True}

    @app.post("/api/cadastro/toggle-ativo")
    async def api_toggle_cadastro_ativo(request: Request):
        body = await request.json()
        section = body.get("section")
        index = body.get("index")
        if section not in ("banks_subsea", "banks_topside") or not isinstance(index, int):
            raise HTTPException(400, "section e index obrigatórios")
        cad = json.loads(cadastro_path.read_text("utf-8"))
        entries = cad.get(section, [])
        if index < 0 or index >= len(entries):
            raise HTTPException(404, "Entrada não encontrada")
        entries[index]["ativo"] = not entries[index].get("ativo", True)
        cadastro_path.write_text(json.dumps(cad, ensure_ascii=False, indent=2), "utf-8")
        conn = db_conn()
        try:
            rebuild_active_view(conn, get_inactive_tag_associados())
        finally:
            conn.close()
        return {"ok": True, "ativo": entries[index]["ativo"]}

    @app.post("/api/cadastro/toggle-anp")
    async def api_toggle_cadastro_anp(request: Request):
        body = await request.json()
        section = body.get("section")
        index = body.get("index")
        if section not in ("banks_subsea", "banks_topside") or not isinstance(index, int):
            raise HTTPException(400, "section e index obrigatórios")
        cad = json.loads(cadastro_path.read_text("utf-8"))
        entries = cad.get(section, [])
        if index < 0 or index >= len(entries):
            raise HTTPException(404, "Entrada não encontrada")
        entries[index]["aprovado_anp"] = not entries[index].get("aprovado_anp", False)
        cadastro_path.write_text(json.dumps(cad, ensure_ascii=False, indent=2), "utf-8")
        return {"ok": True, "aprovado_anp": entries[index]["aprovado_anp"]}

    @app.post("/api/edit/{filename}/{sheet}")
    async def api_edit_cell(filename: str, sheet: str, request: Request):
        path = output_dir / filename
        if not path.exists():
            raise HTTPException(404, "Arquivo não encontrado")
        body = await request.json()
        action = body.get("action", "edit")
        from openpyxl import load_workbook

        workbook = load_workbook(path)
        if sheet not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(404, f'Aba "{sheet}" não encontrada')
        ws = workbook[sheet]
        try:
            if action == "edit":
                row, col = int(body["row"]) + 2, int(body["col"]) + 1
                value = body.get("value", "")
                try:
                    normalized = value.replace(",", ".") if isinstance(value, str) else value
                    value = float(normalized) if isinstance(normalized, str) and normalized.replace(".", "", 1).lstrip("-").isdigit() else value
                except Exception:
                    pass
                # Sanitize against CSV/Excel formula injection (OWASP)
                safe_value = value
                if isinstance(safe_value, str) and safe_value and safe_value[0] in ("=", "+", "-", "@", "\t", "\r"):
                    safe_value = "'" + safe_value
                ws.cell(row, col, safe_value if safe_value != "" else None)
            elif action == "delete_row":
                ws.delete_rows(int(body["row"]) + 2)
            elif action == "insert_row":
                ws.insert_rows(ws.max_row + 1)
            workbook.save(path)
            return {"ok": True}
        except Exception as exc:
            raise HTTPException(400, f"Erro ao editar: {exc}")
        finally:
            workbook.close()

    @app.get("/api/chartdata/{filename}/{sheet}")
    def api_chartdata(filename: str, sheet: str, tag: str = ""):
        path = output_dir / filename
        if not path.exists():
            raise HTTPException(404, "Arquivo não encontrado")
        from openpyxl import load_workbook as lw

        workbook = lw(path, data_only=True)
        if sheet not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(404, f'Aba "{sheet}" não encontrada')
        ws = workbook[sheet]
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        rows_raw = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row)
            tag_idx = headers.index("TAG") if "TAG" in headers else None
            if tag and tag_idx is not None and str(values[tag_idx] or "") != tag:
                continue
            serialized = []
            for value in values:
                if hasattr(value, "isoformat"):
                    serialized.append(value.isoformat())
                elif value is None or (isinstance(value, float) and value != value):
                    serialized.append(None)
                else:
                    serialized.append(value)
            rows_raw.append(serialized)
        workbook.close()
        sheet_type = "hourly" if sheet.startswith("HOURLY") else "daily" if sheet.startswith("DAILY") else "other"
        day_idx = headers.index("Dia ref.") if "Dia ref." in headers else (headers.index("Dia") if "Dia" in headers else None)
        hour_idx = headers.index("Hora") if "Hora" in headers else None
        labels = []
        for idx, row in enumerate(rows_raw):
            if day_idx is not None:
                date_value = row[day_idx]
                if isinstance(date_value, str) and "T" in date_value:
                    parts = date_value.split("T")[0].split("-")
                    date_value = f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else date_value
                if hour_idx is not None and row[hour_idx] is not None:
                    labels.append(f"{date_value} h{int(row[hour_idx]):02d}")
                else:
                    labels.append(str(date_value) if date_value else "?")
            else:
                labels.append(str(idx))
        tags = []
        if "TAG" in headers:
            ti = headers.index("TAG")
            tags = list(dict.fromkeys(row[ti] for row in rows_raw if row[ti]))
        return {"columns": headers, "rows": rows_raw, "labels": labels, "tags": tags, "sheet_type": sheet_type}

    @app.get("/api/user-prefs")
    def api_get_prefs():
        prefs = load_prefs()
        return {"prefs": prefs, "all_metrics": all_metric_names}

    @app.post("/api/user-prefs")
    async def api_save_prefs(request: Request):
        body = await request.json()
        # Read existing file to preserve keys managed by other endpoints (e.g. auto_folder_monitor)
        existing = {}
        if prefs_path.exists():
            try:
                existing = json.loads(prefs_path.read_text("utf-8"))
            except Exception:
                existing = {}
        merged = {**existing, **body}
        prefs_path.write_text(json.dumps(merged, ensure_ascii=False, indent=2), "utf-8")
        return {"ok": True}
