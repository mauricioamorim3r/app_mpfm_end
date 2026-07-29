"""
routes/ai_routes.py
─────────────────────────────────────────────────────────────────────────────
Endpoints de IA para o MPFM App.

GET  /api/ai/status           → lista providers configurados
POST /api/ai/ask              → envia pergunta, retorna resposta em JSON
POST /api/ai/analyze/report   → analisa texto de um relatório MPFM
GET  /api/ai/keys             → retorna quais chaves estão configuradas (sem revelar valores)
POST /api/ai/keys             → salva chaves no .env e atualiza config em memória
"""

from __future__ import annotations

import base64
import ipaddress
import io
import json
import logging
import os
from pathlib import Path
from typing import Any, Dict, Literal, Optional
from datetime import datetime

from fastapi import APIRouter, HTTPException, Request
from pydantic import BaseModel, Field

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/ai", tags=["ai"])


# ─────────────────────────────────────────────────────────────────────────────
# Schemas
# ─────────────────────────────────────────────────────────────────────────────

class ChatMessage(BaseModel):
    role: Literal["user", "assistant"]
    content: str


class AskRequest(BaseModel):
    question: str = Field(..., min_length=1, max_length=24000)
    provider: Optional[Literal["gemini", "kimi"]] = Field(None)
    model: Optional[str] = Field(None)
    system: Optional[str] = Field(None)
    history: list[ChatMessage] = Field(default_factory=list, description="Histórico de turnos anteriores")
    include_app_context: bool = Field(True, description="Injeta resumo MPFM atual no system prompt")
    app_context: dict[str, Any] = Field(default_factory=dict, description="Contexto da tela/filtros enviado pelo frontend")
    max_tokens: int = Field(8192, ge=256, le=32768)
    temperature: float = Field(0.3, ge=0.0, le=1.0)
    conversation_id: Optional[int] = None


class AskResponse(BaseModel):
    content: str
    provider: str
    model: str
    input_tokens: int
    output_tokens: int
    conversation_id: Optional[int] = None
    message_id: Optional[int] = None


def _token_count(value) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


class AIAttachmentPayload(BaseModel):
    name: str = Field(..., max_length=220)
    mime_type: str = Field("application/octet-stream", max_length=160)
    data_base64: str = Field(..., max_length=18_000_000)


class AnalyzeReportRequest(BaseModel):
    report_text: str = Field("", max_length=24000, description="Texto extraído do relatório MPFM")
    provider: Optional[Literal["gemini", "kimi"]] = None
    attachments: list[AIAttachmentPayload] = Field(default_factory=list)
    conversation_id: Optional[int] = None


class CreateAIActionRequest(BaseModel):
    conversation_id: Optional[int] = None
    message_id: Optional[int] = None
    action_type: Literal[
        "record_from_ai",
        "gerar_proposta",
        "abrir_pendencia",
        "revisar_limite_cv",
        "nota_fechamento",
    ] = "record_from_ai"
    target_area: Literal[
        "nota",
        "pvt",
        "alarme",
        "daily_pdf",
        "medicao",
        "painel_operador",
        "fechamento",
        "pendencia",
        "limites_cv",
        "proposta",
        "outro",
    ] = "nota"
    title: str = Field(..., min_length=1, max_length=180)
    summary: str = Field("", max_length=4000)
    source_content: str = Field("", max_length=24000)
    payload: dict[str, Any] = Field(default_factory=dict)


class CreateAINoteRequest(BaseModel):
    conversation_id: Optional[int] = None
    message_id: Optional[int] = None
    title: str = Field(..., min_length=1, max_length=180)
    summary: str = Field("", max_length=4000)
    source_content: str = Field("", max_length=24000)


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/status", summary="Providers de IA configurados")
async def ai_status():
    """Retorna quais providers têm chaves válidas no .env."""
    from services.ai_assistant import providers_status
    status = providers_status()
    return {
        "providers": status,
        "any_ready": any(status.values()),
    }


@router.post("/ask", response_model=AskResponse, summary="Pergunta livre ao modelo")
async def ai_ask(req: AskRequest):
    """Envia uma pergunta ao modelo AI com histórico e contexto da aplicação."""
    from services.ai_assistant import ask_ai

    # Monta system prompt com contexto opcional da aplicação
    system = req.system or ""
    if req.include_app_context and not req.system:
        system = _build_system_with_context(req.question, req.app_context)

    # Converte histórico para format interno
    history = [{"role": m.role, "content": m.content} for m in req.history]

    conversation_id = req.conversation_id or _create_ai_conversation(
        req.question[:90],
        str((req.app_context or {}).get("current_page") or "assistente"),
    )

    try:
        resp = await ask_ai(
            user=req.question,
            system=system,
            history=history,
            provider=req.provider or "gemini",
            model=req.model,
            max_tokens=req.max_tokens,
            temperature=req.temperature,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    except Exception as exc:
        logger.exception("ai_ask error")
        raise HTTPException(status_code=500, detail=f"Erro ao chamar o modelo: {exc}")

    _log_ai_message(
        conversation_id=conversation_id,
        role="user",
        content=req.question,
        metadata={"app_context": req.app_context, "history_count": len(history)},
    )
    assistant_message_id = _log_ai_message(
        conversation_id=conversation_id,
        role="assistant",
        content=resp.content,
        provider=resp.provider,
        model=resp.model,
        input_tokens=_token_count(resp.input_tokens),
        output_tokens=_token_count(resp.output_tokens),
        metadata={"request_type": "ask"},
    )

    return AskResponse(
        content=resp.content,
        provider=resp.provider,
        model=resp.model,
        input_tokens=_token_count(resp.input_tokens),
        output_tokens=_token_count(resp.output_tokens),
        conversation_id=conversation_id,
        message_id=assistant_message_id,
    )


def _ai_db_conn():
    import sqlite3
    import app_config
    conn = sqlite3.connect(str(app_config.DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _json_dumps(value: Any) -> str:
    return json.dumps(value or {}, ensure_ascii=False)


def _json_loads(value: Any, fallback: Any = None) -> Any:
    if value in (None, ""):
        return fallback
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return fallback


def _ensure_methodology_flow_items_table(conn) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS methodology_flow_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER,
            item_type TEXT NOT NULL DEFAULT 'nota',
            scope TEXT DEFAULT 'run',
            item_key TEXT DEFAULT '',
            title TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT 'aberto',
            owner TEXT DEFAULT '',
            due_date TEXT DEFAULT '',
            summary TEXT DEFAULT '',
            payload_json TEXT DEFAULT '{}',
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(run_id) REFERENCES recon_runs(id)
        );
        CREATE INDEX IF NOT EXISTS idx_methodology_flow_items_lookup
            ON methodology_flow_items(active, run_id, item_type, status, item_key);
        """
    )


def _latest_recon_run_id(conn) -> int | None:
    row = conn.execute("SELECT id FROM recon_runs ORDER BY id DESC LIMIT 1").fetchone()
    return int(row["id"]) if row else None


def _flow_context_from_action_payload(payload: dict[str, Any]) -> dict[str, Any]:
    flow_context = payload.get("methodology_flow_context")
    if isinstance(flow_context, dict):
        return flow_context
    app_context = payload.get("app_context") if isinstance(payload.get("app_context"), dict) else {}
    filters = app_context.get("filters") if isinstance(app_context.get("filters"), dict) else {}
    return {
        "run_id": filters.get("flowRunSelect") or None,
        "active_step": filters.get("flowActiveStep") or "",
        "active_hour": filters.get("flowActiveHour") or None,
    }


def _materialize_ai_action_as_flow_item(conn, row, now: str) -> dict[str, Any]:
    _ensure_methodology_flow_items_table(conn)
    action_type = str(row["action_type"] or "record_from_ai")
    target_area = str(row["target_area"] or "nota")
    payload = _json_loads(row["payload_json"], {}) or {}
    flow_context = _flow_context_from_action_payload(payload)
    run_id_raw = flow_context.get("run_id")
    try:
        run_id = int(run_id_raw) if run_id_raw not in (None, "") else None
    except Exception:
        run_id = None
    if not run_id:
        run_id = _latest_recon_run_id(conn)

    item_type_map = {
        "gerar_proposta": "decisao",
        "abrir_pendencia": "pendencia",
        "revisar_limite_cv": "revisao",
        "nota_fechamento": "nota",
        "record_from_ai": "nota",
    }
    status_map = {
        "gerar_proposta": "aberto",
        "abrir_pendencia": "aberto",
        "revisar_limite_cv": "aberto",
        "nota_fechamento": "resolvido",
        "record_from_ai": "aberto",
    }
    flow_payload = {
        "source": "assistente_ia",
        "ai_action_request_id": row["id"],
        "conversation_id": row["conversation_id"],
        "message_id": row["message_id"],
        "action_type": action_type,
        "target_area": target_area,
        "app_context": payload.get("app_context") if isinstance(payload.get("app_context"), dict) else {},
        "methodology_flow_context": flow_context,
        "source_excerpt": str(row["source_content"] or "")[:2000],
    }
    active_step = str(flow_context.get("active_step") or "").strip()
    active_hour = flow_context.get("active_hour")
    if active_step:
        flow_payload["step_id"] = active_step
    if active_hour not in (None, ""):
        flow_payload["hour"] = active_hour

    cur = conn.execute(
        """
        INSERT INTO methodology_flow_items(
            run_id, item_type, scope, item_key, title, status, owner, due_date,
            summary, payload_json, active, created_at, updated_at
        ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            run_id,
            item_type_map.get(action_type, "nota"),
            "etapa" if active_step else "run",
            active_step or target_area,
            str(row["title"] or "Registro IA aprovado")[:180],
            status_map.get(action_type, "aberto"),
            "Assistente IA",
            "",
            str(row["summary"] or row["source_content"] or "")[:4000],
            _json_dumps(flow_payload),
            1,
            now,
            now,
        ),
    )
    flow_item_id = int(cur.lastrowid)
    return {
        "flow_item_id": flow_item_id,
        "run_id": run_id,
        "item_type": item_type_map.get(action_type, "nota"),
        "scope": "etapa" if active_step else "run",
        "item_key": active_step or target_area,
    }


def _get_master_prompt() -> str:
    try:
        import app_config
        return str(getattr(app_config, "AI_MASTER_PROMPT", "") or "").replace("\\n", "\n").strip()
    except Exception:
        return ""


def _compose_base_system() -> str:
    from services.ai_assistant import SYSTEM_MPFM
    master = _get_master_prompt()
    if not master:
        return SYSTEM_MPFM
    return SYSTEM_MPFM + "\n\n=== Prompt master configurado pelo usuário ===\n" + master


def _create_ai_conversation(title: str, source_page: str = "assistente") -> int:
    conn = _ai_db_conn()
    try:
        now = _now()
        cur = conn.execute(
            "INSERT INTO ai_conversations(title, source_page, created_at, updated_at) VALUES(?,?,?,?)",
            (str(title or "Conversa IA")[:180], source_page or "assistente", now, now),
        )
        conn.commit()
        return int(cur.lastrowid)
    finally:
        conn.close()


def _touch_ai_conversation(conversation_id: int) -> None:
    conn = _ai_db_conn()
    try:
        conn.execute("UPDATE ai_conversations SET updated_at=? WHERE id=?", (_now(), conversation_id))
        conn.commit()
    finally:
        conn.close()


def _log_ai_message(
    *,
    conversation_id: int,
    role: str,
    content: str,
    provider: str = "",
    model: str = "",
    input_tokens: int = 0,
    output_tokens: int = 0,
    metadata: dict[str, Any] | None = None,
) -> int:
    conn = _ai_db_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO ai_messages(conversation_id, role, content, provider, model, input_tokens, output_tokens, metadata_json, created_at)
            VALUES(?,?,?,?,?,?,?,?,?)
            """,
            (conversation_id, role, content, provider, model, input_tokens, output_tokens, _json_dumps(metadata), _now()),
        )
        conn.execute("UPDATE ai_conversations SET updated_at=? WHERE id=?", (_now(), conversation_id))
        conn.commit()
        return int(cur.lastrowid)
    finally:
        conn.close()


def _row_dict(row) -> dict[str, Any]:
    return {key: row[key] for key in row.keys()}


def _prepare_ai_attachments(items: list[AIAttachmentPayload]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    prepared: list[dict[str, Any]] = []
    metadata: list[dict[str, Any]] = []
    for item in items[:8]:
        name = item.name.strip() or "anexo"
        mime_type = (item.mime_type or "application/octet-stream").strip()
        try:
            raw = base64.b64decode(item.data_base64, validate=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Anexo inválido ({name}): {exc}")
        meta = {"name": name, "mime_type": mime_type, "size_bytes": len(raw)}
        metadata.append(meta)

        lower_name = name.lower()
        if mime_type.startswith("text/") or lower_name.endswith((".txt", ".csv", ".json", ".md", ".log")):
            text = raw.decode("utf-8", errors="replace")[:24000]
            prepared.append({"name": name, "mime_type": mime_type, "text": text})
        elif lower_name.endswith(".xlsx") or "spreadsheet" in mime_type:
            prepared.append({"name": name, "mime_type": "text/plain", "text": _extract_xlsx_preview(raw, name)})
        elif mime_type.startswith("image/") or mime_type == "application/pdf":
            prepared.append({"name": name, "mime_type": mime_type, "data": raw})
        else:
            prepared.append({
                "name": name,
                "mime_type": "text/plain",
                "text": f"Anexo {name} recebido ({mime_type}, {len(raw)} bytes). O conteúdo binário não foi extraído automaticamente.",
            })
    return prepared, metadata


def _extract_xlsx_preview(raw: bytes, name: str) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        blocks = [f"Arquivo Excel: {name}"]
        for sheet_name in wb.sheetnames[:6]:
            ws = wb[sheet_name]
            blocks.append(f"\n[Aba: {sheet_name}]")
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), max_col=min(ws.max_column, 12), values_only=True):
                values = [str(value) if value is not None else "" for value in row]
                if any(values):
                    blocks.append(" | ".join(values))
        wb.close()
        return "\n".join(blocks)[:24000]
    except Exception as exc:
        return f"Arquivo Excel {name} recebido, mas não foi possível extrair prévia tabular: {exc}"


def _build_system_with_context(question: str = "", app_context: dict[str, Any] | None = None) -> str:
    """Monta system prompt com snapshot atual dos dados principais da aplicação."""
    from services.ai_assistant import RESPONSE_FORMAT_INSTRUCTIONS
    from services.ai_tools import build_tool_context
    base_system = _compose_base_system()
    try:
        import sqlite3
        import app_config
        db_path = str(app_config.DB_PATH)
        ctx_lines = []
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            ctx_lines.append("=== Escopo de consulta disponível neste prompt ===")
            ctx_lines.append(
                "Você pode responder usando o snapshot abaixo da aplicação: medições MPFM/SEP, "
                "alarmes FCS320, prazos, issues de validação, rastreabilidade de arquivos, Painel do Operador/Radar ANP "
                "em tabelas painel_operador_* e ferramentas internas read-only. "
                "A visualização não deve ser restringida automaticamente pelo mês selecionado na tela; consulte todos os períodos disponíveis, "
                "exceto quando o usuário pedir explicitamente um período. Se o usuário pedir detalhe histórico ou granularidade fina, use o contexto das ferramentas internas abaixo; "
                "se ainda faltar filtro ou fonte integrada, peça o filtro mínimo em vez de inventar dados. Qualquer alteração operacional deve virar proposta para aprovação humana "
                "na área do Assistente IA; se aprovada, a autorização vale somente para o item e escopo solicitados."
            )

            # Última data disponível e bancos ativos
            last_date = conn.execute(
                "SELECT MAX(day_ref) as d FROM measurements_curated WHERE row_kind='daily'"
            ).fetchone()
            if not last_date or not last_date["d"]:
                return base_system
            date_ref = last_date["d"]

            # Produção diária por banco (pivot manual)
            rows = conn.execute("""
                SELECT bank,
                    SUM(CASE WHEN metric_name LIKE '%leo%' AND row_kind='daily' AND is_official=1
                             THEN metric_value ELSE 0 END) as oil_t,
                    SUM(CASE WHEN metric_name LIKE 'MPFM corr G%' AND row_kind='daily' AND is_official=1
                             THEN metric_value ELSE 0 END) as gas_t,
                    SUM(CASE WHEN metric_name LIKE '%gua%' AND row_kind='daily' AND is_official=1
                             THEN metric_value ELSE 0 END) as water_t,
                    SUM(CASE WHEN metric_name LIKE '%HC%' AND row_kind='daily' AND is_official=1
                             THEN metric_value ELSE 0 END) as hc_t
                FROM measurements_curated
                WHERE day_ref = ? AND row_kind = 'daily'
                GROUP BY bank ORDER BY bank
            """, (date_ref,)).fetchall()
            ctx_lines.append(f"=== Dados MPFM mais recentes: {date_ref} ===")
            for r in rows:
                ctx_lines.append(
                    f"Banco {r['bank']}: óleo≈{r['oil_t']:.1f}t, gás≈{r['gas_t']:.1f}t, "
                    f"água≈{r['water_t']:.1f}t, HC≈{r['hc_t']:.1f}t"
                )
            # Sumário do mês
            month_row = conn.execute("""
                SELECT strftime('%Y-%m', day_ref) as month,
                       COUNT(DISTINCT bank) as banks,
                       COUNT(DISTINCT day_ref) as days
                FROM measurements_curated
                WHERE row_kind='daily' AND is_official=1
                  AND strftime('%Y-%m', day_ref) = strftime('%Y-%m', ?, 'localtime')
                GROUP BY month
            """, (date_ref,)).fetchone()
            if month_row:
                ctx_lines.append(
                    f"Período {month_row['month']}: {month_row['banks']} bancos, "
                    f"{month_row['days']} dias com dados processados."
                )
            total = conn.execute("SELECT COUNT(*) as n FROM measurements_curated").fetchone()
            ctx_lines.append(f"Base de dados: {total['n']:,} medições históricas disponíveis.")

            # Alarmes do módulo e carga PDF FCS320 consolidada
            alarm_total = conn.execute(
                "SELECT COUNT(*) AS n FROM alarm_records WHERE active=1"
            ).fetchone()["n"]
            if alarm_total:
                ctx_lines.append("\n=== Módulo de alarmes da aplicação ===")
                ctx_lines.append(f"Total ativo no módulo de alarmes: {alarm_total:,} registros.")
                source_rows = conn.execute("""
                    SELECT source_kind, COUNT(*) AS total,
                           SUM(CASE WHEN status_code NOT IN ('closed','cancelled') THEN 1 ELSE 0 END) AS open_n
                    FROM alarm_records
                    WHERE active=1
                    GROUP BY source_kind
                    ORDER BY total DESC
                """).fetchall()
                for row in source_rows:
                    ctx_lines.append(
                        f"Origem {row['source_kind'] or 'n/d'}: {row['total']} ativos, {row['open_n']} em aberto."
                    )

                pdf_summary = conn.execute("""
                    SELECT COUNT(*) AS total,
                           SUM(CASE WHEN record_type='event' THEN 1 ELSE 0 END) AS events,
                           SUM(CASE WHEN record_type='incident' THEN 1 ELSE 0 END) AS incidents,
                           SUM(CASE WHEN status_code NOT IN ('closed','cancelled') THEN 1 ELSE 0 END) AS open_n,
                           SUM(CASE WHEN status_code='closed' THEN 1 ELSE 0 END) AS closed_n,
                           MIN(production_date) AS min_day,
                           MAX(production_date) AS max_day
                    FROM alarm_records
                    WHERE active=1 AND source_kind='pdf'
                """).fetchone()
                if pdf_summary and pdf_summary["total"]:
                    ctx_lines.append("\n=== Carga PDF FCS320 consolidada ===")
                    ctx_lines.append(
                        f"PDF FCS320 ativo: {pdf_summary['total']} registros, "
                        f"{pdf_summary['events']} eventos, {pdf_summary['incidents']} incidentes, "
                        f"{pdf_summary['closed_n']} fechados, {pdf_summary['open_n']} em aberto. "
                        f"Período de produção: {pdf_summary['min_day']} a {pdf_summary['max_day']}."
                    )

                alarm_status_rows = conn.execute("""
                    SELECT record_type, status_code, severity_code, priority_code, COUNT(*) AS n
                    FROM alarm_records
                    WHERE active=1 AND source_kind='pdf'
                    GROUP BY record_type, status_code, severity_code, priority_code
                    ORDER BY record_type, status_code, severity_code, priority_code
                """).fetchall()
                for row in alarm_status_rows[:24]:
                    ctx_lines.append(
                        f"{row['record_type'] or 'tipo?'} | status={row['status_code'] or 'n/d'} | "
                        f"sev={row['severity_code'] or 'n/d'} | prioridade={row['priority_code'] or 'n/d'}: {row['n']}"
                    )

                open_alarms = conn.execute("""
                    SELECT id, record_type, production_date, event_at, measurement_point, tag, title,
                           severity_code, priority_code, status_code, occurrence_count
                    FROM alarm_records
                    WHERE active=1 AND source_kind='pdf' AND status_code NOT IN ('closed','cancelled')
                    ORDER BY production_date DESC, event_at DESC, severity_code, priority_code
                    LIMIT 16
                """).fetchall()
                if open_alarms:
                    ctx_lines.append("Alarmes PDF FCS320 em aberto / tratamento pendente:")
                    for row in open_alarms:
                        ctx_lines.append(
                            f"#{row['id']} {row['record_type']} {row['production_date']} {row['event_at']} "
                            f"{row['measurement_point'] or row['tag'] or 'ponto n/d'} - {row['title']} "
                            f"[{row['severity_code']}/{row['priority_code']}/{row['status_code']}] "
                            f"ocorrencias={row['occurrence_count'] or 0}"
                        )

                recurrent = conn.execute("""
                    SELECT COALESCE(measurement_point, tag, '') AS point, title, record_type, COUNT(*) AS n,
                           SUM(CASE WHEN status_code NOT IN ('closed','cancelled') THEN 1 ELSE 0 END) AS open_n
                    FROM alarm_records
                          WHERE active=1 AND source_kind='pdf'
                    GROUP BY COALESCE(measurement_point, tag, ''), title, record_type
                    HAVING COUNT(*) > 1
                    ORDER BY n DESC
                    LIMIT 10
                """).fetchall()
                if recurrent:
                    ctx_lines.append("Alarmes mais recorrentes:")
                    for row in recurrent:
                        ctx_lines.append(
                            f"{row['point'] or 'ponto n/d'} | {row['record_type']} | {row['title']}: "
                            f"{row['n']} ocorrencias, {row['open_n']} abertas"
                        )

            # Prazos operacionais
            try:
                deadline_total = conn.execute(
                    "SELECT COUNT(*) AS n FROM deadline_items WHERE is_active=1"
                ).fetchone()["n"]
            except sqlite3.OperationalError:
                deadline_total = 0
            if deadline_total:
                ctx_lines.append("\n=== Prazos ativos ===")
                ctx_lines.append(f"Total de prazos ativos: {deadline_total}.")
                deadline_rows = conn.execute("""
                    SELECT id, subject, category, start_date, due_date, periodicity, periodicity_days, notes
                    FROM deadline_items
                    WHERE is_active=1
                    ORDER BY CASE WHEN COALESCE(due_date,'')='' THEN '9999-12-31' ELSE due_date END, id
                    LIMIT 16
                """).fetchall()
                for row in deadline_rows:
                    ctx_lines.append(
                        f"#{row['id']} {row['subject']} | categoria={row['category'] or 'n/d'} | "
                        f"inicio={row['start_date'] or 'n/d'} | vencimento={row['due_date'] or 'n/d'} | "
                        f"periodicidade={row['periodicity'] or 'custom'} {row['periodicity_days'] or ''} | "
                        f"notas={row['notes'] or ''}"
                    )

            # Pendências de validação
            issue_rows = conn.execute("""
                SELECT issue_type, severity, COUNT(*) AS n
                FROM validation_issues
                GROUP BY issue_type, severity
                ORDER BY n DESC
                LIMIT 10
            """).fetchall()
            if issue_rows:
                ctx_lines.append("\n=== Issues de validação ===")
                for row in issue_rows:
                    ctx_lines.append(f"{row['issue_type']} | {row['severity']}: {row['n']}")
        tool_context = build_tool_context(_ai_db_conn, question, app_context or {})
        if ctx_lines:
            return base_system + "\n\n" + RESPONSE_FORMAT_INSTRUCTIONS + "\n\n" + "\n".join(ctx_lines) + "\n\n" + tool_context
    except Exception as e:
        logger.warning("Falha ao buscar contexto MPFM: %s", e)
    from services.ai_assistant import RESPONSE_FORMAT_INSTRUCTIONS
    return _compose_base_system() + "\n\n" + RESPONSE_FORMAT_INSTRUCTIONS


# ─────────────────────────────────────────────────────────────────────────────
# Gestão de chaves de API
# ─────────────────────────────────────────────────────────────────────────────

_ENV_KEY_MAP: dict[str, str] = {
    "default_provider":  "AI_DEFAULT_PROVIDER",
    "gemini_key":        "GEMINI_API_KEY",
    "gemini_model":      "GEMINI_MODEL",
    "kimi_key":          "MOONSHOT_API_KEY",
    "kimi_model":        "MOONSHOT_MODEL",
    "kimi_base_url":     "MOONSHOT_BASE_URL",
    "master_prompt":     "AI_MASTER_PROMPT",
}


def _env_path() -> Path:
    import app_config
    return Path(app_config.BASE_DIR) / ".env"


def _read_env_file() -> dict[str, str]:
    p = _env_path()
    result: dict[str, str] = {}
    if not p.exists():
        return result
    for line in p.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or "=" not in stripped:
            continue
        k, _, v = stripped.partition("=")
        result[k.strip()] = v.strip()
    return result


def _write_env_file(data: dict[str, str]) -> None:
    p = _env_path()
    lines = ["# MPFM Manager — Environment Variables", ""]
    for k, v in data.items():
        lines.append(f"{k}={v}")
    p.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _is_loopback_host(host: str | None) -> bool:
    if not host:
        return False
    if host.lower() == "localhost":
        return True
    try:
        return ipaddress.ip_address(host).is_loopback
    except ValueError:
        return False


def _forwarded_hosts(request: Request) -> list[str]:
    hosts: list[str] = []
    forwarded_for = request.headers.get("x-forwarded-for", "")
    for host in forwarded_for.split(","):
        host = host.strip()
        if host:
            hosts.append(host)
    real_ip = request.headers.get("x-real-ip", "").strip()
    if real_ip:
        hosts.append(real_ip)
    return hosts


def _require_keys_access(request: Request) -> None:
    import app_config

    if app_config.AUTH_ENABLED:
        if bool(getattr(getattr(request, "state", None), "mpfm_authenticated", False)):
            return
        raise HTTPException(status_code=401, detail="Autenticação requerida.")

    client_host = getattr(getattr(request, "client", None), "host", None)
    if _is_loopback_host(client_host) and all(_is_loopback_host(host) for host in _forwarded_hosts(request)):
        return
    raise HTTPException(
        status_code=403,
        detail="Configuração de chaves de IA requer acesso local ou autenticação habilitada.",
    )


@router.get("/keys", summary="Status das chaves configuradas (sem expor valores)")
def ai_keys_status(request: Request):
    """Retorna quais chaves estão presentes no .env (não expõe os valores)."""
    _require_keys_access(request)
    env = _read_env_file()

    def _ok(var: str) -> bool:
        v = env.get(var, "")
        return bool(v and "COLE_" not in v and len(v) > 4)

    return {
        "default_provider": env.get("AI_DEFAULT_PROVIDER", "gemini"),
        "gemini":    {"key": _ok("GEMINI_API_KEY"),    "model": env.get("GEMINI_MODEL", "gemini-2.5-flash")},
        "kimi":      {"key": _ok("MOONSHOT_API_KEY"),  "model": env.get("MOONSHOT_MODEL", "moonshot-v1-8k"), "base_url": env.get("MOONSHOT_BASE_URL", "https://api.moonshot.ai/v1")},
        "master_prompt": env.get("AI_MASTER_PROMPT", "").replace("\\n", "\n"),
    }


class KeysPayload(BaseModel):
    default_provider: Optional[str] = None
    gemini_key:       Optional[str] = None
    gemini_model:     Optional[str] = None
    kimi_key:         Optional[str] = None
    kimi_model:       Optional[str] = None
    kimi_base_url:    Optional[str] = None
    master_prompt:    Optional[str] = None


@router.post("/keys", summary="Salva chaves de API no .env e aplica imediatamente")
def ai_save_keys(payload: KeysPayload, request: Request):
    """
    Recebe chaves e configurações, grava no .env e atualiza os valores em memória
    para que as mudanças tenham efeito imediato sem reiniciar o servidor.
    """
    _require_keys_access(request)
    import app_config

    env = _read_env_file()
    updated: list[str] = []
    for field_name, env_var in _ENV_KEY_MAP.items():
        value = getattr(payload, field_name, None)
        if value is None:
            continue
        value = value.strip()
        if not value:
            env.pop(env_var, None)
            os.environ.pop(env_var, None)
            setattr(app_config, env_var, "")
            updated.append(env_var)
        else:
            env_value = value.replace("\r\n", "\n").replace("\n", "\\n") if field_name == "master_prompt" else value
            env[env_var] = env_value
            os.environ[env_var] = env_value
            setattr(app_config, env_var, value)
            updated.append(env_var)

    _write_env_file(env)
    logger.info("AI keys updated: %s", updated)
    return {"saved": updated, "status": "ok"}


@router.get("/logs", summary="Histórico recente de mensagens da IA")
def ai_logs(limit: int = 40):
    conn = _ai_db_conn()
    try:
        rows = conn.execute(
            """
            SELECT m.id, m.conversation_id, c.title AS conversation_title, m.role, m.content,
                   m.provider, m.model, m.input_tokens, m.output_tokens, m.created_at
            FROM ai_messages m
            LEFT JOIN ai_conversations c ON c.id=m.conversation_id
            ORDER BY m.id DESC
            LIMIT ?
            """,
            (max(1, min(int(limit or 40), 200)),),
        ).fetchall()
        return {"items": [_row_dict(row) for row in rows]}
    finally:
        conn.close()


@router.get("/notes", summary="Notas técnicas salvas a partir da IA")
def ai_notes(limit: int = 30):
    conn = _ai_db_conn()
    try:
        rows = conn.execute(
            """
            SELECT id, context_key, decision_summary, rationale, impact_notes, created_at
            FROM project_decisions
            WHERE context_key LIKE 'ai_note%'
            ORDER BY id DESC
            LIMIT ?
            """,
            (max(1, min(int(limit or 30), 100)),),
        ).fetchall()
        return {"items": [_row_dict(row) for row in rows]}
    finally:
        conn.close()


@router.post("/notes", summary="Salva última resposta da IA como nota técnica")
def ai_create_note(payload: CreateAINoteRequest):
    source_content = payload.source_content or ""
    if not source_content and payload.message_id:
        conn_lookup = _ai_db_conn()
        try:
            row = conn_lookup.execute("SELECT content FROM ai_messages WHERE id=?", (payload.message_id,)).fetchone()
            source_content = row["content"] if row else ""
        finally:
            conn_lookup.close()
    if not source_content.strip():
        raise HTTPException(status_code=400, detail="Não há conteúdo da IA para salvar como nota.")
    context_key = f"ai_note:{payload.conversation_id or 0}:{payload.message_id or 0}"
    impact_parts = []
    if payload.summary.strip():
        impact_parts.append(payload.summary.strip())
    impact_parts.append("Nota técnica salva diretamente pelo usuário a partir do Assistente IA; não altera dados operacionais.")
    conn = _ai_db_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO project_decisions(context_key, decision_summary, rationale, impact_notes, created_at)
            VALUES(?,?,?,?,?)
            """,
            (
                context_key,
                payload.title.strip(),
                source_content.strip(),
                "\n\n".join(impact_parts),
                _now(),
            ),
        )
        conn.commit()
        note_id = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM project_decisions WHERE id=?", (note_id,)).fetchone()
        return {"ok": True, "item": _row_dict(row)}
    finally:
        conn.close()


@router.get("/actions", summary="Propostas da IA pendentes/aprovadas")
def ai_actions(status: str = "pending", limit: int = 50):
    conn = _ai_db_conn()
    try:
        if status == "all":
            rows = conn.execute(
                "SELECT * FROM ai_action_requests ORDER BY id DESC LIMIT ?",
                (max(1, min(int(limit or 50), 200)),),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT * FROM ai_action_requests WHERE status=? ORDER BY id DESC LIMIT ?",
                (status, max(1, min(int(limit or 50), 200))),
            ).fetchall()
        return {"items": [_row_dict(row) for row in rows]}
    finally:
        conn.close()


@router.post("/actions", summary="Cria proposta pendente a partir de resposta da IA")
def ai_create_action(payload: CreateAIActionRequest):
    source_content = payload.source_content or ""
    if not source_content and payload.message_id:
        conn_lookup = _ai_db_conn()
        try:
            row = conn_lookup.execute("SELECT content FROM ai_messages WHERE id=?", (payload.message_id,)).fetchone()
            source_content = row["content"] if row else ""
        finally:
            conn_lookup.close()
    conn = _ai_db_conn()
    try:
        cur = conn.execute(
            """
            INSERT INTO ai_action_requests(
                conversation_id, message_id, action_type, target_area, title, summary,
                source_content, payload_json, status, requested_by, created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                payload.conversation_id,
                payload.message_id,
                payload.action_type,
                payload.target_area,
                payload.title.strip(),
                payload.summary.strip(),
                source_content,
                _json_dumps(payload.payload),
                "pending",
                "user",
                _now(),
            ),
        )
        conn.commit()
        request_id = int(cur.lastrowid)
        row = conn.execute("SELECT * FROM ai_action_requests WHERE id=?", (request_id,)).fetchone()
        return {"ok": True, "item": _row_dict(row)}
    finally:
        conn.close()


@router.post("/actions/{request_id}/approve", summary="Aprova e registra uma proposta da IA")
def ai_approve_action(request_id: int):
    conn = _ai_db_conn()
    try:
        row = conn.execute("SELECT * FROM ai_action_requests WHERE id=?", (request_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Proposta IA não encontrada.")
        if row["status"] != "pending":
            raise HTTPException(status_code=400, detail="Somente propostas pendentes podem ser aprovadas.")
        now = _now()
        action_type = str(row["action_type"] or "record_from_ai")
        target_area = str(row["target_area"] or "nota")
        action_labels = {
            "gerar_proposta": "Proposta técnica aprovada para tratamento rastreável.",
            "abrir_pendencia": "Pendência aprovada para acompanhamento rastreável.",
            "revisar_limite_cv": "Revisão de limite/PAM ou configuração CV aprovada para acompanhamento.",
            "nota_fechamento": "Nota de fechamento aprovada para registro técnico.",
            "record_from_ai": "Registro IA aprovado para auditoria.",
        }
        result = {
            "registered": True,
            "action_type": action_type,
            "target_area": target_area,
            "note": action_labels.get(action_type, action_labels["record_from_ai"])
            + " Alterações operacionais diretas exigem rotina específica do módulo correspondente.",
        }
        flow_result = _materialize_ai_action_as_flow_item(conn, row, now)
        result["methodology_flow_item"] = flow_result
        conn.execute(
            """
            UPDATE ai_action_requests
            SET status='approved', approved_by='user', approved_at=?, applied_at=?, result_json=?
            WHERE id=?
            """,
            (now, now, _json_dumps(result), request_id),
        )
        conn.commit()
        updated = conn.execute("SELECT * FROM ai_action_requests WHERE id=?", (request_id,)).fetchone()
        return {"ok": True, "item": _row_dict(updated)}
    finally:
        conn.close()


@router.post("/actions/{request_id}/archive", summary="Arquiva proposta da IA")
def ai_archive_action(request_id: int):
    conn = _ai_db_conn()
    try:
        conn.execute("UPDATE ai_action_requests SET status='archived' WHERE id=? AND status='pending'", (request_id,))
        conn.commit()
        row = conn.execute("SELECT * FROM ai_action_requests WHERE id=?", (request_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Proposta IA não encontrada.")
        return {"ok": True, "item": _row_dict(row)}
    finally:
        conn.close()


@router.post("/analyze/report", response_model=AskResponse, summary="Analisa texto de relatório MPFM")
async def ai_analyze_report(req: AnalyzeReportRequest):
    """
    Recebe o texto extraído de um relatório MPFM e retorna uma análise
    automática com pontos de atenção, anomalias e sugestões.
    """
    from services.ai_assistant import ask_ai
    from services.ai_assistant import RESPONSE_FORMAT_INSTRUCTIONS

    report_text = (req.report_text or "").strip()
    if not report_text and not req.attachments:
        raise HTTPException(status_code=400, detail="Informe um texto ou carregue ao menos um arquivo para análise.")
    prepared_attachments, attachment_meta = _prepare_ai_attachments(req.attachments)

    system = (
        _compose_base_system()
        + "\n\n"
        "Você é um especialista em metrologia e medição multifásica (MPFM) para produção "
        "offshore de petróleo e gás. Analise o relatório a seguir e forneça:\n"
        "1. Resumo dos principais indicadores (vazões, GOR, BSW, pressão, temperatura)\n"
        "2. Pontos de atenção ou anomalias identificadas\n"
        "3. Sugestões de ação ou investigação\n"
        "Responda em português, de forma técnica e estruturada.\n\n"
        f"{RESPONSE_FORMAT_INSTRUCTIONS}"
    )
    prompt = f"Relatório MPFM / anexos para análise:\n\n{report_text or 'A análise deve considerar os arquivos anexados.'}"
    conversation_id = req.conversation_id or _create_ai_conversation("Análise de relatório MPFM", "assistente")

    try:
        resp = await ask_ai(
            user=prompt,
            system=system,
            provider=req.provider or "gemini",
            max_tokens=8192,
            temperature=0.2,
            attachments=prepared_attachments,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc))
    except Exception as exc:
        logger.exception("ai_analyze_report error")
        raise HTTPException(status_code=500, detail=f"Erro ao analisar relatório: {exc}")

    user_message_id = _log_ai_message(
        conversation_id=conversation_id,
        role="user",
        content=f"[Análise de relatório]\n\n{report_text}".strip(),
        metadata={"attachments": attachment_meta},
    )
    assistant_message_id = _log_ai_message(
        conversation_id=conversation_id,
        role="assistant",
        content=resp.content,
        provider=resp.provider,
        model=resp.model,
        input_tokens=_token_count(resp.input_tokens),
        output_tokens=_token_count(resp.output_tokens),
        metadata={"request_type": "analyze_report", "user_message_id": user_message_id, "attachments": attachment_meta},
    )

    return AskResponse(
        content=resp.content,
        provider=resp.provider,
        model=resp.model,
        input_tokens=_token_count(resp.input_tokens),
        output_tokens=_token_count(resp.output_tokens),
        conversation_id=conversation_id,
        message_id=assistant_message_id,
    )
