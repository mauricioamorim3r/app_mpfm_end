from __future__ import annotations

import glob
import hashlib
import json
import math
import os
import re
import shutil
import sqlite3
import tempfile
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

from fastapi import HTTPException, Request
from fastapi.responses import FileResponse, HTMLResponse

from routes.date_utils import normalize_date_input
from services.importing.sep_folder_scan_service import DEFAULT_FOLDER_NAMES, scan_sep_folder



def register_system_routes(app, ctx: dict) -> None:
    db_conn = ctx["db_conn"]
    process_file_list = ctx["process_file_list"]
    default_density = ctx["default_density"]
    upload_dir = ctx["upload_dir"]
    output_dir = ctx["output_dir"]
    static_dir = ctx["static_dir"]
    work_dir = ctx["work_dir"]
    app_title = ctx["app_title"]
    app_version = ctx["app_version"]
    public_base_url = ctx["public_base_url"]
    db_path = ctx["db_path"]
    is_monthly_workbook_rebuilding = ctx["is_monthly_workbook_rebuilding"]
    prefs_path = work_dir / "user_prefs.json"

    _B05_OLD_FOLDER = "3.1.4_18-FT-1506 PE 4 - Subsea B05"
    _B05_OFFICIAL_FOLDER = "3.1.4_18-FT-1506 PE 4 e PE_EO105 - Subsea B05"

    def _canonical_monitor_folder_path(folder_path: str) -> str:
        """Mantém PE-4A e PE-EO105 na única pasta oficial compartilhada."""
        return str(folder_path or "").strip().replace(_B05_OLD_FOLDER, _B05_OFFICIAL_FOLDER)
    
    allowed_output_patterns = (
        re.compile(r"^MPFM_(JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)_\d{4}\.xlsx$", re.IGNORECASE),
        re.compile(r"^SEP_Dados_\d{4}-\d{2}-\d{2}_a_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE),
        re.compile(r"^dados_producao_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE),
        re.compile(r"^mpfm_export_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.xlsx$", re.IGNORECASE),
    )

    def _now_local_iso() -> str:
        return datetime.now().replace(microsecond=0).isoformat()

    def _is_supported_output_workbook(path: Path) -> bool:
        name = path.name
        if name.lower() == "mpfm.xlsx":
            return True
        return any(pattern.match(name) for pattern in allowed_output_patterns)

    def _resolve_safe(base: Path, name: str) -> Path:
        resolved = (base / name).resolve()
        if not resolved.is_relative_to(base.resolve()):
            raise HTTPException(400, "Caminho de arquivo inválido")
        return resolved

    def _discover_measurement_files(folder: str | Path) -> list[Path]:
        """Return PDF/TXT candidates for production imports, skipping reserved app subfolders."""
        base = Path(folder).resolve()
        reserved_dirs = {"alarme", "alarmes", "alarm", "alarms"}
        all_files = sorted(glob.glob(os.path.join(str(base), "**", "*.pdf"), recursive=True)) + sorted(
            glob.glob(os.path.join(str(base), "**", "*.txt"), recursive=True)
        )
        candidates: list[Path] = []
        for path_str in all_files:
            path = Path(path_str)
            try:
                relative = path.resolve().relative_to(base)
            except ValueError:
                relative = path
            if any(part.lower() in reserved_dirs for part in relative.parts[:-1]):
                continue
            candidates.append(path)
        return candidates

    def _sep_folder_scan_from_body(body: dict) -> tuple[Path, list, dict]:
        folder = str((body or {}).get("folder") or (body or {}).get("source_root") or "").strip()
        if not folder or not os.path.isdir(folder):
            raise HTTPException(400, f"Pasta não encontrada: {folder}")

        date_from = normalize_date_input(str((body or {}).get("date_from") or "").strip())
        date_to = normalize_date_input(str((body or {}).get("date_to") or "").strip())
        if date_from and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
            raise HTTPException(400, "Data inicial inválida. Use YYYY-MM-DD ou DD/MM/YYYY.")
        if date_to and not re.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
            raise HTTPException(400, "Data final inválida. Use YYYY-MM-DD ou DD/MM/YYYY.")
        if date_from and date_to and date_from > date_to:
            raise HTTPException(400, "Data inicial maior que data final.")

        raw_folders = (body or {}).get("folder_names") or DEFAULT_FOLDER_NAMES
        if isinstance(raw_folders, str):
            raw_folders = [item.strip() for item in re.split(r"[,;]", raw_folders)]
        folder_names = [str(item).strip() for item in raw_folders if str(item).strip()] or list(DEFAULT_FOLDER_NAMES)
        include_incomplete = bool((body or {}).get("include_incomplete_days"))

        source_root = Path(folder).resolve()
        _candidates, selected, preview = scan_sep_folder(
            source_root,
            date_from=date_from,
            date_to=date_to,
            folder_names=folder_names,
            include_incomplete_days=include_incomplete,
        )
        return source_root, selected, preview

    def _download_readiness(path: Path) -> dict:
        rebuilding = bool(path.suffix.lower() == ".xlsx" and is_monthly_workbook_rebuilding(path))
        return {
            "is_rebuilding": rebuilding,
            "rebuild_message": "Workbook mensal em atualização assíncrona. Aguarde alguns segundos e atualize a tela." if rebuilding else "",
        }

    def _load_monitor_prefs() -> dict:
        base = {"enabled": False, "stability_seconds": 20, "schedule_enabled": False, "schedule_times": ["09:00", "12:00", "18:00"], "interval_enabled": True, "folders": []}
        if prefs_path.exists():
            try:
                payload = json.loads(prefs_path.read_text("utf-8"))
                current = payload.get("auto_folder_monitor") or {}
                if isinstance(current, dict):
                    base.update({k: current[k] for k in ("enabled", "stability_seconds", "schedule_enabled", "schedule_times", "interval_enabled") if k in current})
                    base["folders"] = current.get("folders") or []
            except Exception:
                pass
        base["enabled"] = bool(base.get("enabled"))
        base["schedule_enabled"] = bool(base.get("schedule_enabled"))
        base["interval_enabled"] = bool(base.get("interval_enabled", True))
        try:
            base["stability_seconds"] = max(5, min(300, int(base.get("stability_seconds") or 20)))
        except Exception:
            base["stability_seconds"] = 20
        raw_st = base.get("schedule_times") or []
        if isinstance(raw_st, str):
            raw_st = [t.strip() for t in raw_st.split(",")]
        valid_st = [t for t in raw_st if re.match(r"^\d{2}:\d{2}$", str(t).strip())]
        base["schedule_times"] = valid_st if valid_st else ["09:00", "12:00", "18:00"]
        folders = []
        for idx, item in enumerate(base.get("folders") or []):
            if not isinstance(item, dict):
                continue
            folder_path = _canonical_monitor_folder_path(item.get("path") or "")
            if not folder_path:
                continue
            duplicate_policy = str(item.get("duplicate_policy") or "skip").strip().lower()
            if duplicate_policy not in {"skip", "overwrite"}:
                duplicate_policy = "skip"
            try:
                interval_seconds = max(30, min(3600, int(item.get("interval_seconds") or 300)))
            except Exception:
                interval_seconds = 300
            folders.append(
                {
                    "id": str(item.get("id") or f"folder-{idx+1}"),
                    "label": str(item.get("label") or "").strip(),
                    "path": folder_path,
                    "active": item.get("active", True) is not False,
                    "interval_seconds": interval_seconds,
                    "duplicate_policy": duplicate_policy,
                }
            )
        base["folders"] = folders
        return base

    def _save_monitor_prefs(config: dict) -> dict:
        current = {}
        if prefs_path.exists():
            try:
                current = json.loads(prefs_path.read_text("utf-8"))
            except Exception:
                current = {}
        raw_st = (config or {}).get("schedule_times") or []
        if isinstance(raw_st, str):
            raw_st = [t.strip() for t in raw_st.split(",")]
        valid_st = [t for t in raw_st if re.match(r"^\d{2}:\d{2}$", str(t).strip())]
        sanitized = _load_monitor_prefs() | {
            "enabled": bool((config or {}).get("enabled")),
            "stability_seconds": max(5, min(300, int((config or {}).get("stability_seconds") or 20))),
            "schedule_enabled": bool((config or {}).get("schedule_enabled")),
            "schedule_times": valid_st if valid_st else ["09:00", "12:00", "18:00"],
            "interval_enabled": bool((config or {}).get("interval_enabled", True)),
            "folders": [],
        }
        for idx, item in enumerate((config or {}).get("folders") or []):
            if not isinstance(item, dict):
                continue
            folder_path = _canonical_monitor_folder_path(item.get("path") or "")
            if not folder_path:
                continue
            duplicate_policy = str(item.get("duplicate_policy") or "skip").strip().lower()
            if duplicate_policy not in {"skip", "overwrite"}:
                duplicate_policy = "skip"
            try:
                interval_seconds = max(30, min(3600, int(item.get("interval_seconds") or 300)))
            except Exception:
                interval_seconds = 300
            sanitized["folders"].append(
                {
                    "id": str(item.get("id") or f"folder-{idx+1}"),
                    "label": str(item.get("label") or "").strip(),
                    "path": folder_path,
                    "active": item.get("active", True) is not False,
                    "interval_seconds": interval_seconds,
                    "duplicate_policy": duplicate_policy,
                }
            )
        current["auto_folder_monitor"] = sanitized
        prefs_path.write_text(json.dumps(current, ensure_ascii=False, indent=2), "utf-8")
        return sanitized

    monitor_state = {
        "lock": threading.Lock(),
        "wake_event": threading.Event(),
        "stop_event": threading.Event(),
        "thread": None,
        "schedule_triggered": set(),
        "status": {
            "running": False,
            "last_cycle_at": "",
            "last_cycle_message": "Monitor automático inativo.",
            "folders": {},
        },
    }

    def _file_sha1(path: Path) -> str:
        digest = hashlib.sha1()
        with path.open("rb") as fh:
            while True:
                chunk = fh.read(1024 * 1024)
                if not chunk:
                    break
                digest.update(chunk)
        return digest.hexdigest()

    def _find_duplicates(payload_items: list[dict]) -> list[dict]:
        if not payload_items:
            return []
        conn = db_conn()
        cur = conn.cursor()
        results = []
        try:
            for item in payload_items:
                file_id = item.get("file_id") or item.get("filename") or ""
                fname = item.get("filename") or ""
                file_hash = (item.get("file_hash") or "").strip().lower()
                row = None
                mode = ""
                if file_hash:
                    row = cur.execute(
                        """SELECT filename, detected_type, content_date, created_at, meter_id, location
                           FROM source_files_raw WHERE file_hash=? ORDER BY id DESC LIMIT 1""",
                        (file_hash,),
                    ).fetchone()
                    mode = "same_content" if row else ""
                    if not row:
                        row = cur.execute(
                            """SELECT source_file AS filename, fluid_kind AS file_type, production_date AS content_date,
                                      updated_at AS created_at, meter_id, location
                               FROM sep_source_files WHERE source_hash=? ORDER BY id DESC LIMIT 1""",
                            (file_hash,),
                        ).fetchone()
                        mode = "same_content" if row else ""
                if not row and fname and not fname.lower().endswith(".txt"):
                    row = cur.execute(
                        """SELECT filename, file_type, content_date, created_at, meter_id, location
                           FROM files_imported WHERE filename=? ORDER BY id DESC LIMIT 1""",
                        (fname,),
                    ).fetchone()
                    mode = "same_name" if row else ""
                if not row and fname and fname.lower().endswith(".txt"):
                    row = cur.execute(
                        """SELECT source_file AS filename, fluid_kind AS file_type, production_date AS content_date,
                                  updated_at AS created_at, meter_id, location
                           FROM sep_source_files WHERE source_file=? ORDER BY id DESC LIMIT 1""",
                        (fname,),
                    ).fetchone()
                    mode = "same_name" if row else ""
                if row:
                    results.append(
                        {
                            "file_id": file_id,
                            "filename": fname,
                            "file_type": row[1] or "",
                            "content_date": row[2] or "",
                            "last_imported": (row[3] or "")[:16].replace("T", " "),
                            "meter_id": row[4] or "",
                            "location": row[5] or "",
                            "duplicate_mode": mode or "same_content",
                        }
                    )
        finally:
            conn.close()
        return results

    def _serialize_folder_status(folder_cfg: dict) -> dict:
        status = monitor_state["status"]["folders"].get(folder_cfg["id"], {})
        return {
            "id": folder_cfg["id"],
            "label": folder_cfg.get("label") or Path(folder_cfg["path"]).name,
            "path": folder_cfg["path"],
            "active": folder_cfg.get("active", True),
            "interval_seconds": folder_cfg.get("interval_seconds", 300),
            "duplicate_policy": folder_cfg.get("duplicate_policy", "skip"),
            "last_scan_at": status.get("last_scan_at", ""),
            "last_action_at": status.get("last_action_at", ""),
            "last_action_trigger": status.get("last_action_trigger", ""),
            "last_result": status.get("last_result", ""),
            "last_scan_result": status.get("last_scan_result", ""),
            "last_found": status.get("last_found", 0),
            "last_processed": status.get("last_processed", 0),
            "last_skipped": status.get("last_skipped", 0),
            "last_error": status.get("last_error", ""),
            "next_scan_at": status.get("next_scan_at_iso", ""),
        }

    def _update_folder_status_summary(
        entry: dict,
        trigger: str,
        processed: int,
        result_text: str,
        *,
        force_display: bool = False,
    ) -> None:
        entry["last_scan_result"] = result_text
        if force_display or trigger == "manual" or processed > 0 or not entry.get("last_action_at"):
            entry["last_result"] = result_text
            entry["last_action_at"] = entry.get("last_scan_at", "")
            entry["last_action_trigger"] = trigger

    def _stable_folder_files(folder: str, stability_seconds: int) -> list[Path]:
        now = time.time()
        stable = []
        for path in _discover_measurement_files(folder):
            try:
                if not path.is_file():
                    continue
                age = now - path.stat().st_mtime
                if age < stability_seconds:
                    continue
                stable.append(path)
            except OSError:
                continue
        return stable

    def _parse_mpfm_pdf_filename(path: Path) -> dict | None:
        match = re.search(r"(?P<bank>B\d{2})_MPFM_(?P<kind>Daily|Hourly)-(?P<date>\d{8})-(?P<time>\d{6})", path.name, re.IGNORECASE)
        if not match:
            return None
        kind = match.group("kind").lower()
        try:
            end_dt = datetime.strptime(f'{match.group("date")}{match.group("time")}', "%Y%m%d%H%M%S")
        except ValueError:
            return None
        if kind == "daily":
            production_day = (end_dt.date() - timedelta(days=1)).isoformat()
            hour = None
        else:
            start_dt = end_dt - timedelta(hours=1)
            production_day = start_dt.date().isoformat()
            hour = start_dt.hour
        return {
            "path": path,
            "name": path.name,
            "bank": match.group("bank").upper(),
            "report_type": kind,
            "production_day": production_day,
            "hour": hour,
        }

    def _active_mpfm_monitor_folders() -> list[dict]:
        return [folder for folder in monitor_prefs.get("folders", []) if folder.get("active", True)]

    def _candidate_existing_mpfm_roots(folder_path: str) -> list[Path]:
        raw = Path(_canonical_monitor_folder_path(folder_path))
        variants: list[Path] = []
        if str(raw):
            variants.append(raw)
            parts = list(raw.parts)
            for idx, part in enumerate(parts):
                if part == "3. Registros de Operação SGM Multifasico" and idx + 1 < len(parts) and parts[idx + 1] == "3.1 Registros Diarios MPFM":
                    try:
                        variants.append(Path(parts[0], *parts[1:idx], *parts[idx + 1:]))
                    except Exception:
                        pass
                    break

        roots: list[Path] = []
        seen: set[str] = set()
        month_match = re.match(r"^(?P<num>\d{2})\.", raw.name or "")
        preferred_months: set[int] = set()
        if month_match:
            try:
                month_num = int(month_match.group("num"))
                preferred_months.update({month_num - 1, month_num, month_num + 1})
            except Exception:
                preferred_months = set()
        for variant in variants:
            if not variant.exists() or not variant.is_dir():
                continue
            scan_roots = [variant]
            if re.match(r"^\d{2}\.", variant.name or "") and variant.parent.is_dir() and re.match(r"^\d{4}$", variant.parent.name or ""):
                siblings = [item for item in variant.parent.iterdir() if item.is_dir() and re.match(r"^\d{2}\.", item.name or "")]
                filtered = []
                for item in siblings:
                    item_match = re.match(r"^(\d{2})\.", item.name or "")
                    if not item_match:
                        continue
                    try:
                        if int(item_match.group(1)) in preferred_months:
                            filtered.append(item)
                    except Exception:
                        continue
                scan_roots = filtered or [variant]
            for root in scan_roots:
                key = str(root.resolve()).casefold()
                if key in seen:
                    continue
                seen.add(key)
                roots.append(root.resolve())
        return sorted(roots, key=lambda item: str(item).casefold())

    def _discover_mpfm_pdf_files_fast(scan_root: Path) -> list[Path]:
        candidates: list[Path] = []
        if scan_root.name.lower() in {"daily", "hourly"}:
            candidates.extend(scan_root.glob("*.pdf"))
        else:
            for child_name in ("Daily", "DAILY", "daily", "Hourly", "HOURLY", "hourly"):
                child = scan_root / child_name
                if child.is_dir():
                    candidates.extend(child.glob("*.pdf"))
            candidates.extend(scan_root.glob("*.pdf"))
        seen: set[str] = set()
        result: list[Path] = []
        for path in candidates:
            if not path.is_file():
                continue
            key = str(path.resolve()).casefold()
            if key in seen:
                continue
            seen.add(key)
            result.append(path.resolve())
        return sorted(result, key=lambda item: item.name.casefold())

    def _scan_latest_day_ingestion(body: dict | None = None) -> dict:
        body = body or {}
        requested_day = normalize_date_input(str(body.get("target_day") or body.get("production_day") or "").strip())
        try:
            days_count = max(1, min(31, int(body.get("days_count") or body.get("last_days") or 1)))
        except Exception:
            days_count = 1
        if requested_day and not re.match(r"^\d{4}-\d{2}-\d{2}$", requested_day):
            raise HTTPException(400, "Data de produção inválida. Use YYYY-MM-DD ou DD/MM/YYYY.")
        configured_folders = _active_mpfm_monitor_folders()
        folder_candidates: list[dict] = []
        folder_summaries: list[dict] = []

        for folder in configured_folders:
            folder_path = str(folder.get("path") or "").strip()
            label = folder.get("label") or Path(folder_path).name
            summary = {
                "folder_id": folder.get("id") or "",
                "label": label,
                "path": folder_path,
                "exists": False,
                "scan_roots": [],
                "daily_found": 0,
                "hourly_found": 0,
                "selected": 0,
                "missing": [],
                "error": "",
            }
            scan_roots = _candidate_existing_mpfm_roots(folder_path)
            summary["exists"] = bool(scan_roots)
            summary["scan_roots"] = [str(path) for path in scan_roots]
            if not scan_roots:
                summary["error"] = f"Pasta não encontrada: {folder_path}"
                folder_summaries.append(summary)
                continue
            try:
                files = []
                for scan_root in scan_roots:
                    files.extend(_discover_mpfm_pdf_files_fast(scan_root))
            except Exception as exc:
                summary["error"] = str(exc)
                folder_summaries.append(summary)
                continue
            for path in files:
                meta = _parse_mpfm_pdf_filename(path)
                if not meta:
                    continue
                meta["folder_id"] = summary["folder_id"]
                meta["folder_label"] = label
                folder_candidates.append(meta)
            folder_summaries.append(summary)

        available_days = sorted({item["production_day"] for item in folder_candidates}, reverse=True)
        if requested_day:
            end_day = datetime.strptime(requested_day, "%Y-%m-%d").date()
            start_day = end_day - timedelta(days=days_count - 1)
            target_days = [(start_day + timedelta(days=offset)).isoformat() for offset in range(days_count)]
        else:
            target_days = available_days[:days_count]
        target_day = target_days[0] if target_days else ""
        target_day_set = set(target_days)
        selected_mpfm = [item for item in folder_candidates if item["production_day"] in target_day_set] if target_days else []
        selected_paths = {str(item["path"].resolve()) for item in selected_mpfm}

        by_folder_type: dict[tuple[str, str], int] = {}
        by_folder_day_type: dict[tuple[str, str, str], int] = {}
        for item in selected_mpfm:
            key = (item.get("folder_id") or "", item["report_type"])
            by_folder_type[key] = by_folder_type.get(key, 0) + 1
            day_key = (item.get("folder_id") or "", item["production_day"], item["report_type"])
            by_folder_day_type[day_key] = by_folder_day_type.get(day_key, 0) + 1
        for summary in folder_summaries:
            summary["daily_found"] = by_folder_type.get((summary["folder_id"], "daily"), 0)
            summary["hourly_found"] = by_folder_type.get((summary["folder_id"], "hourly"), 0)
            summary["selected"] = summary["daily_found"] + summary["hourly_found"]
            if summary["exists"] and target_days:
                for day in sorted(target_days):
                    day_daily = by_folder_day_type.get((summary["folder_id"], day, "daily"), 0)
                    day_hourly = by_folder_day_type.get((summary["folder_id"], day, "hourly"), 0)
                    if day_daily < 1:
                        summary["missing"].append(f"{day} daily")
                    if day_hourly < 24:
                        summary["missing"].append(f"{day} hourly {day_hourly}/24")

        sep_root_raw = str(body.get("sep_root") or body.get("sep_folder") or "").strip()
        raw_folders = body.get("sep_folder_names") or body.get("folder_names") or DEFAULT_FOLDER_NAMES
        if isinstance(raw_folders, str):
            raw_folders = [item.strip() for item in re.split(r"[,;]", raw_folders)]
        sep_selected = []
        sep_preview = {
            "enabled": bool(sep_root_raw),
            "source_root": sep_root_raw,
            "selected_count": 0,
            "candidate_count": 0,
            "complete": False,
            "message": "Pasta SEP não informada; a rotina seguirá apenas com PDFs MPFM.",
        }
        if sep_root_raw:
            if not os.path.isdir(sep_root_raw):
                sep_preview["message"] = f"Pasta SEP não encontrada: {sep_root_raw}"
            elif target_days:
                try:
                    sep_date_from = min(target_days)
                    sep_date_to = max(target_days)
                    _sep_candidates, sep_selected, raw_preview = scan_sep_folder(
                        Path(sep_root_raw),
                        date_from=sep_date_from,
                        date_to=sep_date_to,
                        folder_names=raw_folders,
                        include_incomplete_days=True,
                    )
                    sep_selected = [item for item in sep_selected if item.content_date in target_day_set]
                    stats = raw_preview.get("stats") or {}
                    day_rows = raw_preview.get("days") or []
                    complete_days = {row.get("date") for row in day_rows if row.get("is_complete")}
                    sep_preview = {
                        **raw_preview,
                        "enabled": True,
                        "selected_count": len(sep_selected),
                        "candidate_count": int(stats.get("candidate_count") or 0),
                        "complete": all(day in complete_days for day in target_days),
                        "message": f"{len(sep_selected)} TXT(s) SEP elegível(is) para {sep_date_from} a {sep_date_to}.",
                    }
                except Exception as exc:
                    sep_preview["message"] = f"Falha na varredura SEP: {exc}"

        items = []
        for item in sorted(selected_mpfm, key=lambda row: (row["folder_label"], row["report_type"], row.get("hour") if row.get("hour") is not None else -1, row["name"])):
            items.append(
                {
                    "file_id": str(item["path"].resolve()),
                    "filename": item["name"],
                    "path": str(item["path"].resolve()),
                    "kind": f"mpfm_{item['report_type']}",
                    "bank": item.get("bank") or "",
                    "folder_label": item.get("folder_label") or "",
                    "production_day": item["production_day"],
                    "hour": item.get("hour"),
                    "file_hash": "",
                }
            )
        for item in sep_selected:
            path_key = str(item.path.resolve())
            if path_key in selected_paths:
                continue
            items.append(
                {
                    "file_id": path_key,
                    "filename": item.name,
                    "path": path_key,
                    "kind": item.fluid_kind,
                    "bank": "SEP",
                    "folder_label": item.folder_name,
                    "production_day": item.content_date,
                    "hour": None,
                    "file_hash": "",
                }
            )

        duplicates = _find_duplicates(items)
        return {
            "ok": True,
            "requested_day": requested_day,
            "target_day": target_day,
            "target_days": target_days,
            "days_count": len(target_days) or days_count,
            "has_eligible": bool(items),
            "mpfm_count": len(selected_mpfm),
            "sep_count": len(sep_selected),
            "files_count": len(items),
            "duplicates_count": len(duplicates),
            "folders": folder_summaries,
            "sep": sep_preview,
            "items": items,
            "duplicates": duplicates,
            "message": (f"Sim: {len(items)} arquivo(s) elegível(is) encontrados para {target_days[0] if len(target_days) == 1 else f'{min(target_days)} a {max(target_days)}'}." if items else f"Não: nenhum arquivo elegível encontrado para {target_day or 'o último dia'}.")
        }

    def _process_latest_day_ingestion(body: dict | None = None) -> dict:
        body = body or {}
        preview = _scan_latest_day_ingestion(body)
        if not preview.get("items"):
            return {**preview, "processed": 0, "skipped": 0, "log": [preview.get("message") or "Nenhum arquivo elegível."]}
        overwrite_map = body.get("overwrite_map") or {}
        pairs = []
        skipped_log = []
        for item in preview.get("items") or []:
            path = item.get("path") or ""
            filename = item.get("filename") or Path(path).name
            decision = overwrite_map.get(path) or overwrite_map.get(filename) or overwrite_map.get(item.get("file_id") or "")
            if decision == "skip":
                skipped_log.append(f"⏭️  {filename}  →  ignorado por decisão do usuário")
                continue
            pairs.append((Path(path), filename))
        if not pairs:
            return {**preview, "processed": 0, "skipped": len(skipped_log), "log": skipped_log + ["⚠ Nenhum arquivo para processar (todos ignorados)."]}
        force_overwrite = bool(body.get("force_overwrite") or any(value == "overwrite" for value in overwrite_map.values()))
        result = process_file_list(
            pairs,
            default_density,
            source_type="latest-day",
            source_ref=f"latest-day:{','.join(preview.get('target_days') or [preview.get('target_day') or ''])}",
            force_overwrite=force_overwrite,
        )
        result["log"] = skipped_log + result.get("log", [])
        conn = db_conn()
        cur = conn.cursor()
        last_date = cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or ""
        conn.close()
        return {**preview, **result, "ok": True, "processed": len(pairs), "skipped": len(skipped_log), "last_date": last_date}

    def _run_monitored_folder(folder_cfg: dict, trigger: str = "interval") -> dict:
        with monitor_state["lock"]:
            entry = monitor_state["status"]["folders"].setdefault(folder_cfg["id"], {})
            entry["last_scan_at"] = _now_local_iso()
            entry["last_error"] = ""
            entry["trigger"] = trigger
            if not folder_cfg.get("active", True) and trigger != "manual":
                _update_folder_status_summary(entry, trigger, 0, "Pasta desativada.", force_display=True)
                return {"ok": True, "processed": 0, "skipped": 0, "found": 0, "log": ["Pasta desativada."]}
            folder = (folder_cfg.get("path") or "").strip()
            if not folder or not os.path.isdir(folder):
                entry["last_error"] = f"Pasta não encontrada: {folder}"
                _update_folder_status_summary(entry, trigger, 0, entry["last_error"], force_display=True)
                return {"ok": False, "processed": 0, "skipped": 0, "found": 0, "log": [entry["last_error"]]}
            stable_files = _stable_folder_files(folder, monitor_prefs["stability_seconds"])
            items = [{"file_id": str(path), "filename": path.name, "file_hash": _file_sha1(path)} for path in stable_files]
            duplicates = {dup["file_id"]: dup for dup in _find_duplicates(items)}
            pairs = []
            skipped_log = []
            skipped = 0
            for path in stable_files:
                duplicate = duplicates.get(str(path))
                if duplicate and duplicate.get("duplicate_mode") == "same_content":
                    skipped += 1
                    skipped_log.append(f"⏭️  {path.name}  →  ignorado (mesmo conteúdo já importado)")
                    continue
                if duplicate and folder_cfg.get("duplicate_policy") == "skip":
                    skipped += 1
                    skipped_log.append(f"⏭️  {path.name}  →  ignorado (mesmo nome já importado)")
                    continue
                pairs.append((path, path.name))
            if not pairs:
                entry["last_found"] = len(stable_files)
                entry["last_processed"] = 0
                entry["last_skipped"] = skipped
                _update_folder_status_summary(entry, trigger, 0, "Nenhum arquivo novo elegível.")
                return {"ok": True, "processed": 0, "skipped": skipped, "found": len(stable_files), "log": skipped_log + ["⚠ Nenhum arquivo novo elegível."]}
            result = process_file_list(pairs, default_density, source_type="auto-folder", source_ref=folder)
            entry["last_found"] = len(stable_files)
            entry["last_processed"] = len(pairs)
            entry["last_skipped"] = skipped
            _update_folder_status_summary(
                entry,
                trigger,
                len(pairs),
                f"{len(pairs)} arquivo(s) processado(s), {skipped} ignorado(s)",
            )
            return {
                "ok": True,
                "processed": len(pairs),
                "skipped": skipped,
                "found": len(stable_files),
                "log": skipped_log + result.get("log", []),
            }

    monitor_prefs = _load_monitor_prefs()

    def _refresh_monitor_next_scan(force_folder_id: str | None = None) -> None:
        now = time.time()
        folders = {folder["id"]: folder for folder in monitor_prefs.get("folders", [])}
        for folder_id, status in list(monitor_state["status"]["folders"].items()):
            if folder_id not in folders:
                monitor_state["status"]["folders"].pop(folder_id, None)
        # Stagger initial scan times: daily=300s, hourly=600s between each folder
        accumulated_stagger = 0
        for idx, folder in enumerate(monitor_prefs.get("folders", [])):
            status = monitor_state["status"]["folders"].setdefault(folder["id"], {})
            if folder.get("active", True) is False and not (force_folder_id and folder["id"] == force_folder_id):
                status.pop("next_scan_ts", None)
                status["next_scan_at_iso"] = ""
                continue
            if force_folder_id and folder["id"] == force_folder_id:
                status["next_scan_ts"] = now
            elif "next_scan_ts" not in status:
                status["next_scan_ts"] = now + accumulated_stagger
            _stagger = 600 if int(folder.get("interval_seconds") or 300) >= 600 else 300
            accumulated_stagger += _stagger
            status["next_scan_at_iso"] = datetime.fromtimestamp(status["next_scan_ts"]).isoformat()

    def _monitor_snapshot() -> dict:
        _refresh_monitor_next_scan()
        return {
            "config": monitor_prefs,
            "runtime": {
                "running": monitor_state["status"]["running"],
                "last_cycle_at": monitor_state["status"]["last_cycle_at"],
                "last_cycle_message": monitor_state["status"]["last_cycle_message"],
                "folders": [_serialize_folder_status(folder) for folder in monitor_prefs.get("folders", [])],
            },
        }

    def _monitor_loop() -> None:
        while not monitor_state["stop_event"].is_set():
            if not monitor_prefs.get("enabled"):
                monitor_state["status"]["running"] = False
                monitor_state["status"]["last_cycle_message"] = "Monitor automático desativado."
                monitor_state["wake_event"].wait(timeout=5)
                monitor_state["wake_event"].clear()
                continue
            monitor_state["status"]["running"] = True
            cycle_messages = []
            now = time.time()
            now_dt = datetime.now()
            today_str = now_dt.strftime("%Y-%m-%d")
            current_hhmm = now_dt.strftime("%H:%M")
            active_folders = [folder for folder in monitor_prefs.get("folders", []) if folder.get("active", True)]

            # ── Scheduled time triggers (09:00, 12:00, 18:00, etc.) ──────────
            if monitor_prefs.get("schedule_enabled"):
                triggered = monitor_state["schedule_triggered"]
                for stime in (monitor_prefs.get("schedule_times") or []):
                    key = f"{today_str}|{stime}"
                    if current_hhmm == stime and key not in triggered:
                        triggered.add(key)
                        # Run folders strictly sequentially: finish one, wait 5 min, then start the next.
                        # Never run more than one folder at the same time.
                        _SCHED_GAP_S = 300  # 5 minutes between consecutive folders
                        for seq_idx, folder in enumerate(active_folders):
                            try:
                                result = _run_monitored_folder(folder, trigger=f"agendado-{stime}")
                                fstatus = monitor_state["status"]["folders"].setdefault(folder["id"], {})
                                cycle_messages.append(
                                    f"[{stime}] {folder.get('label') or Path(folder['path']).name}: "
                                    f"{result['processed']} processado(s), {result['skipped']} ignorado(s)"
                                )
                            except Exception as exc:
                                fstatus = monitor_state["status"]["folders"].setdefault(folder["id"], {})
                                fstatus["last_error"] = str(exc)
                                cycle_messages.append(f"[{stime}] {folder.get('label') or Path(folder['path']).name}: erro")
                            # Wait 5 min between folders; no gap after the last one
                            if seq_idx < len(active_folders) - 1:
                                time.sleep(_SCHED_GAP_S)
                # Remove keys from previous days to keep the set small
                monitor_state["schedule_triggered"] = {k for k in triggered if k.startswith(today_str)}

            # ── Interval-based triggers ───────────────────────────────────────
            if monitor_prefs.get("interval_enabled", True):
                for folder in active_folders:
                    status = monitor_state["status"]["folders"].setdefault(folder["id"], {})
                    next_scan_ts = status.get("next_scan_ts", now)
                    if next_scan_ts > now:
                        continue
                    try:
                        result = _run_monitored_folder(folder, trigger="interval")
                        cycle_messages.append(f"{folder.get('label') or Path(folder['path']).name}: {result['processed']} processado(s), {result['skipped']} ignorado(s)")
                    except Exception as exc:
                        status["last_error"] = str(exc)
                        status["last_result"] = f"Erro: {exc}"
                        cycle_messages.append(f"{folder.get('label') or Path(folder['path']).name}: erro")
                    status["next_scan_ts"] = time.time() + max(30, int(folder.get("interval_seconds") or 300))
                    status["next_scan_at_iso"] = datetime.fromtimestamp(status["next_scan_ts"]).isoformat()
            monitor_state["status"]["last_cycle_at"] = _now_local_iso()
            if cycle_messages:
                monitor_state["status"]["last_cycle_message"] = " | ".join(cycle_messages)
            elif not monitor_prefs.get("folders"):
                monitor_state["status"]["last_cycle_message"] = "Nenhuma pasta monitorada configurada."
            elif not active_folders:
                monitor_state["status"]["last_cycle_message"] = "Nenhuma pasta ativa no monitor automático."
            elif not monitor_prefs.get("schedule_enabled") and not monitor_prefs.get("interval_enabled", True):
                monitor_state["status"]["last_cycle_message"] = "Monitor ligado, mas sem agenda nem intervalo habilitados."
            monitor_state["wake_event"].wait(timeout=5)
            monitor_state["wake_event"].clear()

    def _start_monitor_thread() -> None:
        if monitor_state["thread"] and monitor_state["thread"].is_alive():
            return
        monitor_state["stop_event"].clear()
        _refresh_monitor_next_scan()
        thread = threading.Thread(target=_monitor_loop, name="auto-folder-monitor", daemon=True)
        monitor_state["thread"] = thread
        thread.start()

    def _stop_monitor_thread() -> None:
        monitor_state["stop_event"].set()
        monitor_state["wake_event"].set()

    @app.on_event("startup")
    async def on_startup():
        _start_monitor_thread()

    @app.on_event("shutdown")
    async def on_shutdown():
        _stop_monitor_thread()

    @app.get("/static/{filename:path}")
    def api_static(filename: str):
        path = _resolve_safe(static_dir, filename)
        if not path.exists():
            raise HTTPException(404, "Arquivo estático não encontrado")
        media_type = "application/javascript" if str(path).endswith(".js") else None
        return FileResponse(str(path), media_type=media_type)

    @app.get("/twin/{filename:path}")
    def api_twin_static(filename: str):
        twin_dir = Path(__file__).resolve().parent.parent / "twin"
        path = _resolve_safe(twin_dir, filename or "index.html")
        if not path.exists() or not path.is_file():
            raise HTTPException(404, "Arquivo twin não encontrado")
        if path.suffix.lower() == ".html":
            return HTMLResponse(path.read_text("utf-8"))
        return FileResponse(str(path))

    @app.get("/", response_class=HTMLResponse)
    def root():
        idx = Path(__file__).resolve().parent.parent / "index.html"
        return HTMLResponse(idx.read_text("utf-8") if idx.exists() else "<h1>index.html não encontrado</h1>")

    @app.post("/api/admin/repair-banks")
    def api_repair_banks():
        instrument_to_bank = {
            '18FT0506': 'B10', '18FT0306': 'B10', '18FT0106': 'B10',
            '18FT1506': 'B05', '18FT1406': 'B05', '18FT1706': 'B05', '18FT1806': 'B05',
            '18FT0706': 'B15', '18FT0906': 'B15', '18FT1206': 'B15', '18FT1106': 'B15',
            '13FT0167': 'B08', '13FT0217': 'B08',
            '13FT0267': 'B13', '13FT0317': 'B13',
            '13FT0367': 'B03', '13FT0417': 'B03',
            '20FT0244': 'SEP', '20FT0247': 'SEP', '20FT0251': 'SEP'
        }
        tag_to_bank = {
            'PE_2': 'B10', 'PE_8': 'B10', 'PE_9': 'B10',
            'PE_4': 'B05', 'PE_EO10': 'B05', 'PE_EO105': 'B05', 'PE_EO4': 'B05',
            'PE_1': 'B15', 'PI_1': 'B15', 'PI_2': 'B15', 'PW-104DA': 'B15',
            'Riser_P1': 'B08', 'Riser_P2': 'B08',
            'Riser_P3': 'B13', 'Riser_P4': 'B13',
            'Riser_P5': 'B03', 'Riser_P6': 'B03',
        }
        well_to_bank = {
            'PE_2': 'B10', '7-BAC-1-SPS': 'B10',
            'PE_4': 'B05', '7-BAC-5A-SPS': 'B05', 'PE-4A': 'B05',
            'PW-104DA': 'B15', '7-BAC-4D-SPS': 'B15',
        }
        conn = db_conn()
        cur = conn.cursor()
        
        # 1. repair measurements_curated
        bad_banks = ['B12799', 'B145', 'B15153', 'B17525', 'B19109', 'B47', 'B52', 'UNK']
        placeholders = ','.join('?' for _ in bad_banks)
        cur.execute(f"""
            SELECT id, bank, tag, instrument 
            FROM measurements_curated 
            WHERE bank IN ({placeholders}) 
               OR (bank = 'B08' AND (instrument IN ('18FT0506','18FT0306','18FT0106') OR tag IN ('PE_2','PE_8','PE_9')))
        """, bad_banks)
        rows = cur.fetchall()
        
        updates_by_id = []
        for row_id, current_bank, tag, inst in rows:
            correct_bank = instrument_to_bank.get(inst) or tag_to_bank.get(tag)
            if correct_bank and correct_bank != current_bank:
                updates_by_id.append((correct_bank, row_id))
                
        if updates_by_id:
            cur.executemany("UPDATE measurements_curated SET bank = ? WHERE id = ?", updates_by_id)
            
        # 2. repair files_imported
        cur.execute("""
            SELECT DISTINCT f.id, f.filename, f.unit_code, m.bank
            FROM files_imported f
            JOIN measurements_curated m ON f.run_id = m.run_id
            WHERE f.unit_code != m.bank AND m.bank IN ('B03','B05','B08','B10','B13','B15','SEP')
        """)
        file_fixes = cur.fetchall()
        for file_id, filename, old_unit, new_unit in file_fixes:
            cur.execute("UPDATE files_imported SET unit_code = ? WHERE id = ?", (new_unit, file_id))

        # 3. repair source_files_raw
        cur.execute("""
            SELECT DISTINCT s.id, s.filename, s.unit_code, m.bank
            FROM source_files_raw s
            JOIN measurements_curated m ON s.run_id = m.run_id
            WHERE s.unit_code != m.bank AND m.bank IN ('B03','B05','B08','B10','B13','B15','SEP')
        """)
        raw_fixes = cur.fetchall()
        for raw_id, filename, old_unit, new_unit in raw_fixes:
            cur.execute("UPDATE source_files_raw SET unit_code = ? WHERE id = ?", (new_unit, raw_id))

        # 4. repair xml042 tables
        for well, correct_bank in well_to_bank.items():
            cur.execute("""
                UPDATE xml042_documents
                SET bank = ?
                WHERE (well_operator_name = ? OR cod_cadastro_poco = ?) AND bank != ?
            """, (correct_bank, well, well, correct_bank))

            cur.execute("""
                UPDATE xml042_imported_files
                SET bank = ?
                WHERE (well_operator_name = ? OR subsea_tag = ?) AND bank != ?
            """, (correct_bank, well, well, correct_bank))

            cur.execute("""
                UPDATE xml042_imported_rows
                SET bank = ?
                WHERE (well_operator_name = ? OR subsea_tag = ?) AND bank != ?
            """, (correct_bank, well, well, correct_bank))

        conn.commit()
        conn.close()
        return {"status": "ok", "updated_measurements": len(updates_by_id), "file_fixes": len(file_fixes), "raw_fixes": len(raw_fixes)}

    @app.get("/api/health")
    def api_health():
        conn = db_conn()
        try:
            cur = conn.cursor()
            latest_day = cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or ""
            counts = {
                "processing_runs": cur.execute("SELECT COUNT(*) FROM processing_runs").fetchone()[0],
                "source_files_raw": cur.execute("SELECT COUNT(*) FROM source_files_raw").fetchone()[0],
                "measurements_curated": cur.execute("SELECT COUNT(*) FROM measurements_curated").fetchone()[0],
                "validation_issues": cur.execute("SELECT COUNT(*) FROM validation_issues").fetchone()[0],
            }
            status = "ok"
            detail = ""
        except sqlite3.Error as exc:
            # Keep the health endpoint usable even before the local DB schema exists.
            latest_day = ""
            counts = {
                "processing_runs": 0,
                "source_files_raw": 0,
                "measurements_curated": 0,
                "validation_issues": 0,
            }
            status = "degraded"
            detail = str(exc)
        finally:
            conn.close()
        return {
            "status": status,
            "app": app_title,
            "version": app_version,
            "public_base_url": public_base_url,
            "db_path": str(db_path),
            "work_dir": str(work_dir),
            "latest_day_ref": latest_day,
            "counts": counts,
            "detail": detail,
        }

    @app.post("/api/check-duplicates")
    async def api_check_duplicates(request: Request):
        import asyncio
        body = await request.json()
        items = body.get("items", [])
        filenames = body.get("filenames", [])
        if not items and not filenames:
            return {"duplicates": []}
        payload_items = items or [{"file_id": fname, "filename": fname, "file_hash": ""} for fname in filenames]
        duplicates = await asyncio.to_thread(_find_duplicates, payload_items)
        return {"duplicates": duplicates}

    @app.get("/api/browse-folder")
    def api_browse_folder():
        """Opens a native OS folder picker dialog (tkinter) and returns the selected path."""
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)
            selected = filedialog.askdirectory(title="Selecionar pasta para monitorar", parent=root)
            root.destroy()
            if selected:
                return {"ok": True, "path": str(Path(selected))}
            return {"ok": False, "path": ""}
        except Exception as exc:
            raise HTTPException(500, f"Não foi possível abrir o seletor de pasta: {exc}")

    @app.post("/api/check-folder-duplicates")
    def api_check_folder_duplicates(body: dict):
        folder = (body.get("folder") or "").strip()
        if not folder or not os.path.isdir(folder):
            raise HTTPException(400, f"Pasta não encontrada: {folder}")
        all_files = _discover_measurement_files(folder)
        if not all_files:
            return {"duplicates": [], "items": []}
        items = []
        for path in all_files:
            items.append(
                {
                    "file_id": str(path),
                    "filename": path.name,
                    "file_hash": _file_sha1(path),
                }
            )
        return {"items": items, "duplicates": _find_duplicates(items)}

    @app.post("/api/process-files")
    async def api_process_files(request: Request):
        form = await request.form()
        files = form.getlist("files")
        file_manifest_raw = form.get("file_manifest", None)
        file_manifest = json.loads(file_manifest_raw) if file_manifest_raw else []
        overwrite_map_raw = form.get("overwrite_map", None)
        overwrite_map = json.loads(overwrite_map_raw) if overwrite_map_raw else None

        tmp = Path(tempfile.mkdtemp(dir=upload_dir))
        pairs = []
        skipped_log = []
        for idx, file in enumerate(files):
            meta = file_manifest[idx] if idx < len(file_manifest) and isinstance(file_manifest[idx], dict) else {}
            fname = (meta.get("filename") or file.filename or f"upload_{idx}").strip()
            file_id = (meta.get("file_id") or fname or f"upload_{idx}").strip()
            if overwrite_map and (overwrite_map.get(file_id) == "skip" or overwrite_map.get(fname) == "skip"):
                skipped_log.append(f"⏭️  {fname}  →  ignorado (duplicado)")
                continue
            safe_name = Path(fname).name or f"upload_{idx}"
            dest = tmp / f"{idx:04d}_{safe_name}"
            dest.write_bytes(await file.read())
            pairs.append((dest, fname))

        if not pairs:
            try:
                shutil.rmtree(tmp)
            except Exception:
                pass
            return {"ok": True, "log": skipped_log + ["⚠ Nenhum arquivo para processar (todos ignorados)."], "last_date": ""}

        force_overwrite = bool(overwrite_map and any(v == "overwrite" for v in overwrite_map.values()))
        result = process_file_list(pairs, default_density, source_type="upload", source_ref="ui-upload", force_overwrite=force_overwrite)
        result["log"] = skipped_log + result.get("log", [])
        try:
            shutil.rmtree(tmp)
        except Exception:
            pass
        conn = db_conn()
        cur = conn.cursor()
        last_date = cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or ""
        conn.close()
        return {**result, "ok": True, "last_date": last_date}

    @app.post("/api/process-folder")
    def api_process_folder(body: dict):
        folder = body.get("folder", "").strip()
        if not folder or not os.path.isdir(folder):
            raise HTTPException(400, f"Pasta não encontrada: {folder}")
        overwrite_map = body.get("overwrite_map") or {}
        all_files = _discover_measurement_files(folder)
        if not all_files:
            raise HTTPException(400, "Nenhum PDF ou TXT encontrado.")
        pairs = []
        skipped_log = []
        for path in all_files:
            fname = path.name
            if overwrite_map and (overwrite_map.get(str(path)) == "skip" or overwrite_map.get(fname) == "skip"):
                skipped_log.append(f"⏭️  {fname}  →  ignorado (duplicado)")
                continue
            pairs.append((path, fname))
        if not pairs:
            return {"ok": True, "log": skipped_log + ["⚠ Nenhum arquivo para processar (todos ignorados)."], "last_date": ""}
        result = process_file_list(
            pairs,
            default_density,
            source_type="folder",
            source_ref=folder,
            force_overwrite=bool((body or {}).get("force_overwrite")),
        )
        result["log"] = skipped_log + result.get("log", [])
        conn = db_conn()
        cur = conn.cursor()
        last_date = cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or ""
        conn.close()
        return {**result, "ok": True, "last_date": last_date}

    @app.post("/api/latest-day-ingestion/preview")
    def api_latest_day_ingestion_preview(body: dict):
        return _scan_latest_day_ingestion(body or {})

    @app.post("/api/latest-day-ingestion/process")
    def api_latest_day_ingestion_process(body: dict):
        return _process_latest_day_ingestion(body or {})

    @app.post("/api/sep-folder/preview")
    def api_sep_folder_preview(body: dict):
        _source_root, _selected, preview = _sep_folder_scan_from_body(body or {})
        return {"ok": True, "preview": preview}

    @app.post("/api/sep-folder/import")
    def api_sep_folder_import(body: dict):
        source_root, selected, preview = _sep_folder_scan_from_body(body or {})
        if not selected:
            return {"ok": True, "preview": preview, "log": ["Nenhum TXT SEP elegível para importar."], "last_date": ""}

        items = [(item.path, item.name) for item in selected]
        result = process_file_list(
            items,
            default_density,
            source_type="manual-sep-folder",
            source_ref=str(source_root),
            force_overwrite=bool((body or {}).get("force_overwrite")),
        )
        conn = db_conn()
        cur = conn.cursor()
        last_date = cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or ""
        conn.close()
        return {**result, "ok": True, "preview": preview, "last_date": last_date}

    @app.get("/api/auto-folder-monitor")
    def api_get_auto_folder_monitor():
        return _monitor_snapshot()

    @app.post("/api/auto-folder-monitor")
    async def api_save_auto_folder_monitor(request: Request):
        nonlocal monitor_prefs
        body = await request.json()
        monitor_prefs = _save_monitor_prefs(body or {})
        _refresh_monitor_next_scan()
        monitor_state["wake_event"].set()
        return {"ok": True, **_monitor_snapshot()}

    @app.post("/api/auto-folder-monitor/run-now")
    async def api_run_auto_folder_monitor_now(request: Request):
        import asyncio
        body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
        folder_id = str((body or {}).get("folder_id") or "").strip() or None
        configured_folders = list(monitor_prefs.get("folders", []))
        if folder_id:
            folders = [folder for folder in configured_folders if folder["id"] == folder_id]
        else:
            folders = [folder for folder in configured_folders if folder.get("active", True)]
        if not folders:
            message = "Pasta monitorada não encontrada." if folder_id else "Nenhuma pasta ativa selecionada."
            return {"ok": True, "message": message, **_monitor_snapshot()}
        
        def _run_sync():
            results = []
            for folder in folders:
                result = _run_monitored_folder(folder, trigger="manual")
                status = monitor_state["status"]["folders"].setdefault(folder["id"], {})
                if folder.get("active", True):
                    status["next_scan_ts"] = time.time() + max(30, int(folder.get("interval_seconds") or 300))
                    status["next_scan_at_iso"] = datetime.fromtimestamp(status["next_scan_ts"]).isoformat()
                else:
                    status.pop("next_scan_ts", None)
                    status["next_scan_at_iso"] = ""
                results.append(
                    {
                        "folder_id": folder["id"],
                        "label": folder.get("label") or Path(folder["path"]).name,
                        "processed": result.get("processed", 0),
                        "skipped": result.get("skipped", 0),
                        "found": result.get("found", 0),
                        "ok": result.get("ok", False),
                    }
                )
                monitor_state["status"]["last_cycle_at"] = _now_local_iso()
            monitor_state["status"]["last_cycle_message"] = " | ".join(
                f"{item['label']}: {item['processed']} processado(s), {item['skipped']} ignorado(s)" for item in results
            )
            return results

        results = await asyncio.to_thread(_run_sync)
        return {"ok": True, "results": results, **_monitor_snapshot()}

    @app.get("/api/list-outputs")
    def api_list_outputs():
        files = sorted(
            (file for file in output_dir.glob("*.xlsx") if _is_supported_output_workbook(file)),
            key=lambda file: file.stat().st_mtime,
            reverse=True,
        )
        payload = []
        for file in files:
            payload.append(
                {
                    "name": file.name,
                    "size_kb": round(file.stat().st_size / 1024, 1),
                    "modified": datetime.fromtimestamp(file.stat().st_mtime).strftime("%d-%m-%Y %H:%M"),
                    **_download_readiness(file),
                }
            )
        return {"files": payload}

    @app.get("/api/download/{filename}")
    def api_download(filename: str):
        path = _resolve_safe(output_dir, filename)
        if not path.exists():
            raise HTTPException(404, "Arquivo não encontrado")
        readiness = _download_readiness(path)
        if readiness["is_rebuilding"]:
            raise HTTPException(409, readiness["rebuild_message"])
        return FileResponse(str(path), filename=filename)

    @app.delete("/api/outputs/{filename}")
    def api_delete(filename: str):
        path = _resolve_safe(output_dir, filename)
        if path.exists():
            path.unlink()
        return {"ok": True}

    @app.get("/api/sheets/{filename}")
    def api_list_sheets(filename: str):
        path = _resolve_safe(output_dir, filename)
        if not path.exists():
            raise HTTPException(404, "Arquivo não encontrado")
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheets = workbook.sheetnames
        workbook.close()
        return {"sheets": sheets}

    @app.get("/api/view/{filename}")
    def api_view_compat(filename: str):
        return api_list_sheets(filename)

    @app.get("/api/view/{filename}/{sheet}")
    def api_view_sheet(filename: str, sheet: str):
        path = _resolve_safe(output_dir, filename)
        if not path.exists():
            raise HTTPException(404, "Arquivo não encontrado")
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(404, f'Aba "{sheet}" não encontrada')
        ws = workbook[sheet]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 501:
                break
            rows.append(["" if v is None else ("NaN" if isinstance(v, float) and math.isnan(v) else v) for v in row])
        workbook.close()
        headers = rows[0] if rows else []
        data = rows[1:] if len(rows) > 1 else []
        return {"headers": headers, "rows": data, "total": len(data)}
