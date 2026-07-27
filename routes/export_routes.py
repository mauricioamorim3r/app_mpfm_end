from __future__ import annotations

import io
import json
from copy import copy
from collections import defaultdict
from pathlib import Path

from fastapi import HTTPException
from fastapi.responses import Response, StreamingResponse
from routes.date_utils import normalize_date_input, normalize_date_range
from services.excel_template_service import (
    PRODUCTION_EXPORT_TEMPLATE,
    SEP_EXPORT_TEMPLATE,
    center_cell_content,
    center_filled_cells,
    copy_cell_style,
)


def register_export_routes(app, ctx: dict) -> None:
    db_conn = ctx["db_conn"]
    daily_metric_groups = ctx["daily_metric_groups"]
    load_prefs = ctx["load_prefs"]
    sep_detail_headers = ctx["sep_detail_headers"]
    sep_detail_kind = ctx["sep_detail_kind"]

    @app.get("/api/export-sep-excel")
    def api_export_sep_excel(date_from: str = "", date_to: str = "", unit: str = ""):
        date_from, date_to = normalize_date_range(date_from, date_to)
        from datetime import datetime
        import openpyxl
        from openpyxl.utils import get_column_letter

        template_path = SEP_EXPORT_TEMPLATE

        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute(
                "SELECT MAX(day_ref) FROM measurements_active "
                "WHERE row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail') "
                "AND COALESCE(is_official,1)=1"
            ).fetchone()[0] or ""
        if not date_from:
            date_from = date_to

        detail_kind_by_fluid = {
            "oleo": "sep_oleo_detail",
            "gas": "sep_gas_detail",
            "agua": "sep_agua_detail",
        }
        fluid_by_detail_kind = {value: key for key, value in detail_kind_by_fluid.items()}
        detail_sql = (
            "SELECT day_ref, hour_ref, bank, tag, instrument, row_kind, metric_name, metric_value "
            "FROM measurements_active "
            "WHERE row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail') "
            "AND COALESCE(is_official,1)=1 AND day_ref BETWEEN ? AND ? "
        )
        params = [date_from, date_to]
        if unit:
            detail_sql += "AND bank=? "
            params.append(unit)
        detail_sql += "ORDER BY day_ref, row_kind, COALESCE(hour_ref,-1), tag, metric_name"
        detail_rows = [dict(item) for item in cur.execute(detail_sql, params).fetchall()]
        conn.close()

        detail_sheet_config = {
            "oleo": [
                ("Hour", "Hora"),
                ("Pressure_kPa", "Pressure (kpa)"),
                ("Pressure_barg", "Pressure (barg)"),
                ("Temperature_degC", "Temperature (deg c)"),
                ("SD_kg_sm3", "SD (kg/sm³)"),
                ("MD_kg_m3", "MD (kg/m³)"),
                ("IV_m3", "IV (m³)"),
                ("GV_m3", "GV (m³)"),
                ("GSV_sm3", "GSV (sm³)"),
                ("Mass_ton", "Mass (t)"),
                ("NSV_sm3", "NSV (sm³)"),
                ("BSW_pct", "BSW (%)"),
                ("CPL", "CPL"),
                ("CTL", "CTL"),
            ],
            "gas": [
                ("Hour", "Hora"),
                ("Pressure_kPa_g", "Pressure (kpa_g)"),
                ("Temperature_degC", "Temperature (deg c)"),
                ("SD_kg_sm3", "SD (kg/sm³)"),
                ("DT_kg_m3", "DT (kg/m³)"),
                ("GrVol_m3", "Gr. vol. (m³)"),
                ("StVol_m3", "St. vol. (m³)"),
                ("Mass_t", "Mass (t)"),
                ("Energy_GJ", "Energy (gj)"),
                ("DiffPress_kPa", "Diff. press. (kpa)"),
                ("Flowtime_min", "Flowtime (min)"),
            ],
            "agua": [
                ("Hour", "Hora"),
                ("Pressure_kPa", "Pressure (kpa)"),
                ("Temperature_degC", "Temperature (deg c)"),
                ("SD_kg_sm3", "SD (kg/sm³)"),
                ("MD_kg_m3", "MD (kg/m³)"),
                ("IV_m3", "IV (m³)"),
                ("GV_m3", "GV (m³)"),
                ("GSV_sm3", "GSV (sm³)"),
                ("Mass_ton", "Mass (t)"),
                ("NSV_sm3", "NSV (sm³)"),
                ("BSW_pct", "BSW (%)"),
                ("CPL", "CPL"),
                ("CTL", "CTL"),
            ],
        }
        avg_metrics = {
            "oleo": {"Pressure_kPa", "Pressure_barg", "Temperature_degC", "SD_kg_sm3", "MD_kg_m3", "BSW_pct", "CPL", "CTL"},
            "gas": {"Pressure_kPa_g", "Temperature_degC", "SD_kg_sm3", "DT_kg_m3", "DiffPress_kPa"},
            "agua": {"Pressure_kPa", "Temperature_degC", "SD_kg_sm3", "MD_kg_m3", "BSW_pct", "CPL", "CTL"},
        }

        wb = openpyxl.load_workbook(template_path) if template_path.exists() else openpyxl.Workbook()
        if wb.sheetnames == ["Sheet"]:
            wb.active.title = "Separador_Totais"

        def get_or_create_sheet(name: str):
            return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

        def clear_export_region(ws, start_row: int, start_col: int):
            for merged in list(ws.merged_cells.ranges):
                if merged.min_col >= start_col and merged.min_row >= start_row:
                    ws.unmerge_cells(str(merged))
            for row_idx in range(start_row, ws.max_row + 1):
                for col_idx in range(start_col, ws.max_column + 1):
                    ws.cell(row_idx, col_idx).value = None

        def copy_template_row(ws, template_row: int, target_row: int, start_col: int, end_col: int):
            if target_row == template_row and target_row <= ws.max_row:
                return
            for col_idx in range(start_col, end_col + 1):
                copy_cell_style(ws.cell(template_row, col_idx), ws.cell(target_row, col_idx))
            ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height

        def sorted_detail_keys(pivot):
            return sorted(
                pivot.keys(),
                key=lambda item: (
                    item[0] or "",
                    -1 if item[1] is None else int(item[1]),
                    item[2] or "",
                    item[3] or "",
                ),
            )

        detail_store = {}
        for fluid in detail_kind_by_fluid:
            detail_store[fluid] = {
                "pivot": defaultdict(dict),
                "meta": {},
                "grouped": defaultdict(list),
                "day_rows": {},
            }
        for row in detail_rows:
            fluid = fluid_by_detail_kind[row["row_kind"]]
            key = (row["day_ref"], row["hour_ref"], row["tag"], row["instrument"])
            detail_store[fluid]["pivot"][key][row["metric_name"]] = row["metric_value"]
            detail_store[fluid]["meta"][key] = {
                "day_ref": row["day_ref"],
                "hour_ref": row["hour_ref"],
                "bank": row["bank"],
                "tag": row["tag"],
                "instrument": row["instrument"],
            }
        for fluid, payload in detail_store.items():
            for key in sorted_detail_keys(payload["pivot"]):
                item_meta = payload["meta"][key]
                payload["grouped"][item_meta["day_ref"]].append((key, item_meta))
                if item_meta["hour_ref"] is None:
                    payload["day_rows"][item_meta["day_ref"]] = (key, item_meta)

        def metric_for_day(fluid: str, day_ref: str, metric_name: str):
            payload = detail_store[fluid]
            day_entry = payload["day_rows"].get(day_ref)
            if day_entry:
                day_value = payload["pivot"][day_entry[0]].get(metric_name)
                if day_value is not None:
                    return day_value
            values = []
            for key, item_meta in payload["grouped"].get(day_ref, []):
                if item_meta["hour_ref"] is None:
                    continue
                value = payload["pivot"][key].get(metric_name)
                if value is not None:
                    values.append(value)
            if not values:
                return None
            if metric_name in avg_metrics[fluid]:
                return sum(values) / len(values)
            return sum(values)

        def excel_date(day_ref: str):
            if not day_ref:
                return None
            return datetime.strptime(day_ref, "%Y-%m-%d").date()

        totals_sheet_name = "Separador_Totais" if "Separador_Totais" in wb.sheetnames else "Separador"
        ws = get_or_create_sheet(totals_sheet_name)
        if ws.title != "Separador_Totais":
            ws.title = "Separador_Totais"
        clear_export_region(ws, 2, 3)

        total_days = sorted(
            {
                day_ref
                for payload in detail_store.values()
                for day_ref in payload["grouped"].keys()
                if day_ref
            }
        )
        daily_start_row = 3
        daily_end_row = daily_start_row + len(total_days) - 1
        summary_values = {
            3: "-",
            4: "DAY",
            5: "SEP",
            6: "SEP",
            7: "extraido",
            8: f"=AVERAGE(H{daily_start_row}:H{daily_end_row})" if total_days else None,
            9: f"=AVERAGE(I{daily_start_row}:I{daily_end_row})" if total_days else None,
            10: f"=AVERAGE(J{daily_start_row}:J{daily_end_row})" if total_days else None,
            11: f"=AVERAGE(K{daily_start_row}:K{daily_end_row})" if total_days else None,
            12: f"=AVERAGE(L{daily_start_row}:L{daily_end_row})" if total_days else None,
            13: f"=AVERAGE(M{daily_start_row}:M{daily_end_row})" if total_days else None,
            14: "=I2+K2" if total_days else None,
            15: "=I2+K2+M2" if total_days else None,
        }
        for col_idx, value in summary_values.items():
            ws.cell(2, col_idx, value=value)

        for idx, day_ref in enumerate(total_days):
            row_idx = daily_start_row + idx
            oil_mass = metric_for_day("oleo", day_ref, "Mass_ton")
            gas_mass = metric_for_day("gas", day_ref, "Mass_t")
            water_mass = metric_for_day("agua", day_ref, "Mass_ton")
            row_values = {
                3: excel_date(day_ref),
                4: f"=DAY(C{row_idx})",
                5: "SEP",
                6: "SEP",
                7: "extraido",
                8: metric_for_day("oleo", day_ref, "GSV_sm3"),
                9: oil_mass,
                10: metric_for_day("gas", day_ref, "StVol_m3"),
                11: gas_mass,
                12: metric_for_day("agua", day_ref, "GSV_sm3"),
                13: water_mass,
                14: (oil_mass or 0) + (gas_mass or 0) if oil_mass is not None or gas_mass is not None else None,
                15: (oil_mass or 0) + (gas_mass or 0) + (water_mass or 0)
                if oil_mass is not None or gas_mass is not None or water_mass is not None
                else None,
            }
            for col_idx, value in row_values.items():
                ws.cell(row_idx, col_idx, value=value)
        if daily_end_row + 1 <= ws.max_row:
            clear_export_region(ws, daily_end_row + 1, 3)
        ws.auto_filter.ref = f"C1:O{max(3, daily_end_row)}"

        def add_fluid_sheet(name: str, fluid: str):
            ws2 = get_or_create_sheet(name)
            clear_export_region(ws2, 2, 3)
            headers2 = detail_sheet_config[fluid]
            payload = detail_store[fluid]
            row_idx = 2
            last_col = 2 + len(headers2)
            for day_ref in sorted(payload["grouped"]):
                hourly_entries = [
                    (key, item_meta)
                    for key, item_meta in payload["grouped"][day_ref]
                    if item_meta["hour_ref"] is not None
                ]
                day_key, day_meta = payload["day_rows"].get(day_ref, payload["grouped"][day_ref][0])

                if row_idx > ws2.max_row:
                    copy_template_row(ws2, 2, row_idx, 1, last_col)
                meta_cell = ws2.cell(
                    row=row_idx,
                    column=3,
                    value=f"Data: {day_ref}  |  TAG: {day_meta['tag'] or '-'}  |  Meter ID: {day_meta['instrument'] or '-'}",
                )
                copy_cell_style(ws2.cell(2, 3), meta_cell)
                ws2.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=last_col)
                row_idx += 1

                if row_idx > ws2.max_row:
                    copy_template_row(ws2, 3, row_idx, 1, last_col)
                day_values = ["DAY"]
                for metric_key, _label in headers2[1:]:
                    day_values.append(metric_for_day(fluid, day_ref, metric_key))
                for ci, value in enumerate(day_values, 3):
                    ws2.cell(row=row_idx, column=ci, value=value)
                row_idx += 1

                for hour_idx in range(24):
                    if row_idx > ws2.max_row:
                        copy_template_row(ws2, 4 + min(hour_idx, 23), row_idx, 1, last_col)
                    if hour_idx < len(hourly_entries):
                        key, item_meta = hourly_entries[hour_idx]
                        values = [int(item_meta["hour_ref"])]
                        values += [payload["pivot"][key].get(metric_key) for metric_key, _label in headers2[1:]]
                        for ci, value in enumerate(values, 3):
                            ws2.cell(row=row_idx, column=ci, value=value)
                    row_idx += 1
            ws2.auto_filter.ref = f"C1:{get_column_letter(last_col)}{max(2, ws2.max_row)}"
            if row_idx <= ws2.max_row:
                clear_export_region(ws2, row_idx, 3)

        add_fluid_sheet("separador oleo", "oleo")
        add_fluid_sheet("separador gas", "gas")
        add_fluid_sheet("separador agua", "agua")
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=SEP_Dados_{date_from}_a_{date_to}.xlsx"},
        )

    @app.get("/api/export-sep-csv")
    def api_export_sep_csv(date_from: str = "", date_to: str = "", unit: str = ""):
        date_from, date_to = normalize_date_range(date_from, date_to)
        import csv as csv_mod

        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute(
                "SELECT MAX(day_ref) FROM measurements_active WHERE row_kind='sep'"
            ).fetchone()[0] or ""
        if not date_from:
            date_from = date_to
        sql = (
            "SELECT id, day_ref, hour_ref, bank, metric_name, metric_value "
            "FROM measurements_active WHERE row_kind='sep' AND COALESCE(is_official,1)=1 "
            "AND day_ref BETWEEN ? AND ?"
        )
        params = [date_from, date_to]
        if unit:
            sql += " AND bank=?"
            params.append(unit)
        sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), bank, metric_name"
        rows = cur.execute(sql, params).fetchall()
        conn.close()

        conn = db_conn()
        cur = conn.cursor()
        align_map = {}
        for row in cur.execute(
            "SELECT production_date, GROUP_CONCAT(bank, ', ') AS banks "
            "FROM sep_alignments WHERE is_active=1 AND production_date BETWEEN ? AND ? "
            "GROUP BY production_date",
            (date_from, date_to),
        ).fetchall():
            align_map[row["production_date"]] = row["banks"] or ""
        conn.close()

        pivot_map = defaultdict(dict)
        pivot_meta = {}
        metric_set = []
        for row in rows:
            key = (row[1], row[2], row[3], "")
            pivot_map[key][row[4]] = row[5]
            pivot_meta[key] = {
                "day_ref": row[1],
                "hour_ref": row[2],
                "bank": row[3],
                "tag": "",
                "aligned_banks": align_map.get(row[1], ""),
            }
            if row[4] not in metric_set:
                metric_set.append(row[4])

        buf = io.StringIO()
        writer = csv_mod.writer(buf)
        writer.writerow(["Data", "Hora", "Origem SEP", "TAG SEP", "Status SEP", "Bancos alinhados"] + metric_set)
        for key in sorted(
            pivot_map.keys(),
            key=lambda item: (item[0] or "", -1 if item[1] is None else item[1], item[2] or "", item[3] or ""),
        ):
            meta = pivot_meta[key]
            hour_label = "DAY" if meta["hour_ref"] is None else f"{int(meta['hour_ref']):02d}:00"
            row = [
                meta["day_ref"],
                hour_label,
                meta["bank"],
                meta.get("tag", ""),
                "Aplicado" if meta.get("aligned_banks") else "Extraído",
                meta.get("aligned_banks", ""),
            ] + [pivot_map[key].get(metric, "") for metric in metric_set]
            writer.writerow(row)

        period = f"{date_from}_a_{date_to}".replace("-", "")
        return Response(
            content=buf.getvalue().encode("utf-8-sig"),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="SEP_Dados_{period}.csv"'},
        )

    @app.get("/api/cards/export-pdf")
    def api_cards_export_pdf(date_from: str = "", date_to: str = "", bank: str = "", date: str = ""):
        date = normalize_date_input(date)
        date_from, date_to = normalize_date_range(date_from, date_to)
        if date and not date_from and not date_to:
            date_from = date_to = date
        from io import BytesIO
        from reportlab.lib.colors import HexColor, black, white
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas

        conn = db_conn()
        max_day = conn.execute(
            "SELECT MAX(day_ref) FROM measurements_active WHERE row_kind='daily'"
        ).fetchone()[0] or ""
        conn.close()
        if not date_to:
            date_to = max_day
        if not date_from:
            date_from = date_to
        cards = daily_metric_groups(date_from, date_to, bank)
        if not cards:
            raise HTTPException(404, f"Nenhum card disponível para {date_from} até {date_to}.")

        buf = BytesIO()
        page_w, page_h = landscape(A4)
        canvas_obj = canvas.Canvas(buf, pagesize=landscape(A4))
        margin = 24
        gap = 16
        card_w = (page_w - margin * 2 - gap) / 2
        card_h = 270
        x_positions = [margin, margin + card_w + gap]
        y = page_h - margin - card_h
        idx = 0

        def fmt_num(value, precision=2):
            if value is None:
                return "—"
            try:
                return f"{float(value):,.{precision}f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                return str(value)

        def fmt_date(value):
            try:
                year, month, day = str(value).split("-")[:3]
                return f"{day}/{month}/{year}"
            except Exception:
                return str(value or "—")

        def draw_value_box(x, y0, width, height, title, unit, value, head_color, body_color=white, value_color=black, small=False):
            head_h = 28 if not small else 26
            canvas_obj.setFillColor(head_color)
            canvas_obj.rect(x, y0 + height - head_h, width, head_h, fill=1, stroke=0)
            canvas_obj.setFillColor(body_color)
            canvas_obj.rect(x, y0, width, height - head_h, fill=1, stroke=1)
            canvas_obj.setFillColor(white)
            canvas_obj.setFont("Helvetica-Bold", 9 if small else 11)
            canvas_obj.drawCentredString(x + width / 2, y0 + height - 17, title)
            if unit:
                canvas_obj.setFont("Helvetica-Bold", 7 if small else 9)
                canvas_obj.drawCentredString(x + width / 2, y0 + height - 27, unit)
            canvas_obj.setFillColor(value_color)
            canvas_obj.setFont("Helvetica-Bold", 13 if small else 17)
            canvas_obj.drawCentredString(x + width / 2, y0 + 10, value)

        for card in cards:
            if idx and idx % 4 == 0:
                canvas_obj.showPage()
                y = page_h - margin - card_h
            col = idx % 2
            if idx % 2 == 0 and idx % 4 != 0:
                y -= card_h + gap
            x = x_positions[col]
            canvas_obj.setFillColor(HexColor("#ffffff"))
            canvas_obj.rect(x, y, card_w, card_h, fill=1, stroke=1)
            canvas_obj.setFillColor(HexColor("#efefef"))
            canvas_obj.rect(x, y + card_h - 58, card_w, 58, fill=1, stroke=1)
            canvas_obj.setFillColor(HexColor("#000000"))
            canvas_obj.setFont("Helvetica-Bold", 18)
            canvas_obj.drawCentredString(x + card_w / 2, y + card_h - 24, card["title"][:38].upper())
            canvas_obj.setFont("Helvetica-Bold", 15)
            canvas_obj.drawCentredString(x + card_w / 2, y + card_h - 46, fmt_date(card["production_date"]))
            canvas_obj.setFillColor(HexColor("#334155"))
            canvas_obj.setFont("Helvetica-Bold", 9)
            canvas_obj.drawCentredString(
                x + card_w / 2,
                y + card_h - 57,
                " · ".join([part for part in [card.get("bank"), card.get("card_type"), card.get("source"), card.get("loop")] if part]),
            )

            sx = x
            sy = y + card_h - 104
            section_w = card_w
            canvas_obj.setFillColor(HexColor("#5b8fd1"))
            canvas_obj.rect(sx, sy, section_w, 24, fill=1, stroke=1)
            canvas_obj.setFillColor(white)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawCentredString(sx + section_w / 2, sy + 6, "VOLUME")
            vw = section_w / 3
            draw_value_box(sx, sy - 56, vw, 56, "Óleo", "[Sm³]", fmt_num(card["volumes"].get("oil_sm3")), HexColor("#2f5f9f"))
            draw_value_box(sx + vw, sy - 56, vw, 56, "Gás", "[Mil Sm³]", fmt_num(card["volumes"].get("gas_msm3")), HexColor("#2f5f9f"))
            draw_value_box(sx + vw * 2, sy - 56, vw, 56, "Água", "[Sm³]", fmt_num(card["volumes"].get("water_sm3")), HexColor("#2f5f9f"))

            my = sy - 92
            canvas_obj.setFillColor(HexColor("#dfb1b1"))
            canvas_obj.rect(sx, my, section_w, 24, fill=1, stroke=1)
            canvas_obj.setFillColor(black)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawCentredString(sx + section_w / 2, my + 6, "MASSA")
            draw_value_box(sx, my - 56, vw, 56, "Óleo", "[t]", fmt_num(card["masses"].get("oil_t")), HexColor("#a63a3a"))
            draw_value_box(sx + vw, my - 56, vw, 56, "Gás", "[t]", fmt_num(card["masses"].get("gas_t")), HexColor("#a63a3a"))
            draw_value_box(sx + vw * 2, my - 56, vw, 56, "Água", "[t]", fmt_num(card["masses"].get("water_t")), HexColor("#a63a3a"))

            cy = my - 104
            left_w = section_w * 0.66
            right_w = section_w - left_w
            canvas_obj.setFillColor(HexColor("#e5edf5"))
            canvas_obj.rect(sx, cy, left_w, 24, fill=1, stroke=1)
            canvas_obj.setFillColor(black)
            canvas_obj.setFont("Helvetica-Bold", 12)
            canvas_obj.drawCentredString(sx + left_w / 2, cy + 6, "VARIÁVEIS DE CONTROLE")
            canvas_obj.setFillColor(HexColor("#d40000"))
            canvas_obj.rect(sx + left_w, cy, right_w, 24, fill=1, stroke=1)
            canvas_obj.setFillColor(white)
            canvas_obj.drawCentredString(sx + left_w + right_w / 2, cy + 6, "BALANÇO")
            hw = left_w / 2
            draw_value_box(sx, cy - 48, hw, 48, "Velocidade Escoamento", "[m/s]", fmt_num(card["control"].get("flow_velocity_ms")), HexColor("#1774ba"), small=True)
            draw_value_box(sx + hw, cy - 48, hw, 48, "dP", "[ΔP]", fmt_num(card["control"].get("dp_value")), HexColor("#1774ba"), small=True)
            draw_value_box(sx, cy - 96, hw, 48, "Pressão", "[barg]", fmt_num(card["control"].get("pressure_barg")), HexColor("#ef7d00"), small=True)
            draw_value_box(sx + hw, cy - 96, hw, 48, "Temperatura", "[°C]", fmt_num(card["control"].get("temperature_c")), HexColor("#ef7d00"), small=True)

            def bal_color(value):
                if value is None:
                    return HexColor("#e5e7eb")
                return HexColor("#d9ead3") if float(value) <= 7 else HexColor("#f4cccc")

            def bal_text(value):
                return "—" if value is None else f"{round(value):.0f}%".replace(".", ",")

            canvas_obj.setFillColor(HexColor("#d40000"))
            canvas_obj.rect(sx + left_w, cy - 48, right_w, 48, fill=1, stroke=1)
            canvas_obj.setFillColor(white)
            canvas_obj.setFont("Helvetica-Bold", 12)
            canvas_obj.drawCentredString(sx + left_w + right_w / 2, cy - 18, "% HC")
            canvas_obj.setFillColor(bal_color(card["balance"].get("hc_pct")))
            canvas_obj.rect(sx + left_w, cy - 96, right_w, 48, fill=1, stroke=1)
            canvas_obj.setFillColor(black)
            canvas_obj.setFont("Helvetica-Bold", 18)
            canvas_obj.drawCentredString(sx + left_w + right_w / 2, cy - 66, bal_text(card["balance"].get("hc_pct")))
            canvas_obj.setFillColor(HexColor("#d40000"))
            canvas_obj.rect(sx + left_w, cy - 144, right_w, 48, fill=1, stroke=1)
            canvas_obj.setFillColor(white)
            canvas_obj.setFont("Helvetica-Bold", 12)
            canvas_obj.drawCentredString(sx + left_w + right_w / 2, cy - 114, "% Total")
            canvas_obj.setFillColor(bal_color(card["balance"].get("total_pct")))
            canvas_obj.rect(sx + left_w, cy - 192, right_w, 48, fill=1, stroke=1)
            canvas_obj.setFillColor(black)
            canvas_obj.setFont("Helvetica-Bold", 18)
            canvas_obj.drawCentredString(sx + left_w + right_w / 2, cy - 162, bal_text(card["balance"].get("total_pct")))

            fy = y + 8
            fw = section_w / 3
            footer_items = [
                ("MPFM X FISCAL", bal_text(card["balance"].get("mpfm_x_fiscal_pct"))),
                ("BALANÇO GÁS", bal_text(card["balance"].get("balanco_gas_pct"))),
                ("OBSERVAÇÕES", (card.get("observations") or "—")[:26]),
            ]
            for i, (label, value) in enumerate(footer_items):
                canvas_obj.setFillColor(HexColor("#9f9560"))
                canvas_obj.rect(sx + i * fw, fy + 24, fw, 28, fill=1, stroke=1)
                canvas_obj.setFillColor(white)
                canvas_obj.setFont("Helvetica-Bold", 11)
                canvas_obj.drawCentredString(sx + i * fw + fw / 2, fy + 34, label)
                canvas_obj.setFillColor(HexColor("#ffffff"))
                canvas_obj.rect(sx + i * fw, fy, fw, 24, fill=1, stroke=1)
                canvas_obj.setFillColor(black)
                canvas_obj.setFont("Helvetica-Bold", 12)
                canvas_obj.drawCentredString(sx + i * fw + fw / 2, fy + 7, value)
            idx += 1

        canvas_obj.save()
        buf.seek(0)
        fname = f"cards_{date_from}_{date_to}.pdf"
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )

    @app.get("/api/export-producao-excel")
    def api_export_producao_excel(
        date_from: str = "",
        date_to: str = "",
        bank: str = "",
        include_daily: int = 1,
        include_hourly: int = 1,
        include_sep_oil: int = 1,
        include_sep_gas: int = 1,
        include_sep_water: int = 1,
        include_cards: int = 0,
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        import openpyxl

        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to

        def fetch_pivot(row_kind: str):
            sql = (
                "SELECT day_ref, hour_ref, bank, loop, tipo, tag, metric_name, metric_value "
                "FROM measurements_active WHERE day_ref BETWEEN ? AND ? AND row_kind=? "
                "AND COALESCE(is_official,1)=1"
            )
            params = [date_from, date_to, row_kind]
            if bank:
                sql += " AND bank=?"
                params.append(bank)
            sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), bank, tag, metric_name"
            rows = cur.execute(sql, params).fetchall()
            piv = defaultdict(dict)
            metrics = []
            for row in rows:
                key = (row["day_ref"], row["hour_ref"], row["bank"], row["loop"], row["tipo"], row["tag"])
                piv[key][row["metric_name"]] = row["metric_value"]
                if row["metric_name"] not in metrics:
                    metrics.append(row["metric_name"])
            return piv, metrics

        daily_piv, daily_metrics = fetch_pivot("daily")
        hourly_piv, hourly_metrics = fetch_pivot("hourly")
        template_path = PRODUCTION_EXPORT_TEMPLATE
        wb = openpyxl.load_workbook(template_path) if template_path.exists() else openpyxl.Workbook()
        subtitle = f"Período: {date_from} até {date_to}" + (f" · Banco: {bank}" if bank else " · Todos os bancos")

        def template_headers(ws):
            return [ws.cell(4, ci).value for ci in range(1, ws.max_column + 1)]

        def clear_template_data(ws, start_row: int = 5):
            for row_idx in range(start_row, ws.max_row + 1):
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row_idx, col_idx).value = None

        def clone_template_row(ws, template_row: int, target_row: int):
            for col_idx in range(1, ws.max_column + 1):
                copy_cell_style(ws.cell(template_row, col_idx), ws.cell(target_row, col_idx))
            ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height

        def set_sheet_subtitle(ws):
            ws["A2"] = subtitle
            center_cell_content(ws["A2"])

        def write_template_rows(ws, row_dicts):
            headers = template_headers(ws)
            clear_template_data(ws)
            start_row = 5
            style_even_row = start_row
            style_odd_row = min(start_row + 1, max(ws.max_row, start_row))
            for row_idx, row_dict in enumerate(row_dicts, start=start_row):
                if row_idx > ws.max_row:
                    style_row = style_even_row if (row_idx - start_row) % 2 == 0 else style_odd_row
                    clone_template_row(ws, style_row, row_idx)
                for col_idx, header in enumerate(headers, start=1):
                    value = row_dict.get(header, "") if header else ""
                    cell = ws.cell(row_idx, col_idx, value)
                    if value not in (None, ""):
                        center_cell_content(cell)
            center_filled_cells(ws)

        def pivot_rows(piv, include_hour: bool):
            items = []
            for key in sorted(piv.keys(), key=lambda item: (item[0] or "", -1 if item[1] is None else int(item[1]), item[2] or "", item[5] or "")):
                day_ref, hour_ref, bank_name, loop, tipo, tag_name = key
                metric_dict = piv[key]
                row = {
                    "Data": day_ref,
                    "Banco": bank_name,
                    "Loop": loop,
                    "Tipo": tipo,
                    "TAG": tag_name,
                }
                if include_hour:
                    row["Hora"] = "DAY" if hour_ref is None else f"{int(hour_ref):02d}:00"
                row.update(metric_dict)
                items.append(row)
            return items

        ws_daily = wb["DIARIOS"] if "DIARIOS" in wb.sheetnames else wb.create_sheet("DIARIOS")
        set_sheet_subtitle(ws_daily)

        # ── Computed columns for DIARIOS ─────────────────────────────────────
        _SUBSEA_TO_TOPSIDE = {
            'PE_4':     'Riser_P5',
            'PE_2':     'Riser_P2',
            'PW-104DA': 'Riser_P4',
        }
        _CHOKE_PI = {
            'PE_4':     r'\\AFBRA\BRA\Performance Monitoring\ProdWellsSUB\PE_4|Choke|Measured Pos (%)',
            'PE_2':     r'\\AFBRA\BRA\Performance Monitoring\ProdWellsSUB\PE_2|Choke|Measured Pos (%)',
            'PW-104DA': r'\\AFBRA\BRA\Performance Monitoring\ProdWellsSUB\PW_104DA|Choke|Measured Pos (%)',
        }
        _POB_THRESHOLD = 490

        daily_row_items = pivot_rows(daily_piv, include_hour=False) if include_daily else []

        # First pass: build topside HC/Total lookup from the fetched rows
        _top_hc  = {}   # {(day_ref, topside_tag): value}
        _top_tot = {}
        for _it in daily_row_items:
            if _it['TAG'] in _SUBSEA_TO_TOPSIDE.values():
                _k = (_it['Data'], _it['TAG'])
                _top_hc[_k]  = _it.get('MPFM corr HC (t)')
                _top_tot[_k] = _it.get('MPFM corr Total (t)')

        # Second pass: compute and attach 4 new fields
        for _it in daily_row_items:
            _tag = _it['TAG']
            _unc = _it.get('MPFM uncorr Total (t)') or 0

            # 1. Modo de Operação
            if not _unc:
                _it['Modo de Operação'] = None
            else:
                _it['Modo de Operação'] = (
                    'Operou Acima Pob'
                    if (_it.get('Pressão (barg)') or 0) > _POB_THRESHOLD
                    else 'Operou Abaixo PoB'
                )

            # 2. % Desvio HC & 3. % Desvio Total (subsea pairs only)
            _top_tag = _SUBSEA_TO_TOPSIDE.get(_tag)
            if _top_tag:
                _shc  = _it.get('MPFM corr HC (t)') or 0
                _thc  = _top_hc.get((_it['Data'], _top_tag)) or 0
                _it['% Desvio HC'] = (
                    round(_shc / _thc - 1, 6) if _shc and _thc else None
                )
                _stot = _it.get('MPFM corr Total (t)') or 0
                _ttot = _top_tot.get((_it['Data'], _top_tag)) or 0
                _it['% Desvio Total'] = (
                    round(_stot / _ttot - 1, 6) if _stot and _ttot else None
                )
            else:
                _it['% Desvio HC']    = None
                _it['% Desvio Total'] = None

            # 4. Choke % — sentinel; replaced with PITimeDat formula in post-pass
            _it['Choke %'] = '__CHOKE__' if (_tag in _CHOKE_PI and _unc) else None

        # Ensure 4 new headers exist in row 4 of DIARIOS before writing
        _new_col_names = ['Modo de Operação', '% Desvio HC', '% Desvio Total', 'Choke %']
        _existing_hdr  = [ws_daily.cell(4, ci).value for ci in range(1, ws_daily.max_column + 1)]
        for _col_name in _new_col_names:
            if _col_name not in _existing_hdr:
                ws_daily.cell(4, ws_daily.max_column + 1, _col_name)
                _existing_hdr.append(_col_name)

        write_template_rows(ws_daily, daily_row_items)

        # Post-pass: replace sentinel with PITimeDat Excel formula
        _hdr_row4   = [ws_daily.cell(4, ci).value for ci in range(1, ws_daily.max_column + 1)]
        _choke_col  = next((i + 1 for i, h in enumerate(_hdr_row4) if h == 'Choke %'), None)
        _data_col   = next((i + 1 for i, h in enumerate(_hdr_row4) if h == 'Data'), None)
        _tag_col    = next((i + 1 for i, h in enumerate(_hdr_row4) if h == 'TAG'), None)
        if _choke_col and _data_col and _tag_col:
            _data_letter = openpyxl.utils.get_column_letter(_data_col)
            for _ri in range(5, ws_daily.max_row + 1):
                _cell = ws_daily.cell(_ri, _choke_col)
                if _cell.value == '__CHOKE__':
                    _row_tag  = ws_daily.cell(_ri, _tag_col).value
                    _pi_path  = _CHOKE_PI.get(_row_tag)
                    if _pi_path:
                        _cell.value = (
                            f'=PITimeDat("{_pi_path}",{_data_letter}{_ri}+TIME(12,0,0),"","interpolated")'
                        )
                    else:
                        _cell.value = None
        # ─────────────────────────────────────────────────────────────────────

        ws_hourly = wb["HORARIOS"] if "HORARIOS" in wb.sheetnames else wb.create_sheet("HORARIOS")
        set_sheet_subtitle(ws_hourly)
        write_template_rows(ws_hourly, pivot_rows(hourly_piv, include_hour=True) if include_hourly else [])

        fluid_flags = {"oleo": include_sep_oil, "gas": include_sep_gas, "agua": include_sep_water}
        for fluid, title in [("oleo", "SEPARADOR_OLEO"), ("gas", "SEPARADOR_GAS"), ("agua", "SEPARADOR_AGUA")]:
            kind = sep_detail_kind(fluid)
            params = (kind, date_from, date_to) if not bank else (kind, date_from, date_to, bank)
            sql = (
                "SELECT day_ref, hour_ref, tag, instrument, metric_name, metric_value, bank "
                "FROM measurements_active WHERE row_kind=? AND COALESCE(is_official,1)=1 AND day_ref BETWEEN ? AND ?"
            )
            if bank:
                sql += " AND (bank=? OR bank='SEP')"
            sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), tag, metric_name"
            rows = [dict(row) for row in cur.execute(sql, params).fetchall()]
            ws = wb[title] if title in wb.sheetnames else wb.create_sheet(title)
            set_sheet_subtitle(ws)
            piv = defaultdict(dict)
            for row in rows:
                key = (row["day_ref"], row["hour_ref"], row["tag"], row["instrument"], row.get("bank") or "SEP")
                piv[key][row["metric_name"]] = row["metric_value"]
            sep_rows = []
            if fluid_flags[fluid]:
                for key in sorted(piv.keys(), key=lambda item: (item[0] or "", -1 if item[1] is None else int(item[1]), item[2] or "")):
                    day_ref, hour_ref, tag_name, instrument, bank_name = key
                    row = {
                        "Data": day_ref,
                        "Hora": "DAY" if hour_ref is None else int(hour_ref),
                        "TAG": tag_name,
                        "Meter ID": instrument,
                        "Banco": bank_name or "SEP",
                    }
                    row.update(piv[key])
                    sep_rows.append(row)
            write_template_rows(ws, sep_rows)

        if "CARDS_RESUMO" in wb.sheetnames:
            del wb["CARDS_RESUMO"]
        for ws in wb.worksheets:
            center_filled_cells(ws)
        conn.close()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"dados_producao_{date_from}_{date_to}.xlsx"
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/export-excel")
    def api_export_excel(
        date_from: str = "",
        date_to: str = "",
        row_kind: str = "daily",
        bank: str = "",
        metrics: str = "",
        tag: str = "",
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        prefs = load_prefs()
        selected = [metric.strip() for metric in metrics.split(",")] if metrics else prefs.get("selected_metrics", [])

        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to

        sql = (
            "SELECT day_ref, hour_ref, bank, loop, tipo, tag, metric_name, metric_value "
            "FROM measurements_active WHERE day_ref BETWEEN ? AND ?"
        )
        params = [date_from, date_to]
        if row_kind in ("hourly", "daily", "recon", "sep"):
            sql += " AND row_kind=?"
            params.append(row_kind)
        if bank:
            sql += " AND bank=?"
            params.append(bank)
        if tag:
            sql += " AND tag=?"
            params.append(tag)
        if selected:
            placeholders = ",".join("?" * len(selected))
            sql += f" AND metric_name IN ({placeholders})"
            params.extend(selected)
        sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), bank, tag, metric_name"
        rows = cur.execute(sql, params).fetchall()
        conn.close()

        pivot = defaultdict(dict)
        for row in rows:
            key = (row[0], row[1], row[2], row[3], row[4], row[5])
            pivot[key][row[6]] = row[7]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Exportação {date_from}"

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        alt_fill = PatternFill("solid", fgColor="EEF2FF")

        headers = ["Data", "Hora", "Banco", "Loop", "Tipo", "TAG"] + selected
        for ci, header in enumerate(headers, 1):
            cell = ws.cell(1, ci, header)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        for ri, (key, metric_dict) in enumerate(sorted(pivot.items()), 2):
            day, hour, bank_name, loop, tipo, tag_name = key
            base = [day, "" if hour is None else f"{int(hour):02d}:00", bank_name, loop, tipo, tag_name]
            fill = alt_fill if ri % 2 == 0 else None
            for ci, value in enumerate(base, 1):
                cell = ws.cell(ri, ci, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if fill:
                    cell.fill = fill
            for ci, metric in enumerate(selected, len(base) + 1):
                value = metric_dict.get(metric)
                cell = ws.cell(ri, ci, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="right" if value is not None else "left", vertical="center")
                if fill and value is not None:
                    cell.fill = fill

        for ci in range(1, len(headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            values = [str(ws.cell(row, ci).value or "") for row in range(1, min(ws.max_row + 1, 200))]
            ws.column_dimensions[col_letter].width = min(max((len(item) for item in values), default=8) + 3, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"mpfm_export_{date_from}_{date_to}.xlsx"
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.get("/api/export-csv")
    def api_export_csv(
        date_from: str = "",
        date_to: str = "",
        row_kind: str = "daily",
        bank: str = "",
        metrics: str = "",
        tag: str = "",
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        import csv

        prefs = load_prefs()
        selected = [metric.strip() for metric in metrics.split(",")] if metrics else prefs.get("selected_metrics", [])
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to
        sql = (
            "SELECT day_ref, hour_ref, bank, loop, tipo, tag, metric_name, metric_value "
            "FROM measurements_active WHERE day_ref BETWEEN ? AND ?"
        )
        params = [date_from, date_to]
        if row_kind in ("hourly", "daily", "recon"):
            sql += " AND row_kind=?"
            params.append(row_kind)
        if bank:
            sql += " AND bank=?"
            params.append(bank)
        if tag:
            sql += " AND tag=?"
            params.append(tag)
        if selected:
            placeholders = ",".join("?" * len(selected))
            sql += f" AND metric_name IN ({placeholders})"
            params.extend(selected)
        sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), bank, tag, metric_name"
        rows = cur.execute(sql, params).fetchall()
        conn.close()

        pivot = defaultdict(dict)
        for row in rows:
            key = (row[0], row[1], row[2], row[3], row[4], row[5])
            pivot[key][row[6]] = row[7]

        output = io.StringIO()
        writer = csv.writer(output)
        headers = ["Data", "Hora", "Banco", "Loop", "Tipo", "TAG"] + selected
        writer.writerow(headers)
        for (day, hour, bank_name, loop, tipo, tag_name), metric_dict in sorted(pivot.items()):
            row = [day, "" if hour is None else f"{int(hour):02d}:00", bank_name, loop, tipo, tag_name]
            row += [metric_dict.get(metric, "") for metric in selected]
            writer.writerow(row)
        output.seek(0)
        fname = f"mpfm_export_{date_from}_{date_to}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.get("/api/recon/export-excel/{run_id}")
    def api_recon_export_excel(run_id: int):
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        conn = db_conn()
        row = conn.execute("SELECT * FROM recon_runs WHERE id=?", (run_id,)).fetchone()
        conn.close()
        if not row:
            raise HTTPException(404, f"Run {run_id} não encontrada")

        data = dict(row)
        sep_h = json.loads(data.get("sep_hourly_json") or "[]")
        mpfm_h = json.loads(data.get("mpfm_hourly_json") or "[]")
        calc_h = json.loads(data.get("calc_hourly_json") or "[]")
        resumo = json.loads(data.get("resumo_json") or "{}")
        pvt = json.loads(data.get("pvt_snapshot") or "{}")
        analytical = json.loads(data.get("analytical_snapshot") or "{}")

        wb = openpyxl.Workbook()
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="4472C4")
        alt_fill = PatternFill("solid", fgColor="EEF2FF")
        warn_fill = PatternFill("solid", fgColor="FFF3CD")
        ok_fill = PatternFill("solid", fgColor="D4EDDA")
        err_fill = PatternFill("solid", fgColor="F8D7DA")

        def _hdr(ws, headers, row=1):
            for ci, header in enumerate(headers, 1):
                cell = ws.cell(row, ci, header)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[row].height = 28
            ws.freeze_panes = f"A{row + 1}"

        def _cell(ws, row_idx, col_idx, value, fmt=None, fill=None):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left", vertical="center")
            if fmt:
                cell.number_format = fmt
            if fill:
                cell.fill = fill
            return cell

        def _auto_width(ws, max_r=50):
            for col in ws.columns:
                width = max((len(str(cell.value or "")) for cell in list(col)[:max_r]), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(width + 3, 45)

        ws = wb.active
        ws.title = "README"
        readme_lines = [
            "Reconciliação 24h MPFM × Separador",
            "",
            f"Banco/TAG:     {data['bank']} / {data['tag']}",
            f"Dia ref.:      {data['day_ref']}",
            f"Executado em:  {data['run_at']}",
            f"Autor:         {data.get('author', '')}",
            "",
            "• A água medida do separador tem prioridade sobre BSW.",
            "• FE e RS vêm da base PVT e ficam congelados no snapshot da campanha.",
            "• O separador horário usa a leitura detalhada do teste quando existir; sem isso, cai em compatibilidade para a camada SEP resumida.",
            "• KPIs principais: Massa HC e Massa Total (linha vs linha).",
            "• Gás reconstruído por RS é trilha standard/diagnóstica.",
            f"• Condição de referência: {pvt.get('temp_ref_c', 20)} °C / {pvt.get('pres_ref_bar', 1.01325)} bar(a).",
            "",
            f"GOR mode:      {pvt.get('gor_mode', 'unknown')}",
            f"GSV confirmado:{pvt.get('gsv_confirmed', False)}",
            f"BSW campanha:  {analytical.get('bsw_pct', '')}",
            f"Fonte densid.: {analytical.get('density_source', '')}",
            f"Notas:         {data.get('notes', '')}",
        ]
        for i, line in enumerate(readme_lines, 1):
            ws.cell(i, 1, line).font = Font(size=11)
        ws.column_dimensions["A"].width = 70

        ws = wb.create_sheet("Parametros")
        _hdr(ws, ["Parâmetro", "Valor", "Unidade", "Observação"])
        params_rows = [
            ("Temperatura de referência", pvt.get("temp_ref_c", 20), "°C", "Condição de referência"),
            ("Pressão de referência", pvt.get("pres_ref_bar", 1.01325), "bar(a)", "Condição de referência"),
            ("FE - fator de encolhimento", pvt.get("fe"), "-", "Parâmetro PVT externo"),
            ("RS - razão de solubilidade", pvt.get("rs"), "Sm³/Sm³", "Parâmetro PVT externo"),
            ("Densidade óleo standard", pvt.get("rho_oleo_std"), "kg/m³", "Parâmetro PVT externo"),
            ("Densidade gás standard", pvt.get("rho_gas_std"), "kg/m³", "Parâmetro PVT externo"),
            ("Densidade água standard", pvt.get("rho_agua_std"), "kg/m³", "Parâmetro PVT externo"),
            ("GOR mode", pvt.get("gor_mode", "unknown"), "-", "fixed/zero/triphasic/unknown"),
            ("GSV óleo confirmado", str(pvt.get("gsv_confirmed", False)), "-", "Subtração água permitida"),
            ("Limite desvio HC - linha", pvt.get("limite_hc_pct", 5), "%", "Critério KPI principal"),
            ("Limite desvio Total - linha", pvt.get("limite_total_pct", 5), "%", "Critério KPI principal"),
            ("Limite alerta Água - linha", pvt.get("limite_agua_pct", 20), "%", "KPI secundário/QA"),
        ]
        for ri, (param_name, value, unit_name, obs) in enumerate(params_rows, 2):
            _cell(ws, ri, 1, param_name)
            _cell(ws, ri, 2, value, "#,##0.0000" if isinstance(value, float) else None)
            _cell(ws, ri, 3, unit_name)
            _cell(ws, ri, 4, obs)
        _auto_width(ws)

        ws = wb.create_sheet("Analitica")
        _hdr(ws, ["Item", "Valor", "Unidade", "Observação"])
        analytical_rows = [
            ("BSW campanha", analytical.get("bsw_pct"), "%", "Valor analítico associado ao run"),
            ("Densidade coriolis", analytical.get("density_coriolis_kg_m3"), "kg/m³", "Leitura operacional do coriolis"),
            ("Densidade laboratório", analytical.get("density_lab_kg_m3"), "kg/m³", "Valor de laboratório"),
            ("Outra densidade", analytical.get("density_other_kg_m3"), "kg/m³", "Fonte complementar"),
            ("Fonte da densidade", analytical.get("density_source", ""), "-", "Origem declarada pelo usuário"),
            ("Referência analítica", analytical.get("analysis_reference", ""), "-", "Laudo, amostra ou relatório"),
            ("Notas analíticas", analytical.get("analysis_notes", ""), "-", "Observações da campanha"),
        ]
        for ri, (item_name, value, unit_name, obs) in enumerate(analytical_rows, 2):
            _cell(ws, ri, 1, item_name)
            _cell(ws, ri, 2, value, "#,##0.0000" if isinstance(value, float) else None)
            _cell(ws, ri, 3, unit_name)
            _cell(ws, ri, 4, obs)
        _auto_width(ws)

        ws = wb.create_sheet("Separador_Hora")
        _hdr(ws, ["Hora", "Data/Hora", "GSV_sep_Sm3", "Agua_sep_Sm3", "Gas_livre_sep_Sm3", "BSW_user_pct", "BSW_calc_pct"])
        sep_map = {item["hora"]: item for item in sep_h}
        calc_map = {item["hora"]: item for item in calc_h}
        for ri, hour in enumerate(sorted(sep_map.keys()), 2):
            sep_row = sep_map[hour]
            calc_row = calc_map.get(hour, {})
            fill = alt_fill if ri % 2 == 0 else None
            values = [
                hour,
                sep_row.get("dt_str", ""),
                sep_row.get("gsv_sep_sm3"),
                calc_row.get("agua_sep_sm3"),
                sep_row.get("gas_vol_sm3") or sep_row.get("gas_mass_t"),
                sep_row.get("bsw_user_pct"),
                calc_row.get("bsw_calc_pct"),
            ]
            for ci, value in enumerate(values, 1):
                _cell(ws, ri, ci, value, "#,##0.000" if isinstance(value, float) else None, fill)
        _auto_width(ws)

        ws = wb.create_sheet("MPFM_Hora")
        _hdr(
            ws,
            [
                "Hora",
                "Data/Hora",
                "MPFM_Oleo_corr_t",
                "MPFM_Gas_corr_t",
                "MPFM_Agua_corr_t",
                "MPFM_HC_corr_t",
                "MPFM_Total_corr_t",
                "Pressao_barg",
                "Temperatura_C",
                "MPFM_Oleo_ST_t",
                "MPFM_Gas_ST_t",
                "MPFM_Agua_ST_t",
                "MPFM_Oleo_ST_m3",
                "MPFM_Gas_ST_kSm3",
                "MPFM_Agua_ST_m3",
            ],
        )
        mpfm_map = {item["hora"]: item for item in mpfm_h}
        for ri, hour in enumerate(sorted(mpfm_map.keys()), 2):
            mpfm_row = mpfm_map[hour]
            fill = alt_fill if ri % 2 == 0 else None
            values = [
                hour,
                mpfm_row.get("dt_str", ""),
                mpfm_row.get("oleo_corr_t"),
                mpfm_row.get("gas_corr_t"),
                mpfm_row.get("agua_corr_t"),
                mpfm_row.get("hc_corr_t"),
                mpfm_row.get("total_corr_t"),
                mpfm_row.get("pressao_barg"),
                mpfm_row.get("temperatura_c"),
                mpfm_row.get("oleo_st_t"),
                mpfm_row.get("gas_st_ksm3"),
                mpfm_row.get("agua_st_m3"),
                mpfm_row.get("oleo_st_m3"),
                mpfm_row.get("gas_st_ksm3"),
                mpfm_row.get("agua_st_m3"),
            ]
            for ci, value in enumerate(values, 1):
                _cell(ws, ri, ci, value, "#,##0.000" if isinstance(value, float) else None, fill)
        _auto_width(ws)

        ws = wb.create_sheet("Calculo_Ref_Hora")
        calc_headers = [
            "Hora",
            "Data/Hora",
            "GSV_sep_Sm3",
            "Agua_sep_Sm3",
            "Gas_livre_sep_Sm3",
            "BSW_user_pct",
            "BSW_calc_pct",
            "Oleo_base_ref_Sm3",
            "FE",
            "RS",
            "Oleo_std_reconc_Sm3",
            "Gas_dissolvido_Sm3",
            "Gas_total_reconc_Sm3",
            "Rho_oleo_std",
            "Rho_gas_std",
            "Rho_agua_std",
            "Massa_oleo_ref_t",
            "Massa_gas_ref_t",
            "Massa_agua_ref_t",
            "Massa_HC_ref_t",
            "Massa_total_ref_t",
            "Desvio_HC_linha_%",
            "Desvio_Total_linha_%",
            "Desvio_Agua_linha_%",
            "Desvio_Oleo_ST_%",
            "Desvio_Gas_ST_%",
            "QA_gap_BSW_pp",
            "Flag_BSW",
            "Agua_fonte",
            "Gas_fonte",
            "Status_Linha",
            "Status_Standard",
            "Status_Final",
            "QA_flags",
        ]
        _hdr(ws, calc_headers)
        status_fill = {"OK": ok_fill, "ATENÇÃO": warn_fill, "VERIFICAR": err_fill}
        for ri, calc_row in enumerate(sorted(calc_h, key=lambda item: item.get("hora", 0)), 2):
            fill = alt_fill if ri % 2 == 0 else None
            row_fill = status_fill.get(calc_row.get("status_final", ""), fill)
            values = [
                calc_row.get("hora"),
                calc_row.get("dt_str", ""),
                calc_row.get("gsv_sep_sm3"),
                calc_row.get("agua_sep_sm3"),
                calc_row.get("gas_livre_sep_sm3"),
                calc_row.get("bsw_user_pct"),
                calc_row.get("bsw_calc_pct"),
                calc_row.get("oleo_base_ref_sm3"),
                calc_row.get("fe"),
                calc_row.get("rs"),
                calc_row.get("oleo_std_reconc_sm3"),
                calc_row.get("gas_dissolvido_sm3"),
                calc_row.get("gas_total_reconc_sm3"),
                calc_row.get("rho_oleo_std"),
                calc_row.get("rho_gas_std"),
                calc_row.get("rho_agua_std"),
                calc_row.get("massa_oleo_ref_t"),
                calc_row.get("massa_gas_ref_t"),
                calc_row.get("massa_agua_ref_t"),
                calc_row.get("massa_hc_ref_t"),
                calc_row.get("massa_total_ref_t"),
                calc_row.get("desvio_hc_linha_pct"),
                calc_row.get("desvio_total_linha_pct"),
                calc_row.get("desvio_agua_linha_pct"),
                calc_row.get("desvio_oleo_st_pct"),
                calc_row.get("desvio_gas_st_pct"),
                calc_row.get("qa_gap_bsw_pp"),
                calc_row.get("flag_bsw"),
                calc_row.get("agua_fonte"),
                calc_row.get("gas_fonte"),
                calc_row.get("status_linha"),
                calc_row.get("status_standard"),
                calc_row.get("status_final"),
                calc_row.get("qa_flags", ""),
            ]
            for ci, value in enumerate(values, 1):
                is_status = ci >= len(values) - 4
                _cell(ws, ri, ci, value, "#,##0.0000" if isinstance(value, float) else None, row_fill if is_status else fill)

        total_row = len(calc_h) + 3
        ws.cell(total_row, 1, "TOTAL 24h").font = Font(bold=True)
        for ci, value in [
            (17, resumo.get("massa_hc_ref_t")),
            (20, resumo.get("massa_hc_ref_t")),
            (21, resumo.get("massa_total_ref_t")),
            (22, resumo.get("desvio_hc_pct")),
            (23, resumo.get("desvio_total_pct")),
        ]:
            cell = ws.cell(total_row, ci, value)
            cell.font = Font(bold=True)
            cell.number_format = "#,##0.0000"
        ws.cell(
            total_row + 2,
            1,
            f"Nota: HC e Total (linha) são os KPIs principais. Gás ST com cautela (GOR mode: {pvt.get('gor_mode', '?')})",
        )
        _auto_width(ws)

        ws = wb.create_sheet("Resumo_24h")

        def _res_row(ws, row_idx, label, ref, mpfm, desvio, status, fill=None):
            values = [label, ref, mpfm, desvio, status]
            for ci, value in enumerate(values, 1):
                cell = ws.cell(row_idx, ci, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left", vertical="center")
                if isinstance(value, float):
                    cell.number_format = "#,##0.0000"
                if fill:
                    cell.fill = fill
                if ci == 5 and status in status_fill:
                    cell.fill = status_fill[status]
            return row_idx + 1

        row_idx = 1
        ws.cell(row_idx, 1, "Comparação linha vs linha").font = Font(bold=True, size=12)
        row_idx += 1
        _hdr(ws, ["Indicador", "Referência (SEP)", "MPFM", "Desvio (%)", "Status"], row=row_idx)
        row_idx += 1
        row_idx = _res_row(ws, row_idx, "Massa HC (t)", resumo.get("massa_hc_ref_t"), resumo.get("massa_hc_mpfm_t"), resumo.get("desvio_hc_pct"), resumo.get("status_hc", ""))
        row_idx = _res_row(ws, row_idx, "Massa Total (t)", resumo.get("massa_total_ref_t"), resumo.get("massa_total_mpfm_t"), resumo.get("desvio_total_pct"), resumo.get("status_total", ""))
        row_idx = _res_row(ws, row_idx, "Massa Água (t)", resumo.get("massa_agua_ref_t"), resumo.get("massa_agua_mpfm_t"), resumo.get("desvio_agua_pct"), resumo.get("status_agua", ""))

        summary_fill = status_fill.get(resumo.get("status_linha", ""), None)
        ws.cell(row_idx, 1, "Status consolidado linha").font = Font(bold=True)
        ws.cell(row_idx, 5, resumo.get("status_linha", "")).fill = summary_fill or PatternFill()
        row_idx += 2

        ws.cell(row_idx, 1, "Comparação standard vs standard").font = Font(bold=True, size=12)
        row_idx += 1
        _hdr(ws, ["Indicador", "Referência reconciliada", "MPFM ST", "Desvio (%)", "Status"], row=row_idx)
        row_idx += 1
        row_idx = _res_row(ws, row_idx, "Óleo ST (m³)", resumo.get("oleo_std_ref_sm3"), resumo.get("oleo_st_mpfm_m3"), resumo.get("desvio_oleo_st_pct"), resumo.get("status_oleo_st", ""))
        row_idx = _res_row(ws, row_idx, "Gás ST (Sm³)", resumo.get("gas_total_ref_sm3"), resumo.get("gas_st_mpfm_ksm3"), resumo.get("desvio_gas_st_pct"), resumo.get("status_gas_st", ""))
        row_idx = _res_row(ws, row_idx, "Água ST (m³)", resumo.get("agua_ref_sm3"), resumo.get("agua_st_mpfm_m3"), resumo.get("desvio_agua_st_pct"), resumo.get("status_agua_st", ""))
        ws.cell(row_idx, 1, "Status consolidado standard").font = Font(bold=True)
        ws.cell(row_idx, 5, resumo.get("status_standard", "")).fill = status_fill.get(resumo.get("status_standard", ""), PatternFill())
        row_idx += 2

        ws.cell(row_idx, 1, "QA / premissas").font = Font(bold=True, size=12)
        row_idx += 1
        _hdr(ws, ["Item", "Resultado", "Critério", "Status", "Comentário"], row=row_idx)
        row_idx += 1
        qa_rows = [
            (
                "Cobertura horária",
                f"{resumo.get('horas_validas', 0)}/24 ({resumo.get('cobertura_pct', 0):.1f}%)",
                "24/24 horas",
                "OK" if resumo.get("consolidado_completo") else "ATENÇÃO",
                "Consolidado marcado como incompleto" if not resumo.get("consolidado_completo") else "",
            ),
            (
                "Gap médio BSW (p.p.)",
                f"{resumo.get('qa_gap_bsw_medio_pp', '—')}",
                "<= 0,50 p.p.",
                resumo.get("flag_bsw_consolidado", "INDISPONÍVEL"),
                "BSW como fallback/QA, não base principal",
            ),
            (
                "GSV óleo - confirmação",
                "Confirmado" if pvt.get("gsv_confirmed") else "NÃO confirmado",
                "Confirmado para subtração de água",
                "OK" if pvt.get("gsv_confirmed") else "REVISAR",
                "" if pvt.get("gsv_confirmed") else "Oleo_base = GSV-Água bloqueado",
            ),
            (
                "GOR mode",
                pvt.get("gor_mode", "unknown"),
                "Confirmar operacional",
                "OK" if pvt.get("gor_mode") not in ("unknown", None, "") else "ATENÇÃO",
                "GOR fixed: gás ST é diagnóstico, não KPI de linha",
            ),
        ]
        for item, result_value, criterion, status, comment in qa_rows:
            for ci, value in enumerate([item, result_value, criterion, status, comment], 1):
                cell = ws.cell(row_idx, ci, value)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if ci == 4 and status in status_fill:
                    cell.fill = status_fill[status]
            row_idx += 1

        if resumo.get("qa_flags_consolidados"):
            row_idx += 1
            ws.cell(row_idx, 1, "Flags QA").font = Font(bold=True)
            ws.cell(row_idx, 2, resumo.get("qa_flags_consolidados", ""))

        _auto_width(ws)
        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = max(ws.column_dimensions[col].width, 22)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'Recon_{data["bank"]}_{data["tag"]}_{data["day_ref"]}.xlsx'
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
