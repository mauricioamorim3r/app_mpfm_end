from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
import os
import time
from pathlib import Path, PurePosixPath
import re
import shutil
import tempfile
import zipfile

from fastapi import File, HTTPException, Request, UploadFile
from routes.date_utils import normalize_date_input, normalize_date_range, normalize_validation_issue_day_ref
from services.ops import build_dashboard_months
from cache_manager import cached, _cache
from services.ops.monitoring_service import (
    HC_LIMIT_PCT,
    MONITORING_BOOL_FIELDS,
    TOTAL_LIMIT_PCT,
    delete_monitoring_annotation,
    invalidate_months_cache,
    list_monitoring_rows,
    normalize_meter_type,
    upsert_monitoring_annotation,
)


def register_ops_routes(app, ctx: dict) -> None:
    db_conn = ctx["db_conn"]
    load_prefs = ctx["load_prefs"]
    recon_deviation_metrics = ctx["recon_deviation_metrics"]
    month_calendar_status = ctx["month_calendar_status"]
    bbl_from_m3 = ctx["bbl_from_m3"]
    boe_from = ctx["boe_from"]
    build_backup_zip = ctx["build_backup_zip"]
    build_monthly_base_unica = ctx["build_monthly_base_unica"]
    clear_local_data = ctx["clear_local_data"]
    restart_local_data = ctx["restart_local_data"]
    cleanup_workbook = ctx["cleanup_workbook"]
    db_path = ctx["db_path"]
    init_db = ctx["init_db"]
    work_dir = ctx["work_dir"]
    output_dir = ctx["output_dir"]
    month_pt = ctx["month_pt"]
    excel_name = ctx["excel_name"]
    recompute_alignment_resolution = ctx["recompute_alignment_resolution"]
    recompute_card_resolution = ctx["recompute_card_resolution"]
    recompute_sep_source_resolution = ctx["recompute_sep_source_resolution"]
    rebuild_sep_summary_for_day = ctx["rebuild_sep_summary_for_day"]
    rebuild_validation_snapshot_for_month = ctx["rebuild_validation_snapshot_for_month"]
    sanitize_files_imported_history = ctx["sanitize_files_imported_history"]
    delete_all_data_for_day = ctx["delete_all_data_for_day"]
    preview_base_unica_import = ctx["preview_base_unica_import"]
    apply_base_unica_import = ctx["apply_base_unica_import"]
    schedule_monthly_base_unica = ctx["schedule_monthly_base_unica"]
    load_cadastro = ctx["load_cadastro"]
    normalize_tag_name = ctx["normalize_tag_name"]
    load_state = ctx["load_state"]
    save_state = ctx["save_state"]
    load_sep_data_by_day = ctx["load_sep_data_by_day"]
    load_sep_data_by_range = ctx.get("load_sep_data_by_range", load_sep_data_by_day)
    serialize_sep_row = ctx["serialize_sep_row"]

    def _as_float(value, default):
        try:
            parsed = float(value)
            return parsed if parsed == parsed else default
        except Exception:
            return default

    def _summary_conversion_settings():
        prefs = load_prefs() or {}
        raw = prefs.get("summary_conversions") or {}
        show_criterion = raw.get("show_boe_criterion", True)
        if isinstance(show_criterion, str):
            show_criterion = show_criterion.strip().lower() not in {"0", "false", "off", "no"}
        return {
            "oil_m3_to_bbl_factor": _as_float(raw.get("oil_m3_to_bbl_factor"), 6.28981),
            "gas_sm3_per_boe_factor": max(_as_float(raw.get("gas_sm3_per_boe_factor"), 170.0), 0.000001),
            "gas_input_unit": str(raw.get("gas_input_unit") or "Sm³"),
            "gas_boe_mode": str(raw.get("gas_boe_mode") or "Padrão corporativo"),
            "show_boe_criterion": bool(show_criterion),
        }

    def _invalidate_cache(pattern: str | None = None) -> dict:
        """Invalida entradas do cache. Sem padrão, limpa tudo."""
        removed = _cache.invalidate(pattern)
        return {"ok": True, "removed": removed}

    def _parse_target_month(month: str):
        raw = str(month or "").strip()
        if not re.fullmatch(r"\d{4}-\d{2}", raw):
            raise HTTPException(status_code=400, detail="Mês inválido. Use o formato YYYY-MM.")
        yr, mo = raw.split("-")
        if int(mo) < 1 or int(mo) > 12:
            raise HTTPException(status_code=400, detail="Mês inválido. Use o formato YYYY-MM.")
        return yr, mo, f"{yr}-{mo}-01", f"{yr}-{mo}-31"

    def _safe_scalar(cur, sql: str, params=(), default=0):
        try:
            row = cur.execute(sql, params).fetchone()
            return row[0] if row and row[0] is not None else default
        except Exception:
            return default

    def _persist_uploaded_file(upload: UploadFile, default_name: str) -> str:
        suffix = Path(upload.filename or default_name).suffix or Path(default_name).suffix or ".bin"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            while True:
                chunk = upload.file.read(1024 * 1024)
                if not chunk:
                    break
                tmp.write(chunk)
            return tmp.name

    def _persist_uploaded_workbook(upload: UploadFile) -> str:
        return _persist_uploaded_file(upload, "base_unica.xlsx")

    def _backup_restore_target(member_name: str) -> Path | None:
        pure = PurePosixPath(member_name)
        parts = pure.parts
        if not parts or pure.is_absolute() or ".." in parts:
            return None
        if len(parts) == 1:
            name = parts[0]
            if name == Path(db_path).name:
                return Path(db_path)
            if name == "cadastro.json":
                return work_dir / name
            if re.fullmatch(r"state_\d{4}_\d{2}\.json", name):
                return work_dir / name
            return None
        if len(parts) == 2 and parts[0] == "outputs" and parts[1].lower().endswith(".xlsx"):
            return output_dir / parts[1]
        return None

    def _inspect_backup_zip(zip_path: Path) -> dict:
        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                members = zf.namelist()
                bad_entry = zf.testzip()
        except zipfile.BadZipFile as exc:
            raise HTTPException(status_code=400, detail=f"Backup ZIP inválido: {exc}") from exc
        allowed = []
        rejected = []
        for name in members:
            target = _backup_restore_target(name)
            if target is None:
                rejected.append(name)
            else:
                allowed.append({"name": name, "target": str(target)})
        if bad_entry:
            raise HTTPException(status_code=400, detail=f"Backup ZIP corrompido na entrada: {bad_entry}")
        if rejected:
            raise HTTPException(
                status_code=400,
                detail=f"Backup ZIP contém itens não suportados: {', '.join(rejected[:5])}",
            )
        if not any(item["name"] == Path(db_path).name for item in allowed):
            raise HTTPException(status_code=400, detail="Backup ZIP sem arquivo principal do banco local.")
        return {"members": members, "allowed": allowed, "bad_entry": bad_entry}

    def _restore_backup_zip(zip_path: Path):
        inspected = _inspect_backup_zip(zip_path)
        safety_backup = build_backup_zip()
        restart_local_data(keep_backup_zip=True)
        restored = []
        with zipfile.ZipFile(zip_path, "r") as zf:
            for item in inspected["allowed"]:
                target = Path(item["target"])
                target.parent.mkdir(parents=True, exist_ok=True)
                if target.exists():
                    if target.is_file():
                        target.unlink()
                    else:
                        shutil.rmtree(target)
                with zf.open(item["name"], "r") as src, open(target, "wb") as dst:
                    shutil.copyfileobj(src, dst)
                restored.append(str(target))
        init_db()
        return {
            "ok": True,
            "backup_file": Path(zip_path).name,
            "safety_backup_file": safety_backup.name,
            "restored_items": restored,
            "db_path": str(db_path),
            "message": "Backup restaurado com sucesso na aplicação local.",
        }

    def _recovery_result_lines(payload: dict):
        lines = []
        for key, value in payload.items():
            if isinstance(value, dict):
                lines.append(f"{key}:")
                for sk, sv in value.items():
                    lines.append(f"  - {sk}: {sv}")
            else:
                lines.append(f"{key}: {value}")
        return "\n".join(lines)

    def _conversion_payload(oil_m3=None, gas_sm3=None, settings=None):
        settings = settings or _summary_conversion_settings()
        oil_value = _as_float(oil_m3, 0.0)
        gas_value = _as_float(gas_sm3, 0.0)
        oil_bbl = round(oil_value * settings["oil_m3_to_bbl_factor"], 3)
        gas_boe = round(gas_value / settings["gas_sm3_per_boe_factor"], 3)
        return {
            "oil_bbl": oil_bbl,
            "gas_boe": gas_boe,
            "boe_total": round(oil_bbl + gas_boe, 3),
        }

    def _summary_limits():
        conn = db_conn()
        cur = conn.cursor()
        row = cur.execute(
            "SELECT limite_hc_pct, limite_total_pct FROM pvt_params ORDER BY id DESC LIMIT 1"
        ).fetchone()
        conn.close()
        return {
            "hc_pct": float(row[0]) if row and row[0] is not None else 10.0,
            "total_pct": float(row[1]) if row and row[1] is not None else 7.0,
        }

    CHART_FOCUS_PAIRS = [
        {
            "key": "PE4_RISERP5",
            "title": "PE-04 × Riser P5",
            "subsea_bank": "B05",
            "subsea_tag": "18FT1506",
            "subsea_tags": ["18FT1506", "PE_4", "PE_4A", "PE-4A", "PE-04"],
            "topside_bank": "B03",
            "topside_tag": "Riser_P5",
            "topside_tags": ["Riser_P5", "13FT0367"],
            "subsea_label": "Subsea · PE-04",
            "topside_label": "Topside · Riser P5",
        },
        {
            "key": "PE2_RISERP2",
            "title": "PE-02 × Riser P2",
            "subsea_bank": "B10",
            "subsea_tag": "PE_2",
            "subsea_tags": ["PE_2", "18FT0506"],
            "topside_bank": "B08",
            "topside_tag": "Riser_P2",
            "topside_tags": ["Riser_P2", "13FT0217", "13FT0167"],
            "subsea_label": "Subsea · PE-02",
            "topside_label": "Topside · Riser P2",
        },
        {
            "key": "PW104_RISERP4",
            "title": "PW-104DA × Riser P4",
            "subsea_bank": "B15",
            "subsea_tag": "PW-104DA",
            "subsea_tags": ["PW-104DA", "18FT1106"],
            "topside_bank": "B13",
            "topside_tag": "Riser_P4",
            "topside_tags": ["Riser_P4", "13FT0317"],
            "subsea_label": "Subsea · PW-104DA",
            "topside_label": "Topside · Riser P4",
        },
    ]

    CHART_COMPARE_METRICS = {
        "hc": {
            "label": "HC (t)",
            "mpfm_metric": "MPFM corr HC (t)",
            "sep_metric": "hc_t",
            "recon_daily": "Daily HC (t)",
            "recon_sum": "Soma h. HC (t)",
            "recon_delta": "Δ HC (t)",
            "deviation_key": "hc",
        },
        "total": {
            "label": "Total (t)",
            "mpfm_metric": "MPFM corr Total (t)",
            "sep_metric": "total_t",
            "deviation_key": "total",
        },
        "oil": {
            "label": "Óleo (t)",
            "mpfm_metric": "MPFM corr Óleo (t)",
            "sep_metric": "oil_t",
            "recon_daily": "Daily Óleo (t)",
            "recon_sum": "Soma h. Óleo (t)",
            "recon_delta": "Δ Óleo (t)",
        },
        "gas": {
            "label": "Gás (t)",
            "mpfm_metric": "MPFM corr Gás (t)",
            "sep_metric": "gas_t",
            "recon_daily": "Daily Gás (t)",
            "recon_sum": "Soma h. Gás (t)",
            "recon_delta": "Δ Gás (t)",
        },
        "water": {
            "label": "Água (t)",
            "mpfm_metric": "MPFM corr Água (t)",
            "sep_metric": "water_t",
            "recon_daily": "Daily Água (t)",
            "recon_sum": "Soma h. Água (t)",
            "recon_delta": "Δ Água (t)",
        },
        "pressure": {
            "label": "Pressão (barg)",
            "sep_metric": "pressure_barg",
        },
        "temperature": {
            "label": "Temperatura (°C)",
            "sep_metric": "temp",
        },
    }

    CHART_SEPARATOR_TAGS = {
        "20FT0244": "20FT0244 · Gás",
        "20FT0247": "20FT0247 · Óleo",
        "20FT0251": "20FT0251 · Água",
        "SEP": "SEP · Consolidado 24h",
    }

    CHART_SEPARATOR_HOURLY_METRICS = {
        "oil": {"row_kind": "sep_oleo_detail", "metric_name": "Mass_ton", "aggregate": "sum"},
        "gas": {"row_kind": "sep_gas_detail", "metric_name": "Mass_t", "aggregate": "sum"},
        "water": {"row_kind": "sep_agua_detail", "metric_name": "Mass_ton", "aggregate": "sum"},
    }

    CHART_ADVANCED_PROCESS_METRICS = {
        "Pressão (barg)",
        "Temperatura (°C)",
        "Dens. Gás (kg/m³)",
        "Dens. Óleo (kg/m³)",
        "Dens. Água (kg/m³)",
        "Pressure_kPa",
        "Pressure_kPa_g",
        "Pressure_barg",
        "Temperature_degC",
        "SD_kg_sm3",
        "MD_kg_m3",
        "DT_kg_m3",
        "BSW_pct",
        "CPL",
        "CTL",
        "DiffPress_kPa",
        "Flowtime_min",
    }

    CHART_ADVANCED_GROUP_ORDER = {"prod": 0, "pvt": 1, "sep": 2, "proc": 3, "other": 4}

    def _chart_fill_gaps(rows, keys, value_getter):
        data = {}
        for row in rows:
            data[str(row[keys[0]])] = value_getter(row)
        return data

    def _chart_pair_map():
        return {item["key"]: item for item in CHART_FOCUS_PAIRS}

    def _chart_metric_choices(metric_keys):
        return [
            {"value": key, "label": CHART_COMPARE_METRICS[key]["label"]}
            for key in metric_keys
            if key in CHART_COMPARE_METRICS
        ]

    def _chart_format_tag_label(tag: str) -> str:
        raw = str(tag or "").strip()
        normalized = normalize_tag_name(raw)
        if normalized.startswith("RISER_"):
            return raw.replace("_", " ")
        return raw

    def _chart_format_bank_label(bank: str) -> str:
        raw = str(bank or "").strip()
        if raw == "SEP":
            return "SEP · Separador de Teste"
        return raw

    def _chart_format_sep_tag_label(tag: str) -> str:
        raw = str(tag or "").strip()
        return CHART_SEPARATOR_TAGS.get(raw, raw or "SEP")

    def _chart_metric_group(metric_name: str, *, bank: str = "", row_kind: str = "") -> str:
        name = str(metric_name or "")
        if name.startswith("PVT"):
            return "pvt"
        if name.startswith("MPFM"):
            return "prod"
        if bank == "SEP" or str(row_kind).startswith("sep"):
            if name in CHART_ADVANCED_PROCESS_METRICS:
                return "proc"
            return "sep"
        if name in CHART_ADVANCED_PROCESS_METRICS:
            return "proc"
        return "other"

    def _chart_sep_context_row_kinds(kind: str):
        if kind == "daily":
            return ("sep",)
        if kind == "hourly":
            return ("sep_oleo_detail", "sep_gas_detail", "sep_agua_detail")
        return (kind,)

    def _chart_sep_metric_query(cur, metric_name: str, date_from: str, date_to: str, tag: str = "", all_time: bool = False):
        row_kinds = _chart_sep_context_row_kinds("hourly")
        q_marks = ",".join("?" * len(row_kinds))
        sql = f"""
            SELECT day_ref, hour_ref, metric_value
            FROM measurements_active
            WHERE row_kind IN ({q_marks})
              AND bank='SEP'
              AND metric_name=?
              AND COALESCE(is_official,1)=1
        """
        params = list(row_kinds) + [metric_name]
        if not all_time:
            sql += " AND day_ref BETWEEN ? AND ?"
            params.extend([date_from, date_to])
        if tag:
            sql += """
                AND (
                    tag=?
                    OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                )
            """
            params.extend([tag, normalize_tag_name(tag)])
        sql += " ORDER BY day_ref, hour_ref"
        return [dict(r) for r in cur.execute(sql, params).fetchall()]

    def _chart_sep_metric_labels(kind: str, rows):
        labels = []
        for row in rows:
            day_ref = str(row.get("day_ref") or "")
            if kind == "hourly":
                hour_ref = row.get("hour_ref")
                if hour_ref is None:
                    continue
                labels.append(f"{day_ref} {int(hour_ref):02d}:00")
            else:
                labels.append(day_ref)
        return sorted({label for label in labels if label})

    def _chart_sep_metric_map(kind: str, rows):
        value_map = {}
        bucket = defaultdict(list)
        for row in rows:
            day_ref = str(row.get("day_ref") or "")
            if kind == "hourly":
                hour_ref = row.get("hour_ref")
                if hour_ref is None:
                    continue
                label = f"{day_ref} {int(hour_ref):02d}:00"
            else:
                label = day_ref
            bucket[label].append(row.get("metric_value"))
        for label, values in bucket.items():
            clean = [float(v) for v in values if v is not None]
            if not clean:
                value_map[label] = None
            elif kind == "hourly":
                value_map[label] = sum(clean)
            else:
                value_map[label] = clean[-1]
        return value_map

    def _chart_sep_hourly_series(cur, date_from: str, date_to: str, metric_key: str):
        component_rows = {}
        for key, cfg in CHART_SEPARATOR_HOURLY_METRICS.items():
            rows = [
                dict(r)
                for r in cur.execute(
                    """
                    SELECT day_ref, hour_ref, metric_value
                    FROM measurements_active
                    WHERE day_ref BETWEEN ? AND ?
                      AND row_kind=?
                      AND bank='SEP'
                      AND metric_name=?
                      AND hour_ref IS NOT NULL
                      AND COALESCE(is_official,1)=1
                    ORDER BY day_ref, hour_ref
                    """,
                    (date_from, date_to, cfg["row_kind"], cfg["metric_name"]),
                ).fetchall()
            ]
            component_rows[key] = _chart_sep_metric_map("hourly", rows)
        labels = sorted({label for item in component_rows.values() for label in item.keys()})
        if metric_key in component_rows:
            value_map = component_rows[metric_key]
        elif metric_key == "hc":
            value_map = {
                label: (component_rows["oil"].get(label) or 0) + (component_rows["gas"].get(label) or 0)
                for label in labels
            }
        elif metric_key == "total":
            value_map = {
                label: (component_rows["oil"].get(label) or 0) + (component_rows["gas"].get(label) or 0) + (component_rows["water"].get(label) or 0)
                for label in labels
            }
        else:
            value_map = {}
        return labels, value_map

    def _chart_adjustment_meta(source_file: str = "") -> tuple[bool, str]:
        source_name = str(source_file or "")
        is_adjusted = source_name.lower().startswith("manual_adjustment:")
        source_label = source_name.split(":", 1)[1] if is_adjusted and ":" in source_name else source_name
        return is_adjusted, source_label

    def _normalize_monitoring_bool(value):
        raw = str(value or "").strip().lower()
        if raw in {"sim", "yes", "true", "1"}:
            return "Sim"
        if raw in {"não", "nao", "no", "false", "0"}:
            return "Não"
        return ""

    def _sum_metric(cur, row_kind: str, day_ref: str, metric_name: str):
        return cur.execute(
            """
            SELECT COALESCE(SUM(metric_value),0)
            FROM measurements_active
            WHERE row_kind=? AND day_ref=? AND hour_ref IS NULL AND COALESCE(is_official,1)=1 AND metric_name=?
            """,
            (row_kind, day_ref, metric_name),
        ).fetchone()[0] or 0

    def _count_sep_official_fluids(cur, day_ref: str):
        return cur.execute(
            """
            SELECT COUNT(DISTINCT fluid_kind)
            FROM sep_source_files
            WHERE production_date=?
              AND is_active=1
              AND is_official=1
              AND fluid_kind IN ('sep_oleo','sep_gas','sep_agua')
            """,
            (day_ref,),
        ).fetchone()[0] or 0

    def _get_sep_day_summary(cur, day_ref: str):
        gas_sm3_detail = _sum_metric(cur, "sep_gas_detail", day_ref, "StVol_m3") or _sum_metric(
            cur, "sep_gas_detail", day_ref, "GrVol_m3"
        )
        water_sm3_detail = (
            _sum_metric(cur, "sep_agua_detail", day_ref, "GSV_sm3")
            or _sum_metric(cur, "sep_agua_detail", day_ref, "NSV_sm3")
            or _sum_metric(cur, "sep_agua_detail", day_ref, "GV_m3")
            or _sum_metric(cur, "sep_agua_detail", day_ref, "IV_m3")
        )
        summary_count = cur.execute(
            """
            SELECT COUNT(*)
            FROM measurements_active
            WHERE day_ref=? AND row_kind='sep' AND hour_ref IS NULL AND COALESCE(is_official,1)=1
            """,
            (day_ref,),
        ).fetchone()[0] or 0
        row = cur.execute(
            """
            SELECT
                SUM(CASE WHEN metric_name='oil_t' THEN metric_value END) as oil_t,
                SUM(CASE WHEN metric_name='gas_t' THEN metric_value END) as gas_t,
                SUM(CASE WHEN metric_name='water_t' THEN metric_value END) as water_t,
                SUM(CASE WHEN metric_name='hc_t' THEN metric_value END) as hc_t,
                SUM(CASE WHEN metric_name='total_t' THEN metric_value END) as total_t,
                SUM(CASE WHEN metric_name='oil_m3' THEN metric_value END) as oil_m3
            FROM measurements_active
            WHERE day_ref=? AND row_kind='sep' AND hour_ref IS NULL AND COALESCE(is_official,1)=1
            """,
            (day_ref,),
        ).fetchone()
        if summary_count:
            oil_t, gas_t, water_t, hc_t, total_t, oil_m3 = row
            return {
                "oil_t": oil_t or 0,
                "gas_t": gas_t or 0,
                "water_t": water_t or 0,
                "hc_t": hc_t or 0,
                "total_t": total_t or 0,
                "oil_m3": oil_m3 or 0,
                "gas_sm3": gas_sm3_detail or 0,
                "water_sm3": water_sm3_detail or 0,
                "source": "consolidado",
                "present": True,
                "official_fluids": _count_sep_official_fluids(cur, day_ref),
            }

        oil_t = _sum_metric(cur, "sep_oleo_detail", day_ref, "Mass_ton")
        gas_t = _sum_metric(cur, "sep_gas_detail", day_ref, "Mass_t")
        water_t = _sum_metric(cur, "sep_agua_detail", day_ref, "Mass_ton")
        oil_m3 = (
            _sum_metric(cur, "sep_oleo_detail", day_ref, "IV_m3")
            or _sum_metric(cur, "sep_oleo_detail", day_ref, "GV_m3")
            or _sum_metric(cur, "sep_oleo_detail", day_ref, "GSV_sm3")
        )
        detail_count = cur.execute(
            """
            SELECT COUNT(*)
            FROM measurements_active
            WHERE day_ref=?
              AND row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail')
              AND COALESCE(is_official,1)=1
            """,
            (day_ref,),
        ).fetchone()[0] or 0
        official_fluids = _count_sep_official_fluids(cur, day_ref)
        if detail_count or official_fluids:
            return {
                "oil_t": oil_t or 0,
                "gas_t": gas_t or 0,
                "water_t": water_t or 0,
                "hc_t": (oil_t or 0) + (gas_t or 0),
                "total_t": (oil_t or 0) + (gas_t or 0) + (water_t or 0),
                "oil_m3": oil_m3 or 0,
                "gas_sm3": gas_sm3_detail or 0,
                "water_sm3": water_sm3_detail or 0,
                "source": "detalhe_manual",
                "present": True,
                "official_fluids": official_fluids,
            }
        return {
            "oil_t": None,
            "gas_t": None,
            "water_t": None,
            "hc_t": None,
            "total_t": None,
            "oil_m3": None,
            "gas_sm3": None,
            "water_sm3": None,
            "source": "",
            "present": False,
            "official_fluids": 0,
        }

    _SEP_EMPTY_SUMMARY = {
        "oil_t": None,
        "gas_t": None,
        "water_t": None,
        "hc_t": None,
        "total_t": None,
        "oil_m3": None,
        "gas_sm3": None,
        "water_sm3": None,
        "source": "",
        "present": False,
        "official_fluids": 0,
    }

    def _get_sep_month_summary_map(cur, month_start: str, month_end: str) -> dict:
        """Versão em lote de _get_sep_day_summary: resolve o resumo SEP de todos os dias
        do mês com poucas queries (GROUP BY day_ref) em vez de uma rodada de queries por dia.
        Mantém exatamente a mesma lógica de fallback do original."""
        detail_sums: dict[str, dict[tuple[str, str], float]] = defaultdict(dict)
        for day_ref, row_kind, metric_name, total in cur.execute(
            """
            SELECT day_ref, row_kind, metric_name, SUM(metric_value)
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND row_kind IN ('sep_gas_detail','sep_agua_detail','sep_oleo_detail')
              AND hour_ref IS NULL AND COALESCE(is_official,1)=1
              AND metric_name IN ('StVol_m3','GrVol_m3','GSV_sm3','NSV_sm3','GV_m3','IV_m3','Mass_ton','Mass_t','GSV_sm3')
            GROUP BY day_ref, row_kind, metric_name
            """,
            (month_start, month_end),
        ).fetchall():
            detail_sums[day_ref][(row_kind, metric_name)] = total or 0

        sep_stats: dict[str, dict] = {}
        for day_ref, oil_t, gas_t, water_t, hc_t, total_t, oil_m3, cnt in cur.execute(
            """
            SELECT day_ref,
                   SUM(CASE WHEN metric_name='oil_t' THEN metric_value END),
                   SUM(CASE WHEN metric_name='gas_t' THEN metric_value END),
                   SUM(CASE WHEN metric_name='water_t' THEN metric_value END),
                   SUM(CASE WHEN metric_name='hc_t' THEN metric_value END),
                   SUM(CASE WHEN metric_name='total_t' THEN metric_value END),
                   SUM(CASE WHEN metric_name='oil_m3' THEN metric_value END),
                   COUNT(*)
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND row_kind='sep' AND hour_ref IS NULL AND COALESCE(is_official,1)=1
            GROUP BY day_ref
            """,
            (month_start, month_end),
        ).fetchall():
            sep_stats[day_ref] = {
                "oil_t": oil_t, "gas_t": gas_t, "water_t": water_t,
                "hc_t": hc_t, "total_t": total_t, "oil_m3": oil_m3, "cnt": cnt,
            }

        detail_counts: dict[str, int] = {}
        for day_ref, cnt in cur.execute(
            """
            SELECT day_ref, COUNT(*)
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<?
              AND row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail')
              AND COALESCE(is_official,1)=1
            GROUP BY day_ref
            """,
            (month_start, month_end),
        ).fetchall():
            detail_counts[day_ref] = cnt

        official_fluids_map: dict[str, int] = {}
        for day_ref, cnt in cur.execute(
            """
            SELECT production_date, COUNT(DISTINCT fluid_kind)
            FROM sep_source_files
            WHERE production_date>=? AND production_date<?
              AND is_active=1 AND is_official=1
              AND fluid_kind IN ('sep_oleo','sep_gas','sep_agua')
            GROUP BY production_date
            """,
            (month_start, month_end),
        ).fetchall():
            official_fluids_map[day_ref] = cnt

        def _detail(day_ref: str, row_kind: str, metric: str) -> float:
            return detail_sums.get(day_ref, {}).get((row_kind, metric)) or 0

        all_days = set(detail_sums) | set(sep_stats) | set(detail_counts) | set(official_fluids_map)
        result: dict[str, dict] = {}
        for day_ref in all_days:
            gas_sm3_detail = _detail(day_ref, "sep_gas_detail", "StVol_m3") or _detail(day_ref, "sep_gas_detail", "GrVol_m3")
            water_sm3_detail = (
                _detail(day_ref, "sep_agua_detail", "GSV_sm3")
                or _detail(day_ref, "sep_agua_detail", "NSV_sm3")
                or _detail(day_ref, "sep_agua_detail", "GV_m3")
                or _detail(day_ref, "sep_agua_detail", "IV_m3")
            )
            official_fluids = official_fluids_map.get(day_ref, 0)
            stat = sep_stats.get(day_ref)
            if stat and stat["cnt"]:
                result[day_ref] = {
                    "oil_t": stat["oil_t"] or 0,
                    "gas_t": stat["gas_t"] or 0,
                    "water_t": stat["water_t"] or 0,
                    "hc_t": stat["hc_t"] or 0,
                    "total_t": stat["total_t"] or 0,
                    "oil_m3": stat["oil_m3"] or 0,
                    "gas_sm3": gas_sm3_detail or 0,
                    "water_sm3": water_sm3_detail or 0,
                    "source": "consolidado",
                    "present": True,
                    "official_fluids": official_fluids,
                }
                continue

            oil_t = _detail(day_ref, "sep_oleo_detail", "Mass_ton")
            gas_t = _detail(day_ref, "sep_gas_detail", "Mass_t")
            water_t = _detail(day_ref, "sep_agua_detail", "Mass_ton")
            oil_m3 = (
                _detail(day_ref, "sep_oleo_detail", "IV_m3")
                or _detail(day_ref, "sep_oleo_detail", "GV_m3")
                or _detail(day_ref, "sep_oleo_detail", "GSV_sm3")
            )
            detail_count = detail_counts.get(day_ref, 0)
            if detail_count or official_fluids:
                result[day_ref] = {
                    "oil_t": oil_t or 0,
                    "gas_t": gas_t or 0,
                    "water_t": water_t or 0,
                    "hc_t": (oil_t or 0) + (gas_t or 0),
                    "total_t": (oil_t or 0) + (gas_t or 0) + (water_t or 0),
                    "oil_m3": oil_m3 or 0,
                    "gas_sm3": gas_sm3_detail or 0,
                    "water_sm3": water_sm3_detail or 0,
                    "source": "detalhe_manual",
                    "present": True,
                    "official_fluids": official_fluids,
                }
        return result

    def _sync_sep_month_state(target_month: str, *, rebuild_summary: bool = True):
        yr, mo, date_from, date_to = _parse_target_month(target_month)
        state = load_state(yr, mo)
        conn = db_conn()
        conn.row_factory = lambda cursor, row: {
            cursor.description[idx][0]: row[idx] for idx in range(len(cursor.description))
        }
        cur = conn.cursor()
        official_days = [
            row["production_date"]
            for row in cur.execute(
                """
                SELECT production_date
                FROM sep_source_files
                WHERE is_active=1
                  AND is_official=1
                  AND production_date BETWEEN ? AND ?
                  AND fluid_kind IN ('sep_oleo','sep_gas','sep_agua')
                GROUP BY production_date
                HAVING COUNT(DISTINCT fluid_kind)=3
                ORDER BY production_date
                """,
                (date_from, date_to),
            ).fetchall()
        ]
        conn.close()

        month_keys = [key for key in (state.get("sep_by_day") or {}).keys() if str(key).startswith(target_month)]
        for key in month_keys:
            state.setdefault("sep_by_day", {}).pop(key, None)
        state["sep_days"] = [day for day in (state.get("sep_days") or []) if not str(day).startswith(target_month)]

        rebuilt_days = []
        zero_days = []
        synced_days = []
        missing_payload_days = []

        # Carrega todos os payloads SEP do mês de uma vez para evitar N+1
        sep_payloads_by_day = load_sep_data_by_range(date_from, date_to) if official_days else {}

        for production_date in official_days:
            if rebuild_summary:
                rebuild_result = rebuild_sep_summary_for_day(production_date)
                if rebuild_result.get("rebuilt"):
                    rebuilt_days.append(production_date)
            sep_payload = sep_payloads_by_day.get(production_date, load_sep_data_by_day(production_date))
            if sep_payload:
                state.setdefault("sep_by_day", {})[production_date] = serialize_sep_row(sep_payload)
                synced_days.append(production_date)
                day_block = sep_payload.get("DAY", {}) or {}
                if all(abs(float(day_block.get(metric) or 0)) <= 0.000001 for metric in ("oil_t", "gas_t", "water_t", "hc_t", "total_t", "oil_m3")):
                    zero_days.append(production_date)
            else:
                state.setdefault("sep_by_day", {})[production_date] = {}
                missing_payload_days.append(production_date)
            state.setdefault("sep_days", []).append(production_date)

        state["sep_days"] = sorted(set(state.get("sep_days") or []))
        save_state(state)
        return {
            "month": target_month,
            "official_days": len(official_days),
            "synced_days": len(synced_days),
            "rebuilt_days": len(rebuilt_days),
            "zero_days": len(zero_days),
            "missing_payload_days": len(missing_payload_days),
            "days": official_days,
            "zero_day_refs": zero_days,
        }

    @app.get("/api/dashboard")
    @cached(ttl=300, key_prefix="dashboard")
    def api_dashboard():
        return {"months": build_dashboard_months(work_dir, output_dir, month_pt, excel_name)}

    @app.get("/api/ops/summary")
    @cached(ttl=120, key_prefix="ops_summary")
    def api_ops_summary(date: str = ""):
        date = normalize_date_input(date)
        conn = db_conn()
        cur = conn.cursor()
        day = (
            date
            or cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE bank<>'SEP' AND row_kind='daily'").fetchone()[0]
            or cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE bank<>'SEP'").fetchone()[0]
            or cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0]
            or ""
        )
        if not day:
            conn.close()
            return {"date": "", "cards": {}, "mpfm_groups": [], "sep_status": [], "alerts": [], "calendar": []}

        banks_monitored = cur.execute(
            "SELECT COUNT(DISTINCT bank) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND bank<>'' AND bank<>'SEP'",
            (day,),
        ).fetchone()[0]
        hours_received = cur.execute(
            "SELECT COUNT(DISTINCT printf('%s-%02d', bank, hour_ref)) FROM measurements_active WHERE day_ref=? AND row_kind='hourly' AND bank<>'' AND bank<>'SEP' AND hour_ref IS NOT NULL",
            (day,),
        ).fetchone()[0]
        hours_expected = banks_monitored * 24
        alerts_count = cur.execute("SELECT COUNT(*) FROM validation_issues WHERE day_ref=?", (day,)).fetchone()[0]
        points_count = cur.execute(
            "SELECT COUNT(DISTINCT bank || '|' || tag) FROM measurements_active WHERE day_ref=? AND bank<>'' AND bank<>'SEP' AND tag<>''",
            (day,),
        ).fetchone()[0]
        total_hc_t = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='MPFM corr HC (t)'",
            (day,),
        ).fetchone()[0] or 0
        total_oil_t = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='MPFM corr Óleo (t)'",
            (day,),
        ).fetchone()[0] or 0
        total_gas_t = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='MPFM corr Gás (t)'",
            (day,),
        ).fetchone()[0] or 0
        total_water_t = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='MPFM corr Água (t)'",
            (day,),
        ).fetchone()[0] or 0

        devs = recon_deviation_metrics(day)
        got = {r[0] for r in cur.execute("SELECT file_type FROM files_imported WHERE content_date=? AND ext='txt'", (day,)).fetchall()}
        sep_complete = len(got.intersection({"sep_oleo", "sep_agua", "sep_gas"})) == 3

        groups = []
        for row in cur.execute(
            """
            SELECT bank,
                   SUM(CASE WHEN row_kind='daily' THEN 1 ELSE 0 END) AS daily_rows,
                   COUNT(DISTINCT CASE WHEN row_kind='hourly' THEN hour_ref END) AS hours,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) AS total_t
            FROM measurements_active
            WHERE day_ref=? AND bank<>'' AND bank<>'SEP'
            GROUP BY bank ORDER BY bank
            """,
            (day,),
        ):
            bank, daily_rows, hours, total_t = row
            status = "ok" if daily_rows and hours >= 24 else ("attention" if hours > 0 else "missing")
            groups.append({"bank": bank, "daily_rows": daily_rows, "hours": hours, "total_t": round(total_t or 0, 3), "status": status})

        sep = []
        for key, label in [("sep_oleo", "Óleo"), ("sep_agua", "Água"), ("sep_gas", "Gás")]:
            sep.append({"key": key, "label": label, "present": key in got, "status": "ok" if key in got else "missing"})
        sep.append({"key": "sep_consolidado", "label": "Consolidação SEP", "present": sep_complete, "status": "ok" if sep_complete else "attention"})

        alerts = [dict(r) for r in cur.execute("SELECT severity, issue_type, ref_key, details FROM validation_issues WHERE day_ref=? ORDER BY id DESC LIMIT 8", (day,)).fetchall()]
        total_oil_m3 = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='PVT @20 vol Óleo (m³)'",
            (day,),
        ).fetchone()[0] or 0
        total_gas_sm3 = cur.execute(
            "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref=? AND row_kind='daily' AND metric_name='PVT @20 vol Gás (Sm³)'",
            (day,),
        ).fetchone()[0] or 0
        cards_payload = {
            "banks_monitored": banks_monitored,
            "points": points_count,
            "hours_expected": hours_expected,
            "hours_received": hours_received,
            "hours_missing": max(hours_expected - hours_received, 0),
            "alerts": alerts_count,
            "total_hc_t": round(total_hc_t, 3),
            "total_oil_t": round(total_oil_t, 3),
            "total_gas_t": round(total_gas_t, 3),
            "total_water_t": round(total_water_t, 3),
            "total_oil_bbl_d": bbl_from_m3(total_oil_m3),
            "total_boe_d": boe_from(total_oil_m3, total_gas_sm3),
            "max_dev_hc_pct": devs.get("max_dev_hc_pct"),
            "max_dev_total_pct": devs.get("max_dev_total_pct"),
            "sep_complete": sep_complete,
        }
        conn.close()
        return {
            "date": day,
            "cards": cards_payload,
            "mpfm_groups": groups,
            "sep_status": sep,
            "alerts": alerts,
            "calendar": month_calendar_status(day),
            "last_refresh": datetime.now().isoformat(timespec="seconds"),
        }

    _month_summary_cache: dict[str, tuple[float, dict]] = {}
    _MONTH_SUMMARY_TTL = 20.0

    @app.get("/api/ops/month-summary")
    @cached(ttl=180, key_prefix="ops_month_summary")
    def api_ops_month_summary(month: str = ""):
        cache_key = month or "__latest__"
        cached = _month_summary_cache.get(cache_key)
        if cached and (time.time() - cached[0]) < _MONTH_SUMMARY_TTL:
            return cached[1]
        conn = db_conn()
        cur = conn.cursor()
        conversion = _summary_conversion_settings()
        criteria = _summary_limits()
        if not month:
            last = cur.execute(
                "SELECT MAX(day_ref) FROM measurements_active WHERE row_kind IN ('daily','hourly') AND day_ref NOT LIKE '0000%'"
            ).fetchone()[0] or ""
            month = last[:7] if last else ""
        if not month:
            conn.close()
            return {
                "month": "",
                "months_available": [],
                "production": {},
                "by_bank": [],
                "by_tag": [],
                "daily": [],
                "sep": {},
                "alerts": {},
                "last_runs": [],
            }

        import calendar as _cal

        yr, mo = int(month[:4]), int(month[5:7])
        month_start = f"{yr:04d}-{mo:02d}-01"
        _nyr, _nmo = (yr + 1, 1) if mo == 12 else (yr, mo + 1)
        month_end = f"{_nyr:04d}-{_nmo:02d}-01"

        months_available = [
            row[0]
            for row in cur.execute(
                "SELECT DISTINCT substr(day_ref,1,7) m FROM measurements_active WHERE row_kind IN ('daily','hourly') AND day_ref NOT LIKE '0000%' ORDER BY m DESC"
            ).fetchall()
        ]

        production = {}
        metric_specs = [
            ("MPFM corr Óleo (t)", "oil_t"),
            ("MPFM corr Gás (t)", "gas_t"),
            ("MPFM corr HC (t)", "hc_t"),
            ("MPFM corr Água (t)", "water_t"),
            ("MPFM corr Total (t)", "total_t"),
            ("PVT @20 vol Óleo (m³)", "oil_m3"),
            ("PVT @20 vol Gás (Sm³)", "gas_sm3"),
            ("PVT @20 vol Água (m³)", "water_m3"),
            ("PVT vol Óleo (m³)", "oil_m3_line"),
            ("PVT vol Gás (Sm³)", "gas_sm3_line"),
            ("PVT vol Água (m³)", "water_m3_line"),
        ]
        for metric, key in metric_specs:
            value = cur.execute(
                "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref>=? AND day_ref<? AND row_kind='daily' AND metric_name=?",
                (month_start, month_end, metric),
                ).fetchone()[0]
            production[key] = round(value or 0, 3)

        if production["hc_t"] == 0:
            fallback_specs = [
                ("MPFM corr HC (t)", "hc_t"),
                ("MPFM corr Óleo (t)", "oil_t"),
                ("MPFM corr Gás (t)", "gas_t"),
                ("MPFM corr Água (t)", "water_t"),
                ("PVT @20 vol Óleo (m³)", "oil_m3"),
                ("PVT @20 vol Gás (Sm³)", "gas_sm3"),
                ("PVT @20 vol Água (m³)", "water_m3"),
            ]
            for metric, key in fallback_specs:
                value = cur.execute(
                    "SELECT COALESCE(SUM(metric_value),0) FROM measurements_active WHERE day_ref>=? AND day_ref<? AND row_kind='hourly' AND metric_name=?",
                    (month_start, month_end, metric),
                ).fetchone()[0]
                production[key] = round(value or 0, 3)

        production["total_t"] = round((production.get("oil_t") or 0) + (production.get("gas_t") or 0) + (production.get("water_t") or 0), 3)

        production.update(_conversion_payload(production.get("oil_m3"), production.get("gas_sm3"), conversion))
        production["boe"] = production["boe_total"]

        days_with_data = cur.execute(
            "SELECT COUNT(DISTINCT day_ref) FROM measurements_active WHERE day_ref>=? AND day_ref<? AND row_kind IN ('daily','hourly') AND bank<>''",
            (month_start, month_end),
        ).fetchone()[0]

        # Pré-carrega totais horários por banco para evitar N+1 no loop by_bank
        hourly_bank_totals: dict[str, dict[str, float]] = {}
        for bank, metric, value in cur.execute(
            """
            SELECT bank, metric_name, COALESCE(SUM(metric_value),0)
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND bank<>'' AND row_kind='hourly'
              AND metric_name IN ('MPFM corr HC (t)', 'MPFM corr Óleo (t)', 'MPFM corr Gás (t)', 'MPFM corr Água (t)')
            GROUP BY bank, metric_name
            """,
            (month_start, month_end),
        ).fetchall():
            hourly_bank_totals.setdefault(bank, {})[metric] = float(value or 0)

        by_bank = []
        for row in cur.execute(
            """
            SELECT bank,
                   COUNT(DISTINCT day_ref) as days,
                   COUNT(DISTINCT CASE WHEN row_kind='hourly' THEN day_ref||hour_ref END) as hours,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Óleo (t)' THEN metric_value ELSE 0 END) as oil_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Gás (t)' THEN metric_value ELSE 0 END) as gas_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr HC (t)' THEN metric_value ELSE 0 END) as hc_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Água (t)' THEN metric_value ELSE 0 END) as water_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) as total_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oil_m3,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gas_sm3,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as water_m3
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND bank<>'' AND row_kind IN ('daily','hourly')
            GROUP BY bank ORDER BY bank
            """,
            (month_start, month_end),
        ):
            bank, days, hours, oil, gas, hc, water, total, oil_m3, gas_sm3, water_m3 = row
            if hc == 0 and hours > 0:
                htotals = hourly_bank_totals.get(bank, {})
                hc = htotals.get("MPFM corr HC (t)", hc)
                oil = htotals.get("MPFM corr Óleo (t)", oil)
                gas = htotals.get("MPFM corr Gás (t)", gas)
                water = htotals.get("MPFM corr Água (t)", water)
            by_bank.append(
                {
                    "bank": bank,
                    "days": days,
                    "hours": hours,
                    "oil_t": round(oil or 0, 1),
                    "gas_t": round(gas or 0, 1),
                    "hc_t": round(hc or 0, 1),
                    "water_t": round(water or 0, 1),
                    "total_t": round((oil or 0) + (gas or 0) + (water or 0), 1),
                    "oil_m3": round(oil_m3 or 0, 3),
                    "gas_sm3": round(gas_sm3 or 0, 1),
                    "water_m3": round(water_m3 or 0, 3),
                    "oil_bbl": _conversion_payload(oil_m3, gas_sm3, conversion)["oil_bbl"],
                    "gas_boe": _conversion_payload(oil_m3, gas_sm3, conversion)["gas_boe"],
                    "boe_total": _conversion_payload(oil_m3, gas_sm3, conversion)["boe_total"],
                    "boe": _conversion_payload(oil_m3, gas_sm3, conversion)["boe_total"],
                }
            )

        by_tag = []
        for row in cur.execute(
            """
            SELECT bank, tag,
                   COUNT(DISTINCT day_ref) as days,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr HC (t)' THEN metric_value ELSE 0 END) as hc_daily,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr HC (t)' THEN metric_value ELSE 0 END) as hc_hourly,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Óleo (t)' THEN metric_value ELSE 0 END) as oil_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Gás (t)' THEN metric_value ELSE 0 END) as gas_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Água (t)' THEN metric_value ELSE 0 END) as water_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) as total_t,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oil_m3,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gas_sm3,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as water_m3
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND bank<>'' AND tag<>'' AND row_kind IN ('daily','hourly')
            GROUP BY bank, tag ORDER BY bank, tag
            """,
            (month_start, month_end),
        ):
            bank, tag, days, hc_daily, hc_hourly, oil, gas, water, total, oil_m3, gas_sm3, water_m3 = row
            hc = hc_daily if hc_daily > 0 else hc_hourly
            by_tag.append(
                {
                    "bank": bank,
                    "tag": tag,
                    "days": days,
                    "hc_t": round(hc or 0, 1),
                    "oil_t": round(oil or 0, 1),
                    "gas_t": round(gas or 0, 1),
                    "water_t": round(water or 0, 1),
                    "total_t": round((total if total not in (None, 0) else (oil or 0) + (gas or 0) + (water or 0)), 1),
                    "oil_m3": round(oil_m3 or 0, 3),
                    "gas_sm3": round(gas_sm3 or 0, 1),
                    "water_m3": round(water_m3 or 0, 3),
                    "oil_bbl": _conversion_payload(oil_m3, gas_sm3, conversion)["oil_bbl"],
                    "gas_boe": _conversion_payload(oil_m3, gas_sm3, conversion)["gas_boe"],
                    "boe_total": _conversion_payload(oil_m3, gas_sm3, conversion)["boe_total"],
                    "boe": _conversion_payload(oil_m3, gas_sm3, conversion)["boe_total"],
                }
            )

        tag_daily = []
        bank_daily_map = {}
        for row in cur.execute(
            """
            SELECT day_ref, bank, tag,
                   COUNT(DISTINCT CASE WHEN row_kind='hourly' THEN hour_ref END) as hours,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Óleo (t)' THEN metric_value ELSE 0 END) as oil_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Gás (t)' THEN metric_value ELSE 0 END) as gas_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Água (t)' THEN metric_value ELSE 0 END) as water_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr HC (t)' THEN metric_value ELSE 0 END) as hc_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) as total_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oilv_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gasv_d,
                   SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as waterv_d,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Óleo (t)' THEN metric_value ELSE 0 END) as oil_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Gás (t)' THEN metric_value ELSE 0 END) as gas_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Água (t)' THEN metric_value ELSE 0 END) as water_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr HC (t)' THEN metric_value ELSE 0 END) as hc_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) as total_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oilv_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gasv_h,
                   SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as waterv_h
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND bank<>'' AND bank<>'SEP' AND tag<>'' AND row_kind IN ('daily','hourly')
            GROUP BY day_ref, bank, tag
            ORDER BY day_ref, bank, tag
            """,
            (month_start, month_end),
        ).fetchall():
            (
                day_ref,
                bank,
                tag,
                hours,
                oil_d,
                gas_d,
                water_d,
                hc_d,
                total_d,
                oilv_d,
                gasv_d,
                waterv_d,
                oil_h,
                gas_h,
                water_h,
                hc_h,
                total_h,
                oilv_h,
                gasv_h,
                waterv_h,
            ) = row
            use_hourly = (hc_d or 0) == 0 and (total_d or 0) == 0 and (oilv_d or 0) == 0 and hours > 0
            oil_val   = (oil_h   if use_hourly else oil_d)   or 0
            gas_val   = (gas_h   if use_hourly else gas_d)   or 0
            water_val = (water_h if use_hourly else water_d) or 0
            item = {
                "day": day_ref,
                "bank": bank,
                "tag": tag,
                "hours": hours or 0,
                "oil_t":   round(oil_val,   3),
                "gas_t":   round(gas_val,   3),
                "water_t": round(water_val, 3),
                "hc_t":    round((hc_h if use_hourly else hc_d) or 0, 3),
                "hc_calc": round(oil_val + gas_val, 3),  # óleo+gás calculado (deve == hc_t do PDF)
                "total_t": round(oil_val + gas_val + water_val, 3),
                "oil_m3":  round((oilv_h if use_hourly else oilv_d) or 0, 3),
                "gas_sm3": round((gasv_h if use_hourly else gasv_d) or 0, 3),
                "water_m3": round((waterv_h if use_hourly else waterv_d) or 0, 3),
            }
            item.update(_conversion_payload(item["oil_m3"], item["gas_sm3"], conversion))
            item["boe"] = item["boe_total"]
            tag_daily.append(item)

            bank_key = (day_ref, bank)
            current = bank_daily_map.setdefault(
                bank_key,
                {
                    "day": day_ref,
                    "bank": bank,
                    "hours": 0,
                    "tags_count": 0,
                    "oil_t": 0.0,
                    "gas_t": 0.0,
                    "water_t": 0.0,
                    "hc_t": 0.0,
                    "total_t": 0.0,
                    "oil_m3": 0.0,
                    "gas_sm3": 0.0,
                    "water_m3": 0.0,
                },
            )
            current["hours"] = max(current["hours"], item["hours"])
            current["tags_count"] += 1
            for key in ("oil_t", "gas_t", "water_t", "hc_t", "total_t", "oil_m3", "gas_sm3", "water_m3"):
                current[key] += item[key] or 0.0

        bank_daily = []
        for item in sorted(bank_daily_map.values(), key=lambda row: (row["bank"], row["day"])):
            item["oil_t"] = round(item["oil_t"], 3)
            item["gas_t"] = round(item["gas_t"], 3)
            item["water_t"] = round(item["water_t"], 3)
            item["hc_t"] = round(item["hc_t"], 3)
            item["total_t"] = round(item["oil_t"] + item["gas_t"] + item["water_t"], 3)
            item["oil_m3"] = round(item["oil_m3"], 3)
            item["gas_sm3"] = round(item["gas_sm3"], 3)
            item["water_m3"] = round(item["water_m3"], 3)
            item.update(_conversion_payload(item["oil_m3"], item["gas_sm3"], conversion))
            item["boe"] = item["boe_total"]
            bank_daily.append(item)

        daily = []
        _mpfm_rows_by_day: dict[str, list] = defaultdict(list)
        for row in cur.execute(
            """
            SELECT day_ref, bank, tag,
                SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Óleo (t)'  THEN metric_value ELSE 0 END) as oil_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Gás (t)'   THEN metric_value ELSE 0 END) as gas_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Água (t)'  THEN metric_value ELSE 0 END) as water_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr HC (t)'    THEN metric_value ELSE 0 END) as hc_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='MPFM corr Total (t)' THEN metric_value ELSE 0 END) as total_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oilv_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gasv_d,
                SUM(CASE WHEN row_kind='daily' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as waterv_d,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Óleo (t)' THEN metric_value ELSE 0 END) as oil_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Gás (t)'  THEN metric_value ELSE 0 END) as gas_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Água (t)' THEN metric_value ELSE 0 END) as water_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr HC (t)'   THEN metric_value ELSE 0 END) as hc_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='MPFM corr Total (t)'THEN metric_value ELSE 0 END) as total_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Óleo (m³)' THEN metric_value ELSE 0 END) as oilv_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Gás (Sm³)' THEN metric_value ELSE 0 END) as gasv_h,
                SUM(CASE WHEN row_kind='hourly' AND metric_name='PVT @20 vol Água (m³)' THEN metric_value ELSE 0 END) as waterv_h,
                COUNT(DISTINCT CASE WHEN row_kind='hourly' THEN hour_ref END) as hrs
            FROM measurements_active WHERE day_ref>=? AND day_ref<? AND bank<>'' AND tag<>''
            GROUP BY day_ref, bank, tag
            """,
            (month_start, month_end),
        ).fetchall():
            _mpfm_rows_by_day[row[0]].append(row[1:])

        sep_month_map = _get_sep_month_summary_map(cur, month_start, month_end)

        # Master list of expected active MPFM tags from cadastro or fallback
        cad = load_cadastro() or {}
        active_subsea = [
            {
                "bank": b.get("bank_code") or b.get("bank"),
                "tag": b.get("sistema") or b.get("poco_equinor"),
                "sensor_tag": b.get("tag_associado") or "",
                "tipo": "Subsea",
            }
            for b in cad.get("banks_subsea", [])
            if b.get("ativo", True)
        ]
        if not active_subsea:
            active_subsea = [
                {"bank": "B10", "tag": "PE_2", "sensor_tag": "18FT0506", "tipo": "Subsea"},
                {"bank": "B15", "tag": "PW-104DA", "sensor_tag": "18FT1106", "tipo": "Subsea"},
                {"bank": "B05", "tag": "PE_4", "sensor_tag": "18FT1506", "tipo": "Subsea"},
            ]

        active_topside = [
            {
                "bank": b.get("bank_code") or b.get("bank"),
                "tag": b.get("sistema"),
                "sensor_tag": b.get("tag_associado") or "",
                "tipo": "Topside",
            }
            for b in cad.get("banks_topside", [])
            if b.get("ativo", True)
        ]
        if not active_topside:
            active_topside = [
                {"bank": "B08", "tag": "Riser_P2", "sensor_tag": "13FT0217", "tipo": "Topside"},
                {"bank": "B13", "tag": "Riser_P4", "sensor_tag": "13FT0317", "tipo": "Topside"},
                {"bank": "B03", "tag": "Riser_P5", "sensor_tag": "13FT0367", "tipo": "Topside"},
            ]

        # Map MPFM coverage per (day_ref, bank, tag) for the whole month
        mpfm_cov: dict[tuple[str, str, str], tuple[bool, int]] = {}
        for day_ref_c, b_code, t_name, d_cnt, h_hrs in cur.execute(
            """
            SELECT day_ref, bank, tag,
                   SUM(CASE WHEN row_kind='daily' THEN 1 ELSE 0 END) as d_cnt,
                   COUNT(DISTINCT CASE WHEN row_kind='hourly' THEN hour_ref END) as h_hrs
            FROM measurements_active
            WHERE day_ref>=? AND day_ref<? AND bank<>'' AND bank<>'SEP' AND tag<>''
            GROUP BY day_ref, bank, tag
            """,
            (month_start, month_end),
        ).fetchall():
            mpfm_cov[(day_ref_c, b_code, t_name)] = (bool(d_cnt > 0), int(h_hrs or 0))

        # Map Separador presence per day
        sep_p: dict[str, dict[str, bool]] = defaultdict(lambda: {"oleo": False, "gas": False, "agua": False})
        for day_ref_s, fk in cur.execute(
            "SELECT DISTINCT production_date, fluid_kind FROM sep_source_files WHERE production_date>=? AND production_date<? AND is_active=1 AND is_official=1",
            (month_start, month_end),
        ).fetchall():
            if fk == "sep_oleo": sep_p[day_ref_s]["oleo"] = True
            elif fk == "sep_gas": sep_p[day_ref_s]["gas"] = True
            elif fk == "sep_agua": sep_p[day_ref_s]["agua"] = True

        for day_ref_s, ft in cur.execute(
            "SELECT DISTINCT content_date, file_type FROM files_imported WHERE content_date>=? AND content_date<? AND file_type LIKE 'sep_%'",
            (month_start, month_end),
        ).fetchall():
            if ft == "sep_oleo": sep_p[day_ref_s]["oleo"] = True
            elif ft == "sep_gas": sep_p[day_ref_s]["gas"] = True
            elif ft == "sep_agua": sep_p[day_ref_s]["agua"] = True

        for day_ref_s, rk in cur.execute(
            "SELECT DISTINCT day_ref, row_kind FROM measurements_active WHERE day_ref>=? AND day_ref<? AND row_kind LIKE 'sep_%'",
            (month_start, month_end),
        ).fetchall():
            if rk == "sep_oleo_detail": sep_p[day_ref_s]["oleo"] = True
            elif rk == "sep_gas_detail": sep_p[day_ref_s]["gas"] = True
            elif rk == "sep_agua_detail": sep_p[day_ref_s]["agua"] = True

        for day_number in range(1, _cal.monthrange(yr, mo)[1] + 1):
            day_ref = f"{yr:04d}-{mo:02d}-{day_number:02d}"
            mpfm_rows = _mpfm_rows_by_day.get(day_ref, [])

            sep_summary = sep_month_map.get(day_ref, _SEP_EMPTY_SUMMARY)

            # Build detailed coverage payload for the day
            s_flags = dict(sep_p[day_ref])
            if sep_summary.get("present"):
                if sep_summary.get("oil_t") is not None or sep_summary.get("oil_m3") is not None:
                    s_flags["oleo"] = True
                if sep_summary.get("gas_t") is not None or sep_summary.get("gas_sm3") is not None:
                    s_flags["gas"] = True
                if sep_summary.get("water_t") is not None or sep_summary.get("water_sm3") is not None:
                    s_flags["agua"] = True

            sep_details = {
                "oleo": {"tag": "20FT0247", "label": "Óleo", "present": bool(s_flags["oleo"])},
                "gas":  {"tag": "20FT0244", "label": "Gás",  "present": bool(s_flags["gas"])},
                "agua": {"tag": "20FT0251", "label": "Água", "present": bool(s_flags["agua"])},
                "present_count": sum(1 for k in ("oleo", "gas", "agua") if s_flags[k]),
            }

            topside_details = []
            for t in active_topside:
                key = (day_ref, t["bank"], t["tag"])
                d_has, hrs = mpfm_cov.get(key, (False, 0))
                topside_details.append({
                    "bank": t["bank"],
                    "tag": t["tag"],
                    "sensor_tag": t.get("sensor_tag") or "",
                    "daily": d_has,
                    "hourly": hrs > 0,
                    "hours": hrs,
                })

            subsea_details = []
            for s in active_subsea:
                key = (day_ref, s["bank"], s["tag"])
                d_has, hrs = mpfm_cov.get(key, (False, 0))
                subsea_details.append({
                    "bank": s["bank"],
                    "tag": s["tag"],
                    "sensor_tag": s.get("sensor_tag") or "",
                    "daily": d_has,
                    "hourly": hrs > 0,
                    "hours": hrs,
                })

            mpfm_details = {
                "topside": topside_details,
                "subsea": subsea_details,
                "topside_daily_cnt": sum(1 for item in topside_details if item["daily"]),
                "topside_hourly_cnt": sum(1 for item in topside_details if item["hourly"]),
                "subsea_daily_cnt": sum(1 for item in subsea_details if item["daily"]),
                "subsea_hourly_cnt": sum(1 for item in subsea_details if item["hourly"]),
            }

            if mpfm_rows:
                use_hourly = sum(row[5] for row in mpfm_rows) == 0
                oil_t = sum(row[10] if use_hourly else row[2] for row in mpfm_rows)
                gas_t = sum(row[11] if use_hourly else row[3] for row in mpfm_rows)
                water_t = sum(row[12] if use_hourly else row[4] for row in mpfm_rows)
                hc_t = sum(row[13] if use_hourly else row[5] for row in mpfm_rows)
                total_t = oil_t + gas_t + water_t
                oil_m3 = sum(row[15] if use_hourly else row[7] for row in mpfm_rows)
                gas_sm3 = sum(row[16] if use_hourly else row[8] for row in mpfm_rows)
                water_m3 = sum(row[17] if use_hourly else row[9] for row in mpfm_rows)
                banks_day = len({row[0] for row in mpfm_rows})
                max_hours = max((row[18] for row in mpfm_rows), default=0)

                sep_oil = sep_summary["oil_t"]
                sep_gas = sep_summary["gas_t"]
                sep_water = sep_summary["water_t"]
                sep_hc = sep_summary["hc_t"]
                sep_total = sep_summary["total_t"]
                sep_oil_m3 = sep_summary["oil_m3"]
                sep_gas_sm3 = sep_summary["gas_sm3"]
                sep_water_sm3 = sep_summary["water_sm3"]
                sep_present = bool(sep_summary.get("present"))
                sep_zero = sep_present and all(
                    abs(float(value or 0)) <= 0.000001
                    for value in (sep_oil, sep_gas, sep_water, sep_hc, sep_total, sep_oil_m3)
                )

                daily.append(
                    {
                        "day": day_ref,
                        "d": day_number,
                        "mpfm_oil": round(oil_t, 1),
                        "mpfm_gas": round(gas_t, 1),
                        "mpfm_water": round(water_t, 3),
                        "mpfm_hc": round(hc_t, 1),
                        "mpfm_total": round(total_t, 1),
                        "mpfm_oil_m3": round(oil_m3, 3),
                        "mpfm_gas_sm3": round(gas_sm3, 1),
                        "mpfm_water_m3": round(water_m3, 3),
                        "mpfm_oil_bbl": _conversion_payload(oil_m3, gas_sm3, conversion)["oil_bbl"],
                        "mpfm_gas_boe": _conversion_payload(oil_m3, gas_sm3, conversion)["gas_boe"],
                        "mpfm_boe": _conversion_payload(oil_m3, gas_sm3, conversion)["boe_total"],
                        "sep_oil": round(sep_oil, 1) if sep_oil is not None else None,
                        "sep_gas": round(sep_gas, 1) if sep_gas is not None else None,
                        "sep_water": round(sep_water, 3) if sep_water is not None else None,
                        "sep_hc": round(sep_hc, 1) if sep_hc is not None else None,
                        "sep_total": round(sep_total, 1) if sep_total is not None else None,
                        "sep_oil_m3": round(sep_oil_m3, 3) if sep_oil_m3 is not None else None,
                        "sep_gas_sm3": round(sep_gas_sm3, 3) if sep_gas_sm3 is not None else None,
                        "sep_water_sm3": round(sep_water_sm3, 3) if sep_water_sm3 is not None else None,
                        "sep_source": sep_summary["source"],
                        "sep_present": sep_present,
                        "sep_zero_day": sep_zero,
                        "sep_official_fluids": sep_summary.get("official_fluids", 0),
                        "sep_details": sep_details,
                        "mpfm_details": mpfm_details,
                        "banks": banks_day,
                        "max_hrs": max_hours,
                        "has_data": True,
                    }
                )
            else:
                sep_oil = sep_summary["oil_t"]
                sep_hc = sep_summary["hc_t"]
                sep_gas = sep_summary["gas_t"]
                sep_water = sep_summary["water_t"]
                sep_total = sep_summary["total_t"]
                has_sep = sep_hc is not None
                sep_oil_m3 = sep_summary["oil_m3"]
                sep_gas_sm3 = sep_summary["gas_sm3"]
                sep_water_sm3 = sep_summary["water_sm3"]
                sep_zero = has_sep and all(
                    abs(float(value or 0)) <= 0.000001
                    for value in (sep_oil, sep_gas, sep_water, sep_hc, sep_total, sep_oil_m3)
                )
                daily.append(
                    {
                        "day": day_ref,
                        "d": day_number,
                        "mpfm_oil": None,
                        "mpfm_gas": None,
                        "mpfm_water": None,
                        "mpfm_hc": None,
                        "mpfm_total": None,
                        "mpfm_oil_m3": None,
                        "mpfm_gas_sm3": None,
                        "mpfm_water_m3": None,
                        "mpfm_oil_bbl": None,
                        "mpfm_gas_boe": None,
                        "mpfm_boe": None,
                        "sep_oil": round(sep_oil, 1) if sep_oil is not None else None,
                        "sep_gas": round(sep_gas, 1) if sep_gas is not None else None,
                        "sep_water": round(sep_water, 3) if sep_water is not None else None,
                        "sep_hc": round(sep_hc, 1) if sep_hc is not None else None,
                        "sep_total": round(sep_total, 1) if sep_total is not None else None,
                        "sep_oil_m3": round(sep_oil_m3, 3) if sep_oil_m3 is not None else None,
                        "sep_gas_sm3": round(sep_gas_sm3, 3) if sep_gas_sm3 is not None else None,
                        "sep_water_sm3": round(sep_water_sm3, 3) if sep_water_sm3 is not None else None,
                        "sep_source": sep_summary["source"],
                        "sep_present": has_sep,
                        "sep_zero_day": sep_zero,
                        "sep_official_fluids": sep_summary.get("official_fluids", 0),
                        "sep_details": sep_details,
                        "mpfm_details": mpfm_details,
                        "banks": 0,
                        "max_hrs": 0,
                        "has_data": has_sep,
                    }
                )

        sep_days = sum(1 for row in daily if row.get("sep_hc") is not None)
        sep_hc = sum((row.get("sep_hc") or 0) for row in daily)
        sep_oil_m3 = sum((row.get("sep_oil_m3") or 0) for row in daily)

        total_alerts = cur.execute(
            "SELECT COUNT(*) FROM validation_issues WHERE day_ref>=? AND day_ref<?",
            (month_start, month_end),
        ).fetchone()[0]
        err_alerts = cur.execute(
            "SELECT COUNT(*) FROM validation_issues WHERE day_ref>=? AND day_ref<? AND severity='error'",
            (month_start, month_end),
        ).fetchone()[0]

        last_runs = [
            dict(row)
            for row in cur.execute(
                "SELECT id, started_at, source_type, source_ref, files_count, status FROM processing_runs ORDER BY id DESC LIMIT 5"
            ).fetchall()
        ]

        conn.close()
        result = {
            "month": month,
            "months_available": months_available,
            "production": production,
            "conversion": conversion,
            "criteria": criteria,
            "days_with_data": days_with_data,
            "by_bank": by_bank,
            "by_tag": by_tag,
            "bank_daily": bank_daily,
            "tag_daily": tag_daily,
            "daily": daily,
            "sep": {"days": sep_days, "hc_t": round(sep_hc or 0, 1), "oil_m3": round(sep_oil_m3 or 0, 3)},
            "alerts": {"total": total_alerts, "error": err_alerts},
            "last_runs": last_runs,
        }
        _month_summary_cache[cache_key] = (time.time(), result)
        return result

    _POCO_RISER_METRICS = [
        "MPFM corr Óleo (t)", "MPFM corr Gás (t)", "MPFM corr Água (t)",
        "MPFM corr HC (t)", "MPFM corr Total (t)",
        "PVT @20 vol Óleo (m³)", "PVT @20 vol Gás (Sm³)", "PVT @20 vol Água (m³)",
        "PVT @20 mass Óleo (t)", "PVT @20 mass Gás (t)", "PVT @20 mass Água (t)",
    ]

    def _normalize_poco_riser_source_kind(source_kind: str = "") -> str:
        return "hourly" if str(source_kind or "").strip().lower() == "hourly" else "daily"

    def _poco_riser_source_label(source_kind: str) -> str:
        return "Horários (somatório do dia)" if source_kind == "hourly" else "Diários"

    def _poco_riser_status(source_kind: str, present: bool, hours: int = 0) -> dict:
        if source_kind == "hourly":
            if not present:
                return {"label": "Sem horário", "badge": "warn", "missing": True, "hours": 0}
            label = f"{int(hours or 0)}/24 h"
            badge = "ok" if int(hours or 0) >= 24 else "warn"
            return {"label": label, "badge": badge, "missing": False, "hours": int(hours or 0)}
        if not present:
            return {"label": "Sem diário", "badge": "warn", "missing": True, "hours": None}
        return {"label": "Diário OK", "badge": "ok", "missing": False, "hours": None}

    def _poco_riser_diario_payload(date_from: str = "", date_to: str = "", source_kind: str = "daily"):
        source_kind = _normalize_poco_riser_source_kind(source_kind)
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute(
                "SELECT MAX(day_ref) FROM measurements_active WHERE row_kind=?",
                (source_kind,),
            ).fetchone()[0] or ""
        if not date_from:
            date_from = date_to
        if not date_from or not date_to:
            conn.close()
            return {
                "date_from": date_from,
                "date_to": date_to,
                "source_kind": source_kind,
                "source_label": _poco_riser_source_label(source_kind),
                "deviation_formula": "((Poço / Riser) - 1) × 100",
                "deviation_reference": "Riser",
                "limits": _summary_limits(),
                "pairs": [],
                "rows": [],
            }

        limits = _summary_limits()

        tags = set()
        for pair in CHART_FOCUS_PAIRS:
            tags.update(pair.get("subsea_tags") or [pair["subsea_tag"]])
            tags.update(pair.get("topside_tags") or [pair["topside_tag"]])

        tags = sorted(tags)
        tag_ph = ",".join("?" * len(tags))
        metric_ph = ",".join("?" * len(_POCO_RISER_METRICS))
        rows = cur.execute(
            f"""
            SELECT day_ref, tag, metric_name, SUM(metric_value)
            FROM measurements_active
            WHERE row_kind=? AND day_ref BETWEEN ? AND ?
              AND tag IN ({tag_ph}) AND metric_name IN ({metric_ph})
            GROUP BY day_ref, tag, metric_name
            """,
            [source_kind, date_from, date_to] + tags + _POCO_RISER_METRICS,
        ).fetchall()
        presence_rows = cur.execute(
            f"""
            SELECT day_ref, tag, COUNT(*) as row_count,
                   COUNT(DISTINCT CASE WHEN hour_ref IS NOT NULL THEN hour_ref END) as hours
            FROM measurements_active
            WHERE row_kind=? AND day_ref BETWEEN ? AND ?
              AND tag IN ({tag_ph})
            GROUP BY day_ref, tag
            """,
            [source_kind, date_from, date_to] + tags,
        ).fetchall()
        conn.close()

        aliases = {}
        for pair in CHART_FOCUS_PAIRS:
            for side in ("subsea", "topside"):
                canonical = pair[f"{side}_tag"]
                for alias in pair.get(f"{side}_tags") or [canonical]:
                    aliases[alias] = canonical

        data = {}
        for day_ref, tag, metric_name, value in rows:
            canonical_tag = aliases.get(tag, tag)
            data.setdefault((day_ref, canonical_tag), {})[metric_name] = value

        presence = {}
        for day_ref, tag, row_count, hours in presence_rows:
            canonical_tag = aliases.get(tag, tag)
            current = presence.setdefault((day_ref, canonical_tag), {"present": False, "hours": 0})
            current["present"] = current["present"] or bool(row_count)
            current["hours"] = max(current["hours"], int(hours or 0))

        def _v(day_ref, tag, metric_name):
            return (data.get((day_ref, tag)) or {}).get(metric_name)

        days = []
        try:
            d0 = datetime.strptime(date_from, "%Y-%m-%d")
            d1 = datetime.strptime(date_to, "%Y-%m-%d")
            cursor_d = d0
            while cursor_d <= d1:
                days.append(cursor_d.strftime("%Y-%m-%d"))
                cursor_d += timedelta(days=1)
        except ValueError:
            days = []

        result_rows = []
        for pair in CHART_FOCUS_PAIRS:
            poco_tag = pair["subsea_tag"]
            riser_tag = pair["topside_tag"]
            poco_label = pair["subsea_label"].replace("Subsea · ", "")
            riser_label = pair["topside_label"].replace("Topside · ", "")
            for day_ref in days:
                poco_presence = presence.get((day_ref, poco_tag), {"present": False, "hours": 0})
                riser_presence = presence.get((day_ref, riser_tag), {"present": False, "hours": 0})
                has_poco = bool(poco_presence["present"])
                has_riser = bool(riser_presence["present"])
                if not has_poco and not has_riser:
                    continue

                oleo_corr = _v(day_ref, poco_tag, "MPFM corr Óleo (t)")
                gas_corr = _v(day_ref, poco_tag, "MPFM corr Gás (t)")
                agua_corr = _v(day_ref, poco_tag, "MPFM corr Água (t)")
                hc_corr = _v(day_ref, poco_tag, "MPFM corr HC (t)")
                total_corr = _v(day_ref, poco_tag, "MPFM corr Total (t)")

                oleo_vol20 = _v(day_ref, poco_tag, "PVT @20 vol Óleo (m³)")
                gas_vol20 = _v(day_ref, poco_tag, "PVT @20 vol Gás (Sm³)")
                agua_vol20 = _v(day_ref, poco_tag, "PVT @20 vol Água (m³)")
                oleo_m20 = _v(day_ref, poco_tag, "PVT @20 mass Óleo (t)")
                gas_m20 = _v(day_ref, poco_tag, "PVT @20 mass Gás (t)")
                agua_m20 = _v(day_ref, poco_tag, "PVT @20 mass Água (t)")

                riser_oleo_corr = _v(day_ref, riser_tag, "MPFM corr Óleo (t)")
                riser_gas_corr = _v(day_ref, riser_tag, "MPFM corr Gás (t)")
                riser_agua_corr = _v(day_ref, riser_tag, "MPFM corr Água (t)")
                riser_hc = _v(day_ref, riser_tag, "MPFM corr HC (t)")
                riser_total = _v(day_ref, riser_tag, "MPFM corr Total (t)")
                riser_oleo_vol20 = _v(day_ref, riser_tag, "PVT @20 vol Óleo (m³)")
                riser_gas_vol20 = _v(day_ref, riser_tag, "PVT @20 vol Gás (Sm³)")
                riser_agua_vol20 = _v(day_ref, riser_tag, "PVT @20 vol Água (m³)")
                riser_oleo_m20 = _v(day_ref, riser_tag, "PVT @20 mass Óleo (t)")
                riser_gas_m20 = _v(day_ref, riser_tag, "PVT @20 mass Gás (t)")
                riser_agua_m20 = _v(day_ref, riser_tag, "PVT @20 mass Água (t)")

                desvio_hc = round(((hc_corr / riser_hc) - 1) * 100, 2) if hc_corr is not None and riser_hc not in (None, 0) else None
                desvio_total = round(((total_corr / riser_total) - 1) * 100, 2) if total_corr is not None and riser_total not in (None, 0) else None

                result_rows.append({
                    "day": day_ref,
                    "pair_key": pair["key"],
                    "poco_tag": poco_tag,
                    "poco_label": poco_label,
                    "riser_tag": riser_tag,
                    "riser_label": riser_label,
                    "source_kind": source_kind,
                    "has_counterpart": has_poco and has_riser,
                    "poco_source_status": _poco_riser_status(source_kind, has_poco, poco_presence["hours"]),
                    "riser_source_status": _poco_riser_status(source_kind, has_riser, riser_presence["hours"]),
                    "poco_mpfm": {
                        "oil_corr_t": oleo_corr,
                        "gas_corr_t": gas_corr,
                        "water_corr_t": agua_corr,
                        "hc_corr_t": hc_corr,
                        "total_corr_t": total_corr,
                    },
                    "riser_mpfm": {
                        "oil_corr_t": riser_oleo_corr,
                        "gas_corr_t": riser_gas_corr,
                        "water_corr_t": riser_agua_corr,
                        "hc_corr_t": riser_hc,
                        "total_corr_t": riser_total,
                    },
                    "reference_20c": {
                        "poco": {
                            "oil_vol20_m3": oleo_vol20,
                            "gas_vol20_sm3": gas_vol20,
                            "water_vol20_m3": agua_vol20,
                            "oil_mass20_t": oleo_m20,
                            "gas_mass20_t": gas_m20,
                            "water_mass20_t": agua_m20,
                        },
                        "riser": {
                            "oil_vol20_m3": riser_oleo_vol20,
                            "gas_vol20_sm3": riser_gas_vol20,
                            "water_vol20_m3": riser_agua_vol20,
                            "oil_mass20_t": riser_oleo_m20,
                            "gas_mass20_t": riser_gas_m20,
                            "water_mass20_t": riser_agua_m20,
                        },
                    },
                    "oleo": {"massa_corr_t": oleo_corr, "vol20_m3": oleo_vol20, "massa20_t": oleo_m20},
                    "gas": {"massa_corr_t": gas_corr, "vol20_sm3": gas_vol20, "massa20_t": gas_m20},
                    "agua": {"massa_corr_t": agua_corr, "vol20_m3": agua_vol20, "massa20_t": agua_m20},
                    "hc_corr_t": hc_corr,
                    "total_corr_t": total_corr,
                    "riser_hc_corr_t": riser_hc,
                    "riser_total_corr_t": riser_total,
                    "desvio_hc_pct": desvio_hc,
                    "desvio_total_pct": desvio_total,
                    "alerta_hc": desvio_hc is not None and abs(desvio_hc) > limits["hc_pct"],
                    "alerta_total": desvio_total is not None and abs(desvio_total) > limits["total_pct"],
                })

        result_rows.sort(key=lambda item: (item["day"], item["pair_key"]))
        return {
            "date_from": date_from,
            "date_to": date_to,
            "source_kind": source_kind,
            "source_label": _poco_riser_source_label(source_kind),
            "deviation_formula": "((Poço / Riser) - 1) × 100",
            "deviation_reference": "Riser",
            "limits": limits,
            "pairs": [
                {"key": p["key"], "poco": p["subsea_label"].replace("Subsea · ", ""), "riser": p["topside_label"].replace("Topside · ", "")}
                for p in CHART_FOCUS_PAIRS
            ],
            "rows": result_rows,
        }

    @app.get("/api/ops/poco-riser-diario")
    def api_poco_riser_diario(date_from: str = "", date_to: str = "", source_kind: str = "daily"):
        return _poco_riser_diario_payload(date_from, date_to, source_kind)

    @app.get("/api/ops/poco-riser-diario/export-excel")
    def api_poco_riser_diario_export_excel(date_from: str = "", date_to: str = "", source_kind: str = "daily"):
        import io

        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from fastapi.responses import StreamingResponse

        payload = _poco_riser_diario_payload(date_from, date_to, source_kind)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Massas Corrigidas"

        headers = [
            "Data", "Poço", "Riser", "Arquivo Poço", "Arquivo Riser",
            "Poço - Óleo corr. (t)", "Poço - Gás corr. (t)", "Poço - Água corr. (t)", "Poço - HC corr. (t)", "Poço - Total corr. (t)",
            "Riser - Óleo corr. (t)", "Riser - Gás corr. (t)", "Riser - Água corr. (t)", "Riser - HC corr. (t)", "Riser - Total corr. (t)",
            "% Desvio HC", "% Desvio Total",
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="002060")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        alert_fill = PatternFill("solid", fgColor="F8CBAD")
        thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

        for row in payload["rows"]:
            ws.append([
                row["day"], row["poco_label"], row["riser_label"],
                row["poco_source_status"]["label"], row["riser_source_status"]["label"],
                row["poco_mpfm"]["oil_corr_t"], row["poco_mpfm"]["gas_corr_t"], row["poco_mpfm"]["water_corr_t"], row["poco_mpfm"]["hc_corr_t"], row["poco_mpfm"]["total_corr_t"],
                row["riser_mpfm"]["oil_corr_t"], row["riser_mpfm"]["gas_corr_t"], row["riser_mpfm"]["water_corr_t"], row["riser_mpfm"]["hc_corr_t"], row["riser_mpfm"]["total_corr_t"],
                row["desvio_hc_pct"], row["desvio_total_pct"],
            ])
            excel_row = ws.max_row
            for ci in range(1, len(headers) + 1):
                ws.cell(excel_row, ci).border = thin_border
            if row["alerta_hc"]:
                ws.cell(excel_row, 16).fill = alert_fill
                ws.cell(excel_row, 16).font = Font(bold=True, color="9C0006")
            if row["alerta_total"]:
                ws.cell(excel_row, 17).fill = alert_fill
                ws.cell(excel_row, 17).font = Font(bold=True, color="9C0006")

        for ci, header in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ws.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        ws.freeze_panes = "A2"

        ref_ws = wb.create_sheet("Referencia 20C")
        ref_headers = [
            "Data", "Poço", "Riser",
            "Poço - Óleo vol20 (m³)", "Poço - Óleo mass20 (t)",
            "Poço - Gás vol20 (Sm³)", "Poço - Gás mass20 (t)",
            "Poço - Água vol20 (m³)", "Poço - Água mass20 (t)",
            "Riser - Óleo vol20 (m³)", "Riser - Óleo mass20 (t)",
            "Riser - Gás vol20 (Sm³)", "Riser - Gás mass20 (t)",
            "Riser - Água vol20 (m³)", "Riser - Água mass20 (t)",
        ]
        ref_ws.append(ref_headers)
        for cell in ref_ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in payload["rows"]:
            ref = row["reference_20c"]
            ref_ws.append([
                row["day"], row["poco_label"], row["riser_label"],
                ref["poco"]["oil_vol20_m3"], ref["poco"]["oil_mass20_t"],
                ref["poco"]["gas_vol20_sm3"], ref["poco"]["gas_mass20_t"],
                ref["poco"]["water_vol20_m3"], ref["poco"]["water_mass20_t"],
                ref["riser"]["oil_vol20_m3"], ref["riser"]["oil_mass20_t"],
                ref["riser"]["gas_vol20_sm3"], ref["riser"]["gas_mass20_t"],
                ref["riser"]["water_vol20_m3"], ref["riser"]["water_mass20_t"],
            ])
            excel_row = ref_ws.max_row
            for ci in range(1, len(ref_headers) + 1):
                ref_ws.cell(excel_row, ci).border = thin_border
        for ci, header in enumerate(ref_headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ref_ws.column_dimensions[col_letter].width = max(len(header) + 2, 14)
        ref_ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"poco_riser_{payload['source_kind']}_{payload['date_from']}_{payload['date_to']}.xlsx"
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    @app.get("/api/ops/desvio-mensal")
    def api_desvio_mensal(month: str = ""):
        conn = db_conn()
        cur = conn.cursor()
        if not month:
            last = cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE metric_name='Desvio HC (%)'").fetchone()[0] or ""
            month = last[:7] if last else ""
        if not month:
            conn.close()
            return {"month": month, "days": [], "limite_hc": 10.0, "limite_total": 7.0}

        import calendar as _cal

        yr, mo = int(month[:4]), int(month[5:7])
        m_from = f"{month}-01"
        m_to = f"{month}-{_cal.monthrange(yr, mo)[1]:02d}"

        lim_hc = cur.execute("SELECT limite_hc_pct FROM pvt_params ORDER BY id DESC LIMIT 1").fetchone()
        lim_total = cur.execute("SELECT limite_total_pct FROM pvt_params ORDER BY id DESC LIMIT 1").fetchone()
        limite_hc = lim_hc[0] if lim_hc else 10.0
        limite_total = lim_total[0] if lim_total else 7.0

        rows = cur.execute(
            """
            SELECT day_ref, bank, tag, metric_name,
                   AVG(metric_value) as avg_val, COUNT(*) as n_runs
            FROM measurements_active
            WHERE metric_name IN ('Desvio HC (%)', 'Desvio Total (%)')
              AND day_ref BETWEEN ? AND ?
              AND metric_value > -99
            GROUP BY day_ref, bank, tag, metric_name
            ORDER BY day_ref, bank, tag, metric_name
            """,
            (m_from, m_to),
        ).fetchall()

        from collections import defaultdict

        pivot = defaultdict(dict)
        for day, bank, tag, metric, avg_val, n in rows:
            key = (day, bank, tag)
            if "HC" in metric:
                pivot[key]["desvio_hc"] = round(avg_val, 2)
                pivot[key]["n_runs"] = n
            else:
                pivot[key]["desvio_total"] = round(avg_val, 2)

        result = []
        for (day, bank, tag), vals in sorted(pivot.items()):
            dev_hc = vals.get("desvio_hc")
            dev_total = vals.get("desvio_total")
            result.append(
                {
                    "day": day,
                    "bank": bank,
                    "tag": tag,
                    "desvio_hc": dev_hc,
                    "desvio_total": dev_total,
                    "ok_hc": abs(dev_hc) <= limite_hc if dev_hc is not None else None,
                    "ok_total": abs(dev_total) <= limite_total if dev_total is not None else None,
                    "n_runs": vals.get("n_runs", 1),
                }
            )

        conn.close()
        return {"month": month, "days": result, "limite_hc": limite_hc, "limite_total": limite_total}

    @app.get("/api/ops/processing-history")
    def api_processing_history(limit: int = 30):
        conn = db_conn()
        try:
            cur = conn.cursor()

            # Subquery limita runs ANTES do JOIN, evitando full scan de files_imported
            sql = """
                SELECT
                    pr.id, pr.started_at, pr.finished_at, pr.source_type,
                    pr.source_ref, pr.files_count, pr.status, pr.notes_json,
                    fi.filename, fi.file_type, fi.content_date, fi.processed_ok, fi.message
                FROM (
                    SELECT * FROM processing_runs ORDER BY id DESC LIMIT ?
                ) pr
                LEFT JOIN files_imported fi ON fi.run_id = pr.id
                ORDER BY pr.id DESC, fi.id
            """
            rows = cur.execute(sql, (limit,)).fetchall()

            # Agrupa resultados por run_id
            runs_dict = {}
            for row in rows:
                run_id = row[0]
                if run_id not in runs_dict:
                    runs_dict[run_id] = {
                        "id": run_id,
                        "started_at": row[1],
                        "finished_at": row[2],
                        "source_type": row[3],
                        "source_ref": row[4],
                        "files_count": row[5],
                        "status": row[6],
                        "notes_json": row[7],
                        "files": [],
                        "months_updated": set()
                    }

                # Adiciona file se existir (LEFT JOIN pode retornar NULL)
                if row[8]:  # filename exists
                    file_data = {
                        "filename": row[8],
                        "file_type": row[9],
                        "content_date": row[10],
                        "processed_ok": row[11],
                        "message": row[12]
                    }
                    runs_dict[run_id]["files"].append(file_data)

                    # Coleta meses
                    if row[10] and row[10][:4] not in ("", "0000"):
                        runs_dict[run_id]["months_updated"].add(row[10][:7])

            # Converte para lista
            runs = []
            for run_data in runs_dict.values():
                run_data["months_updated"] = sorted(run_data["months_updated"])
                runs.append(run_data)

            return {"runs": runs}
        finally:
            conn.close()

    @app.get("/api/ops/mpfm-monitoring")
    def api_ops_mpfm_monitoring(
        month: str = "",
        bank: str = "",
        tag: str = "",
        meter_type: str = "",
        event_status: str = "",
        only_outside_limits: str = "0",
    ):
        if not month:
            conn = db_conn()
            month = conn.execute(
                "SELECT MAX(substr(day_ref,1,7)) FROM measurements_active WHERE day_ref<>'' AND bank<>'' AND bank<>'SEP'"
            ).fetchone()[0] or ""
            conn.close()
        return list_monitoring_rows(
            db_conn,
            month=month,
            bank=bank,
            tag=tag,
            meter_type=meter_type,
            event_status=event_status,
            only_outside_limits=str(only_outside_limits).strip().lower() in {"1", "true", "yes", "sim", "on"},
            load_cadastro_fn=load_cadastro,
            normalize_tag_name=normalize_tag_name,
        )

    @app.post("/api/ops/mpfm-monitoring")
    async def api_ops_mpfm_monitoring_upsert(request: Request):
        payload = dict(await request.json())
        payload["meter_type"] = normalize_meter_type(payload.get("meter_type") or "")
        for field in MONITORING_BOOL_FIELDS:
            if field in payload:
                payload[field] = _normalize_monitoring_bool(payload.get(field))
        try:
            item_id = upsert_monitoring_annotation(db_conn, payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        _invalidate_cache("ops_mpfm_monitoring")
        _invalidate_cache("monitoring_summary")
        return {"ok": True, "id": item_id}

    @app.delete("/api/ops/mpfm-monitoring/{item_id}")
    def api_ops_mpfm_monitoring_delete(item_id: int):
        delete_monitoring_annotation(db_conn, item_id)
        _invalidate_cache("ops_mpfm_monitoring")
        _invalidate_cache("monitoring_summary")
        return {"ok": True, "id": item_id}

    @app.get("/api/ops/mpfm-data")
    def api_ops_mpfm_data(
        date_from: str = "",
        date_to: str = "",
        row_kind: str = "daily",
        bank: str = "",
        metric: str = "",
        tag: str = "",
        q: str = "",
        limit: int = 500,
        offset: int = 0  # ✅ Novo parâmetro de paginação
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to

        # ✅ OTIMIZADO: Cache de metadados (evita 3 queries repetidas)
        from cache_manager import cached

        @cached(ttl=1800, key_prefix='mpfm_metadata')
        def get_dropdown_metadata():
            banks = [r[0] for r in cur.execute("SELECT DISTINCT bank FROM measurements_active WHERE bank<>'' ORDER BY bank").fetchall()]
            metrics = [r[0] for r in cur.execute("SELECT DISTINCT metric_name FROM measurements_active ORDER BY metric_name").fetchall()]
            tags = [r[0] for r in cur.execute("SELECT DISTINCT tag FROM measurements_active WHERE tag<>'' ORDER BY tag").fetchall()]
            return {"banks": banks, "metrics": metrics, "tags": tags}

        # Query principal com COUNT para paginação
        sql_count = "SELECT COUNT(*) FROM measurements_active WHERE day_ref BETWEEN ? AND ?"
        sql_data = """
          SELECT id, day_ref, hour_ref, bank, tag, sheet_name, row_kind, metric_name, metric_value, metric_unit, source_file
          FROM measurements_active WHERE day_ref BETWEEN ? AND ?
        """
        params = [date_from, date_to]

        if row_kind in ("hourly", "daily", "recon"):
            sql_count += " AND row_kind=?"
            sql_data += " AND row_kind=?"
            params.append(row_kind)
        if bank:
            sql_count += " AND bank=?"
            sql_data += " AND bank=?"
            params.append(bank)
        if metric:
            sql_count += " AND metric_name=?"
            sql_data += " AND metric_name=?"
            params.append(metric)
        if tag:
            tag_normalized = normalize_tag_name(tag)
            tag_filter = """
             AND (
               tag=?
               OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
             )
            """
            sql_count += tag_filter
            sql_data += tag_filter
            params += [tag, tag_normalized]
        if q:
            normalized_q = normalize_tag_name(q)
            q_filter = """
             AND (
               bank LIKE ?
               OR tag LIKE ?
               OR metric_name LIKE ?
               OR source_file LIKE ?
               OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '') LIKE ?
               OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(source_file,'')), ' ', ''), '-', ''), '_', '') LIKE ?
             )
            """
            sql_count += q_filter
            sql_data += q_filter
            qq = f"%{q}%"
            params += [qq, qq, qq, qq, f"%{normalized_q}%", f"%{normalized_q}%"]

        # ✅ OTIMIZADO: Paginação real com LIMIT e OFFSET
        sql_data += " ORDER BY day_ref DESC, COALESCE(hour_ref,-1) DESC, bank, tag, metric_name LIMIT ? OFFSET ?"
        params_data = params + [limit, offset]

        # Executa queries
        total = cur.execute(sql_count, params).fetchone()[0]
        rows = [dict(r) for r in cur.execute(sql_data, params_data).fetchall()]

        # Metadados cacheados
        metadata = get_dropdown_metadata()

        conn.close()

        for row in rows:
            source_name = str(row.get("source_file") or "")
            source_lower = source_name.lower()
            is_adjusted = source_lower.startswith("manual_adjustment:")
            row["is_adjusted"] = bool(is_adjusted)
            row["adjustment_source"] = source_name.split(":", 1)[1] if is_adjusted and ":" in source_name else ""
            row["source_kind"] = "ajustado" if is_adjusted else "manual" if source_lower.startswith("manual") else "arquivo"

        # ✅ OTIMIZADO: Retorna informação de paginação
        return {
            "rows": rows,
            "banks": metadata["banks"],
            "metrics": metadata["metrics"],
            "tags": metadata["tags"],
            "date_from": date_from,
            "date_to": date_to,
            "pagination": {
                "offset": offset,
                "limit": limit,
                "total": total,
                "has_more": offset + limit < total,
                "page": offset // limit + 1 if limit > 0 else 1,
                "total_pages": (total + limit - 1) // limit if limit > 0 else 1
            }
        }

    @app.get("/api/ops/sep-data")
    @cached(ttl=120, key_prefix="ops_sep_data")
    def api_ops_sep_data(
        date_from: str = "",
        date_to: str = "",
        limit: int = 200,
        offset: int = 0,
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = (
                cur.execute("SELECT MAX(content_date) FROM files_imported WHERE ext='txt'").fetchone()[0]
                or cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE row_kind='sep' AND bank='SEP'").fetchone()[0]
                or ""
            )
        if not date_from:
            date_from = date_to
        align_map = {}
        for a in cur.execute(
            """
            SELECT production_date, GROUP_CONCAT(bank, ', ') AS banks
            FROM sep_alignments
            WHERE is_active=1 AND production_date BETWEEN ? AND ?
            GROUP BY production_date
            """,
            (date_from, date_to),
        ).fetchall():
            align_map[a["production_date"]] = a["banks"] or ""

        # Paginação: conta total de arquivos txt no período
        total_rows = cur.execute(
            "SELECT COUNT(*) FROM files_imported WHERE ext='txt' AND content_date BETWEEN ? AND ?",
            (date_from, date_to),
        ).fetchone()[0]

        rows = []
        for r in cur.execute(
            """
            SELECT content_date, filename, file_type, unit_code, meter_id, location, message
            FROM files_imported WHERE ext='txt' AND content_date BETWEEN ? AND ?
            ORDER BY content_date DESC, file_type, filename LIMIT ? OFFSET ?
            """,
            (date_from, date_to, limit, offset),
        ).fetchall():
            d = dict(r)
            d["aligned_banks"] = align_map.get(d["content_date"], "")
            d["sep_status"] = "aplicado" if d["aligned_banks"] else "extraido"
            rows.append(d)

        days = []
        for r in cur.execute(
            """
            SELECT content_date,
                   SUM(CASE WHEN file_type='sep_oleo' THEN 1 ELSE 0 END) oleo,
                   SUM(CASE WHEN file_type='sep_agua' THEN 1 ELSE 0 END) agua,
                   SUM(CASE WHEN file_type='sep_gas' THEN 1 ELSE 0 END) gas
            FROM files_imported WHERE ext='txt' AND content_date BETWEEN ? AND ?
            GROUP BY content_date ORDER BY content_date DESC LIMIT ? OFFSET ?
            """,
            (date_from, date_to, limit, offset),
        ).fetchall():
            d = dict(r)
            d["aligned_banks"] = align_map.get(d["content_date"], "")
            trio_ok = bool(d["oleo"] and d["agua"] and d["gas"])
            d["status"] = "ok" if trio_ok else "attention"
            d["aplicacao"] = "aplicado" if d["aligned_banks"] else "extraido"
            d["status_label"] = ("Trio completo" if trio_ok else "Trio incompleto") + (" · aplicado" if d["aligned_banks"] else " · extraído")
            days.append(d)

        if not rows and not days:
            sep_rows = [
                dict(r)
                for r in cur.execute(
                    """
                    SELECT day_ref, metric_name, metric_value
                    FROM measurements_active
                    WHERE row_kind='sep' AND bank='SEP' AND COALESCE(is_official,1)=1
                      AND day_ref BETWEEN ? AND ?
                    ORDER BY day_ref DESC, metric_name LIMIT ? OFFSET ?
                    """,
                    (date_from, date_to, limit, offset),
                ).fetchall()
            ]
            sep_by_day = {}
            for item in sep_rows:
                sep_by_day.setdefault(item["day_ref"], {})[item["metric_name"]] = item["metric_value"]

            recovered_days = sorted(sep_by_day.keys(), reverse=True)
            for day_ref in recovered_days:
                metrics = sep_by_day.get(day_ref, {})
                aligned_banks = align_map.get(day_ref, "")
                zero_day = all(
                    abs(float(metrics.get(metric) or 0)) <= 0.000001
                    for metric in ("oil_t", "gas_t", "water_t", "hc_t", "total_t", "oil_m3")
                )
                days.append(
                    {
                        "content_date": day_ref,
                        "oleo": 1,
                        "agua": 1,
                        "gas": 1,
                        "aligned_banks": aligned_banks,
                        "status": "ok",
                        "aplicacao": "aplicado" if aligned_banks else "reconstituido",
                        "status_label": ("Dia zerado" if zero_day else "Resumo SEP recuperado do Excel")
                        + (" · aplicado" if aligned_banks else " · sem TXT original"),
                        "recovered_from_excel": True,
                    }
                )
                rows.append(
                    {
                        "content_date": day_ref,
                        "filename": "BASE_UNICA_MES",
                        "file_type": "recuperado_do_excel",
                        "unit_code": "SEP",
                        "meter_id": "SEP",
                        "location": "Recuperado do Excel",
                        "message": "Histórico bruto TXT indisponível; origem recomposta da BASE_UNICA_MES.",
                        "aligned_banks": aligned_banks,
                        "sep_status": "aplicado" if aligned_banks else "reconstituido",
                        "recovered_from_excel": True,
                    }
                )
            rows = rows[:limit]
            total_rows = len(rows)
        conn.close()
        return {
            "rows": rows,
            "days": days,
            "date_from": date_from,
            "date_to": date_to,
            "fallback_used": bool(days and any(d.get("recovered_from_excel") for d in days)),
            "pagination": {"limit": limit, "offset": offset, "total": total_rows},
        }

    @app.get("/api/ops/alerts")
    @cached(ttl=90, key_prefix="ops_alerts")
    def api_ops_alerts(
        date_from: str = "",
        date_to: str = "",
        severity: str = "",
        limit: int = 200,
        offset: int = 0,
    ):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = (
                cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE day_ref LIKE '____-__-__'").fetchone()[0]
                or cur.execute("SELECT MAX(substr(created_at,1,10)) FROM validation_issues").fetchone()[0]
                or ""
            )
        if not date_from:
            date_from = date_to

        # Paginação: aplicada depois de juntar issues e devs (simplificado por filtro dinâmico)
        vi_rows = []
        for r in cur.execute(
            """
            SELECT created_at, excel_file, issue_type, severity, ref_key,
                   COALESCE(day_ref,'') as day_ref, details
            FROM validation_issues
            WHERE issue_type <> 'sep_duplicate_candidate'
            ORDER BY id DESC
            """,
        ).fetchall():
            item = dict(r)
            item["day_ref"] = normalize_validation_issue_day_ref(item.get("day_ref", ""), item.get("created_at", ""))
            if item["day_ref"] and item["day_ref"] < date_from:
                continue
            if item["day_ref"] and item["day_ref"] > date_to:
                continue
            vi_rows.append(item)

        dev_rows = []
        hc_limit = 10.0
        total_limit = 7.0
        for r in cur.execute(
            """
            SELECT day_ref, hour_ref, bank, tag, metric_name, metric_value
            FROM measurements_active
            WHERE day_ref BETWEEN ? AND ?
              AND metric_name IN ('Desvio HC (%)', 'Desvio Total (%)')
              AND metric_value IS NOT NULL
              AND ABS(metric_value) > ?
            ORDER BY day_ref DESC, COALESCE(hour_ref,-1) DESC, bank
            """,
            (date_from, date_to, min(hc_limit, total_limit)),
        ).fetchall():
            d = dict(r)
            is_hc = d["metric_name"] == "Desvio HC (%)"
            limit_pct = hc_limit if is_hc else total_limit
            if abs(d["metric_value"] or 0) <= limit_pct:
                continue
            sev = "error" if abs(d["metric_value"]) > limit_pct * 2 else "warn"
            hr = f" h{int(d['hour_ref']):02d}" if d.get("hour_ref") is not None else ""
            dev_rows.append(
                {
                    "created_at": d["day_ref"],
                    "excel_file": f"{d['bank']}/{d['tag']}",
                    "issue_type": "desvio_alto",
                    "severity": sev,
                    "ref_key": f"{d['bank']}/{d['tag']}{hr}",
                    "day_ref": d["day_ref"],
                    "details": f"{d['metric_name']}: {d['metric_value']:+.1f}% (limite ±{limit_pct:.0f}%)",
                }
            )

        all_rows = vi_rows + dev_rows
        if severity:
            all_rows = [r for r in all_rows if r["severity"] == severity]
        all_rows.sort(key=lambda row: (row.get("day_ref") or row.get("created_at") or "", row.get("created_at") or ""), reverse=True)
        total_rows = len(all_rows)
        all_rows = all_rows[offset : offset + limit]
        conn.close()
        crit = sum(1 for r in all_rows if r["severity"] == "error")
        warn = sum(1 for r in all_rows if r["severity"] == "warn")
        info = sum(1 for r in all_rows if r["severity"] == "info")
        return {
            "rows": all_rows,
            "date_from": date_from,
            "date_to": date_to,
            "counts": {"error": crit, "warn": warn, "info": info},
            "pagination": {"limit": limit, "offset": offset, "total": total_rows},
        }

    @app.get("/api/ops/chart-meta")
    @cached(ttl=120, key_prefix="ops_chart_meta")
    def api_ops_chart_meta(date_from: str = "", date_to: str = "", row_kind: str = "daily", bank: str = "", tag: str = ""):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        effective_kind = row_kind if row_kind in ("hourly", "daily", "recon") else "daily"
        if not date_to:
            date_to = (
                cur.execute("SELECT MAX(day_ref) FROM measurements_active WHERE row_kind=?", (effective_kind,)).fetchone()[0]
                or cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0]
                or ""
            )
        if not date_from:
            date_from = date_to
        normalized_tag = normalize_tag_name(tag) if tag else ""

        if effective_kind == "recon":
            bank_sql = """
                SELECT DISTINCT bank
                FROM measurements_active
                WHERE row_kind='recon'
                  AND COALESCE(is_official,1)=1
                  AND bank<>''
                ORDER BY bank
            """
            banks = [{"value": row[0], "label": _chart_format_bank_label(row[0])} for row in cur.execute(bank_sql).fetchall()]
            tag_sql = """
                SELECT DISTINCT tag
                FROM measurements_active
                WHERE row_kind='recon'
                  AND COALESCE(is_official,1)=1
            """
            tag_params = []
            if bank:
                tag_sql += " AND bank=?"
                tag_params.append(bank)
            tag_sql += " ORDER BY tag"
            tags = [
                {"value": row[0], "label": _chart_format_tag_label(row[0])}
                for row in cur.execute(tag_sql, tag_params).fetchall()
                if row[0]
            ]
            base_sql = """
                SELECT DISTINCT metric_name
                FROM measurements_active
                WHERE row_kind='recon'
                  AND COALESCE(is_official,1)=1
            """
            base_params = []
            current_sql = base_sql + " AND day_ref BETWEEN ? AND ?"
            current_params = [date_from, date_to]
            if bank:
                base_sql += " AND bank=?"
                current_sql += " AND bank=?"
                base_params.append(bank)
                current_params.append(bank)
            if tag:
                tag_clause = """
                    AND (
                        tag=?
                        OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                    )
                """
                base_sql += tag_clause
                current_sql += tag_clause
                base_params.extend([tag, normalized_tag])
                current_params.extend([tag, normalized_tag])
            base_sql += " ORDER BY metric_name"
            current_sql += " ORDER BY metric_name"
            all_metrics = {row[0] for row in cur.execute(base_sql, base_params).fetchall() if row[0]}
            current_metrics = {row[0] for row in cur.execute(current_sql, current_params).fetchall() if row[0]}
        else:
            regular_kind = "hourly" if effective_kind == "hourly" else "daily"
            banks = [
                {"value": row[0], "label": _chart_format_bank_label(row[0])}
                for row in cur.execute(
                    f"""
                    SELECT DISTINCT bank
                    FROM measurements_active
                    WHERE row_kind=?
                      AND COALESCE(is_official,1)=1
                      AND bank<>''
                    ORDER BY bank
                    """,
                    (regular_kind,),
                ).fetchall()
                if row[0]
            ]
            sep_exists = bool(
                cur.execute(
                    """
                    SELECT 1
                    FROM measurements_active
                    WHERE row_kind IN ({})
                      AND bank='SEP'
                      AND COALESCE(is_official,1)=1
                    LIMIT 1
                    """.format(",".join("?" * len(_chart_sep_context_row_kinds(effective_kind)))),
                    _chart_sep_context_row_kinds(effective_kind),
                ).fetchone()
            )
            if sep_exists:
                banks.append({"value": "SEP", "label": _chart_format_bank_label("SEP")})
            banks = sorted(banks, key=lambda item: item["value"])

            if bank == "SEP":
                row_kinds = _chart_sep_context_row_kinds(effective_kind)
                q_marks = ",".join("?" * len(row_kinds))
                tag_sql = f"""
                    SELECT DISTINCT tag
                    FROM measurements_active
                    WHERE row_kind IN ({q_marks})
                      AND bank='SEP'
                      AND COALESCE(is_official,1)=1
                """
                tag_params = list(row_kinds)
                if tag:
                    tag_sql += """
                        AND (
                            tag=?
                            OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                        )
                    """
                    tag_params.extend([tag, normalized_tag])
                tag_sql += " ORDER BY tag"
                tags = [
                    {"value": row[0], "label": _chart_format_sep_tag_label(row[0])}
                    for row in cur.execute(tag_sql, tag_params).fetchall()
                    if row[0]
                ]

                metric_sql = f"""
                    SELECT DISTINCT metric_name
                    FROM measurements_active
                    WHERE row_kind IN ({q_marks})
                      AND bank='SEP'
                      AND COALESCE(is_official,1)=1
                """
                metric_params_all = list(row_kinds)
                metric_sql_current = metric_sql + " AND day_ref BETWEEN ? AND ?"
                metric_params_current = list(row_kinds) + [date_from, date_to]
                if tag:
                    tag_clause = """
                        AND (
                            tag=?
                            OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                        )
                    """
                    metric_sql += tag_clause
                    metric_sql_current += tag_clause
                    metric_params_all.extend([tag, normalized_tag])
                    metric_params_current.extend([tag, normalized_tag])
                metric_sql += " ORDER BY metric_name"
                metric_sql_current += " ORDER BY metric_name"
                all_metrics = {row[0] for row in cur.execute(metric_sql, metric_params_all).fetchall() if row[0]}
                current_metrics = {row[0] for row in cur.execute(metric_sql_current, metric_params_current).fetchall() if row[0]}
            else:
                tag_sql = """
                    SELECT DISTINCT tag
                    FROM measurements_active
                    WHERE row_kind=?
                      AND COALESCE(is_official,1)=1
                """
                tag_params = [regular_kind]
                if bank:
                    tag_sql += " AND bank=?"
                    tag_params.append(bank)
                tag_sql += " ORDER BY tag"
                tags = [
                    {"value": row[0], "label": _chart_format_tag_label(row[0])}
                    for row in cur.execute(tag_sql, tag_params).fetchall()
                    if row[0]
                ]

                metric_sql = """
                    SELECT DISTINCT metric_name
                    FROM measurements_active
                    WHERE row_kind=?
                      AND COALESCE(is_official,1)=1
                """
                metric_params_all = [regular_kind]
                metric_sql_current = metric_sql + " AND day_ref BETWEEN ? AND ?"
                metric_params_current = [regular_kind, date_from, date_to]
                if bank:
                    metric_sql += " AND bank=?"
                    metric_sql_current += " AND bank=?"
                    metric_params_all.append(bank)
                    metric_params_current.append(bank)
                if tag:
                    tag_clause = """
                        AND (
                            tag=?
                            OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                        )
                    """
                    metric_sql += tag_clause
                    metric_sql_current += tag_clause
                    metric_params_all.extend([tag, normalized_tag])
                    metric_params_current.extend([tag, normalized_tag])
                metric_sql += " ORDER BY metric_name"
                metric_sql_current += " ORDER BY metric_name"
                all_metrics = {row[0] for row in cur.execute(metric_sql, metric_params_all).fetchall() if row[0]}
                current_metrics = {row[0] for row in cur.execute(metric_sql_current, metric_params_current).fetchall() if row[0]}

        metrics = sorted(
            [
                {
                    "value": metric_name,
                    "label": metric_name,
                    "group": _chart_metric_group(metric_name, bank=bank, row_kind=effective_kind if bank != "SEP" else f"sep_{effective_kind}"),
                    "has_data": metric_name in current_metrics,
                }
                for metric_name in all_metrics
            ],
            key=lambda item: (
                CHART_ADVANCED_GROUP_ORDER.get(item["group"], 99),
                0 if item["has_data"] else 1,
                item["label"],
            ),
        )
        conn.close()
        return {
            "date_from": date_from,
            "date_to": date_to,
            "row_kind": effective_kind,
            "banks": banks,
            "tags": tags,
            "metrics": metrics,
        }

    @app.get("/api/ops/chart-presets-meta")
    def api_ops_chart_presets_meta(date_from: str = "", date_to: str = ""):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to

        daily_points = {
            (row["bank"], row["tag"])
            for row in cur.execute(
                """
                SELECT DISTINCT bank, tag
                FROM measurements_active
                WHERE day_ref BETWEEN ? AND ?
                  AND row_kind='daily'
                  AND COALESCE(is_official,1)=1
                """,
                (date_from, date_to),
            ).fetchall()
        }
        recon_targets = [
            {
                "value": f"{row['bank']}|{row['tag']}",
                "label": f"{row['bank']} · {_chart_format_tag_label(row['tag'])}",
            }
            for row in cur.execute(
                """
                SELECT DISTINCT bank, tag
                FROM measurements_active
                WHERE day_ref BETWEEN ? AND ?
                  AND row_kind='recon'
                  AND COALESCE(is_official,1)=1
                  AND bank<>''
                  AND tag<>''
                ORDER BY bank, tag
                """,
                (date_from, date_to),
            ).fetchall()
        ]
        aligned_targets = [
            {
                "value": f"{row['bank']}|{row['mpfm_tag']}",
                "label": f"{row['bank']} · {_chart_format_tag_label(row['mpfm_tag'])}",
            }
            for row in cur.execute(
                """
                SELECT DISTINCT bank, mpfm_tag
                FROM sep_alignments
                WHERE production_date BETWEEN ? AND ?
                  AND is_active=1
                  AND COALESCE(is_official,1)=1
                  AND bank<>''
                  AND COALESCE(mpfm_tag,'')<>''
                ORDER BY bank, mpfm_tag
                """,
                (date_from, date_to),
            ).fetchall()
        ]
        separator_available = bool(
            cur.execute(
                """
                SELECT 1
                FROM measurements_active
                WHERE day_ref BETWEEN ? AND ?
                  AND row_kind='sep'
                  AND bank='SEP'
                  AND COALESCE(is_official,1)=1
                LIMIT 1
                """,
                (date_from, date_to),
            ).fetchone()
        )
        focus_pairs = []
        for pair in CHART_FOCUS_PAIRS:
            available = (
                (pair["subsea_bank"], pair["subsea_tag"]) in daily_points and
                (pair["topside_bank"], pair["topside_tag"]) in daily_points
            )
            focus_pairs.append({"value": pair["key"], "label": pair["title"], "available": available})
        conn.close()
        return {
            "date_from": date_from,
            "date_to": date_to,
            "focus_pairs": focus_pairs,
            "recon_targets": recon_targets,
            "aligned_targets": aligned_targets,
            "separator_available": separator_available,
            "separator_metrics_daily": _chart_metric_choices(["oil", "gas", "water", "hc", "total", "pressure", "temperature"]),
            "separator_metrics_hourly": _chart_metric_choices(["oil", "gas", "water", "hc", "total"]),
            "compare_metrics": _chart_metric_choices(["oil", "gas", "water", "hc", "total"]),
        }

    @app.get("/api/ops/chart-preset-series")
    def api_ops_chart_preset_series(date_from: str = "", date_to: str = "", preset: str = "", target: str = "", metric_key: str = "hc", separator_mode: str = "daily"):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to
        metric_cfg = CHART_COMPARE_METRICS.get(metric_key) or CHART_COMPARE_METRICS["hc"]
        labels, datasets = [], []
        message = ""
        response_kind = "daily"

        if preset == "subsea_topside":
            pair = _chart_pair_map().get(target)
            if not pair:
                conn.close()
                raise HTTPException(400, "Par inválido para comparação Subsea × Topside")
            rows = [
                dict(r)
                for r in cur.execute(
                    """
                    SELECT day_ref, bank, tag, metric_value, source_file
                    FROM measurements_active
                    WHERE day_ref BETWEEN ? AND ?
                      AND row_kind='daily'
                      AND metric_name=?
                      AND COALESCE(is_official,1)=1
                      AND ((bank=? AND tag=?) OR (bank=? AND tag=?))
                    ORDER BY day_ref
                    """,
                    (
                        date_from,
                        date_to,
                        metric_cfg["mpfm_metric"],
                        pair["subsea_bank"],
                        pair["subsea_tag"],
                        pair["topside_bank"],
                        pair["topside_tag"],
                    ),
                ).fetchall()
            ]
            subsea_map = {row["day_ref"]: row["metric_value"] for row in rows if row["tag"] == pair["subsea_tag"]}
            topside_map = {row["day_ref"]: row["metric_value"] for row in rows if row["tag"] == pair["topside_tag"]}
            subsea_adj = {row["day_ref"]: _chart_adjustment_meta(row.get("source_file")) for row in rows if row["tag"] == pair["subsea_tag"]}
            topside_adj = {row["day_ref"]: _chart_adjustment_meta(row.get("source_file")) for row in rows if row["tag"] == pair["topside_tag"]}
            labels = sorted(set(subsea_map.keys()) | set(topside_map.keys()))
            datasets = [
                {
                    "label": pair["subsea_label"],
                    "values": [subsea_map.get(label) for label in labels],
                    "adjusted": [subsea_adj.get(label, (False, ""))[0] for label in labels],
                    "sources": [subsea_adj.get(label, (False, ""))[1] for label in labels],
                },
                {
                    "label": pair["topside_label"],
                    "values": [topside_map.get(label) for label in labels],
                    "adjusted": [topside_adj.get(label, (False, ""))[0] for label in labels],
                    "sources": [topside_adj.get(label, (False, ""))[1] for label in labels],
                },
            ]
            message = pair["title"]
        elif preset == "separator_test":
            sep_mode = "hourly" if separator_mode == "hourly" else "daily"
            if sep_mode == "hourly":
                labels, value_map = _chart_sep_hourly_series(cur, date_from, date_to, metric_key)
                response_kind = "hourly"
                if not labels and metric_key not in {"oil", "gas", "water", "hc", "total"}:
                    message = "Modo horário do Separador disponível para óleo, gás, água, HC e total."
                else:
                    message = "Separador de Teste · Horário"
            else:
                rows = [
                    dict(r)
                    for r in cur.execute(
                        """
                        SELECT day_ref, metric_value, source_file
                        FROM measurements_active
                        WHERE day_ref BETWEEN ? AND ?
                          AND row_kind='sep'
                          AND bank='SEP'
                          AND metric_name=?
                          AND COALESCE(is_official,1)=1
                        ORDER BY day_ref
                        """,
                        (date_from, date_to, metric_cfg["sep_metric"]),
                    ).fetchall()
                ]
                value_map = {row["day_ref"]: row["metric_value"] for row in rows}
                labels = sorted(value_map.keys())
                message = "Separador de Teste · 24h consolidado"
            datasets = [{"label": f"Separador de Teste · {metric_cfg['label']}", "values": [value_map.get(label) for label in labels]}]
        elif preset == "mpfm_sep":
            if "|" not in target:
                conn.close()
                raise HTTPException(400, "Selecione um ponto alinhado ao separador")
            bank, mpfm_tag = target.split("|", 1)
            aligned_days = [
                row["production_date"]
                for row in cur.execute(
                    """
                    SELECT DISTINCT production_date
                    FROM sep_alignments
                    WHERE production_date BETWEEN ? AND ?
                      AND bank=?
                      AND mpfm_tag=?
                      AND is_active=1
                      AND COALESCE(is_official,1)=1
                    ORDER BY production_date
                    """,
                    (date_from, date_to, bank, mpfm_tag),
                ).fetchall()
            ]
            if aligned_days:
                q_marks = ",".join("?" * len(aligned_days))
                mpfm_rows = [
                    dict(r)
                    for r in cur.execute(
                        f"""
                        SELECT day_ref, metric_value
                        FROM measurements_active
                        WHERE row_kind='daily'
                          AND bank=?
                          AND tag=?
                          AND metric_name=?
                          AND COALESCE(is_official,1)=1
                          AND day_ref IN ({q_marks})
                        ORDER BY day_ref
                        """,
                        [bank, mpfm_tag, metric_cfg["mpfm_metric"], *aligned_days],
                    ).fetchall()
                ]
                sep_rows = [
                    dict(r)
                    for r in cur.execute(
                        f"""
                        SELECT day_ref, metric_value
                        FROM measurements_active
                        WHERE row_kind='sep'
                          AND bank='SEP'
                          AND metric_name=?
                          AND COALESCE(is_official,1)=1
                          AND day_ref IN ({q_marks})
                        ORDER BY day_ref
                        """,
                        [metric_cfg["sep_metric"], *aligned_days],
                    ).fetchall()
                ]
                mpfm_map = {row["day_ref"]: row["metric_value"] for row in mpfm_rows}
                mpfm_adj = {row["day_ref"]: _chart_adjustment_meta(row.get("source_file")) for row in mpfm_rows}
                sep_map = {row["day_ref"]: row["metric_value"] for row in sep_rows}
                labels = sorted(aligned_days)
                datasets = [
                    {
                        "label": f"{bank} · {_chart_format_tag_label(mpfm_tag)}",
                        "values": [mpfm_map.get(label) for label in labels],
                        "adjusted": [mpfm_adj.get(label, (False, ""))[0] for label in labels],
                        "sources": [mpfm_adj.get(label, (False, ""))[1] for label in labels],
                    },
                    {"label": f"Separador de Teste · {metric_cfg['label']}", "values": [sep_map.get(label) for label in labels]},
                ]
                message = f"{bank} · {_chart_format_tag_label(mpfm_tag)} × Separador"
            else:
                message = "Sem alinhamentos ativos do Separador no período selecionado."
        elif preset == "recon":
            if "|" not in target:
                conn.close()
                raise HTTPException(400, "Selecione um ponto para o preset de reconciliação")
            bank, mpfm_tag = target.split("|", 1)
            if not metric_cfg.get("recon_daily"):
                conn.close()
                raise HTTPException(400, "Métrica inválida para reconciliação")
            rows = [
                dict(r)
                for r in cur.execute(
                    """
                    SELECT day_ref, metric_name, metric_value
                    FROM measurements_active
                    WHERE day_ref BETWEEN ? AND ?
                      AND row_kind='recon'
                      AND bank=?
                      AND tag=?
                      AND metric_name IN (?, ?, ?)
                      AND COALESCE(is_official,1)=1
                    ORDER BY day_ref, metric_name
                    """,
                    (
                        date_from,
                        date_to,
                        bank,
                        mpfm_tag,
                        metric_cfg["recon_daily"],
                        metric_cfg["recon_sum"],
                        metric_cfg["recon_delta"],
                    ),
                ).fetchall()
            ]
            by_name = {
                metric_cfg["recon_daily"]: {},
                metric_cfg["recon_sum"]: {},
                metric_cfg["recon_delta"]: {},
            }
            for row in rows:
                by_name[row["metric_name"]][row["day_ref"]] = row["metric_value"]
            labels = sorted({row["day_ref"] for row in rows})
            datasets = [
                {"label": "Daily", "values": [by_name[metric_cfg["recon_daily"]].get(label) for label in labels]},
                {"label": "Soma hourly", "values": [by_name[metric_cfg["recon_sum"]].get(label) for label in labels]},
                {"label": "Δ", "values": [by_name[metric_cfg["recon_delta"]].get(label) for label in labels]},
            ]
            message = f"{bank} · {_chart_format_tag_label(mpfm_tag)} · Reconciliação"
        else:
            conn.close()
            raise HTTPException(400, "Preset de gráfico inválido")

        conn.close()
        return {
            "date_from": date_from,
            "date_to": date_to,
            "preset": preset,
            "metric_key": metric_key,
            "metric_label": metric_cfg["label"],
            "labels": labels,
            "datasets": datasets,
            "message": message,
            "deviation_key": metric_cfg.get("deviation_key", ""),
            "kind": response_kind,
        }

    @app.get("/api/ops/chart-series")
    def api_ops_chart_series(date_from: str = "", date_to: str = "", row_kind: str = "daily", bank: str = "", metric: str = "", tag: str = ""):
        date_from, date_to = normalize_date_range(date_from, date_to)
        conn = db_conn()
        cur = conn.cursor()
        if not date_to:
            date_to = cur.execute("SELECT MAX(day_ref) FROM measurements_active").fetchone()[0] or ""
        if not date_from:
            date_from = date_to
        labels = []
        values = []
        adjusted = []
        sources = []
        if bank == "SEP" and row_kind in ("daily", "hourly"):
            if row_kind == "daily":
                sql = """
                    SELECT day_ref, hour_ref, metric_value
                    FROM measurements_active
                    WHERE day_ref BETWEEN ? AND ?
                      AND row_kind='sep'
                      AND bank='SEP'
                      AND COALESCE(is_official,1)=1
                """
                params = [date_from, date_to]
                if metric:
                    sql += " AND metric_name=?"
                    params.append(metric)
                if tag:
                    sql += """
                        AND (
                            tag=?
                            OR REPLACE(REPLACE(REPLACE(UPPER(COALESCE(tag,'')), ' ', ''), '-', ''), '_', '')=?
                        )
                    """
                    params.extend([tag, normalize_tag_name(tag)])
                sql += " ORDER BY day_ref, hour_ref"
                data = cur.execute(sql, params).fetchall()
                for d, h, v in data:
                    labels.append(d)
                    values.append(v)
                    adjusted.append(False)
                    sources.append("")
            else:
                rows = _chart_sep_metric_query(cur, metric, date_from, date_to, tag=tag)
                value_map = _chart_sep_metric_map("hourly", rows)
                labels = sorted(value_map.keys())
                values = [value_map.get(label) for label in labels]
                adjusted = [False for _ in labels]
                sources = ["" for _ in labels]
        else:
            sql = "SELECT day_ref, hour_ref, metric_value, source_file FROM measurements_active WHERE day_ref BETWEEN ? AND ?"
            params = [date_from, date_to]
            if row_kind in ("hourly", "daily", "recon"):
                sql += " AND row_kind=?"
                params.append(row_kind)
            if bank:
                sql += " AND bank=?"
                params.append(bank)
            if metric:
                sql += " AND metric_name=?"
                params.append(metric)
            if tag:
                sql += " AND tag=?"
                params.append(tag)
            sql += " ORDER BY day_ref, hour_ref"
            data = cur.execute(sql, params).fetchall()
            for d, h, v, source_file in data:
                source_name = str(source_file or "")
                is_adjusted = source_name.lower().startswith("manual_adjustment:")
                labels.append(f"{d} {int(h):02d}:00" if row_kind == "hourly" and h is not None else d)
                values.append(v)
                adjusted.append(bool(is_adjusted))
                sources.append(source_name.split(":", 1)[1] if is_adjusted and ":" in source_name else source_name)
        conn.close()
        return {"labels": labels, "values": values, "adjusted": adjusted, "sources": sources, "date_from": date_from, "date_to": date_to}

    @app.get("/api/monitoring-summary")
    @cached(ttl=90, key_prefix="monitoring_summary")
    def api_monitoring_summary(
        limit: int = 50,
        offset: int = 0,
        runs_limit: int = 20,
        runs_offset: int = 0,
    ):
        conn = db_conn()
        cur = conn.cursor()
        total_issues = cur.execute("SELECT COUNT(*) FROM validation_issues").fetchone()[0]
        issues = []
        for row in cur.execute(
            "SELECT created_at, excel_file, issue_type, severity, ref_key, day_ref, details FROM validation_issues ORDER BY id DESC LIMIT ? OFFSET ?",
            (limit, offset),
        ).fetchall():
            item = dict(row)
            item["day_ref"] = normalize_validation_issue_day_ref(item.get("day_ref", ""), item.get("created_at", ""))
            issues.append(item)
        total_runs = cur.execute("SELECT COUNT(*) FROM processing_runs").fetchone()[0]
        runs = [dict(r) for r in cur.execute(
            "SELECT id, started_at, finished_at, source_type, source_ref, files_count, status FROM processing_runs ORDER BY id DESC LIMIT ? OFFSET ?",
            (runs_limit, runs_offset),
        ).fetchall()]
        count = cur.execute("SELECT COUNT(*) FROM measurements_active").fetchone()[0]
        conn.close()
        return {
            "runs": runs,
            "issues": issues,
            "measurements_count": count,
            "pagination": {
                "issues": {"limit": limit, "offset": offset, "total": total_issues},
                "runs": {"limit": runs_limit, "offset": runs_offset, "total": total_runs},
            },
        }

    @app.post("/api/admin/clear-data")
    def api_admin_clear_data(body: dict = None):
        keep_backup_zip = True
        if body and isinstance(body, dict):
            keep_backup_zip = bool(body.get("keep_backup_zip", True))
        result = clear_local_data(keep_backup_zip=keep_backup_zip)
        _invalidate_cache()
        return result

    @app.post("/api/admin/restart-db")
    def api_admin_restart_db(body: dict = None):
        keep_backup_zip = True
        if body and isinstance(body, dict):
            keep_backup_zip = bool(body.get("keep_backup_zip", True))
        result = restart_local_data(keep_backup_zip=keep_backup_zip)
        _invalidate_cache()
        return result

    @app.post("/api/admin/backup-and-clear")
    def api_admin_backup_and_clear():
        zpath = build_backup_zip()
        result = clear_local_data(keep_backup_zip=True)
        _invalidate_cache()
        return {
            **result,
            "backup_file": zpath.name,
        }

    @app.post("/api/admin/backup-only")
    def api_admin_backup_only():
        zpath = build_backup_zip()
        return {
            "ok": True,
            "backup_file": zpath.name,
            "message": "Backup gerado sem alterar a base local.",
            "db_path": str(db_path),
        }

    @app.get("/api/admin/cache/stats")
    def api_admin_cache_stats():
        return _cache.get_stats()

    @app.post("/api/admin/cache/invalidate")
    def api_admin_cache_invalidate(body: dict | None = None):
        pattern = (body or {}).get("pattern")
        return _invalidate_cache(pattern)

    @app.get("/api/admin/recovery/diagnostics")
    def api_admin_recovery_diagnostics(month: str = ""):
        target_month = month or datetime.now().strftime("%Y-%m")
        yr, mo, date_from, date_to = _parse_target_month(target_month)
        workbook_path = output_dir / excel_name(yr, mo)
        state_path = work_dir / f"state_{yr}_{mo}.json"
        diagnostics = {
            "month": target_month,
            "db_path": str(db_path),
            "work_dir": str(work_dir),
            "output_dir": str(output_dir),
            "workbook_exists": workbook_path.exists(),
            "workbook_name": workbook_path.name,
            "workbook_updated_at": datetime.fromtimestamp(workbook_path.stat().st_mtime).isoformat(timespec="seconds") if workbook_path.exists() else "",
            "state_exists": state_path.exists(),
            "state_file": state_path.name,
        }
        conn = None
        try:
            conn = db_conn()
            cur = conn.cursor()
            diagnostics["db_ok"] = True
            diagnostics["database"] = {
                "journal_mode": _safe_scalar(cur, "PRAGMA journal_mode", default=""),
                "busy_timeout_ms": _safe_scalar(cur, "PRAGMA busy_timeout", default=0),
                "measurements_total": _safe_scalar(cur, "SELECT COUNT(*) FROM measurements_active"),
                "measurements_month": _safe_scalar(cur, "SELECT COUNT(*) FROM measurements_active WHERE day_ref BETWEEN ? AND ?", (date_from, date_to)),
                "issues_month": _safe_scalar(cur, "SELECT COUNT(*) FROM validation_issues WHERE day_ref BETWEEN ? AND ?", (date_from, date_to)),
                "runs_total": _safe_scalar(cur, "SELECT COUNT(*) FROM processing_runs"),
                "runs_month": _safe_scalar(cur, "SELECT COUNT(*) FROM processing_runs WHERE source_ref LIKE ?", (f"{yr}-{mo}%",)),
                "sep_alignments_month": _safe_scalar(cur, "SELECT COUNT(*) FROM sep_alignments WHERE is_active=1 AND production_date BETWEEN ? AND ?", (date_from, date_to)),
                "sep_sources_month": _safe_scalar(cur, "SELECT COUNT(*) FROM sep_source_files WHERE is_active=1 AND production_date BETWEEN ? AND ?", (date_from, date_to)),
                "sep_official_days_month": _safe_scalar(
                    cur,
                    """
                    SELECT COUNT(*)
                    FROM (
                        SELECT production_date
                        FROM sep_source_files
                        WHERE is_active=1
                          AND is_official=1
                          AND production_date BETWEEN ? AND ?
                          AND fluid_kind IN ('sep_oleo','sep_gas','sep_agua')
                        GROUP BY production_date
                        HAVING COUNT(DISTINCT fluid_kind)=3
                    )
                    """,
                    (date_from, date_to),
                ),
                "sep_summary_days_month": _safe_scalar(cur, "SELECT COUNT(DISTINCT day_ref) FROM measurements_active WHERE row_kind='sep' AND day_ref BETWEEN ? AND ?", (date_from, date_to)),
                "sep_detail_days_month": _safe_scalar(cur, "SELECT COUNT(DISTINCT day_ref) FROM measurements_active WHERE row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail') AND day_ref BETWEEN ? AND ?", (date_from, date_to)),
                "daily_cards_month": _safe_scalar(cur, "SELECT COUNT(*) FROM daily_cards WHERE is_active=1 AND production_date BETWEEN ? AND ?", (date_from, date_to)),
                "latest_day_ref": _safe_scalar(cur, "SELECT MAX(day_ref) FROM measurements_active", default=""),
            }
            diagnostics["hourly_coverage"] = [
                {
                    "bank": row["bank"],
                    "daily_days": row["daily_days"],
                    "full_days": row["full_days"],
                    "partial_days": row["partial_days"],
                    "missing_days": row["missing_days"],
                }
                for row in cur.execute(
                    """
                    WITH daily_days AS (
                        SELECT bank, day_ref
                        FROM measurements_active
                        WHERE row_kind='daily' AND bank<>'' AND bank<>'SEP' AND day_ref BETWEEN ? AND ?
                        GROUP BY bank, day_ref
                    ),
                    hourly_days AS (
                        SELECT bank, day_ref, COUNT(DISTINCT hour_ref) AS n_hours
                        FROM measurements_active
                        WHERE row_kind='hourly' AND bank<>'' AND bank<>'SEP' AND day_ref BETWEEN ? AND ?
                        GROUP BY bank, day_ref
                    )
                    SELECT
                        d.bank,
                        COUNT(*) AS daily_days,
                        SUM(CASE WHEN COALESCE(h.n_hours, 0) = 24 THEN 1 ELSE 0 END) AS full_days,
                        SUM(CASE WHEN COALESCE(h.n_hours, 0) BETWEEN 1 AND 23 THEN 1 ELSE 0 END) AS partial_days,
                        SUM(CASE WHEN COALESCE(h.n_hours, 0) = 0 THEN 1 ELSE 0 END) AS missing_days
                    FROM daily_days d
                    LEFT JOIN hourly_days h ON h.bank = d.bank AND h.day_ref = d.day_ref
                    GROUP BY d.bank
                    ORDER BY d.bank
                    """,
                    (date_from, date_to, date_from, date_to),
                ).fetchall()
            ]
        except Exception as exc:
            diagnostics["db_ok"] = False
            diagnostics["database_error"] = str(exc)
        finally:
            if conn:
                conn.close()
        diagnostics["summary"] = _recovery_result_lines(diagnostics.get("database", {})) if diagnostics.get("db_ok") else diagnostics.get("database_error", "")
        return diagnostics

    @app.post("/api/admin/recovery/test-db")
    def api_admin_recovery_test_db():
        conn = None
        started = datetime.now()
        try:
            conn = db_conn()
            cur = conn.cursor()
            cur.execute("SELECT 1").fetchone()
            result = {
                "ok": True,
                "journal_mode": _safe_scalar(cur, "PRAGMA journal_mode", default=""),
                "busy_timeout_ms": _safe_scalar(cur, "PRAGMA busy_timeout", default=0),
                "db_path": str(db_path),
                "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
            }
        except Exception as exc:
            result = {
                "ok": False,
                "db_path": str(db_path),
                "elapsed_ms": int((datetime.now() - started).total_seconds() * 1000),
                "error": str(exc),
            }
        finally:
            if conn:
                conn.close()
        return result

    @app.post("/api/admin/recovery/rebuild-month")
    def api_admin_recovery_rebuild_month(body: dict | None = None):
        target_month = str((body or {}).get("month") or datetime.now().strftime("%Y-%m"))
        yr, mo, _, _ = _parse_target_month(target_month)
        workbook_path = output_dir / excel_name(yr, mo)
        queued = schedule_monthly_base_unica(workbook_path, yr, mo)
        _invalidate_cache()
        return {
            "ok": True,
            "month": target_month,
            "queued": queued,
            "workbook_name": workbook_path.name,
            "workbook_exists": workbook_path.exists(),
            "workbook_path": str(workbook_path),
        }

    @app.post("/api/admin/recovery/recompute-month")
    def api_admin_recovery_recompute_month(body: dict | None = None):
        target_month = str((body or {}).get("month") or datetime.now().strftime("%Y-%m"))
        yr, mo, date_from, date_to = _parse_target_month(target_month)
        conn = db_conn()
        conn.row_factory = lambda cursor, row: {
            cursor.description[idx][0]: row[idx] for idx in range(len(cursor.description))
        }
        cur = conn.cursor()
        alignment_rows = cur.execute(
            """
            SELECT DISTINCT production_date, bank
            FROM sep_alignments
            WHERE is_active=1 AND production_date BETWEEN ? AND ?
            ORDER BY production_date, bank
            """,
            (date_from, date_to),
        ).fetchall()
        source_rows = cur.execute(
            """
            SELECT DISTINCT production_date, fluid_kind, meter_id
            FROM sep_source_files
            WHERE is_active=1 AND production_date BETWEEN ? AND ?
            ORDER BY production_date, fluid_kind, meter_id
            """,
            (date_from, date_to),
        ).fetchall()
        card_rows = cur.execute(
            """
            SELECT DISTINCT production_date, bank, card_type, COALESCE(tag,'') AS tag, COALESCE(instrument,'') AS instrument
            FROM daily_cards
            WHERE is_active=1 AND production_date BETWEEN ? AND ?
            ORDER BY production_date, bank, card_type, tag, instrument
            """,
            (date_from, date_to),
        ).fetchall()
        sep_days = [
            row["production_date"]
            for row in cur.execute(
                """
                SELECT DISTINCT production_date
                FROM sep_source_files
                WHERE is_active=1 AND production_date BETWEEN ? AND ?
                ORDER BY production_date
                """,
                (date_from, date_to),
            ).fetchall()
        ]
        conn.close()

        for row in alignment_rows:
            recompute_alignment_resolution(row["production_date"], row["bank"])
        for row in source_rows:
            recompute_sep_source_resolution(row["production_date"], row["fluid_kind"], row["meter_id"])
        sep_sync = _sync_sep_month_state(target_month, rebuild_summary=True)
        for row in card_rows:
            recompute_card_resolution(
                row["production_date"],
                row["bank"],
                row["card_type"],
                row["tag"],
                row["instrument"],
            )
        validation_snapshot = rebuild_validation_snapshot_for_month(target_month)

        workbook_path = output_dir / excel_name(yr, mo)
        queued = schedule_monthly_base_unica(workbook_path, yr, mo)
        _invalidate_cache()
        return {
            "ok": True,
            "month": target_month,
            "recomputed": {
                "sep_alignments": len(alignment_rows),
                "sep_source_groups": len(source_rows),
                "sep_summary_days": sep_sync.get("rebuilt_days", 0),
                "daily_card_groups": len(card_rows),
            },
            "sep_state_sync": sep_sync,
            "validation_snapshot": validation_snapshot.get("recomputed", {}),
            "queued": queued,
            "workbook_name": workbook_path.name,
            "workbook_exists": workbook_path.exists(),
        }

    @app.post("/api/admin/recovery/sync-sep-month")
    def api_admin_recovery_sync_sep_month(body: dict | None = None):
        target_month = str((body or {}).get("month") or datetime.now().strftime("%Y-%m"))
        _parse_target_month(target_month)
        result = _sync_sep_month_state(target_month, rebuild_summary=True)
        _invalidate_cache()
        return {
            "ok": True,
            "month": target_month,
            "sep_state_sync": result,
        }

    @app.post("/api/admin/recovery/sanitize-import-history")
    def api_admin_recovery_sanitize_import_history(body: dict | None = None):
        target_month = str((body or {}).get("month") or datetime.now().strftime("%Y-%m"))
        _, _, _, _ = _parse_target_month(target_month)
        result = sanitize_files_imported_history(target_month)
        _invalidate_cache()
        return result

    @app.post("/api/admin/recovery/delete-day")
    def api_admin_recovery_delete_day(body: dict | None = None):
        target_date = normalize_date_input(str((body or {}).get("date") or "").strip())
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", target_date):
            raise HTTPException(status_code=400, detail="Dia inválido. Use o formato YYYY-MM-DD.")
        result = delete_all_data_for_day(target_date)
        if not result.get("ok"):
            raise HTTPException(status_code=400, detail=result.get("error") or "Falha ao apagar o dia selecionado.")
        _invalidate_cache()
        return result

    @app.post("/api/admin/recovery/base-unica-import/preview")
    async def api_admin_recovery_base_unica_preview(request: Request, file: UploadFile = File(...)):
        if not str(file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx válido.")
        temp_path = _persist_uploaded_workbook(file)
        try:
            result = preview_base_unica_import(
                db_conn,
                Path(temp_path),
                str(request.query_params.get("month") or "").strip(),
            )
            return result
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

    @app.post("/api/admin/recovery/base-unica-import/apply")
    async def api_admin_recovery_base_unica_apply(request: Request, file: UploadFile = File(...)):
        if not str(file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Envie um arquivo .xlsx válido.")
        temp_path = _persist_uploaded_workbook(file)
        try:
            result = apply_base_unica_import(
                db_conn_fn=db_conn,
                workbook_path=Path(temp_path),
                build_backup_zip_fn=build_backup_zip,
                load_state_fn=load_state,
                save_state_fn=save_state,
                rebuild_validation_snapshot_for_month_fn=rebuild_validation_snapshot_for_month,
                schedule_monthly_base_unica_fn=schedule_monthly_base_unica,
                output_dir=output_dir,
                excel_name_fn=excel_name,
                serialize_sep_row_fn=serialize_sep_row,
                target_month=str(request.query_params.get("month") or "").strip(),
            )
            _invalidate_cache()
            return result
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc))
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

    @app.post("/api/admin/recovery/backup-zip/restore")
    async def api_admin_recovery_restore_backup_zip(file: UploadFile = File(...)):
        if not str(file.filename or "").lower().endswith(".zip"):
            raise HTTPException(status_code=400, detail="Envie um arquivo .zip válido.")
        temp_path = _persist_uploaded_file(file, "mpfm_backup.zip")
        try:
            result = _restore_backup_zip(Path(temp_path))
            _invalidate_cache()
            return result
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass

    @app.post("/api/admin/pi-vision/import")
    async def api_admin_pi_vision_import(request: Request):
        """
        Importa leituras PI Vision do Excel gerado pelo coletor para pi_vision_readings.
        Body JSON opcional: { "excel_path": "...", "only_authorized_variables": true }
        """
        from services.importing.pi_vision_import_service import import_pi_excel
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        excel_path = body.get("excel_path") or None
        only_auth = bool(body.get("only_authorized_variables", True))
        conn = db_conn()
        try:
            result = import_pi_excel(conn, excel_path=excel_path, only_authorized_variables=only_auth)
        finally:
            conn.close()
        return result

    @app.get("/api/admin/pi-vision/summary")
    def api_admin_pi_vision_summary():
        """Resumo das leituras PI armazenadas em pi_vision_readings."""
        from services.importing.pi_vision_import_service import pi_readings_summary
        conn = db_conn()
        try:
            return pi_readings_summary(conn)
        finally:
            conn.close()
