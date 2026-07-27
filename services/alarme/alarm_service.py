from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from pypdf import PdfReader

_TIMESTAMP_RE = re.compile(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$")
_BANK_RE = re.compile(r"Bank(\d{2})", re.IGNORECASE)
_STREAM_RE = re.compile(r"Stream(\d{2})", re.IGNORECASE)
_METER_RE = re.compile(r"Meter([A-Z])", re.IGNORECASE)


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _clean_lines(text: str) -> list[str]:
    return [line.strip() for line in text.splitlines() if line and line.strip()]


def _is_timestamp(value: str) -> bool:
    return bool(_TIMESTAMP_RE.match(value.strip()))


def _severity(priority: str, description: str = "", detailed_state: str = "") -> str:
    value = f"{priority} {description} {detailed_state}".lower()
    if "critical" in value or "communication" in value or "failed" in value:
        return "critical"
    if "general" in value or "warning" in value or "deviation" in value or "alarm" in value:
        return "warning"
    return "info"


def _priority(severity: str, description: str = "") -> str:
    value = description.lower()
    if severity == "critical":
        return "urgent" if "communication" in value else "high"
    if severity == "warning":
        return "medium"
    return "low"


def _category(description: str, tag: str) -> str:
    value = f"{description} {tag}".lower()
    if "communication" in value or "not connected" in value:
        return "communication"
    if "deviation" in value or "dev" in value:
        return "deviation"
    if "voltage" in value or "electr" in value:
        return "critical_variable"
    if "flow" in value or "alarm" in value or "warning" in value:
        return "warning"
    return "other"


def _family(tag: str) -> str:
    value = tag.lower()
    if "communication" in value or "modbus" in value:
        return "Comunicacao"
    if "dp" in value:
        return "DP"
    if "trdev" in value or "transmitter" in value:
        return "Transmissor"
    if "flow" in value:
        return "Vazao"
    if "volt" in value:
        return "Eletronica"
    return "FCS320"


def _point(tag: str) -> tuple[str, str, str, str, str]:
    bank_match = _BANK_RE.search(tag)
    stream_match = _STREAM_RE.search(tag)
    meter_match = _METER_RE.search(tag)
    bank = f"B{bank_match.group(1)}" if bank_match else ""
    stream = f"S{stream_match.group(1)}" if stream_match else ""
    meter = f"Meter{meter_match.group(1).upper()}" if meter_match else ""
    measurement_point = " ".join(part for part in (bank, stream) if part)
    instrument = " ".join(part for part in (measurement_point, meter) if part)
    meter_type = "subsea" if bank else "system"
    return bank, measurement_point, tag, instrument, meter_type


def _record_from_parts(path: Path, parts: dict[str, str], index: int) -> dict:
    tag = parts.get("tag") or parts.get("object_id") or ""
    description = parts.get("description") or tag or "Alarme FCS320"
    event_at = parts.get("timestamp", "")
    severity = _severity(parts.get("priority", ""), description, parts.get("detailed_state", ""))
    bank, measurement_point, tag_value, instrument, meter_type = _point(tag)
    production_date = event_at[:10] if event_at else ""
    state = parts.get("state", "")
    detailed = parts.get("detailed_state", "")
    title = f"{description} - {tag_value}" if tag_value else description
    external = f"{path.stem}:{index}:{event_at}:{tag_value}:{parts.get('signal_number', '')}:{state}"
    payload = {key: value for key, value in parts.items() if value}
    return {
        "source_kind": "pdf",
        "source_ref": str(path),
        "record_type": "event",
        "source_sheet": "FCS320 PDF",
        "external_code": external,
        "title": title,
        "message": description,
        "category_code": _category(description, tag_value),
        "family_code": _family(tag_value),
        "severity_code": severity,
        "priority_code": _priority(severity, description),
        "status_code": "open" if state.lower() == "on" or severity == "critical" else "monitoring",
        "bank": bank,
        "measurement_point": measurement_point,
        "tag": tag_value,
        "instrument": instrument,
        "meter_type": meter_type,
        "measurement_state": " ".join(part for part in (state, detailed) if part),
        "event_at": event_at,
        "detected_at": event_at,
        "production_date": production_date,
        "occurrence_count": 1,
        "distinct_alarm_count": 1,
        "impact": detailed,
        "immediate_action": "Verificar comunicacao/status do medidor." if severity == "critical" else "Acompanhar recorrencia do alarme.",
        "reference": path.name,
        "payload_json": json.dumps(payload, ensure_ascii=False),
        "active": 1,
    }


def _parse_fcs320_pdf(path: str | Path) -> list[dict]:
    path = Path(path)
    lines = _clean_lines(_text(path))
    rows: list[dict] = []
    index = 0
    i = 0
    while i < len(lines):
        line = lines[i]
        priority_before = ""
        if line in {"Critical", "General", "Warning", "Info"} and i + 1 < len(lines) and _is_timestamp(lines[i + 1]):
            priority_before = line
            i += 1
            line = lines[i]
        if not _is_timestamp(line):
            i += 1
            continue
        timestamp = line
        remaining = len(lines) - i
        if remaining < 7:
            break
        object_id = lines[i + 1]
        tag = lines[i + 2]
        description = lines[i + 3]
        state = lines[i + 4]
        cursor = i + 5
        acknowledge = ""
        detailed = ""
        priority = priority_before
        if priority_before:
            detailed = lines[cursor] if cursor < len(lines) else ""
            cursor += 1
        else:
            acknowledge = lines[cursor] if cursor < len(lines) else ""
            cursor += 1
            detailed = lines[cursor] if cursor < len(lines) else ""
            cursor += 1
            priority = lines[cursor] if cursor < len(lines) else ""
            cursor += 1
        source_id = lines[cursor] if cursor < len(lines) else ""
        signal_number = lines[cursor + 1] if cursor + 1 < len(lines) else ""
        index += 1
        rows.append(_record_from_parts(path, {
            "timestamp": timestamp,
            "object_id": object_id,
            "tag": tag,
            "description": description,
            "state": state,
            "acknowledge_status": acknowledge,
            "detailed_state": detailed,
            "priority": priority,
            "source_id": source_id,
            "signal_number": signal_number,
        }, index))
        i = cursor + 2
    return rows


def _summary(rows: list[dict]) -> dict:
    meters = sorted({row.get("measurement_point") or row.get("tag") for row in rows if row.get("measurement_point") or row.get("tag")})
    return {"rows": len(rows), "meters": meters[:25]}


def _repo_summary(rows: list[dict]) -> dict:
    return {
        "total_active": len([row for row in rows if int(row.get("active") or 0) == 1]),
        "events": sum(1 for row in rows if row.get("record_type") == "event"),
        "incidents": sum(1 for row in rows if row.get("record_type") == "incident"),
        "open": sum(1 for row in rows if row.get("status_code") == "open"),
        "critical": sum(1 for row in rows if row.get("severity_code") == "critical"),
        "in_progress": sum(1 for row in rows if row.get("status_code") == "in_progress"),
        "monitoring": sum(1 for row in rows if row.get("status_code") == "monitoring"),
        "actions_open": 0,
        "overdue": 0,
    }


def _monitoring(rows: list[dict]) -> dict:
    by_day: dict[str, list[dict]] = defaultdict(list)
    by_month: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        day = row.get("production_date") or ""
        if day:
            by_day[day].append(row)
            by_month[day[:7]].append(row)

    def build(groups: dict[str, list[dict]]) -> dict:
        latest_period = sorted(groups)[-1] if groups else ""
        latest_rows = []
        totals = Counter()
        peak = ""
        peak_count = 0
        for period, period_rows in sorted(groups.items()):
            point_counts = Counter(row.get("measurement_point") or row.get("tag") or "Sistema" for row in period_rows)
            point, point_total = point_counts.most_common(1)[0] if point_counts else ("", 0)
            if point_total > peak_count:
                peak, peak_count = point, point_total
            critical = sum(1 for row in period_rows if row.get("severity_code") == "critical")
            open_count = sum(1 for row in period_rows if row.get("status_code") == "open")
            item = {
                "period": period,
                "measurement_point": point,
                "instrument": "FCS320",
                "snapshots": len(period_rows),
                "events": len(period_rows),
                "incidents_started": len({(row.get("tag"), row.get("message")) for row in period_rows}),
                "critical": critical,
                "open_count": open_count,
                "reserved_bits": 0,
            }
            if period == latest_period:
                latest_rows.append(item)
            for key in ("snapshots", "events", "incidents_started", "critical", "open_count", "reserved_bits"):
                totals[key] += int(item.get(key) or 0)
        return {
            "row_count": len(groups),
            "latest_period": latest_period,
            "latest_rows": latest_rows,
            "peak_measurement_point": peak,
            "totals": dict(totals),
            "window": "PDF FCS320",
        }

    return {"daily": build(by_day), "monthly": build(by_month)}


def _incident_rows(events: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for row in events:
        grouped[(row.get("tag", ""), row.get("message", ""), row.get("production_date", ""))].append(row)
    incidents = []
    for index, ((tag, message, day), rows) in enumerate(grouped.items(), start=1):
        if len(rows) < 2 and not any(row.get("severity_code") == "critical" for row in rows):
            continue
        first = sorted(rows, key=lambda row: row.get("event_at") or "")[0]
        incident = dict(first)
        incident.update({
            "record_type": "incident",
            "external_code": f"incident:{day}:{tag}:{message}:{index}",
            "title": f"Incidente - {message} - {tag}" if tag else f"Incidente - {message}",
            "occurrence_count": len(rows),
            "distinct_alarm_count": len({row.get("external_code") for row in rows}),
            "status_code": "open" if any(row.get("status_code") == "open" for row in rows) else "monitoring",
        })
        incidents.append(incident)
    return incidents


def preview_alarm_pdf_import(paths: Iterable[str | Path]) -> dict:
    source_paths = [Path(path) for path in paths if Path(path).exists()]
    events: list[dict] = []
    source_files = []
    for path in source_paths:
        parsed = _parse_fcs320_pdf(path)
        events.extend(parsed)
        source_files.append({"name": path.name, "path": str(path), "rows": len(parsed)})
    incidents = _incident_rows(events)
    return {
        "path": ";".join(str(path) for path in source_paths),
        "model_type": "raw-derived",
        "sheets_found": {"FCS320 PDF": bool(source_paths)},
        "event_source_sheet": "FCS320 PDF",
        "raw_event_rows": len(events),
        "derived_event_rows": len(events),
        "incident_rows": len(incidents),
        "events": _summary(events),
        "incidents": _summary(incidents),
        "monitoring": _monitoring(events),
        "source_files": source_files,
        "references_detected": {
            "families": sorted({row.get("family_code") for row in events if row.get("family_code")}),
            "categories": sorted({row.get("category_code") for row in events if row.get("category_code")}),
            "measurement_states": sorted({row.get("measurement_state") for row in events if row.get("measurement_state")})[:20],
        },
    }


def import_alarm_pdfs(paths: Iterable[str | Path], repo) -> dict:
    source_paths = [Path(path) for path in paths if Path(path).exists()]
    events: list[dict] = []
    for path in source_paths:
        events.extend(_parse_fcs320_pdf(path))
    incidents = _incident_rows(events)
    rows = [*events, *incidents]
    imported = 0
    updated = 0
    for row in rows:
        alarm_id = repo.save_alarm(row)
        if alarm_id:
            imported += 1
    preview = preview_alarm_pdf_import(source_paths)
    return {
        "ok": True,
        "path": ";".join(str(path) for path in source_paths),
        "imported": imported,
        "updated": updated,
        "skipped": 0,
        "references_created": 0,
        "summary": _repo_summary(rows),
        "preview": preview,
    }


def normalize_alarm_payload(body: dict) -> dict:
    payload = dict(body or {})
    payload.setdefault("source_kind", "manual")
    payload.setdefault("record_type", "event")
    payload.setdefault("status_code", "open")
    payload.setdefault("severity_code", "warning")
    payload.setdefault("priority_code", "medium")
    payload.setdefault("active", 1)
    return payload


def normalize_alarm_action_payload(body: dict) -> dict:
    payload = dict(body or {})
    payload.setdefault("action_type", "corrective")
    payload.setdefault("status_code", "open")
    payload.setdefault("active", 1)
    return payload


def inspect_alarm_workbook(path: str | Path) -> dict:
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(max_row=4, values_only=True))
        sheets.append({
            "name": ws.title,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "header": list(rows[0]) if rows else [],
            "preview_rows": [list(row) for row in rows],
        })
    wb.close()
    return {"path": str(path), "sheets": sheets}


def preview_alarm_workbook_import(path: str | Path) -> dict:
    inspected = inspect_alarm_workbook(path)
    return {
        "path": str(path),
        "model_type": "workbook",
        "sheets_found": {sheet["name"]: True for sheet in inspected.get("sheets", [])},
        "event_source_sheet": inspected.get("sheets", [{}])[0].get("name", "") if inspected.get("sheets") else "",
        "raw_event_rows": 0,
        "derived_event_rows": 0,
        "incident_rows": 0,
        "events": {"rows": 0, "meters": []},
        "incidents": {"rows": 0, "meters": []},
        "monitoring": {"daily": {"row_count": 0, "totals": {}}, "monthly": {"row_count": 0, "totals": {}}},
        "references_detected": {"families": [], "categories": [], "measurement_states": []},
    }


def import_alarm_workbook(path: str | Path, repo) -> dict:
    preview = preview_alarm_workbook_import(path)
    return {"ok": True, "path": str(path), "imported": 0, "updated": 0, "skipped": 0, "references_created": 0, "preview": preview}
