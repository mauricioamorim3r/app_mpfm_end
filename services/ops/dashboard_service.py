from __future__ import annotations

import json
import math


def build_dashboard_months(work_dir, output_dir, month_pt: dict, excel_name) -> list[dict]:
    result = []
    for state_file in sorted(work_dir.glob("state_*.json")):
        try:
            state = json.loads(state_file.read_text("utf-8"))
            yr, mo = state.get("yr", ""), state.get("mo", "")
            if not yr or not mo:
                continue
            fname = excel_name(yr, mo)
            fpath = output_dir / fname
            if not fpath.exists():
                continue

            from openpyxl import load_workbook

            wb = load_workbook(fpath, read_only=True, data_only=True)
            total_oil = 0.0
            total_gas = 0.0
            daily_points_map = {}
            for sheet_name in wb.sheetnames:
                if not sheet_name.startswith("DAILY_"):
                    continue
                ws = wb[sheet_name]
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                oil_idx = headers.index("MPFM corr Óleo (t)") if "MPFM corr Óleo (t)" in headers else None
                gas_idx = headers.index("MPFM corr Gás (t)") if "MPFM corr Gás (t)" in headers else None
                day_idx = headers.index("Dia") if "Dia" in headers else None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    oil_value = row[oil_idx] if oil_idx is not None else None
                    gas_value = row[gas_idx] if gas_idx is not None else None
                    day_value = row[day_idx] if day_idx is not None else None
                    if isinstance(oil_value, (int, float)) and not (isinstance(oil_value, float) and math.isnan(oil_value)):
                        total_oil += oil_value
                    if isinstance(gas_value, (int, float)) and not (isinstance(gas_value, float) and math.isnan(gas_value)):
                        total_gas += gas_value
                    if day_value and isinstance(oil_value, (int, float)) and not (isinstance(oil_value, float) and math.isnan(oil_value)):
                        day_str = str(day_value)[:10]
                        existing = daily_points_map.setdefault(day_str, {"dia": day_str, "oil": 0.0, "gas": 0.0})
                        existing["oil"] = round(existing["oil"] + oil_value, 3)
                        existing["gas"] = round(existing["gas"] + (gas_value or 0), 3)
            wb.close()

            processed_hours = state.get("processed_hours", {})
            processed = state.get("processed", [])

            import calendar

            n_days = calendar.monthrange(int(yr), int(mo))[1]
            days_with_data = sum(1 for day in range(1, n_days + 1) if processed_hours.get(f"{day:02d}_{mo}"))
            total_hours = sum(len(value) for value in processed_hours.values())
            days_complete = sum(
                1
                for day in range(1, n_days + 1)
                if len(processed_hours.get(f"{day:02d}_{mo}", [])) == 24
                and any(f"_{day:02d}_{mo}" in item for item in processed)
            )

            result.append(
                {
                    "file": fname,
                    "yr": yr,
                    "mo": mo,
                    "label": month_pt.get(mo, mo) + "/" + yr,
                    "total_oil_t": round(total_oil, 2),
                    "total_gas_t": round(total_gas, 2),
                    "days_with_data": days_with_data,
                    "days_complete": days_complete,
                    "n_days": n_days,
                    "total_hours_received": total_hours,
                    "daily_points": sorted(daily_points_map.values(), key=lambda item: item["dia"]),
                    "day_status": {
                        f"{day:02d}": {
                            "hours": len(processed_hours.get(f"{day:02d}_{mo}", [])),
                            "has_daily": any(f"_{day:02d}_{mo}" in item for item in processed),
                        }
                        for day in range(1, n_days + 1)
                    },
                    "last_run": state.get("last_run", ""),
                }
            )
        except Exception:
            pass
    return result
