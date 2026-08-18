"""
Serviço de importação de leituras PI Vision → pi_vision_readings.

Lê o arquivo Excel gerado pelo PI Vision Collector (PI_EXTRACT_TOTAL) e
persiste cada linha em pi_vision_readings, permitindo que o relatório DB
exiba dados estruturados de PI.

Formato esperado do Excel (colunas presentes no PI_EXTRACT_TOTAL):
    DataHora     — timestamp da leitura (string ou datetime)
    Medidor      — identificador do medidor / tag
    Variavel     — nome da variável (WLR, Temperature, Pressure, etc.)
    Canal        — canal do medidor
    Grupo        — grupo (ex: "LogQualidade" = leitura de qualidade)
    Valor        — valor numérico da leitura

Colunas opcionais que podem vir pré-incluídas pelo coletor:
    PI Dia Coleta, PI Inicio, PI Final, PI Status Coleta, PI Arquivo Origem
"""
from __future__ import annotations

import os
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

# ── Constantes ────────────────────────────────────────────────────────────────
_REQUIRED_COLS = {"Medidor", "Variavel", "DataHora", "Valor"}
_PI_EXTRACT_SHEET = "PI_EXTRACT_TOTAL"
_QUALITY_GROUP = "LogQualidade"

# Colunas autorizadas para importação (variáveis de condição de contorno)
PI_CONTOUR_VARIABLES = frozenset([
    "WLR", "WVF", "GVF", "GOR", "Temperature", "Pressure",
    "dP Inlet", "dP Outlet", "Velocity", "Water Conductivity",
    "Meter Status 1", "Meter Status 2", "Flow Calculation Warn.",
    "Calculation Mode", "Continuous Phase", "Water Conductivity Input",
])

_DEFAULT_PI_EXCEL = Path(
    os.environ.get("BASE_UNICA_PI_OUTPUT", "")
    or r"C:\PI_Vision_Collector\saida_v4\Historico_V49_Geometrico.xlsx"
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_ts(val) -> str:
    if val is None or str(val).strip() in ("", "nan", "None"):
        return ""
    s = str(val).strip()
    # Tenta converter para ISO sem timezone
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except ValueError:
            pass
    return s[:19]  # trunca como fallback


def _day_ref(ts: str) -> str:
    return ts[:10] if len(ts) >= 10 else ""


def _float_or_none(val):
    try:
        v = float(val)
        return None if v != v else v  # NaN guard
    except (TypeError, ValueError):
        return None


def _iter_rows(excel_path: Path, sheet: str) -> Iterator[dict]:
    """Lê o Excel e itera as linhas como dicts. Requer openpyxl ou xlrd."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return
        headers = [str(c).strip() if c is not None else "" for c in rows[0]]
        for row in rows[1:]:
            yield dict(zip(headers, row))
    except ImportError:
        raise RuntimeError("openpyxl não instalado — execute: pip install openpyxl")


# ── Serviço principal ─────────────────────────────────────────────────────────

def import_pi_excel(
    db_conn: sqlite3.Connection,
    excel_path: str | Path | None = None,
    run_id: int | None = None,
    sheet: str = _PI_EXTRACT_SHEET,
    only_authorized_variables: bool = True,
    batch_size: int = 2000,
) -> dict:
    """
    Importa o Excel PI Vision para a tabela pi_vision_readings.

    Parâmetros
    ----------
    db_conn : sqlite3.Connection
        Conexão com o banco (mpfm_local.db).
    excel_path : path opcional
        Arquivo Excel de saída do coletor PI. Padrão: BASE_UNICA_PI_OUTPUT.
    run_id : int opcional
        ID do run de processamento para rastreabilidade.
    sheet : str
        Nome da aba no Excel. Padrão: 'PI_EXTRACT_TOTAL'.
    only_authorized_variables : bool
        Se True, importa apenas as variáveis em PI_CONTOUR_VARIABLES.
    batch_size : int
        Tamanho do batch de INSERT.

    Retorna
    -------
    dict com: inserted, skipped, errors, source_file, elapsed_s
    """
    path = Path(excel_path) if excel_path else _DEFAULT_PI_EXCEL
    if not path.exists():
        return {"ok": False, "error": f"Arquivo não encontrado: {path}", "inserted": 0}

    t0 = datetime.now(timezone.utc)
    cur = db_conn.cursor()
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")

    stats = {"inserted": 0, "skipped": 0, "errors": 0, "source_file": str(path)}
    batch: list[tuple] = []

    def flush():
        if not batch:
            return
        cur.executemany(
            """
            INSERT OR IGNORE INTO pi_vision_readings
                (run_id, tag, variable_name, channel, group_name,
                 timestamp, day_ref, value, quality, source, source_file, created_at)
            VALUES (?,?,?,?,?, ?,?,?,?,?,?,?)
            """,
            batch,
        )
        stats["inserted"] += cur.rowcount
        batch.clear()

    for row in _iter_rows(path, sheet):
        medidor = str(row.get("Medidor") or "").strip()
        variavel = str(row.get("Variavel") or "").strip()
        data_hora = _normalize_ts(row.get("DataHora"))
        valor = _float_or_none(row.get("Valor"))

        if not medidor or not variavel or not data_hora:
            stats["skipped"] += 1
            continue

        if only_authorized_variables and variavel not in PI_CONTOUR_VARIABLES:
            stats["skipped"] += 1
            continue

        canal = str(row.get("Canal") or "").strip()
        grupo = str(row.get("Grupo") or "").strip()
        quality = "Error" if grupo == _QUALITY_GROUP else "Good"
        day = _day_ref(data_hora)

        batch.append((
            run_id, medidor, variavel, canal, grupo,
            data_hora, day, valor, quality, "pi_vision", str(path), now_iso,
        ))

        if len(batch) >= batch_size:
            flush()

    flush()
    db_conn.commit()
    elapsed = (datetime.now(timezone.utc) - t0).total_seconds()
    stats.update({"ok": True, "elapsed_s": round(elapsed, 2)})
    return stats


def import_pi_excel_auto(db_conn: sqlite3.Connection, run_id: int | None = None) -> dict:
    """Importa usando o caminho padrão (BASE_UNICA_PI_OUTPUT env var)."""
    return import_pi_excel(db_conn, excel_path=_DEFAULT_PI_EXCEL, run_id=run_id)


def pi_readings_summary(db_conn: sqlite3.Connection) -> dict:
    """Retorna resumo das leituras PI armazenadas."""
    cur = db_conn.cursor()
    total = cur.execute("SELECT COUNT(*) FROM pi_vision_readings").fetchone()[0]
    if not total:
        return {"total": 0, "tags": [], "variables": [], "date_range": None}
    tags = [r[0] for r in cur.execute(
        "SELECT DISTINCT tag FROM pi_vision_readings ORDER BY tag"
    ).fetchall()]
    variables = [r[0] for r in cur.execute(
        "SELECT DISTINCT variable_name FROM pi_vision_readings ORDER BY variable_name"
    ).fetchall()]
    date_range = cur.execute(
        "SELECT MIN(day_ref), MAX(day_ref) FROM pi_vision_readings WHERE day_ref != ''"
    ).fetchone()
    return {
        "total": total,
        "tags": tags,
        "variables": variables,
        "date_range": {"from": date_range[0], "to": date_range[1]} if date_range else None,
    }
