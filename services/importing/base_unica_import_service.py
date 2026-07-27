from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
import json
import sqlite3

from openpyxl import load_workbook

from services.importing.monthly_workbook_service import BASE_UNICA_COLUMNS


BASE_UNICA_REQUIRED_COLUMNS = {
    "ProductionDate",
    "Granularity",
    "Origin",
    "Bank",
    "Loop",
    "Tipo",
    "Tag",
    "Instrumento",
    "SourceFile",
    "IsOfficial",
}

BASE_UNICA_META_COLUMNS = {
    "ProductionDate",
    "Hour",
    "Granularity",
    "Origin",
    "SourceType",
    "Area",
    "System",
    "Bank",
    "Loop",
    "Tipo",
    "Entity",
    "Tag",
    "Instrumento",
    "PI Tag",
    "SEP TAG",
    "SEP Medidor",
    "SEP Local",
    "SEP Status",
    "Bancos alinhados",
    "Fonte",
    "SourceFile",
    "IsOfficial",
}

SEP_EXPORT_TO_METRIC = {
    "SEP Óleo Vol. Bruto (m³) CV": "oil_m3",
    "SEP Óleo (t) CV": "oil_t",
    "SEP Gás (t) CV": "gas_t",
    "SEP Água (t) CV": "water_t",
    "SEP HC (t)": "hc_t",
    "SEP Total (t)": "total_t",
    "SEP Temperatura Méd. (°C)": "temp",
    "SEP Pressão Méd. (barg)": "pressure_barg",
}

RECON_EXPORT_TO_METRIC = {
    "Recon Cobertura": "Cobertura",
    "Recon Horas": "Horas",
    "Recon Daily Gás (t)": "Daily Gás (t)",
    "Recon Daily Óleo (t)": "Daily Óleo (t)",
    "Recon Daily HC (t)": "Daily HC (t)",
    "Recon Daily Água (t)": "Daily Água (t)",
    "Recon Soma h. Gás (t)": "Soma h. Gás (t)",
    "Recon Soma h. Óleo (t)": "Soma h. Óleo (t)",
    "Recon Soma h. HC (t)": "Soma h. HC (t)",
    "Recon Soma h. Água (t)": "Soma h. Água (t)",
    "Recon Δ Gás (t)": "Δ Gás (t)",
    "Recon Δ Óleo (t)": "Δ Óleo (t)",
    "Recon Δ HC (t)": "Δ HC (t)",
    "Recon Δ Água (t)": "Δ Água (t)",
    "Status Gás": "Status Gás",
    "Status Óleo": "Status Óleo",
    "Status HC": "Status HC",
    "Status Água": "Status Água",
}


def _as_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _as_date_text(value) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    raw = _as_text(value)
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw[:10], fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    return ""


def _as_hour_ref(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        try:
            iv = int(value)
            return iv if 0 <= iv <= 23 else None
        except Exception:
            return None
    raw = _as_text(value)
    if not raw:
        return None
    if hasattr(value, "hour"):
        try:
            return int(value.hour)
        except Exception:
            pass
    if ":" in raw:
        raw = raw.split(":", 1)[0]
    try:
        iv = int(raw)
        return iv if 0 <= iv <= 23 else None
    except Exception:
        return None


def _as_float(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        try:
            parsed = float(value)
            return parsed if parsed == parsed else None
        except Exception:
            return None
    raw = _as_text(value)
    if not raw:
        return None
    if "," in raw and "." in raw:
        if raw.rfind(",") > raw.rfind("."):
            raw = raw.replace(".", "").replace(",", ".")
        else:
            raw = raw.replace(",", "")
    else:
        raw = raw.replace(",", ".")
    try:
        parsed = float(raw)
        return parsed if parsed == parsed else None
    except Exception:
        return None


def _as_int_flag(value) -> int:
    raw = _as_text(value).lower()
    if raw in {"", "1", "true", "sim", "yes", "y"}:
        return 1
    if raw in {"0", "false", "nao", "não", "no", "n"}:
        return 0
    try:
        return 1 if float(raw) else 0
    except Exception:
        return 1


def _row_kind(origin: str, granularity: str) -> str:
    origin_norm = _as_text(origin).upper()
    granularity_norm = _as_text(granularity).lower()
    if origin_norm == "SEP":
        return "sep"
    if origin_norm == "RECON":
        return "recon"
    return "hourly" if granularity_norm == "hourly" else "daily"


def _metric_name_from_column(row_kind: str, column: str):
    if column in BASE_UNICA_META_COLUMNS:
        return None
    if row_kind == "sep":
        return SEP_EXPORT_TO_METRIC.get(column)
    if row_kind == "recon":
        return RECON_EXPORT_TO_METRIC.get(column)
    if column.startswith("SEP ") or column.startswith("Recon "):
        return None
    return column


def _metric_entries_from_row(row_kind: str, values: dict):
    out = []
    for column in BASE_UNICA_COLUMNS:
        metric_name = _metric_name_from_column(row_kind, column)
        if not metric_name:
            continue
        parsed = _as_float(values.get(column))
        if parsed is None:
            continue
        out.append((metric_name, parsed))
    return out


def parse_base_unica_workbook(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "BASE_UNICA_MES" not in wb.sheetnames:
        wb.close()
        raise ValueError("A planilha não contém a aba BASE_UNICA_MES.")
    ws = wb["BASE_UNICA_MES"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [_as_text(cell) for cell in (header_row or [])]
    header_index = {name: idx for idx, name in enumerate(headers) if name}
    missing = [column for column in BASE_UNICA_REQUIRED_COLUMNS if column not in header_index]
    if missing:
        wb.close()
        raise ValueError(f"Colunas obrigatórias ausentes na BASE_UNICA_MES: {', '.join(missing)}")

    month = ""
    row_groups = []
    metric_rows = []
    sep_daily = defaultdict(dict)
    days = set()
    banks = set()
    origins = defaultdict(int)
    metric_names = set()

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = {header: row[idx] if idx < len(row) else None for header, idx in header_index.items()}
        production_date = _as_date_text(values.get("ProductionDate"))
        if not production_date:
            continue
        current_month = production_date[:7]
        if not month:
            month = current_month
        elif current_month != month:
            wb.close()
            raise ValueError("A BASE_UNICA_MES deve conter apenas um único mês.")

        row_kind = _row_kind(values.get("Origin"), values.get("Granularity"))
        hour_ref = _as_hour_ref(values.get("Hour"))
        bank = "SEP" if row_kind == "sep" else _as_text(values.get("Bank")).upper()
        loop = _as_text(values.get("Loop"))
        tipo = _as_text(values.get("Tipo"))
        tag = _as_text(values.get("Tag"))
        instrument = _as_text(values.get("Instrumento") or values.get("SEP Medidor"))
        source_file = _as_text(values.get("SourceFile")) or workbook_path.name
        is_official = _as_int_flag(values.get("IsOfficial"))

        entries = _metric_entries_from_row(row_kind, values)
        if not entries:
            continue

        row_groups.append(
            {
                "production_date": production_date,
                "hour_ref": hour_ref,
                "row_kind": row_kind,
                "bank": bank,
                "loop": loop,
                "tipo": tipo,
                "tag": tag,
                "instrument": instrument,
                "source_file": source_file,
                "is_official": is_official,
                "row_number": row_number,
                "metrics": len(entries),
            }
        )
        days.add(production_date)
        if bank:
            banks.add(bank)
        origins[row_kind] += 1

        for metric_name, metric_value in entries:
            metric_names.add(metric_name)
            metric_rows.append(
                {
                    "day_ref": production_date,
                    "hour_ref": hour_ref,
                    "row_kind": row_kind,
                    "bank": bank,
                    "loop": loop,
                    "tipo": tipo,
                    "tag": tag,
                    "instrument": instrument,
                    "metric_name": metric_name,
                    "metric_value": metric_value,
                    "metric_unit": "",
                    "source_file": source_file,
                    "sheet_name": "BASE_UNICA_MES",
                    "is_official": is_official,
                }
            )
            if row_kind == "sep" and hour_ref is None:
                sep_daily[production_date][metric_name] = metric_value

    wb.close()

    if not month or not metric_rows:
        raise ValueError("Nenhuma linha válida foi encontrada na BASE_UNICA_MES.")

    return {
        "month": month,
        "row_groups": row_groups,
        "metric_rows": metric_rows,
        "sep_daily": dict(sep_daily),
        "summary": {
            "month": month,
            "line_groups": len(row_groups),
            "metric_rows": len(metric_rows),
            "days": len(days),
            "banks": sorted(banks),
            "origins": dict(origins),
            "metric_names": len(metric_names),
        },
    }


def preview_base_unica_import(db_conn_fn, workbook_path: Path) -> dict:
    parsed = parse_base_unica_workbook(workbook_path)
    month = parsed["month"]
    conn = db_conn_fn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    current_rows = cur.execute(
        """
        SELECT day_ref, hour_ref, row_kind, COALESCE(bank,'') AS bank, COALESCE(loop,'') AS loop,
               COALESCE(tipo,'') AS tipo, COALESCE(tag,'') AS tag, COALESCE(instrument,'') AS instrument,
               COALESCE(metric_name,'') AS metric_name, COALESCE(metric_value,0) AS metric_value
        FROM measurements_curated
        WHERE substr(day_ref,1,7)=?
        """,
        (month,),
    ).fetchall()
    conn.close()

    current_map = {
        (
            row["day_ref"],
            row["hour_ref"],
            row["row_kind"],
            row["bank"],
            row["loop"],
            row["tipo"],
            row["tag"],
            row["instrument"],
            row["metric_name"],
        ): float(row["metric_value"] or 0)
        for row in current_rows
    }
    incoming_map = {
        (
            row["day_ref"],
            row["hour_ref"],
            row["row_kind"],
            row["bank"],
            row["loop"],
            row["tipo"],
            row["tag"],
            row["instrument"],
            row["metric_name"],
        ): float(row["metric_value"] or 0)
        for row in parsed["metric_rows"]
    }

    current_keys = set(current_map)
    incoming_keys = set(incoming_map)
    new_keys = incoming_keys - current_keys
    removed_keys = current_keys - incoming_keys
    shared_keys = incoming_keys & current_keys
    changed_keys = {
        key for key in shared_keys
        if round(float(incoming_map[key]), 6) != round(float(current_map[key]), 6)
    }

    return {
        "ok": True,
        "mode": "preview",
        "file_name": workbook_path.name,
        "month": month,
        "import_summary": parsed["summary"],
        "diff": {
            "existing_metric_rows": len(current_map),
            "incoming_metric_rows": len(incoming_map),
            "new_metric_rows": len(new_keys),
            "removed_metric_rows": len(removed_keys),
            "changed_metric_rows": len(changed_keys),
            "unchanged_metric_rows": len(shared_keys) - len(changed_keys),
        },
        "affected": {
            "days": sorted({row["day_ref"] for row in parsed["metric_rows"]}),
            "banks": parsed["summary"]["banks"],
            "origins": parsed["summary"]["origins"],
        },
    }


def apply_base_unica_import(
    *,
    db_conn_fn,
    workbook_path: Path,
    build_backup_zip_fn,
    load_state_fn,
    save_state_fn,
    rebuild_validation_snapshot_for_month_fn,
    schedule_monthly_base_unica_fn,
    output_dir: Path,
    excel_name_fn,
    serialize_sep_row_fn,
):
    parsed = parse_base_unica_workbook(workbook_path)
    month = parsed["month"]
    year = month[:4]
    mon = month[5:7]
    date_from = f"{month}-01"
    date_to = f"{month}-31"
    workbook_out = output_dir / excel_name_fn(year, mon)
    backup_path = build_backup_zip_fn()
    now = datetime.now().isoformat(timespec="seconds")

    conn = db_conn_fn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    xml_rows = cur.execute(
        "SELECT id, file_path FROM xml042_documents WHERE substr(production_day,1,7)=?",
        (month,),
    ).fetchall()
    deleted_xml_paths = [Path(str(row["file_path"] or "").strip()) for row in xml_rows if str(row["file_path"] or "").strip()]

    run_ids = {
        int(row["run_id"])
        for row in cur.execute("SELECT DISTINCT run_id FROM measurements_curated WHERE substr(day_ref,1,7)=? AND run_id IS NOT NULL", (month,)).fetchall()
        if row["run_id"] is not None
    }
    run_ids.update(
        int(row["run_id"])
        for row in cur.execute("SELECT DISTINCT run_id FROM files_imported WHERE substr(content_date,1,7)=? AND run_id IS NOT NULL", (month,)).fetchall()
        if row["run_id"] is not None
    )
    run_ids.update(
        int(row["run_id"])
        for row in cur.execute("SELECT DISTINCT run_id FROM source_files_raw WHERE substr(content_date,1,7)=? AND run_id IS NOT NULL", (month,)).fetchall()
        if row["run_id"] is not None
    )
    source_file_ids = [
        int(row["id"])
        for row in cur.execute("SELECT id FROM source_files_raw WHERE substr(content_date,1,7)=?", (month,)).fetchall()
    ]
    sep_source_ids = [
        int(row["id"])
        for row in cur.execute("SELECT id FROM sep_source_files WHERE substr(production_date,1,7)=?", (month,)).fetchall()
    ]
    daily_card_ids = [
        int(row["id"])
        for row in cur.execute("SELECT id FROM daily_cards WHERE substr(production_date,1,7)=?", (month,)).fetchall()
    ]

    deleted = {
        "measurements_curated": 0,
        "validation_issues": 0,
        "daily_cards": 0,
        "daily_card_edits": 0,
        "mpfm_monitoring_daily": 0,
        "tpoc_daily_potential_curated": 0,
        "xml042_documents": 0,
        "recon_runs": 0,
        "sep_alignments": 0,
        "measurements_raw": 0,
        "parsing_events_raw": 0,
        "files_imported": 0,
        "source_files_raw": 0,
        "sep_source_files": 0,
        "processing_runs_pruned": 0,
    }

    try:
        cur.execute("BEGIN")
        run_id = cur.execute(
            """
            INSERT INTO processing_runs(started_at, source_type, source_ref, density, files_count, status, notes_json)
            VALUES (?, ?, ?, ?, ?, 'running', ?)
            """,
            (
                now,
                "excel-base-unica-import",
                workbook_path.name,
                0.0,
                1,
                json.dumps({"month": month, "mode": "base_unica_import"}, ensure_ascii=False),
            ),
        ).lastrowid

        deleted["measurements_curated"] = cur.execute(
            "DELETE FROM measurements_curated WHERE substr(day_ref,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["validation_issues"] = cur.execute(
            "DELETE FROM validation_issues WHERE substr(day_ref,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["recon_runs"] = cur.execute(
            "DELETE FROM recon_runs WHERE substr(day_ref,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["mpfm_monitoring_daily"] = cur.execute(
            "DELETE FROM mpfm_monitoring_daily WHERE substr(production_date,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["tpoc_daily_potential_curated"] = cur.execute(
            "DELETE FROM tpoc_daily_potential_curated WHERE substr(production_day,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["sep_alignments"] = cur.execute(
            "DELETE FROM sep_alignments WHERE substr(production_date,1,7)=?",
            (month,),
        ).rowcount or 0

        if daily_card_ids:
            placeholders = ",".join("?" * len(daily_card_ids))
            deleted["daily_card_edits"] = cur.execute(
                f"DELETE FROM daily_card_edits WHERE daily_card_id IN ({placeholders})",
                daily_card_ids,
            ).rowcount or 0
        deleted["daily_cards"] = cur.execute(
            "DELETE FROM daily_cards WHERE substr(production_date,1,7)=?",
            (month,),
        ).rowcount or 0

        if source_file_ids:
            placeholders = ",".join("?" * len(source_file_ids))
            deleted["parsing_events_raw"] = cur.execute(
                f"DELETE FROM parsing_events_raw WHERE source_file_raw_id IN ({placeholders})",
                source_file_ids,
            ).rowcount or 0
        if source_file_ids and sep_source_ids:
            placeholders_sf = ",".join("?" * len(source_file_ids))
            placeholders_sep = ",".join("?" * len(sep_source_ids))
            deleted["measurements_raw"] = cur.execute(
                f"DELETE FROM measurements_raw WHERE source_file_raw_id IN ({placeholders_sf}) OR source_record_id IN ({placeholders_sep}) OR substr(content_date,1,7)=?",
                source_file_ids + sep_source_ids + [month],
            ).rowcount or 0
        elif source_file_ids:
            placeholders_sf = ",".join("?" * len(source_file_ids))
            deleted["measurements_raw"] = cur.execute(
                f"DELETE FROM measurements_raw WHERE source_file_raw_id IN ({placeholders_sf}) OR substr(content_date,1,7)=?",
                source_file_ids + [month],
            ).rowcount or 0
        elif sep_source_ids:
            placeholders_sep = ",".join("?" * len(sep_source_ids))
            deleted["measurements_raw"] = cur.execute(
                f"DELETE FROM measurements_raw WHERE source_record_id IN ({placeholders_sep}) OR substr(content_date,1,7)=?",
                sep_source_ids + [month],
            ).rowcount or 0
        else:
            deleted["measurements_raw"] = cur.execute(
                "DELETE FROM measurements_raw WHERE substr(content_date,1,7)=?",
                (month,),
            ).rowcount or 0

        deleted["files_imported"] = cur.execute(
            "DELETE FROM files_imported WHERE substr(content_date,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["sep_source_files"] = cur.execute(
            "DELETE FROM sep_source_files WHERE substr(production_date,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["source_files_raw"] = cur.execute(
            "DELETE FROM source_files_raw WHERE substr(content_date,1,7)=?",
            (month,),
        ).rowcount or 0
        deleted["xml042_documents"] = cur.execute(
            "DELETE FROM xml042_documents WHERE substr(production_day,1,7)=?",
            (month,),
        ).rowcount or 0

        rows_to_insert = [
            (
                run_id,
                row["source_file"],
                workbook_path.name,
                row["sheet_name"],
                row["row_kind"],
                row["day_ref"],
                row["hour_ref"],
                row["bank"],
                row["loop"],
                row["tipo"],
                row["tag"],
                row["instrument"],
                row["metric_name"],
                row["metric_value"],
                row["metric_unit"],
                row["is_official"],
                now,
            )
            for row in parsed["metric_rows"]
        ]
        cur.executemany(
            """
            INSERT INTO measurements_curated(
                run_id, source_file, excel_file, sheet_name, row_kind, day_ref, hour_ref,
                bank, loop, tipo, tag, instrument, metric_name, metric_value, metric_unit,
                is_official, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows_to_insert,
        )

        for orphan_run_id in run_ids:
            has_remaining = (
                cur.execute("SELECT 1 FROM files_imported WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM source_files_raw WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM measurements_curated WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM measurements_raw WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM validation_issues WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM parsing_events_raw WHERE run_id=? LIMIT 1", (orphan_run_id,)).fetchone()
            )
            if has_remaining:
                continue
            deleted["processing_runs_pruned"] += cur.execute(
                "DELETE FROM processing_runs WHERE id=?",
                (orphan_run_id,),
            ).rowcount or 0

        cur.execute(
            "UPDATE processing_runs SET finished_at=?, status='ok', notes_json=? WHERE id=?",
            (
                datetime.now().isoformat(timespec="seconds"),
                json.dumps(
                    {
                        "month": month,
                        "mode": "base_unica_import",
                        "row_groups": parsed["summary"]["line_groups"],
                        "metric_rows": parsed["summary"]["metric_rows"],
                    },
                    ensure_ascii=False,
                ),
                run_id,
            ),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        raise
    finally:
        if conn:
            conn.close()

    for xml_path in deleted_xml_paths:
        try:
            if xml_path.exists():
                xml_path.unlink()
        except Exception:
            pass

    state = load_state_fn(year, mon)
    state["processed"] = []
    state["processed_hours_by_key"] = {}
    state["processed_hours"] = {}
    state["sep_by_day"] = {
        day: serialize_sep_row_fn({"DAY": metrics})
        for day, metrics in parsed["sep_daily"].items()
    }
    state["sep_days"] = sorted(parsed["sep_daily"].keys())
    state["last_run"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_state_fn(state)

    validation_snapshot = rebuild_validation_snapshot_for_month_fn(month)
    queued = schedule_monthly_base_unica_fn(workbook_out, year, mon)

    return {
        "ok": True,
        "mode": "apply",
        "file_name": workbook_path.name,
        "month": month,
        "backup_file": backup_path.name if backup_path else "",
        "deleted": deleted,
        "imported": parsed["summary"],
        "validation_snapshot": validation_snapshot.get("recomputed", {}),
        "queued": queued,
        "workbook_name": workbook_out.name,
        "workbook_exists": workbook_out.exists(),
    }
