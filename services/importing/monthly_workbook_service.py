from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import posixpath
from tempfile import NamedTemporaryFile
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET
import warnings

from routes.date_utils import normalize_validation_issue_day_ref
from services.excel_template_service import (
    MONTHLY_WORKBOOK_TEMPLATE,
    clear_value_region,
    center_cell_content,
    center_filled_cells,
    ensure_workbook_from_template,
    reset_sheet_from_template,
    reset_tabular_sheet_from_template,
    seed_row_from_template,
)

warnings.filterwarnings(
    "ignore",
    message="Conditional Formatting extension is not supported and will be removed",
    category=UserWarning,
)

BASE_UNICA_COLUMNS = [
    "ProductionDate","Hour","Granularity","Origin","SourceType","Area","System","Bank","Loop","Tipo","Entity","Tag","Instrumento","PI Tag",
    "MPFM uncorr Gás (t)","MPFM uncorr Óleo (t)","MPFM uncorr HC (t)","MPFM uncorr Água (t)","MPFM uncorr Total (t)",
    "MPFM corr Gás (t)","MPFM corr Óleo (t)","MPFM corr HC (t)","MPFM corr Água (t)","MPFM corr Total (t)",
    "PVT mass Gás (t)","PVT mass Óleo (t)","PVT mass HC (t)","PVT mass Água (t)","PVT mass Total (t)",
    "PVT vol Gás (Sm³)","PVT vol Óleo (m³)","PVT vol HC (m³)","PVT vol Água (m³)","PVT vol Total (m³)",
    "PVT @20 mass Gás (t)","PVT @20 mass Óleo (t)","PVT @20 mass HC (t)","PVT @20 mass Água (t)","PVT @20 mass Total (t)",
    "PVT @20 vol Gás (Sm³)","PVT @20 vol Óleo (m³)","PVT @20 vol HC (m³)","PVT @20 vol Água (m³)","PVT @20 vol Total (m³)",
    "Pressão (barg)","Temperatura (°C)","Dens. Gás (kg/m³)","Dens. Óleo (kg/m³)","Dens. Água (kg/m³)",
    "SEP TAG","SEP Medidor","SEP Local","SEP Status","Bancos alinhados",
    "SEP Óleo Vol. Bruto (m³) CV","SEP Óleo (t) CV","SEP Gás (t) CV","SEP Água (t) CV","SEP HC (t)","SEP Total (t)","SEP Temperatura Méd. (°C)","SEP Pressão Méd. (barg)",
    "Recon Cobertura","Recon Horas","Recon Daily Gás (t)","Recon Daily Óleo (t)","Recon Daily HC (t)","Recon Daily Água (t)",
    "Recon Soma h. Gás (t)","Recon Soma h. Óleo (t)","Recon Soma h. HC (t)","Recon Soma h. Água (t)",
    "Recon Δ Gás (t)","Recon Δ Óleo (t)","Recon Δ HC (t)","Recon Δ Água (t)",
    "Status Gás","Status Óleo","Status HC","Status Água","Fonte","SourceFile","IsOfficial"
]


def excel_name(month_pt: dict, year, month):
    return f"MPFM_{month_pt.get(month, month)}_{year}.xlsx"


def cleanup_workbook(workbook_path: Path):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(workbook_path)
        _cleanup_workbook_in_memory(wb)
        wb.save(workbook_path)
        wb.close()
        _normalize_base_unica_table_xml(workbook_path)
    except Exception:
        pass


def _last_filled_row(ws, start_row: int, start_col: int, column_count: int) -> int:
    last_row = start_row
    end_col = start_col + max(column_count - 1, 0)
    for row_idx in range(ws.max_row, start_row - 1, -1):
        for col_idx in range(start_col, end_col + 1):
            value = ws.cell(row_idx, col_idx).value
            if value not in (None, ""):
                return row_idx
    return last_row


def _rebuild_base_unica_table(ws, header_row: int = 1, start_col: int = 2, column_names=None, data_rows: int | None = None) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.filters import AutoFilter
    from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

    _col_list = list(column_names) if column_names is not None else []
    headers = [str(name) for name in (_col_list if _col_list else BASE_UNICA_COLUMNS)]
    if not headers:
        return

    last_row = max(header_row, header_row + max(data_rows or 0, 0))
    last_col_idx = start_col + len(headers) - 1
    last_col = get_column_letter(last_col_idx)
    table_ref = f"{get_column_letter(start_col)}{header_row}:{last_col}{last_row}"

    ws.auto_filter.ref = table_ref
    ws._tables.clear()

    table = Table(displayName="tblBaseUnicaMes", ref=table_ref)
    table.autoFilter = AutoFilter(ref=table_ref)
    table.tableColumns = [TableColumn(id=idx, name=header) for idx, header in enumerate(headers, start=1)]
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _cleanup_workbook_in_memory(wb):
    keep = {
        "BASE_UNICA_MES",
        "DAILYS",
        "HOURLYS",
        "RECON",
        "STATUS_MES",
        "ALERTAS_MES",
    }
    auxiliary_hidden = set()
    remove = [ws.title for ws in wb.worksheets if ws.title not in keep]
    for title in remove:
        del wb[title]
    for ws in wb.worksheets:
        if ws.title in auxiliary_hidden:
            ws.sheet_state = "hidden"
    if "BASE_UNICA_MES" in wb.sheetnames:
        base_ws = wb["BASE_UNICA_MES"]
        data_last_row = _last_filled_row(base_ws, 1, 2, len(BASE_UNICA_COLUMNS))
        _rebuild_base_unica_table(
            base_ws,
            header_row=1,
            start_col=2,
            column_names=BASE_UNICA_COLUMNS,
            data_rows=max(data_last_row - 1, 0),
        )


def _table_part_for_sheet(xlsx_path: Path, sheet_name: str) -> str | None:
    sheet_part = _sheet_part_map(xlsx_path).get(sheet_name)
    if not sheet_part:
        return None

    rels_path = f"xl/worksheets/_rels/{Path(sheet_part).stem}.xml.rels"
    table_rel_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/table"
    ns_rel = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}

    with ZipFile(xlsx_path) as zf:
        if rels_path not in zf.namelist():
            return None
        rels_xml = ET.fromstring(zf.read(rels_path))

    for rel in rels_xml.findall("rel:Relationship", ns_rel):
        if rel.attrib.get("Type") != table_rel_type:
            continue
        target = rel.attrib.get("Target", "")
        if not target:
            continue
        normalized_target = posixpath.normpath(posixpath.join(posixpath.dirname(sheet_part), target))
        return normalized_target.lstrip("/")
    return None


def _normalize_base_unica_table_xml(workbook_path: Path) -> None:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        return

    table_part = _table_part_for_sheet(workbook_path, "BASE_UNICA_MES")
    if not table_part:
        return

    ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ET.register_namespace("", ns["main"])

    with ZipFile(workbook_path) as zf:
        if table_part not in zf.namelist():
            return
        table_root = ET.fromstring(zf.read(table_part))

    table_ref = table_root.attrib.get("ref")
    if not table_ref:
        return

    auto_filter = table_root.find("main:autoFilter", ns)
    if auto_filter is None:
        auto_filter = ET.SubElement(table_root, f"{{{ns['main']}}}autoFilter")
    auto_filter.attrib["ref"] = table_ref

    table_columns = table_root.find("main:tableColumns", ns)
    if table_columns is not None:
        columns = list(table_columns)
        table_columns.attrib["count"] = str(len(columns))
        for index, column in enumerate(columns, start=1):
            column.attrib["id"] = str(index)

    # Preserve the OOXML-standard XML declaration required by Excel.
    # ET.tostring with encoding="unicode" returns a str (no declaration, no BOM);
    # we then encode to UTF-8 bytes and prepend the mandatory declaration with
    # standalone='yes' exactly as openpyxl originally writes it.
    _xml_decl = b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n"
    _table_body = ET.tostring(table_root, encoding="unicode").encode("utf-8")
    replacements = {
        table_part: _xml_decl + _table_body,
    }

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        temp_path = Path(tmp_file.name)
    try:
        with ZipFile(workbook_path) as source_zip, ZipFile(temp_path, "w", ZIP_DEFLATED) as target_zip:
            for item in source_zip.infolist():
                payload = replacements.get(item.filename)
                if payload is None:
                    payload = source_zip.read(item.filename)
                target_zip.writestr(item.filename, payload, compress_type=ZIP_DEFLATED)
        temp_path.replace(workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


PRESERVE_TEMPLATE_SHEETS = {
    "P1 Choke minuto % 1-31_01",
}


def _sheet_part_map(xlsx_path: Path):
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    workbook_rel_ns = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
    with ZipFile(xlsx_path) as zf:
        workbook_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        workbook_rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {}
    for rel in workbook_rels_xml.findall("rel:Relationship", ns_rel):
        target = rel.attrib["Target"].lstrip("/")
        rel_map[rel.attrib["Id"]] = target if target.startswith("xl/") else f"xl/{target}"
    sheet_map = {}
    for sheet in workbook_xml.find("main:sheets", ns_main):
        title = sheet.attrib["name"]
        rel_id = sheet.attrib.get(workbook_rel_ns)
        if rel_id and rel_id in rel_map:
            sheet_map[title] = rel_map[rel_id]
    return sheet_map


def _preserve_template_sheet_parts(workbook_path: Path, template_path: Path, sheet_names: set[str]) -> None:
    if not workbook_path.exists() or not template_path.exists():
        return
    template_parts = _sheet_part_map(template_path)
    workbook_parts = _sheet_part_map(workbook_path)
    replacements = {}
    with ZipFile(template_path) as template_zip:
        for sheet_name in sheet_names:
            template_part = template_parts.get(sheet_name)
            workbook_part = workbook_parts.get(sheet_name)
            if not template_part or not workbook_part:
                continue
            replacements[workbook_part] = template_zip.read(template_part)
            template_rel = f"xl/worksheets/_rels/{Path(template_part).stem}.xml.rels"
            workbook_rel = f"xl/worksheets/_rels/{Path(workbook_part).stem}.xml.rels"
            if template_rel in template_zip.namelist():
                replacements[workbook_rel] = template_zip.read(template_rel)

    if not replacements:
        return

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        temp_path = Path(tmp_file.name)
    try:
        with ZipFile(workbook_path) as source_zip, ZipFile(temp_path, "w", ZIP_DEFLATED) as target_zip:
            copied = set()
            for item in source_zip.infolist():
                if item.filename in replacements:
                    target_zip.writestr(
                        item.filename,
                        replacements[item.filename],
                        compress_type=ZIP_DEFLATED,
                    )
                    copied.add(item.filename)
                else:
                    target_zip.writestr(
                        item.filename,
                        source_zip.read(item.filename),
                        compress_type=ZIP_DEFLATED,
                    )
            for name, payload in replacements.items():
                if name not in copied:
                    target_zip.writestr(name, payload, compress_type=ZIP_DEFLATED)
        temp_path.replace(workbook_path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


# Sheets that are always managed by the application — never treated as user annotation sheets.
_APP_MANAGED_SHEET_NAMES = {
    "BASE_UNICA_MES", "DAILYS", "HOURLYS", "RECON",
    "STATUS_MES", "CARDS_RESUMO", "ALERTAS_MES",
}


def _save_user_hidden_sheets(workbook_path: Path) -> list:
    """Return cell data for hidden sheets that are not managed by the app."""
    if not workbook_path.exists():
        return []
    try:
        from openpyxl import load_workbook as _lw
        source_wb = _lw(workbook_path, read_only=True, data_only=True)
        saved = []
        for ws in source_wb.worksheets:
            if ws.sheet_state in ("hidden", "veryHidden") and ws.title not in _APP_MANAGED_SHEET_NAMES:
                saved.append({
                    "title": ws.title,
                    "state": ws.sheet_state,
                    "rows": list(ws.iter_rows(values_only=True)),
                })
        source_wb.close()
        return saved
    except Exception:
        return []


def _restore_user_hidden_sheets(wb, saved_sheets: list) -> None:
    """Re-add previously saved hidden sheets into the workbook."""
    for sheet_data in saved_sheets:
        title = sheet_data["title"]
        if title in wb.sheetnames:
            continue
        new_ws = wb.create_sheet(title)
        new_ws.sheet_state = sheet_data["state"]
        for row_idx, row in enumerate(sheet_data["rows"], start=1):
            for col_idx, value in enumerate(row, start=1):
                if value is not None:
                    new_ws.cell(row=row_idx, column=col_idx, value=value)


def build_monthly_base_unica(
    db_conn_fn,
    workbook_path: Path,
    year: str,
    month: str,
    *,
    write_cards_to_workbook_fn,
    serialize_sep_row_fn,
    month_pt: dict,
    engine,
    load_state_fn=None,
):
    import pandas as pd
    from openpyxl import Workbook, load_workbook

    ensure_workbook_from_template(workbook_path, MONTHLY_WORKBOOK_TEMPLATE)

    # Keep month scope for the workbook file name, but read directly from
    # measurements_curated to avoid implicit view-level filtering.
    date_from, date_to = f"{year}-{month}-01", f"{year}-{month}-31"
    month_prefix = f"{year}-{month}-"
    conn = db_conn_fn()
    cur = conn.cursor()
    rows = cur.execute(
        """
        SELECT day_ref, hour_ref, row_kind, bank, loop, tipo, tag, instrument, metric_name, metric_value, source_file
        FROM measurements_curated
        WHERE day_ref LIKE ?
          AND row_kind IN ('hourly','daily','recon','sep')
        ORDER BY day_ref, COALESCE(hour_ref,-1), row_kind, bank, tag, metric_name
        """,
        (f"{month_prefix}%",),
    ).fetchall()
    aligns = {
        row["production_date"]: row["banks"]
        for row in cur.execute(
            """
            SELECT production_date, GROUP_CONCAT(bank, ', ') as banks
            FROM sep_alignments WHERE is_active=1 AND production_date BETWEEN ? AND ?
            GROUP BY production_date
            """,
            (date_from, date_to),
        ).fetchall()
    }
    alerts = []
    for row in cur.execute(
        """
        SELECT day_ref, severity, issue_type, ref_key, details, created_at
        FROM validation_issues
        ORDER BY id
        """
    ).fetchall():
        item = dict(row)
        item["day_ref"] = normalize_validation_issue_day_ref(item.get("day_ref", ""), item.get("created_at", ""))
        if item["day_ref"] and item["day_ref"] < date_from:
            continue
        if item["day_ref"] and item["day_ref"] > date_to:
            continue
        alerts.append(item)
    conn.close()

    state = load_state_fn(year, month) if load_state_fn else {"yr": year, "mo": month}
    status_df = engine.build_status_sheet(state)

    if not rows and not alerts:
        ensure_workbook_from_template(workbook_path, MONTHLY_WORKBOOK_TEMPLATE)
        wb = load_workbook(workbook_path)
        template_wb = load_workbook(MONTHLY_WORKBOOK_TEMPLATE) if MONTHLY_WORKBOOK_TEMPLATE.exists() else None
        _replace_sheet_in_workbook(wb, template_wb, "STATUS_MES", status_df)
        _cleanup_workbook_in_memory(wb)
        wb.save(workbook_path)
        wb.close()
        if template_wb:
            template_wb.close()
        return

    piv = defaultdict(dict)
    meta = {}
    for row in rows:
        key = (row["day_ref"], row["hour_ref"], row["row_kind"], row["bank"], row["loop"], row["tipo"], row["tag"], row["instrument"], row["source_file"])
        piv[key][row["metric_name"]] = row["metric_value"]
        meta[key] = {
            "ProductionDate": row["day_ref"],
            "Hour": "" if row["hour_ref"] is None else f"{int(row['hour_ref']):02d}:00",
            "Granularity": "Hourly" if row["row_kind"] == "hourly" else "Daily" if row["row_kind"] == "daily" else "Recon" if row["row_kind"] == "recon" else ("Hourly" if row["hour_ref"] is not None else "Daily"),
            "Origin": "SEP" if row["row_kind"] == "sep" else ("RECON" if row["row_kind"] == "recon" else "MPFM"),
            "SourceType": "TXT" if row["row_kind"] == "sep" else ("CALC" if row["row_kind"] == "recon" else "PDF"),
            "Area": "",
            "System": "",
            "Bank": "" if row["bank"] == "SEP" else (row["bank"] or ""),
            "Loop": row["loop"] or "",
            "Tipo": row["tipo"] or "",
            "Entity": row["tag"] or "",
            "Tag": row["tag"] or "",
            "Instrumento": row["instrument"] or "",
            "Fonte": "Separador" if row["row_kind"] == "sep" else ("Reconciliação" if row["row_kind"] == "recon" else "MPFM"),
            "SourceFile": row["source_file"] or "",
            "IsOfficial": 1,
        }
    out_rows = []
    for key in sorted(piv.keys(), key=lambda item: (item[0], -1 if item[1] is None else item[1], item[2], item[3], item[6])):
        out_row = {column: "" for column in BASE_UNICA_COLUMNS}
        out_row.update(meta[key])
        values = piv[key]
        for column in BASE_UNICA_COLUMNS:
            if column in values:
                out_row[column] = values[column]
        if meta[key]["Origin"] == "SEP":
            out_row["SEP TAG"] = meta[key]["Tag"] or "SEP"
            out_row["SEP Medidor"] = meta[key]["Instrumento"]
            out_row["SEP Local"] = meta[key]["Loop"] or meta[key]["Tipo"]
            out_row["SEP Status"] = "Aplicado" if aligns.get(meta[key]["ProductionDate"]) else "Extraído"
            out_row["Bancos alinhados"] = aligns.get(meta[key]["ProductionDate"], "")
            metric_map = {
                "oil_m3": "SEP Óleo Vol. Bruto (m³) CV",
                "oil_t": "SEP Óleo (t) CV",
                "gas_t": "SEP Gás (t) CV",
                "water_t": "SEP Água (t) CV",
                "hc_t": "SEP HC (t)",
                "total_t": "SEP Total (t)",
                "temp": "SEP Temperatura Méd. (°C)",
                "pressure_barg": "SEP Pressão Méd. (barg)",
            }
            for metric, column in metric_map.items():
                if metric in values:
                    out_row[column] = values[metric]
        if meta[key]["Origin"] == "RECON":
            recon_map = {
                "Cobertura": "Recon Cobertura",
                "Horas": "Recon Horas",
                "Daily Gás (t)": "Recon Daily Gás (t)",
                "Daily Óleo (t)": "Recon Daily Óleo (t)",
                "Daily HC (t)": "Recon Daily HC (t)",
                "Daily Água (t)": "Recon Daily Água (t)",
                "Soma h. Gás (t)": "Recon Soma h. Gás (t)",
                "Soma h. Óleo (t)": "Recon Soma h. Óleo (t)",
                "Soma h. HC (t)": "Recon Soma h. HC (t)",
                "Soma h. Água (t)": "Recon Soma h. Água (t)",
                "Δ Gás (t)": "Recon Δ Gás (t)",
                "Δ Óleo (t)": "Recon Δ Óleo (t)",
                "Δ HC (t)": "Recon Δ HC (t)",
                "Δ Água (t)": "Recon Δ Água (t)",
                "Status Gás": "Status Gás",
                "Status Óleo": "Status Óleo",
                "Status HC": "Status HC",
                "Status Água": "Status Água",
            }
            for metric, column in recon_map.items():
                if metric in values:
                    out_row[column] = values[metric]
        out_rows.append(out_row)

    df = pd.DataFrame(out_rows, columns=BASE_UNICA_COLUMNS) if out_rows else pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    _user_hidden_sheets = _save_user_hidden_sheets(workbook_path)
    template_wb = load_workbook(MONTHLY_WORKBOOK_TEMPLATE) if MONTHLY_WORKBOOK_TEMPLATE.exists() else None
    if template_wb:
        wb = load_workbook(MONTHLY_WORKBOOK_TEMPLATE)
    else:
        wb = load_workbook(workbook_path) if workbook_path.exists() else Workbook()

    _replace_sheet_in_workbook(wb, template_wb, "BASE_UNICA_MES", df)

    df_daily = df[df["Granularity"].eq("Daily") & ~df["Origin"].eq("RECON")].copy()
    df_hourly = df[df["Granularity"].eq("Hourly") & ~df["Origin"].eq("RECON")].copy()
    df_recon = df[df["Origin"].eq("RECON")].copy()
    _replace_sheet_in_workbook(wb, template_wb, "DAILYS", df_daily if not df_daily.empty else pd.DataFrame(columns=BASE_UNICA_COLUMNS))
    _replace_sheet_in_workbook(wb, template_wb, "HOURLYS", df_hourly if not df_hourly.empty else pd.DataFrame(columns=BASE_UNICA_COLUMNS))
    _replace_sheet_in_workbook(wb, template_wb, "RECON", df_recon if not df_recon.empty else pd.DataFrame(columns=BASE_UNICA_COLUMNS))

    alerts_df = pd.DataFrame(alerts, columns=["day_ref", "severity", "issue_type", "ref_key", "details", "created_at"])
    _replace_sheet_in_workbook(wb, template_wb, "ALERTAS_MES", alerts_df)
    _replace_sheet_in_workbook(wb, template_wb, "STATUS_MES", status_df)
    for ws in wb.worksheets:
        center_filled_cells(ws)
    _cleanup_workbook_in_memory(wb)
    if _user_hidden_sheets:
        _restore_user_hidden_sheets(wb, _user_hidden_sheets)
    wb.save(workbook_path)
    wb.close()
    _normalize_base_unica_table_xml(workbook_path)
    if template_wb:
        template_wb.close()
        # Restoring raw template sheet XML breaks sharedStrings coherence in the
        # saved workbook and produces invalid .xlsx files. Keep the workbook
        # generated by openpyxl until we have a safe merge strategy for those
        # advanced template parts.


def _replace_sheet_in_workbook(wb, template_wb, sheet_name: str, df):
    from openpyxl.utils import get_column_letter

    lightweight_template_sheets = {
        "BASE_UNICA_MES",
        "DAILYS",
        "HOURLYS",
        "RECON",
        "ALERTAS_MES",
        "STATUS_MES",
    }
    start_col_map = {
        "BASE_UNICA_MES": 2,
        "DAILYS": 3,
        "HOURLYS": 3,
        "RECON": 3,
        "ALERTAS_MES": 3,
        "STATUS_MES": 1,
    }

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        template_ws = template_wb[sheet_name] if template_wb and sheet_name in template_wb.sheetnames else None
        start_col = start_col_map.get(sheet_name, 3 if template_ws else 1)
    elif template_wb and sheet_name in template_wb.sheetnames:
        if sheet_name in lightweight_template_sheets:
            ws, template_ws = reset_tabular_sheet_from_template(wb, template_wb, sheet_name)
        else:
            ws, template_ws = reset_sheet_from_template(wb, template_wb, sheet_name)
        start_col = start_col_map.get(sheet_name, 3)
    else:
        ws = wb.create_sheet(sheet_name)
        template_ws = None
        start_col = 1

    header_row = 1
    data_row = 2
    clear_value_region(ws, header_row, start_col)

    for column_offset, column in enumerate(df.columns):
        cell = ws.cell(header_row, start_col + column_offset, column)
        center_cell_content(cell)

    for row_offset, row in enumerate(df.itertuples(index=False), start=data_row):
        if template_ws and row_offset > template_ws.max_row:
            style_row = data_row if (row_offset - data_row) % 2 == 0 else min(data_row + 1, template_ws.max_row)
            seed_row_from_template(
                ws,
                template_ws,
                row_offset,
                style_row,
                start_col,
                start_col + max(len(df.columns) - 1, 0),
            )
        for column_offset, value in enumerate(row):
            cell = ws.cell(row_offset, start_col + column_offset, value)
            if value not in (None, ""):
                center_cell_content(cell)

    if df.columns.size:
        last_col = get_column_letter(start_col + len(df.columns) - 1)
        ws.auto_filter.ref = f"{get_column_letter(start_col)}{header_row}:{last_col}{max(ws.max_row, header_row)}"
        if sheet_name == "BASE_UNICA_MES":
            _rebuild_base_unica_table(
                ws,
                header_row=header_row,
                start_col=start_col,
                column_names=df.columns,
                data_rows=int(df.shape[0]),
            )
