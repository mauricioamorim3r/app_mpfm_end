"""
Captura Choke % de um Excel de produção já resolvido (com PITimeDat calculado)
e persiste em well_choke_history.

Fluxo:
  1. Usuário gera o Excel de produção via /api/export-producao-excel
  2. Abre o arquivo no Excel com PI DataLink instalado → fórmulas PITimeDat resolvem
  3. Salva o arquivo como .xlsx (valores fixados)
  4. Faz upload para POST /api/admin/choke-history/import-excel
  5. Este serviço lê a coluna "Choke %" da aba DIARIOS e insere em well_choke_history

Alternativamente, pode ser chamado com --excel-path via linha de comando.
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator


_SHEET_NAME = "DIARIOS"
_COL_TAG = "TAG"
_COL_DATE = "Data"
_COL_CHOKE = "Choke %"
_HEADER_ROW = 4  # linha 4 do template tem os cabeçalhos (1-based)

_CHOKE_TAGS = frozenset(["PE_4", "PE_2", "PW-104DA"])


def _iter_diarios_rows(excel_path: Path) -> Iterator[dict]:
    """Lê a aba DIARIOS e itera linhas com TAG, Data e Choke %."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl não instalado — execute: pip install openpyxl")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if _SHEET_NAME not in wb.sheetnames:
        return

    ws = wb[_SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))

    if len(all_rows) < _HEADER_ROW:
        return

    headers = [str(c).strip() if c is not None else "" for c in all_rows[_HEADER_ROW - 1]]

    try:
        tag_idx   = headers.index(_COL_TAG)
        date_idx  = headers.index(_COL_DATE)
        choke_idx = headers.index(_COL_CHOKE)
    except ValueError:
        return

    for row in all_rows[_HEADER_ROW:]:
        if not row or all(c is None for c in row):
            continue
        tag  = str(row[tag_idx]  or "").strip()
        date = row[date_idx]
        chk  = row[choke_idx]
        yield {"tag": tag, "date": date, "choke": chk}


def _normalize_date(val) -> str:
    if val is None:
        return ""
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return s[:10]


def _float_or_none(val):
    try:
        v = float(val)
        return None if v != v else round(v, 4)
    except (TypeError, ValueError):
        return None


def import_choke_from_excel(
    db_conn: sqlite3.Connection,
    excel_path: str | Path,
    source_label: str = "excel_resolved",
) -> dict:
    """
    Lê a aba DIARIOS de um Excel resolvido e insere Choke % em well_choke_history.

    Retorna dict com: inserted, skipped, errors, source_file, elapsed_s
    """
    path = Path(excel_path)
    if not path.exists():
        return {"ok": False, "error": f"Arquivo não encontrado: {path}", "inserted": 0}

    t0 = datetime.now(timezone.utc)
    cur = db_conn.cursor()
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")

    stats = {"inserted": 0, "skipped": 0, "errors": 0, "source_file": str(path)}
    batch: list[tuple] = []

    for row in _iter_diarios_rows(path):
        tag = row["tag"]
        if not tag or tag not in _CHOKE_TAGS:
            stats["skipped"] += 1
            continue

        day_ref = _normalize_date(row["date"])
        if not day_ref:
            stats["skipped"] += 1
            continue

        choke_pct = _float_or_none(row["choke"])
        if choke_pct is None:
            stats["skipped"] += 1
            continue

        batch.append((tag, day_ref, choke_pct, source_label, now_iso))

    if batch:
        cur.executemany(
            """
            INSERT OR REPLACE INTO well_choke_history
                (tag, day_ref, choke_pct, source, created_at)
            VALUES (?,?,?,?,?)
            """,
            batch,
        )
        stats["inserted"] = cur.rowcount
        db_conn.commit()

    elapsed = (datetime.now(timezone.utc) - t0).total_seconds()
    stats.update({"ok": True, "elapsed_s": round(elapsed, 2)})
    return stats


def choke_history_summary(db_conn: sqlite3.Connection) -> dict:
    """Resumo do histórico de Choke % armazenado."""
    cur = db_conn.cursor()
    total = cur.execute("SELECT COUNT(*) FROM well_choke_history").fetchone()[0]
    if not total:
        return {"total": 0, "tags": [], "date_range": None}
    tags = [r[0] for r in cur.execute(
        "SELECT DISTINCT tag FROM well_choke_history ORDER BY tag"
    ).fetchall()]
    date_range = cur.execute(
        "SELECT MIN(day_ref), MAX(day_ref) FROM well_choke_history WHERE day_ref != ''"
    ).fetchone()
    return {
        "total": total,
        "tags": tags,
        "date_range": {"from": date_range[0], "to": date_range[1]} if date_range else None,
    }
