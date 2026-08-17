from __future__ import annotations

import hashlib
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_XML042_CNPJ8 = "04028583"
DEFAULT_AUTHOR = "local-user"
DEFAULT_XML042_DEST_DIR = r"C:\Users\MAUAM\OneDrive - Equinor\DPB FPSO Bacalhau - Metering - 02 MULTIPHASE MANAGEMENT SYSTEM\3. Registros de Operação SGM Multifasico\3.7 Registros XML 042"


def _norm_tag(normalize_tag_name, value: str) -> str:
    return normalize_tag_name(str(value or "").strip())


def _parse_month_bounds(month: str) -> tuple[str, str]:
    raw = str(month or "").strip()
    if len(raw) != 7 or raw[4] != "-":
        return "", ""
    return f"{raw}-01", f"{raw}-31"


def _to_float(value):
    try:
        raw = str(value).strip() if value is not None else ""
        if not raw:
            return None
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        parsed = float(raw)
        return parsed if parsed == parsed else None
    except Exception:
        return None


def _format_decimal_ptbr(value, digits: int = 5) -> str:
    parsed = _to_float(value)
    if parsed is None:
        return ""
    return f"{parsed:.{digits}f}".replace(".", ",")


def _format_xml_datetime(day_iso: str) -> str:
    dt = datetime.strptime(str(day_iso or ""), "%Y-%m-%d")
    return dt.strftime("%d/%m/%Y 00:00:00")


def _parse_xml_datetime_to_iso(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Data inválida no XML 042: {raw}")


def _xml_text(parent, tag_name: str) -> str:
    node = parent.find(tag_name)
    return str(node.text or "").strip() if node is not None else ""


def parse_xml042_import(content: bytes, filename: str, repo=None) -> dict:
    try:
        root = ET.fromstring(content)
    except ET.ParseError as exc:
        raise ValueError(f"XML inválido em {filename}: {exc}") from exc

    test_node = root.find(".//TESTE_POCO")
    if test_node is None:
        raise ValueError(f"XML {filename} não contém TESTE_POCO")

    cod_cadastro_poco = _xml_text(test_node, "COD_CADASTRO_POCO")
    if not cod_cadastro_poco:
        raise ValueError(f"XML {filename} sem COD_CADASTRO_POCO")

    production_day = _parse_xml_datetime_to_iso(_xml_text(test_node, "DHA_TESTE"))
    month_ref = production_day[:7]
    oil_sm3 = _to_float(_xml_text(test_node, "MED_POTENCIAL_OLEO"))
    gas_1000sm3 = _to_float(_xml_text(test_node, "MED_POTENCIAL_GAS"))
    water_sm3 = _to_float(_xml_text(test_node, "MED_POTENCIAL_AGUA"))

    enriched = {}
    if repo is not None:
        enriched = repo.get_document_by_key(production_day, cod_cadastro_poco) or {}
        if not enriched:
            enriched = repo.get_catalog_by_code(cod_cadastro_poco) or {}

    return {
        "month_ref": month_ref,
        "production_day": production_day,
        "cod_cadastro_poco": cod_cadastro_poco,
        "well_operator_name": str(
            enriched.get("well_operator_name")
            or enriched.get("well_anp_name")
            or ""
        ).strip(),
        "subsea_tag": str(enriched.get("subsea_tag") or "").strip(),
        "bank": str(enriched.get("bank") or "").strip().upper(),
        "ind_tipo_teste": _xml_text(test_node, "IND_TIPO_TESTE"),
        "dha_teste": _xml_text(test_node, "DHA_TESTE"),
        "dha_aplicacao": _xml_text(test_node, "DHA_APLICACAO"),
        "ind_valido": _xml_text(test_node, "IND_VALIDO"),
        "oil_sm3": oil_sm3,
        "gas_1000sm3": gas_1000sm3,
        "water_sm3": water_sm3,
        "file_hash": hashlib.sha256(content).hexdigest(),
        "payload_json": {
            "root_tag": root.tag,
            "filename": filename,
            "oil_sm3": oil_sm3,
            "gas_1000sm3": gas_1000sm3,
            "water_sm3": water_sm3,
        },
    }


def summarize_imported_xml042(rows: list[dict], files: list[dict]) -> dict:
    codes = {str(row.get("cod_cadastro_poco") or "").strip() for row in rows if row.get("cod_cadastro_poco")}
    return {
        "rows": len(rows),
        "codes": len(codes),
        "files": len(files),
        "latest_day": max((row.get("production_day") or "" for row in rows), default=""),
    }


def build_xml042_import_workbook(rows: list[dict]) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XML042_IMPORTADOS"
    headers = [
        "Mês",
        "Data de produção",
        "Código do poço",
        "Poço",
        "TAG subsea",
        "Banco",
        "Tipo de teste",
        "Válido",
        "Óleo (Sm³)",
        "Gás (mil Sm³)",
        "Água (Sm³)",
        "Arquivo",
        "Importado em",
    ]
    thin = Side(style="thin", color="D7E0EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="1F4E78")
    alt = PatternFill("solid", fgColor="F7FAFD")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        values = [
            row.get("month_ref") or "",
            row.get("production_day") or "",
            row.get("cod_cadastro_poco") or "",
            row.get("well_operator_name") or "",
            row.get("subsea_tag") or "",
            row.get("bank") or "",
            row.get("ind_tipo_teste") or "",
            row.get("ind_valido") or "",
            row.get("oil_sm3"),
            row.get("gas_1000sm3"),
            row.get("water_sm3"),
            row.get("filename") or "",
            str(row.get("imported_at") or "").replace("T", " ")[:16],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt

    widths = [12, 16, 18, 22, 18, 10, 15, 10, 14, 16, 14, 34, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_xml042_seed_rows(cadastro: dict) -> list[dict]:
    fixed_codes = {
        ("PE_2", "18FT0506"): {
            "cod_cadastro_poco": "86316029925",
            "well_anp_name": "7-BAC-1-SPS",
        },
        ("PW-104DA", "18FT1106"): {
            "cod_cadastro_poco": "86316030246",
            "well_anp_name": "7-BAC-4D-SPS",
        },
        ("PW_104DA", "18FT1106"): {
            "cod_cadastro_poco": "86316030246",
            "well_anp_name": "7-BAC-4D-SPS",
        },
        ("PE_4", "18FT1506"): {
            "cod_cadastro_poco": "86316030256",
            "well_anp_name": "7-BAC-5A-SPS",
        },
    }
    rows = []
    for entry in cadastro.get("banks_subsea", []):
        operator_name = str(entry.get("sistema") or "").strip()
        subsea_tag = str(entry.get("tag_associado") or "").strip()
        if not operator_name or not subsea_tag:
            continue
        fixed = fixed_codes.get((operator_name, subsea_tag))
        if not fixed:
            continue
        rows.append(
            {
                "well_operator_name": operator_name,
                "well_anp_name": fixed["well_anp_name"],
                "cod_cadastro_poco": fixed["cod_cadastro_poco"],
                "subsea_tag": subsea_tag,
                "cod_campo": "4735",
                "campo": "BACALHAU",
                "cod_instalacao": "38480",
                "instalacao": "FPSO BACALHAU",
                "enabled_042": True,
                "active": True,
                "notes": f"Seed inicial a partir do cadastro local ({entry.get('bank_code','')}).",
            }
        )
    return rows


def _candidate_key(row: dict) -> tuple[str, str, str, str]:
    return (
        str(row.get("production_day") or ""),
        str(row.get("bank") or "").strip().upper(),
        str(row.get("well_operator_name") or "").strip(),
        str(row.get("subsea_tag") or "").strip(),
    )


def build_xml042_preview(candidate: dict) -> str:
    return (
        '<?xml version="1.0" encoding="ISO-8859-1"?>\n'
        '<a042 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="042.xsd">\n'
        "  <LISTA_TESTE_POCO>\n"
        "    <TESTE_POCO>\n"
        f"      <COD_CADASTRO_POCO>{candidate['catalog']['cod_cadastro_poco']}</COD_CADASTRO_POCO>\n"
        "      <IND_TIPO_TESTE>M</IND_TIPO_TESTE>\n"
        f"      <DHA_TESTE>{_format_xml_datetime(candidate['production_day'])}</DHA_TESTE>\n"
        f"      <DHA_APLICACAO>{_format_xml_datetime(candidate['production_day'])}</DHA_APLICACAO>\n"
        "      <IND_VALIDO>S</IND_VALIDO>\n"
        f"      <MED_POTENCIAL_OLEO>{_format_decimal_ptbr(candidate['oil_sm3'])}</MED_POTENCIAL_OLEO>\n"
        f"      <MED_POTENCIAL_GAS>{_format_decimal_ptbr(candidate['gas_1000sm3'])}</MED_POTENCIAL_GAS>\n"
        f"      <MED_POTENCIAL_AGUA>{_format_decimal_ptbr(candidate['water_sm3'])}</MED_POTENCIAL_AGUA>\n"
        "      <LISTA_SEPARADOR/>\n"
        "    </TESTE_POCO>\n"
        "  </LISTA_TESTE_POCO>\n"
        "</a042>\n"
    )


def _safe_filename_part(value: str) -> str:
    return "".join(ch for ch in str(value or "").strip() if ch.isalnum() or ch in ("-", "_")) or "sem_ref"


def _build_anp_xml042_filename(cnpj8: str, folder: Path, target_dir: Optional[Path] = None) -> str:
    """Build an ANP-compliant XML042 filename: 042_<CNPJ8>_<AAAAMMDDHHmmSS>.xml.

    The ANP naming convention does not include the well code. When multiple files
    are generated in the same second, advance the timestamp by one second until
    the filename is free in both the app output folder and the optional SGM target
    folder. This preserves the accepted format without overwriting files.
    """
    current = datetime.now().replace(microsecond=0)
    for offset in range(86400):
        stamp = (current + timedelta(seconds=offset)).strftime("%Y%m%d%H%M%S")
        filename = f"042_{cnpj8}_{stamp}.xml"
        if (folder / filename).exists():
            continue
        if target_dir is not None and (target_dir / filename).exists():
            continue
        return filename
    raise RuntimeError("Não foi possível criar nome XML042 único no padrão ANP.")


def list_xml042_candidates(
    repo,
    *,
    month: str,
    production_day: str = "",
    bank: str = "",
    status: str = "",
    normalize_tag_name,
):
    date_from, date_to = _parse_month_bounds(month)
    if not date_from:
        return {
            "month": "",
            "rows": [],
            "summary": {},
            "days": [],
            "banks": [],
            "statuses": ["elegivel", "gerado", "pendente", "sem cadastro", "cadastro ambiguo", "desabilitado", "valor critico ausente"],
        }
    measurements = repo.list_daily_subsea_candidates(date_from, date_to, bank=bank, production_day=production_day)
    curated_map = {
        _candidate_key(row): row
        for row in repo.list_curated_rows(date_from, date_to, production_day=production_day)
    }
    catalog_rows = repo.list_catalog(active_only=False)
    catalog_map: dict[tuple[str, str], list[dict]] = {}
    for row in catalog_rows:
        key = (
            _norm_tag(normalize_tag_name, row.get("well_operator_name")),
            _norm_tag(normalize_tag_name, row.get("subsea_tag")),
        )
        catalog_map.setdefault(key, []).append(row)

    documents_list = repo.list_documents(month)
    documents_map = {
        (doc["production_day"], str(doc.get("cod_cadastro_poco") or "").strip()): doc
        for doc in documents_list
    }

    rows = []
    for item in measurements:
        oil_sm3 = _to_float(item.get("oil_sm3"))
        gas_sm3 = _to_float(item.get("gas_sm3"))
        water_sm3 = _to_float(item.get("water_sm3"))
        match_key = (
            _norm_tag(normalize_tag_name, item.get("well_operator_name")),
            _norm_tag(normalize_tag_name, item.get("subsea_tag")),
        )
        matches = catalog_map.get(match_key, [])
        qa_flags = []
        match = None
        if not matches:
            match_status = "sem cadastro"
            qa_flags.append("catalog_missing")
        elif len(matches) > 1:
            match_status = "cadastro ambiguo"
            qa_flags.append("catalog_ambiguous")
        else:
            match = matches[0]
            if not match.get("active", 1) or not match.get("enabled_042", 1):
                match_status = "desabilitado"
                qa_flags.append("catalog_disabled")
            else:
                match_status = "elegivel"
        if oil_sm3 is None or gas_sm3 is None or water_sm3 is None:
            match_status = "valor critico ausente"
            qa_flags.append("critical_blank")
            
        cod_poco = str(match.get("cod_cadastro_poco") or "").strip() if match else ""
        doc = documents_map.get((item["production_day"], cod_poco)) if cod_poco else None

        candidate = {
            "production_day": item["production_day"],
            "bank": item["bank"],
            "well_operator_name": item["well_operator_name"],
            "subsea_tag": item["subsea_tag"],
            "loop": item.get("loop") or "",
            "hours_available": int(item.get("hours_available") or 0),
            "oil_sm3": oil_sm3,
            "gas_sm3": gas_sm3,
            "gas_1000sm3": round(gas_sm3 / 1000.0, 5) if gas_sm3 is not None else None,
            "water_sm3": water_sm3,
            "oil_t": _to_float(item.get("oil_t")),
            "gas_t": _to_float(item.get("gas_t")),
            "water_t": _to_float(item.get("water_t")),
            "catalog_match_status": match_status,
            "qa_flags": qa_flags,
            "catalog": {
                "id": match.get("id") if match else None,
                "well_anp_name": match.get("well_anp_name") if match else "",
                "cod_cadastro_poco": match.get("cod_cadastro_poco") if match else "",
                "cod_campo": match.get("cod_campo") if match else "",
                "campo": match.get("campo") if match else "",
                "cod_instalacao": match.get("cod_instalacao") if match else "",
                "instalacao": match.get("instalacao") if match else "",
                "enabled_042": bool(match.get("enabled_042")) if match else False,
                "active": bool(match.get("active")) if match else False,
            },
            "eligible": match_status == "elegivel",
            "generated": bool(doc),
            "generated_at": doc.get("generated_at", "") if doc else "",
            "generated_filename": doc.get("filename", "") if doc else "",
            "document_id": doc.get("id") if doc else None,
            "file_path": doc.get("file_path", "") if doc else "",
            "source_daily_row_ref": f'{item["production_day"]}|{item["bank"]}|{item["well_operator_name"]}|{item["subsea_tag"]}',
        }
        curated = curated_map.get(_candidate_key(candidate)) or {}
        candidate["approved"] = bool(curated.get("approved_at"))
        candidate["approved_at"] = curated.get("approved_at", "")
        candidate["approved_by_user"] = curated.get("approved_by_user", "")
        candidate["preview_xml"] = build_xml042_preview(candidate) if candidate["eligible"] else ""
        rows.append(candidate)

    if status:
        if status == "gerado":
            rows = [row for row in rows if row.get("generated")]
        elif status == "pendente":
            rows = [row for row in rows if row["eligible"] and not row.get("generated")]
        else:
            rows = [row for row in rows if row["catalog_match_status"] == status]

    rows.sort(key=lambda row: (row["production_day"], row["bank"], row["well_operator_name"]), reverse=True)
    return {
        "month": month,
        "rows": rows,
        "days": sorted({row["production_day"] for row in rows}, reverse=True),
        "banks": sorted({row["bank"] for row in rows}),
        "statuses": ["elegivel", "gerado", "pendente", "sem cadastro", "cadastro ambiguo", "desabilitado", "valor critico ausente"],
        "summary": {
            "rows": len(rows),
            "eligible": sum(1 for row in rows if row["eligible"]),
            "approved": sum(1 for row in rows if row["approved"]),
            "generated": sum(1 for row in rows if row.get("generated")),
            "pending_eligible": sum(1 for row in rows if row["eligible"] and not row.get("generated")),
            "not_eligible": sum(1 for row in rows if not row["eligible"]),
            "default_target_dir": DEFAULT_XML042_DEST_DIR,
        },
    }


def generate_xml042_document(
    repo,
    *,
    candidate: dict,
    output_dir: Path,
    cnpj8: str = DEFAULT_XML042_CNPJ8,
    author: str = DEFAULT_AUTHOR,
    target_dir: Optional[Path] = None,
):
    if not candidate.get("eligible"):
        raise ValueError("Linha não elegível para emissão.")
    if not candidate.get("approved"):
        raise ValueError("Linha ainda não aprovada.")
    xml_text = build_xml042_preview(candidate)
    generated_at = datetime.now().isoformat(timespec="seconds")
    folder = output_dir / "xml042"
    folder.mkdir(parents=True, exist_ok=True)
    dest_path = target_dir if target_dir is not None else Path(DEFAULT_XML042_DEST_DIR)
    filename = _build_anp_xml042_filename(cnpj8, folder, dest_path)
    file_path = folder / filename
    data = xml_text.encode("iso-8859-1", errors="xmlcharrefreplace")
    file_path.write_bytes(data)
    file_hash = hashlib.sha256(data).hexdigest()

    saved_to_target_dir = False
    target_file_path = ""
    try:
        if dest_path and str(dest_path).strip():
            dest_path.mkdir(parents=True, exist_ok=True)
            target_file = dest_path / filename
            target_file.write_bytes(data)
            saved_to_target_dir = True
            target_file_path = str(target_file)
    except Exception as exc:
        print(f"[WARNING] Não foi possível salvar cópia do XML em {dest_path}: {exc}")

    payload = {
        "production_day": candidate["production_day"],
        "cod_cadastro_poco": candidate["catalog"]["cod_cadastro_poco"],
        "well_operator_name": candidate["well_operator_name"],
        "subsea_tag": candidate["subsea_tag"],
        "bank": candidate["bank"],
        "filename": filename,
        "file_path": str(file_path),
        "file_hash": file_hash,
        "status": "generated",
        "generated_at": generated_at,
        "generated_by": author,
        "payload_json": {
            "oil_sm3": candidate["oil_sm3"],
            "gas_sm3": candidate["gas_sm3"],
            "gas_1000sm3": candidate["gas_1000sm3"],
            "water_sm3": candidate["water_sm3"],
            "catalog": candidate["catalog"],
            "xml_preview": xml_text,
            "target_file_path": target_file_path,
            "saved_to_target_dir": saved_to_target_dir,
        },
    }
    previous = repo.get_document_by_key(payload["production_day"], payload["cod_cadastro_poco"])
    document_id = repo.save_document(payload)
    if previous:
        previous_path = Path(previous.get("file_path") or "")
        if previous_path and previous_path.exists() and previous_path != file_path:
            try:
                previous_path.unlink()
            except Exception:
                pass
    return {
        "id": document_id,
        "filename": filename,
        "file_path": str(file_path),
        "target_file_path": target_file_path,
        "saved_to_target_dir": saved_to_target_dir,
        "download_url": f"/api/xml042/download/{document_id}",
        "xml_preview": xml_text,
        "generated_at": generated_at,
    }
