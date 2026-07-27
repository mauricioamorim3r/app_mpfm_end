from __future__ import annotations

import json
import hashlib
import re
import sqlite3
import time
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import Any

from .daily_checklist_service import DailyChecklistService


class PainelOperadorStagingService:
    """Adapter for the copied Painel do Operador module and local audit decisions."""

    DEFAULT_MAX_LIST_ITEMS = 200
    INDEX_EXCLUDED_DIRS = {
        ".git",
        "node_modules",
        "dist",
        "release",
        ".playwright-mcp",
        "__pycache__",
        ".pytest_cache",
    }
    INDEX_LOW_VALUE_EXTENSIONS = {".pyc", ".vsix"}
    ANP_EXPORT_FILES = {
        "Óleo Linear.xlsx": {"family": "a001", "export_name": "Óleo Linear", "record_kind": "linear_oil"},
        "Gás Linear.xlsx": {"family": "a002", "export_name": "Gás Linear", "record_kind": "linear_gas"},
        "Gás Diferencial.xlsx": {"family": "a003", "export_name": "Gás Diferencial", "record_kind": "differential_gas"},
        "Falha de Medição.xlsx": {"family": "a039", "export_name": "Falha de Medição", "record_kind": "measurement_failure"},
        "BSW em Linha.xlsx": {"family": "a040", "export_name": "BSW em Linha", "record_kind": "inline_bsw"},
    }
    TAG_RE = re.compile(r"\b(?:\d{2}[A-Z]{2}\d{4}[A-Z]?|[A-Z]{2}_?\d{1,3}[A-Z]?|PW_?\d{2,3}[A-Z]?|PE_?\d{1,3})\b", re.IGNORECASE)
    ISO_DATE_RE = re.compile(r"\b(20\d{2})[-_\.](\d{2})[-_\.](\d{2})\b")
    COMPACT_DATE_RE = re.compile(r"\b(20\d{2})(\d{2})(\d{2})(?:\d{6})?\b")
    BR_DATE_RE = re.compile(r"\b(\d{2})[-_\.](\d{2})[-_\.](20\d{2})\b")
    STAGING_VIEWS = {
        "sources": {
            "table": "painel_operador_sources",
            "date_column": "source_date",
            "search_columns": ("source_name", "source_path", "local_path", "family", "family_name", "source_kind"),
            "filters": {"kind": "source_kind", "family": "family", "file_exists": "file_exists"},
            "order": "source_date DESC, id DESC",
        },
        "points": {
            "table": "painel_operador_measurement_points",
            "date_column": "point_date",
            "search_columns": ("tag", "family", "family_name", "fluid", "principal", "secondary", "meter_type", "computador_vazao"),
            "filters": {"family": "family", "tag": "tag", "fluid": "fluid", "meter_type": "meter_type", "active": "active_status"},
            "order": "point_date DESC, family, tag",
        },
        "comparisons": {
            "table": "painel_operador_comparisons",
            "date_column": "comparison_date",
            "search_columns": ("tag", "family", "family_name", "fluid", "status", "note", "raw_source", "xml_source"),
            "filters": {"family": "family", "tag": "tag", "fluid": "fluid", "status": "status"},
            "order": "comparison_date DESC, family, tag",
        },
        "evidence": {
            "table": "painel_operador_evidence",
            "date_column": "event_at",
            "search_columns": ("title", "requirement_id", "status", "evidence_state", "source_path", "target_id", "target_type"),
            "filters": {"kind": "evidence_kind", "status": "status", "requirement_id": "requirement_id", "target_id": "target_id"},
            "order": "event_at DESC, id DESC",
        },
        "alerts": {
            "table": "painel_operador_alerts",
            "date_column": "alert_date",
            "search_columns": ("title", "detail", "area", "target_id", "status", "severity", "source_path"),
            "filters": {"kind": "alert_kind", "severity": "severity", "status": "status", "area": "area", "target_id": "target_id"},
            "order": "alert_date DESC, severity DESC, id DESC",
        },
        "proposals": {
            "table": "painel_operador_proposals",
            "date_column": "created_at_source",
            "search_columns": (
                "proposal_id",
                "proposal_kind",
                "domain",
                "title",
                "target_id",
                "field_name",
                "confidence",
                "risk",
                "status",
                "source_type",
                "source_name",
                "evidence_state",
                "evidence_text",
                "recommended_action",
            ),
            "filters": {
                "kind": "proposal_kind",
                "status": "status",
                "severity": "risk",
                "area": "domain",
                "target_id": "target_id",
                "evidence_state": "evidence_state",
                "confidence": "confidence",
                "source_type": "source_type",
            },
            "order": "created_at_source DESC, risk ASC, id DESC",
        },
        "calendar": {
            "table": "painel_operador_calendar_days",
            "date_column": "calendar_date",
            "search_columns": ("calendar_date", "status", "closing_status", "xml_families_json", "missing_xml_families_json"),
            "filters": {"status": "status", "loaded": "loaded"},
            "order": "calendar_date DESC",
        },
        "pendencies": {
            "table": "painel_operador_calendar_pendencies",
            "date_column": "calendar_date",
            "search_columns": ("pendency_id", "pendency_type", "severity", "status", "title", "detail", "recommended_action", "resolution_mode"),
            "filters": {"kind": "pendency_type", "severity": "severity", "status": "status", "target_id": "pendency_id"},
            "order": "calendar_date DESC, severity DESC, id DESC",
        },
    }
    FILE_INDEX_VIEW = {
        "table": "painel_operador_file_index",
        "date_column": "inferred_date",
        "search_columns": (
            "relative_path",
            "filename",
            "category",
            "document_kind",
            "source_group",
            "inferred_tag",
            "inferred_family",
            "ignore_reason",
        ),
        "filters": {
            "category": "category",
            "document_kind": "document_kind",
            "source_group": "source_group",
            "extension": "extension",
            "tag": "inferred_tag",
            "family": "inferred_family",
            "ignored": "ignored",
            "is_duplicate": "is_duplicate",
            "parse_priority": "parse_priority",
        },
        "order": "ignored ASC, parse_priority DESC, inferred_date DESC, relative_path ASC",
    }
    ANP_EXPORT_VIEW = {
        "table": "painel_operador_anp_export_rows",
        "date_column": "reference_date",
        "search_columns": (
            "source_file",
            "export_name",
            "record_kind",
            "installation",
            "installation_code",
            "tag",
            "element_tag",
            "serial_number",
            "failure_code",
            "failure_type",
            "notification_type",
            "received_file",
        ),
        "filters": {
            "family": "family",
            "tag": "tag",
            "record_kind": "record_kind",
            "source_file": "source_file",
            "failure_type": "failure_type",
            "notification_type": "notification_type",
        },
        "order": "reference_date DESC, family ASC, tag ASC, row_number ASC",
    }

    def __init__(self, module_root: Path):
        self.module_root = Path(module_root)
        self.dashboard_data_path = self.module_root / "dashboard-anp-radar" / "src" / "data" / "dashboard-data.json"
        self.sqlite_path = self.module_root / "dashboard-anp-radar" / "data" / "radar-anp.sqlite"
        self.config_path = self.module_root / "dashboard-anp-radar" / "config" / "data-sources.json"
        self.daily_checklist = DailyChecklistService()
        # Cache em memória para dashboard-data.json (TTL 60s para evitar releituras desnecessárias)
        self._dashboard_cache: dict[str, Any] | None = None
        self._dashboard_cache_mtime: float = 0.0
        self._dashboard_cache_ts: float = 0.0

    def ihm_reports(
        self,
        *,
        date_from: str = "",
        date_to: str = "",
        fluid: str = "",
        tag: str = "",
        limit: int = 500,
    ) -> dict[str, Any]:
        """Retorna dados dos IHM Daily Reports (Oil/Gas/Water) do dashboard-data.json."""
        data = self._load_dashboard_data()
        ihm = data.get("ihmReports") or {}
        rows: list[dict[str, Any]] = list(ihm.get("rows") or [])

        if date_from:
            rows = [r for r in rows if str(r.get("production_date", "")) >= date_from]
        if date_to:
            rows = [r for r in rows if str(r.get("production_date", "")) <= date_to]
        if fluid:
            rows = [r for r in rows if str(r.get("fluid", "")).lower() == fluid.lower()]
        if tag:
            rows = [r for r in rows if tag.upper() in str(r.get("tag", "")).upper()]

        rows = rows[:limit]
        days: list[dict[str, Any]] = list(ihm.get("days") or [])
        if date_from:
            days = [d for d in days if str(d.get("date", "")) >= date_from]
        if date_to:
            days = [d for d in days if str(d.get("date", "")) <= date_to]

        return {
            "record_type": "ihm-reports",
            "filters": {"date_from": date_from, "date_to": date_to, "fluid": fluid, "tag": tag},
            "summary": ihm.get("summary") or {},
            "days": days,
            "rows": rows,
            "total": len(rows),
        }

    def gas_balance_ihm(self, *, date_from: str = "", date_to: str = "") -> dict[str, Any]:
        """Retorna dados do GasBalance IHM do dashboard-data.json."""
        data = self._load_dashboard_data()
        gb = data.get("gasBalance") or {}
        rows: list[dict[str, Any]] = list(gb.get("rows") or [])

        if date_from:
            rows = [r for r in rows if str(r.get("production_date", "")) >= date_from]
        if date_to:
            rows = [r for r in rows if str(r.get("production_date", "")) <= date_to]

        return {
            "record_type": "gas-balance-ihm",
            "summary": gb.get("summary") or {},
            "rows": rows,
            "total": len(rows),
        }

    def status(self) -> dict[str, Any]:
        dashboard_info = self._file_info(self.dashboard_data_path)
        sqlite_info = self._file_info(self.sqlite_path)
        config_info = self._file_info(self.config_path)
        return {
            "ok": dashboard_info["exists"] and config_info["exists"],
            "module_root": str(self.module_root),
            "dashboard_data": dashboard_info,
            "sqlite": sqlite_info,
            "config": config_info,
            "database": self.database_summary(),
        }

    def contract(self) -> dict[str, Any]:
        data = self._load_dashboard_data()
        blocks = []
        for key, value in data.items():
            info: dict[str, Any] = {"key": key, "type": type(value).__name__}
            if isinstance(value, list):
                info["count"] = len(value)
                info["sample_keys"] = self._sample_keys(value)
            elif isinstance(value, dict):
                info["count"] = len(value)
                info["keys"] = list(value.keys())[:50]
            else:
                info["value"] = value
            blocks.append(info)
        return {
            "module_root": str(self.module_root),
            "dashboard_data_path": str(self.dashboard_data_path),
            "generated_at": data.get("meta", {}).get("generatedAt") if isinstance(data.get("meta"), dict) else "",
            "blocks": blocks,
        }

    def data(self, block_names: list[str] | None = None, max_list_items: int | None = None) -> dict[str, Any]:
        data = self._load_dashboard_data()
        selected_names = [name for name in (block_names or []) if name]
        if not selected_names:
            selected_names = [
                "meta",
                "kpis",
                "families",
                "files",
                "comparisons",
                "operatorPanelHealth",
                "regulatoryMatrix",
                "eventEvidenceRadar",
                "changeProposals",
                "operationalCalendar",
                "bsw",
                "failures",
                "mpfm",
                "alerts",
                "database",
            ]
        limit = self._normalize_limit(max_list_items)
        payload = {}
        truncated = {}
        for name in selected_names:
            if name not in data:
                continue
            payload[name], truncated[name] = self._bounded_value(data[name], limit)
        return {
            "module_root": str(self.module_root),
            "source": str(self.dashboard_data_path),
            "selected_blocks": list(payload.keys()),
            "max_list_items": limit,
            "truncated": truncated,
            "data": payload,
        }

    def data_sources(self, *, validate: bool = False) -> dict[str, Any]:
        config = self._load_data_sources_config()
        sources = []
        validation_cache: dict[tuple[str, bool], dict[str, Any]] = {}
        for source in config.get("sources") or []:
            if not isinstance(source, dict):
                continue
            normalized = self._normalize_data_source(source)
            if validate:
                normalized["validation"] = self._validate_source_paths(normalized, validation_cache)
            sources.append(normalized)
        return {
            "schemaVersion": config.get("schemaVersion", 1),
            "workspaceRoot": str(config.get("workspaceRoot") or self.module_root),
            "configPath": str(self.config_path),
            "sources": sources,
        }

    def save_data_source(self, source_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        source_id = str(source_id or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9_.-]+", source_id):
            raise ValueError("ID de fonte inválido.")
        config = self._load_data_sources_config()
        sources = [item for item in (config.get("sources") or []) if isinstance(item, dict)]
        source_index = next((idx for idx, item in enumerate(sources) if str(item.get("id") or "") == source_id), None)
        if source_index is None:
            sources.append({"id": source_id})
            source_index = len(sources) - 1

        current = dict(sources[source_index])
        for key in ("label", "description", "kind"):
            if key in payload:
                current[key] = str(payload.get(key) or "").strip()
        if "recursive" in payload:
            current["recursive"] = bool(payload.get("recursive"))
        if "paths" in payload:
            raw_paths = payload.get("paths") or []
            if isinstance(raw_paths, str):
                raw_paths = [line.strip() for line in raw_paths.splitlines()]
            paths = [str(path).strip() for path in raw_paths if str(path or "").strip()]
            current["paths"] = paths
        current.setdefault("label", source_id)
        current.setdefault("description", "")
        current.setdefault("kind", "folder")
        current.setdefault("recursive", True)
        current.setdefault("paths", [])
        sources[source_index] = current
        config["sources"] = sources
        config.setdefault("schemaVersion", 1)
        config.setdefault("workspaceRoot", str(self.module_root))
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        self.config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        normalized = self._normalize_data_source(current)
        normalized["validation"] = self._validate_source_paths(normalized, {})
        return {"ok": True, "source": normalized, "configPath": str(self.config_path)}

    def validate_data_sources(self) -> dict[str, Any]:
        return self.data_sources(validate=True)

    def database_summary(self) -> dict[str, Any]:
        if not self.sqlite_path.exists():
            return {"exists": False, "tables": []}
        try:
            conn = sqlite3.connect(str(self.sqlite_path))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()
            tables = []
            for row in rows:
                table = row["name"]
                try:
                    count = cur.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
                except sqlite3.Error:
                    count = None
                tables.append({"name": table, "rows": count})
            conn.close()
            return {"exists": True, "path": str(self.sqlite_path), "tables": tables}
        except sqlite3.Error as exc:
            return {"exists": True, "path": str(self.sqlite_path), "error": str(exc), "tables": []}

    def staging_summary(self, db_conn_fn) -> dict[str, Any]:
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            latest = cur.execute(
                """
                SELECT id, started_at, finished_at, source_data_hash, status, counts_json
                FROM painel_operador_sync_runs
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()
            tables = {}
            for table in self._staging_tables():
                tables[table] = cur.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
            return {
                "latest_sync": dict(latest) if latest else None,
                "tables": tables,
            }
        finally:
            conn.close()

    def file_index_summary(self, db_conn_fn) -> dict[str, Any]:
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            latest = cur.execute(
                """
                SELECT id, started_at, finished_at, module_root, status, total_files,
                       indexed_files, ignored_files, duplicate_files, total_size_bytes,
                       counts_json, notes
                FROM painel_operador_file_index_runs
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()
            if not latest:
                return {"latest_run": None, "tables": {"painel_operador_file_index": 0}, "categories": []}
            categories = [
                dict(row)
                for row in cur.execute(
                    """
                    SELECT category, document_kind, COUNT(*) AS count,
                           COALESCE(SUM(file_size_bytes), 0) AS size_bytes
                    FROM painel_operador_file_index
                    WHERE ignored=0
                    GROUP BY category, document_kind
                    ORDER BY count DESC, category
                    """
                ).fetchall()
            ]
            return {
                "latest_run": dict(latest),
                "tables": {
                    "painel_operador_file_index": cur.execute("SELECT COUNT(*) FROM painel_operador_file_index").fetchone()[0]
                },
                "categories": categories,
            }
        finally:
            conn.close()

    def scan_file_index(self, db_conn_fn, *, hash_files: bool = True) -> dict[str, Any]:
        started_at = self._now()
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                INSERT INTO painel_operador_file_index_runs(
                    started_at, finished_at, module_root, status, notes
                )
                VALUES(?,?,?,?,?)
                """,
                (started_at, started_at, str(self.module_root), "running", "Varredura iniciada."),
            )
            index_run_id = int(cur.lastrowid)
            cur.execute("DELETE FROM painel_operador_file_index")

            rows = []
            ignored_dirs = 0
            total_size = 0
            category_counts: dict[str, int] = {}
            extension_counts: dict[str, int] = {}
            for path in self._iter_index_files():
                relative_path = str(path.relative_to(self.module_root))
                parts = set(path.relative_to(self.module_root).parts)
                if parts & self.INDEX_EXCLUDED_DIRS:
                    ignored_dirs += 1
                    continue
                try:
                    stat = path.stat()
                except OSError:
                    continue
                total_size += stat.st_size
                classification = self._classify_index_file(path)
                extension_counts[classification["extension"] or "[none]"] = extension_counts.get(classification["extension"] or "[none]", 0) + 1
                category_counts[classification["category"]] = category_counts.get(classification["category"], 0) + 1
                file_hash = ""
                if hash_files and not classification["ignored"]:
                    file_hash = self._sha1_file(path)
                elif not classification["ignored"]:
                    file_hash = self._payload_hash({"relative_path": relative_path, "size": stat.st_size, "modified": stat.st_mtime})
                duplicate_key = file_hash if hash_files else self._stable_key(path.name.lower(), stat.st_size)
                rows.append(
                    {
                        **classification,
                        "index_run_id": index_run_id,
                        "stable_key": self._stable_key("file-index", relative_path),
                        "relative_path": relative_path,
                        "full_path": str(path),
                        "filename": path.name,
                        "file_size_bytes": stat.st_size,
                        "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
                        "file_hash": file_hash,
                        "duplicate_key": duplicate_key,
                        "payload_json": self._json(
                            {
                                "relative_path": relative_path,
                                "parts": list(path.relative_to(self.module_root).parts),
                                "hash_mode": "sha1" if file_hash and hash_files else "metadata",
                            }
                        ),
                    }
                )

            duplicate_counts: dict[str, int] = {}
            for row in rows:
                if row["ignored"]:
                    continue
                duplicate_counts[row["duplicate_key"]] = duplicate_counts.get(row["duplicate_key"], 0) + 1
            duplicate_files = 0
            for row in rows:
                is_dup = (not row["ignored"]) and duplicate_counts.get(row["duplicate_key"], 0) > 1
                row["is_duplicate"] = 1 if is_dup else 0
                if is_dup:
                    duplicate_files += 1
                self._insert_file_index_row(cur, row)

            indexed_files = sum(1 for row in rows if not row["ignored"])
            ignored_files = sum(1 for row in rows if row["ignored"])
            counts = {
                "categories": category_counts,
                "extensions": extension_counts,
                "ignored_dirs": ignored_dirs,
            }
            finished_at = self._now()
            cur.execute(
                """
                UPDATE painel_operador_file_index_runs
                SET finished_at=?, status='ok', total_files=?, indexed_files=?,
                    ignored_files=?, duplicate_files=?, total_size_bytes=?,
                    counts_json=?, notes=?
                WHERE id=?
                """,
                (
                    finished_at,
                    len(rows),
                    indexed_files,
                    ignored_files,
                    duplicate_files,
                    total_size,
                    self._json(counts),
                    "Indice de fontes atualizado.",
                    index_run_id,
                ),
            )
            conn.commit()
            return {
                "ok": True,
                "index_run_id": index_run_id,
                "started_at": started_at,
                "finished_at": finished_at,
                "total_files": len(rows),
                "indexed_files": indexed_files,
                "ignored_files": ignored_files,
                "duplicate_files": duplicate_files,
                "total_size_bytes": total_size,
                "counts": counts,
            }
        except Exception as exc:
            conn.rollback()
            raise exc
        finally:
            conn.close()

    def list_file_index(
        self,
        db_conn_fn,
        *,
        q: str = "",
        date_from: str = "",
        date_to: str = "",
        filters: dict[str, Any] | None = None,
        limit: int = 100,
        offset: int = 0,
        include_payload: bool = False,
    ) -> dict[str, Any]:
        return self._list_table_records(
            db_conn_fn,
            self.FILE_INDEX_VIEW,
            "file-index",
            q=q,
            date_from=date_from,
            date_to=date_to,
            filters=filters,
            limit=limit,
            offset=offset,
            include_payload=include_payload,
        )

    def xml_validation(
        self,
        db_conn_fn,
        *,
        q: str = "",
        date_from: str = "",
        date_to: str = "",
        family: str = "",
        tag: str = "",
        kind: str = "",
        status: str = "",
        limit: int = 100,
        offset: int = 0,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        offset = max(0, self._int_or_zero(offset))
        filters = {
            "q": str(q or "").strip(),
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
            "kind": str(kind or "").strip(),
            "status": str(status or "").strip(),
        }
        where = ["ignored=0", "(LOWER(extension)='.xml' OR LOWER(document_kind) LIKE '%xml%')"]
        params: list[Any] = []
        if filters["date_from"]:
            where.append("inferred_date >= ?")
            params.append(filters["date_from"])
        if filters["date_to"]:
            where.append("inferred_date <= ?")
            params.append(filters["date_to"])
        if filters["family"]:
            where.append("inferred_family = ?")
            params.append(filters["family"])
        if filters["tag"]:
            where.append("inferred_tag = ?")
            params.append(filters["tag"])
        if filters["kind"]:
            where.append("document_kind = ?")
            params.append(filters["kind"])
        if filters["q"]:
            like = f"%{filters['q']}%"
            where.append(
                """(
                    COALESCE(relative_path, '') LIKE ?
                    OR COALESCE(filename, '') LIKE ?
                    OR COALESCE(document_kind, '') LIKE ?
                    OR COALESCE(source_group, '') LIKE ?
                    OR COALESCE(inferred_family, '') LIKE ?
                    OR COALESCE(inferred_tag, '') LIKE ?
                )"""
            )
            params.extend([like] * 6)

        where_sql = " WHERE " + " AND ".join(where)
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            if not self._table_exists(cur, "painel_operador_file_index"):
                return {
                    "record_type": "xml-validation",
                    "total": 0,
                    "returned": 0,
                    "limit": limit,
                    "offset": offset,
                    "filters": {key: value for key, value in filters.items() if value},
                    "summary": {"total": 0, "ok": 0, "warning": 0, "critical": 0, "by_kind": []},
                    "items": [],
                }
            total_candidates = cur.execute(
                f"SELECT COUNT(*) FROM painel_operador_file_index {where_sql}",
                params,
            ).fetchone()[0]
            if filters["status"]:
                candidate_limit = min(max(offset + (limit * 4), 500), 1000)
                rows = cur.execute(
                    f"""
                    SELECT *
                    FROM painel_operador_file_index
                    {where_sql}
                    ORDER BY inferred_date DESC, document_kind ASC, relative_path ASC
                    LIMIT ?
                    """,
                    params + [candidate_limit],
                ).fetchall()
            else:
                candidate_limit = limit
                rows = cur.execute(
                    f"""
                    SELECT *
                    FROM painel_operador_file_index
                    {where_sql}
                    ORDER BY inferred_date DESC, document_kind ASC, relative_path ASC
                    LIMIT ? OFFSET ?
                    """,
                    params + [limit, offset],
                ).fetchall()
            diagnostics = [self._xml_validation_item(cur, dict(row)) for row in rows]
            if filters["status"]:
                diagnostics = [item for item in diagnostics if item["status"] == filters["status"]]
                total = len(diagnostics)
                page = diagnostics[offset : offset + limit]
            else:
                total = total_candidates
                page = diagnostics
            summary = self._xml_validation_summary(diagnostics)
            summary["cataloged_total"] = total_candidates
            summary["evaluated"] = len(diagnostics)
            summary["status_scope"] = "filtered_batch" if filters["status"] else "current_page"
            summary["candidate_limit"] = candidate_limit
            return {
                "record_type": "xml-validation",
                "total": total,
                "returned": len(page),
                "limit": limit,
                "offset": offset,
                "filters": {key: value for key, value in filters.items() if value},
                "summary": summary,
                "items": page,
            }
        finally:
            conn.close()

    def anp_export_summary(self, db_conn_fn) -> dict[str, Any]:
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            latest = cur.execute(
                """
                SELECT id, started_at, finished_at, module_root, status, total_files,
                       total_rows, counts_json, notes
                FROM painel_operador_anp_export_runs
                ORDER BY id DESC
                LIMIT 1
                """
            ).fetchone()
            if not latest:
                return {"latest_import": None, "tables": {"painel_operador_anp_export_rows": 0}, "groups": []}
            groups = [
                dict(row)
                for row in cur.execute(
                    """
                    SELECT family, record_kind, source_file, COUNT(*) AS count,
                           MIN(NULLIF(reference_date, '')) AS first_date,
                           MAX(NULLIF(reference_date, '')) AS last_date,
                           COUNT(DISTINCT tag) AS tags_count
                    FROM painel_operador_anp_export_rows
                    GROUP BY family, record_kind, source_file
                    ORDER BY family, source_file
                    """
                ).fetchall()
            ]
            return {
                "latest_import": dict(latest),
                "tables": {
                    "painel_operador_anp_export_rows": cur.execute("SELECT COUNT(*) FROM painel_operador_anp_export_rows").fetchone()[0]
                },
                "groups": groups,
            }
        finally:
            conn.close()

    def daily_checklist_summary(self, db_conn_fn) -> dict[str, Any]:
        return self.daily_checklist.summary(db_conn_fn)

    def inspect_daily_checklist(self, source_path: str, *, include_rows: bool = False) -> dict[str, Any]:
        return self.daily_checklist.inspect_workbook(source_path, include_rows=include_rows)

    def import_daily_checklist(self, db_conn_fn, source_path: str) -> dict[str, Any]:
        return self.daily_checklist.import_workbook(db_conn_fn, source_path)

    def list_daily_checklist_rows(
        self,
        db_conn_fn,
        *,
        sheet_name: str = "",
        date_from: str = "",
        date_to: str = "",
        tag: str = "",
        q: str = "",
        limit: int = 120,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.list_rows(
            db_conn_fn,
            sheet_name=sheet_name,
            date_from=date_from,
            date_to=date_to,
            tag=tag,
            q=q,
            limit=limit,
            offset=offset,
        )

    def tank_balance(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        q: str = "",
        status: str = "",
        limit: int = 120,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.tank_balance(
            db_conn_fn,
            date_from=date_from,
            date_to=date_to,
            q=q,
            status=status,
            limit=limit,
            offset=offset,
        )

    def offspec_tank(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        q: str = "",
        status: str = "",
        limit: int = 120,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.offspec_tank(
            db_conn_fn,
            date_from=date_from,
            date_to=date_to,
            q=q,
            status=status,
            limit=limit,
            offset=offset,
        )

    def quality_samples(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        q: str = "",
        status: str = "",
        limit: int = 160,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.quality_samples(
            db_conn_fn,
            date_from=date_from,
            date_to=date_to,
            q=q,
            status=status,
            limit=limit,
            offset=offset,
        )

    def mpfm_fiscal_oil(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        q: str = "",
        status: str = "",
        limit: int = 160,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.mpfm_fiscal_oil(
            db_conn_fn,
            date_from=date_from,
            date_to=date_to,
            q=q,
            status=status,
            limit=limit,
            offset=offset,
        )

    def gas_balance(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        q: str = "",
        status: str = "",
        limit: int = 160,
        offset: int = 0,
    ) -> dict[str, Any]:
        return self.daily_checklist.gas_balance(
            db_conn_fn,
            date_from=date_from,
            date_to=date_to,
            q=q,
            status=status,
            limit=limit,
            offset=offset,
        )

    def import_anp_exports(self, db_conn_fn) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("Dependencia openpyxl indisponivel para ler exports ANP.") from exc

        started_at = self._now()
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                INSERT INTO painel_operador_anp_export_runs(
                    started_at, finished_at, module_root, status, notes
                )
                VALUES(?,?,?,?,?)
                """,
                (started_at, started_at, str(self.module_root), "running", "Importacao ANP iniciada."),
            )
            import_run_id = int(cur.lastrowid)
            cur.execute("DELETE FROM painel_operador_anp_export_rows")

            counts: dict[str, dict[str, Any]] = {}
            total_files = 0
            total_rows = 0
            for filename, meta in self.ANP_EXPORT_FILES.items():
                path = self.module_root / filename
                source_count = {"rows": 0, "status": "missing", "family": meta["family"], "record_kind": meta["record_kind"]}
                counts[filename] = source_count
                if not path.exists():
                    continue
                total_files += 1
                source_count["status"] = "ok"
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    sheet = workbook["Export"] if "Export" in workbook.sheetnames else workbook.worksheets[0]
                    rows_iter = sheet.iter_rows(values_only=True)
                    headers = next(rows_iter, None)
                    if not headers:
                        source_count["status"] = "empty"
                        continue
                    header_map = {self._normalize_header(value): idx for idx, value in enumerate(headers)}
                    for row_number, row in enumerate(rows_iter, start=2):
                        if not any(value not in (None, "") for value in row):
                            continue
                        parsed = self._parse_anp_export_row(filename, meta, path, sheet.title, row_number, row, header_map)
                        self._insert_anp_export_row(cur, import_run_id, parsed)
                        source_count["rows"] += 1
                        total_rows += 1
                finally:
                    workbook.close()

            finished_at = self._now()
            cur.execute(
                """
                UPDATE painel_operador_anp_export_runs
                SET finished_at=?, status='ok', total_files=?, total_rows=?,
                    counts_json=?, notes=?
                WHERE id=?
                """,
                (
                    finished_at,
                    total_files,
                    total_rows,
                    self._json(counts),
                    "Exports ANP importados para staging.",
                    import_run_id,
                ),
            )
            conn.commit()
            return {
                "ok": True,
                "import_run_id": import_run_id,
                "started_at": started_at,
                "finished_at": finished_at,
                "total_files": total_files,
                "total_rows": total_rows,
                "counts": counts,
            }
        except Exception as exc:
            conn.rollback()
            raise exc
        finally:
            conn.close()

    def list_anp_exports(
        self,
        db_conn_fn,
        *,
        q: str = "",
        date_from: str = "",
        date_to: str = "",
        filters: dict[str, Any] | None = None,
        limit: int = 100,
        offset: int = 0,
        include_payload: bool = False,
    ) -> dict[str, Any]:
        return self._list_table_records(
            db_conn_fn,
            self.ANP_EXPORT_VIEW,
            "anp-exports",
            q=q,
            date_from=date_from,
            date_to=date_to,
            filters=filters,
            limit=limit,
            offset=offset,
            include_payload=include_payload,
        )

    def compare_anp_staging(
        self,
        db_conn_fn,
        *,
        family: str = "",
        tag: str = "",
        record_kind: str = "",
        status: str = "",
        date_from: str = "",
        date_to: str = "",
        limit: int = 100,
        offset: int = 0,
        tolerance: float = 0.001,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        offset = max(0, self._int_or_zero(offset))
        tolerance = max(0.0, self._float_or_none(tolerance) or 0.0)
        where = []
        params: list[Any] = []
        filters = {
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
            "record_kind": str(record_kind or "").strip(),
            "status": str(status or "").strip(),
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
        }
        if filters["family"]:
            where.append("family = ?")
            params.append(filters["family"])
        if filters["tag"]:
            where.append("tag = ?")
            params.append(filters["tag"])
        if filters["record_kind"]:
            where.append("record_kind = ?")
            params.append(filters["record_kind"])
        if filters["date_from"]:
            where.append("reference_date >= ?")
            params.append(filters["date_from"])
        if filters["date_to"]:
            where.append("reference_date <= ?")
            params.append(filters["date_to"])

        status_filter = filters["status"]
        summary_where = where[:]
        summary_params = params[:]
        if status_filter:
            summary_where.append("match_status = ?")
            summary_params.append(status_filter)
        base_where_sql = " WHERE " + " AND ".join(where) if where else ""
        summary_where_sql = " WHERE " + " AND ".join(summary_where) if summary_where else ""

        comparison_sql = f"""
        WITH combined AS (
            SELECT
                a.id AS anp_id,
                c.id AS staging_id,
                COALESCE(NULLIF(a.reference_date, ''), NULLIF(c.comparison_date, '')) AS reference_date,
                COALESCE(NULLIF(a.family, ''), NULLIF(c.family, '')) AS family,
                COALESCE(NULLIF(a.tag, ''), NULLIF(c.tag, '')) AS tag,
                a.record_kind,
                a.source_file,
                a.source_path,
                a.volume_corrigido,
                a.bsw_percent,
                a.failure_code,
                a.failure_type,
                c.status AS staging_status,
                c.anp_corrigido AS staging_anp_corrigido,
                c.raw_corrigido AS staging_raw_corrigido,
                c.xml_corrigido AS staging_xml_corrigido,
                c.note AS staging_note
            FROM painel_operador_anp_export_rows a
            LEFT JOIN painel_operador_comparisons c
                ON c.comparison_date = a.reference_date
               AND c.family = a.family
               AND c.tag = a.tag
            UNION ALL
            SELECT
                NULL AS anp_id,
                c.id AS staging_id,
                c.comparison_date AS reference_date,
                c.family,
                c.tag,
                '' AS record_kind,
                '' AS source_file,
                '' AS source_path,
                NULL AS volume_corrigido,
                NULL AS bsw_percent,
                '' AS failure_code,
                '' AS failure_type,
                c.status AS staging_status,
                c.anp_corrigido AS staging_anp_corrigido,
                c.raw_corrigido AS staging_raw_corrigido,
                c.xml_corrigido AS staging_xml_corrigido,
                c.note AS staging_note
            FROM painel_operador_comparisons c
            LEFT JOIN painel_operador_anp_export_rows a
                ON a.reference_date = c.comparison_date
               AND a.family = c.family
               AND a.tag = c.tag
            WHERE a.id IS NULL
        ),
        scored AS (
            SELECT
                *,
                CASE
                    WHEN record_kind IN ('linear_oil', 'linear_gas', 'differential_gas') THEN volume_corrigido
                    WHEN record_kind = 'inline_bsw' THEN bsw_percent
                    ELSE NULL
                END AS anp_value,
                staging_anp_corrigido AS staging_value
            FROM combined
        ),
        final AS (
            SELECT
                *,
                CASE
                    WHEN anp_id IS NULL THEN 'staging_only'
                    WHEN staging_id IS NULL THEN 'anp_only'
                    WHEN record_kind NOT IN ('linear_oil', 'linear_gas', 'differential_gas') THEN 'not_comparable'
                    WHEN anp_value IS NULL OR staging_value IS NULL THEN 'not_comparable'
                    WHEN ABS(anp_value - staging_value) <= ? THEN 'matched'
                    ELSE 'value_mismatch'
                END AS match_status,
                CASE
                    WHEN anp_value IS NULL OR staging_value IS NULL THEN NULL
                    ELSE anp_value - staging_value
                END AS delta,
                CASE
                    WHEN anp_value IS NULL OR staging_value IS NULL THEN NULL
                    ELSE ABS(anp_value - staging_value)
                END AS delta_abs
            FROM scored
        )
        SELECT * FROM final
        """
        base_params = [tolerance]
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            total = cur.execute(f"SELECT COUNT(*) FROM ({comparison_sql}){summary_where_sql}", base_params + summary_params).fetchone()[0]
            rows = cur.execute(
                f"SELECT * FROM ({comparison_sql}){summary_where_sql} ORDER BY reference_date DESC, family ASC, tag ASC LIMIT ? OFFSET ?",
                base_params + summary_params + [limit, offset],
            ).fetchall()
            groups = [
                dict(row)
                for row in cur.execute(
                    f"""
                    SELECT match_status, COUNT(*) AS count
                    FROM ({comparison_sql}){base_where_sql}
                    GROUP BY match_status
                    ORDER BY count DESC, match_status ASC
                    """,
                    base_params + params,
                ).fetchall()
            ]
            items = [dict(row) for row in rows]
            return {
                "record_type": "anp-staging-comparison",
                "total": total,
                "limit": limit,
                "offset": offset,
                "returned": len(items),
                "tolerance": tolerance,
                "filters": {key: value for key, value in filters.items() if value},
                "summary": groups,
                "items": items,
            }
        finally:
            conn.close()

    def measured_data(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        family: str = "",
        tag: str = "",
        source: str = "",
        limit: int = 120,
        offset: int = 0,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        offset = max(0, self._int_or_zero(offset))
        filters = {
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
            "source": str(source or "").strip(),
        }
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            daily = self._measured_daily_summary(cur, filters)
            rows = self._measured_rows(cur, filters, limit=limit, offset=offset)
            totals = {
                "days": len(daily),
                "rows_returned": len(rows["items"]),
                "rows_total": rows["total"],
                "fiscal_volume_m3": sum(self._float_or_none(row.get("fiscal_volume_m3")) or 0 for row in daily),
                "anp_volume_m3": sum(self._float_or_none(row.get("anp_volume_m3")) or 0 for row in daily),
                "mpfm_corr_hc_t": sum(self._float_or_none(row.get("mpfm_corr_hc_t")) or 0 for row in daily),
            }
            return {
                "record_type": "measured-data",
                "filters": {key: value for key, value in filters.items() if value},
                "limit": limit,
                "offset": offset,
                "totals": totals,
                "daily": daily,
                "rows": rows["items"],
                "total": rows["total"],
                "returned": len(rows["items"]),
                "notes": [
                    "Volumes fiscal/ANP estão em m3 quando vindos do Painel ANP/Radar.",
                    "MPFM diário está em toneladas nas métricas corr/uncorr existentes; delta direto requer normalização de unidade/densidade.",
                ],
            }
        finally:
            conn.close()

    def measurement_point_dossiers(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        family: str = "",
        tag: str = "",
        q: str = "",
        limit: int = 50,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        filters = {
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
            "q": str(q or "").strip(),
        }
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            tags = self._dossier_tag_keys(cur, filters, limit=limit)
            items = [self._measurement_point_dossier(cur, row["tag"], filters) for row in tags]
            items = [item for item in items if item]
            summary = {
                "points": len(items),
                "critical": sum(1 for item in items if item.get("health") == "critical"),
                "warning": sum(1 for item in items if item.get("health") == "warning"),
                "ok": sum(1 for item in items if item.get("health") == "ok"),
                "partial": sum(1 for item in items if item.get("health") == "partial"),
                "with_limits": sum(1 for item in items if int(item.get("limits", {}).get("count") or 0) > 0),
                "with_fiscal": sum(1 for item in items if int(item.get("fiscal", {}).get("rows") or 0) > 0),
                "with_anp": sum(1 for item in items if int(item.get("anp", {}).get("rows") or 0) > 0),
                "with_mpfm": sum(1 for item in items if int(item.get("mpfm", {}).get("rows") or 0) > 0),
            }
            return {
                "record_type": "measurement-point-dossiers",
                "filters": {key: value for key, value in filters.items() if value},
                "limit": limit,
                "returned": len(items),
                "summary": summary,
                "items": items,
                "notes": [
                    "Dossies unem staging Radar, exports ANP, limites/PAM, arquivos catalogados, propostas, evidencias e MPFM por TAG exata.",
                    "Fiscal/Radar e ANP estao em m3; MPFM permanece em t ate existir normalizacao validada por densidade/unidade.",
                    "TAGs fiscais podem nao ter equivalente direto em MPFM; nesse caso o dossie sinaliza cobertura parcial em vez de forcar comparacao indevida.",
                ],
            }
        finally:
            conn.close()

    def production_days(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        family: str = "",
        tag: str = "",
        category: str = "",
        limit: int = 90,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        filters = {
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
            "category": str(category or "").strip(),
        }
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            day_result = self._production_day_keys(cur, filters, limit=limit)
            dates = [row["production_date"] for row in day_result["items"]]
            if not dates:
                return {
                    "record_type": "production-days",
                    "filters": {key: value for key, value in filters.items() if value},
                    "limit": limit,
                    "total": 0,
                    "returned": 0,
                    "totals": {"days": 0, "complete_days": 0, "attention_days": 0, "file_count": 0, "open_pending_count": 0},
                    "items": [],
                }

            files = self._production_files_by_day(cur, dates, filters)
            fiscal = self._production_fiscal_by_day(cur, dates, filters)
            anp = self._production_anp_by_day(cur, dates, filters)
            mpfm = self._production_mpfm_by_day(cur, dates, filters)
            calendar = self._production_calendar_by_day(cur, dates)

            items = []
            for date_value in dates:
                row = {
                    "production_date": date_value,
                    **files.get(date_value, {}),
                    **fiscal.get(date_value, {}),
                    **anp.get(date_value, {}),
                    **mpfm.get(date_value, {}),
                    **calendar.get(date_value, {}),
                }
                defaults = {
                    "file_count": 0,
                    "file_tags_count": 0,
                    "daily_report_files": 0,
                    "fiscal_document_files": 0,
                    "anp_xml_files": 0,
                    "anp_export_files": 0,
                    "evidence_files": 0,
                    "calibration_files": 0,
                    "uncertainty_files": 0,
                    "pi_files": 0,
                    "fiscal_rows": 0,
                    "fiscal_tags_count": 0,
                    "fiscal_volume_m3": 0,
                    "fiscal_warning_rows": 0,
                    "anp_rows": 0,
                    "anp_tags_count": 0,
                    "anp_volume_m3": 0,
                    "bsw_rows": 0,
                    "failure_rows": 0,
                    "mpfm_rows": 0,
                    "mpfm_tags_count": 0,
                    "mpfm_hc_t": 0,
                    "mpfm_oil_t": 0,
                    "mpfm_gas_t": 0,
                    "mpfm_water_t": 0,
                    "open_pending_count": 0,
                    "resolved_pending_count": 0,
                    "calendar_status": "",
                    "loaded": None,
                    "file_samples": [],
                }
                for key, value in defaults.items():
                    row.setdefault(key, value)
                row["status"] = self._production_day_status(row)
                items.append(row)

            totals = {
                "days": len(items),
                "complete_days": sum(1 for row in items if row["status"] == "complete"),
                "attention_days": sum(1 for row in items if row["status"] == "attention"),
                "file_count": sum(int(row.get("file_count") or 0) for row in items),
                "open_pending_count": sum(int(row.get("open_pending_count") or 0) for row in items),
                "fiscal_volume_m3": sum(self._float_or_none(row.get("fiscal_volume_m3")) or 0 for row in items),
                "anp_volume_m3": sum(self._float_or_none(row.get("anp_volume_m3")) or 0 for row in items),
                "mpfm_hc_t": sum(self._float_or_none(row.get("mpfm_hc_t")) or 0 for row in items),
            }
            return {
                "record_type": "production-days",
                "filters": {key: value for key, value in filters.items() if value},
                "limit": limit,
                "total": day_result["total"],
                "returned": len(items),
                "totals": totals,
                "items": items,
                "notes": [
                    "Status completo exige presença de dados Fiscal/Radar, export ANP e MPFM no dia.",
                    "Volumes Fiscal/Radar e ANP estão em m3; MPFM HC está em t.",
                ],
            }
        finally:
            conn.close()

    def process_technical_monitor(self, db_conn_fn) -> dict[str, Any]:
        started_at = self._now()
        data = self._load_dashboard_data()
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cur.execute("DELETE FROM painel_operador_measurement_limits WHERE source_type='radar_limit_monitor'")
            cur.execute("DELETE FROM painel_operador_cv_config_snapshots")
            cur.execute("DELETE FROM painel_operador_cv_config_changes")

            limit_count = self._sync_measurement_limits(cur, data.get("limitMonitors") or [])
            files = self._technical_xml_files(cur)
            snapshot_count = 0
            file_counts = {"cv_parameters_xml": 0, "cv_security_xml": 0, "skipped_duplicates": 0, "parse_errors": 0}
            seen_files: set[tuple[str, str, str, str]] = set()
            for file_row in files:
                flow_computer = self._flow_computer_from_relative_path(file_row["relative_path"], file_row["inferred_tag"])
                dedupe_key = (
                    file_row["inferred_date"] or "",
                    flow_computer,
                    file_row["document_kind"] or "",
                    file_row["filename"] or "",
                )
                if dedupe_key in seen_files:
                    file_counts["skipped_duplicates"] += 1
                    continue
                seen_files.add(dedupe_key)
                try:
                    parsed_rows = self._parse_cv_config_file(file_row, flow_computer)
                except (ET.ParseError, OSError, UnicodeDecodeError):
                    file_counts["parse_errors"] += 1
                    continue
                if not parsed_rows:
                    continue
                self._insert_cv_snapshot_rows(cur, parsed_rows)
                snapshot_count += len(parsed_rows)
                file_counts[file_row["document_kind"]] = file_counts.get(file_row["document_kind"], 0) + 1

            change_count = self._build_cv_config_changes(cur)
            finished_at = self._now()
            conn.commit()
            return {
                "ok": True,
                "started_at": started_at,
                "finished_at": finished_at,
                "measurement_limits": limit_count,
                "cv_snapshot_rows": snapshot_count,
                "cv_change_rows": change_count,
                "files": file_counts,
                "notes": [
                    "Snapshots de Security.xml registram apenas usuario/nivel/permissao; hashes de senha e PIN nao sao persistidos.",
                    "Processamento e idempotente: tabelas derivadas de Limites/CV sao reconstruidas a cada execucao.",
                ],
            }
        except Exception as exc:
            conn.rollback()
            raise exc
        finally:
            conn.close()

    def technical_monitor(
        self,
        db_conn_fn,
        *,
        date_from: str = "",
        date_to: str = "",
        family: str = "",
        tag: str = "",
        limit: int = 120,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        filters = {
            "date_from": str(date_from or "").strip(),
            "date_to": str(date_to or "").strip(),
            "family": str(family or "").strip(),
            "tag": str(tag or "").strip(),
        }
        data = self._load_dashboard_data()
        limit_monitors = self._filter_dashboard_rows(data.get("limitMonitors") or [], filters, limit=limit)
        uncertainty_monitor = self._filter_dashboard_rows(data.get("uncertaintyMonitor") or [], filters, limit=limit)
        event_radar = data.get("eventEvidenceRadar") if isinstance(data.get("eventEvidenceRadar"), dict) else {}
        event_changes = self._filter_event_rows(event_radar.get("events") or [], filters, limit=limit)
        change_proposals = self._filter_dashboard_rows(data.get("changeProposals") or [], filters, limit=limit)

        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            cv_files = self._technical_cv_files(cur, filters, limit=limit)
            persisted_changes = self._technical_persisted_changes(cur, filters, limit=limit)
            checklist_ranges = self._technical_checklist_ranges(cur, filters, limit=limit)
            configured_limits = self._technical_configured_limits(cur, filters, limit=limit)
            cv_diagnostics = self._technical_cv_change_diagnostics(cur, filters)
            daily = self._measured_daily_summary(cur, {**filters, "source": ""})
            trend = self._technical_trend_rows(cur, filters, daily, limit_monitors, event_changes, persisted_changes)
            registry = self._technical_registry_summary(cur)
            return {
                "record_type": "technical-monitor",
                "filters": {key: value for key, value in filters.items() if value},
                "parameterization": {
                    "module": "Painel Operador > Limites & CV",
                    "limit_table": "painel_operador_measurement_limits",
                    "snapshot_table": "painel_operador_cv_config_snapshots",
                    "change_table": "painel_operador_cv_config_changes",
                    "editable_scope": [
                        "faixa calibrada por TAG/metrica/unidade",
                        "limites PAM por TAG/metrica/unidade",
                        "validade e evidencia do limite aprovado",
                        "severidade para mudancas em parametros CV",
                    ],
                    "guardrails": [
                        "nao comparar m3 com t sem normalizacao por densidade",
                        "alteracoes operacionais devem virar proposta auditavel antes de atualizar cadastro",
                        "frontend apenas apresenta; regra e historico ficam no SQLite",
                    ],
                },
                "registry": registry,
                "cv_files": cv_files,
                "rules": event_radar.get("rules") or [],
                "limit_monitors": limit_monitors,
                "uncertainty_monitor": uncertainty_monitor,
                "event_changes": event_changes,
                "persisted_changes": persisted_changes,
                "checklist_ranges": checklist_ranges,
                "configured_limits": configured_limits,
                "cv_diagnostics": cv_diagnostics,
                "change_proposals": change_proposals,
                "trend": trend,
                "summary": {
                    "limit_monitors": len(limit_monitors),
                    "uncertainty_rows": len(uncertainty_monitor),
                    "event_changes": len(event_changes),
                    "persisted_changes": len(persisted_changes),
                    "configured_limits": len(configured_limits),
                    "checklist_ranges": len(checklist_ranges),
                    "checklist_range_changes": sum(1 for row in checklist_ranges if row.get("range_changed")),
                    "change_proposals": len(change_proposals),
                    "cv_file_groups": len(cv_files.get("groups") or []),
                    "cv_compared_pairs": cv_diagnostics.get("compared_pairs", 0),
                    "cv_changed_pairs": cv_diagnostics.get("changed_pairs", 0),
                    "trend_days": len(trend),
                },
                "notes": [
                    "Limites/PAM aprovados devem ser parametrizados no SQLite, nao no JavaScript.",
                    "Parameters.xml/Security.xml entram como snapshots; AlarmsAndEvents entra como trilha de mudancas.",
                    "A tendencia usa Fiscal/Radar e ANP em m3; MPFM segue em t ate existir normalizacao validada.",
                ],
            }
        finally:
            conn.close()

    def _filter_dashboard_rows(self, rows: list[Any], filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        filtered = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            row_date = str(item.get("date") or item.get("createdAt") or item.get("created_at") or "")[:10]
            if filters.get("date_from") and row_date and row_date < filters["date_from"]:
                continue
            if filters.get("date_to") and row_date and row_date > filters["date_to"]:
                continue
            if filters.get("family") and str(item.get("family") or "").lower() != filters["family"].lower():
                continue
            if filters.get("tag"):
                item_tag = str(item.get("tag") or item.get("targetId") or item.get("target_id") or "")
                if item_tag.lower() != filters["tag"].lower():
                    continue
            filtered.append(item)
            if len(filtered) >= limit:
                break
        return filtered

    def _filter_event_rows(self, rows: list[Any], filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        filtered = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            event_date = str(item.get("timestamp") or "")[:10]
            if filters.get("date_from") and event_date and event_date < filters["date_from"]:
                continue
            if filters.get("date_to") and event_date and event_date > filters["date_to"]:
                continue
            if filters.get("tag"):
                tags = [str(value).lower() for value in (item.get("tags") or [])]
                if filters["tag"].lower() not in tags and filters["tag"].lower() not in str(item.get("parameter") or "").lower():
                    continue
            filtered.append(item)
            if len(filtered) >= limit:
                break
        return filtered

    def _sync_measurement_limits(self, cur, rows: list[Any]) -> int:
        inserted = 0
        now = self._now()
        metric_specs = (
            ("pam_flow_rate", "pam", "PAM / vazao operacional", "pam"),
            ("pressure", "pressure", "Faixa de pressao", "calibrated"),
            ("temperature", "temperature", "Faixa de temperatura", "calibrated"),
            ("differential", "differential", "Faixa diferencial", "calibrated"),
        )
        for item in rows:
            if not isinstance(item, dict):
                continue
            for metric_name, payload_key, label, limit_kind in metric_specs:
                payload = item.get(payload_key)
                if not isinstance(payload, dict):
                    continue
                lower = self._float_or_none(payload.get("lower"))
                upper = self._float_or_none(payload.get("upper"))
                if lower is None and upper is None:
                    continue
                calibrated_min = lower if limit_kind == "calibrated" else None
                calibrated_max = upper if limit_kind == "calibrated" else None
                pam_min = lower if limit_kind == "pam" else None
                pam_max = upper if limit_kind == "pam" else None
                source = str(item.get("source") or "")
                cur.execute(
                    """
                    INSERT INTO painel_operador_measurement_limits(
                        active, family, tag, metric_name, value_unit,
                        calibrated_min, calibrated_max, pam_min, pam_max,
                        alarm_low, alarm_high, valid_from, valid_to,
                        source_type, source_path, evidence_ref, approval_status,
                        notes, payload_json, created_at, updated_at
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        1,
                        str(item.get("family") or ""),
                        str(item.get("tag") or ""),
                        metric_name,
                        str(payload.get("unit") or ""),
                        calibrated_min,
                        calibrated_max,
                        pam_min,
                        pam_max,
                        lower,
                        upper,
                        str(item.get("date") or ""),
                        "",
                        "radar_limit_monitor",
                        source,
                        source,
                        "imported_from_radar",
                        label,
                        self._json({"monitor": item, "metric": metric_name, "latestValue": payload.get("value"), "latestStatus": payload.get("status")}),
                        now,
                        now,
                    ),
                )
                inserted += 1
        return inserted

    def _technical_xml_files(self, cur) -> list[dict[str, Any]]:
        if not self._table_exists(cur, "painel_operador_file_index"):
            return []
        rows = cur.execute(
            """
            SELECT inferred_date, inferred_tag, filename, document_kind,
                   relative_path, full_path, file_hash, modified_at
            FROM painel_operador_file_index
            WHERE ignored=0
              AND document_kind IN ('cv_parameters_xml','cv_security_xml')
              AND full_path<>''
            ORDER BY inferred_date ASC, document_kind ASC, LENGTH(relative_path) ASC, relative_path ASC
            """
        ).fetchall()
        return [dict(row) for row in rows]

    def _flow_computer_from_relative_path(self, relative_path: str, fallback: str = "") -> str:
        matches = re.findall(r"(?i)(?:^|[\\/])(FC\d{2})(?=[\\/])", str(relative_path or ""))
        if matches:
            return matches[-1].upper()
        match = re.search(r"\bFC\d{2}\b", str(fallback or ""), re.IGNORECASE)
        return match.group(0).upper() if match else str(fallback or "")

    def _parse_cv_config_file(self, file_row: dict[str, Any], flow_computer: str) -> list[dict[str, Any]]:
        path = Path(file_row["full_path"])
        if not path.exists():
            return []
        root = ET.parse(path).getroot()
        base = {
            "production_date": file_row.get("inferred_date") or "",
            "flow_computer": flow_computer,
            "tag": flow_computer,
            "source_file": file_row.get("full_path") or "",
            "file_hash": file_row.get("file_hash") or self._sha1_file(path),
            "created_at": self._now(),
        }
        if file_row.get("document_kind") == "cv_parameters_xml":
            parsed = []
            for element in root.iter("t"):
                name = str(element.attrib.get("n") or "").strip()
                if not name:
                    continue
                label = str(element.attrib.get("t") or "").strip()
                unit = str(element.attrib.get("u") or "").strip()
                value = str(element.attrib.get("v") or "").strip()
                parsed.append(
                    {
                        **base,
                        "parameter_path": name,
                        "parameter_name": name,
                        "parameter_value": value,
                        "value_unit": unit,
                        "payload_json": self._json({"kind": "parameter", "label": label, "unit_hint": unit}),
                    }
                )
            return parsed

        if file_row.get("document_kind") == "cv_security_xml":
            parsed = []
            for user in root.findall("./users/u"):
                username = str(user.attrib.get("n") or "").strip()
                if not username:
                    continue
                parsed.append(
                    {
                        **base,
                        "parameter_path": f"security.users.{username}.level",
                        "parameter_name": f"security.users.{username}.level",
                        "parameter_value": str(user.attrib.get("l") or "").strip(),
                        "value_unit": "access_level",
                        "payload_json": self._json({"kind": "security_user", "username": username, "full_name": user.attrib.get("fn") or "", "redacted": ["pw", "pin"]}),
                    }
                )
            levels = root.find("./levels")
            if levels is not None:
                for level in list(levels):
                    name = str(level.tag or "").strip()
                    if not name:
                        continue
                    parsed.append(
                        {
                            **base,
                            "parameter_path": f"security.levels.{name}",
                            "parameter_name": f"security.levels.{name}",
                            "parameter_value": str(level.attrib.get("v") or "").strip(),
                            "value_unit": "access_level",
                            "payload_json": self._json({"kind": "security_level", "permission": name}),
                        }
                    )
            return parsed
        return []

    def _xml_validation_item(self, cur, row: dict[str, Any]) -> dict[str, Any]:
        path = Path(row.get("full_path") or "")
        issues: list[str] = []
        status = "ok"
        xml_info = {
            "exists": path.exists(),
            "well_formed": False,
            "root": "",
            "elements": 0,
            "error": "",
        }
        if not xml_info["exists"]:
            status = "critical"
            issues.append("Arquivo XML não localizado no caminho indexado.")
        else:
            try:
                if (row.get("file_size_bytes") or 0) > 262_144:
                    xml_info["well_formed"] = True
                    xml_info["root"] = self._expected_xml_root_name(str(row.get("document_kind") or ""))
                    xml_info["shallow_check"] = True
                    xml_info["metadata_check"] = True
                else:
                    root = ET.parse(path).getroot()
                    xml_info["elements"] = sum(1 for _ in root.iter())
                    xml_info["well_formed"] = True
                    xml_info["root"] = self._xml_local_name(root.tag)
            except ET.ParseError as exc:
                status = "critical"
                xml_info["error"] = str(exc)
                issues.append("XML malformado ou incompleto.")
            except OSError as exc:
                status = "critical"
                xml_info["error"] = str(exc)
                issues.append("Falha ao ler o arquivo XML.")

        document_kind = str(row.get("document_kind") or "")
        inferred_date = str(row.get("inferred_date") or "")
        inferred_family = str(row.get("inferred_family") or "")
        inferred_tag = str(row.get("inferred_tag") or "")
        related = self._xml_related_counts(cur, row)

        if row.get("is_duplicate"):
            status = self._merge_xml_status(status, "warning")
            issues.append("Possível duplicidade pelo hash ou metadados do arquivo.")
        if document_kind in {"anp_fiscal_xml", "anp_failure_or_bsw_xml"} and not inferred_family:
            status = self._merge_xml_status(status, "warning")
            issues.append("Família ANP não inferida pelo nome/caminho.")
        if document_kind == "anp_fiscal_xml" and not inferred_date:
            status = self._merge_xml_status(status, "warning")
            issues.append("Dia de produção não inferido pelo pacote diário.")
        if document_kind == "anp_fiscal_xml" and related["comparison_rows"] == 0:
            status = self._merge_xml_status(status, "warning")
            issues.append("Sem comparação fiscal/XML/ANP correspondente no staging.")
        if document_kind in {"cv_parameters_xml", "cv_security_xml"} and related["cv_snapshot_rows"] == 0:
            status = self._merge_xml_status(status, "warning")
            issues.append("XML de CV ainda não processado no monitor Limites/CV.")
        if document_kind == "technical_xml" and not inferred_tag:
            status = self._merge_xml_status(status, "warning")
            issues.append("XML técnico sem tag inferida; revisar classificação antes de automatizar uso.")

        recommended_actions = self._xml_recommended_actions(
            row,
            document_kind=document_kind,
            status=status,
            related=related,
            issues=issues,
        )

        return {
            "id": row.get("id"),
            "relative_path": row.get("relative_path") or "",
            "full_path": row.get("full_path") or "",
            "filename": row.get("filename") or "",
            "category": row.get("category") or "",
            "document_kind": document_kind,
            "source_group": row.get("source_group") or "",
            "inferred_date": inferred_date,
            "inferred_family": inferred_family,
            "inferred_tag": inferred_tag,
            "file_size_bytes": row.get("file_size_bytes") or 0,
            "modified_at": row.get("modified_at") or "",
            "is_duplicate": int(row.get("is_duplicate") or 0),
            "status": status,
            "issues": issues,
            "note": " ".join(issues) if issues else "XML válido e coerente com os vínculos disponíveis.",
            "xml": xml_info,
            "related": related,
            "recommended_actions": recommended_actions,
        }

    def _xml_recommended_actions(
        self,
        row: dict[str, Any],
        *,
        document_kind: str,
        status: str,
        related: dict[str, int],
        issues: list[str],
    ) -> list[dict[str, Any]]:
        actions: list[dict[str, Any]] = []

        def add(code: str, label: str, detail: str = "", priority: str = "secondary") -> None:
            if any(action["code"] == code for action in actions):
                return
            actions.append({"code": code, "label": label, "detail": detail, "priority": priority})

        if status == "critical":
            add("open_files", "Revisar fonte", "Abrir o catálogo filtrado por este XML.", "primary")
            add("reindex_files", "Reindexar", "Atualizar a leitura dos caminhos configurados.")
            return actions

        if document_kind in {"cv_parameters_xml", "cv_security_xml"}:
            if related.get("cv_snapshot_rows", 0) == 0:
                add("process_technical", "Processar Limites/CV", "Extrair snapshots e mudanças dos XMLs de CV.", "primary")
            else:
                add("open_technical", "Abrir Limites/CV", "Ver snapshots, mudanças e limites ligados ao CV.", "primary")

        if document_kind == "anp_fiscal_xml":
            if related.get("comparison_rows", 0) == 0:
                add("sync_contract", "Sincronizar staging", "Recriar comparações fiscal/XML/ANP antes da apuração.", "primary")
            add("open_compare", "Abrir comparação", "Filtrar comparação ANP x staging por dia e família.")

        if document_kind == "anp_failure_or_bsw_xml":
            add("open_anp", "Abrir exports ANP", "Conferir falhas/BSW importados para a família.")

        if row.get("is_duplicate"):
            add("open_files", "Revisar duplicidade", "Abrir o catálogo filtrado por este arquivo.")

        if document_kind == "technical_xml" or any("tag inferida" in issue.lower() for issue in issues):
            add("open_files", "Revisar classificação", "Abrir o catálogo para validar tipo, tag e fonte.")

        if not actions:
            if document_kind == "anp_fiscal_xml":
                add("open_compare", "Abrir comparação", "Ver fechamento fiscal/XML/ANP.")
            elif document_kind in {"cv_parameters_xml", "cv_security_xml"}:
                add("open_technical", "Abrir Limites/CV", "Ver processamento técnico do CV.")
            else:
                add("open_files", "Abrir fonte", "Ver arquivo no catálogo.")
        return actions[:3]

    def _xml_related_counts(self, cur, row: dict[str, Any]) -> dict[str, int]:
        inferred_date = str(row.get("inferred_date") or "")
        inferred_family = str(row.get("inferred_family") or "")
        inferred_tag = str(row.get("inferred_tag") or "")
        full_path = str(row.get("full_path") or "")
        counts = {"comparison_rows": 0, "anp_export_rows": 0, "cv_snapshot_rows": 0}
        if inferred_date and inferred_family and self._table_exists(cur, "painel_operador_comparisons"):
            counts["comparison_rows"] = cur.execute(
                """
                SELECT COUNT(*)
                FROM painel_operador_comparisons
                WHERE comparison_date=? AND family=?
                  AND (?='' OR tag=?)
                """,
                (inferred_date, inferred_family, inferred_tag, inferred_tag),
            ).fetchone()[0]
        if inferred_family and self._table_exists(cur, "painel_operador_anp_export_rows"):
            if inferred_date:
                counts["anp_export_rows"] = cur.execute(
                    """
                    SELECT COUNT(*)
                    FROM painel_operador_anp_export_rows
                    WHERE family=? AND reference_date=?
                    """,
                    (inferred_family, inferred_date),
                ).fetchone()[0]
            else:
                counts["anp_export_rows"] = cur.execute(
                    "SELECT COUNT(*) FROM painel_operador_anp_export_rows WHERE family=?",
                    (inferred_family,),
                ).fetchone()[0]
        if full_path and self._table_exists(cur, "painel_operador_cv_config_snapshots"):
            counts["cv_snapshot_rows"] = cur.execute(
                "SELECT COUNT(*) FROM painel_operador_cv_config_snapshots WHERE source_file=?",
                (full_path,),
            ).fetchone()[0]
        return counts

    def _xml_validation_summary(self, items: list[dict[str, Any]]) -> dict[str, Any]:
        by_kind: dict[str, dict[str, Any]] = {}
        summary = {"total": len(items), "ok": 0, "warning": 0, "critical": 0, "by_kind": []}
        for item in items:
            status = item.get("status") or "warning"
            if status not in {"ok", "warning", "critical"}:
                status = "warning"
            summary[status] += 1
            kind = item.get("document_kind") or "unknown"
            group = by_kind.setdefault(kind, {"document_kind": kind, "count": 0, "ok": 0, "warning": 0, "critical": 0})
            group["count"] += 1
            group[status] += 1
        summary["by_kind"] = sorted(by_kind.values(), key=lambda item: (-item["count"], item["document_kind"]))
        return summary

    def _merge_xml_status(self, current: str, candidate: str) -> str:
        order = {"ok": 0, "warning": 1, "critical": 2}
        return candidate if order.get(candidate, 0) > order.get(current, 0) else current

    def _xml_local_name(self, tag_name: Any) -> str:
        text = str(tag_name or "")
        return text.rsplit("}", 1)[-1] if "}" in text else text

    def _expected_xml_root_name(self, document_kind: str) -> str:
        if document_kind == "cv_parameters_xml":
            return "tags"
        if document_kind == "cv_security_xml":
            return "security"
        return "não verificado"

    def _xml_root_fast(self, path: Path):
        parser = ET.iterparse(path, events=("start",))
        for _, root in parser:
            close = getattr(parser, "close", None)
            if callable(close):
                close()
            return root
        raise ET.ParseError("XML sem elemento raiz.")

    def _insert_cv_snapshot_rows(self, cur, rows: list[dict[str, Any]]) -> None:
        cur.executemany(
            """
            INSERT INTO painel_operador_cv_config_snapshots(
                production_date, flow_computer, tag, source_file, file_hash,
                parameter_path, parameter_name, parameter_value, value_unit,
                payload_json, created_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            [
                (
                    row["production_date"],
                    row["flow_computer"],
                    row["tag"],
                    row["source_file"],
                    row["file_hash"],
                    row["parameter_path"],
                    row["parameter_name"],
                    row["parameter_value"],
                    row["value_unit"],
                    row["payload_json"],
                    row["created_at"],
                )
                for row in rows
            ],
        )

    def _build_cv_config_changes(self, cur) -> int:
        rows = cur.execute(
            """
            SELECT production_date, flow_computer, tag, parameter_name,
                   parameter_value, value_unit, MIN(source_file) AS source_file
            FROM painel_operador_cv_config_snapshots
            WHERE production_date<>'' AND parameter_name<>''
            GROUP BY production_date, flow_computer, tag, parameter_name
            ORDER BY flow_computer ASC, tag ASC, parameter_name ASC, production_date ASC
            """
        ).fetchall()
        now = self._now()
        inserts = []
        previous_by_key: dict[tuple[str, str, str], sqlite3.Row] = {}
        for row in rows:
            key = (row["flow_computer"] or "", row["tag"] or "", row["parameter_name"] or "")
            previous = previous_by_key.get(key)
            if previous and str(previous["parameter_value"]) != str(row["parameter_value"]):
                delta = self._numeric_delta(previous["parameter_value"], row["parameter_value"])
                severity = self._cv_change_severity(row["parameter_name"])
                payload = {
                    "value_unit": row["value_unit"] or "",
                    "numeric_delta": delta,
                    "previous_source": previous["source_file"] or "",
                    "current_source": row["source_file"] or "",
                }
                inserts.append(
                    (
                        previous["production_date"],
                        row["production_date"],
                        row["flow_computer"] or "",
                        row["tag"] or "",
                        row["parameter_name"] or "",
                        str(previous["parameter_value"] or ""),
                        str(row["parameter_value"] or ""),
                        delta,
                        "modified",
                        severity,
                        row["source_file"] or "",
                        row["source_file"] or "",
                        self._json(payload),
                        now,
                    )
                )
            previous_by_key[key] = row
        if inserts:
            cur.executemany(
                """
                INSERT INTO painel_operador_cv_config_changes(
                    previous_date, current_date, flow_computer, tag, parameter_name,
                    previous_value, current_value, delta_value, change_type, severity,
                    source_file, evidence_ref, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                inserts,
            )
        return len(inserts)

    def _numeric_delta(self, old_value: Any, new_value: Any) -> float | None:
        old_num = self._float_or_none(old_value)
        new_num = self._float_or_none(new_value)
        if old_num is None or new_num is None:
            return None
        return new_num - old_num

    def _cv_change_severity(self, parameter_name: str) -> str:
        text = str(parameter_name or "").lower()
        if "security." in text or "write" in text or "suppress" in text or "ackalarms" in text:
            return "critical"
        critical_terms = ("hihilim", "lololim", "hi limit", "low limit", "limit", "pam", "cutoff", "range")
        if any(term in text for term in critical_terms):
            return "warning"
        review_terms = ("density", "bsw", "viscos", "compress", "meter", "factor", "k_", "pvt", "chromat")
        if any(term in text for term in review_terms):
            return "warning"
        return "info"

    def _technical_registry_summary(self, cur) -> dict[str, Any]:
        tables = {}
        for table in (
            "painel_operador_measurement_limits",
            "painel_operador_cv_config_snapshots",
            "painel_operador_cv_config_changes",
        ):
            tables[table] = {
                "exists": self._table_exists(cur, table),
                "rows": cur.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] if self._table_exists(cur, table) else 0,
            }
        return {"tables": tables}

    def _technical_configured_limits(self, cur, filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        if not self._table_exists(cur, "painel_operador_measurement_limits"):
            return []
        where = []
        params: list[Any] = []
        if filters.get("family"):
            where.append("family = ?")
            params.append(filters["family"])
        if filters.get("tag"):
            where.append("tag = ?")
            params.append(filters["tag"])
        if filters.get("date_from"):
            where.append("(valid_to = '' OR valid_to >= ?)")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("(valid_from = '' OR valid_from <= ?)")
            params.append(filters["date_to"])
        where_sql = " WHERE " + " AND ".join(where) if where else ""
        rows = cur.execute(
            f"""
            SELECT id, active, family, tag, metric_name, value_unit,
                   calibrated_min, calibrated_max, pam_min, pam_max,
                   alarm_low, alarm_high, valid_from, valid_to,
                   source_type, source_path, evidence_ref, approval_status,
                   notes, created_at, updated_at
            FROM painel_operador_measurement_limits
            {where_sql}
            ORDER BY active DESC,
                     CASE approval_status WHEN 'approved' THEN 0 WHEN 'aprovado' THEN 0 WHEN 'review' THEN 1 ELSE 2 END,
                     tag ASC, metric_name ASC, valid_from DESC, id DESC
            LIMIT ?
            """,
            params + [limit],
        ).fetchall()
        return [dict(row) for row in rows]

    def save_measurement_limit(self, db_conn_fn, payload: dict[str, Any]) -> dict[str, Any]:
        payload = payload or {}
        tag = str(payload.get("tag") or "").strip()
        metric_name = str(payload.get("metric_name") or "").strip()
        if not tag:
            raise ValueError("TAG do limite é obrigatória.")
        if not metric_name:
            raise ValueError("Métrica do limite é obrigatória.")
        now = self._now()
        row_id = int(payload.get("id") or 0)
        values = {
            "active": 1 if str(payload.get("active", "1")).lower() not in {"0", "false", "inativo", "inactive"} else 0,
            "family": str(payload.get("family") or "").strip(),
            "tag": tag,
            "metric_name": metric_name,
            "value_unit": str(payload.get("value_unit") or "").strip(),
            "calibrated_min": self._float_or_none(payload.get("calibrated_min")),
            "calibrated_max": self._float_or_none(payload.get("calibrated_max")),
            "pam_min": self._float_or_none(payload.get("pam_min")),
            "pam_max": self._float_or_none(payload.get("pam_max")),
            "alarm_low": self._float_or_none(payload.get("alarm_low")),
            "alarm_high": self._float_or_none(payload.get("alarm_high")),
            "valid_from": str(payload.get("valid_from") or "").strip(),
            "valid_to": str(payload.get("valid_to") or "").strip(),
            "source_type": str(payload.get("source_type") or "manual_approved").strip() or "manual_approved",
            "source_path": str(payload.get("source_path") or "").strip(),
            "evidence_ref": str(payload.get("evidence_ref") or "").strip(),
            "approval_status": str(payload.get("approval_status") or "approved").strip() or "approved",
            "notes": str(payload.get("notes") or "").strip(),
            "payload_json": self._json({"saved_from": "painel_operador_limit_form", "raw": payload}),
        }
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            if row_id:
                exists = cur.execute(
                    "SELECT id FROM painel_operador_measurement_limits WHERE id=?",
                    (row_id,),
                ).fetchone()
                if not exists:
                    raise ValueError("Limite informado não foi encontrado.")
                cur.execute(
                    """
                    UPDATE painel_operador_measurement_limits
                    SET active=?, family=?, tag=?, metric_name=?, value_unit=?,
                        calibrated_min=?, calibrated_max=?, pam_min=?, pam_max=?,
                        alarm_low=?, alarm_high=?, valid_from=?, valid_to=?,
                        source_type=?, source_path=?, evidence_ref=?, approval_status=?,
                        notes=?, payload_json=?, updated_at=?
                    WHERE id=?
                    """,
                    (
                        values["active"], values["family"], values["tag"], values["metric_name"], values["value_unit"],
                        values["calibrated_min"], values["calibrated_max"], values["pam_min"], values["pam_max"],
                        values["alarm_low"], values["alarm_high"], values["valid_from"], values["valid_to"],
                        values["source_type"], values["source_path"], values["evidence_ref"], values["approval_status"],
                        values["notes"], values["payload_json"], now, row_id,
                    ),
                )
                action = "updated"
            else:
                cur.execute(
                    """
                    INSERT INTO painel_operador_measurement_limits(
                        active, family, tag, metric_name, value_unit,
                        calibrated_min, calibrated_max, pam_min, pam_max,
                        alarm_low, alarm_high, valid_from, valid_to,
                        source_type, source_path, evidence_ref, approval_status,
                        notes, payload_json, created_at, updated_at
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        values["active"], values["family"], values["tag"], values["metric_name"], values["value_unit"],
                        values["calibrated_min"], values["calibrated_max"], values["pam_min"], values["pam_max"],
                        values["alarm_low"], values["alarm_high"], values["valid_from"], values["valid_to"],
                        values["source_type"], values["source_path"], values["evidence_ref"], values["approval_status"],
                        values["notes"], values["payload_json"], now, now,
                    ),
                )
                row_id = int(cur.lastrowid)
                action = "created"
            conn.commit()
            row = cur.execute(
                """
                SELECT id, active, family, tag, metric_name, value_unit,
                       calibrated_min, calibrated_max, pam_min, pam_max,
                       alarm_low, alarm_high, valid_from, valid_to,
                       source_type, source_path, evidence_ref, approval_status,
                       notes, created_at, updated_at
                FROM painel_operador_measurement_limits
                WHERE id=?
                """,
                (row_id,),
            ).fetchone()
            return {"ok": True, "action": action, "limit": dict(row)}
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _technical_cv_files(self, cur, filters: dict[str, str], *, limit: int) -> dict[str, Any]:
        kinds = ("cv_parameters_xml", "cv_security_xml", "cv_alarm_event_txt", "cv_run_daily_txt", "cv_run_24hours_txt", "cv_run_hourly_txt")
        where = [f"document_kind IN ({','.join('?' for _ in kinds)})", "ignored=0"]
        params: list[Any] = list(kinds)
        if filters.get("date_from"):
            where.append("inferred_date >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("inferred_date <= ?")
            params.append(filters["date_to"])
        if filters.get("tag"):
            where.append("(inferred_tag = ? OR inferred_tag = '')")
            params.append(filters["tag"])
        where_sql = " AND ".join(where)
        groups = [
            dict(row)
            for row in cur.execute(
                f"""
                SELECT document_kind, COUNT(*) AS count,
                       MIN(NULLIF(inferred_date, '')) AS first_date,
                       MAX(NULLIF(inferred_date, '')) AS last_date,
                       COUNT(DISTINCT NULLIF(inferred_tag, '')) AS tags_count,
                       SUM(file_size_bytes) AS size_bytes
                FROM painel_operador_file_index
                WHERE {where_sql}
                GROUP BY document_kind
                ORDER BY document_kind
                """,
                params,
            ).fetchall()
        ]
        files = [
            dict(row)
            for row in cur.execute(
                f"""
                SELECT inferred_date, inferred_tag, filename, document_kind,
                       relative_path, full_path, file_hash, modified_at
                FROM painel_operador_file_index
                WHERE {where_sql}
                ORDER BY inferred_date DESC, document_kind ASC, filename ASC
                LIMIT ?
                """,
                params + [limit],
            ).fetchall()
        ]
        return {"groups": groups, "files": files}

    def _technical_checklist_ranges(self, cur, filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        if not self._table_exists(cur, "painel_operador_daily_checklist_rows"):
            return []
        latest_run = cur.execute("SELECT MAX(id) FROM painel_operador_daily_checklist_runs").fetchone()
        latest_run_id = int(latest_run[0] or 0) if latest_run else 0
        if not latest_run_id:
            return []
        where = ["import_run_id=?", "record_domain='flow_computer_range'"]
        params: list[Any] = [latest_run_id]
        if filters.get("date_from"):
            where.append("record_date >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("record_date <= ?")
            params.append(filters["date_to"])
        if filters.get("tag"):
            where.append("tag = ?")
            params.append(filters["tag"])
        rows = cur.execute(
            f"""
            SELECT sheet_name, record_date, tag, status, title, payload_json
            FROM painel_operador_daily_checklist_rows
            WHERE {' AND '.join(where)}
            ORDER BY tag ASC, record_date ASC, row_number ASC
            """,
            params,
        ).fetchall()
        previous_by_tag: dict[str, dict[str, Any]] = {}
        parsed: list[dict[str, Any]] = []
        for row in rows:
            payload = self._loads_json(row["payload_json"])
            values = payload.get("values") if isinstance(payload, dict) else {}
            if not isinstance(values, dict):
                values = {}
            tag = str(row["tag"] or row["sheet_name"] or "").replace("-Fx", "")
            current = {
                "date": row["record_date"] or "",
                "tag": tag,
                "sheet_name": row["sheet_name"] or "",
                "pressure_min": self._float_or_none(values.get("E")),
                "pressure_max": self._float_or_none(values.get("F")),
                "temperature_min": self._float_or_none(values.get("H")),
                "temperature_max": self._float_or_none(values.get("I")),
                "qcorr_min": self._float_or_none(values.get("O")),
                "qcorr_max": self._float_or_none(values.get("P")),
                "bsw_min": self._float_or_none(values.get("X")),
                "bsw_max": self._float_or_none(values.get("W")),
                "comment": str(values.get("T") or "").strip(),
                "source": "daily_checklist_fx",
            }
            signature = (
                current["pressure_min"],
                current["pressure_max"],
                current["temperature_min"],
                current["temperature_max"],
                current["qcorr_min"],
                current["qcorr_max"],
                current["bsw_min"],
                current["bsw_max"],
            )
            if not any(value is not None for value in signature):
                continue
            previous = previous_by_tag.get(tag)
            current["range_changed"] = bool(previous and previous.get("signature") != signature)
            current["previous_date"] = previous.get("date", "") if previous else ""
            current["status"] = "changed" if current["range_changed"] else "ok"
            if not tag or not current["date"]:
                current["status"] = "warning"
            normalized_comment = current["comment"].strip().lower()
            if normalized_comment in {"0", "0.0", "1", "1.0", "-", "ok"}:
                current["comment"] = ""
                normalized_comment = ""
            if normalized_comment:
                current["status"] = "warning" if current["status"] == "ok" else current["status"]
            current["signature"] = signature
            previous_by_tag[tag] = current
            parsed.append(current)
        parsed = sorted(parsed, key=lambda item: (item.get("date") or "", item.get("tag") or ""), reverse=True)
        for item in parsed:
            item.pop("signature", None)
        return parsed[:limit]

    def _technical_cv_change_diagnostics(self, cur, filters: dict[str, str]) -> dict[str, Any]:
        if not self._table_exists(cur, "painel_operador_cv_config_snapshots"):
            return {"production_dates": 0, "flow_computers": 0, "parameters": 0, "compared_pairs": 0, "changed_pairs": 0}
        where = ["production_date<>''", "parameter_name<>''"]
        params: list[Any] = []
        if filters.get("date_from"):
            where.append("production_date >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("production_date <= ?")
            params.append(filters["date_to"])
        if filters.get("tag"):
            where.append("(tag = ? OR flow_computer = ?)")
            params.extend([filters["tag"], filters["tag"]])
        where_sql = " AND ".join(where)
        stats = dict(
            cur.execute(
                f"""
                SELECT COUNT(DISTINCT production_date) AS production_dates,
                       MIN(production_date) AS first_date,
                       MAX(production_date) AS last_date,
                       COUNT(DISTINCT flow_computer) AS flow_computers,
                       COUNT(DISTINCT parameter_name) AS parameters,
                       COUNT(*) AS snapshot_rows
                FROM painel_operador_cv_config_snapshots
                WHERE {where_sql}
                """,
                params,
            ).fetchone()
        )
        rows = cur.execute(
            f"""
            SELECT production_date, flow_computer, tag, parameter_name, parameter_value
            FROM painel_operador_cv_config_snapshots
            WHERE {where_sql}
            ORDER BY flow_computer ASC, tag ASC, parameter_name ASC, production_date ASC
            """,
            params,
        ).fetchall()
        compared = 0
        changed = 0
        previous_by_key: dict[tuple[str, str, str], sqlite3.Row] = {}
        for row in rows:
            key = (row["flow_computer"] or "", row["tag"] or "", row["parameter_name"] or "")
            previous = previous_by_key.get(key)
            if previous:
                compared += 1
                if str(previous["parameter_value"] or "") != str(row["parameter_value"] or ""):
                    changed += 1
            previous_by_key[key] = row
        stats["compared_pairs"] = compared
        stats["changed_pairs"] = changed
        return stats

    def _technical_persisted_changes(self, cur, filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        if not self._table_exists(cur, "painel_operador_cv_config_changes"):
            return []
        where = []
        params: list[Any] = []
        if filters.get("date_from"):
            where.append('"current_date" >= ?')
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append('"current_date" <= ?')
            params.append(filters["date_to"])
        if filters.get("tag"):
            where.append("(tag = ? OR flow_computer = ?)")
            params.extend([filters["tag"], filters["tag"]])
        where_sql = " WHERE " + " AND ".join(where) if where else ""
        rows = cur.execute(
            f"""
            SELECT previous_date, current_date, flow_computer, tag, parameter_name,
                   previous_value, current_value, delta_value, change_type,
                   severity, source_file, evidence_ref
            FROM painel_operador_cv_config_changes
            {where_sql}
            ORDER BY "current_date" DESC,
                     CASE severity WHEN 'critical' THEN 0 WHEN 'warning' THEN 1 ELSE 2 END,
                     flow_computer ASC, parameter_name ASC
            LIMIT ?
            """,
            params + [limit],
        ).fetchall()
        return [dict(row) for row in rows]

    def _technical_trend_rows(
        self,
        cur,
        filters: dict[str, str],
        daily: list[dict[str, Any]],
        limit_monitors: list[dict[str, Any]],
        event_changes: list[dict[str, Any]],
        persisted_changes: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        trend = sorted(daily, key=lambda row: row.get("measurement_date") or "")[-120:]
        cv_counts = self._technical_counts_by_day(
            cur,
            filters,
            "painel_operador_file_index",
            "inferred_date",
            "ignored=0 AND document_kind IN ('cv_parameters_xml','cv_security_xml','cv_alarm_event_txt','cv_run_daily_txt','cv_run_24hours_txt','cv_run_hourly_txt')",
        )
        out_of_range = self._technical_counts_by_day(
            cur,
            filters,
            "painel_operador_measurement_points",
            "point_date",
            "in_range=0",
        )
        change_counts: dict[str, int] = {}
        for event in event_changes:
            day = str(event.get("timestamp") or "")[:10]
            if day:
                change_counts[day] = change_counts.get(day, 0) + 1
        for change in persisted_changes:
            day = str(change.get("current_date") or "")[:10]
            if day:
                change_counts[day] = change_counts.get(day, 0) + 1
        limit_counts: dict[str, int] = {}
        for row in limit_monitors:
            day = str(row.get("date") or "")[:10]
            status = str(row.get("status") or "").lower()
            if day and status and status not in {"ok", "normal", "dentro"}:
                limit_counts[day] = limit_counts.get(day, 0) + 1
        for row in trend:
            day = str(row.get("measurement_date") or "")[:10]
            row["cv_file_count"] = cv_counts.get(day, 0)
            row["config_change_count"] = change_counts.get(day, 0)
            row["out_of_range_points"] = out_of_range.get(day, 0)
            row["limit_alert_count"] = limit_counts.get(day, 0)
        return trend

    def _technical_counts_by_day(self, cur, filters: dict[str, str], table: str, date_column: str, extra_where: str) -> dict[str, int]:
        if not self._table_exists(cur, table):
            return {}
        where = [extra_where, f"{date_column}<>''"]
        params: list[Any] = []
        if filters.get("date_from"):
            where.append(f"{date_column} >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append(f"{date_column} <= ?")
            params.append(filters["date_to"])
        if filters.get("tag") and table == "painel_operador_measurement_points":
            where.append("tag = ?")
            params.append(filters["tag"])
        rows = cur.execute(
            f"""
            SELECT {date_column} AS day, COUNT(*) AS count
            FROM {table}
            WHERE {' AND '.join(where)}
            GROUP BY {date_column}
            """,
            params,
        ).fetchall()
        return {row["day"]: int(row["count"] or 0) for row in rows}

    def _dossier_tag_keys(self, cur, filters: dict[str, str], *, limit: int) -> list[dict[str, Any]]:
        clauses: list[str] = []
        outer_params: list[Any] = []
        inner_params: list[Any] = []
        if filters.get("date_from"):
            clauses.append("latest_date >= ?")
            outer_params.append(filters["date_from"])
        if filters.get("date_to"):
            clauses.append("latest_date <= ?")
            outer_params.append(filters["date_to"])
        if filters.get("family"):
            clauses.append("family = ?")
            outer_params.append(filters["family"])
        if filters.get("tag"):
            clauses.append("tag = ?")
            outer_params.append(filters["tag"])
        if filters.get("q"):
            like = f"%{filters['q']}%"
            clauses.append("(tag LIKE ? OR family LIKE ? OR family_name LIKE ?)")
            outer_params.extend([like, like, like])
        where_sql = " WHERE " + " AND ".join(clauses) if clauses else ""
        mpfm_union = ""
        if filters.get("tag"):
            inner_params.extend([filters["tag"], filters["tag"]])
            mpfm_union = """
                UNION ALL
                SELECT tag, '' AS family, tipo AS family_name, MAX(day_ref) AS latest_date, 'mpfm' AS source
                FROM measurements_curated
                WHERE COALESCE(tag,'')<>'' AND row_kind='daily' AND COALESCE(is_official,1)=1
                  AND (tag=? OR instrument=?)
                GROUP BY tag, tipo
            """
        rows = cur.execute(
            f"""
            WITH tag_sources AS (
                SELECT tag, family, family_name, MAX(point_date) AS latest_date, 'point' AS source
                FROM painel_operador_measurement_points
                WHERE COALESCE(tag,'')<>''
                GROUP BY tag, family, family_name
                UNION ALL
                SELECT tag, family, family_name, MAX(comparison_date) AS latest_date, 'fiscal' AS source
                FROM painel_operador_comparisons
                WHERE COALESCE(tag,'')<>''
                GROUP BY tag, family, family_name
                UNION ALL
                SELECT tag, family, export_name AS family_name, MAX(reference_date) AS latest_date, 'anp' AS source
                FROM painel_operador_anp_export_rows
                WHERE COALESCE(tag,'')<>''
                GROUP BY tag, family, export_name
                UNION ALL
                SELECT tag, family, '' AS family_name, MAX(valid_from) AS latest_date, 'limit' AS source
                FROM painel_operador_measurement_limits
                WHERE COALESCE(tag,'')<>'' AND COALESCE(active,1)=1
                GROUP BY tag, family
                {mpfm_union}
            ),
            ranked AS (
                SELECT
                    tag,
                    MAX(NULLIF(family,'')) AS family,
                    MAX(NULLIF(family_name,'')) AS family_name,
                    MAX(NULLIF(latest_date,'')) AS latest_date,
                    COUNT(DISTINCT source) AS source_count
                FROM tag_sources
                GROUP BY tag
            )
            SELECT tag, COALESCE(family,'') AS family, COALESCE(family_name,'') AS family_name,
                   COALESCE(latest_date,'') AS latest_date, source_count
            FROM ranked
            {where_sql}
            ORDER BY source_count DESC, latest_date DESC, tag ASC
            LIMIT ?
            """,
            inner_params + outer_params + [limit],
        ).fetchall()
        return [dict(row) for row in rows]

    def _measurement_point_dossier(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        tag = str(tag or "").strip()
        if not tag:
            return {}
        point = self._dossier_latest_point(cur, tag, filters)
        fiscal = self._dossier_fiscal(cur, tag, filters)
        anp = self._dossier_anp(cur, tag, filters)
        mpfm = self._dossier_mpfm(cur, tag, filters) if filters.get("tag") else {
            "rows": 0,
            "samples": [],
            "coverage": "not_checked_without_explicit_tag",
        }
        limits = self._dossier_limits(cur, tag, filters)
        files = self._dossier_files(cur, tag, filters)
        proposals = self._dossier_proposals(cur, tag, filters)
        evidence = self._dossier_evidence(cur, tag, filters)
        critical_signals = int(limits.get("critical") or 0) + int(fiscal.get("warnings") or 0) + int(anp.get("failure_rows") or 0)
        warning_signals = int(limits.get("warning") or 0) + int(proposals.get("pending") or 0)
        if critical_signals:
            health = "critical"
        elif warning_signals:
            health = "warning"
        elif int(fiscal.get("rows") or 0) and int(anp.get("rows") or 0):
            health = "ok"
        else:
            health = "partial"
        latest_candidates = [
            point.get("point_date"),
            fiscal.get("last_date"),
            anp.get("last_date"),
            mpfm.get("last_date"),
        ]
        return {
            "tag": tag,
            "family": point.get("family") or fiscal.get("family") or anp.get("family") or "",
            "family_name": point.get("family_name") or fiscal.get("family_name") or anp.get("family_name") or "",
            "fluid": point.get("fluid") or "",
            "meter_type": point.get("meter_type") or "",
            "computador_vazao": point.get("computador_vazao") or "",
            "active_status": point.get("active_status") or "",
            "latest_date": max([str(value) for value in latest_candidates if value] or [""]),
            "health": health,
            "signals": {"critical": critical_signals, "warning": warning_signals},
            "latest_point": point,
            "limits": limits,
            "fiscal": fiscal,
            "anp": anp,
            "mpfm": mpfm,
            "files": files,
            "proposals": proposals,
            "evidence": evidence,
        }

    def _dossier_date_where(self, column: str, filters: dict[str, str], params: list[Any]) -> list[str]:
        where = []
        if filters.get("date_from"):
            where.append(f"substr({column},1,10) >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append(f"substr({column},1,10) <= ?")
            params.append(filters["date_to"])
        return where

    def _dossier_latest_point(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["tag=?"]
        where += self._dossier_date_where("point_date", filters, params)
        row = cur.execute(
            f"""
            SELECT point_date, family, family_name, tag, fluid, meter_type, active_status,
                   computador_vazao, volume_corrigido, volume_bruto, volume_liquido,
                   temperatura, pressao, in_range
            FROM painel_operador_measurement_points
            WHERE {' AND '.join(where)}
            ORDER BY point_date DESC, id DESC
            LIMIT 1
            """,
            params,
        ).fetchone()
        return dict(row) if row else {"tag": tag}

    def _dossier_fiscal(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["tag=?"]
        where += self._dossier_date_where("comparison_date", filters, params)
        row = cur.execute(
            f"""
            SELECT MAX(family) AS family, MAX(family_name) AS family_name,
                   COUNT(*) AS rows, MIN(comparison_date) AS first_date, MAX(comparison_date) AS last_date,
                   SUM(COALESCE(raw_corrigido,0)) AS raw_m3,
                   SUM(COALESCE(xml_corrigido,0)) AS xml_m3,
                   SUM(COALESCE(anp_corrigido,0)) AS anp_m3,
                   SUM(CASE WHEN status<>'' AND status<>'ok' THEN 1 ELSE 0 END) AS warnings
            FROM painel_operador_comparisons
            WHERE {' AND '.join(where)}
            """,
            params,
        ).fetchone()
        sample_rows = cur.execute(
            f"""
            SELECT comparison_date, family, status, raw_corrigido, xml_corrigido,
                   anp_corrigido, note, xml_source_local, raw_source_local
            FROM painel_operador_comparisons
            WHERE {' AND '.join(where)}
            ORDER BY comparison_date DESC, id DESC
            LIMIT 8
            """,
            params,
        ).fetchall()
        result = dict(row) if row else {}
        result["samples"] = [dict(item) for item in sample_rows]
        return result

    def _dossier_anp(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["tag=?"]
        where += self._dossier_date_where("reference_date", filters, params)
        row = cur.execute(
            f"""
            SELECT MAX(family) AS family, MAX(export_name) AS family_name,
                   COUNT(*) AS rows, MIN(reference_date) AS first_date, MAX(reference_date) AS last_date,
                   SUM(CASE WHEN record_kind IN ('linear_oil','linear_gas','differential_gas') THEN COALESCE(volume_corrigido,0) ELSE 0 END) AS volume_m3,
                   SUM(CASE WHEN record_kind='inline_bsw' THEN 1 ELSE 0 END) AS bsw_rows,
                   SUM(CASE WHEN record_kind='measurement_failure' THEN 1 ELSE 0 END) AS failure_rows,
                   COUNT(DISTINCT record_kind) AS kinds
            FROM painel_operador_anp_export_rows
            WHERE {' AND '.join(where)}
            """,
            params,
        ).fetchone()
        sample_rows = cur.execute(
            f"""
            SELECT reference_date, family, export_name, record_kind, volume_corrigido,
                   bsw_percent, failure_code, failure_type, source_file
            FROM painel_operador_anp_export_rows
            WHERE {' AND '.join(where)}
            ORDER BY reference_date DESC, id DESC
            LIMIT 8
            """,
            params,
        ).fetchall()
        result = dict(row) if row else {}
        result["samples"] = [dict(item) for item in sample_rows]
        return result

    def _dossier_mpfm(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag, tag]
        where = ["(tag=? OR instrument=?)", "row_kind='daily'", "COALESCE(is_official,1)=1"]
        where += self._dossier_date_where("day_ref", filters, params)
        row = cur.execute(
            f"""
            SELECT COUNT(*) AS rows, MIN(day_ref) AS first_date, MAX(day_ref) AS last_date,
                   COUNT(DISTINCT bank) AS banks,
                   SUM(CASE WHEN metric_name='MPFM corr HC (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS hc_t,
                   SUM(CASE WHEN metric_name='MPFM corr Óleo (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS oil_t,
                   SUM(CASE WHEN metric_name='MPFM corr Gás (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS gas_t,
                   SUM(CASE WHEN metric_name='MPFM corr Água (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS water_t
            FROM measurements_curated
            WHERE {' AND '.join(where)}
            """,
            params,
        ).fetchone()
        sample_rows = cur.execute(
            f"""
            SELECT day_ref, bank, tipo, tag, instrument, metric_name, metric_value,
                   metric_unit, COALESCE(source_file, excel_file, '') AS source_path
            FROM measurements_curated
            WHERE {' AND '.join(where)}
            ORDER BY day_ref DESC, bank ASC, metric_name ASC
            LIMIT 8
            """,
            params,
        ).fetchall()
        result = dict(row) if row else {}
        result["samples"] = [dict(item) for item in sample_rows]
        return result

    def _dossier_limits(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["tag=?", "COALESCE(active,1)=1"]
        if filters.get("family"):
            where.append("(family=? OR family='')")
            params.append(filters["family"])
        rows = [dict(row) for row in cur.execute(
            f"""
            SELECT family, tag, metric_name, value_unit, calibrated_min, calibrated_max,
                   pam_min, pam_max, alarm_low, alarm_high, valid_from, valid_to,
                   approval_status, source_type, evidence_ref, notes, payload_json
            FROM painel_operador_measurement_limits
            WHERE {' AND '.join(where)}
            ORDER BY metric_name ASC, valid_from DESC, id DESC
            """,
            params,
        ).fetchall()]
        critical = 0
        warning = 0
        for row in rows:
            payload = self._loads_json(row.get("payload_json"), {})
            status = str(payload.get("latestStatus") or payload.get("status") or row.get("approval_status") or "").lower()
            row["latest_status"] = status
            if status == "critical":
                critical += 1
            elif status and status not in {"ok", "approved", "aprovado"}:
                warning += 1
        return {"count": len(rows), "critical": critical, "warning": warning, "items": rows[:8]}

    def _dossier_files(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["ignored=0", "inferred_tag=?"]
        where += self._dossier_date_where("inferred_date", filters, params)
        row = cur.execute(
            f"""
            SELECT COUNT(*) AS count, COUNT(DISTINCT category) AS categories,
                   MIN(NULLIF(inferred_date,'')) AS first_date, MAX(NULLIF(inferred_date,'')) AS last_date,
                   SUM(file_size_bytes) AS size_bytes
            FROM painel_operador_file_index
            WHERE {' AND '.join(where)}
            """,
            params,
        ).fetchone()
        samples = [dict(item) for item in cur.execute(
            f"""
            SELECT inferred_date, category, document_kind, filename, relative_path, file_size_bytes
            FROM painel_operador_file_index
            WHERE {' AND '.join(where)}
            ORDER BY inferred_date DESC, parse_priority DESC, filename ASC
            LIMIT 8
            """,
            params,
        ).fetchall()]
        result = dict(row) if row else {}
        result["samples"] = samples
        return result

    def _dossier_proposals(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        params: list[Any] = [tag]
        where = ["target_id=?"]
        where += self._dossier_date_where("created_at_source", filters, params)
        rows = [dict(row) for row in cur.execute(
            f"""
            SELECT proposal_id, domain, title, target_id, field_name, confidence,
                   risk, status, evidence_state, recommended_action, created_at_source
            FROM painel_operador_proposals
            WHERE {' AND '.join(where)}
            ORDER BY created_at_source DESC, id DESC
            LIMIT 8
            """,
            params,
        ).fetchall()]
        return {
            "count": len(rows),
            "pending": sum(1 for row in rows if str(row.get("status") or "") == "pending_authorization"),
            "items": rows,
        }

    def _dossier_evidence(self, cur, tag: str, filters: dict[str, str]) -> dict[str, Any]:
        like = f"%{tag}%"
        params: list[Any] = [tag, like, like, like]
        where = ["(target_id=? OR title LIKE ? OR source_path LIKE ? OR payload_json LIKE ?)"]
        where += self._dossier_date_where("event_at", filters, params)
        rows = [dict(row) for row in cur.execute(
            f"""
            SELECT evidence_kind, event_at, requirement_id, title, status,
                   evidence_state, target_id, target_type, source_path, local_path
            FROM painel_operador_evidence
            WHERE {' AND '.join(where)}
            ORDER BY event_at DESC, id DESC
            LIMIT 8
            """,
            params,
        ).fetchall()]
        return {"count": len(rows), "items": rows}

    def _production_day_keys(self, cur, filters: dict[str, str], *, limit: int) -> dict[str, Any]:
        params: list[Any] = []
        subqueries: list[str] = []

        file_where = ["ignored=0", "inferred_date<>''"]
        if filters.get("category"):
            file_where.append("category=?")
            params.append(filters["category"])
        if filters.get("family"):
            file_where.append("(inferred_family=? OR category='daily_report')")
            params.append(filters["family"])
        if filters.get("tag"):
            file_where.append("(inferred_tag=? OR category='daily_report')")
            params.append(filters["tag"])
        subqueries.append(f"SELECT inferred_date AS production_date FROM painel_operador_file_index WHERE {' AND '.join(file_where)}")

        fiscal_where = ["comparison_date<>''"]
        if filters.get("family"):
            fiscal_where.append("family=?")
            params.append(filters["family"])
        if filters.get("tag"):
            fiscal_where.append("tag=?")
            params.append(filters["tag"])
        subqueries.append(f"SELECT comparison_date AS production_date FROM painel_operador_comparisons WHERE {' AND '.join(fiscal_where)}")

        anp_where = ["reference_date<>''"]
        if filters.get("family"):
            anp_where.append("family=?")
            params.append(filters["family"])
        if filters.get("tag"):
            anp_where.append("tag=?")
            params.append(filters["tag"])
        subqueries.append(f"SELECT reference_date AS production_date FROM painel_operador_anp_export_rows WHERE {' AND '.join(anp_where)}")

        mpfm_where = ["row_kind='daily'", "COALESCE(is_official,1)=1", "day_ref<>''"]
        if filters.get("tag"):
            mpfm_where.append("(tag=? OR instrument=?)")
            params.extend([filters["tag"], filters["tag"]])
        if not filters.get("category"):
            subqueries.append(f"SELECT day_ref AS production_date FROM measurements_curated WHERE {' AND '.join(mpfm_where)}")

        if not filters.get("family") and not filters.get("tag") and not filters.get("category"):
            subqueries.append("SELECT calendar_date AS production_date FROM painel_operador_calendar_days WHERE calendar_date<>''")

        where = []
        outer_params = list(params)
        if filters.get("date_from"):
            where.append("production_date >= ?")
            outer_params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("production_date <= ?")
            outer_params.append(filters["date_to"])
        where_sql = " WHERE " + " AND ".join(where) if where else ""
        union_sql = " UNION ALL ".join(subqueries)
        total = cur.execute(
            f"SELECT COUNT(DISTINCT production_date) FROM ({union_sql}){where_sql}",
            outer_params,
        ).fetchone()[0]
        rows = cur.execute(
            f"""
            SELECT production_date
            FROM ({union_sql})
            {where_sql}
            GROUP BY production_date
            ORDER BY production_date DESC
            LIMIT ?
            """,
            outer_params + [limit],
        ).fetchall()
        return {"total": total, "items": [dict(row) for row in rows]}

    def _day_in_clause(self, dates: list[str]) -> str:
        return ",".join("?" for _ in dates)

    def _production_files_by_day(self, cur, dates: list[str], filters: dict[str, str]) -> dict[str, dict[str, Any]]:
        where = [f"inferred_date IN ({self._day_in_clause(dates)})", "ignored=0"]
        params: list[Any] = list(dates)
        if filters.get("category"):
            where.append("category=?")
            params.append(filters["category"])
        if filters.get("family"):
            where.append("(inferred_family=? OR category='daily_report')")
            params.append(filters["family"])
        if filters.get("tag"):
            where.append("(inferred_tag=? OR category='daily_report')")
            params.append(filters["tag"])
        grouped = {
            row["production_date"]: dict(row)
            for row in cur.execute(
                f"""
                SELECT
                    inferred_date AS production_date,
                    COUNT(*) AS file_count,
                    COUNT(DISTINCT inferred_tag) AS file_tags_count,
                    SUM(CASE WHEN category='daily_report' THEN 1 ELSE 0 END) AS daily_report_files,
                    SUM(CASE WHEN category IN ('fiscal_document','anp_xml') THEN 1 ELSE 0 END) AS fiscal_document_files,
                    SUM(CASE WHEN document_kind IN ('anp_fiscal_xml','anp_fiscal_archive','anp_failure_or_bsw_xml') THEN 1 ELSE 0 END) AS anp_xml_files,
                    SUM(CASE WHEN category='anp_operator_export' THEN 1 ELSE 0 END) AS anp_export_files,
                    SUM(CASE WHEN category IN ('evidence_document','regulatory_document','operational_document','memorial_document') THEN 1 ELSE 0 END) AS evidence_files,
                    SUM(CASE WHEN category='calibration_certificate' THEN 1 ELSE 0 END) AS calibration_files,
                    SUM(CASE WHEN category='uncertainty_document' THEN 1 ELSE 0 END) AS uncertainty_files,
                    SUM(CASE WHEN category='pi_timeseries' THEN 1 ELSE 0 END) AS pi_files
                FROM painel_operador_file_index
                WHERE {' AND '.join(where)}
                GROUP BY inferred_date
                """,
                params,
            ).fetchall()
        }
        sample_rows = cur.execute(
            f"""
            SELECT inferred_date AS production_date, filename, category, document_kind
            FROM painel_operador_file_index
            WHERE {' AND '.join(where)}
            ORDER BY inferred_date DESC, parse_priority DESC, category ASC, filename ASC
            LIMIT 800
            """,
            params,
        ).fetchall()
        for row in sample_rows:
            bucket = grouped.setdefault(row["production_date"], {"production_date": row["production_date"]})
            samples = bucket.setdefault("file_samples", [])
            if len(samples) < 4:
                samples.append({"filename": row["filename"], "category": row["category"], "document_kind": row["document_kind"]})
        return grouped

    def _production_fiscal_by_day(self, cur, dates: list[str], filters: dict[str, str]) -> dict[str, dict[str, Any]]:
        where = [f"comparison_date IN ({self._day_in_clause(dates)})"]
        params: list[Any] = list(dates)
        if filters.get("family"):
            where.append("family=?")
            params.append(filters["family"])
        if filters.get("tag"):
            where.append("tag=?")
            params.append(filters["tag"])
        return {
            row["production_date"]: dict(row)
            for row in cur.execute(
                f"""
                SELECT
                    comparison_date AS production_date,
                    COUNT(*) AS fiscal_rows,
                    COUNT(DISTINCT tag) AS fiscal_tags_count,
                    SUM(COALESCE(anp_corrigido,0)) AS fiscal_volume_m3,
                    SUM(CASE WHEN status<>'' AND status<>'ok' THEN 1 ELSE 0 END) AS fiscal_warning_rows
                FROM painel_operador_comparisons
                WHERE {' AND '.join(where)}
                GROUP BY comparison_date
                """,
                params,
            ).fetchall()
        }

    def _production_anp_by_day(self, cur, dates: list[str], filters: dict[str, str]) -> dict[str, dict[str, Any]]:
        where = [f"reference_date IN ({self._day_in_clause(dates)})"]
        params: list[Any] = list(dates)
        if filters.get("family"):
            where.append("family=?")
            params.append(filters["family"])
        if filters.get("tag"):
            where.append("tag=?")
            params.append(filters["tag"])
        return {
            row["production_date"]: dict(row)
            for row in cur.execute(
                f"""
                SELECT
                    reference_date AS production_date,
                    COUNT(*) AS anp_rows,
                    COUNT(DISTINCT tag) AS anp_tags_count,
                    SUM(CASE WHEN record_kind IN ('linear_oil','linear_gas','differential_gas') THEN COALESCE(volume_corrigido,0) ELSE 0 END) AS anp_volume_m3,
                    SUM(CASE WHEN record_kind='inline_bsw' THEN 1 ELSE 0 END) AS bsw_rows,
                    SUM(CASE WHEN record_kind='measurement_failure' THEN 1 ELSE 0 END) AS failure_rows
                FROM painel_operador_anp_export_rows
                WHERE {' AND '.join(where)}
                GROUP BY reference_date
                """,
                params,
            ).fetchall()
        }

    def _production_mpfm_by_day(self, cur, dates: list[str], filters: dict[str, str]) -> dict[str, dict[str, Any]]:
        where = [f"day_ref IN ({self._day_in_clause(dates)})", "row_kind='daily'", "COALESCE(is_official,1)=1"]
        params: list[Any] = list(dates)
        if filters.get("tag"):
            where.append("(tag=? OR instrument=?)")
            params.extend([filters["tag"], filters["tag"]])
        return {
            row["production_date"]: dict(row)
            for row in cur.execute(
                f"""
                SELECT
                    day_ref AS production_date,
                    COUNT(*) AS mpfm_rows,
                    COUNT(DISTINCT tag) AS mpfm_tags_count,
                    SUM(CASE WHEN metric_name='MPFM corr HC (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS mpfm_hc_t,
                    SUM(CASE WHEN metric_name='MPFM corr Óleo (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS mpfm_oil_t,
                    SUM(CASE WHEN metric_name='MPFM corr Gás (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS mpfm_gas_t,
                    SUM(CASE WHEN metric_name='MPFM corr Água (t)' THEN COALESCE(metric_value,0) ELSE 0 END) AS mpfm_water_t
                FROM measurements_curated
                WHERE {' AND '.join(where)}
                GROUP BY day_ref
                """,
                params,
            ).fetchall()
        }

    def _production_calendar_by_day(self, cur, dates: list[str]) -> dict[str, dict[str, Any]]:
        rows = cur.execute(
            f"""
            SELECT
                calendar_date AS production_date,
                status AS calendar_status,
                loaded,
                points_count AS calendar_points,
                open_pending_count,
                resolved_pending_count,
                missing_xml_families_json
            FROM painel_operador_calendar_days
            WHERE calendar_date IN ({self._day_in_clause(dates)})
            """,
            list(dates),
        ).fetchall()
        return {row["production_date"]: dict(row) for row in rows}

    def _production_day_status(self, row: dict[str, Any]) -> str:
        if int(row.get("open_pending_count") or 0) > 0 or int(row.get("fiscal_warning_rows") or 0) > 0:
            return "attention"
        if int(row.get("fiscal_rows") or 0) > 0 and int(row.get("anp_rows") or 0) > 0 and int(row.get("mpfm_rows") or 0) > 0:
            return "complete"
        if int(row.get("file_count") or 0) > 0 or int(row.get("fiscal_rows") or 0) > 0 or int(row.get("anp_rows") or 0) > 0 or int(row.get("mpfm_rows") or 0) > 0:
            return "partial"
        return "empty"

    def _measured_daily_summary(self, cur, filters: dict[str, str]) -> list[dict[str, Any]]:
        date_where = []
        params: list[Any] = []
        if filters.get("date_from"):
            date_where.append("measurement_date >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            date_where.append("measurement_date <= ?")
            params.append(filters["date_to"])
        if filters.get("family"):
            date_where.append("(family = ? OR source = 'mpfm_daily')")
            params.append(filters["family"])
        if filters.get("tag"):
            date_where.append("(tag = ? OR instrument = ?)")
            params.extend([filters["tag"], filters["tag"]])
        source_filter = filters.get("source")
        if source_filter:
            date_where.append("source = ?")
            params.append(source_filter)
        where_sql = " WHERE " + " AND ".join(date_where) if date_where else ""
        rows = cur.execute(
            f"""
            WITH measured AS (
                SELECT
                    'fiscal_radar' AS source,
                    comparison_date AS measurement_date,
                    family,
                    tag,
                    '' AS instrument,
                    anp_corrigido AS fiscal_volume_m3,
                    NULL AS anp_volume_m3,
                    NULL AS mpfm_corr_hc_t,
                    status
                FROM painel_operador_comparisons
                UNION ALL
                SELECT
                    'anp_export' AS source,
                    reference_date AS measurement_date,
                    family,
                    tag,
                    '' AS instrument,
                    NULL AS fiscal_volume_m3,
                    CASE
                        WHEN record_kind IN ('linear_oil', 'linear_gas', 'differential_gas') THEN volume_corrigido
                        ELSE NULL
                    END AS anp_volume_m3,
                    NULL AS mpfm_corr_hc_t,
                    '' AS status
                FROM painel_operador_anp_export_rows
                UNION ALL
                SELECT
                    'mpfm_daily' AS source,
                    day_ref AS measurement_date,
                    '' AS family,
                    tag,
                    instrument,
                    NULL AS fiscal_volume_m3,
                    NULL AS anp_volume_m3,
                    CASE WHEN metric_name = 'MPFM corr HC (t)' THEN metric_value ELSE NULL END AS mpfm_corr_hc_t,
                    '' AS status
                FROM measurements_curated
                WHERE row_kind='daily' AND COALESCE(is_official,1)=1
            )
            SELECT
                measurement_date,
                COUNT(*) AS row_count,
                COUNT(DISTINCT source) AS sources_count,
                COUNT(DISTINCT tag) AS tags_count,
                SUM(COALESCE(fiscal_volume_m3, 0)) AS fiscal_volume_m3,
                SUM(COALESCE(anp_volume_m3, 0)) AS anp_volume_m3,
                SUM(COALESCE(mpfm_corr_hc_t, 0)) AS mpfm_corr_hc_t,
                SUM(CASE WHEN source='fiscal_radar' THEN 1 ELSE 0 END) AS fiscal_rows,
                SUM(CASE WHEN source='anp_export' THEN 1 ELSE 0 END) AS anp_rows,
                SUM(CASE WHEN source='mpfm_daily' THEN 1 ELSE 0 END) AS mpfm_rows,
                SUM(CASE WHEN status='ok' THEN 1 ELSE 0 END) AS ok_rows,
                SUM(CASE WHEN status<>'' AND status<>'ok' THEN 1 ELSE 0 END) AS warning_rows
            FROM measured
            {where_sql}
            GROUP BY measurement_date
            ORDER BY measurement_date DESC
            LIMIT 120
            """,
            params,
        ).fetchall()
        return [dict(row) for row in rows]

    def _measured_rows(self, cur, filters: dict[str, str], *, limit: int, offset: int) -> dict[str, Any]:
        where = []
        params: list[Any] = []
        if filters.get("date_from"):
            where.append("measurement_date >= ?")
            params.append(filters["date_from"])
        if filters.get("date_to"):
            where.append("measurement_date <= ?")
            params.append(filters["date_to"])
        if filters.get("family"):
            where.append("(family = ? OR source = 'mpfm_daily')")
            params.append(filters["family"])
        if filters.get("tag"):
            where.append("(tag = ? OR instrument = ?)")
            params.extend([filters["tag"], filters["tag"]])
        if filters.get("source"):
            where.append("source = ?")
            params.append(filters["source"])
        where_sql = " WHERE " + " AND ".join(where) if where else ""
        base_sql = """
            WITH measured AS (
                SELECT
                    'fiscal_radar' AS source,
                    comparison_date AS measurement_date,
                    family,
                    family_name,
                    tag,
                    fluid,
                    '' AS record_kind,
                    '' AS bank,
                    '' AS instrument,
                    raw_corrigido,
                    xml_corrigido,
                    anp_corrigido,
                    NULL AS volume_corrigido,
                    NULL AS volume_bruto,
                    NULL AS volume_liquido,
                    NULL AS bsw_percent,
                    NULL AS pressure_kpa,
                    NULL AS temperature_c,
                    '' AS metric_name,
                    NULL AS metric_value,
                    '' AS metric_unit,
                    status,
                    xml_source_local AS source_path
                FROM painel_operador_comparisons
                UNION ALL
                SELECT
                    'anp_export' AS source,
                    reference_date AS measurement_date,
                    family,
                    export_name AS family_name,
                    tag,
                    '' AS fluid,
                    record_kind,
                    '' AS bank,
                    '' AS instrument,
                    NULL AS raw_corrigido,
                    NULL AS xml_corrigido,
                    NULL AS anp_corrigido,
                    volume_corrigido,
                    volume_bruto,
                    volume_liquido,
                    bsw_percent,
                    pressure_kpa,
                    temperature_c,
                    '' AS metric_name,
                    NULL AS metric_value,
                    'm3/%/kPa/C' AS metric_unit,
                    '' AS status,
                    source_path
                FROM painel_operador_anp_export_rows
                UNION ALL
                SELECT
                    'mpfm_daily' AS source,
                    day_ref AS measurement_date,
                    '' AS family,
                    tipo AS family_name,
                    tag,
                    '' AS fluid,
                    row_kind AS record_kind,
                    bank,
                    instrument,
                    NULL AS raw_corrigido,
                    NULL AS xml_corrigido,
                    NULL AS anp_corrigido,
                    NULL AS volume_corrigido,
                    NULL AS volume_bruto,
                    NULL AS volume_liquido,
                    NULL AS bsw_percent,
                    NULL AS pressure_kpa,
                    NULL AS temperature_c,
                    metric_name,
                    metric_value,
                    metric_unit,
                    '' AS status,
                    COALESCE(source_file, excel_file, '') AS source_path
                FROM measurements_curated
                WHERE row_kind='daily' AND COALESCE(is_official,1)=1
            )
            SELECT * FROM measured
        """
        total = cur.execute(f"SELECT COUNT(*) FROM ({base_sql}){where_sql}", params).fetchone()[0]
        rows = cur.execute(
            f"SELECT * FROM ({base_sql}){where_sql} ORDER BY measurement_date DESC, source ASC, family ASC, tag ASC LIMIT ? OFFSET ?",
            params + [limit, offset],
        ).fetchall()
        return {"total": total, "items": [dict(row) for row in rows]}

    def list_staging_records(
        self,
        db_conn_fn,
        record_type: str,
        *,
        q: str = "",
        date_from: str = "",
        date_to: str = "",
        filters: dict[str, Any] | None = None,
        limit: int = 100,
        offset: int = 0,
        include_payload: bool = False,
    ) -> dict[str, Any]:
        view = self.STAGING_VIEWS.get(str(record_type or "").strip().lower())
        if not view:
            raise ValueError(f"Tipo de staging inválido: {record_type}")
        return self._list_table_records(
            db_conn_fn,
            view,
            record_type,
            q=q,
            date_from=date_from,
            date_to=date_to,
            filters=filters,
            limit=limit,
            offset=offset,
            include_payload=include_payload,
        )

    def decide_calendar_pendency(self, db_conn_fn, pendency_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        raw_id = str(pendency_id or "").strip()
        if not raw_id:
            raise ValueError("Pendência não informada.")
        status = str((payload or {}).get("status") or "").strip()
        if status not in {"open", "resolved", "deferred", "ignored"}:
            raise ValueError("Status de pendência inválido.")
        resolution_mode = str((payload or {}).get("resolution_mode") or "").strip()
        if not resolution_mode:
            resolution_mode = {
                "resolved": "manual_close",
                "deferred": "manual_defer",
                "ignored": "manual_ignore",
            }.get(status, "")
        closed_by = str((payload or {}).get("closed_by") or "usuario_local").strip()
        decision_note = str((payload or {}).get("decision_note") or "").strip()
        closed_at = self._now() if status in {"resolved", "deferred", "ignored"} else ""

        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            row = cur.execute(
                """
                SELECT *
                FROM painel_operador_calendar_pendencies
                WHERE CAST(id AS TEXT)=? OR pendency_id=?
                ORDER BY id DESC
                LIMIT 1
                """,
                (raw_id, raw_id),
            ).fetchone()
            if not row:
                raise ValueError(f"Pendência não encontrada: {raw_id}")
            db_id = int(row["id"])
            target_id = str(row["pendency_id"] or raw_id)
            cur.execute(
                """
                UPDATE painel_operador_calendar_pendencies
                SET status=?, resolution_mode=?, closed_by=?, closed_at=?, decision_note=?
                WHERE id=?
                """,
                (status, resolution_mode, closed_by, closed_at, decision_note, db_id),
            )
            self._insert_decision_audit(
                cur,
                target_type="calendar_pendency",
                target_id=target_id,
                target_db_id=db_id,
                previous_status=str(row["status"] or ""),
                decision_status=status,
                decision_mode=resolution_mode,
                decided_by=closed_by,
                decided_at=closed_at or self._now(),
                decision_note=decision_note,
                payload=payload,
            )
            conn.commit()
            updated = cur.execute(
                "SELECT * FROM painel_operador_calendar_pendencies WHERE id=?",
                (db_id,),
            ).fetchone()
            return {
                "ok": True,
                "record_type": "calendar_pendency",
                "item": self._serialize_staging_row(dict(updated), include_payload=True),
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def decide_proposal(self, db_conn_fn, proposal_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        raw_id = str(proposal_id or "").strip()
        if not raw_id:
            raise ValueError("Proposta não informada.")
        status = str((payload or {}).get("status") or "").strip()
        if status not in {"authorized", "rejected", "deferred"}:
            raise ValueError("Status de proposta inválido.")
        decision_mode = str((payload or {}).get("decision_mode") or "").strip() or f"manual_{status}"
        decided_by = str((payload or {}).get("decided_by") or (payload or {}).get("authorized_by") or "usuario_local").strip()
        decision_note = str((payload or {}).get("decision_note") or "").strip()
        decided_at = self._now()

        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            row = cur.execute(
                """
                SELECT *
                FROM painel_operador_proposals
                WHERE CAST(id AS TEXT)=? OR proposal_id=?
                ORDER BY id DESC
                LIMIT 1
                """,
                (raw_id, raw_id),
            ).fetchone()
            if not row:
                raise ValueError(f"Proposta não encontrada: {raw_id}")
            db_id = int(row["id"])
            target_id = str(row["proposal_id"] or raw_id)
            audit_trail = self._loads_json(row["audit_trail_json"], [])
            if not isinstance(audit_trail, list):
                audit_trail = []
            audit_trail.append(
                {
                    "status": status,
                    "decisionMode": decision_mode,
                    "decidedBy": decided_by,
                    "decidedAt": decided_at,
                    "decisionNote": decision_note,
                }
            )
            cur.execute(
                """
                UPDATE painel_operador_proposals
                SET status=?, authorized_by=?, authorized_at=?, decision_note=?, audit_trail_json=?
                WHERE id=?
                """,
                (status, decided_by, decided_at, decision_note, self._json(audit_trail), db_id),
            )
            self._insert_decision_audit(
                cur,
                target_type="proposal",
                target_id=target_id,
                target_db_id=db_id,
                previous_status=str(row["status"] or ""),
                decision_status=status,
                decision_mode=decision_mode,
                decided_by=decided_by,
                decided_at=decided_at,
                decision_note=decision_note,
                payload=payload,
            )
            conn.commit()
            updated = cur.execute(
                "SELECT * FROM painel_operador_proposals WHERE id=?",
                (db_id,),
            ).fetchone()
            return {
                "ok": True,
                "record_type": "proposal",
                "item": self._serialize_staging_row(dict(updated), include_payload=True),
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _list_table_records(
        self,
        db_conn_fn,
        view: dict[str, Any],
        record_type: str,
        *,
        q: str = "",
        date_from: str = "",
        date_to: str = "",
        filters: dict[str, Any] | None = None,
        limit: int = 100,
        offset: int = 0,
        include_payload: bool = False,
    ) -> dict[str, Any]:
        limit = self._normalize_query_limit(limit)
        offset = max(0, self._int_or_zero(offset))
        where = []
        params: list[Any] = []
        date_column = view["date_column"]
        if date_from:
            where.append(f"{date_column} >= ?")
            params.append(str(date_from))
        if date_to:
            where.append(f"{date_column} <= ?")
            params.append(str(date_to))
        q = str(q or "").strip()
        if q:
            like = f"%{q}%"
            search_parts = [f"COALESCE({column}, '') LIKE ?" for column in view["search_columns"]]
            where.append("(" + " OR ".join(search_parts) + ")")
            params.extend([like] * len(search_parts))
        active_filters = {}
        for key, column in view["filters"].items():
            value = (filters or {}).get(key)
            if value is None or value == "":
                continue
            active_filters[key] = value
            where.append(f"{column} = ?")
            params.append(value)

        where_sql = " WHERE " + " AND ".join(where) if where else ""
        table = view["table"]
        order = view["order"]
        conn = db_conn_fn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        try:
            total = cur.execute(f'SELECT COUNT(*) FROM "{table}"{where_sql}', params).fetchone()[0]
            rows = cur.execute(
                f'SELECT * FROM "{table}"{where_sql} ORDER BY {order} LIMIT ? OFFSET ?',
                params + [limit, offset],
            ).fetchall()
            items = [self._serialize_staging_row(dict(row), include_payload=include_payload) for row in rows]
            return {
                "record_type": record_type,
                "total": total,
                "limit": limit,
                "offset": offset,
                "returned": len(items),
                "filters": {
                    "q": q,
                    "date_from": date_from,
                    "date_to": date_to,
                    **active_filters,
                },
                "items": items,
            }
        finally:
            conn.close()

    def sync_to_staging(self, db_conn_fn) -> dict[str, Any]:
        data = self._load_dashboard_data()
        started_at = datetime.now().isoformat(timespec="seconds")
        source_hash = self._sha1_file(self.dashboard_data_path)
        conn = db_conn_fn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                INSERT INTO painel_operador_sync_runs(
                    started_at, finished_at, source_data_path, source_data_hash, status, counts_json, notes
                )
                VALUES(?,?,?,?,?,?,?)
                """,
                (
                    started_at,
                    started_at,
                    str(self.dashboard_data_path),
                    source_hash,
                    "running",
                    "{}",
                    "Sincronizacao iniciada.",
                ),
            )
            sync_run_id = int(cur.lastrowid)
            for table in self._staging_tables():
                cur.execute(f'DELETE FROM "{table}"')

            counts = {
                "sources": self._sync_sources(cur, sync_run_id, data),
                "measurement_points": self._sync_measurement_points(cur, sync_run_id, data),
                "comparisons": self._sync_comparisons(cur, sync_run_id, data),
                "evidence": self._sync_evidence(cur, sync_run_id, data),
                "alerts": self._sync_alerts(cur, sync_run_id, data),
                "proposals": self._sync_proposals(cur, sync_run_id, data),
                "calendar": self._sync_calendar(cur, sync_run_id, data),
            }
            finished_at = datetime.now().isoformat(timespec="seconds")
            cur.execute(
                """
                UPDATE painel_operador_sync_runs
                SET finished_at=?, status='ok', counts_json=?, notes=?
                WHERE id=?
                """,
                (
                    finished_at,
                    json.dumps(counts, ensure_ascii=False, sort_keys=True),
                    "Snapshot de staging atualizado.",
                    sync_run_id,
                ),
            )
            conn.commit()
            return {
                "ok": True,
                "sync_run_id": sync_run_id,
                "source_data_hash": source_hash,
                "started_at": started_at,
                "finished_at": finished_at,
                "counts": counts,
            }
        except Exception as exc:
            conn.rollback()
            raise exc
        finally:
            conn.close()

    def _load_dashboard_data(self) -> dict[str, Any]:
        if not self.dashboard_data_path.exists():
            raise FileNotFoundError(f"Dados consolidados do Painel do Operador não encontrados: {self.dashboard_data_path}")
        # Cache simples: reusa se arquivo não mudou e TTL < 60s
        try:
            mtime = self.dashboard_data_path.stat().st_mtime
        except OSError:
            mtime = 0.0
        now = time.monotonic()
        if (
            self._dashboard_cache is not None
            and mtime == self._dashboard_cache_mtime
            and (now - self._dashboard_cache_ts) < 60.0
        ):
            return self._dashboard_cache
        data = json.loads(self.dashboard_data_path.read_text("utf-8"))
        self._dashboard_cache = data
        self._dashboard_cache_mtime = mtime
        self._dashboard_cache_ts = now
        return data

    def _load_data_sources_config(self) -> dict[str, Any]:
        if not self.config_path.exists():
            return {"schemaVersion": 1, "workspaceRoot": str(self.module_root), "sources": []}
        return json.loads(self.config_path.read_text("utf-8"))

    def _normalize_data_source(self, source: dict[str, Any]) -> dict[str, Any]:
        paths = source.get("paths") or []
        if isinstance(paths, str):
            paths = [paths]
        return {
            "id": str(source.get("id") or ""),
            "label": str(source.get("label") or source.get("id") or ""),
            "description": str(source.get("description") or ""),
            "kind": str(source.get("kind") or "folder"),
            "recursive": bool(source.get("recursive", True)),
            "paths": [str(path) for path in paths if str(path or "").strip()],
        }

    def _validate_source_paths(
        self,
        source: dict[str, Any],
        cache: dict[tuple[str, bool], dict[str, Any]],
    ) -> dict[str, Any]:
        path_results = []
        total_files = 0
        total_size = 0
        existing_paths = 0
        extension_counts: dict[str, int] = {}
        samples: list[str] = []
        recursive = bool(source.get("recursive"))
        for raw_path in source.get("paths") or []:
            path = Path(str(raw_path)).expanduser()
            cache_key = (str(path), recursive)
            result = cache.get(cache_key)
            if result is None:
                result = self._scan_source_path(path, recursive=recursive)
                cache[cache_key] = result
            path_results.append({**result, "path": str(path)})
            if result.get("exists"):
                existing_paths += 1
            total_files += int(result.get("file_count") or 0)
            total_size += int(result.get("total_size_bytes") or 0)
            for extension, count in (result.get("extensions") or {}).items():
                extension_counts[extension] = extension_counts.get(extension, 0) + int(count or 0)
            for sample in result.get("samples") or []:
                if len(samples) < 8:
                    samples.append(sample)
        if not path_results:
            status = "missing_path"
        elif existing_paths == 0:
            status = "not_found"
        elif total_files == 0:
            status = "empty"
        else:
            status = "ok"
        return {
            "status": status,
            "paths_count": len(path_results),
            "existing_paths": existing_paths,
            "file_count": total_files,
            "total_size_bytes": total_size,
            "extensions": dict(sorted(extension_counts.items(), key=lambda item: (-item[1], item[0]))[:12]),
            "samples": samples,
            "paths": path_results,
        }

    def _scan_source_path(self, path: Path, *, recursive: bool) -> dict[str, Any]:
        exists = path.exists()
        is_dir = exists and path.is_dir()
        is_file = exists and path.is_file()
        if not exists:
            return {"exists": False, "is_dir": False, "is_file": False, "file_count": 0, "total_size_bytes": 0, "extensions": {}, "samples": []}
        if is_file:
            stat = path.stat()
            return {
                "exists": True,
                "is_dir": False,
                "is_file": True,
                "file_count": 1,
                "total_size_bytes": stat.st_size,
                "extensions": {path.suffix.lower() or "(sem extensão)": 1},
                "samples": [path.name],
            }
        file_count = 0
        total_size = 0
        extension_counts: dict[str, int] = {}
        samples: list[str] = []
        if recursive:
            items = self._iter_source_path_files(path)
        else:
            items = (item for item in path.glob("*") if item.is_file())
        for item in items:
            try:
                stat = item.stat()
            except OSError:
                continue
            file_count += 1
            total_size += stat.st_size
            extension = item.suffix.lower() or "(sem extensão)"
            extension_counts[extension] = extension_counts.get(extension, 0) + 1
            if len(samples) < 8:
                try:
                    samples.append(str(item.relative_to(path)))
                except ValueError:
                    samples.append(item.name)
        return {
            "exists": True,
            "is_dir": is_dir,
            "is_file": is_file,
            "file_count": file_count,
            "total_size_bytes": total_size,
            "extensions": dict(sorted(extension_counts.items(), key=lambda item: (-item[1], item[0]))[:12]),
            "samples": samples,
        }

    def _iter_source_path_files(self, root: Path):
        stack = [root]
        while stack:
            current = stack.pop()
            try:
                for child in current.iterdir():
                    if child.is_dir():
                        if child.name in self.INDEX_EXCLUDED_DIRS:
                            continue
                        stack.append(child)
                    elif child.is_file():
                        yield child
            except OSError:
                continue

    def _iter_index_files(self):
        stack = [self.module_root]
        while stack:
            current = stack.pop()
            try:
                for child in current.iterdir():
                    if child.is_dir():
                        if child.name in self.INDEX_EXCLUDED_DIRS:
                            continue
                        stack.append(child)
                    elif child.is_file():
                        yield child
            except OSError:
                continue

    def _insert_file_index_row(self, cur, row: dict[str, Any]) -> None:
        cur.execute(
            """
            INSERT INTO painel_operador_file_index(
                index_run_id, stable_key, relative_path, full_path, filename,
                extension, file_size_bytes, modified_at, file_hash, duplicate_key,
                is_duplicate, category, document_kind, source_group, inferred_date,
                inferred_tag, inferred_family, parse_priority, ignored, ignore_reason,
                payload_json, created_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                row["index_run_id"],
                row["stable_key"],
                row["relative_path"],
                row["full_path"],
                row["filename"],
                row["extension"],
                row["file_size_bytes"],
                row["modified_at"],
                row["file_hash"],
                row["duplicate_key"],
                row["is_duplicate"],
                row["category"],
                row["document_kind"],
                row["source_group"],
                row["inferred_date"],
                row["inferred_tag"],
                row["inferred_family"],
                row["parse_priority"],
                row["ignored"],
                row["ignore_reason"],
                row["payload_json"],
                self._now(),
            ),
        )

    def _classify_index_file(self, path: Path) -> dict[str, Any]:
        rel = str(path.relative_to(self.module_root))
        text = rel.replace("/", "\\")
        lower = text.lower()
        ext = path.suffix.lower()
        category = "other"
        document_kind = "unknown"
        source_group = "painel_operador"
        priority = "normal"
        ignored = 0
        ignore_reason = ""

        if ext in self.INDEX_LOW_VALUE_EXTENSIONS:
            ignored = 1
            ignore_reason = "Extensao tecnica sem valor operacional direto."
            priority = "ignore"
        elif "dashboard-anp-radar\\node_modules\\" in lower or "\\.git\\" in lower:
            ignored = 1
            ignore_reason = "Artefato tecnico excluido."
            priority = "ignore"

        if "fpsO-bacalhau_daily reports".lower() in lower:
            category = "daily_report"
            source_group = "daily_reports"
            if "\\01 - cv_reports\\" in lower:
                document_kind = self._cv_report_kind(path.name)
                priority = "high"
            elif "\\03 - ihm_reports\\" in lower:
                document_kind = "ihm_report"
                priority = "high"
            elif "\\05 - xml\\" in lower:
                document_kind = "anp_fiscal_xml" if ext == ".xml" else "anp_fiscal_archive"
                priority = "high"
        elif re.search(r"(^|\\)00[1-4]_", text, re.IGNORECASE) and ext in {".xml", ".zip"}:
            category = "anp_xml"
            document_kind = "anp_fiscal_xml" if ext == ".xml" else "anp_fiscal_archive"
            source_group = "anp_submission"
            priority = "high"
        elif re.search(r"(^|\\)0(39|40)_", text, re.IGNORECASE) and ext == ".xml":
            category = "anp_xml"
            document_kind = "anp_failure_or_bsw_xml"
            source_group = "anp_submission"
            priority = "high"
        elif path.name in {"Óleo Linear.xlsx", "Gás Linear.xlsx", "Gás Diferencial.xlsx", "BSW em Linha.xlsx", "Falha de Medição.xlsx"}:
            category = "anp_operator_export"
            document_kind = "anp_export"
            source_group = "painel_anp"
            priority = "high"
        elif ext == ".csv" and re.search(r"BAC_SUB|BAC_Fiscal|Metering Bacalhau", path.name, re.IGNORECASE):
            category = "pi_timeseries"
            document_kind = "pi_csv"
            source_group = "performance_monitoring"
            priority = "high"
        elif ext in {".xlsx", ".xlsm"} and re.search(r"Relação de CV|Calibration|Calibracao|Calibração|Validacao de Corrida|Validação|Checklist|MPFM_", text, re.IGNORECASE):
            category = "technical_workbook"
            document_kind = "calibration_or_registry_workbook"
            source_group = "technical_registry"
            priority = "high"
        elif ext in {".pdf", ".docx", ".rtf", ".msg"}:
            category = self._document_category(text)
            document_kind = "evidence_document"
            source_group = "document_evidence"
            priority = "normal"
        elif ext == ".zip":
            category = "archive"
            document_kind = "transport_or_source_archive"
            source_group = "archive"
            priority = "low"
        elif ext == ".xml":
            category = "technical_xml"
            document_kind = "technical_xml"
            source_group = "technical_registry"
            priority = "normal"
        elif ext in {".py", ".mjs", ".jsx", ".css", ".html", ".json", ".md", ".bat", ".ps1"} and "dashboard-anp-radar" in lower:
            category = "radar_app"
            document_kind = "app_source_or_config"
            source_group = "radar_anp"
            priority = "low"

        inferred_tag = self._infer_tag(text)
        if category == "daily_report" and ext == ".txt":
            content_tag = self._infer_tag_from_file(path)
            if content_tag and (not inferred_tag or inferred_tag.startswith("FC")):
                inferred_tag = content_tag
        inferred_date = self._infer_date(text)
        daily_folder_match = re.search(r"daily reports[_\s-]+(\d{4})-(\d{2})-(\d{2})", text, re.IGNORECASE)
        if category == "daily_report" and daily_folder_match:
            inferred_date = f"{daily_folder_match.group(1)}-{daily_folder_match.group(2)}-{daily_folder_match.group(3)}"

        return {
            "extension": ext,
            "category": category,
            "document_kind": document_kind,
            "source_group": source_group,
            "inferred_date": inferred_date,
            "inferred_tag": inferred_tag,
            "inferred_family": self._infer_family(text, path.name),
            "parse_priority": priority,
            "ignored": ignored,
            "ignore_reason": ignore_reason,
        }

    def _cv_report_kind(self, name: str) -> str:
        lower = name.lower()
        if "run_daily" in lower:
            return "cv_run_daily_txt"
        if "run_24hours" in lower:
            return "cv_run_24hours_txt"
        if "run_hourly" in lower:
            return "cv_run_hourly_txt"
        if "alarmsandevents" in lower:
            return "cv_alarm_event_txt"
        if lower == "parameters.xml":
            return "cv_parameters_xml"
        if lower == "security.xml":
            return "cv_security_xml"
        return "cv_report"

    def _document_category(self, text: str) -> str:
        lower = text.lower()
        if "memorial" in lower or "descritivo" in lower or "functional description" in lower:
            return "memorial_document"
        if "uncert" in lower or "incerteza" in lower or "ucg" in lower:
            return "uncertainty_document"
        if "calib" in lower or "cert" in lower or "labm" in lower or "rfl" in lower:
            return "calibration_certificate"
        if "book anp" in lower or "manual xml" in lower or "rtm" in lower or "portaria" in lower:
            return "regulatory_document"
        if "fiscal" in lower:
            return "fiscal_document"
        if "operational" in lower:
            return "operational_document"
        return "evidence_document"

    def _infer_date(self, text: str) -> str:
        match = self.ISO_DATE_RE.search(text)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        match = self.COMPACT_DATE_RE.search(text)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        match = self.BR_DATE_RE.search(text)
        if match:
            return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        return ""

    def _infer_tag(self, text: str) -> str:
        matches = [match.group(0).upper().replace("-", "_") for match in self.TAG_RE.finditer(text)]
        if not matches:
            return ""
        for token in matches:
            if "FT" in token:
                return token
        for token in matches:
            if token.startswith(("PE", "PW")):
                return token
        for token in matches:
            if "JN" in token:
                return token
        return matches[0]

    def _infer_tag_from_file(self, path: Path) -> str:
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")[:4096]
        except OSError:
            return ""
        return self._infer_tag(text)

    def _infer_family(self, text: str, filename: str) -> str:
        lower = text.lower()
        if re.search(r"(^|\\)001_", text):
            return "a001"
        if re.search(r"(^|\\)002_", text):
            return "a002"
        if re.search(r"(^|\\)003_", text):
            return "a003"
        if re.search(r"(^|\\)004_", text):
            return "a004"
        if re.search(r"(^|\\)039_", text):
            return "a039"
        if "óleo linear" in lower or "oleo linear" in lower:
            return "a001"
        if "gás linear" in lower or "gas linear" in lower:
            return "a002"
        if "gás diferencial" in lower or "gas diferencial" in lower:
            return "a003"
        if "falha de medição" in lower or "falha de medicao" in lower:
            return "a039"
        if "bsw" in lower:
            return "a040"
        return ""

    def _staging_tables(self) -> tuple[str, ...]:
        return (
            "painel_operador_sources",
            "painel_operador_measurement_points",
            "painel_operador_comparisons",
            "painel_operador_evidence",
            "painel_operador_alerts",
            "painel_operador_proposals",
            "painel_operador_calendar_days",
            "painel_operador_calendar_pendencies",
        )

    def _sync_sources(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        for idx, row in enumerate(data.get("files") or []):
            if not isinstance(row, dict):
                continue
            source_path = str(row.get("path") or "")
            local_path = self._localize_path(source_path)
            local_file = Path(local_path) if local_path else None
            exists = bool(local_file and local_file.exists() and local_file.is_file())
            file_size = local_file.stat().st_size if exists and local_file else 0
            file_hash = self._sha1_file(local_file) if exists and local_file else self._payload_hash(row)
            source_name = Path(local_path or source_path).name
            stable_key = self._stable_key("source", idx, row.get("date"), row.get("family"), source_path, source_name)
            cur.execute(
                """
                INSERT INTO painel_operador_sources(
                    sync_run_id, stable_key, source_kind, source_date, family, family_name,
                    source_path, local_path, source_name, records_count, file_exists,
                    file_size_bytes, file_hash, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    str(row.get("kind") or ""),
                    str(row.get("date") or ""),
                    str(row.get("family") or ""),
                    str(row.get("familyName") or ""),
                    source_path,
                    local_path,
                    source_name,
                    self._int_or_zero(row.get("records")),
                    1 if exists else 0,
                    file_size,
                    file_hash,
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_measurement_points(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        for idx, row in enumerate(data.get("latestPoints") or []):
            if not isinstance(row, dict):
                continue
            stable_key = self._stable_key("point", idx, row.get("date"), row.get("family"), row.get("tag"))
            cur.execute(
                """
                INSERT INTO painel_operador_measurement_points(
                    sync_run_id, stable_key, point_date, family, family_name, tag, fluid,
                    principal, secondary, meter_type, active_status, computador_vazao,
                    volume_corrigido, volume_bruto, volume_liquido, temperatura, pressao,
                    in_range, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    str(row.get("date") or ""),
                    str(row.get("family") or ""),
                    str(row.get("familyName") or ""),
                    str(row.get("tag") or ""),
                    str(row.get("fluid") or ""),
                    str(row.get("principal") or ""),
                    str(row.get("secondary") or ""),
                    str(row.get("meterType") or ""),
                    str(row.get("active") or ""),
                    str(row.get("computadorVazao") or ""),
                    self._float_or_none(row.get("volumeCorrigido")),
                    self._float_or_none(row.get("volumeBruto")),
                    self._float_or_none(row.get("volumeLiquido")),
                    self._float_or_none(row.get("temperatura")),
                    self._float_or_none(row.get("pressao")),
                    1 if row.get("inRange") else 0,
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_comparisons(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        for idx, row in enumerate(data.get("comparisons") or []):
            if not isinstance(row, dict):
                continue
            raw_source = str(row.get("rawSource") or "")
            xml_source = str(row.get("xmlSource") or "")
            stable_key = self._stable_key("comparison", idx, row.get("date"), row.get("family"), row.get("tag"), raw_source, xml_source)
            cur.execute(
                """
                INSERT INTO painel_operador_comparisons(
                    sync_run_id, stable_key, comparison_date, family, family_name, tag, fluid,
                    status, raw_ok, anp_ok, raw_corrigido, xml_corrigido, anp_corrigido,
                    raw_source, raw_source_local, xml_source, xml_source_local, note,
                    payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    str(row.get("date") or ""),
                    str(row.get("family") or ""),
                    str(row.get("familyName") or ""),
                    str(row.get("tag") or ""),
                    str(row.get("fluid") or ""),
                    str(row.get("status") or ""),
                    1 if row.get("rawOk") else 0,
                    1 if row.get("anpOk") else 0,
                    self._float_or_none(row.get("rawCorrigido")),
                    self._float_or_none(row.get("xmlCorrigido")),
                    self._float_or_none(row.get("anpCorrigido")),
                    raw_source,
                    self._localize_path(raw_source),
                    xml_source,
                    self._localize_path(xml_source),
                    "" if row.get("note") is None else str(row.get("note")),
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_evidence(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        radar = data.get("eventEvidenceRadar") if isinstance(data.get("eventEvidenceRadar"), dict) else {}
        for idx, row in enumerate(radar.get("events") or []):
            if not isinstance(row, dict):
                continue
            source_path = str(row.get("source") or "")
            stable_key = self._stable_key("event", idx, row.get("timestamp"), row.get("message"), source_path)
            cur.execute(
                """
                INSERT INTO painel_operador_evidence(
                    sync_run_id, stable_key, evidence_kind, event_at, requirement_id,
                    title, status, evidence_state, source_path, local_path, target_id,
                    target_type, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    "event",
                    str(row.get("timestamp") or ""),
                    "",
                    str(row.get("message") or "")[:500],
                    str(row.get("status") or ""),
                    str(row.get("evidenceState") or ""),
                    source_path,
                    self._localize_path(source_path),
                    ",".join(str(item) for item in (row.get("tags") or [])),
                    str(row.get("parameter") or ""),
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1

        matrix = data.get("regulatoryMatrix") if isinstance(data.get("regulatoryMatrix"), dict) else {}
        for idx, row in enumerate(matrix.get("rows") or []):
            if not isinstance(row, dict):
                continue
            requirement_id = str(row.get("ID") or row.get("id") or "")
            title = str(row.get("Requisito / Atividade") or row.get("title") or "")
            stable_key = self._stable_key("requirement", idx, requirement_id, title)
            cur.execute(
                """
                INSERT INTO painel_operador_evidence(
                    sync_run_id, stable_key, evidence_kind, event_at, requirement_id,
                    title, status, evidence_state, source_path, local_path, target_id,
                    target_type, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    "requirement",
                    "",
                    requirement_id,
                    title,
                    str(row.get("Categoria") or ""),
                    "matrix",
                    str(matrix.get("source") or ""),
                    self._localize_path(str(matrix.get("source") or "")),
                    str(row.get("Sistema Aplicável") or ""),
                    str(row.get("Subcategoria") or ""),
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_alerts(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        for idx, row in enumerate(data.get("alerts") or []):
            if not isinstance(row, dict):
                continue
            stable_key = self._stable_key("alert", idx, row.get("date"), row.get("title"), row.get("area"))
            cur.execute(
                """
                INSERT INTO painel_operador_alerts(
                    sync_run_id, stable_key, alert_kind, severity, alert_date, title,
                    detail, area, target_id, status, source_path, local_path,
                    payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    "alert",
                    str(row.get("severity") or ""),
                    str(row.get("date") or ""),
                    str(row.get("title") or ""),
                    str(row.get("detail") or ""),
                    str(row.get("area") or ""),
                    "",
                    "",
                    "",
                    "",
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1

        for idx, row in enumerate(data.get("changeProposals") or []):
            if not isinstance(row, dict):
                continue
            source_path = str(row.get("sourcePath") or "")
            stable_key = self._stable_key("proposal", idx, row.get("id"), row.get("title"), row.get("targetId"))
            cur.execute(
                """
                INSERT INTO painel_operador_alerts(
                    sync_run_id, stable_key, alert_kind, severity, alert_date, title,
                    detail, area, target_id, status, source_path, local_path,
                    payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    "proposal",
                    str(row.get("risk") or ""),
                    str(row.get("createdAt") or "")[:10],
                    str(row.get("title") or ""),
                    str(row.get("recommendedAction") or row.get("evidenceText") or ""),
                    str(row.get("domain") or ""),
                    str(row.get("targetId") or ""),
                    str(row.get("status") or ""),
                    source_path,
                    self._localize_path(source_path),
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_proposals(self, cur, sync_run_id: int, data: dict[str, Any]) -> int:
        count = 0
        for idx, row in enumerate(data.get("changeProposals") or []):
            if not isinstance(row, dict):
                continue
            source_path = str(row.get("sourcePath") or "")
            proposal_id = str(row.get("id") or "")
            stable_key = self._stable_key("proposal", proposal_id, idx, row.get("title"), row.get("targetId"))
            decision = self._latest_decision_audit(cur, "proposal", proposal_id)
            status = str(row.get("status") or "")
            authorized_by = str(row.get("authorizedBy") or "")
            authorized_at = str(row.get("authorizedAt") or "")
            decision_note = str(row.get("decisionNote") or "")
            audit_trail = row.get("auditTrail") or []
            if decision:
                status = str(decision.get("decision_status") or status)
                authorized_by = str(decision.get("decided_by") or authorized_by)
                authorized_at = str(decision.get("decided_at") or authorized_at)
                decision_note = str(decision.get("decision_note") or decision_note)
                if not isinstance(audit_trail, list):
                    audit_trail = []
                audit_trail = [
                    *audit_trail,
                    {
                        "status": status,
                        "decisionMode": decision.get("decision_mode") or "",
                        "decidedBy": authorized_by,
                        "decidedAt": authorized_at,
                        "decisionNote": decision_note,
                        "reappliedFromAudit": True,
                    },
                ]
            cur.execute(
                """
                INSERT INTO painel_operador_proposals(
                    sync_run_id, stable_key, proposal_id, proposal_kind, domain,
                    title, target_type, target_id, field_name, current_value_json,
                    proposed_value_json, unit, confidence, risk, status,
                    requires_approval, source_type, source_path, local_path,
                    source_name, evidence_state, evidence_text, recommended_action,
                    created_at_source, authorized_by, authorized_at, decision_note,
                    audit_trail_json, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    proposal_id,
                    str(row.get("kind") or ""),
                    str(row.get("domain") or ""),
                    str(row.get("title") or ""),
                    str(row.get("targetType") or ""),
                    str(row.get("targetId") or ""),
                    str(row.get("field") or ""),
                    self._json(row.get("currentValue")),
                    self._json(row.get("proposedValue")),
                    "" if row.get("unit") is None else str(row.get("unit")),
                    str(row.get("confidence") or ""),
                    str(row.get("risk") or ""),
                    status,
                    1 if row.get("requiresApproval") else 0,
                    str(row.get("sourceType") or ""),
                    source_path,
                    self._localize_path(source_path),
                    str(row.get("sourceName") or ""),
                    str(row.get("evidenceState") or ""),
                    str(row.get("evidenceText") or ""),
                    str(row.get("recommendedAction") or ""),
                    str(row.get("createdAt") or ""),
                    authorized_by,
                    authorized_at,
                    decision_note,
                    self._json(audit_trail),
                    self._json(row),
                    self._now(),
                ),
            )
            count += 1
        return count

    def _sync_calendar(self, cur, sync_run_id: int, data: dict[str, Any]) -> dict[str, int]:
        calendar = data.get("operationalCalendar") if isinstance(data.get("operationalCalendar"), dict) else {}
        days_count = 0
        pendencies_count = 0
        for idx, row in enumerate(calendar.get("days") or []):
            if not isinstance(row, dict):
                continue
            calendar_date = str(row.get("date") or "")
            stable_key = self._stable_key("calendar-day", idx, calendar_date)
            cur.execute(
                """
                INSERT INTO painel_operador_calendar_days(
                    sync_run_id, stable_key, calendar_date, day_number, loaded,
                    status, closing_status, points_count, xml_families_json,
                    package_families_json, missing_xml_families_json, raw_pending,
                    anp_pending, alert_count, pending_count, open_pending_count,
                    resolved_pending_count, payload_json, created_at
                )
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    sync_run_id,
                    stable_key,
                    calendar_date,
                    self._int_or_zero(row.get("day")),
                    1 if row.get("loaded") else 0,
                    str(row.get("status") or ""),
                    str(row.get("closingStatus") or ""),
                    self._int_or_zero(row.get("points")),
                    self._json(row.get("xmlFamilies") or []),
                    self._json(row.get("packageFamilies") or []),
                    self._json(row.get("missingXmlFamilies") or []),
                    self._int_or_zero(row.get("rawPending")),
                    self._int_or_zero(row.get("anpPending")),
                    self._int_or_zero(row.get("alertCount")),
                    self._int_or_zero(row.get("pendingCount")),
                    self._int_or_zero(row.get("openPendingCount")),
                    self._int_or_zero(row.get("resolvedPendingCount")),
                    self._json(row),
                    self._now(),
                ),
            )
            days_count += 1

            for item_idx, item in enumerate(row.get("pendingItems") or []):
                if not isinstance(item, dict):
                    continue
                pendency_id = str(item.get("id") or "")
                pendency_key = self._stable_key("calendar-pendency", pendency_id, calendar_date, item_idx, item.get("title"))
                decision = self._latest_decision_audit(cur, "calendar_pendency", pendency_id)
                status = str(item.get("status") or "")
                resolution_mode = str(item.get("resolutionMode") or "")
                closed_by = str(item.get("closedBy") or "")
                closed_at = str(item.get("closedAt") or "")
                decision_note = str(item.get("decisionNote") or "")
                if decision:
                    status = str(decision.get("decision_status") or status)
                    resolution_mode = str(decision.get("decision_mode") or resolution_mode)
                    closed_by = str(decision.get("decided_by") or closed_by)
                    closed_at = str(decision.get("decided_at") or closed_at)
                    decision_note = str(decision.get("decision_note") or decision_note)
                cur.execute(
                    """
                    INSERT INTO painel_operador_calendar_pendencies(
                        sync_run_id, stable_key, pendency_id, calendar_date,
                        pendency_type, severity, status, title, detail,
                        recommended_action, resolution_mode, closed_by,
                        closed_at, decision_note, payload_json, created_at
                    )
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        sync_run_id,
                        pendency_key,
                        pendency_id,
                        calendar_date,
                        str(item.get("type") or ""),
                        str(item.get("severity") or ""),
                        status,
                        str(item.get("title") or ""),
                        str(item.get("detail") or ""),
                        str(item.get("recommendedAction") or ""),
                        resolution_mode,
                        closed_by,
                        closed_at,
                        decision_note,
                        self._json(item),
                        self._now(),
                    ),
                )
                pendencies_count += 1
        return {"days": days_count, "pendencies": pendencies_count}

    def _parse_anp_export_row(
        self,
        filename: str,
        meta: dict[str, str],
        path: Path,
        sheet_name: str,
        row_number: int,
        row: tuple[Any, ...],
        header_map: dict[str, int],
    ) -> dict[str, Any]:
        def get(*names: str) -> Any:
            for name in names:
                idx = header_map.get(self._normalize_header(name))
                if idx is not None and idx < len(row):
                    return row[idx]
            return None

        period_start = self._date_text(get("Início Período Medição", "Inicio Período Medição"))
        period_end = self._date_text(get("Fim Período Medição"))
        event_at = self._date_text(get("Data & Hora Medição", "Data & Hora da Ocorrência"))
        detected_at = self._date_text(get("Data & Hora da Detecção"))
        returned_at = self._date_text(get("Data & Hora do Retorno"))
        collection_at = self._date_text(get("Data e Hora da Coleta"))
        file_loaded_at = self._date_text(get("Data de Carga do Arquivo"))
        reference_date = period_start or event_at or detected_at or returned_at or file_loaded_at
        tag = self._text_value(get("Tag do Ponto Medição", "Tag do ponto", "Tag do Ponto"))
        element_tag = self._text_value(get("Tag do Elemento de Medição"))

        payload = {
            "source_file": filename,
            "sheet_name": sheet_name,
            "row_number": row_number,
            "headers": list(header_map.keys()),
            "values": {key: self._jsonable_cell(row[idx] if idx < len(row) else None) for key, idx in header_map.items()},
        }
        stable_key = self._stable_key(
            "anp-export",
            filename,
            row_number,
            meta["family"],
            tag,
            period_start,
            event_at,
            self._text_value(get("Código da Falha")),
        )
        return {
            "stable_key": stable_key,
            "source_file": filename,
            "source_path": str(path),
            "sheet_name": sheet_name,
            "row_number": row_number,
            "family": meta["family"],
            "export_name": meta["export_name"],
            "record_kind": meta["record_kind"],
            "installation": self._text_value(get("Instalação")),
            "installation_code": self._text_value(get("Código da Instalação", "Código Instalação", "Instalação")),
            "tag": tag,
            "element_tag": element_tag,
            "serial_number": self._text_value(get("Nº de Série", "No de Série", "N de Série")),
            "reference_date": reference_date[:10],
            "period_start": period_start,
            "period_end": period_end,
            "event_at": event_at,
            "detected_at": detected_at,
            "returned_at": returned_at,
            "collection_at": collection_at,
            "file_loaded_at": file_loaded_at,
            "volume_corrigido": self._float_or_none(get("Volume Bruto Corrigido (m3)")),
            "volume_bruto": self._float_or_none(get("Volume Bruto (m3)")),
            "volume_liquido": self._float_or_none(get("Volume Liquido (m3)", "Volume Líquido (m3)")),
            "bsw_percent": self._float_or_none(get("% BSW")),
            "bsw_max_percent": self._float_or_none(get("% Máximo BSW")),
            "pressure_kpa": self._float_or_none(get("Pressão Estática (kPa)", "Pressão Estatica (kPa)")),
            "temperature_c": self._float_or_none(get("Temperatura Fluido  (°C)", "Temperatura (°C)")),
            "failure_code": self._text_value(get("Código da Falha")),
            "notification_type": self._text_value(get("Tipo de Notificação")),
            "failure_type": self._text_value(get("Tipo de Falha")),
            "received_file": self._text_value(get("Arquivo Recebido")),
            "payload_json": self._json(payload),
        }

    def _insert_anp_export_row(self, cur, import_run_id: int, row: dict[str, Any]) -> None:
        cur.execute(
            """
            INSERT INTO painel_operador_anp_export_rows(
                import_run_id, stable_key, source_file, source_path, sheet_name,
                row_number, family, export_name, record_kind, installation,
                installation_code, tag, element_tag, serial_number, reference_date,
                period_start, period_end, event_at, detected_at, returned_at,
                collection_at, file_loaded_at, volume_corrigido, volume_bruto,
                volume_liquido, bsw_percent, bsw_max_percent, pressure_kpa,
                temperature_c, failure_code, notification_type, failure_type,
                received_file, payload_json, created_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                import_run_id,
                row["stable_key"],
                row["source_file"],
                row["source_path"],
                row["sheet_name"],
                row["row_number"],
                row["family"],
                row["export_name"],
                row["record_kind"],
                row["installation"],
                row["installation_code"],
                row["tag"],
                row["element_tag"],
                row["serial_number"],
                row["reference_date"],
                row["period_start"],
                row["period_end"],
                row["event_at"],
                row["detected_at"],
                row["returned_at"],
                row["collection_at"],
                row["file_loaded_at"],
                row["volume_corrigido"],
                row["volume_bruto"],
                row["volume_liquido"],
                row["bsw_percent"],
                row["bsw_max_percent"],
                row["pressure_kpa"],
                row["temperature_c"],
                row["failure_code"],
                row["notification_type"],
                row["failure_type"],
                row["received_file"],
                row["payload_json"],
                self._now(),
            ),
        )

    def _file_info(self, path: Path) -> dict[str, Any]:
        if not path.exists():
            return {"exists": False, "path": str(path), "size_bytes": 0, "modified_at": ""}
        stat = path.stat()
        return {
            "exists": True,
            "path": str(path),
            "size_bytes": stat.st_size,
            "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds"),
        }

    def _sample_keys(self, rows: list[Any]) -> list[str]:
        for row in rows:
            if isinstance(row, dict):
                return list(row.keys())[:50]
        return []

    def _normalize_limit(self, value: int | None) -> int:
        try:
            parsed = int(value if value is not None else self.DEFAULT_MAX_LIST_ITEMS)
        except (TypeError, ValueError):
            parsed = self.DEFAULT_MAX_LIST_ITEMS
        return max(1, min(parsed, 2000))

    def _normalize_query_limit(self, value: int | None) -> int:
        try:
            parsed = int(value if value is not None else 100)
        except (TypeError, ValueError):
            parsed = 100
        return max(1, min(parsed, 500))

    def _bounded_value(self, value: Any, limit: int) -> tuple[Any, bool]:
        if isinstance(value, list) and len(value) > limit:
            return value[:limit], True
        return value, False

    def _serialize_staging_row(self, row: dict[str, Any], *, include_payload: bool) -> dict[str, Any]:
        if include_payload:
            raw = row.get("payload_json")
            try:
                row["payload"] = json.loads(raw) if raw else {}
            except json.JSONDecodeError:
                row["payload"] = {}
        row.pop("payload_json", None)
        return row

    def _localize_path(self, raw_path: str) -> str:
        if not raw_path:
            return ""
        text = str(raw_path)
        old_roots = (
            r"C:\Users\mauri\OneDrive\Documentos\Painel_Operador",
            r"C:\Users\mauri\OneDrive\Painel_Operador",
        )
        for old_root in old_roots:
            if text.startswith(old_root):
                suffix = text[len(old_root) :].lstrip("\\/")
                return str(self.module_root / suffix)
        return text

    def _sha1_file(self, path: Path) -> str:
        digest = hashlib.sha1()
        with Path(path).open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _payload_hash(self, value: Any) -> str:
        return hashlib.sha1(self._json(value).encode("utf-8")).hexdigest()

    def _stable_key(self, *parts: Any) -> str:
        clean = "|".join(str(part or "") for part in parts)
        return hashlib.sha1(clean.encode("utf-8")).hexdigest()

    def _json(self, value: Any) -> str:
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)

    def _loads_json(self, value: Any, default: Any = None) -> Any:
        if value in (None, ""):
            return {} if default is None else default
        try:
            return json.loads(value) if isinstance(value, str) else value
        except Exception:
            return {} if default is None else default

    def _insert_decision_audit(
        self,
        cur,
        *,
        target_type: str,
        target_id: str,
        target_db_id: int,
        previous_status: str,
        decision_status: str,
        decision_mode: str,
        decided_by: str,
        decided_at: str,
        decision_note: str,
        payload: dict[str, Any] | None = None,
    ) -> None:
        cur.execute(
            """
            INSERT INTO painel_operador_decision_audit(
                target_type, target_id, target_db_id, previous_status,
                decision_status, decision_mode, decided_by, decided_at,
                decision_note, payload_json, created_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                target_type,
                target_id,
                target_db_id,
                previous_status,
                decision_status,
                decision_mode,
                decided_by,
                decided_at,
                decision_note,
                self._json(payload or {}),
                self._now(),
            ),
        )

    def _latest_decision_audit(self, cur, target_type: str, target_id: str) -> dict[str, Any]:
        if not target_id or not self._table_exists(cur, "painel_operador_decision_audit"):
            return {}
        row = cur.execute(
            """
            SELECT *
            FROM painel_operador_decision_audit
            WHERE target_type=? AND target_id=?
            ORDER BY decided_at DESC, id DESC
            LIMIT 1
            """,
            (target_type, target_id),
        ).fetchone()
        return dict(row) if row else {}

    def _now(self) -> str:
        return datetime.now().isoformat(timespec="seconds")

    def _table_exists(self, cur, table: str) -> bool:
        return bool(cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone())

    def _normalize_header(self, value: Any) -> str:
        text = unicodedata.normalize("NFKD", str(value or ""))
        text = "".join(char for char in text if not unicodedata.combining(char))
        text = re.sub(r"[^a-zA-Z0-9]+", "_", text.lower()).strip("_")
        return text

    def _jsonable_cell(self, value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        return value

    def _text_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        return str(value).strip()

    def _date_text(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        if isinstance(value, datetime):
            return value.isoformat(timespec="seconds")
        text = str(value).strip()
        return text

    def _int_or_zero(self, value: Any) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    def _float_or_none(self, value: Any) -> float | None:
        try:
            parsed = float(value)
            return parsed if parsed == parsed else None
        except (TypeError, ValueError):
            return None
