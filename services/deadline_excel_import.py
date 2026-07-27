from __future__ import annotations

from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


DEFAULT_RANP44_DEADLINES_PATH = (
    r"C:\Users\mauri\Downloads\controle_prazos_RANP44_MPFM_v5_1_corrigido_tabelas.xlsx"
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _iso_date(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and 20000 <= float(value) <= 70000:
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return ""
    raw = str(value).strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _int_or_zero(value: Any) -> int:
    if value is None or value == "":
        return 0
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def _add_days(iso_value: str, days: int) -> str:
    if not iso_value or not days:
        return ""
    try:
        return (datetime.strptime(iso_value, "%Y-%m-%d").date() + timedelta(days=days)).isoformat()
    except ValueError:
        return ""


def _row_maps(ws, header_row: int, start_row: int):
    headers = [_text(cell.value) for cell in ws[header_row]]
    for excel_row in range(start_row, ws.max_row + 1):
        values = [ws.cell(excel_row, col + 1).value for col in range(len(headers))]
        if not any(_text(v) for v in values):
            continue
        yield excel_row, {headers[idx]: values[idx] for idx in range(len(headers)) if headers[idx]}


def _notes(*parts: str) -> str:
    return " | ".join(part for part in (_text(p) for p in parts) if part)


def _base_item(path: Path, source_ref: str, **overrides) -> dict:
    item = {
        "subject": "",
        "category": "RANP 44",
        "start_date": "",
        "due_date": "",
        "periodicity": "custom",
        "periodicity_days": 0,
        "notes": "",
        "icon": "deadlines",
        "source_ref": source_ref,
        "source_file": str(path),
        "norm_ref": "",
        "evidence_required": "",
        "responsible_area": "Medição fiscal / MPFM",
        "trigger_event": "",
        "risk_level": "",
        "recommended_action": "",
        "completion_date": "",
        "source_status": "",
    }
    item.update(overrides)
    return item


def build_ranp44_deadline_items(workbook_path: str | Path) -> tuple[list[dict], dict]:
    path = Path(workbook_path or DEFAULT_RANP44_DEADLINES_PATH)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    wb = load_workbook(path, data_only=True, read_only=False)
    items: list[dict] = []
    sheets_seen = [{"name": ws.title, "rows": ws.max_row, "cols": ws.max_column} for ws in wb.worksheets]
    skipped: list[str] = []

    if "01_Base_RANP44" in wb.sheetnames:
        ws = wb["01_Base_RANP44"]
        for _, row in _row_maps(ws, 1, 2):
            req_id = _text(row.get("ID requisito"))
            obligation = _text(row.get("Obrigação"))
            if not req_id or not obligation:
                continue
            days = _int_or_zero(row.get("Prazo"))
            unit = _text(row.get("Unidade"))
            items.append(
                _base_item(
                    path,
                    f"01_Base_RANP44:{req_id}",
                    subject=obligation,
                    category="Matriz normativa RANP 44",
                    periodicity_days=days,
                    norm_ref=_notes(req_id, _text(row.get("Item RANP 44"))),
                    trigger_event=_text(row.get("Data-base normativa")),
                    recommended_action=_text(row.get("Atividade controlada")),
                    evidence_required=_text(row.get("Evidência esperada")),
                    notes=_notes(
                        f"Prazo normativo: {days} {unit}" if days else f"Prazo normativo: {unit or 'sem numero fixo'}",
                        row.get("Comentário de aplicação"),
                    ),
                    source_status="Template normativo",
                )
            )

    if "03_Comissionamento" in wb.sheetnames:
        ws = wb["03_Comissionamento"]
        for _, row in _row_maps(ws, 1, 2):
            row_id = _text(row.get("ID"))
            tag = _text(row.get("TAG"))
            if not row_id or not tag or row_id.startswith("="):
                continue
            d0 = _iso_date(row.get("D0 conservador (ANP/ofício)"))
            due_60 = _iso_date(row.get("Deadline 60d conservador")) or _add_days(d0, 60)
            if not due_60:
                skipped.append(f"{row_id}: sem prazo 60d calculável")
                continue
            completion_60 = _iso_date(row.get("Data conclusão / relatório final"))
            items.append(
                _base_item(
                    path,
                    f"03_Comissionamento:{row_id}:60d",
                    subject=f"{tag} - conclusão/pós-comissionamento 60d",
                    category="Comissionamento RANP 44",
                    start_date=d0,
                    due_date=due_60,
                    periodicity_days=60,
                    norm_ref="RANP 44 item 9.3",
                    trigger_event=_text(row.get("Evento controlado")),
                    risk_level=_text(row.get("Risco regulatório")),
                    recommended_action=_text(row.get("Próxima ação automática")),
                    evidence_required=_text(row.get("Evidência")) or "Relatório final/pós-comissionamento",
                    completion_date=completion_60,
                    source_status=_text(row.get("Status 60d conservador")),
                    notes=_notes(
                        row.get("Poço/Riser"),
                        row.get("Ofício ANP"),
                        row.get("Fonte/justificativa D0 conservador"),
                        f"Dias extrapolados 60d: {_text(row.get('Dias extrap. 60d conservador'))}",
                        row.get("Observações / tese / lacuna"),
                    ),
                )
            )

    if "04_Ciclos_30d" in wb.sheetnames:
        ws = wb["04_Ciclos_30d"]
        for _, row in _row_maps(ws, 1, 2):
            cycle_id = _text(row.get("ID ciclo"))
            tag = _text(row.get("TAG"))
            if not cycle_id or not tag or cycle_id.startswith("="):
                continue
            d0 = _iso_date(row.get("D0 conservador"))
            due_30 = _iso_date(row.get("Deadline 30d conservador")) or _add_days(d0, 30)
            if not due_30:
                skipped.append(f"{cycle_id}: sem prazo 30d calculável")
                continue
            evidence_date = _iso_date(row.get("Data envio/evidência"))
            items.append(
                _base_item(
                    path,
                    f"04_Ciclos_30d:{cycle_id}",
                    subject=f"{tag} - relatório de desempenho 30d ciclo {_text(row.get('Ciclo nº')) or '1'}",
                    category="Ciclo 30d RANP 44",
                    start_date=d0,
                    due_date=due_30,
                    periodicity_days=30,
                    norm_ref="RANP 44 item 7.3.1",
                    trigger_event="Início de operação/comissionamento",
                    risk_level=_text(row.get("Risco")),
                    recommended_action=_text(row.get("Ação automática")),
                    evidence_required=_text(row.get("Evidência usada")) or "Relatório de avaliação de desempenho",
                    completion_date=evidence_date,
                    source_status=_text(row.get("Status")),
                    notes=_notes(
                        row.get("Sistema"),
                        f"Dias extrapolados: {_text(row.get('Dias extrap.'))}",
                        row.get("Conclusão automática"),
                        row.get("Interpretação"),
                    ),
                )
            )

    if "05_Auto_Defesa" in wb.sheetnames:
        ws = wb["05_Auto_Defesa"]
        rows = {row.get("ID"): row for _, row in _row_maps(ws, 1, 2)}
        ai2 = rows.get("AI-002") or {}
        ai3 = rows.get("AI-003") or {}
        if ai2 or ai3:
            start = _iso_date(ai2.get("Data")) or _iso_date(ai3.get("Data"))
            due = ""
            content = _text(ai3.get("Conteúdo relevante"))
            if "08/07/2026" in content:
                due = "2026-07-08"
            items.append(
                _base_item(
                    path,
                    "05_Auto_Defesa:AI-002",
                    subject="Defesa Ofício 677/2026 - auto/DF 6077305",
                    category="Defesa regulatória",
                    start_date=start,
                    due_date=due,
                    periodicity_days=15,
                    norm_ref=_text(ai2.get("Base legal/requisito")) or "Decreto 2.953/1999",
                    trigger_event=_text(ai2.get("Identificação")),
                    risk_level="Alto",
                    recommended_action="Confirmar data real de recebimento e controlar defesa",
                    evidence_required=_text(ai2.get("Evidência")) or "DOC-011",
                    source_status=_text(ai2.get("Gravidade/prazo")),
                    notes=_notes(ai2.get("Conteúdo relevante"), ai2.get("Observação"), ai3.get("Observação")),
                )
            )

    summary = {
        "path": str(path),
        "sheets_seen": sheets_seen,
        "items_built": len(items),
        "skipped": skipped,
        "by_category": {},
    }
    for item in items:
        category = item.get("category") or "Sem categoria"
        summary["by_category"][category] = summary["by_category"].get(category, 0) + 1
    return items, summary


def import_ranp44_deadlines(db_conn, workbook_path: str | Path, dry_run: bool = False) -> dict:
    items, summary = build_ranp44_deadline_items(workbook_path)
    if dry_run:
        return {**summary, "dry_run": True, "inserted": 0, "updated": 0, "examples": items[:5]}

    conn = db_conn()
    cur = conn.cursor()
    now = datetime.now().isoformat(timespec="seconds")
    inserted = 0
    updated = 0
    for item in items:
        existing = cur.execute(
            """
            SELECT id FROM deadline_items
            WHERE source_file=? AND source_ref=? AND COALESCE(is_active,1)=1
            """,
            (item["source_file"], item["source_ref"]),
        ).fetchone()
        vals = (
            item["subject"],
            item["category"],
            item["start_date"],
            item["due_date"],
            item["periodicity"],
            int(item["periodicity_days"] or 0),
            item["notes"],
            item["icon"],
            item["source_ref"],
            item["source_file"],
            item["norm_ref"],
            item["evidence_required"],
            item["responsible_area"],
            item["trigger_event"],
            item["risk_level"],
            item["recommended_action"],
            item["completion_date"],
            item["source_status"],
        )
        if existing:
            cur.execute(
                """
                UPDATE deadline_items
                SET subject=?, category=?, start_date=?, due_date=?, periodicity=?, periodicity_days=?,
                    notes=?, icon=?, source_ref=?, source_file=?, norm_ref=?, evidence_required=?,
                    responsible_area=?, trigger_event=?, risk_level=?, recommended_action=?,
                    completion_date=?, source_status=?, updated_at=?
                WHERE id=?
                """,
                vals + (now, existing["id"]),
            )
            updated += 1
        else:
            cur.execute(
                """
                INSERT INTO deadline_items(
                    subject, category, start_date, due_date, periodicity, periodicity_days,
                    notes, icon, source_ref, source_file, norm_ref, evidence_required,
                    responsible_area, trigger_event, risk_level, recommended_action,
                    completion_date, source_status, is_active, created_at, updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?,?)
                """,
                vals + (now, now),
            )
            inserted += 1
    conn.commit()
    conn.close()
    return {**summary, "dry_run": False, "inserted": inserted, "updated": updated, "examples": items[:5]}
