from __future__ import annotations

import json
import math
import os
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter

SHEET_DAILY = "DIARIO"
SHEET_HOURLY = "HORARIO"
SHEET_INFO = "INSTRUCOES"

DIMENSION_COLUMNS = ["data", "hora", "banco", "loop", "tipo_medidor", "tag"]
VALUE_COLUMNS = [
    "mpfm_uncorr_gas_t",
    "mpfm_uncorr_oleo_t",
    "mpfm_uncorr_hc_t",
    "mpfm_uncorr_agua_t",
    "mpfm_uncorr_total_t",
    "mpfm_corr_gas_t",
    "mpfm_corr_oleo_t",
    "mpfm_corr_hc_t",
    "mpfm_corr_agua_t",
    "mpfm_corr_total_t",
    "pvt_mass_gas_t",
    "pvt_mass_oleo_t",
    "pvt_mass_agua_t",
    "pvt20_mass_gas_t",
    "pvt20_mass_oleo_t",
    "pvt20_mass_agua_t",
    "pvt20_vol_gas_sm3",
    "pvt20_vol_oleo_m3",
    "pvt20_vol_agua_m3",
]
ADJUSTMENT_COLUMNS = ["ajustar", "motivo_ajuste", "responsavel_ajuste", "observacao_ajuste"]
EXPORT_COLUMNS = DIMENSION_COLUMNS + VALUE_COLUMNS + ADJUSTMENT_COLUMNS

METRIC_MAP = {
    "mpfm_uncorr_gas_t": "MPFM uncorr Gás (t)",
    "mpfm_uncorr_oleo_t": "MPFM uncorr Óleo (t)",
    "mpfm_uncorr_hc_t": "MPFM uncorr HC (t)",
    "mpfm_uncorr_agua_t": "MPFM uncorr Água (t)",
    "mpfm_uncorr_total_t": "MPFM uncorr Total (t)",
    "mpfm_corr_gas_t": "MPFM corr Gás (t)",
    "mpfm_corr_oleo_t": "MPFM corr Óleo (t)",
    "mpfm_corr_hc_t": "MPFM corr HC (t)",
    "mpfm_corr_agua_t": "MPFM corr Água (t)",
    "mpfm_corr_total_t": "MPFM corr Total (t)",
    "pvt_mass_gas_t": "PVT mass Gás (t)",
    "pvt_mass_oleo_t": "PVT mass Óleo (t)",
    "pvt_mass_agua_t": "PVT mass Água (t)",
    "pvt20_mass_gas_t": "PVT @20 mass Gás (t)",
    "pvt20_mass_oleo_t": "PVT @20 mass Óleo (t)",
    "pvt20_mass_agua_t": "PVT @20 mass Água (t)",
    "pvt20_vol_gas_sm3": "PVT @20 vol Gás (Sm³)",
    "pvt20_vol_oleo_m3": "PVT @20 vol Óleo (m³)",
    "pvt20_vol_agua_m3": "PVT @20 vol Água (m³)",
}
METRIC_TO_EXPORT = {value: key for key, value in METRIC_MAP.items()}

REQUIRED_IMPORT_COLUMNS = set(DIMENSION_COLUMNS + VALUE_COLUMNS)


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def normalize_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    raw = str(value).strip()
    if not raw or raw.lower() in {"nan", "nat", "none"}:
        return ""
    raw = raw.split(" ", 1)[0].replace(".", "/").replace("-", "/")
    for fmt in ("%Y/%m/%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return pd.to_datetime(value).strftime("%Y-%m-%d")
    except Exception:
        return ""


def normalize_hour(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    raw = str(value).strip()
    if not raw or raw.lower() in {"nan", "nat", "none", "day", "diario", "diário", "daily", "-"}:
        return None
    if ":" in raw:
        raw = raw.split(":", 1)[0]
    try:
        parsed = int(float(raw))
    except Exception:
        return None
    if parsed < 0 or parsed > 24:
        return None
    return 0 if parsed == 24 else parsed


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    raw = str(value).strip()
    if not raw or raw.lower() in {"nan", "nat", "none", "-"}:
        return None
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        parsed = float(raw)
    except Exception:
        return None
    return parsed if not math.isnan(parsed) else None


def is_truthy(value: Any) -> bool:
    raw = str(value or "").strip().lower()
    return raw in {"1", "true", "sim", "s", "yes", "y", "x", "ajustar", "corrigir", "ok"}


def _row_kind_for_sheet(sheet_name: str) -> str:
    return "hourly" if sheet_name.upper() == SHEET_HOURLY else "daily"


def _fetch_adjustment_rows(conn, date_from: str, date_to: str, bank: str = "", tag: str = "") -> dict[str, list[dict[str, Any]]]:
    sql = """
        SELECT day_ref, hour_ref, row_kind, bank, loop, tipo, tag, instrument, metric_name, metric_value
        FROM measurements_active
        WHERE day_ref BETWEEN ? AND ?
          AND row_kind IN ('daily','hourly')
    """
    params: list[Any] = [date_from, date_to]
    if bank:
        sql += " AND bank=?"
        params.append(bank)
    if tag:
        sql += " AND tag=?"
        params.append(tag)
    metric_placeholders = ",".join("?" for _ in METRIC_TO_EXPORT)
    sql += f" AND metric_name IN ({metric_placeholders})"
    params.extend(METRIC_TO_EXPORT.keys())
    sql += " ORDER BY day_ref, COALESCE(hour_ref,-1), bank, tag, metric_name"

    rows_by_kind: dict[str, dict[tuple[Any, ...], dict[str, Any]]] = {"daily": {}, "hourly": {}}
    for row in conn.execute(sql, params).fetchall():
        row_kind = str(row["row_kind"] or "")
        if row_kind not in rows_by_kind:
            continue
        hour_ref = row["hour_ref"] if row_kind == "hourly" else None
        key = (row["day_ref"], hour_ref, row["bank"] or "", row["loop"] or "", row["tipo"] or "", row["tag"] or "")
        current = rows_by_kind[row_kind].setdefault(
            key,
            {
                "data": row["day_ref"],
                "hora": "" if hour_ref is None else int(hour_ref),
                "banco": row["bank"] or "",
                "loop": row["loop"] or "",
                "tipo_medidor": row["tipo"] or "",
                "tag": row["tag"] or "",
                **{col: None for col in VALUE_COLUMNS},
                "ajustar": "",
                "motivo_ajuste": "",
                "responsavel_ajuste": "",
                "observacao_ajuste": "",
            },
        )
        export_col = METRIC_TO_EXPORT.get(row["metric_name"])
        if export_col:
            current[export_col] = row["metric_value"]
    return {kind: list(payload.values()) for kind, payload in rows_by_kind.items()}


def _style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    adjustment_fill = PatternFill("solid", fgColor="FFF2CC")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, col_name in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        cell.fill = adjustment_fill if col_name in ADJUSTMENT_COLUMNS else header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = 14
        if col_name in {"data", "motivo_ajuste", "observacao_ajuste"}:
            width = 22
        elif col_name in VALUE_COLUMNS:
            width = 18
        elif col_name in {"tag", "responsavel_ajuste"}:
            width = 20
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.column >= len(DIMENSION_COLUMNS) + 1 and cell.column <= len(DIMENSION_COLUMNS) + len(VALUE_COLUMNS):
                cell.number_format = "0.00000"
            if cell.column <= len(DIMENSION_COLUMNS):
                cell.protection = Protection(locked=True)
            else:
                cell.protection = Protection(locked=False)
    ws.protection.sheet = False


def _write_sheet(wb, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in EXPORT_COLUMNS])
    _style_sheet(ws)


def export_adjustment_workbook(db_conn_fn, date_from: str, date_to: str, bank: str = "", tag: str = "") -> str:
    conn = db_conn_fn()
    conn.row_factory = conn.row_factory
    if not date_to:
        date_to = conn.execute("SELECT MAX(day_ref) FROM measurements_active WHERE row_kind IN ('daily','hourly')").fetchone()[0] or ""
    if not date_from:
        date_from = date_to
    rows = _fetch_adjustment_rows(conn, date_from, date_to, bank, tag)
    conn.close()

    wb = openpyxl.Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    ws_info = wb.create_sheet(SHEET_INFO)
    info_rows = [
        ["Arquivo", "Registro de ajustes MPFM"],
        ["Janela", f"{date_from} a {date_to}"],
        ["Uso", "Edite somente as colunas de valores e marque ajustar=Sim para importar uma linha."],
        ["Diário", "Aba DIARIO: uma linha por data/banco/TAG."],
        ["Horário", "Aba HORARIO: uma linha por data/hora/banco/TAG."],
        ["Regra de importação", "Para cada célula alterada em linha marcada, a aplicação atualiza a métrica oficial existente e registra auditoria do valor anterior/novo."],
    ]
    for item in info_rows:
        ws_info.append(item)
    ws_info.column_dimensions["A"].width = 24
    ws_info.column_dimensions["B"].width = 120
    ws_info["A1"].font = Font(bold=True)

    _write_sheet(wb, SHEET_DAILY, rows["daily"])
    _write_sheet(wb, SHEET_HOURLY, rows["hourly"])

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_path = temp.name
    temp.close()
    try:
        wb.save(temp_path)
    except Exception:
        try:
            os.unlink(temp_path)
        except FileNotFoundError:
            pass
        raise
    return temp_path


def _ensure_adjustment_columns(conn) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS mpfm_adjustment_imports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            imported_at TEXT NOT NULL,
            source_file TEXT NOT NULL,
            source_hash TEXT DEFAULT '',
            author TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            rows_seen INTEGER DEFAULT 0,
            rows_marked INTEGER DEFAULT 0,
            metrics_changed INTEGER DEFAULT 0,
            status TEXT DEFAULT 'applied',
            summary_json TEXT DEFAULT '{}'
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS mpfm_adjustment_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            import_id INTEGER NOT NULL,
            day_ref TEXT NOT NULL,
            hour_ref INTEGER,
            row_kind TEXT NOT NULL,
            bank TEXT DEFAULT '',
            loop TEXT DEFAULT '',
            tipo TEXT DEFAULT '',
            tag TEXT DEFAULT '',
            metric_name TEXT NOT NULL,
            old_value REAL,
            new_value REAL,
            reason TEXT DEFAULT '',
            responsible TEXT DEFAULT '',
            observation TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            FOREIGN KEY(import_id) REFERENCES mpfm_adjustment_imports(id)
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mpfm_adjustment_items_lookup ON mpfm_adjustment_items(day_ref, row_kind, bank, tag, metric_name)")


def _read_adjustment_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=object)
    except ValueError:
        return []
    df.columns = [str(col).strip().lower() for col in df.columns]
    missing = sorted(REQUIRED_IMPORT_COLUMNS - set(df.columns))
    if missing:
        raise ValueError(f"Aba {sheet_name} sem colunas obrigatórias: {', '.join(missing)}")
    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        item = {col: row.get(col) for col in df.columns}
        if not normalize_date(item.get("data")):
            continue
        rows.append(item)
    return rows


def _current_value(conn, row_kind: str, day_ref: str, hour_ref: int | None, bank: str, tag: str, metric_name: str) -> tuple[int | None, float | None]:
    if hour_ref is None:
        hour_filter = "hour_ref IS NULL"
        params: list[Any] = [row_kind, day_ref, bank, tag, metric_name]
    else:
        hour_filter = "hour_ref=?"
        params = [row_kind, day_ref, hour_ref, bank, tag, metric_name]
    sql = f"""
        SELECT id, metric_value
        FROM measurements_curated
        WHERE row_kind=? AND day_ref=? AND {hour_filter}
          AND COALESCE(bank,'')=? AND COALESCE(tag,'')=? AND metric_name=?
          AND COALESCE(is_official,1)=1
        ORDER BY id DESC
        LIMIT 1
    """
    row = conn.execute(sql, params).fetchone()
    if not row:
        return None, None
    return int(row["id"]), float(row["metric_value"] or 0)


def import_adjustment_workbook(db_conn_fn, workbook_path: Path, author: str = "", notes: str = "", apply: bool = True, source_name: str = "") -> dict[str, Any]:
    display_name = Path(source_name or workbook_path.name).name
    daily_rows = _read_adjustment_sheet(workbook_path, SHEET_DAILY)
    hourly_rows = _read_adjustment_sheet(workbook_path, SHEET_HOURLY)
    source_hash = ""
    try:
        import hashlib
        source_hash = hashlib.sha256(workbook_path.read_bytes()).hexdigest()
    except Exception:
        pass

    conn = db_conn_fn()
    conn.row_factory = conn.row_factory
    _ensure_adjustment_columns(conn)
    rows_seen = len(daily_rows) + len(hourly_rows)
    rows_marked = 0
    changes: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []

    for sheet_name, rows in ((SHEET_DAILY, daily_rows), (SHEET_HOURLY, hourly_rows)):
        row_kind = _row_kind_for_sheet(sheet_name)
        for row in rows:
            if not is_truthy(row.get("ajustar")):
                continue
            rows_marked += 1
            day_ref = normalize_date(row.get("data"))
            hour_ref = None if row_kind == "daily" else normalize_hour(row.get("hora"))
            bank = str(row.get("banco") or "").strip().upper()
            loop = str(row.get("loop") or "").strip()
            tipo = str(row.get("tipo_medidor") or "").strip()
            tag = str(row.get("tag") or "").strip()
            if not day_ref or not bank or not tag:
                skipped.append({"sheet": sheet_name, "day_ref": day_ref, "bank": bank, "tag": tag, "reason": "chave incompleta"})
                continue
            if row_kind == "hourly" and hour_ref is None:
                skipped.append({"sheet": sheet_name, "day_ref": day_ref, "bank": bank, "tag": tag, "reason": "hora inválida"})
                continue
            for col, metric_name in METRIC_MAP.items():
                new_value = to_float(row.get(col))
                if new_value is None:
                    continue
                old_id, old_value = _current_value(conn, row_kind, day_ref, hour_ref, bank, tag, metric_name)
                if old_value is not None and round(float(old_value), 6) == round(float(new_value), 6):
                    continue
                changes.append(
                    {
                        "sheet": sheet_name,
                        "day_ref": day_ref,
                        "hour_ref": hour_ref,
                        "row_kind": row_kind,
                        "bank": bank,
                        "loop": loop,
                        "tipo": tipo,
                        "tag": tag,
                        "metric_name": metric_name,
                        "metric_unit": metric_name.split("(")[-1].rstrip(")") if "(" in metric_name else "",
                        "old_id": old_id,
                        "old_value": old_value,
                        "new_value": new_value,
                        "reason": str(row.get("motivo_ajuste") or "").strip(),
                        "responsible": str(row.get("responsavel_ajuste") or author or "").strip(),
                        "observation": str(row.get("observacao_ajuste") or notes or "").strip(),
                    }
                )

    if not apply:
        conn.close()
        return {
            "ok": True,
            "mode": "preview",
            "file_name": display_name,
            "rows_seen": rows_seen,
            "rows_marked": rows_marked,
            "metrics_changed": len(changes),
            "skipped": skipped[:200],
            "sample_changes": changes[:50],
        }

    created_at = now_iso()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO processing_runs(started_at, finished_at, source_type, source_ref, files_count, status, notes_json)
        VALUES(?,?,?,?,?,?,?)
        """,
        (created_at, created_at, "mpfm_adjustment", display_name, 1, "finished", json.dumps({"author": author, "notes": notes}, ensure_ascii=False)),
    )
    run_id = int(cur.lastrowid)
    cur.execute(
        """
        INSERT INTO mpfm_adjustment_imports(imported_at, source_file, source_hash, author, notes, rows_seen, rows_marked, metrics_changed, status, summary_json)
        VALUES(?,?,?,?,?,?,?,?,?,?)
        """,
        (
            created_at,
            display_name,
            source_hash,
            author,
            notes,
            rows_seen,
            rows_marked,
            len(changes),
            "applied",
            json.dumps({"skipped": skipped[:200]}, ensure_ascii=False),
        ),
    )
    import_id = int(cur.lastrowid)

    for change in changes:
        if change["old_id"]:
            cur.execute(
                """
                UPDATE measurements_curated
                SET metric_value=?, source_file=?, excel_file=?, sheet_name=?, run_id=?
                WHERE id=?
                """,
                (change["new_value"], f"manual_adjustment:{display_name}", display_name, change["sheet"], run_id, change["old_id"]),
            )
        else:
            cur.execute(
                """
                INSERT INTO measurements_curated(
                    run_id, source_file, source_record_id, excel_file, sheet_name, row_kind, day_ref, hour_ref,
                    bank, loop, tipo, tag, instrument, metric_name, metric_value, metric_unit, is_official, created_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    run_id,
                    f"manual_adjustment:{display_name}",
                    import_id,
                    display_name,
                    change["sheet"],
                    change["row_kind"],
                    change["day_ref"],
                    change["hour_ref"],
                    change["bank"],
                    change["loop"],
                    change["tipo"],
                    change["tag"],
                    "",
                    change["metric_name"],
                    change["new_value"],
                    change["metric_unit"],
                    1,
                    created_at,
                ),
            )
        cur.execute(
            """
            INSERT INTO mpfm_adjustment_items(
                import_id, day_ref, hour_ref, row_kind, bank, loop, tipo, tag, metric_name,
                old_value, new_value, reason, responsible, observation, created_at
            ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            (
                import_id,
                change["day_ref"],
                change["hour_ref"],
                change["row_kind"],
                change["bank"],
                change["loop"],
                change["tipo"],
                change["tag"],
                change["metric_name"],
                change["old_value"],
                change["new_value"],
                change["reason"],
                change["responsible"],
                change["observation"],
                created_at,
            ),
        )

    conn.commit()
    conn.close()
    return {
        "ok": True,
        "mode": "apply",
        "file_name": display_name,
        "import_id": import_id,
        "run_id": run_id,
        "rows_seen": rows_seen,
        "rows_marked": rows_marked,
        "metrics_changed": len(changes),
        "skipped": skipped[:200],
    }
