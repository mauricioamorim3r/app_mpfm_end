from __future__ import annotations
import argparse
import shutil
from datetime import date, datetime, time, timedelta
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

TEMPLATE = "dados para PETEC.xlsx"
MASTER = "BASE_UNICA_TOTAL.xlsx"
OUTPUT_DIR = "Excel_PETEC_Gerado"
SEP_ROOT_DEFAULT = __import__("os").environ.get("BASE_UNICA_SEP_ROOT", "")
SEP_PREFIX = {"separador óleo ": "SEP Óleo ", "separador gás": "SEP Gás ", "separador agua": "SEP Água "}

def parse_datetime(value: str, end_of_day: bool = False) -> datetime:
    text = value.strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return datetime.combine(parsed, time.max if end_of_day else time.min)
        except ValueError:
            pass
    raise ValueError(f"Data/hora inválida: {value}. Use DD/MM/AAAA HH:MM:SS")

def normalize_window(first: datetime, last: datetime) -> tuple[datetime, datetime, str]:
    if last < first and first - last == timedelta(hours=24):
        return last, first, "As datas foram invertidas; a janela de 24 horas foi normalizada em ordem cronológica."
    return first, last, ""

def norm(value) -> str:
    return "" if value is None else " ".join(str(value).replace("\n", " ").split()).lower()

def column(df, name):
    target = norm(name)
    return next((c for c in df.columns if norm(c) == target), None)

def row_datetime(production_date, hour) -> datetime | None:
    parsed_date = pd.to_datetime(production_date, errors="coerce")
    if pd.isna(parsed_date):
        return None
    if isinstance(hour, time):
        parsed_time = hour
    else:
        text = str(hour).strip()
        parsed_time = None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                parsed_time = datetime.strptime(text, fmt).time()
                break
            except ValueError:
                pass
        if parsed_time is None:
            try:
                numeric_hour = float(hour)
                parsed_time = time(int(numeric_hour) % 24, int(round((numeric_hour % 1) * 60)))
            except (TypeError, ValueError):
                parsed_time = time.min
    return datetime.combine(parsed_date.date(), parsed_time)

def load_window(path: Path, first: datetime, last: datetime) -> pd.DataFrame:
    # Read-only streaming avoids materializing the 36k-row master workbook.
    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book["BASE_UNICA_TOTAL"]
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    selected = [i for i, value in enumerate(headers) if value in {"ProductionDate", "Hour", "Granularity", "Origin", "Tipo", "Bank", "Tag", "Instrumento"} or str(value).startswith(("MPFM ", "PVT ", "Press", "Temper", "Dens.", "SEP "))]
    names = [headers[i] for i in selected]
    records = []
    date_index = names.index("ProductionDate")
    hour_index = names.index("Hour") if "Hour" in names else None
    for row in rows:
        raw_datetime = row_datetime(row[selected[date_index]], row[selected[hour_index]] if hour_index is not None else 0)
        if raw_datetime is not None and first <= raw_datetime <= last:
            records.append([row[i] for i in selected])
    book.close()
    return pd.DataFrame(records, columns=names)

def load_window_from_db(db_path: Path, first: datetime, last: datetime) -> pd.DataFrame:
    """Lê a janela de medições do banco SQLite (measurements_curated) em vez do Excel."""
    import sqlite3

    day_from = first.date().isoformat()
    day_to = last.date().isoformat()

    conn = sqlite3.connect(db_path)
    try:
        df_long = pd.read_sql_query(
            """
            SELECT day_ref, hour_ref, bank, loop, tipo, tag, instrument,
                   metric_name, metric_value, row_kind
            FROM measurements_curated
            WHERE row_kind = 'hourly'
              AND day_ref >= ? AND day_ref <= ?
            """,
            conn,
            params=(day_from, day_to),
        )
    finally:
        conn.close()

    if df_long.empty:
        return pd.DataFrame(columns=["ProductionDate", "Hour", "Granularity", "Origin", "Tipo", "Bank", "Tag", "Instrumento"])

    # Filter exact datetime window
    df_long["_dt"] = pd.to_datetime(df_long["day_ref"]) + pd.to_timedelta(
        df_long["hour_ref"].fillna(0).astype(int), unit="h"
    )
    df_long = df_long[df_long["_dt"].between(first, last)].copy()
    df_long.drop(columns=["_dt"], inplace=True)

    if df_long.empty:
        return pd.DataFrame(columns=["ProductionDate", "Hour", "Granularity", "Origin", "Tipo", "Bank", "Tag", "Instrumento"])

    # Pivot long → wide (each metric_name becomes a column)
    idx_cols = ["day_ref", "hour_ref", "bank", "loop", "tipo", "tag", "instrument", "row_kind"]
    df_wide = df_long.pivot_table(
        index=idx_cols,
        columns="metric_name",
        values="metric_value",
        aggfunc="first",
    ).reset_index()
    df_wide.columns.name = None

    df_wide.rename(columns={
        "day_ref": "ProductionDate",
        "hour_ref": "Hour",
        "tipo": "Tipo",
        "bank": "Bank",
        "tag": "Tag",
        "instrument": "Instrumento",
    }, inplace=True)
    df_wide["Granularity"] = "HOURLY"
    df_wide["Origin"] = "MPFM"
    df_wide.drop(columns=["row_kind", "loop"], errors="ignore", inplace=True)

    return df_wide


def load_sep_window(sep_root: Path, first: datetime, last: datetime) -> list[dict]:
    """Lê os TXT reais do SEP para as 24 posições da janela solicitada."""
    from gerar_base_unica_standalone import find_sep_files_for_day, parse_sep_txt_set

    values = {}
    day = first.date()
    while day <= last.date():
        paths = find_sep_files_for_day(sep_root, day.isoformat())
        if all(phase in paths for phase in ("oleo", "gas", "agua")):
            parsed = parse_sep_txt_set(paths["oleo"], paths["gas"], paths["agua"])
            for hour, data in parsed.items():
                if hour == "DAY":
                    continue
                timestamp = datetime.combine(day, time.min) + timedelta(hours=int(hour) - 1)
                if first <= timestamp < last:
                    values[timestamp] = data
        day += timedelta(days=1)
    return [values.get(first + timedelta(hours=offset), {}) for offset in range(24)]

def fill_mpfm(ws, df: pd.DataFrame) -> int:
    headers = [c.value for c in ws[1]]
    rows = df[(df["Origin"].astype(str).str.upper() == "MPFM") & (df["Granularity"].astype(str).str.upper() == "HOURLY")].copy()
    rows = rows.sort_values(["ProductionDate", "Hour", "Bank", "Tag"], na_position="last")
    if ws.max_row > 1: ws.delete_rows(2, ws.max_row)
    for _, item in rows.iterrows():
        values = []
        for header in headers:
            h = norm(header)
            if h == "data": value = item.get("ProductionDate")
            elif h == "hora": value = item.get("Hour")
            elif h == "banco": value = item.get("Bank")
            elif h == "tipo": value = item.get("Tipo")
            elif h == "tag": value = item.get("Tag") or item.get("Instrumento")
            else:
                source = column(df, header); value = item.get(source) if source else None
            values.append(value)
        ws.append(values)
    ws.freeze_panes = "A2"
    return len(rows)

def fill_sep(ws, sep_rows: list[dict], phase: str) -> None:
    headers = [norm(c.value) for c in ws[2]]; rules = [norm(c.value) for c in ws[3]]
    for excel_row in range(4, 28):
        offset = excel_row - 4; ws.cell(excel_row, 1).value = offset + 1
        row = sep_rows[offset] if offset < len(sep_rows) else {}
        for number, header in enumerate(headers[1:], 2):
            key = f"{phase}_{header_key(header)}"
            value = row.get(key)
            ws.cell(excel_row, number).value = value if value is not None and not pd.isna(value) else None

def header_key(value: str) -> str:
    return {"pressure (kpa)": "pressure_kpa", "pressure (barg)": "pressure_barg", "pressure (kpa_g)": "pressure_kpa_g", "temperature (deg c)": "temp", "sd (kg/sm³)": "sd", "md (kg/m³)": "md", "dt (kg/m³)": "dt", "iv (m³)": "iv_m3", "gv (m³)": "gv_m3", "gsv (sm³)": "gsv_sm3", "gr. vol. (m³)": "gr_vol_m3", "st. vol. (m³)": "st_vol_m3", "mass (t)": "mass_t", "energy (gj)": "energy_gj", "nsv (sm³)": "nsv_sm3", "bsw (%)": "bsw_pct", "cpl": "cpl", "ctl": "ctl", "diff. press. (kpa)": "diff_press_kpa", "flowtime (min)": "flowtime_min"}.get(value, value)

def main() -> int:
    parser = argparse.ArgumentParser(description="Preenche o template PETEC sem alterar a Base Única.")
    parser.add_argument("--date-from", default=""); parser.add_argument("--date-to", default=""); parser.add_argument("--output", default=""); parser.add_argument("--sep-root", default=SEP_ROOT_DEFAULT)
    parser.add_argument("--bank", default="", help="Filtra banco MPFM (ex.: B05).")
    parser.add_argument("--tag", default="", help="Filtra Tag MPFM.")
    parser.add_argument("--instrument", default="", help="Filtra Instrumento MPFM.")
    parser.add_argument("--allow-missing-sep", action="store_true", help="Permite gerar as abas SEP vazias de forma explícita.")
    parser.add_argument("--from-db", action="store_true", help="Busca dados MPFM do banco SQLite local em vez do Excel.")
    parser.add_argument("--db-path", default="", help="Caminho para mpfm_local.db (padrão: <projeto>/data/mpfm_local.db).")
    args = parser.parse_args()
    try:
        first = parse_datetime(args.date_from or input("Data/hora inicial (DD/MM/AAAA HH:MM:SS): "))
        last = parse_datetime(args.date_to or input("Data/hora final (DD/MM/AAAA HH:MM:SS): "), end_of_day=True)
        first, last, warning = normalize_window(first, last)
        if warning: print(f"AVISO PETEC: {warning}")
        if last < first or last - first > timedelta(hours=24): raise ValueError("A janela PETEC deve ter no máximo 24 horas.")
        root = Path(__file__).resolve().parent; template = root / TEMPLATE; master = root / MASTER
        output_dir = root / OUTPUT_DIR
        output = Path(args.output) if args.output else output_dir / f"PETEC_preenchido_{first:%Y%m%d_%H%M%S}_{last:%Y%m%d_%H%M%S}.xlsx"
        output.parent.mkdir(parents=True, exist_ok=True)
        if output.resolve() == master.resolve(): raise ValueError("A saída não pode substituir BASE_UNICA_TOTAL.xlsx.")
        sep_root = Path(args.sep_root)
        if not sep_root.is_dir() and not args.allow_missing_sep:
            raise ValueError("Informe uma pasta SEP válida em --sep-root ou use --allow-missing-sep conscientemente.")
        shutil.copy2(template, output)
        if args.from_db:
            _db = Path(args.db_path) if args.db_path else Path(__file__).resolve().parents[1] / "data" / "mpfm_local.db"
            if not _db.exists():
                raise FileNotFoundError(f"Banco não encontrado: {_db}. Use --db-path para especificar o caminho.")
            print(f"Fonte: banco SQLite -> {_db}")
            data = load_window_from_db(_db, first, last)
        else:
            data = load_window(master, first, last)
        for column_name, requested in (("Bank", args.bank), ("Tag", args.tag), ("Instrumento", args.instrument)):
            if requested:
                data = data[data[column_name].astype(str).str.strip().str.upper().eq(requested.strip().upper())].copy()
        book = load_workbook(output)
        count = fill_mpfm(book["MPFM"], data)
        sep_rows = load_sep_window(sep_root, first, last) if sep_root.is_dir() else [{} for _ in range(24)]
        if not args.allow_missing_sep and not any(sep_rows):
            raise ValueError("Nenhum conjunto TXT de óleo/gás/água do SEP foi encontrado na janela solicitada.")
        fill_sep(book["separador óleo "], sep_rows, "oil")
        fill_sep(book["separador gás"], sep_rows, "gas")
        fill_sep(book["separador agua"], sep_rows, "water")
        for sheet in book.worksheets: sheet.sheet_view.showGridLines = False
        book.save(output); print(f"PETEC preenchido: {output}"); print(f"Registros MPFM: {count}. BASE_UNICA_TOTAL.xlsx não foi alterada."); return 0
    except Exception as exc: print(f"ERRO PETEC: {exc}"); return 1

if __name__ == "__main__": raise SystemExit(main())
