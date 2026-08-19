#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Gerador Base_Unica — versão 100% AUTOCONTIDA (sem depender do repositório
app_mpfm_end, do banco de dados ou do server.py).

Este único arquivo contém tudo que precisa para rodar em qualquer máquina:
  - parsing dos PDFs MPFM (Daily/Hourly)         -> copiado de mpfm_engine.py
  - parsing dos TXT do Separador de Testes (SEP) -> copiado de mpfm_engine.py
  - localização dos TXT do SEP por Meter ID      -> lógica própria, simples
  - montagem das linhas no formato Base_Unica    -> lógica própria
    - atualização incremental de uma Base_Unica total consolidada

DEPENDÊNCIAS PYTHON (instale com pip se a máquina não tiver):
    pip install pandas numpy pypdf openpyxl

COMO CONFIGURAR OS CAMINHOS
----------------------------
Existem 3 formas (use a que preferir):

1) Editar as constantes na seção "CONFIG" logo abaixo (mais simples).
2) Passar por linha de comando, ex.:
     python gerar_base_unica_standalone.py --mpfm-root "D:\Relatorios\MPFM" ^
         --sep-root "D:\Relatorios\SEP" --output "D:\saida.xlsx" --days 5 ^
         --master-output "D:\BASE_UNICA_TOTAL.xlsx"
3) Se não configurar nem passar argumentos, o script pergunta os caminhos
   interativamente (input) na hora de rodar.

BASE_UNICA_TOTAL INCREMENTAL
----------------------------
Por padrão, depois de gerar o Excel individual da análise, o script também
atualiza `BASE_UNICA_TOTAL.xlsx` ao lado deste arquivo. A integração é
incremental: histórico + nova análise são combinados e linhas repetidas são
substituídas pela versão mais recente usando a chave técnica:
ProductionDate, Hour, Granularity, Origin, SourceType, Bank, Entity, Tag,
Instrumento. Use `--no-master` para desativar ou `--master-output` para escolher
outro caminho.

ESTRUTURA DE PASTAS ESPERADA
-----------------------------
MPFM_ROOT/
  <pasta do banco B03>/<ano>/<NN. NomeMes>/Daily/B03_MPFM_Daily-YYYYMMDD-......pdf
  <pasta do banco B03>/<ano>/<NN. NomeMes>/Hourly/B03_MPFM_Hourly-YYYYMMDD-HHMMSS...pdf
  ... (um subdiretório por banco, ver BANK_FOLDERS abaixo)

SEP_ROOT/
  <ano>/<NN. NomeMes>/FPSO-Bacalhau_Daily reports_YYYY-MM-DD/01 - CV_Reports/FC13/Run_24Hours*.txt
  ... (idem para FC14 e FC17)

Se a estrutura de pastas da sua instalação for diferente, ajuste as funções
`bank_month_dir()` e `sep_day_dir()` na seção "CAMINHOS DE PASTA" abaixo.
"""

from __future__ import annotations

import argparse
import base64
import gc
import hashlib
import html
import importlib.util
import json
import os
import shutil
import re
import socket
import subprocess
import sys
import tempfile
import time
import traceback
import webbrowser
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from pypdf import PdfReader

# ═════════════════════════════════════════════════════════════════════════
# CONFIG — ajuste aqui (ou use os argumentos de linha de comando / prompts)
# ═════════════════════════════════════════════════════════════════════════

# Caminhos padrão. Não embutir caminhos pessoais no pacote: o usuário pode
# informar os caminhos por CLI, na execução assistida ou por variáveis de
# ambiente BASE_UNICA_MPFM_ROOT e BASE_UNICA_SEP_ROOT.
MPFM_ROOT = os.environ.get("BASE_UNICA_MPFM_ROOT", "").strip()
SEP_ROOT = os.environ.get("BASE_UNICA_SEP_ROOT", "").strip()

# Nome da subpasta de cada banco dentro de MPFM_ROOT.
BANK_FOLDERS = {
    "B03": "3.1.1_13-FT-0367 Riser P5 - Topside B03",
    "B08": "3.1.2_13-FT-0167 Riser P2 - Topside B08",
    "B13": "3.1.3_13-FT-0317 Riser P4 - Topside B13",
    "B05": "3.1.4_18-FT-1506 PE 4 e PE_EO105 - Subsea B05",
    "B10": "3.1.5_18-FT-0506 PE 2 - Subsea B10",
    "B15": "3.1.6_18-FT-1106 PW_104DA - Subsea B15",
}

MPFM_INSTRUMENT_METADATA = {
    # B03 – Topside North
    "13FT0367": {"bank": "B03", "tag": "13FT0367", "entity": "Riser_P5",  "tipo": "Topside", "loop": "North"},
    "13FT0417": {"bank": "B03", "tag": "13FT0417", "entity": "Riser_P6",  "tipo": "Topside", "loop": "North"},
    # B05 – Subsea North  (PE_4, PE_4A, PE-4 e PE-04 são aliases do ponto 18FT1506)
    "18FT1506": {"bank": "B05", "tag": "18FT1506", "entity": "PE_4",      "tipo": "Subsea",  "loop": "North"},
    "18FT1706": {"bank": "B05", "tag": "18FT1706", "entity": "PE_EO105",  "tipo": "Subsea",  "loop": "North"},
    "18FT1406": {"bank": "B05", "tag": "18FT1406", "entity": "PE_EO10",   "tipo": "Subsea",  "loop": "North"},
    "18FT1806": {"bank": "B05", "tag": "18FT1806", "entity": "PE_EO4",    "tipo": "Subsea",  "loop": "North"},
    # B08 – Topside South
    "13FT0167": {"bank": "B08", "tag": "13FT0167", "entity": "Riser_P1",  "tipo": "Topside", "loop": "South"},
    "13FT0217": {"bank": "B08", "tag": "13FT0217", "entity": "Riser_P2",  "tipo": "Topside", "loop": "South"},
    # B10 – Subsea South
    "18FT0506": {"bank": "B10", "tag": "18FT0506", "entity": "PE_2",      "tipo": "Subsea",  "loop": "South"},
    "18FT0306": {"bank": "B10", "tag": "18FT0306", "entity": "PE_8",      "tipo": "Subsea",  "loop": "South"},
    "18FT0106": {"bank": "B10", "tag": "18FT0106", "entity": "PE_9",      "tipo": "Subsea",  "loop": "South"},
    # B13 – Topside West
    "13FT0267": {"bank": "B13", "tag": "13FT0267", "entity": "Riser_P3",  "tipo": "Topside", "loop": "West"},
    "13FT0317": {"bank": "B13", "tag": "13FT0317", "entity": "Riser_P4",  "tipo": "Topside", "loop": "West"},
    # B15 – Subsea West
    "18FT0706": {"bank": "B15", "tag": "18FT0706", "entity": "PE_1",      "tipo": "Subsea",  "loop": "West"},
    "18FT0906": {"bank": "B15", "tag": "18FT0906", "entity": "PI_1",      "tipo": "Subsea",  "loop": "West"},
    "18FT1206": {"bank": "B15", "tag": "18FT1206", "entity": "PI_2",      "tipo": "Subsea",  "loop": "West"},
    "18FT1106": {"bank": "B15", "tag": "18FT1106", "entity": "PW-104DA",  "tipo": "Subsea",  "loop": "West"},
}

MPFM_IDENTITY_ALIASES = {
    "18FT1506": "18FT1506", "PE4": "18FT1506", "PE04": "18FT1506", "PE4A": "18FT1506",
    "18FT1706": "18FT1706", "PEEO105": "18FT1706",
    "18FT1406": "18FT1406", "PEEO10": "18FT1406",
    "18FT1806": "18FT1806", "PEEO4": "18FT1806",
    "13FT0367": "13FT0367", "RISERP5": "13FT0367",
    "13FT0417": "13FT0417", "RISERP6": "13FT0417",
    "13FT0167": "13FT0167", "RISERP1": "13FT0167",
    "13FT0217": "13FT0217", "RISERP2": "13FT0217",
    "13FT0267": "13FT0267", "RISERP3": "13FT0267",
    "13FT0317": "13FT0317", "RISERP4": "13FT0317",
    "18FT0506": "18FT0506", "PE2": "18FT0506",
    "18FT0306": "18FT0306", "PE8": "18FT0306",
    "18FT0106": "18FT0106", "PE9": "18FT0106",
    "18FT0706": "18FT0706", "PE1": "18FT0706",
    "18FT0906": "18FT0906", "PI1": "18FT0906",
    "18FT1206": "18FT1206", "PI2": "18FT1206",
    "18FT1106": "18FT1106", "PW104DA": "18FT1106",
}

# Pontos fora do escopo operacional atual. Permanecem bloqueados tanto na
# extração quanto nos dashboards; para uma campanha autorizada, adicione a
# chave normalizada em AUTHORIZED_EXTRACTION_MPFMS, sem alterar o parser.
EXCLUDED_EXTRACTION_MPFMS = {
    "PEEO10", "RISERP1", "RISERP8", "RISERP9", "PE1", "PI1", "PI2",
}
AUTHORIZED_EXTRACTION_MPFMS: set[str] = set()

# Instrumento (TAG) físico de cada seção, usado para atribuir corretamente a
# origem de uma seção de PDF quando o relatório de um banco vem com seções de
# mesmo banco dentro do mesmo arquivo (ex.: PE_4/18FT1506 e
# PE_EO105/18FT1706 no PDF Daily/Hourly do B05).
INSTRUMENT_TO_BANK = {
    "13FT0367": "B03", "13FT0417": "B03",
    "18FT1506": "B05", "18FT1706": "B05", "18FT1406": "B05", "18FT1806": "B05",
    "13FT0167": "B08", "13FT0217": "B08",
    "18FT0506": "B10", "18FT0306": "B10", "18FT0106": "B10",
    "13FT0267": "B13", "13FT0317": "B13",
    "18FT0706": "B15", "18FT0906": "B15", "18FT1206": "B15", "18FT1106": "B15",
}


def _normalize_instrument(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def _mpfm_extraction_enabled(tag="", instrument="") -> bool:
    """Retorna se um ponto está autorizado para entrar na Base_Unica."""
    candidates = {_normalize_instrument(tag), _normalize_instrument(instrument)} - {""}
    return not any(
        candidate in EXCLUDED_EXTRACTION_MPFMS and candidate not in AUTHORIZED_EXTRACTION_MPFMS
        for candidate in candidates
    )


# Pontos ocultos APENAS na publicação HTML. Diferente de
# EXCLUDED_EXTRACTION_MPFMS, aqui a extração continua normal: os dados seguem
# na Base_Unica e nas planilhas geradas, apenas não são publicados nos painéis.
# Para liberar um ponto no dashboard, mova a chave normalizada para
# DASHBOARD_AUTHORIZED_POINTS — liberação por código, sem tocar no parser.
DASHBOARD_HIDDEN_POINTS = {
    "RISERP6", "PEEO10", "PEEO4", "PE8", "PE9", "PE1", "PI1", "PI2",
}
DASHBOARD_AUTHORIZED_POINTS: set[str] = set()

# Nome operacional de cada ponto de medição, usado nos filtros, cards e tabelas
# do HTML no lugar do código do banco. Chave = TAG/instrumento normalizado.
POINT_LABELS = {
    "RISERP2": "Riser P2",
    "RISERP3": "Riser P3",
    "RISERP4": "Riser P4",
    "RISERP5": "Riser P5",
    "RISERP6": "Riser P6",
    "13FT0167": "Riser P2",
    "13FT0217": "Riser P2",
    "13FT0267": "Riser P3",
    "13FT0317": "Riser P4",
    "13FT0367": "Riser P5",
    "13FT0417": "Riser P6",
    "18FT1506": "Poço PE-04",
    "18FT1706": "Poço PE-EO105",
    "18FT1406": "Poço PE-EO10",
    "18FT1806": "Poço PE-EO4",
    "18FT0506": "Poço PE-02",
    "18FT1106": "Poço PW-104",
    "PE2": "Poço PE-02",
    "PE4": "Poço PE-04",
    "PE04": "Poço PE-04",
    "PEEO105": "Poço PE-EO105",
    "PEEO10": "Poço PE-EO10",
    "PEEO4": "Poço PE-EO4",
    "PW104DA": "Poço PW-104",
}


def _dashboard_point_visible(tag="", instrument="") -> bool:
    """Retorna se um ponto pode ser publicado nos painéis HTML."""
    candidates = {_normalize_instrument(tag), _normalize_instrument(instrument)} - {""}
    return not any(
        candidate in DASHBOARD_HIDDEN_POINTS and candidate not in DASHBOARD_AUTHORIZED_POINTS
        for candidate in candidates
    )


def _point_key(tag="", instrument="") -> str:
    """Chave estável de um ponto de medição para agrupar e filtrar no HTML."""
    for value in (tag, instrument):
        norm = _normalize_instrument(value)
        if norm:
            return MPFM_IDENTITY_ALIASES.get(norm, norm)
    return ""


def _point_name(tag="", instrument="") -> str:
    """Nome operacional amigável de um ponto (sem TAG nem banco)."""
    for value in (tag, instrument):
        name = POINT_LABELS.get(_normalize_instrument(value), "")
        if name:
            return name
    raw = str(tag or instrument or "").strip()
    return re.sub(r"[_\-]+", " ", raw).strip() or "Sem TAG"


def _point_display_label(tag="", bank="", instrument="") -> str:
    """Rótulo operacional de um ponto: nome amigável + TAG + banco."""
    raw_tag = str(tag or instrument or "").strip()
    details = " · ".join(part for part in (raw_tag, str(bank or "").strip()) if part)
    name = _point_name(tag, instrument)
    return f"{name} ({details})" if details else name


def _resolve_tag_bank(instrument):
    """Resolve o banco de uma seção de PDF pelo instrumento (TAG). Retorna
    None quando o instrumento não é um dos bancos monitorados (nesse caso o
    chamador decide, de forma leniente ou estrita, a quem atribuir a linha)."""
    return INSTRUMENT_TO_BANK.get(_normalize_instrument(instrument))


def _resolve_mpfm_metadata(instrument="", entity="", tag="") -> dict | None:
    """Resolve aliases operacionais para uma única identidade física."""
    for value in (instrument, entity, tag):
        canonical_instrument = MPFM_IDENTITY_ALIASES.get(_normalize_instrument(value))
        if canonical_instrument in MPFM_INSTRUMENT_METADATA:
            return MPFM_INSTRUMENT_METADATA[canonical_instrument]
    return None


def _tag_belongs_to_bank(instrument, unit_code: str, strict: bool = False) -> bool:
    """Decide se uma seção de PDF (identificada pelo instrumento) deve virar
    linha para `unit_code`. Uso leniente (strict=False, padrão): instrumentos
    desconhecidos continuam atribuídos ao banco nativo do arquivo, preservando
    o comportamento histórico. Uso estrito (strict=True): usado ao reaproveitar
    um relatório compartilhado — instrumentos desconhecidos podem ser
    ignorados para não atribuir uma seção a um banco incorreto."""
    resolved = _resolve_tag_bank(instrument)
    if resolved is None:
        return not strict
    return resolved == unit_code

COMPARISON_PAIRS = [
    {"pair": "PE-04 × Riser P5", "subsea_bank": "B05", "subsea_instruments": ["18FT1506"], "subsea_label": "PE-04 Subsea", "topside_bank": "B03", "topside_instruments": ["13FT0367"], "topside_label": "Riser P5 Topside"},
    {"pair": "PE-02 × Riser P2", "subsea_bank": "B10", "subsea_instruments": ["18FT0506"], "subsea_label": "PE-02 Subsea", "topside_bank": "B08", "topside_instruments": ["13FT0217"], "topside_label": "Riser P2 Topside"},
    {"pair": "PW-104 × Riser P4", "subsea_bank": "B15", "subsea_instruments": ["18FT1106"], "subsea_label": "PW-104 Subsea", "topside_bank": "B13", "topside_instruments": ["13FT0317"], "topside_label": "Riser P4 Topside"},
]
COMPARISON_BANKS = {bank for pair in COMPARISON_PAIRS for bank in (pair["subsea_bank"], pair["topside_bank"])}


def _comparison_info_for_row(row) -> dict:
    """Identifica um lado do par pelo banco e pelo instrumento oficial."""
    bank = str(row.get("Bank", "")).strip().upper()
    row_ids = {
        _normalize_instrument(row.get("Instrumento", "")),
        _normalize_instrument(row.get("Tag", "")),
    }
    row_ids.discard("")
    for pair in COMPARISON_PAIRS:
        for role, prefix in (("Subsea", "subsea"), ("Topside", "topside")):
            if bank != str(pair[f"{prefix}_bank"]).upper():
                continue
            allowed = {_normalize_instrument(value) for value in pair.get(f"{prefix}_instruments", [])}
            if allowed and not row_ids.intersection(allowed):
                continue
            return {"pair": pair["pair"], "side": pair[f"{prefix}_label"], "role": role}
    return {}


def _comparison_side_mask(df: pd.DataFrame, pair: dict, prefix: str) -> pd.Series:
    bank_mask = df["Bank"].astype(str).str.upper().eq(str(pair[f"{prefix}_bank"]).upper())
    allowed = {_normalize_instrument(value) for value in pair.get(f"{prefix}_instruments", [])}
    if not allowed:
        return bank_mask
    instrument = df.get("Instrumento", pd.Series("", index=df.index)).map(_normalize_instrument)
    tag = df.get("Tag", pd.Series("", index=df.index)).map(_normalize_instrument)
    return bank_mask & (instrument.isin(allowed) | tag.isin(allowed))

DASHBOARD_VARIABLES = {
    "mpfm_uncorr_hc": {"label": "MPFM uncorr HC (t)", "mpfm": "MPFM uncorr HC (t)", "sep": "SEP HC (t)", "aggregation": "sum"},
    "mpfm_uncorr_total": {"label": "MPFM uncorr Total (t)", "mpfm": "MPFM uncorr Total (t)", "sep": "SEP Total (t)", "aggregation": "sum"},
    "mpfm_corr_hc": {"label": "MPFM corr HC (t)", "mpfm": "MPFM corr HC (t)", "sep": "SEP HC (t)", "aggregation": "sum"},
    "mpfm_corr_total": {"label": "MPFM corr Total (t)", "mpfm": "MPFM corr Total (t)", "sep": "SEP Total (t)", "aggregation": "sum"},
    "pvt20_mass_gas": {"label": "PVT @20 mass Gás (t)", "mpfm": "PVT @20 mass Gás (t)", "sep": "SEP Gás Mass (t)", "aggregation": "sum"},
    "pvt20_mass_oil": {"label": "PVT @20 mass Óleo (t)", "mpfm": "PVT @20 mass Óleo (t)", "sep": "SEP Óleo Mass (t)", "aggregation": "sum"},
    "pvt20_mass_water": {"label": "PVT @20 mass Água (t)", "mpfm": "PVT @20 mass Água (t)", "sep": "SEP Água Mass (t)", "aggregation": "sum"},
    "pvt20_vol_gas": {"label": "PVT @20 vol Gás (Sm³)", "mpfm": "PVT @20 vol Gás (Sm³)", "sep": "SEP Gás St. Vol. (m³)", "aggregation": "sum"},
    "pvt20_vol_oil": {"label": "PVT @20 vol Óleo (m³)", "mpfm": "PVT @20 vol Óleo (m³)", "sep": "SEP Óleo NSV (sm³)", "aggregation": "sum"},
    "pvt20_vol_water": {"label": "PVT @20 vol Água (m³)", "mpfm": "PVT @20 vol Água (m³)", "sep": "SEP Água NSV (sm³)", "aggregation": "sum"},
    "pressao": {"label": "Pressão (barg)", "mpfm": "Pressão (barg)", "sep": "SEP Pressão Méd. (barg)", "aggregation": "avg"},
    "temperatura": {"label": "Temperatura (°C)", "mpfm": "Temperatura (°C)", "sep": "SEP Temperatura Méd. (°C)", "aggregation": "avg"},
}

OFFICIAL_DEVIATION_SPECS = [
    {"metric": "HC", "label": "Massa HC", "mpfm": "MPFM corr HC (t)", "ref": "SEP HC (t)", "limit": 10.0, "class": "CRÍTICO RANP 44"},
    {"metric": "Total", "label": "Massa Total", "mpfm": "MPFM corr Total (t)", "ref": "SEP Total (t)", "limit": 7.0, "class": "CRÍTICO RANP 44"},
    {"metric": "Óleo", "label": "Fase Óleo", "mpfm": "MPFM corr Óleo (t)", "ref": "SEP Óleo Mass (t)", "limit": None, "class": "DIAGNÓSTICO"},
    {"metric": "Gás", "label": "Fase Gás", "mpfm": "MPFM corr Gás (t)", "ref": "SEP Gás Mass (t)", "limit": None, "class": "DIAGNÓSTICO"},
    {"metric": "Água", "label": "Fase Água", "mpfm": "MPFM corr Água (t)", "ref": "SEP Água Mass (t)", "limit": None, "class": "DIAGNÓSTICO"},
]

# Métricas comparadas entre os dois lados de um par Subsea × Topside.
# (chave, rótulo, coluna da Base_Unica, limite RANP 44 em % ou None para
# métrica apenas diagnóstica). As fases entram em volume padrão 20 °C / 1 atm,
# que é a base em que as fases são reconciliadas operacionalmente.
PAIR_DEVIATION_SPECS = [
    ("HC", "Massa HC", "MPFM corr HC (t)", 10.0),
    ("Total", "Massa Total", "MPFM corr Total (t)", 7.0),
    ("Óleo mass", "Massa Óleo", "MPFM corr Óleo (t)", None),
    ("Gás mass", "Massa Gás", "MPFM corr Gás (t)", None),
    ("Água mass", "Massa Água", "MPFM corr Água (t)", None),
    ("Óleo vol", "Volume Óleo @20 °C/1 atm", "PVT @20 vol Óleo (m³)", None),
    ("Gás vol", "Volume Gás @20 °C/1 atm", "PVT @20 vol Gás (Sm³)", None),
    ("Água vol", "Volume Água @20 °C/1 atm", "PVT @20 vol Água (m³)", None),
]

DAYS_COUNT = 5              # quantos dias (mais recentes disponíveis) exportar
SEP_ALIGNED_BANK = "B10"    # compatibilidade de CLI; não vincula mais SEP a MPFM
AUTOMATION_ROOT = Path(__file__).resolve().parent
EXCEL_OUTPUT_DIR = AUTOMATION_ROOT / "EXCEL_GERADOS"
HTML_OUTPUT_DIR = AUTOMATION_ROOT / "HTML_GERADOS"
OUTPUT_PATH = r""           # vazio: salva em EXCEL_GERADOS
MASTER_OUTPUT_PATH = r""    # vazio: salva BASE_UNICA_TOTAL.xlsx em EXCEL_GERADOS
UPDATE_MASTER = True         # atualiza arquivo incremental total após cada execução
MONTHS_LOOKBACK = 12         # meses máximos pesquisados para encontrar os dias disponíveis
PDF_WORKERS = 4              # processos paralelos para PDFs em exportações históricas

MONTH_PT = {
    1: "01. Janeiro", 2: "02. Fevereiro", 3: "03. Março", 4: "04. Abril",
    5: "05. Maio", 6: "06. Junho", 7: "07. Julho", 8: "08. Agosto",
    9: "09. Setembro", 10: "10. Outubro", 11: "11. Novembro", 12: "12. Dezembro",
}

# Meter ID -> fluido, usado para identificar os TXT do Separador de Testes
TARGET_PHASE_BY_METER = {
    "20FT0247": "oleo",
    "20FT0251": "agua",
    "20FT0244": "gas",
}
SEP_FC_FOLDERS = ["FC13", "FC14", "FC17"]

ALARM_FOLDER_NAME = "3.1.7_ALARMES_FCS_320"
EVENT_FOLDER_NAME = "3.1.8_EVENTOS_FCS320"
ALARM_EVENT_SHEET_NAME = "ALARMES_EVENTOS"
ALARM_EVENT_MASTER_SHEET_NAME = "ALARMES_EVENTOS_TOTAL"
ALARM_EVENT_COLUMNS = [
    "ProductionDate", "Timestamp", "RecordType", "SourceKind", "Bank", "Priority",
    "Object", "Description", "State", "DetailedState", "SourceID", "SignalNumber",
    "Instrumento", "SourceFile", "IssueFlag",
]

PI_COLLECTOR_ROOT = os.environ.get("BASE_UNICA_PI_ROOT", r"C:\PI_Vision_Collector")
PI_COLLECTOR_CONFIG = "config_v48.json"
PI_COLLECTOR_RUNNER = "run_v48.py"
PI_EXTRACT_OUTPUT = os.environ.get("BASE_UNICA_PI_OUTPUT", r"C:\PI_Vision_Collector\saida_v4\Historico_V49_Geometrico.xlsx")
PI_PERIOD_OUTPUT = os.environ.get("BASE_UNICA_PI_PERIOD_OUTPUT", r"C:\PI_Vision_Collector\saida_v4\Periodo_Coleta_V49.json")
PI_DAILY_CONTROL_ROOT = os.environ.get("BASE_UNICA_DAILY_CONTROL_ROOT", r"C:\DailyControl")
PI_DAILY_CONTROL_COLLECTOR = "collector_daily_control_v5_2.py"
PI_DAILY_CONTROL_CONFIG = "config_daily_control_v5_2.json"
PI_DAILY_CONTROL_NORMALIZER = "normalizer_daily_control_v5_1.py"
PI_DAILY_CONTROL_NORMALIZER_CORE = "normalizer_daily_control_v5_corrigido.py"
PI_DAILY_CONTROL_OUTPUT = os.environ.get("BASE_UNICA_DAILY_CONTROL_OUTPUT", r"C:\DailyControl\output\Historico_Daily_Control_V5.xlsx")
PI_DAILY_CONTROL_EXPECTED_JSONS = 7
PI_CDP_HOST = "127.0.0.1"
PI_CDP_PORT = 9222
PI_EDGE_START_TIMEOUT_SECONDS = 30
PI_EDGE_PROFILE_DIR = ""  # vazio: %LOCALAPPDATA%\PI_Vision_CDP
PI_VISION_URL = "https://pivision.equinor.com/PIVision/#/Displays/54854/Metering-Monitor"
PI_DAILY_CONTROL_URL = "https://pivision.equinor.com/PIVision/#/Displays/56466/Metering-Daily-Control"
PI_OPEN_DAILY_CONTROL = True
PI_CAPTURE_DAILY_CONTROL = True
PI_DAILY_CONTROL_VARIABLES = [
    "WLR", "WVF", "GVF", "GOR", "Temperature", "Pressure", "dP Inlet", "dP Outlet",
    "Velocity", "Water Conductivity", "Meter Status 1", "Meter Status 2",
    "Flow Calculation Warn.", "Calculation Mode", "Continuous Phase", "Water Conductivity Input",
]
# Contrato mínimo do PI para condições de contorno. O filtro é aplicado ao
# coletor normalizado e também ao carregamento de arquivos legados, evitando
# que variáveis fora do escopo reapareçam no PI_EXTRACT_TOTAL.
PI_CONTOUR_VARIABLES = frozenset(PI_DAILY_CONTROL_VARIABLES)
PI_SHEET_NAME = "PI_EXTRACT"
PI_MASTER_SHEET_NAME = "PI_EXTRACT_TOTAL"
PI_RETRIES = 2
PI_RETRY_PERIOD_SETTLE_INCREMENT_SECONDS = 20
PI_RETRY_COLLECTOR_SETTLE_INCREMENT_SECONDS = 10
PI_COLLECTION_TIMEOUT_SECONDS = int(os.environ.get("PI_COLLECTION_TIMEOUT_SECONDS", "900") or "900")
EMAIL_AUTOMATION_DIR = "OUTRAS_AUTOMACOES"
EMAIL_DOWNLOAD_SCRIPT = "baixar_zip_email.py"
EMAIL_ORGANIZE_SCRIPT = "dpb_bacalhau_todos_zips.py"
XML042_AUTOMATION_DIR = "XML042_STANDALONE_PACOTE"
XML042_CONFIG_FILE = "config_xml042_standalone.json"
XML042_SCRIPT_FILE = "gerar_xml042_standalone.py"
LOGO_FILENAME = "logo.png"


def _dashboard_logo_data_uri() -> str:
    """Embute o logo no HTML para o dashboard funcionar como arquivo único."""
    logo_path = Path(__file__).resolve().parent / LOGO_FILENAME
    if not logo_path.exists():
        return ""
    try:
        encoded = base64.b64encode(logo_path.read_bytes()).decode("ascii")
        return f"data:image/png;base64,{encoded}"
    except OSError:
        return ""
XML042_BAT_FILE = "executar_xml042_standalone.bat"

OPERATION_MODES = {
    "1": {"label": "Automação completa", "email": True, "pi": True, "base": True},
    "2": {"label": "PI + Base_Unica", "email": False, "pi": True, "base": True},
    "3": {"label": "Baixar PDF/TXT + Base_Unica", "email": True, "pi": False, "base": True},
    "4": {"label": "Baixar PDF/TXT + PI", "email": True, "pi": True, "base": False},
    "5": {"label": "Apenas Baixar PDF/TXT", "email": True, "pi": False, "base": False},
    "6": {"label": "Apenas PI", "email": False, "pi": True, "base": False},
    "7": {"label": "Apenas Base_Unica", "email": False, "pi": False, "base": True},
}


# ═════════════════════════════════════════════════════════════════════════
# MPFM ENGINE — copiado/adaptado de mpfm_engine.py (parsing de PDF e TXT)
# ═════════════════════════════════════════════════════════════════════════

def _to_float(x):
    x = str(x).strip().replace('\u00A0', '').replace(' ', '')
    if x in ('-', '', 'None', 'nan', 'NaN'):
        return np.nan
    try:
        return float(x.replace(',', '.'))
    except Exception:
        return np.nan


def _flat_text(pdf_path):
    reader = PdfReader(pdf_path)
    raw = '\n'.join(page.extract_text() or '' for page in reader.pages)
    return re.sub(r'\s+', ' ', raw)


NUM = r'([\-\d]+(?:\.\d+)?)'
C5 = ['gas', 'oil', 'hc', 'water', 'total']
C4 = ['gas', 'oil', 'hc', 'water']


def _nan_dict(keys):
    return {k: np.nan for k in keys}


def _parse_row(label_pat, text, ncols=5):
    cols = (C5 if ncols == 5 else C4)[:ncols]
    m = re.search(
        label_pat + r'[^\[]*\[(?:t|Sm[³3]?)\]\s*' + r'\s+'.join([NUM] * ncols),
        text, re.IGNORECASE
    )
    if m:
        return {c: _to_float(m.group(i + 1)) for i, c in enumerate(cols)}
    return _nan_dict(cols)


def _parse_fwa(text):
    """Extrai Pressão, Temperatura e Densidades (Flow Weighted Averages)."""
    def _find_meter_value(label, unit_pat):
        m = re.search(
            label + r'\s*\n?\s*' + unit_pat +
            r'\s*\n(?:-[^\n]*\n){0,2}\s*(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
        m = re.search(
            label + r'\s*' + unit_pat + r'(?:\s*-\s*){1,3}(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
        return np.nan

    def _extract_density(text):
        m = re.search(
            r'Density\s*\[kg/m\S*\]\s*(.*?)'
            r'(?=Production\s+Total|PE_|Riser\s+P|\Z)',
            text, re.DOTALL | re.IGNORECASE
        )
        if not m:
            return np.nan, np.nan, np.nan
        tokens = []
        for part in re.split(r'\s+', m.group(1).strip()):
            if re.fullmatch(r'-|\d+(?:\.\d+)?', part):
                tokens.append(part)
                if len(tokens) == 4:
                    break
            elif part:
                break

        def t2f(t):
            return np.nan if t == '-' else _to_float(t)
        gas = t2f(tokens[0]) if len(tokens) > 0 else np.nan
        oil = t2f(tokens[1]) if len(tokens) > 1 else np.nan
        water = t2f(tokens[3]) if len(tokens) > 3 else np.nan
        return gas, oil, water

    pres = _find_meter_value(r'Pressure', r'\[barg\]')
    temp = _find_meter_value(r'Temperature', r'\[.{1,3}C\]')
    dens_gas, dens_oil, dens_water = _extract_density(text)
    return {
        'pressure': pres, 'temperature': temp,
        'dens_gas': dens_gas, 'dens_oil': dens_oil, 'dens_water': dens_water,
    }


def _parse_metrics(block, ncols=5):
    corr = _parse_row(r'MPFM\s+corrected\s+mass', block, ncols)
    all_zero = all(
        (v == 0 or (isinstance(v, float) and (v == 0.0 or np.isnan(v))))
        for v in corr.values()
    )
    fwa = {k: np.nan for k in ('pressure', 'temperature', 'dens_gas', 'dens_oil', 'dens_water')} \
        if all_zero else _parse_fwa(block)
    return {
        'mpfm_uncorr': _parse_row(r'MPFM\s+uncorrected\s+mass', block, ncols),
        'mpfm_corr': corr,
        'pvt_mass': _parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*(?!@)', block, ncols),
        'pvt_vol': _parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*(?!@)', block, ncols),
        'pvt20_mass': _parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*@20', block, ncols),
        'pvt20_vol': _parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*@20', block, ncols),
        'fwa': fwa,
    }


TAG_RE = re.compile(
    r'((?:Riser\s+P\d+|P[EI]_[A-Z0-9]+-?|PW[-_][A-Z0-9]+-?))\s*-\s*(\d{2}FT\d+)\s+'
    r'Production\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)
AREA_TOTAL_RE = re.compile(
    r'(North|South)\s*-\s*(Topside|Subsea)\s+MPFM\s+Production\s+Total\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)


def parse_pdf(pdf_path, report_type='daily'):
    text = _flat_text(pdf_path)

    if 'FPSO South' in text[:300]:
        fpso_side = 'South'
    elif 'FPSO West' in text[:300]:
        fpso_side = 'West'
    elif 'FPSO North' in text[:300]:
        fpso_side = 'North'
    else:
        fpso_side = 'South'
    unit_type = 'Subsea' if 'Subsea' in text[:300] else 'Topside'

    if report_type == 'daily':
        dm = re.search(r'Daily Report from\s+([\d.]+)\s+[\d:]+\s+to\s+([\d.]+)', text)
        date_from = dm.group(1).replace('.', '-') if dm else None
        date_to = dm.group(2).replace('.', '-') if dm else None
        hour = None
        dt_from = dt_to = None
    else:
        hm = re.search(r'Hourly Report from\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)\s+to\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)', text)
        dt_from = f"{hm.group(1).replace('.', '-')} {hm.group(2)}" if hm else None
        dt_to = f"{hm.group(3).replace('.', '-')} {hm.group(4)}" if hm else None
        date_from = hm.group(1).replace('.', '-') if hm else None
        date_to = None
        hour = int(hm.group(2).split(':')[0]) if hm else None

    headers = list(TAG_RE.finditer(text))
    tags = {}
    for i, h in enumerate(headers):
        tag = h.group(1).strip().replace(' ', '_').rstrip('-')
        instr = h.group(2)
        block = text[h.start(): headers[i + 1].start() if i + 1 < len(headers) else None]
        tags[tag] = {'instrument': instr, 'metrics': _parse_metrics(block, 5)}

    am = AREA_TOTAL_RE.search(text)
    area_total = {}
    if am:
        block = text[am.start():]
        area_total = {'area': am.group(1), 'unit_type': am.group(2), 'metrics': _parse_metrics(block, 4)}

    return {
        'pdf_path': pdf_path, 'date_from': date_from, 'date_to': date_to,
        'dt_from': dt_from, 'dt_to': dt_to, 'hour': hour,
        'fpso_side': fpso_side, 'unit_type': unit_type,
        'tags': tags, 'area_total': area_total,
    }


def _parse_pdf_worker(task):
    pdf_path, report_type = task
    try:
        return pdf_path, parse_pdf(pdf_path, report_type), ""
    except Exception as exc:
        return pdf_path, None, str(exc)


def parse_pdf_batch(paths, report_type: str, workers: int):
    tasks = [(str(path), report_type) for path in paths]
    if workers == 1 or len(tasks) < 2:
        return [_parse_pdf_worker(task) for task in tasks]
    with ProcessPoolExecutor(max_workers=workers) as executor:
        return list(executor.map(_parse_pdf_worker, tasks, chunksize=8))


HOURLY_COLS = [
    'Dia ref.', 'Hora', 'Bank', 'Loop', 'Tipo', 'TAG', 'Instrumento', 'DT Início', 'DT Fim',
    'MPFM uncorr Gás (t)', 'MPFM uncorr Óleo (t)', 'MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)', 'MPFM uncorr Total (t)',
    'MPFM corr Gás (t)', 'MPFM corr Óleo (t)', 'MPFM corr HC (t)',
    'MPFM corr Água (t)', 'MPFM corr Total (t)',
    'PVT mass Gás (t)', 'PVT mass Óleo (t)', 'PVT mass Água (t)',
    'PVT vol Gás (Sm³)', 'PVT vol Óleo (m³)', 'PVT vol Água (m³)',
    'PVT @20 mass Gás (t)', 'PVT @20 mass Óleo (t)', 'PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)', 'PVT @20 vol Óleo (m³)', 'PVT @20 vol Água (m³)',
    'Pressão (barg)', 'Temperatura (°C)',
    'Dens. Gás (kg/m³)', 'Dens. Óleo (kg/m³)', 'Dens. Água (kg/m³)', 'Fonte',
]

DAILY_COLS = [
    'Bank', 'Loop', 'Tipo', 'TAG', 'Instrumento', 'Dia',
    'MPFM uncorr Gás (t)', 'MPFM uncorr Óleo (t)', 'MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)', 'MPFM uncorr Total (t)',
    'MPFM corr Gás (t)', 'MPFM corr Óleo (t)', 'MPFM corr HC (t)',
    'MPFM corr Água (t)', 'MPFM corr Total (t)',
    'PVT mass Gás (t)', 'PVT mass Óleo (t)', 'PVT mass Água (t)',
    'PVT vol Gás (Sm³)', 'PVT vol Óleo (m³)', 'PVT vol Água (m³)',
    'PVT @20 mass Gás (t)', 'PVT @20 mass Óleo (t)', 'PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)', 'PVT @20 vol Óleo (m³)', 'PVT @20 vol Água (m³)',
    'Pressão (barg)', 'Temperatura (°C)',
    'Dens. Gás (kg/m³)', 'Dens. Óleo (kg/m³)', 'Dens. Água (kg/m³)', 'Fonte (Daily)',
]

SEP_PHASE_COLUMNS = [
    "SEP Óleo Pressure (kPa)", "SEP Óleo Pressure (barg)", "SEP Óleo Temperature (°C)",
    "SEP Óleo SD (kg/sm³)", "SEP Óleo MD (kg/m³)", "SEP Óleo IV (m³)", "SEP Óleo GV (m³)",
    "SEP Óleo GSV (sm³)", "SEP Óleo Mass (t)", "SEP Óleo NSV (sm³)", "SEP Óleo BSW (%)",
    "SEP Óleo CPL", "SEP Óleo CTL",
    "SEP Água Pressure (kPa)", "SEP Água Pressure (barg)", "SEP Água Temperature (°C)",
    "SEP Água SD (kg/sm³)", "SEP Água MD (kg/m³)", "SEP Água IV (m³)", "SEP Água GV (m³)",
    "SEP Água GSV (sm³)", "SEP Água Mass (t)", "SEP Água NSV (sm³)", "SEP Água BSW (%)",
    "SEP Água CPL", "SEP Água CTL",
    "SEP Gás Pressure (kPa_g)", "SEP Gás Pressure (barg)", "SEP Gás Temperature (°C)",
    "SEP Gás SD (kg/sm³)", "SEP Gás DT (kg/m³)", "SEP Gás Gr. Vol. (m³)",
    "SEP Gás St. Vol. (m³)", "SEP Gás Mass (t)", "SEP Gás Energy (GJ)",
    "SEP Gás Diff. press. (kPa)", "SEP Gás Flowtime (min)",
]

SEP_COLS_HOURLY = [
    'SEP Temperatura Méd. (°C)', 'SEP Pressão Méd. (barg)', 'SEP Óleo Vol. Bruto (m³) CV',
    'SEP Óleo (t) CV', 'SEP Gás (t) CV', 'SEP Água (t) CV', 'SEP HC (t)', 'SEP Total (t)',
    *SEP_PHASE_COLUMNS,
    'Desvio HC (%)', 'Desvio Total (%)', 'Desvio Óleo (%)', 'Desvio Gás (%)', 'Desvio Água (%)',
]


def _mrow(m):
    u = m['mpfm_uncorr']
    c = m['mpfm_corr']
    pm = m['pvt_mass']
    pv = m['pvt_vol']
    p2 = m['pvt20_mass']
    v2 = m['pvt20_vol']
    f = m['fwa']
    return [
        u.get('gas'), u.get('oil'), u.get('hc'), u.get('water'), u.get('total'),
        c.get('gas'), c.get('oil'), c.get('hc'), c.get('water'), c.get('total'),
        pm.get('gas'), pm.get('oil'), pm.get('water'),
        pv.get('gas'), pv.get('oil'), pv.get('water'),
        p2.get('gas'), p2.get('oil'), p2.get('water'),
        v2.get('gas'), v2.get('oil'), v2.get('water'),
        f['pressure'], f['temperature'],
        f['dens_gas'], f['dens_oil'], f['dens_water'],
    ]


def build_daily_df(daily, unit_code, strict=False):
    import os
    rows = []
    for tag, td in daily['tags'].items():
        if not _mpfm_extraction_enabled(tag, td.get('instrument')):
            continue
        # Um PDF pode trazer várias seções (ex.: PE_4/18FT1506 e
        # PE_EO105/18FT1706 no B05). Só gera linha para o banco a que o
        # instrumento (TAG) realmente pertence.
        if not _tag_belongs_to_bank(td.get('instrument'), unit_code, strict):
            continue
        base = [unit_code, daily['fpso_side'], daily['unit_type'], tag, td['instrument'], daily['date_from']]
        rows.append(base + _mrow(td['metrics']) + [os.path.basename(daily['pdf_path'])])
    return pd.DataFrame(rows, columns=DAILY_COLS)


def build_recon_df(daily, hourly_records, unit_code, abs_tol=0.5, strict=False):
    hours_found = sorted(set(r['hour'] for r in hourly_records if r['hour'] is not None))
    hours_str = ','.join(f'{h:02d}' for h in hours_found)
    n = len(hours_found)
    coverage = 'OK (24/24h)' if n == 24 else (f'PARCIAL ({n}/24h)' if n > 0 else 'SEM HORÁRIOS')

    rows = []
    for tag, td in daily['tags'].items():
        if not _tag_belongs_to_bank(td.get('instrument'), unit_code, strict):
            continue
        d_c = td['metrics']['mpfm_corr']
        d_pv = td['metrics']['pvt_vol']
        s = {k: 0.0 for k in ['gas', 'oil', 'hc', 'water', 'total', 'pv_gas', 'pv_oil', 'pv_water']}
        for rec in hourly_records:
            if tag not in rec['tags']:
                continue
            hm = rec['tags'][tag]['metrics']

            def v(x):
                return x if isinstance(x, float) and not np.isnan(x) else 0.0
            for col in C5:
                s[col] += v(hm['mpfm_corr'].get(col, 0.0))
            s['pv_gas'] += v(hm['pvt_vol'].get('gas', 0.0))
            s['pv_oil'] += v(hm['pvt_vol'].get('oil', 0.0))
            s['pv_water'] += v(hm['pvt_vol'].get('water', 0.0))

        def delta(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)):
                return np.nan
            return round(sv - dv, 4)

        def status(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)):
                return '-'
            if n == 0:
                return 'SEM DADOS'
            d = abs(sv - dv)
            return 'OK' if d <= max(abs_tol, 0.0005 * abs(dv)) else 'VERIFICAR'

        row = {
            'Bank': unit_code, 'Loop': daily['fpso_side'], 'Tipo': daily['unit_type'],
            'TAG': tag, 'Instrumento': td['instrument'], 'Dia': daily['date_from'],
            'Cobertura': coverage, 'Horas': hours_str if n > 0 else '-',
            'Daily Gás (t)': d_c.get('gas'), 'Soma h. Gás (t)': round(s['gas'], 4),
            'Δ Gás (t)': delta(d_c.get('gas'), s['gas']), 'Status Gás': status(d_c.get('gas'), s['gas']),
            'Daily Óleo (t)': d_c.get('oil'), 'Soma h. Óleo (t)': round(s['oil'], 4),
            'Δ Óleo (t)': delta(d_c.get('oil'), s['oil']), 'Status Óleo': status(d_c.get('oil'), s['oil']),
            'Daily HC (t)': d_c.get('hc'), 'Soma h. HC (t)': round(s['hc'], 4),
            'Δ HC (t)': delta(d_c.get('hc'), s['hc']), 'Status HC': status(d_c.get('hc'), s['hc']),
            'Daily Água (t)': d_c.get('water'), 'Soma h. Água (t)': round(s['water'], 4),
            'Δ Água (t)': delta(d_c.get('water'), s['water']), 'Status Água': status(d_c.get('water'), s['water']),
            'Daily PVT Gás (Sm³)': d_pv.get('gas'), 'Soma h. PVT Gás (Sm³)': round(s['pv_gas'], 1),
            'Δ PVT Gás': delta(d_pv.get('gas'), s['pv_gas']),
            'Daily PVT Óleo (m³)': d_pv.get('oil'), 'Soma h. PVT Óleo (m³)': round(s['pv_oil'], 4),
            'Δ PVT Óleo': delta(d_pv.get('oil'), s['pv_oil']),
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _desvio(mpfm_val, sep_val):
    try:
        if mpfm_val is None or sep_val is None:
            return np.nan
        if np.isnan(float(mpfm_val)) or np.isnan(float(sep_val)):
            return np.nan
        if float(sep_val) == 0:
            return np.nan
        return round((float(mpfm_val) - float(sep_val)) / float(sep_val) * 100, 2)
    except Exception:
        return np.nan


def _sep_row(sep_hour, mpfm_hc_t, mpfm_total_t, mpfm_oil_t=np.nan, mpfm_gas_t=np.nan, mpfm_water_t=np.nan):
    if not sep_hour:
        return [np.nan] * len(SEP_COLS_HOURLY)
    oil_t = sep_hour.get('oil_t', np.nan)
    gas_t = sep_hour.get('gas_t', np.nan)
    wat_t = sep_hour.get('water_t', np.nan)
    hc_t = sep_hour.get('hc_t', np.nan)
    tot_t = sep_hour.get('total_t', np.nan)

    def _r(v):
        return round(v, 4) if (v is not None and not (isinstance(v, float) and np.isnan(v))) else np.nan
    return [
        sep_hour.get('temp', np.nan), sep_hour.get('pressure_barg', np.nan), sep_hour.get('oil_m3', np.nan),
        _r(oil_t), _r(gas_t), _r(wat_t), _r(hc_t), _r(tot_t),
        sep_hour.get('oil_pressure_kpa', np.nan), sep_hour.get('oil_pressure_barg', np.nan), sep_hour.get('oil_temp', np.nan),
        sep_hour.get('oil_sd', np.nan), sep_hour.get('oil_md', np.nan), sep_hour.get('oil_iv_m3', np.nan),
        sep_hour.get('oil_gv_m3', np.nan), sep_hour.get('oil_gsv_sm3', np.nan), sep_hour.get('oil_mass_t', np.nan), sep_hour.get('oil_nsv_sm3', np.nan),
        sep_hour.get('oil_bsw_pct', np.nan), sep_hour.get('oil_cpl', np.nan), sep_hour.get('oil_ctl', np.nan),
        sep_hour.get('water_pressure_kpa', np.nan), sep_hour.get('water_pressure_barg', np.nan), sep_hour.get('water_temp', np.nan),
        sep_hour.get('water_sd', np.nan), sep_hour.get('water_md', np.nan), sep_hour.get('water_iv_m3', np.nan),
        sep_hour.get('water_gv_m3', np.nan), sep_hour.get('water_gsv_sm3', np.nan), sep_hour.get('water_mass_t', np.nan), sep_hour.get('water_nsv_sm3', np.nan),
        sep_hour.get('water_bsw_pct', np.nan), sep_hour.get('water_cpl', np.nan), sep_hour.get('water_ctl', np.nan),
        sep_hour.get('gas_pressure_kpa_g', np.nan), sep_hour.get('gas_pressure_barg', np.nan), sep_hour.get('gas_temp', np.nan),
        sep_hour.get('gas_sd', np.nan), sep_hour.get('gas_dt', np.nan), sep_hour.get('gas_gr_vol_m3', np.nan),
        sep_hour.get('gas_st_vol_m3', np.nan), sep_hour.get('gas_mass_t', np.nan), sep_hour.get('gas_energy_gj', np.nan),
        sep_hour.get('gas_diff_press_kpa', np.nan), sep_hour.get('gas_flowtime_min', np.nan),
        _desvio(mpfm_hc_t, hc_t), _desvio(mpfm_total_t, tot_t),
        _desvio(mpfm_oil_t, oil_t), _desvio(mpfm_gas_t, gas_t), _desvio(mpfm_water_t, wat_t),
    ]


def build_hourly_df_with_sep(hourly_records, unit_code, sep_data=None, strict=False):
    import os
    cols = HOURLY_COLS + (SEP_COLS_HOURLY if sep_data else [])
    rows = []
    for rec in sorted(hourly_records, key=lambda r: r['hour'] or 0):
        for tag, td in rec['tags'].items():
            if not _mpfm_extraction_enabled(tag, td.get('instrument')):
                continue
            if not _tag_belongs_to_bank(td.get('instrument'), unit_code, strict):
                continue
            base = [rec['date_from'], rec['hour'], unit_code, rec['fpso_side'], rec['unit_type'],
                    tag, td['instrument'], rec['dt_from'], rec['dt_to']]
            mpfm_vals = _mrow(td['metrics'])
            row = base + mpfm_vals + [os.path.basename(rec['pdf_path'])]
            if sep_data:
                pdf_hour = rec['hour']
                txt_key = 24 if pdf_hour == 0 else pdf_hour
                sh = sep_data.get(txt_key)
                c = td['metrics']['mpfm_corr']
                mpfm_hc = c.get('hc')
                mpfm_tot = c.get('total')
                row += _sep_row(sh, mpfm_hc, mpfm_tot, c.get('oil'), c.get('gas'), c.get('water'))
            rows.append(row)
    return pd.DataFrame(rows, columns=cols)


_SEP_TOKEN_SPLIT_DECIMALS = (5, 4, 3, 6, 2)


def _is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def _recover_concatenated_float_token(token):
    raw = str(token or '').strip()
    if raw.count('.') < 2:
        return None
    if not re.fullmatch(r'[+-]?\d+(?:\.\d+)+', raw):
        return None
    first_dot = raw.find('.')
    for frac_len in _SEP_TOKEN_SPLIT_DECIMALS:
        split_at = first_dot + 1 + frac_len
        if split_at <= first_dot + 1 or split_at >= len(raw):
            continue
        left = raw[:split_at]
        right = raw[split_at:]
        if not right or not right[0].isdigit() or right.count('.') != 1:
            continue
        if _is_number(left) and _is_number(right):
            return left, right
    return None


def _parse_sep_float_token(raw_value, *, file_path, row_key, field_name, line):
    token = str(raw_value or '').strip()
    try:
        return float(token)
    except ValueError:
        recovered = _recover_concatenated_float_token(token)
        if not recovered:
            raise
        primary_token, _overflow_token = recovered
        return float(primary_token)


def _parse_hourly_txt_oleo(path):
    """Lê Run_24Hours*_OLEO.txt (ou _AGUA.txt, mesmo formato).
    Retorna {hour_int: {temp, pressure_barg, gv_m3, mass_t}, 'DAY': {...}}.
    hour_int: 1..24 (24 = última hora, equivale ao PDF hour=0 do dia seguinte)."""
    result = {}

    def _parse_liquid_values(parts, *, row_key, line):
        values = [p for p in parts[1:] if _is_number(p) or _recover_concatenated_float_token(p)]
        if len(values) >= 13:
            pressure_kpa = _parse_sep_float_token(values[0], file_path=path, row_key=row_key, field_name='pressure_kpa', line=line)
            pressure_barg = _parse_sep_float_token(values[1], file_path=path, row_key=row_key, field_name='pressure_barg', line=line)
            offset = 2
        elif len(values) >= 9:
            pressure_kpa = _parse_sep_float_token(values[0], file_path=path, row_key=row_key, field_name='pressure_kpa', line=line)
            pressure_barg = pressure_kpa / 100.0
            offset = 1
        else:
            return None
        return {
            'pressure_kpa': pressure_kpa,
            'pressure_barg': pressure_barg,
            'temp': _parse_sep_float_token(values[offset + 0], file_path=path, row_key=row_key, field_name='temp', line=line),
            'sd': _parse_sep_float_token(values[offset + 1], file_path=path, row_key=row_key, field_name='sd', line=line),
            'md': _parse_sep_float_token(values[offset + 2], file_path=path, row_key=row_key, field_name='md', line=line),
            'iv_m3': _parse_sep_float_token(values[offset + 3], file_path=path, row_key=row_key, field_name='iv_m3', line=line),
            'gv_m3': _parse_sep_float_token(values[offset + 4], file_path=path, row_key=row_key, field_name='gv_m3', line=line),
            'gsv_sm3': _parse_sep_float_token(values[offset + 5], file_path=path, row_key=row_key, field_name='gsv_sm3', line=line),
            'mass_t': _parse_sep_float_token(values[offset + 6], file_path=path, row_key=row_key, field_name='mass_t', line=line),
            'nsv_sm3': _parse_sep_float_token(values[offset + 7], file_path=path, row_key=row_key, field_name='nsv_sm3', line=line) if len(values) > offset + 7 else np.nan,
            'bsw_pct': _parse_sep_float_token(values[offset + 8], file_path=path, row_key=row_key, field_name='bsw_pct', line=line) if len(values) > offset + 8 else np.nan,
            'cpl': _parse_sep_float_token(values[offset + 9], file_path=path, row_key=row_key, field_name='cpl', line=line) if len(values) > offset + 9 else np.nan,
            'ctl': _parse_sep_float_token(values[offset + 10], file_path=path, row_key=row_key, field_name='ctl', line=line) if len(values) > offset + 10 else np.nan,
        }

    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAY':
                parsed = _parse_liquid_values(parts, row_key=key, line=line)
                if parsed:
                    result['DAY'] = parsed
            elif key.isdigit():
                h = int(key)
                parsed = _parse_liquid_values(parts, row_key=key, line=line)
                if parsed:
                    result[h] = parsed
    return result


def _parse_hourly_txt_gas(path):
    """Lê Run_24Hours*_GAS.txt. Retorna {hour_int: {mass_t}, 'DAY': {mass_t}}.
    O campo Mass no arquivo de gás está em kg -> convertido aqui para toneladas."""
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAILY':
                nums = [p for p in parts[1:] if _is_number(p)]
                if len(nums) >= 5:
                    result['DAY'] = {
                        'gr_vol_m3': _parse_sep_float_token(nums[0], file_path=path, row_key=key, field_name='gr_vol_m3', line=line),
                        'st_vol_m3': _parse_sep_float_token(nums[1], file_path=path, row_key=key, field_name='st_vol_m3', line=line),
                        'mass_t': _parse_sep_float_token(nums[2], file_path=path, row_key=key, field_name='mass_raw', line=line) / 1000.0,
                        'energy_gj': _parse_sep_float_token(nums[3], file_path=path, row_key=key, field_name='energy_gj', line=line),
                        'flowtime_min': _parse_sep_float_token(nums[4], file_path=path, row_key=key, field_name='flowtime_min', line=line),
                        'diff_press_kpa': _parse_sep_float_token(nums[5], file_path=path, row_key=key, field_name='diff_press_kpa', line=line) if len(nums) >= 6 else np.nan,
                    }
            elif key.isdigit():
                h = int(key)
                if len(parts) >= 8:
                    result[h] = {
                        'pressure_kpa_g': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_kpa_g', line=line),
                        'pressure_barg': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_kpa_g', line=line) / 100.0,
                        'temp': _parse_sep_float_token(parts[2], file_path=path, row_key=key, field_name='temp', line=line),
                        'sd': _parse_sep_float_token(parts[3], file_path=path, row_key=key, field_name='sd', line=line),
                        'dt': _parse_sep_float_token(parts[4], file_path=path, row_key=key, field_name='dt', line=line),
                        'gr_vol_m3': _parse_sep_float_token(parts[5], file_path=path, row_key=key, field_name='gr_vol_m3', line=line),
                        'st_vol_m3': _parse_sep_float_token(parts[6], file_path=path, row_key=key, field_name='st_vol_m3', line=line),
                        'mass_t': _parse_sep_float_token(parts[7], file_path=path, row_key=key, field_name='mass_raw', line=line) / 1000.0,
                        'energy_gj': _parse_sep_float_token(parts[8], file_path=path, row_key=key, field_name='energy_gj', line=line) if len(parts) >= 9 else np.nan,
                        'diff_press_kpa': _parse_sep_float_token(parts[9], file_path=path, row_key=key, field_name='diff_press_kpa', line=line) if len(parts) >= 10 else np.nan,
                        'flowtime_min': _parse_sep_float_token(parts[10], file_path=path, row_key=key, field_name='flowtime_min', line=line) if len(parts) >= 11 else np.nan,
                    }
    return result


def parse_sep_txt_set(oleo_path, gas_path, agua_path):
    """Lê os 3 TXT do separador de testes e monta um dict unificado:
    {hour: {temp, pressure_barg, oil_m3, oil_t, gas_t, water_t, hc_t, total_t}, 'DAY': {...}}."""
    oleo_data = _parse_hourly_txt_oleo(oleo_path)
    gas_data = _parse_hourly_txt_gas(gas_path)
    agua_data = _parse_hourly_txt_oleo(agua_path)

    combined = {}
    all_keys = sorted(set(list(oleo_data.keys()) + list(gas_data.keys()) + list(agua_data.keys())),
                       key=lambda x: 999 if x == 'DAY' else x)
    for k in all_keys:
        o = oleo_data.get(k, {})
        g = gas_data.get(k, {})
        w = agua_data.get(k, {})
        oil_m3 = o.get('gv_m3', np.nan)
        oil_t = o.get('mass_t', np.nan)
        gas_t = g.get('mass_t', np.nan)
        wat_t = w.get('mass_t', np.nan)

        def _s(a, b):
            if np.isnan(a) or np.isnan(b):
                return np.nan
            return a + b

        hc_t = _s(oil_t, gas_t)
        total_t = _s(hc_t, wat_t)
        combined[k] = {
            'temp': o.get('temp', np.nan), 'pressure_barg': o.get('pressure_barg', np.nan),
            'oil_m3': oil_m3, 'oil_t': oil_t, 'gas_t': gas_t, 'water_t': wat_t,
            'hc_t': hc_t, 'total_t': total_t,
            'oil_pressure_kpa': o.get('pressure_kpa', np.nan),
            'oil_pressure_barg': o.get('pressure_barg', np.nan),
            'oil_temp': o.get('temp', np.nan),
            'oil_sd': o.get('sd', np.nan),
            'oil_md': o.get('md', np.nan),
            'oil_iv_m3': o.get('iv_m3', np.nan),
            'oil_gv_m3': o.get('gv_m3', np.nan),
            'oil_gsv_sm3': o.get('gsv_sm3', np.nan),
            'oil_mass_t': o.get('mass_t', np.nan),
            'oil_nsv_sm3': o.get('nsv_sm3', np.nan),
            'oil_bsw_pct': o.get('bsw_pct', np.nan),
            'oil_cpl': o.get('cpl', np.nan),
            'oil_ctl': o.get('ctl', np.nan),
            'water_pressure_kpa': w.get('pressure_kpa', np.nan),
            'water_pressure_barg': w.get('pressure_barg', np.nan),
            'water_temp': w.get('temp', np.nan),
            'water_sd': w.get('sd', np.nan),
            'water_md': w.get('md', np.nan),
            'water_iv_m3': w.get('iv_m3', np.nan),
            'water_gv_m3': w.get('gv_m3', np.nan),
            'water_gsv_sm3': w.get('gsv_sm3', np.nan),
            'water_mass_t': w.get('mass_t', np.nan),
            'water_nsv_sm3': w.get('nsv_sm3', np.nan),
            'water_bsw_pct': w.get('bsw_pct', np.nan),
            'water_cpl': w.get('cpl', np.nan),
            'water_ctl': w.get('ctl', np.nan),
            'gas_pressure_kpa_g': g.get('pressure_kpa_g', np.nan),
            'gas_pressure_barg': g.get('pressure_barg', np.nan),
            'gas_temp': g.get('temp', np.nan),
            'gas_sd': g.get('sd', np.nan),
            'gas_dt': g.get('dt', np.nan),
            'gas_gr_vol_m3': g.get('gr_vol_m3', np.nan),
            'gas_st_vol_m3': g.get('st_vol_m3', np.nan),
            'gas_mass_t': g.get('mass_t', np.nan),
            'gas_energy_gj': g.get('energy_gj', np.nan),
            'gas_diff_press_kpa': g.get('diff_press_kpa', np.nan),
            'gas_flowtime_min': g.get('flowtime_min', np.nan),
        }
    return combined


def _sep_desvio_pct(mpfm_value, sep_value):
    try:
        if mpfm_value in (None, "") or sep_value in (None, ""):
            return ""
        mpfm_f = float(mpfm_value)
        sep_f = float(sep_value)
        if sep_f == 0:
            return ""
        return round((mpfm_f - sep_f) / sep_f * 100, 2)
    except Exception:
        return ""


# ═════════════════════════════════════════════════════════════════════════
# BASE_UNICA — colunas e mapeamentos (copiado de monthly_workbook_service.py)
# ═════════════════════════════════════════════════════════════════════════

BASE_UNICA_COLUMNS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Area", "System", "Bank", "Loop", "Tipo", "Entity", "Tag", "Instrumento", "PI Tag",
    "MPFM uncorr Gás (t)", "MPFM uncorr Óleo (t)", "MPFM uncorr HC (t)", "MPFM uncorr Água (t)", "MPFM uncorr Total (t)",
    "MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr HC (t)", "MPFM corr Água (t)", "MPFM corr Total (t)",
    "PVT mass Gás (t)", "PVT mass Óleo (t)", "PVT mass HC (t)", "PVT mass Água (t)", "PVT mass Total (t)",
    "PVT vol Gás (Sm³)", "PVT vol Óleo (m³)", "PVT vol HC (m³)", "PVT vol Água (m³)", "PVT vol Total (m³)",
    "PVT @20 mass Gás (t)", "PVT @20 mass Óleo (t)", "PVT @20 mass HC (t)", "PVT @20 mass Água (t)", "PVT @20 mass Total (t)",
    "PVT @20 vol Gás (Sm³)", "PVT @20 vol Óleo (m³)", "PVT @20 vol HC (m³)", "PVT @20 vol Água (m³)", "PVT @20 vol Total (m³)",
    "Pressão (barg)", "Temperatura (°C)", "Dens. Gás (kg/m³)", "Dens. Óleo (kg/m³)", "Dens. Água (kg/m³)",
    "SEP TAG", "SEP Medidor", "SEP Local", "SEP Status", "Bancos alinhados",
    "SEP Óleo Vol. Bruto (m³) CV", "SEP Óleo (t) CV", "SEP Gás (t) CV", "SEP Água (t) CV", "SEP HC (t)", "SEP Total (t)", "SEP Temperatura Méd. (°C)", "SEP Pressão Méd. (barg)",
    "SEP Óleo Pressure (kPa)", "SEP Óleo Pressure (barg)", "SEP Óleo Temperature (°C)", "SEP Óleo SD (kg/sm³)", "SEP Óleo MD (kg/m³)", "SEP Óleo IV (m³)", "SEP Óleo GV (m³)", "SEP Óleo GSV (sm³)", "SEP Óleo Mass (t)", "SEP Óleo NSV (sm³)", "SEP Óleo BSW (%)", "SEP Óleo CPL", "SEP Óleo CTL",
    "SEP Água Pressure (kPa)", "SEP Água Pressure (barg)", "SEP Água Temperature (°C)", "SEP Água SD (kg/sm³)", "SEP Água MD (kg/m³)", "SEP Água IV (m³)", "SEP Água GV (m³)", "SEP Água GSV (sm³)", "SEP Água Mass (t)", "SEP Água NSV (sm³)", "SEP Água BSW (%)", "SEP Água CPL", "SEP Água CTL",
    "SEP Gás Pressure (kPa_g)", "SEP Gás Pressure (barg)", "SEP Gás Temperature (°C)", "SEP Gás SD (kg/sm³)", "SEP Gás DT (kg/m³)", "SEP Gás Gr. Vol. (m³)", "SEP Gás St. Vol. (m³)", "SEP Gás Mass (t)", "SEP Gás Energy (GJ)", "SEP Gás Diff. press. (kPa)", "SEP Gás Flowtime (min)",
    "Desvio HC (%)", "Desvio Total (%)", "Desvio Óleo (%)", "Desvio Gás (%)", "Desvio Água (%)",
    "Recon Cobertura", "Recon Horas", "Recon Daily Gás (t)", "Recon Daily Óleo (t)", "Recon Daily HC (t)", "Recon Daily Água (t)",
    "Recon Soma h. Gás (t)", "Recon Soma h. Óleo (t)", "Recon Soma h. HC (t)", "Recon Soma h. Água (t)",
    "Recon Δ Gás (t)", "Recon Δ Óleo (t)", "Recon Δ HC (t)", "Recon Δ Água (t)",
    "Status Gás", "Status Óleo", "Status HC", "Status Água", "Fonte", "SourceFile", "IsOfficial",
]

MANUAL_SHEET_NAME = "COMPARATIVO_MANUAL"
COMPARATIVO_TOTAL_SHEET_NAME = "COMPARATIVO_TOTAL"
COMPARATIVO_PARES_SHEET_NAME = "COMPARATIVO_PARES"
LEGACY_COMPARATIVO_SEP_SHEET_NAME = "COMPARATIVO_SEP_LIVRE"
ALIGNED_COMPARATIVO_SEP_SHEET_NAME = "COMPARATIVO_SEP_ALINHADO"

PI_METER_TO_BANK = {
    "PE_2": "B10",
    "P2": "B08",
    "PW_104DA": "B15",
    "P4": "B13",
    "PE_4": "B05",
    "PE_4A": "B05",
    "PE-4": "B05",
    "PE-04": "B05",
    "PE_04": "B05",
    "P5": "B03",
    "PE_EO105": "B05",
    "PE-EO105": "B05",
    "18FT1706": "B05",
}

COMPARATIVO_COLUMNS = [
    "Data", "Hora", "Banco", "TAG", "Origem", "Pressão MPFM (barg)", "Temperatura MPFM (°C)",
    "Densidade Óleo (kg/m³)", "Densidade Gás (kg/m³)", "Densidade Água (kg/m³)",
    "Massa HC Não Corrigida (t)", "Massa Total Não Corrigida (t)", "Massa Óleo Não Corrigida (t)",
    "Massa Gás Não Corrigida (t)", "Massa Água Não Corrigida (t)",
    "Massa HC Corrigida (t)", "Massa Total Corrigida (t)", "Massa Óleo Corrigida (t)",
    "Massa Gás Corrigida (t)", "Massa Água Corrigida (t)",
    "Massa HC Padrão (t)", "Massa Total Padrão (t)", "Massa Óleo Padrão (t)",
    "Massa Gás Padrão (t)", "Massa Água Padrão (t)",
    "Volume Óleo STD 20°C, 1 atm (m³)", "Volume Gás STD 20°C, 1 atm (Sm³)",
    "Volume Água STD 20°C, 1 atm (m³)",
    "Velocidade Escoamento (m/s)", "GVF (%)", "ΔP - Inlet (mbar)", "ΔP - Outlet (mbar)", "WVF (%)", "WLR (%)", "GOR",
    "Water Conductivity (mS/cm)", "Water Conductivity Input (mS/cm)", "Meter Status 1", "Meter Status 2", "Flow Calculation Warn.",
    "Choke O?", "Pressão MPFM acima do ponto de bolha?", "Continuous Phase", "Calculation Mode", "Observações",
]

COMPARATIVO_NOTES = [
    ("Metodologia do desvio", ""),
    ("PE-02 × Riser P2", "% Desvio das métricas de massa = ((MPFM corrigido − referência) / referência) × 100. HC = óleo + gás; Total = óleo + gás + água. As fases comparadas individualmente são óleo, gás e água."),
    ("Volumes", "% Desvio = ((volume de fase MPFM corrigido @20 °C/1 atm − volume de fase referência @20 °C/1 atm) / volume de fase referência @20 °C/1 atm) × 100."),
    ("PE-04 × Riser P5", "Aplicar a mesma metodologia de desvio HC e Total."),
    ("PW-104 × Riser P4", "Aplicar a mesma metodologia de desvio HC e Total."),
    ("Referência", "Riser (topside). Limites de aceitação: ±10% (HC) e ±7% (Total)."),
    ("Separador de Testes", "Não possui trilha de correção independente; os valores de massa são repetidos nas colunas Corrigida e Padrão."),
]

RECON_MAP = {
    "Cobertura": "Recon Cobertura", "Horas": "Recon Horas",
    "Daily Gás (t)": "Recon Daily Gás (t)", "Daily Óleo (t)": "Recon Daily Óleo (t)",
    "Daily HC (t)": "Recon Daily HC (t)", "Daily Água (t)": "Recon Daily Água (t)",
    "Soma h. Gás (t)": "Recon Soma h. Gás (t)", "Soma h. Óleo (t)": "Recon Soma h. Óleo (t)",
    "Soma h. HC (t)": "Recon Soma h. HC (t)", "Soma h. Água (t)": "Recon Soma h. Água (t)",
    "Δ Gás (t)": "Recon Δ Gás (t)", "Δ Óleo (t)": "Recon Δ Óleo (t)",
    "Δ HC (t)": "Recon Δ HC (t)", "Δ Água (t)": "Recon Δ Água (t)",
    "Status Gás": "Status Gás", "Status Óleo": "Status Óleo",
    "Status HC": "Status HC", "Status Água": "Status Água",
}

SEP_COLS_PASSTHROUGH = [
    "SEP Temperatura Méd. (°C)", "SEP Pressão Méd. (barg)", "SEP Óleo Vol. Bruto (m³) CV",
    "SEP Óleo (t) CV", "SEP Gás (t) CV", "SEP Água (t) CV", "SEP HC (t)", "SEP Total (t)",
    *SEP_PHASE_COLUMNS,
    "Desvio HC (%)", "Desvio Total (%)", "Desvio Óleo (%)", "Desvio Gás (%)", "Desvio Água (%)",
]

_MPFM_METRIC_COLS = [
    "MPFM uncorr Gás (t)", "MPFM uncorr Óleo (t)", "MPFM uncorr HC (t)", "MPFM uncorr Água (t)", "MPFM uncorr Total (t)",
    "MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr HC (t)", "MPFM corr Água (t)", "MPFM corr Total (t)",
    "PVT mass Gás (t)", "PVT mass Óleo (t)", "PVT mass Água (t)",
    "PVT vol Gás (Sm³)", "PVT vol Óleo (m³)", "PVT vol Água (m³)",
    "PVT @20 mass Gás (t)", "PVT @20 mass Óleo (t)", "PVT @20 mass Água (t)",
    "PVT @20 vol Gás (Sm³)", "PVT @20 vol Óleo (m³)", "PVT @20 vol Água (m³)",
    "Pressão (barg)", "Temperatura (°C)", "Dens. Gás (kg/m³)", "Dens. Óleo (kg/m³)", "Dens. Água (kg/m³)",
]


def _new_row() -> dict:
    return {col: "" for col in BASE_UNICA_COLUMNS}


def _apply_mpfm_metadata(row: dict) -> dict:
    """Canonicaliza banco, entidade e TAG pelo instrumento físico."""
    metadata = _resolve_mpfm_metadata(row.get("Instrumento"), row.get("Entity"), row.get("Tag"))
    if metadata:
        row["Bank"] = metadata["bank"]
        row["Tipo"] = metadata["tipo"]
        row["Loop"] = metadata["loop"]
        row["Entity"] = metadata["entity"]
        row["Tag"] = metadata["entity"]
    elif row.get("Instrumento"):
        row["Tag"] = row.get("Instrumento")
    return row


def add_comparativo_sheet(workbook_path: Path, manual_rows: int = 200, replace_existing: bool = True):
    """Cria a aba fixa de preenchimento manual e preserva a aba Base_Unica."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(workbook_path)
    if MANUAL_SHEET_NAME in workbook.sheetnames:
        if not replace_existing:
            workbook.close()
            return
        del workbook[MANUAL_SHEET_NAME]
    sheet = workbook.create_sheet(MANUAL_SHEET_NAME)
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = True

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COMPARATIVO_COLUMNS))
    title = sheet.cell(row=1, column=1, value="COMPARATIVO MANUAL — PREENCHIMENTO PELO USUÁRIO")
    title.font = Font(bold=True, color="FFFFFF", size=12)
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center")
    sheet.cell(row=2, column=1, value="Campos mantidos em branco para preenchimento manual; fórmulas poderão ser incluídas posteriormente.")
    sheet.cell(row=2, column=1).font = Font(italic=True, color="666666")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, column_name in enumerate(COMPARATIVO_COLUMNS, start=1):
        cell = sheet.cell(row=3, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(14, min(len(column_name) + 3, 35))

    for row_index in range(4, 4 + manual_rows):
        for column_index in range(1, len(COMPARATIVO_COLUMNS) + 1):
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(vertical="center")

    notes_row = 6 + manual_rows
    sheet.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=len(COMPARATIVO_COLUMNS))
    notes_title = sheet.cell(row=notes_row, column=1, value="METODOLOGIA E REFERÊNCIAS")
    notes_title.font = Font(bold=True, color="FFFFFF")
    notes_title.fill = PatternFill("solid", fgColor="1F4E78")
    notes_title.alignment = Alignment(horizontal="center")

    for row_offset, (topic, description) in enumerate(COMPARATIVO_NOTES, start=1):
        row_index = notes_row + row_offset
        sheet.cell(row=row_index, column=1, value=topic).font = Font(bold=True)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=len(COMPARATIVO_COLUMNS))
        note_cell = sheet.cell(row=row_index, column=2, value=description)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(workbook_path)
    workbook.close()


def format_workbook(workbook_path: Path) -> None:
    """Aplica acabamento consistente às abas automáticas sem alterar fórmulas/dados."""
    from openpyxl import load_workbook
    from openpyxl.formatting.formatting import ConditionalFormattingList
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not workbook_path or not workbook_path.exists():
        return
    workbook = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="003B5C")
    header_font = Font(bold=True, color="FFFFFF")
    status_fill = PatternFill("solid", fgColor="FFF2CC")
    critical_fill = PatternFill("solid", fgColor="F4CCCC")
    good_fill = PatternFill("solid", fgColor="D9EAD3")
    for sheet in workbook.worksheets:
        if sheet.max_row < 1 or sheet.max_column < 1:
            continue
        # Reseta as regras antes de reaplicar: sem isso, cada execução acumula
        # novas regras sobre as antigas (nunca removidas), o que ao longo de
        # dezenas/centenas de execuções deixa o arquivo enorme e o load/save
        # extremamente lento.
        sheet.conditional_formatting = ConditionalFormattingList()
        sheet.sheet_view.showGridLines = False
        header_row = 3 if sheet.title == MANUAL_SHEET_NAME else 1
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        sheet.print_title_rows = f"${header_row}:${header_row}"
        headers = [str(sheet.cell(header_row, col).value or "") for col in range(1, sheet.max_column + 1)]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(header_row, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            values = [str(sheet.cell(row, col).value or "") for row in range(header_row + 1, min(sheet.max_row, header_row + 80) + 1)]
            width = min(max([len(header), *(len(value) for value in values)] or [12]) + 2, 38)
            sheet.column_dimensions[get_column_letter(col)].width = max(12, width)
            lower = header.lower()
            if "status" in lower or lower in {"classe", "issueflag"} or lower.startswith("gatilho "):
                data_range = f"{get_column_letter(col)}{header_row + 1}:{get_column_letter(col)}{sheet.max_row}"
                sheet.conditional_formatting.add(data_range, CellIsRule(operator="equal", formula=['"OK"'], fill=good_fill))
                sheet.conditional_formatting.add(data_range, CellIsRule(operator="equal", formula=['"CONFORME"'], fill=good_fill))
                sheet.conditional_formatting.add(data_range, CellIsRule(operator="equal", formula=['"ATIVO"'], fill=critical_fill))
                sheet.conditional_formatting.add(data_range, CellIsRule(operator="equal", formula=['"FORA DO LIMITE"'], fill=critical_fill))
                sheet.conditional_formatting.add(data_range, CellIsRule(operator="equal", formula=['"VERIFICAR"'], fill=status_fill))
    workbook.save(workbook_path)
    workbook.close()


def _to_float_or_blank(value):
    if value in (None, ""):
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, str):
        value = value.strip().replace(".", "").replace(",", ".") if "," in value else value.strip()
    try:
        return float(value)
    except Exception:
        return value


def _pi_day_iso(value) -> str:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    for pattern in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text[:19] if "%H" in pattern else text[:10], pattern).strftime("%Y-%m-%d")
        except Exception:
            pass
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%Y-%m-%d")
    except Exception:
        pass
    return text[:10]


def _build_pi_comparativo_lookup(df_pi: pd.DataFrame | None) -> dict:
    """Transforma PI_EXTRACT em lookup por (dia, banco) para variáveis operacionais.

    Quando a linha não traz um "PI Dia Coleta" único (caso do modo de coleta
    PERIODO_ESCOLHIDO, que não possui granularidade diária), o valor agregado
    do período é replicado para todos os dias do intervalo
    [PI Inicio, PI Final] do respectivo banco. Isso é uma aproximação
    documentada: o valor real é o mesmo em todo o período, então cada dia da
    janela recebe a mesma leitura até que uma extração diária esteja disponível.
    """
    if df_pi is None or df_pi.empty:
        return {}
    required = {"Medidor", "Grupo", "Variavel", "Canal", "Valor"}
    if not required.issubset(df_pi.columns):
        return {}
    df = df_pi.copy().where(pd.notna(df_pi), "")
    has_day_col = "PI Dia Coleta" in df.columns
    lookup = {}
    for _, row in df.iterrows():
        meter = str(row.get("Medidor", "")).strip()
        bank = PI_METER_TO_BANK.get(meter)
        if not bank:
            continue
        day_single = _pi_day_iso(row.get("PI Dia Coleta", "")) if has_day_col else ""
        if day_single:
            days = [day_single]
        else:
            start = _pi_day_iso(row.get("PI Inicio", ""))
            if not start:
                continue
            end = _pi_day_iso(row.get("PI Final", "")) or start
            try:
                days = [d.strftime("%Y-%m-%d") for d in pd.date_range(start, end)]
            except Exception:
                days = [start]
            if not days:
                days = [start]
        group = str(row.get("Grupo", "")).strip()
        variable = str(row.get("Variavel", "")).strip()
        channel = str(row.get("Canal", "")).strip()
        value = _to_float_or_blank(row.get("Valor", ""))
        for day in days:
            bucket = lookup.setdefault((day, bank), {})
            if group == "Other" and variable == "Velocity" and channel in {"Mix", "Liquid"}:
                bucket.setdefault("Velocidade Escoamento (m/s)", value)
            elif group == "Measured Fractions" and variable == "GVF":
                bucket["GVF (%)"] = value
            elif group == "Measured Fractions" and variable in {"WVF", "VVF"}:
                bucket["WVF (%)"] = value
            elif group == "Measured Fractions" and variable == "WLR":
                bucket["WLR (%)"] = value
            elif group == "Process Variables" and variable == "dP Inlet" and channel in {"Used", "A"}:
                bucket.setdefault("ΔP - Inlet (mbar)", value)
            elif group == "Other" and variable == "dP" and channel == "Inlet":
                bucket.setdefault("ΔP - Inlet (mbar)", value)
            elif group == "Process Variables" and variable == "dP Outlet":
                bucket.setdefault("ΔP - Outlet (mbar)", value)
            elif group == "Measured Fractions" and variable == "GOR":
                bucket["GOR"] = value
            elif group == "Process Variables" and variable == "Water Conductivity Input":
                bucket["Water Conductivity Input (mS/cm)"] = value
            elif group == "Process Variables" and variable == "Water Conductivity":
                bucket.setdefault("Water Conductivity (mS/cm)", value)
            elif group == "Status" and variable in {"Meter Status 1", "Meter Status 2", "Flow Calculation Warn."}:
                bucket[variable] = value
            elif group == "Calculation Modes" and variable == "Continuous Phase":
                current = bucket.setdefault("Continuous Phase", {})
                if isinstance(current, dict):
                    current[channel or "mPm"] = value
            elif group == "Calculation Modes" and variable == "Calculation Mode":
                current = bucket.setdefault("Calculation Mode", {})
                if isinstance(current, dict):
                    current[channel or "mPm"] = value
    for bucket in lookup.values():
        cp = bucket.get("Continuous Phase")
        if isinstance(cp, dict):
            bucket["Continuous Phase"] = "; ".join(f"{k}={v}" for k, v in cp.items())
        cm = bucket.get("Calculation Mode")
        if isinstance(cm, dict):
            bucket["Calculation Mode"] = "; ".join(f"{k}={v}" for k, v in cm.items())
    return lookup


def build_comparativo_total_df(df_base: pd.DataFrame | None, df_pi: pd.DataFrame | None = None) -> pd.DataFrame:
    """Monta aba automática COMPARATIVO_TOTAL a partir da Base_Unica + PI_EXTRACT."""
    if df_base is None or df_base.empty:
        return pd.DataFrame(columns=COMPARATIVO_COLUMNS)
    base = _normalize_master_columns(df_base)
    base = base.where(pd.notna(base), "")
    mask = (
        (base["Origin"].astype(str) == "MPFM")
        & (base["SourceType"].astype(str) == "PDF")
        & (base["Granularity"].astype(str) == "Daily")
    )
    base = base.loc[mask].copy()
    pi_lookup = _build_pi_comparativo_lookup(df_pi)
    rows = []
    for _, r in base.iterrows():
        day = str(r.get("ProductionDate", ""))[:10]
        bank = str(r.get("Bank", ""))
        pi = pi_lookup.get((day, bank), {})
        rows.append({
            "Data": day,
            "Hora": r.get("Hour", ""),
            "Banco": bank,
            "TAG": r.get("Tag", ""),
            "Origem": "BASE_UNICA_TOTAL + PI_EXTRACT_TOTAL" if pi_lookup else "BASE_UNICA_TOTAL",
            "Pressão MPFM (barg)": r.get("Pressão (barg)", ""),
            "Temperatura MPFM (°C)": r.get("Temperatura (°C)", ""),
            "Densidade Óleo (kg/m³)": r.get("Dens. Óleo (kg/m³)", ""),
            "Densidade Gás (kg/m³)": r.get("Dens. Gás (kg/m³)", ""),
            "Densidade Água (kg/m³)": r.get("Dens. Água (kg/m³)", ""),
            "Massa HC Não Corrigida (t)": r.get("MPFM uncorr HC (t)", ""),
            "Massa Total Não Corrigida (t)": r.get("MPFM uncorr Total (t)", ""),
            "Massa Óleo Não Corrigida (t)": r.get("MPFM uncorr Óleo (t)", ""),
            "Massa Gás Não Corrigida (t)": r.get("MPFM uncorr Gás (t)", ""),
            "Massa Água Não Corrigida (t)": r.get("MPFM uncorr Água (t)", ""),
            "Massa HC Corrigida (t)": r.get("MPFM corr HC (t)", ""),
            "Massa Total Corrigida (t)": r.get("MPFM corr Total (t)", ""),
            "Massa Óleo Corrigida (t)": r.get("MPFM corr Óleo (t)", ""),
            "Massa Gás Corrigida (t)": r.get("MPFM corr Gás (t)", ""),
            "Massa Água Corrigida (t)": r.get("MPFM corr Água (t)", ""),
            "Massa HC Padrão (t)": r.get("PVT @20 mass HC (t)", ""),
            "Massa Total Padrão (t)": r.get("PVT @20 mass Total (t)", ""),
            "Massa Óleo Padrão (t)": r.get("PVT @20 mass Óleo (t)", ""),
            "Massa Gás Padrão (t)": r.get("PVT @20 mass Gás (t)", ""),
            "Massa Água Padrão (t)": r.get("PVT @20 mass Água (t)", ""),
            "Volume Óleo STD 20°C, 1 atm (m³)": r.get("PVT @20 vol Óleo (m³)", ""),
            "Volume Gás STD 20°C, 1 atm (Sm³)": r.get("PVT @20 vol Gás (Sm³)", ""),
            "Volume Água STD 20°C, 1 atm (m³)": r.get("PVT @20 vol Água (m³)", ""),
            "Velocidade Escoamento (m/s)": pi.get("Velocidade Escoamento (m/s)", ""),
            "GVF (%)": pi.get("GVF (%)", ""),
            "ΔP - Inlet (mbar)": pi.get("ΔP - Inlet (mbar)", ""),
            "ΔP - Outlet (mbar)": pi.get("ΔP - Outlet (mbar)", ""),
            "WVF (%)": pi.get("WVF (%)", ""),
            "WLR (%)": pi.get("WLR (%)", ""),
            "GOR": pi.get("GOR", ""),
            "Water Conductivity (mS/cm)": pi.get("Water Conductivity (mS/cm)", ""),
            "Water Conductivity Input (mS/cm)": pi.get("Water Conductivity Input (mS/cm)", ""),
            "Meter Status 1": pi.get("Meter Status 1", ""),
            "Meter Status 2": pi.get("Meter Status 2", ""),
            "Flow Calculation Warn.": pi.get("Flow Calculation Warn.", ""),
            "Choke O?": "",
            "Pressão MPFM acima do ponto de bolha?": "",
            "Continuous Phase": pi.get("Continuous Phase", ""),
            "Calculation Mode": pi.get("Calculation Mode", ""),
            "Observações": "Preenchido automaticamente; colunas Choke/Ponto de bolha dependem de critério operacional manual.",
        })
    out = pd.DataFrame(rows)
    for col in COMPARATIVO_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    out = out[COMPARATIVO_COLUMNS].where(pd.notna(out), "")
    out.sort_values(by=["Data", "Banco", "TAG"], inplace=True, key=lambda col: col.astype(str))
    return out


def write_comparativo_total_sheet(workbook_path: Path, df_base: pd.DataFrame | None, df_pi: pd.DataFrame | None = None) -> pd.DataFrame:
    """Grava/substitui a aba automática COMPARATIVO_TOTAL."""
    df_comparativo = build_comparativo_total_df(df_base, df_pi)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_comparativo.to_excel(writer, sheet_name=COMPARATIVO_TOTAL_SHEET_NAME, index=False)
    else:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            df_comparativo.to_excel(writer, sheet_name=COMPARATIVO_TOTAL_SHEET_NAME, index=False)
    return df_comparativo


COMPARATIVO_PARES_COLUMNS = [
    "Dia", "Hora", "Granularidade", "Par", "Tag/Subsea", "Métrica",
    "Subsea valor", "Topside referência", "Unidade", "Desvio oficial (%)",
    "Limite", "Status", "Convenção do desvio", "Dias consecutivos fora", "Gatilho 3 dias", "Gatilho 6 dias", "Gatilho 10 dias",
]

def _remove_separator_comparison_sheets(workbook_path: Path) -> None:
    """Remove comparações SEP materializadas; a consulta existe somente no HTML."""
    if not workbook_path.exists():
        return
    from openpyxl import load_workbook
    workbook = load_workbook(workbook_path)
    changed = False
    for sheet_name in (LEGACY_COMPARATIVO_SEP_SHEET_NAME, ALIGNED_COMPARATIVO_SEP_SHEET_NAME):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
            changed = True
    if changed:
        workbook.save(workbook_path)
    workbook.close()


def build_comparativo_pares_df(
    df_base: pd.DataFrame | None,
    target_days: list | None = None,
    aligned_bank: str = SEP_ALIGNED_BANK,
) -> pd.DataFrame:
    """Cria trilha auditável dos desvios oficiais entre Subsea e Topside."""
    if df_base is None or df_base.empty:
        return pd.DataFrame(columns=COMPARATIVO_PARES_COLUMNS)
    base = _normalize_master_columns(df_base)
    days = target_days or sorted(base["ProductionDate"].astype(str).unique())
    official = _official_deviation_rows(base, days, aligned_bank)
    pair_rows = [row for row in official if "Subsea × Topside" in str(row.get("Métrica", ""))]
    result = []
    for row in pair_rows:
        metric_key = str(row.get("MetricaChave", ""))
        unit = "t" if metric_key in {"HC", "Total"} else ("Sm³ @20 °C/1 atm" if metric_key == "Gás vol" else "m³ @20 °C/1 atm")
        result.append({
            "Dia": row.get("Dia", ""),
            "Hora": row.get("Hora", ""),
            "Granularidade": row.get("Granularidade", ""),
            "Par": row.get("Banco", ""),
            "Tag/Subsea": row.get("Tag", ""),
            "Métrica": row.get("Métrica", ""),
            "Subsea valor": _to_float_or_blank(row.get("MPFMNum", "")),
            "Topside referência": _to_float_or_blank(row.get("ReferenciaNum", "")),
            "Unidade": unit,
            "Desvio oficial (%)": _to_float_or_blank(row.get("DesvioNum", "")),
            "Limite": row.get("Limite", ""),
            "Status": row.get("Status", ""),
            "Convenção do desvio": "(Subsea corrigido − Topside referência) / Topside referência × 100",
            "Dias consecutivos fora": row.get("Dias consecutivos fora", 0),
            "Gatilho 3 dias": row.get("Gatilho 3 dias", "—"),
            "Gatilho 6 dias": row.get("Gatilho 6 dias", "—"),
            "Gatilho 10 dias": row.get("Gatilho 10 dias", "—"),
        })
    present_pairs = {str(row.get("Par", "")) for row in result}
    for pair in COMPARISON_PAIRS:
        if pair["pair"] not in present_pairs:
            result.append({
                "Dia": "",
                "Hora": "",
                "Granularidade": "",
                "Par": pair["pair"],
                "Tag/Subsea": f"{pair['subsea_label']} × {pair['topside_label']}",
                "Métrica": "STATUS DO PAR",
                "Subsea valor": "",
                "Topside referência": "",
                "Unidade": "t",
                "Desvio oficial (%)": "",
                "Limite": "±10% HC / ±7% Total",
                "Status": "SEM DADOS VÁLIDOS — verificar linha zerada/poço fechado",
                "Convenção do desvio": "(Subsea corrigido − Topside referência) / Topside referência × 100",
                "Dias consecutivos fora": 0,
                "Gatilho 3 dias": "—",
                "Gatilho 6 dias": "—",
                "Gatilho 10 dias": "—",
            })
    out = pd.DataFrame(result, columns=COMPARATIVO_PARES_COLUMNS)
    if not out.empty:
        out.sort_values(by=["Dia", "Par", "Granularidade", "Hora", "Métrica"], inplace=True, key=lambda col: col.astype(str))
    return out.where(pd.notna(out), "")


def write_comparativo_pares_sheet(
    workbook_path: Path,
    df_base: pd.DataFrame | None,
    target_days: list | None = None,
    aligned_bank: str = SEP_ALIGNED_BANK,
) -> pd.DataFrame:
    df_pairs = build_comparativo_pares_df(df_base, target_days, aligned_bank)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_pairs.to_excel(writer, sheet_name=COMPARATIVO_PARES_SHEET_NAME, index=False)
    else:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            df_pairs.to_excel(writer, sheet_name=COMPARATIVO_PARES_SHEET_NAME, index=False)
    return df_pairs


def refresh_master_comparativo_pares(
    master_path: Path,
    aligned_bank: str = SEP_ALIGNED_BANK,
    df_base: pd.DataFrame | None = None,
) -> dict:
    if not master_path or not master_path.exists():
        return {"rows": 0}
    if df_base is None:
        df_base = read_master_base_unica(master_path)
    df_pairs = write_comparativo_pares_sheet(master_path, df_base, None, aligned_bank)
    return {"rows": len(df_pairs)}


def refresh_master_comparativo_total(master_path: Path, df_base: pd.DataFrame | None = None) -> dict:
    """Reconstrói COMPARATIVO_TOTAL usando BASE_UNICA_TOTAL e PI_EXTRACT_TOTAL."""
    if not master_path or not master_path.exists():
        return {"rows": 0}
    if df_base is None:
        df_base = read_master_base_unica(master_path)
    try:
        df_pi = pd.read_excel(master_path, sheet_name=PI_MASTER_SHEET_NAME, dtype=object)
        df_pi = df_pi.where(pd.notna(df_pi), "")
    except Exception:
        df_pi = pd.DataFrame()
    df_comp = write_comparativo_total_sheet(master_path, df_base, df_pi)
    return {"rows": len(df_comp)}


def hourly_df_to_rows(df, sep_merged: bool) -> list:
    if sep_merged:
        raise ValueError("A vinculação automática MPFM × SEP foi desativada; use a consulta do HTML.")
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia ref."], "Hour": r["Hora"], "Granularity": "Hourly",
            "Origin": "MPFM", "SourceType": "PDF", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "MPFM", "SourceFile": r["Fonte"], "IsOfficial": 1,
        })
        _apply_mpfm_metadata(row)
        for col in _MPFM_METRIC_COLS:
            row[col] = r[col]
        if sep_merged and _row_has_production(row):
            for col in SEP_COLS_PASSTHROUGH:
                if col in df.columns:
                    row[col] = r[col]
            row["SEP TAG"] = "SEP_Dados"
            row["SEP Medidor"] = "20FT0244/20FT0247/20FT0251"
            row["SEP Local"] = "Separador de Testes"
            row["SEP Status"] = "Alinhado"
            row["Bancos alinhados"] = row["Bank"]
        elif sep_merged:
            row["SEP Status"] = "Não aplicável — sem produção"
        rows.append(row)
    return rows


def daily_df_to_rows(df, sep_day: dict | None) -> list:
    if sep_day:
        raise ValueError("A vinculação automática MPFM × SEP foi desativada; use a consulta do HTML.")
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia"], "Hour": "", "Granularity": "Daily",
            "Origin": "MPFM", "SourceType": "PDF", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "MPFM", "SourceFile": r["Fonte (Daily)"], "IsOfficial": 1,
        })
        _apply_mpfm_metadata(row)
        for col in _MPFM_METRIC_COLS:
            row[col] = r[col]
        if sep_day and _row_has_production(row):
            mpfm_hc = row["MPFM corr HC (t)"]
            mpfm_tot = row["MPFM corr Total (t)"]
            mpfm_oil = row["MPFM corr Óleo (t)"]
            mpfm_gas = row["MPFM corr Gás (t)"]
            mpfm_water = row["MPFM corr Água (t)"]
            row["SEP Temperatura Méd. (°C)"] = sep_day.get("temp", "")
            row["SEP Pressão Méd. (barg)"] = sep_day.get("pressure_barg", "")
            row["SEP Óleo Vol. Bruto (m³) CV"] = sep_day.get("oil_m3", "")
            row["SEP Óleo (t) CV"] = sep_day.get("oil_t", "")
            row["SEP Gás (t) CV"] = sep_day.get("gas_t", "")
            row["SEP Água (t) CV"] = sep_day.get("water_t", "")
            row["SEP HC (t)"] = sep_day.get("hc_t", "")
            row["SEP Total (t)"] = sep_day.get("total_t", "")
            row["SEP Óleo Pressure (kPa)"] = sep_day.get("oil_pressure_kpa", "")
            row["SEP Óleo Pressure (barg)"] = sep_day.get("oil_pressure_barg", "")
            row["SEP Óleo Temperature (°C)"] = sep_day.get("oil_temp", "")
            row["SEP Óleo SD (kg/sm³)"] = sep_day.get("oil_sd", "")
            row["SEP Óleo MD (kg/m³)"] = sep_day.get("oil_md", "")
            row["SEP Óleo IV (m³)"] = sep_day.get("oil_iv_m3", "")
            row["SEP Óleo GV (m³)"] = sep_day.get("oil_gv_m3", "")
            row["SEP Óleo GSV (sm³)"] = sep_day.get("oil_gsv_sm3", "")
            row["SEP Óleo Mass (t)"] = sep_day.get("oil_mass_t", "")
            row["SEP Óleo NSV (sm³)"] = sep_day.get("oil_nsv_sm3", "")
            row["SEP Óleo BSW (%)"] = sep_day.get("oil_bsw_pct", "")
            row["SEP Óleo CPL"] = sep_day.get("oil_cpl", "")
            row["SEP Óleo CTL"] = sep_day.get("oil_ctl", "")
            row["SEP Água Pressure (kPa)"] = sep_day.get("water_pressure_kpa", "")
            row["SEP Água Pressure (barg)"] = sep_day.get("water_pressure_barg", "")
            row["SEP Água Temperature (°C)"] = sep_day.get("water_temp", "")
            row["SEP Água SD (kg/sm³)"] = sep_day.get("water_sd", "")
            row["SEP Água MD (kg/m³)"] = sep_day.get("water_md", "")
            row["SEP Água IV (m³)"] = sep_day.get("water_iv_m3", "")
            row["SEP Água GV (m³)"] = sep_day.get("water_gv_m3", "")
            row["SEP Água GSV (sm³)"] = sep_day.get("water_gsv_sm3", "")
            row["SEP Água Mass (t)"] = sep_day.get("water_mass_t", "")
            row["SEP Água NSV (sm³)"] = sep_day.get("water_nsv_sm3", "")
            row["SEP Água BSW (%)"] = sep_day.get("water_bsw_pct", "")
            row["SEP Água CPL"] = sep_day.get("water_cpl", "")
            row["SEP Água CTL"] = sep_day.get("water_ctl", "")
            row["SEP Gás Pressure (kPa_g)"] = sep_day.get("gas_pressure_kpa_g", "")
            row["SEP Gás Pressure (barg)"] = sep_day.get("gas_pressure_barg", "")
            row["SEP Gás Temperature (°C)"] = sep_day.get("gas_temp", "")
            row["SEP Gás SD (kg/sm³)"] = sep_day.get("gas_sd", "")
            row["SEP Gás DT (kg/m³)"] = sep_day.get("gas_dt", "")
            row["SEP Gás Gr. Vol. (m³)"] = sep_day.get("gas_gr_vol_m3", "")
            row["SEP Gás St. Vol. (m³)"] = sep_day.get("gas_st_vol_m3", "")
            row["SEP Gás Mass (t)"] = sep_day.get("gas_mass_t", "")
            row["SEP Gás Energy (GJ)"] = sep_day.get("gas_energy_gj", "")
            row["SEP Gás Diff. press. (kPa)"] = sep_day.get("gas_diff_press_kpa", "")
            row["SEP Gás Flowtime (min)"] = sep_day.get("gas_flowtime_min", "")
            row["Desvio HC (%)"] = _sep_desvio_pct(mpfm_hc, sep_day.get("hc_t"))
            row["Desvio Total (%)"] = _sep_desvio_pct(mpfm_tot, sep_day.get("total_t"))
            row["Desvio Óleo (%)"] = _sep_desvio_pct(mpfm_oil, sep_day.get("oil_t"))
            row["Desvio Gás (%)"] = _sep_desvio_pct(mpfm_gas, sep_day.get("gas_t"))
            row["Desvio Água (%)"] = _sep_desvio_pct(mpfm_water, sep_day.get("water_t"))
            row["SEP TAG"] = "SEP_Dados"
            row["SEP Medidor"] = "20FT0244/20FT0247/20FT0251"
            row["SEP Local"] = "Separador de Testes"
            row["SEP Status"] = "Alinhado"
            row["Bancos alinhados"] = row["Bank"]
        elif sep_day:
            row["SEP Status"] = "Não aplicável — sem produção"
        rows.append(row)
    return rows


def sep_data_to_rows(day: str, sep_data: dict | None) -> list:
    """Cria linhas próprias do separador para facilitar filtro por Bank=SEP."""
    if not sep_data:
        return []
    rows = []
    for key in sorted(sep_data.keys(), key=lambda value: 999 if value == "DAY" else int(value)):
        sep_rec = sep_data.get(key) or {}
        row = _new_row()
        hour = "" if key == "DAY" else (0 if int(key) == 24 else int(key))
        row.update({
            "ProductionDate": day,
            "Hour": hour,
            "Granularity": "Daily" if key == "DAY" else "Hourly",
            "Origin": "SEP",
            "SourceType": "TXT",
            "Bank": "SEP",
            "Loop": "Separador de Testes",
            "Tipo": "Separador",
            "Entity": "TAG 20VA121",
            "Tag": "TAG 20VA121",
            "Instrumento": "20VA121",
            "PI Tag": "20VA121",
            "SEP TAG": "TAG 20VA121",
            "SEP Medidor": "20VA121",
            "SEP Local": "Separador de Testes",
            "SEP Status": "Dado próprio do separador",
            "Bancos alinhados": "",
            "Fonte": "SEP",
            "SourceFile": "Run_24Hours FC13/FC14/FC17",
            "IsOfficial": 1,
        })
        for column, value in zip(SEP_COLS_HOURLY, _sep_row(sep_rec, np.nan, np.nan)):
            if column.startswith("Desvio"):
                continue
            row[column] = value
        rows.append(row)
    return rows


def recon_df_to_rows(df) -> list:
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia"], "Hour": "", "Granularity": "Daily",
            "Origin": "RECON", "SourceType": "CALC", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "Reconciliação", "SourceFile": "", "IsOfficial": 1,
        })
        for metric, column in RECON_MAP.items():
            if metric in df.columns:
                row[column] = r[metric]
        _apply_mpfm_metadata(row)
        rows.append(row)
    return rows


MASTER_SHEET_NAME = "BASE_UNICA_TOTAL"
MPFM_MEASUREMENTS_SHEET_NAME = "MPFM_MEDICOES"
SEP_OIL_SHEET_NAME = "SEP_CV_OLEO"
SEP_GAS_SHEET_NAME = "SEP_CV_GAS"
SEP_WATER_SHEET_NAME = "SEP_CV_AGUA"
RECONCILIATION_SHEET_NAME = "RECONCILIACAO"
METER_CATALOG_SHEET_NAME = "CADASTRO_MEDIDORES"
IMPORT_SOURCES_SHEET_NAME = "FONTES_IMPORTADAS"
IMPORT_LOG_SHEET_NAME = "LOG_IMPORTACAO"

# A Base Única histórica deixa de usar a aba larga como fonte primária. Estas
# listas definem as tabelas persistidas; BASE_UNICA_TOTAL é mantida apenas como
# compatibilidade durante a transição e para publicações legadas.
BASE_IDENTITY_COLUMNS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType",
    "Area", "System", "Bank", "Loop", "Tipo", "Entity", "Tag",
    "Instrumento", "PI Tag", "Fonte", "SourceFile", "IsOfficial",
]
SEP_IDENTITY_COLUMNS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType",
    "SEP TAG", "SEP Medidor", "SEP Local", "SEP Status", "SourceFile",
]
SEP_OIL_COLUMNS = [col for col in BASE_UNICA_COLUMNS if col.startswith("SEP Óleo ")]
SEP_GAS_COLUMNS = [col for col in BASE_UNICA_COLUMNS if col.startswith("SEP Gás ")]
SEP_WATER_COLUMNS = [col for col in BASE_UNICA_COLUMNS if col.startswith("SEP Água ")]
RECONCILIATION_COLUMNS = [
    *BASE_IDENTITY_COLUMNS,
    *[col for col in BASE_UNICA_COLUMNS if col.startswith("Desvio ") or col.startswith("Recon ") or col.startswith("Status ")],
]
MPFM_MEASUREMENT_COLUMNS = [
    *BASE_IDENTITY_COLUMNS,
    *[
        col for col in BASE_UNICA_COLUMNS
        if col not in BASE_IDENTITY_COLUMNS
        and not col.startswith("SEP ")
        and not col.startswith("Desvio ")
        and not col.startswith("Recon ")
        and not col.startswith("Status ")
        and col not in {"Bancos alinhados"}
    ],
]
IMPORT_SOURCES_COLUMNS = [
    "ImportID", "Arquivo", "Hash SHA-256", "Origem", "Aba", "Período inicial",
    "Período final", "Granularidade", "Registros", "Importado em", "Observações",
]
IMPORT_LOG_COLUMNS = [
    "ImportID", "Executado em", "Destino", "Registros lidos", "Inseridos",
    "Atualizados", "Duplicados", "Rejeitados", "Mensagem",
]
MASTER_DEDUP_KEYS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType",
    "Bank", "Entity", "Tag", "Instrumento",
]


def _normalize_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Garante a ordem e o conjunto esperado de colunas da Base_Unica."""
    if df is None or df.empty:
        return pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    out = df.copy()
    missing = [col for col in BASE_UNICA_COLUMNS if col not in out.columns]
    if missing:
        out = pd.concat([out, pd.DataFrame("", index=out.index, columns=missing)], axis=1)
    out = out[BASE_UNICA_COLUMNS]
    # Migra também históricos gravados como B17 ou com aliases PE_4A/PE-4,
    # para que uma republicação nunca reintroduza o banco inexistente.
    normalized_identity = {
        column: out[column].where(pd.notna(out[column]), "").astype(str).map(_normalize_instrument)
        for column in ("Instrumento", "Entity", "Tag")
    }
    already_resolved = pd.Series(False, index=out.index)
    # Primeira passagem: o instrumento físico tem precedência absoluta sobre
    # aliases possivelmente incorretos herdados em Entity/Tag.
    for instrument, metadata in MPFM_INSTRUMENT_METADATA.items():
        identity_mask = ~already_resolved & normalized_identity["Instrumento"].eq(instrument)
        out.loc[identity_mask, "Bank"] = metadata["bank"]
        out.loc[identity_mask, "Tipo"] = metadata["tipo"]
        out.loc[identity_mask, "Loop"] = metadata["loop"]
        out.loc[identity_mask, "Entity"] = metadata["entity"]
        out.loc[identity_mask, "Tag"] = metadata["entity"]
        out.loc[identity_mask, "Instrumento"] = instrument
        already_resolved |= identity_mask
    # Segunda passagem: aliases resolvem apenas linhas sem instrumento físico
    # reconhecido, evitando que PE_4 antigo vença um Instrumento=18FT1706.
    for instrument, metadata in MPFM_INSTRUMENT_METADATA.items():
        aliases = {alias for alias, canonical in MPFM_IDENTITY_ALIASES.items() if canonical == instrument}
        alias_match = normalized_identity["Entity"].isin(aliases) | normalized_identity["Tag"].isin(aliases)
        identity_mask = ~already_resolved & alias_match
        out.loc[identity_mask, "Bank"] = metadata["bank"]
        out.loc[identity_mask, "Tipo"] = metadata["tipo"]
        out.loc[identity_mask, "Loop"] = metadata["loop"]
        out.loc[identity_mask, "Entity"] = metadata["entity"]
        out.loc[identity_mask, "Tag"] = metadata["entity"]
        out.loc[identity_mask, "Instrumento"] = instrument
        already_resolved |= identity_mask
    # Migração de históricos: o SEP é fonte independente. Remove das linhas
    # MPFM qualquer massa, desvio ou marcador de alinhamento herdado, sem
    # alterar as linhas Origin=SEP extraídas diretamente dos TXT.
    mpfm_mask = out["Origin"].where(pd.notna(out["Origin"]), "").astype(str).eq("MPFM")
    sep_link_columns = [
        "SEP TAG", "SEP Medidor", "SEP Local", "SEP Status", "Bancos alinhados",
        *SEP_COLS_PASSTHROUGH,
    ]
    sep_link_columns = list(dict.fromkeys(sep_link_columns))
    out[sep_link_columns] = out[sep_link_columns].astype(object)
    out.loc[mpfm_mask, sep_link_columns] = ""
    return out


def _canonical_master_key_value(value, key: str) -> str:
    """Normaliza chaves lidas do Excel com tipos equivalentes (ex.: 12 e 12.0)."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "nat", "none"}:
        return ""
    if key == "Hour":
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.notna(numeric) and float(numeric).is_integer():
            return str(int(numeric))
    return text


def update_master_base_unica(master_path: Path, df_new: pd.DataFrame, replace_days: list | None = None) -> dict:
    """Integra incrementalmente a execução atual em um arquivo Base_Unica total.

    Regra: atualiza somente a aba BASE_UNICA_TOTAL. A aba COMPARATIVO_MANUAL,
    usada para preenchimento do usuário, é criada na primeira execução e depois
    preservada sem sobrescrever valores digitados. O histórico automatizado é
    combinado com a nova execução e deduplicado por chave técnica, mantendo a
    versão mais recente.
    """
    master_path.parent.mkdir(parents=True, exist_ok=True)
    df_current = pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    previous_rows = 0

    if master_path.exists():
        try:
            df_current = pd.read_excel(master_path, sheet_name=MASTER_SHEET_NAME, dtype=object)
        except ValueError:
            from openpyxl import load_workbook

            workbook = load_workbook(master_path, read_only=True, data_only=True)
            try:
                candidate_sheets = [name for name in workbook.sheetnames if name not in {MANUAL_SHEET_NAME, COMPARATIVO_TOTAL_SHEET_NAME}]
            finally:
                workbook.close()
            if not candidate_sheets:
                df_current = pd.DataFrame(columns=BASE_UNICA_COLUMNS)
            else:
                df_current = pd.read_excel(master_path, sheet_name=candidate_sheets[0], dtype=object)
        df_current = _normalize_master_columns(df_current)
        df_current = df_current.where(pd.notna(df_current), "")
        previous_rows = len(df_current)

    df_increment = _normalize_master_columns(df_new)
    df_increment = df_increment.where(pd.notna(df_increment), "")
    replaced_rows = 0
    if replace_days:
        replace_days_str = {str(day) for day in replace_days}
        keep_mask = ~df_current["ProductionDate"].astype(str).isin(replace_days_str)
        replaced_rows = int((~keep_mask).sum())
        df_current = df_current.loc[keep_mask].copy()
    combined = pd.concat([df_current, df_increment], ignore_index=True)

    for key in MASTER_DEDUP_KEYS:
        combined[key] = combined[key].map(lambda value, key=key: _canonical_master_key_value(value, key))

    combined.drop_duplicates(subset=MASTER_DEDUP_KEYS, keep="last", inplace=True)
    combined.sort_values(
        by=["ProductionDate", "Bank", "Granularity", "Hour", "Tag", "Origin"],
        inplace=True,
        na_position="last",
        key=lambda col: col.astype(str) if col.name != "Hour" else col,
    )
    combined = combined[BASE_UNICA_COLUMNS]

    if master_path.exists():
        with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name=MASTER_SHEET_NAME, index=False)
    else:
        combined.to_excel(master_path, sheet_name=MASTER_SHEET_NAME, index=False)

    add_comparativo_sheet(master_path, replace_existing=False)
    write_comparativo_total_sheet(master_path, combined, pd.DataFrame())
    return {
        "previous_rows": previous_rows,
        "increment_rows": len(df_increment),
        "total_rows": len(combined),
        "replaced_or_duplicate_rows": previous_rows + len(df_increment) - len(combined),
        "replaced_rows": replaced_rows,
        # Evita que quem chamou precise reabrir o arquivo inteiro do disco de
        # novo só para obter o mesmo conteúdo que acabamos de gravar.
        "df": combined,
    }


def read_master_base_unica(master_path: Path) -> pd.DataFrame:
    """Lê a aba consolidada da Base_Unica total, se existir."""
    if not master_path or not master_path.exists():
        return pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    try:
        df = pd.read_excel(master_path, sheet_name=MASTER_SHEET_NAME, dtype=object)
    except ValueError:
        from openpyxl import load_workbook

        workbook = load_workbook(master_path, read_only=True, data_only=True)
        try:
            candidate_sheets = [name for name in workbook.sheetnames if name not in {MANUAL_SHEET_NAME, COMPARATIVO_TOTAL_SHEET_NAME}]
        finally:
            workbook.close()
        if not candidate_sheets:
            return pd.DataFrame(columns=BASE_UNICA_COLUMNS)
        df = pd.read_excel(master_path, sheet_name=candidate_sheets[0], dtype=object)
    df = _normalize_master_columns(df)
    return df.where(pd.notna(df), "")


def _read_normalized_sheet(master_path: Path | None, sheet_name: str, columns: list[str]) -> pd.DataFrame:
    """Lê uma tabela normalizada e garante seu contrato de colunas."""
    frame = _read_master_sheet(master_path, sheet_name)
    if frame.empty:
        return pd.DataFrame(columns=columns)
    for column in columns:
        if column not in frame.columns:
            frame[column] = ""
    return frame[columns].where(pd.notna(frame[columns]), "")


def _write_or_replace_sheet(master_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    mode = "a" if master_path.exists() else "w"
    writer_args = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_args["if_sheet_exists"] = "replace"
    with pd.ExcelWriter(master_path, **writer_args) as writer:
        frame.to_excel(writer, sheet_name=sheet_name, index=False)


def _normalized_separator_context(master_path: Path | None) -> pd.DataFrame:
    """Reconstrói em memória o Separador a partir das três fases do C.V.

    Nenhuma aba consolidada é gravada: HC e Total são sempre derivados das
    massas de óleo, gás e água no carregamento do dashboard.
    """
    phase_specs = [
        (SEP_OIL_SHEET_NAME, SEP_OIL_COLUMNS, "óleo"),
        (SEP_GAS_SHEET_NAME, SEP_GAS_COLUMNS, "gás"),
        (SEP_WATER_SHEET_NAME, SEP_WATER_COLUMNS, "água"),
    ]
    keys = ["ProductionDate", "Hour", "Granularity"]
    phase_frames = []
    for sheet_name, phase_columns, _phase in phase_specs:
        frame = _read_normalized_sheet(master_path, sheet_name, [*SEP_IDENTITY_COLUMNS, *phase_columns])
        if not frame.empty:
            phase_frames.append(frame)
    if not phase_frames:
        return pd.DataFrame(columns=BASE_UNICA_COLUMNS)

    basis = pd.concat([frame[SEP_IDENTITY_COLUMNS] for frame in phase_frames], ignore_index=True)
    basis = basis.drop_duplicates(subset=keys, keep="last")
    output = basis.copy()
    for frame, (_sheet_name, phase_columns, _phase) in zip(phase_frames, [spec for spec in phase_specs if True]):
        payload = frame[[*keys, *phase_columns]].drop_duplicates(subset=keys, keep="last")
        output = output.merge(payload, on=keys, how="left")

    def _number(column: str) -> pd.Series:
        return pd.to_numeric(output.get(column, pd.Series(index=output.index, dtype=object)), errors="coerce")

    oil = _number("SEP Óleo Mass (t)")
    gas = _number("SEP Gás Mass (t)")
    water = _number("SEP Água Mass (t)")
    output["SEP Óleo (t) CV"] = oil
    output["SEP Gás (t) CV"] = gas
    output["SEP Água (t) CV"] = water
    output["SEP HC (t)"] = oil + gas
    output["SEP Total (t)"] = oil + gas + water
    output["Origin"] = "SEP"
    output["Fonte"] = "C.V. / Separador de Testes"
    output["IsOfficial"] = 1
    return _normalize_master_columns(output)


def read_dashboard_context(master_path: Path) -> pd.DataFrame:
    """Fornece a visão unificada ao HTML sem duplicar dados no Excel.

    Prefere o modelo normalizado quando ele existir. Mantém fallback integral
    para BASE_UNICA_TOTAL, permitindo migração gradual e reversível.
    """
    mpfm = _read_normalized_sheet(master_path, MPFM_MEASUREMENTS_SHEET_NAME, MPFM_MEASUREMENT_COLUMNS)
    recon = _read_normalized_sheet(master_path, RECONCILIATION_SHEET_NAME, RECONCILIATION_COLUMNS)
    sep = _normalized_separator_context(master_path)
    if mpfm.empty and recon.empty and sep.empty:
        return read_master_base_unica(master_path)
    context = pd.concat([mpfm, recon, sep], ignore_index=True, sort=False)
    context = _normalize_master_columns(context)
    for key in MASTER_DEDUP_KEYS:
        context[key] = context[key].map(lambda value, key=key: _canonical_master_key_value(value, key))
    return context.drop_duplicates(subset=MASTER_DEDUP_KEYS, keep="last").where(lambda value: pd.notna(value), "")


def migrate_master_to_normalized_model(master_path: Path, output_path: Path | None = None) -> dict:
    """Cria as tabelas normalizadas a partir da Base legada sem apagar abas.

    A função é idempotente: repetir a migração substitui apenas as abas do
    modelo novo, sempre reconstruídas da BASE_UNICA_TOTAL atual.
    """
    source_path = master_path
    destination_path = output_path or master_path
    legacy = read_master_base_unica(source_path)
    if legacy.empty:
        raise ValueError("BASE_UNICA_TOTAL não possui registros para migrar.")
    legacy = legacy.where(pd.notna(legacy), "")
    mpfm = legacy[legacy["Origin"].astype(str).eq("MPFM")].copy()
    recon = legacy[legacy["Origin"].astype(str).eq("RECON")].copy()
    sep = legacy[legacy["Origin"].astype(str).eq("SEP")].copy()
    mpfm = mpfm.reindex(columns=MPFM_MEASUREMENT_COLUMNS, fill_value="")
    mpfm_instrument_mask = mpfm["Instrumento"].where(pd.notna(mpfm["Instrumento"]), "").astype(str).str.strip().ne("")
    mpfm.loc[mpfm_instrument_mask, "Tag"] = mpfm.loc[mpfm_instrument_mask, "Instrumento"].astype(str).str.strip()
    recon = recon.reindex(columns=RECONCILIATION_COLUMNS, fill_value="")

    common = sep.reindex(columns=SEP_IDENTITY_COLUMNS, fill_value="")
    oil = pd.concat([common, sep.reindex(columns=SEP_OIL_COLUMNS, fill_value="")], axis=1)
    gas = pd.concat([common, sep.reindex(columns=SEP_GAS_COLUMNS, fill_value="")], axis=1)
    water = pd.concat([common, sep.reindex(columns=SEP_WATER_COLUMNS, fill_value="")], axis=1)

    catalog_rows = []
    actual_meters = legacy[legacy["Origin"].astype(str).eq("MPFM")][
        ["Instrumento", "Bank", "Entity", "Tag", "Tipo", "Loop"]
    ].drop_duplicates()
    for _, meter in actual_meters.iterrows():
        catalog_rows.append({
            "Instrumento": meter.get("Instrumento", ""), "Banco": meter.get("Bank", ""),
            "Entidade": meter.get("Entity", ""), "TAG": meter.get("Instrumento", "") or meter.get("Tag", ""),
            "Tipo": meter.get("Tipo", ""), "Loop": meter.get("Loop", ""),
            "Ativo": "SIM", "Observações": "Migrado da Base Única legada",
        })
    for instrument, metadata in MPFM_INSTRUMENT_METADATA.items():
        if any(str(row["Instrumento"]) == instrument for row in catalog_rows):
            continue
        catalog_rows.append({
            "Instrumento": instrument, "Banco": metadata["bank"], "Entidade": metadata["entity"],
            "TAG": instrument, "Tipo": metadata["tipo"], "Loop": metadata["loop"],
            "Ativo": "SIM", "Observações": "Cadastro técnico da automação",
        })
    catalog = pd.DataFrame(catalog_rows)
    source = pd.DataFrame([{
        "ImportID": f"LEGACY-{datetime.now():%Y%m%d%H%M%S}", "Arquivo": source_path.name,
        "Hash SHA-256": _source_frame_fingerprint(legacy), "Origem": "MIGRAÇÃO LEGADA",
        "Aba": MASTER_SHEET_NAME, "Período inicial": str(legacy["ProductionDate"].min()),
        "Período final": str(legacy["ProductionDate"].max()), "Granularidade": "Daily + Hourly",
        "Registros": len(legacy), "Importado em": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "Observações": "Estrutura normalizada criada sem remover a aba legada.",
    }], columns=IMPORT_SOURCES_COLUMNS)
    log = pd.DataFrame([{
        "ImportID": source.iloc[0]["ImportID"], "Executado em": source.iloc[0]["Importado em"],
        "Destino": "Modelo normalizado", "Registros lidos": len(legacy), "Inseridos": len(legacy),
        "Atualizados": 0, "Duplicados": 0, "Rejeitados": 0,
        "Mensagem": "Migração inicial da BASE_UNICA_TOTAL; origem legada preservada.",
    }], columns=IMPORT_LOG_COLUMNS)

    if destination_path.resolve() != source_path.resolve():
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination_path)
    # Abrir/salvar o workbook uma única vez é essencial: a Base contém o
    # histórico PI e reabrir o Excel para cada aba deixa a migração muito lenta.
    normalized_sheets = [
        (MPFM_MEASUREMENTS_SHEET_NAME, mpfm), (SEP_OIL_SHEET_NAME, oil),
        (SEP_GAS_SHEET_NAME, gas), (SEP_WATER_SHEET_NAME, water),
        (RECONCILIATION_SHEET_NAME, recon), (METER_CATALOG_SHEET_NAME, catalog),
        (IMPORT_SOURCES_SHEET_NAME, source), (IMPORT_LOG_SHEET_NAME, log),
    ]
    with pd.ExcelWriter(destination_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, frame in normalized_sheets:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    return {"mpfm": len(mpfm), "sep": len(sep), "recon": len(recon), "total": len(legacy), "output": str(destination_path)}


def _read_master_sheet(master_path: Path | None, sheet_name: str) -> pd.DataFrame:
    """Lê uma aba histórica sem interromper a geração se ela ainda não existir."""
    if not master_path or not master_path.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(master_path, sheet_name=sheet_name, dtype=object)
        return df.where(pd.notna(df), "")
    except (ValueError, FileNotFoundError, OSError):
        return pd.DataFrame()


def read_master_alarm_events(master_path: Path | None) -> pd.DataFrame:
    return _normalize_alarm_event_columns(_read_master_sheet(master_path, ALARM_EVENT_MASTER_SHEET_NAME))


def read_master_pi_extract(master_path: Path | None) -> pd.DataFrame:
    df = _read_master_sheet(master_path, PI_MASTER_SHEET_NAME)
    if "Variavel" in df.columns:
        df = df[df["Variavel"].astype(str).str.strip().isin(PI_CONTOUR_VARIABLES)].copy()
    return df


def loaded_days_in_master(master_path: Path, target_days: list, daily_by_bank: dict) -> set:
    """Retorna dias já consolidados para todos os bancos disponíveis no dia.

    A checagem usa linhas Daily/MPFM/PDF, que são a evidência mais barata de que
    o dia já passou pela automação. Se faltar qualquer banco esperado, o dia é
    reprocessado para completar a base.
    """
    df = read_master_base_unica(master_path)
    if df.empty:
        return set()

    for col in ("ProductionDate", "Granularity", "Origin", "SourceType", "Bank", "Instrumento"):
        df[col] = df[col].astype(str)

    loaded = set()
    daily_mask = (
        (df["Granularity"] == "Daily")
        & (df["Origin"] == "MPFM")
        & (df["SourceType"] == "PDF")
    )
    df_daily = df.loc[daily_mask]
    for day in target_days:
        expected_banks = {bank for bank, recs in daily_by_bank.items() if day in recs}
        if not expected_banks:
            continue
        day_rows = df_daily.loc[df_daily["ProductionDate"] == day]
        banks_in_master = set(day_rows["Bank"].astype(str))
        complete = expected_banks.issubset(banks_in_master)
        for bank in expected_banks:
            record = daily_by_bank.get(bank, {}).get(day)
            metadata = record[1] if isinstance(record, (tuple, list)) and len(record) > 1 and isinstance(record[1], dict) else {}
            expected_instruments = {
                _normalize_instrument(tag)
                for tag in metadata.get("tags", [])
                if _mpfm_extraction_enabled(tag, tag) and _tag_belongs_to_bank(tag, bank)
            }
            expected_instruments.discard("")
            if not expected_instruments:
                continue
            present = set(day_rows.loc[day_rows["Bank"].astype(str) == bank, "Instrumento"].map(_normalize_instrument))
            if not expected_instruments.issubset(present):
                complete = False
                break
        if complete:
            loaded.add(day)
    return loaded


def _normalize_alarm_event_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=ALARM_EVENT_COLUMNS)
    out = df.copy()
    for col in ALARM_EVENT_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[ALARM_EVENT_COLUMNS]


def update_master_alarm_events(master_path: Path, df_new: pd.DataFrame, replace_days: list | None = None) -> dict:
    """Atualiza a aba consolidada de alarmes/eventos, substituindo dias reprocessados."""
    if df_new is None or df_new.empty:
        return {"previous_rows": 0, "increment_rows": 0, "total_rows": 0, "replaced_rows": 0}
    master_path.parent.mkdir(parents=True, exist_ok=True)
    df_current = pd.DataFrame(columns=ALARM_EVENT_COLUMNS)
    previous_rows = 0
    if master_path.exists():
        try:
            df_current = pd.read_excel(master_path, sheet_name=ALARM_EVENT_MASTER_SHEET_NAME, dtype=object)
        except ValueError:
            df_current = pd.DataFrame(columns=ALARM_EVENT_COLUMNS)
        df_current = _normalize_alarm_event_columns(df_current).where(pd.notna(df_current), "")
        previous_rows = len(df_current)

    df_increment = _normalize_alarm_event_columns(df_new)
    df_increment = df_increment.where(pd.notna(df_increment), "")
    replaced_rows = 0
    if replace_days:
        replace_days_str = {str(day) for day in replace_days}
        keep_mask = ~df_current["ProductionDate"].astype(str).isin(replace_days_str)
        replaced_rows = int((~keep_mask).sum())
        df_current = df_current.loc[keep_mask].copy()

    combined = pd.concat([df_current, df_increment], ignore_index=True)
    for col in ("ProductionDate", "Timestamp", "RecordType", "SourceKind", "Object", "Description", "SourceFile"):
        combined[col] = combined[col].where(pd.notna(combined[col]), "").astype(str)
    combined.drop_duplicates(
        subset=["ProductionDate", "Timestamp", "RecordType", "SourceKind", "Object", "Description", "SourceFile"],
        keep="last", inplace=True,
    )
    combined.sort_values(by=["ProductionDate", "Timestamp", "RecordType", "Bank"], inplace=True, key=lambda col: col.astype(str))
    combined = combined[ALARM_EVENT_COLUMNS]

    if master_path.exists():
        with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name=ALARM_EVENT_MASTER_SHEET_NAME, index=False)
    else:
        combined.to_excel(master_path, sheet_name=ALARM_EVENT_MASTER_SHEET_NAME, index=False)
    return {
        "previous_rows": previous_rows,
        "increment_rows": len(df_increment),
        "total_rows": len(combined),
        "replaced_rows": replaced_rows,
    }


# ═════════════════════════════════════════════════════════════════════════
# CAMINHOS DE PASTA — ajuste aqui se a estrutura da sua instalação for diferente
# ═════════════════════════════════════════════════════════════════════════

_FILENAME_DATE_RE = re.compile(r"-(\d{8})-")


def _filename_date_iso(path: Path):
    m = _FILENAME_DATE_RE.search(path.name)
    if not m:
        return None
    raw = m.group(1)
    return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"


def _month_dir(parent: Path, month: int) -> Path:
    """Retorna a pasta do mês aceitando pequenas variações no nome.

    Ex.: "08. Agosto" (padrão) e "03.Março" (sem espaço após o ponto).
    """
    expected = parent / MONTH_PT[month]
    if expected.is_dir():
        return expected
    prefix = f"{month:02d}."
    matches = [path for path in parent.glob(f"{prefix}*") if path.is_dir()]
    return matches[0] if matches else expected


def _year_dir(root: Path, year: int) -> Path:
    """Aceita raiz no nível pai (.../Daily Reports) ou já no ano (.../2026)."""
    if root.name == str(year):
        return root
    return root / str(year)


def bank_month_dir(mpfm_root: Path, bank_code: str, year: int, month: int, sub: str) -> Path:
    configured_root = mpfm_root / BANK_FOLDERS[bank_code]
    bank_year_dir = _year_dir(configured_root, year)
    return _month_dir(bank_year_dir, month) / sub


def sep_day_dir(sep_root: Path, day_iso: str, fc_folder: str) -> Path:
    dt = datetime.strptime(day_iso, "%Y-%m-%d")
    sep_year_dir = _year_dir(sep_root, dt.year)
    day_folder_name = f"FPSO-Bacalhau_Daily reports_{day_iso}"
    day_folder = _month_dir(sep_year_dir, dt.month) / day_folder_name
    candidates = [
        day_folder / "01 - CV_Reports" / fc_folder,
        day_folder / day_folder_name / "01 - CV_Reports" / fc_folder,
    ]
    for candidate in candidates:
        if candidate.is_dir():
            return candidate
    return candidates[0]


def _months_to_scan(reference: datetime, back: int = 1):
    months = []
    y, m = reference.year, reference.month
    for _ in range(back + 1):
        months.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return months


def _months_in_range(date_from: datetime, date_to: datetime):
    months = []
    year, month = date_from.year, date_from.month
    while (year, month) <= (date_to.year, date_to.month):
        months.append((year, month))
        month += 1
        if month == 13:
            year += 1
            month = 1
    return months


def _parse_date_argument(value: str, option_name: str):
    if not value:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    raise ValueError(f"{option_name} deve usar DD/MM/AAAA ou AAAA-MM-DD.")


def _pi_period_for_day(day_iso: str) -> tuple[datetime, datetime]:
    start = datetime.strptime(str(day_iso), "%Y-%m-%d")
    end = start + timedelta(days=1) - timedelta(seconds=1)
    return start, end


def _iso_days_between(date_from: datetime, date_to: datetime) -> list[str]:
    days = []
    cursor = date_from
    while cursor.date() <= date_to.date():
        days.append(cursor.strftime("%Y-%m-%d"))
        cursor += timedelta(days=1)
    return days


def _read_pi_period_record(period_path: Path) -> dict:
    if not period_path.exists():
        return {}
    try:
        return json.loads(period_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}


def _load_pi_extract(pi_output_path: Path, period_record: dict | None = None) -> pd.DataFrame:
    """Lê a saída normalizada do PI Vision e acrescenta metadados da coleta."""
    if not pi_output_path.exists():
        raise FileNotFoundError(f"Arquivo PI não encontrado: {pi_output_path}")
    df = pd.read_excel(pi_output_path, dtype=object)
    df = df.where(pd.notna(df), "")
    if "Variavel" in df.columns:
        df = df[df["Variavel"].astype(str).str.strip().isin(PI_CONTOUR_VARIABLES)].copy()
    period_record = period_record or {}
    df.insert(0, "PI Inicio", period_record.get("inicio", ""))
    df.insert(1, "PI Final", period_record.get("final", ""))
    df.insert(2, "PI Dia Coleta", period_record.get("inicio", "")[:10])
    df.insert(3, "PI Modo", period_record.get("modo", ""))
    df.insert(4, "PI Registrado Em", period_record.get("registrado_em", ""))
    df.insert(5, "PI URL Aplicada", period_record.get("url_aplicada", ""))
    df.insert(6, "PI Arquivo Origem", str(pi_output_path))
    return df


def _daily_control_v5_paths(pi_root: Path) -> dict[str, Path]:
    """Resolve caminhos da revisão Daily Control V5 enviada pelo usuário."""
    daily_root = Path(os.environ.get("PI_DAILY_CONTROL_ROOT", "").strip() or PI_DAILY_CONTROL_ROOT)
    history = Path(os.environ.get("PI_DAILY_CONTROL_OUTPUT", "").strip()) if os.environ.get("PI_DAILY_CONTROL_OUTPUT", "").strip() else daily_root / "output" / "Historico_Daily_Control_V5.xlsx"
    return {
        "daily_root": daily_root,
        "collector": pi_root / PI_DAILY_CONTROL_COLLECTOR,
        "config": pi_root / PI_DAILY_CONTROL_CONFIG,
        "input": daily_root / "input",
        "old": daily_root / "input" / "old",
        "output": daily_root / "output",
        "normalizer": daily_root / PI_DAILY_CONTROL_NORMALIZER,
        "normalizer_core": daily_root / PI_DAILY_CONTROL_NORMALIZER_CORE,
        "history": history,
    }


def _daily_control_v5_available(pi_root: Path) -> bool:
    paths = _daily_control_v5_paths(pi_root)
    required = ("collector", "config", "normalizer", "normalizer_core")
    return all(paths[name].exists() for name in required)


def _daily_control_meta(row: pd.Series) -> dict:
    return {
        "RunID": row.get("RunID", ""),
        "DataHora": row.get("ExtractionDateTime", ""),
        "Medidor": str(row.get("Medidor", "")).replace("\\_", "_"),
        "Tag": row.get("Tag", ""),
        "Tipo": row.get("Tipo", ""),
        "SourceFile": row.get("SourceFile", ""),
    }


def _daily_control_status(row: pd.Series) -> str:
    status = str(row.get("OverallStatus", "")).strip().upper()
    if status in {"GOOD", "OK", "SUCESSO", "APROVADO"}:
        return "OK"
    if status in {"WARNING", "BAD", "ERRO", "FALHA"}:
        return status
    return "OK"


def _load_daily_control_v5_extract(history_path: Path, start: datetime, end: datetime) -> pd.DataFrame:
    """Converte o histórico Daily Control V5 em linhas compatíveis com PI_EXTRACT."""
    if not history_path.exists():
        raise FileNotFoundError(f"Histórico Daily Control V5 não encontrado: {history_path}")
    try:
        sheets = pd.read_excel(history_path, sheet_name=None, dtype=object)
    except ValueError as exc:
        raise ValueError(f"Histórico Daily Control V5 inválido: {history_path}") from exc
    period_start = start.strftime("%d/%m/%Y %H:%M:%S")
    period_end = end.strftime("%d/%m/%Y %H:%M:%S")
    day_iso = start.strftime("%Y-%m-%d")
    registered_at = datetime.now().astimezone().isoformat()
    rows = []

    def emit(meta: dict, group: str, variable: str, channel: str, value, unit: str = "", source_sheet: str = ""):
        if variable not in PI_CONTOUR_VARIABLES:
            return
        if value in (None, ""):
            return
        try:
            if pd.isna(value):
                return
        except Exception:
            pass
        rows.append({
            "PI Inicio": period_start,
            "PI Final": period_end,
            "PI Dia Coleta": day_iso,
            "PI Modo": "DAILY_CONTROL_V5",
            "PI Registrado Em": registered_at,
            "PI URL Aplicada": PI_DAILY_CONTROL_URL,
            "PI Arquivo Origem": str(history_path),
            "PI Status Coleta": _daily_control_status(pd.Series(meta)),
            "RunID": meta.get("RunID", ""),
            "DataHora": meta.get("DataHora", ""),
            "Medidor": meta.get("Medidor", ""),
            "Tag": meta.get("Tag", ""),
            "Tipo": meta.get("Tipo", ""),
            "Grupo": group,
            "Variavel": variable,
            "Canal": channel,
            "Valor": value,
            "Unidade": unit,
            "Qualidade": "GOOD" if str(value).strip().lower() not in {"bad", "nan", "none", ""} else "BAD",
            "PI Source Sheet": source_sheet,
            "SourceFile": meta.get("SourceFile", ""),
        })

    mappings = {
        "FlowRates": [],
        "Envelope": [
            ("WLR", "Measured Fractions", "WLR", "Used", "%"),
            ("WVF", "Measured Fractions", "WVF", "Used", "%"),
            ("GVF", "Measured Fractions", "GVF", "Used", "%"),
            ("GOR", "Measured Fractions", "GOR", "Used", ""),
            ("Temperature_C", "Process Variables", "Temperature", "Used", "degC"),
            ("Pressure_bar", "Process Variables", "Pressure", "Used", "bar"),
            ("DP_Inlet_mbar", "Process Variables", "dP Inlet", "Used", "mbar"),
            ("DP_Outlet_mbar", "Process Variables", "dP Outlet", "Used", "mbar"),
            ("Velocity_Gas_ms", "Other", "Velocity", "Gas", "m/s"),
            ("Velocity_Mix_ms", "Other", "Velocity", "Mix", "m/s"),
            ("Velocity_Liq_ms", "Other", "Velocity", "Liquid", "m/s"),
            ("WaterConductivity_Used_mScm", "Process Variables", "Water Conductivity", "Used", "mS/cm"),
            ("WaterConductivity_Input_mScm", "Process Variables", "Water Conductivity Input", "Input", "mS/cm"),
        ],
        "Status": [
            ("MeterStatus1", "Status", "Meter Status 1", "", ""),
            ("MeterStatus2", "Status", "Meter Status 2", "", ""),
            ("FlowCalculationWarning", "Status", "Flow Calculation Warn.", "", ""),
        ],
        "CalculationModes": [
            ("ContinuousPhaseOil", "Calculation Modes", "Continuous Phase", "Oil", ""),
            ("ContinuousPhaseWater", "Calculation Modes", "Continuous Phase", "Water", ""),
            ("CalculationModeMultiphase", "Calculation Modes", "Calculation Mode", "Multiphase", ""),
            ("CalculationModeWetGas", "Calculation Modes", "Calculation Mode", "Wet Gas", ""),
            ("CalculationModeAutoSwitch", "Calculation Modes", "Calculation Mode", "Auto Switch", ""),
            ("WaterConductivityInputMode", "Calculation Modes", "Water Conductivity Input", "", ""),
        ],
        "Accumulators": [],
    }
    for sheet, sheet_mappings in mappings.items():
        df = sheets.get(sheet, pd.DataFrame())
        if df.empty:
            continue
        df = df.where(pd.notna(df), "")
        for _, row in df.iterrows():
            meta = _daily_control_meta(row)
            meta["OverallStatus"] = row.get("OverallStatus", "")
            for col, group, variable, channel, unit in sheet_mappings:
                if col in row.index:
                    emit(meta, group, variable, channel, row.get(col), unit, sheet)

    quality = sheets.get("LogQualidade", pd.DataFrame())
    if quality is not None and not quality.empty:
        quality = quality.where(pd.notna(quality), "")
        for _, row in quality.iterrows():
            meta = _daily_control_meta(row)
            rows.append({
                "PI Inicio": period_start,
                "PI Final": period_end,
                "PI Dia Coleta": day_iso,
                "PI Modo": "DAILY_CONTROL_V5",
                "PI Registrado Em": registered_at,
                "PI URL Aplicada": PI_DAILY_CONTROL_URL,
                "PI Arquivo Origem": str(history_path),
                "PI Status Coleta": "FALHA",
                "PI Erro": row.get("Erro", ""),
                "RunID": meta.get("RunID", ""),
                "DataHora": meta.get("DataHora", ""),
                "Medidor": meta.get("Medidor", ""),
                "Tag": meta.get("Tag", ""),
                "Tipo": meta.get("Tipo", ""),
                "Grupo": "LogQualidade",
                "Variavel": "Erro",
                "Canal": "",
                "Valor": row.get("Erro", ""),
                "Qualidade": "BAD",
                "PI Source Sheet": "LogQualidade",
                "SourceFile": meta.get("SourceFile", ""),
            })
    return pd.DataFrame(rows).where(lambda data: pd.notna(data), "")


def _archive_daily_control_input(input_dir: Path, old_dir: Path) -> None:
    input_dir.mkdir(parents=True, exist_ok=True)
    old_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for file in input_dir.glob("*.json"):
        target = old_dir / f"{stamp}_{file.name}"
        shutil.move(str(file), str(target))


def _run_daily_control_v5_collection(pi_root: Path, start: datetime, end: datetime) -> pd.DataFrame:
    """Executa a revisão PI Daily Control V5.2 + normalizador V5.1."""
    paths = _daily_control_v5_paths(pi_root)
    missing = [name for name in ("collector", "config", "normalizer", "normalizer_core") if not paths[name].exists()]
    if missing:
        details = ", ".join(f"{name}={paths[name]}" for name in missing)
        raise FileNotFoundError(f"Daily Control V5 não instalado/completo: {details}")
    _ensure_pi_edge_cdp(pi_root, paths["config"])
    _archive_daily_control_input(paths["input"], paths["old"])
    start_text = start.strftime("%d/%m/%Y %H:%M:%S")
    end_text = end.strftime("%d/%m/%Y %H:%M:%S")
    collector_cmd = [
        sys.executable,
        str(paths["collector"]),
        "--config",
        str(paths["config"]),
        "--start",
        start_text,
        "--end",
        end_text,
        "--output-dir",
        str(paths["input"]),
    ]
    print(f"\n[INFO] Coletando PI Daily Control V5 no período {start_text} a {end_text}...")
    try:
        completed = subprocess.run(collector_cmd, cwd=str(pi_root), timeout=PI_COLLECTION_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired as exc:
        raise TimeoutError(f"Coleta Daily Control V5 excedeu {PI_COLLECTION_TIMEOUT_SECONDS}s sem concluir.") from exc
    if completed.returncode != 0:
        raise RuntimeError(f"Coleta Daily Control V5 retornou código {completed.returncode}.")
    json_count = len(list(paths["input"].glob("*.json")))
    if json_count != PI_DAILY_CONTROL_EXPECTED_JSONS:
        raise RuntimeError(f"Daily Control V5 gerou {json_count} JSON(s); esperado: {PI_DAILY_CONTROL_EXPECTED_JSONS}.")
    paths["output"].mkdir(parents=True, exist_ok=True)
    normalizer_cmd = [
        sys.executable,
        str(paths["normalizer"]),
        "--core-script",
        str(paths["normalizer_core"]),
        "--input-dir",
        str(paths["input"]),
        "--output-dir",
        str(paths["output"]),
        "--period-start",
        start_text,
        "--period-end",
        end_text,
    ]
    try:
        completed = subprocess.run(normalizer_cmd, cwd=str(paths["daily_root"]), timeout=PI_COLLECTION_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired as exc:
        raise TimeoutError(f"Normalização Daily Control V5 excedeu {PI_COLLECTION_TIMEOUT_SECONDS}s sem concluir.") from exc
    if completed.returncode != 0:
        raise RuntimeError(f"Normalização Daily Control V5 retornou código {completed.returncode}.")
    df_pi = _load_daily_control_v5_extract(paths["history"], start, end)
    print(f"[OK] PI Daily Control V5 importado: {len(df_pi)} linha(s) de {paths['history']}")
    return df_pi


def _latest_pi_failure_record(pi_output_path: Path) -> dict:
    """Lê o diagnóstico de reprovação mais recente do coletor PI, quando existir."""
    diag_dir = pi_output_path.parent / "diagnostico_v47"
    if not diag_dir.exists():
        return {}
    candidates = sorted(diag_dir.glob("Validacao_V49_REPROVADA_*.json"), key=lambda path: path.stat().st_mtime, reverse=True)
    if not candidates:
        return {}
    latest = candidates[0]
    try:
        data = json.loads(latest.read_text(encoding="utf-8-sig"))
    except Exception:
        data = {}
    data["diagnostico_arquivo"] = str(latest)
    return data


def _pi_failure_dataframe(day_iso: str, start: datetime, end: datetime, attempt: int, exc: Exception, pi_output_path: Path) -> pd.DataFrame:
    """Cria uma linha rastreável na aba PI_EXTRACT para dias não coletados."""
    failure = _latest_pi_failure_record(pi_output_path)
    errors = failure.get("erros") if isinstance(failure.get("erros"), list) else []
    row = {
        "PI Inicio": start.strftime("%d/%m/%Y %H:%M:%S"),
        "PI Final": end.strftime("%d/%m/%Y %H:%M:%S"),
        "PI Dia Coleta": day_iso,
        "PI Modo": "PERIODO_ESCOLHIDO",
        "PI Registrado Em": datetime.now().astimezone().isoformat(),
        "PI URL Aplicada": "",
        "PI Arquivo Origem": str(pi_output_path),
        "PI Status Coleta": "FALHA",
        "PI Tentativas": attempt,
        "PI Erro": str(exc),
        "PI Diagnóstico": failure.get("diagnostico_arquivo", ""),
        "PI Validação Erros": " | ".join(str(item) for item in errors),
        "PI Objetos": failure.get("objetos", ""),
        "PI Campos": failure.get("campos", ""),
    }
    return pd.DataFrame([row]).where(lambda data: pd.notna(data), "")


def _pi_retry_config(pi_root: Path, config_name: str, attempt: int) -> tuple[str, Path | None]:
    """Cria uma configuração temporária com esperas maiores nas retentativas."""
    if attempt <= 1 or _daily_control_v5_available(pi_root):
        return config_name, None
    config_path = Path(config_name)
    if not config_path.is_absolute():
        config_path = pi_root / config_path
    cfg = json.loads(config_path.read_text(encoding="utf-8-sig"))
    retry_index = attempt - 1
    cfg["period_settle_seconds"] = int(cfg.get("period_settle_seconds", 12)) + retry_index * PI_RETRY_PERIOD_SETTLE_INCREMENT_SECONDS
    cfg["settle_seconds"] = int(cfg.get("settle_seconds", 8)) + retry_index * PI_RETRY_COLLECTOR_SETTLE_INCREMENT_SECONDS
    tmp = tempfile.NamedTemporaryFile("w", suffix=".json", prefix="base_unica_pi_retry_", dir=str(pi_root), delete=False, encoding="utf-8")
    try:
        json.dump(cfg, tmp, ensure_ascii=False, indent=2)
    finally:
        tmp.close()
    return tmp.name, Path(tmp.name)


def _pi_edge_executable() -> Path | None:
    """Localiza o executável do Microsoft Edge em instalações Windows comuns."""
    configured = os.environ.get("PI_EDGE_PATH", "").strip()
    if configured and Path(configured).exists():
        return Path(configured)
    candidates = []
    for env_name in ("PROGRAMFILES", "PROGRAMFILES(X86)", "LOCALAPPDATA"):
        root = os.environ.get(env_name, "").strip()
        if root:
            candidates.append(Path(root) / "Microsoft" / "Edge" / "Application" / "msedge.exe")
    found = shutil.which("msedge.exe") or shutil.which("msedge")
    if found:
        candidates.insert(0, Path(found))
    return next((path for path in candidates if path.exists()), None)


def _pi_configured_url(config_path: Path) -> str:
    """Procura no config PI a URL do display PI Vision."""
    configured_from_env = os.environ.get("PI_VISION_URL", "").strip()
    if configured_from_env.lower().startswith(("http://", "https://")):
        return configured_from_env
    try:
        config = json.loads(config_path.read_text(encoding="utf-8-sig"))
    except Exception:
        return ""

    def walk(value):
        if isinstance(value, dict):
            for key, item in value.items():
                key_text = str(key).lower()
                if isinstance(item, str) and item.lower().startswith(("http://", "https://")):
                    if any(token in key_text for token in ("url", "vision", "display", "meter")):
                        return item.strip()
                found = walk(item)
                if found:
                    return found
        elif isinstance(value, list):
            for item in value:
                found = walk(item)
                if found:
                    return found
        return ""

    configured = walk(config) or ""
    # O config também pode conter a URL local do CDP. Ela nunca é o display.
    if configured and not re.search(r"127\.0\.0\.1:9222|localhost:9222|/json/(list|version)", configured, re.I):
        return configured
    return PI_VISION_URL


def _pi_display_urls(config_path: Path) -> list[str]:
    urls = [_pi_configured_url(config_path)]
    daily_control = os.environ.get("PI_DAILY_CONTROL_URL", "").strip() or PI_DAILY_CONTROL_URL
    if PI_OPEN_DAILY_CONTROL and daily_control.lower().startswith(("http://", "https://")):
        urls.append(daily_control)
    return list(dict.fromkeys(urls))


def _pi_cdp_available(host: str = PI_CDP_HOST, port: int = PI_CDP_PORT) -> bool:
    try:
        with socket.create_connection((host, port), timeout=1.5):
            return True
    except OSError:
        return False


def _pi_open_configured_tab(config_path: Path, host: str = PI_CDP_HOST, port: int = PI_CDP_PORT) -> None:
    """Garante as abas Metering Monitor e Daily Control no browser CDP."""
    target_urls = _pi_display_urls(config_path)
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as playwright:
            browser = playwright.chromium.connect_over_cdp(f"http://{host}:{port}")
            contexts = browser.contexts
            context = contexts[0] if contexts else browser.new_context()
            for target_url in target_urls:
                pages = context.pages
                target = next((page for page in pages if target_url.rstrip("/") in page.url.rstrip("/")), None)
                if target is None:
                    target = context.new_page()
                    target.goto(target_url, wait_until="domcontentloaded", timeout=60000)
                    print(f"[OK] Aba PI Vision aberta automaticamente: {target_url}")
                else:
                    print(f"[OK] Aba PI Vision já disponível: {target.url}")
    except Exception as exc:
        print(f"[WARN] Edge/CDP respondeu, mas não foi possível preparar a aba PI Vision automaticamente: {exc}")


def _pi_capture_daily_control_snapshot(pi_root: Path, day: str, host: str = PI_CDP_HOST, port: int = PI_CDP_PORT) -> None:
    """Salva evidência visual/HTML do display Daily Control para mapeamento futuro."""
    if not PI_CAPTURE_DAILY_CONTROL:
        return
    output_dir = pi_root / "saida_v4" / "daily_control_evidence"
    output_dir.mkdir(parents=True, exist_ok=True)
    image_path = output_dir / f"Metering-Daily-Control_{day}.png"
    html_path = output_dir / f"Metering-Daily-Control_{day}.html"
    manifest_path = output_dir / "variables_manifest.json"
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as playwright:
            browser = playwright.chromium.connect_over_cdp(f"http://{host}:{port}")
            context = browser.contexts[0] if browser.contexts else browser.new_context()
            page = next((item for item in context.pages if "Displays/56466" in item.url or "Metering-Daily-Control" in item.url), None)
            if page is None:
                print("[WARN] Display Daily Control não foi localizado para captura.")
                return
            page.wait_for_timeout(5000)
            page.screenshot(path=str(image_path), full_page=True)
            html_path.write_text(page.content(), encoding="utf-8")
            manifest_path.write_text(json.dumps({"display_url": PI_DAILY_CONTROL_URL, "day": day, "variables": PI_DAILY_CONTROL_VARIABLES}, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"📸 Evidência Daily Control salva: {image_path}")
    except Exception as exc:
        print(f"[WARN] Não foi possível capturar o Daily Control para mapeamento: {exc}")


def _ensure_pi_edge_cdp(pi_root: Path, config_path: Path) -> None:
    """Abre Edge com CDP quando necessário e prepara a aba antes do coletor."""
    host = os.environ.get("PI_CDP_HOST", PI_CDP_HOST).strip() or PI_CDP_HOST
    try:
        port = int(os.environ.get("PI_CDP_PORT", PI_CDP_PORT))
    except ValueError:
        port = PI_CDP_PORT
    if not _pi_cdp_available(host, port):
        edge = _pi_edge_executable()
        if edge is None:
            raise FileNotFoundError("Microsoft Edge não encontrado. Defina PI_EDGE_PATH ou instale o Edge.")
        profile = os.environ.get("PI_EDGE_PROFILE_DIR", "").strip() or PI_EDGE_PROFILE_DIR.strip()
        if not profile:
            local_app_data = os.environ.get("LOCALAPPDATA", str(Path.home()))
            profile = str(Path(local_app_data) / "PI_Vision_CDP")
        Path(profile).mkdir(parents=True, exist_ok=True)
        target_url = _pi_configured_url(config_path)
        command = [str(edge), f"--remote-debugging-port={port}", f"--user-data-dir={profile}", "--new-window"]
        command.append(target_url or "about:blank")
        print(f"[INFO] Edge PI Vision não estava disponível; abrindo com CDP em {host}:{port}...")
        creationflags = getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
        subprocess.Popen(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, creationflags=creationflags)
        deadline = time.monotonic() + PI_EDGE_START_TIMEOUT_SECONDS
        while time.monotonic() < deadline and not _pi_cdp_available(host, port):
            time.sleep(0.5)
        if not _pi_cdp_available(host, port):
            raise RuntimeError(f"Edge foi iniciado, mas a porta CDP {host}:{port} não respondeu em {PI_EDGE_START_TIMEOUT_SECONDS}s.")
        print(f"[OK] Edge disponível para automação em {host}:{port}.")
    _pi_open_configured_tab(config_path, host, port)


def write_pi_extract_sheet(workbook_path: Path, df_pi: pd.DataFrame, sheet_name: str = PI_SHEET_NAME):
    """Grava a extração PI e um resumo auditável sem apagar as demais abas."""
    summary = pd.DataFrame([
        {"Item": "Linhas extraídas", "Valor": len(df_pi)},
        {"Item": "Medidores", "Valor": df_pi["Medidor"].nunique() if "Medidor" in df_pi.columns else 0},
        {"Item": "Variáveis autorizadas", "Valor": df_pi["Variavel"].nunique() if "Variavel" in df_pi.columns else 0},
        {"Item": "Dias coletados", "Valor": df_pi["PI Dia Coleta"].nunique() if "PI Dia Coleta" in df_pi.columns else 0},
        {"Item": "Período", "Valor": f"{df_pi['PI Dia Coleta'].min()} a {df_pi['PI Dia Coleta'].max()}" if "PI Dia Coleta" in df_pi.columns and not df_pi.empty else "—"},
        {"Item": "Fonte(s)", "Valor": "; ".join(sorted(df_pi["PI Arquivo Origem"].dropna().astype(str).unique())) if "PI Arquivo Origem" in df_pi.columns else "—"},
        {"Item": "URL(s) aplicada(s)", "Valor": "; ".join(sorted(df_pi["PI URL Aplicada"].dropna().astype(str).unique())) if "PI URL Aplicada" in df_pi.columns else "—"},
    ])
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    if workbook_path.exists():
        with pd.ExcelWriter(workbook_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_pi.to_excel(writer, sheet_name=sheet_name, index=False)
            summary.to_excel(writer, sheet_name="PI_EXTRACT_RESUMO", index=False)
    else:
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            df_pi.to_excel(writer, sheet_name=sheet_name, index=False)
            summary.to_excel(writer, sheet_name="PI_EXTRACT_RESUMO", index=False)


def update_master_pi_extract(master_path: Path, df_pi: pd.DataFrame) -> dict:
    """Atualiza a aba consolidada do PI no mesmo arquivo Base_Unica total."""
    if df_pi is None or df_pi.empty:
        return {"previous_rows": 0, "increment_rows": 0, "total_rows": 0}
    master_path.parent.mkdir(parents=True, exist_ok=True)
    if "Variavel" in df_pi.columns:
        df_pi = df_pi[df_pi["Variavel"].astype(str).str.strip().isin(PI_CONTOUR_VARIABLES)].copy()
    previous = pd.DataFrame()
    previous_rows = 0
    if master_path.exists():
        try:
            previous = pd.read_excel(master_path, sheet_name=PI_MASTER_SHEET_NAME, dtype=object)
            previous = previous.where(pd.notna(previous), "")
            if "Variavel" in previous.columns:
                previous = previous[previous["Variavel"].astype(str).str.strip().isin(PI_CONTOUR_VARIABLES)].copy()
            previous_rows = len(previous)
        except ValueError:
            previous = pd.DataFrame()
    all_columns = list(dict.fromkeys(list(previous.columns) + list(df_pi.columns)))
    previous = previous.reindex(columns=all_columns) if not previous.empty else pd.DataFrame(columns=all_columns)
    increment = df_pi.reindex(columns=all_columns)
    increment = increment.where(pd.notna(increment), "")
    combined = pd.concat([previous, increment], ignore_index=True)
    dedup_keys = [col for col in ["PI Inicio", "PI Final", "DataHora", "Medidor", "Grupo", "Variavel", "Canal"] if col in combined.columns]
    if dedup_keys:
        for col in dedup_keys:
            combined[col] = combined[col].where(pd.notna(combined[col]), "").astype(str)
        combined.drop_duplicates(subset=dedup_keys, keep="last", inplace=True)
    if master_path.exists():
        with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name=PI_MASTER_SHEET_NAME, index=False)
            pd.DataFrame([
                {"Item": "Linhas totais", "Valor": len(combined)},
                {"Item": "Medidores", "Valor": combined["Medidor"].nunique() if "Medidor" in combined.columns else 0},
                {"Item": "Variáveis autorizadas", "Valor": combined["Variavel"].nunique() if "Variavel" in combined.columns else 0},
                {"Item": "Dias coletados", "Valor": combined["PI Dia Coleta"].nunique() if "PI Dia Coleta" in combined.columns else 0},
            ]).to_excel(writer, sheet_name="PI_EXTRACT_RESUMO", index=False)
    else:
        with pd.ExcelWriter(master_path, engine="openpyxl") as writer:
            combined.to_excel(writer, sheet_name=PI_MASTER_SHEET_NAME, index=False)
            pd.DataFrame([{"Item": "Linhas totais", "Valor": len(combined)}]).to_excel(writer, sheet_name="PI_EXTRACT_RESUMO", index=False)
    return {"previous_rows": previous_rows, "increment_rows": len(increment), "total_rows": len(combined)}


def run_pi_collection(
    pi_root: Path,
    config_name: str,
    output_path: Path,
    period_path: Path,
    start: datetime,
    end: datetime,
) -> pd.DataFrame:
    """Executa a coleta PI Vision; prefere Daily Control V5 e usa V4.9 como fallback."""
    if not pi_root.is_dir():
        raise FileNotFoundError(f"Pasta do coletor PI não encontrada: {pi_root}")
    if _daily_control_v5_available(pi_root):
        return _run_daily_control_v5_collection(pi_root, start, end)
    runner = pi_root / PI_COLLECTOR_RUNNER
    config = Path(config_name)
    if not config.is_absolute():
        config = pi_root / config
    if not runner.exists():
        raise FileNotFoundError(f"Executor PI não encontrado: {runner}")
    if not config.exists():
        raise FileNotFoundError(f"Configuração PI não encontrada: {config}")
    cmd = [
        sys.executable,
        str(runner),
        "--config",
        str(config),
        "--start",
        start.strftime("%d/%m/%Y %H:%M:%S"),
        "--end",
        end.strftime("%d/%m/%Y %H:%M:%S"),
    ]
    _ensure_pi_edge_cdp(pi_root, config)
    try:
        _pi_capture_daily_control_snapshot(pi_root, start.strftime("%Y-%m-%d"))
    except Exception as exc:
        print(f"[WARN] Captura Daily Control não interromperá a coleta PI: {exc}")
    print(f"\n[INFO] Abrindo/coletando PI Vision no período {start:%d/%m/%Y %H:%M:%S} a {end:%d/%m/%Y %H:%M:%S}...")
    try:
        completed = subprocess.run(cmd, cwd=str(pi_root), timeout=PI_COLLECTION_TIMEOUT_SECONDS)
    except subprocess.TimeoutExpired as exc:
        raise TimeoutError(f"Coleta PI Vision V4.9 excedeu {PI_COLLECTION_TIMEOUT_SECONDS}s sem concluir.") from exc
    if completed.returncode != 0:
        raise RuntimeError(f"Coleta PI retornou código {completed.returncode}.")
    period_record = _read_pi_period_record(period_path)
    df_pi = _load_pi_extract(output_path, period_record)
    print(f"[OK] PI Vision coletado/importado: {len(df_pi)} linha(s) de {output_path}")
    return df_pi


def run_pi_collection_for_days(
    pi_root: Path,
    config_name: str,
    pi_output_path: Path,
    period_path: Path,
    workbook_path: Path,
    days_iso: list,
    retries: int = PI_RETRIES,
) -> pd.DataFrame:
    """Executa PI Vision por dia quando a janela passa de 24h e acumula a aba PI_EXTRACT.

    Alguns coletores PI interpretam uma janela grande como leitura única início/fim.
    Para preservar a leitura diária, esta função faz uma captura por dia, aguarda
    a conclusão de cada processo, importa a saída normalizada e regrava a aba
    PI_EXTRACT acumulada antes de passar ao próximo dia.
    """
    ordered_days = sorted(str(day) for day in days_iso)
    if not ordered_days:
        return pd.DataFrame()
    if len(ordered_days) > 1:
        print(f"\n[INFO] Janela PI maior que 24h detectada: {len(ordered_days)} dia(s). A coleta será feita dia a dia.")
    frames = []
    failed_days = []
    for index, day in enumerate(ordered_days, start=1):
        start, end = _pi_period_for_day(day)
        print(f"\n[INFO] PI dia {index}/{len(ordered_days)} — {day}")
        df_day = None
        last_exc = None
        max_attempts = max(1, int(retries) + 1)
        for attempt in range(1, max_attempts + 1):
            retry_config_name, retry_config_path = _pi_retry_config(pi_root, config_name, attempt)
            if attempt > 1:
                print(
                    f"[RETRY] Retentativa PI {attempt}/{max_attempts} para {day}: "
                    "aumentando a espera para a tela terminar de carregar."
                )
            try:
                df_day = run_pi_collection(pi_root, retry_config_name, pi_output_path, period_path, start, end)
                break
            except Exception as exc:
                last_exc = exc
                print(f"[WARN] PI {day} tentativa {attempt}/{max_attempts} falhou: {exc}")
            finally:
                if retry_config_path is not None:
                    try:
                        retry_config_path.unlink(missing_ok=True)
                    except Exception:
                        pass
        if df_day is None:
            failed_days.append(day)
            df_day = _pi_failure_dataframe(day, start, end, max_attempts, last_exc or RuntimeError("Falha PI desconhecida"), pi_output_path)
            print(f"[ERROR] PI {day} não foi aprovado após {max_attempts} tentativa(s). A falha será registrada na aba {PI_SHEET_NAME} e a automação seguirá.")
        elif "PI Status Coleta" not in df_day.columns:
            df_day.insert(7, "PI Status Coleta", "OK")
        df_day.insert(0, "PI Ordem Coleta", index)
        frames.append(df_day)
        df_accumulated = pd.concat(frames, ignore_index=True).where(lambda data: pd.notna(data), "")
        write_pi_extract_sheet(workbook_path, df_accumulated, PI_SHEET_NAME)
        print(f"💾 Dia PI {day} registrado na aba {PI_SHEET_NAME}; acumulado: {len(df_accumulated)} linha(s).")
        if day in failed_days and index < len(ordered_days):
            if not _prompt_yes_no("Deseja continuar a coleta PI para o próximo dia?", default=True):
                print("[SKIP] Etapa PI encerrada pelo usuário. A automação seguirá para a próxima etapa com o que já foi registrado.")
                break
    if failed_days:
        print(f"[WARN] Coleta PI concluída com falha(s) registrada(s) para: {', '.join(failed_days)}")
    return pd.concat(frames, ignore_index=True).where(lambda data: pd.notna(data), "")


def run_email_download_automation(mpfm_root: Path, sep_root: Path, continue_on_warning: bool = False) -> dict:
    """Executa as automações de e-mail/ZIP antes do PI/Base_Unica."""
    automation_dir = Path(__file__).resolve().parent / EMAIL_AUTOMATION_DIR
    download_script = automation_dir / EMAIL_DOWNLOAD_SCRIPT
    organize_script = automation_dir / EMAIL_ORGANIZE_SCRIPT
    if not automation_dir.is_dir():
        raise FileNotFoundError(f"Pasta de automações não encontrada: {automation_dir}")
    if not download_script.exists():
        raise FileNotFoundError(f"Script de download de e-mail não encontrado: {download_script}")
    if not organize_script.exists():
        raise FileNotFoundError(f"Script de organização de ZIPs não encontrado: {organize_script}")

    env = os.environ.copy()
    env["DPB_WORKSPACE_ROOT"] = str(mpfm_root)
    env["DPB_ZIP_FOLDER"] = str(mpfm_root / "Zip")
    sep_destination = sep_root.parent if re.fullmatch(r"20\d{2}", sep_root.name) else sep_root
    env["DPB_DAILY_REPORTS_DESTINATION"] = str(sep_destination)
    # O downloader também pode chamar a automação legada, mas isso duplicaria
    # o processamento. O fluxo standalone organiza os arquivos logo depois.
    env["DPB_EMAIL_RUN_MAIN"] = "0"
    env.setdefault("PYTHONIOENCODING", "utf-8")

    results = {}
    timeout_seconds = max(60, int(os.environ.get("BASE_UNICA_EMAIL_TIMEOUT_SECONDS", "900")))
    print("\n📥 Executando automação de e-mail: baixar ZIPs de Daily Reports/Configuration/FCVs/Eventos...")
    download = subprocess.run([sys.executable, str(download_script)], cwd=str(automation_dir), env=env, timeout=timeout_seconds)
    results["download_returncode"] = download.returncode
    if download.returncode not in (0, 2, 3):
        raise RuntimeError(f"Automação de download de e-mail retornou código {download.returncode}.")
    if download.returncode in (2, 3):
        print(f"[WARN] Download de e-mail finalizou com atenção/código {download.returncode}; seguindo para organização dos ZIPs.")

    print("\n[INFO] Organizando ZIPs/PDFs baixados nas pastas MPFM/Alarmes/Eventos...")
    organize = subprocess.run([sys.executable, str(organize_script)], cwd=str(automation_dir), env=env, timeout=timeout_seconds)
    results["organize_returncode"] = organize.returncode
    if organize.returncode not in (0, 2, 3):
        raise RuntimeError(f"Automação de organização de ZIPs retornou código {organize.returncode}.")
    if organize.returncode in (2, 3):
        print(f"[WARN] Organização de ZIPs finalizou com atenção/código {organize.returncode}. Pode haver itens faltantes a verificar.")
        if not continue_on_warning:
            print("   A execução continuará porque códigos 2/3 indicam ausência de ZIPs ou pendência de completude, não falha crítica.")
    print("[OK] Automação de e-mail/organização concluída.")
    return results


# ═════════════════════════════════════════════════════════════════════════
# DESCOBERTA — dias disponíveis (Daily PDF) e leitura de Hourly PDF
# ═════════════════════════════════════════════════════════════════════════

def _candidate_daily_paths(
    mpfm_root: Path, bank_code: str, months: list, keep: int | None = None,
    date_from: str | None = None, date_to: str | None = None,
) -> list:
    paths = []
    last_report_name_day = ""
    if date_to:
        last_report_name_day = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    for year, month in months:
        daily_dir = bank_month_dir(mpfm_root, bank_code, year, month, "Daily")
        if daily_dir.is_dir():
            for path in daily_dir.glob(f"{bank_code}_MPFM_Daily-*.pdf"):
                file_day = _filename_date_iso(path)
                if date_from and file_day and file_day < date_from:
                    continue
                if last_report_name_day and file_day and file_day > last_report_name_day:
                    continue
                paths.append(path)
    paths.sort(key=lambda p: _filename_date_iso(p) or "", reverse=True)
    return paths[:keep] if keep else paths


def discover_daily_records(
    mpfm_root: Path, bank_code: str, months: list, keep: int | None = None,
    date_from: str | None = None, date_to: str | None = None, workers: int = 1,
) -> dict:
    out = {}
    paths = _candidate_daily_paths(mpfm_root, bank_code, months, keep, date_from, date_to)
    for pdf_name, record, error in parse_pdf_batch(paths, "daily", workers):
        pdf_path = Path(pdf_name)
        if error:
            print(f"  [WARN] falha ao ler {pdf_path.name}: {error}")
            continue
        day = record.get("date_from")
        if not day:
            continue
        if date_from and day < date_from:
            continue
        if date_to and day > date_to:
            continue
        out.setdefault(day, (pdf_path, record))
    return out


def discover_hourly_records_for_days(mpfm_root: Path, bank_code: str, months: list, target_days: set, workers: int = 1) -> dict:
    relevant_names = set(target_days)
    for d in target_days:
        dt = datetime.strptime(d, "%Y-%m-%d") + timedelta(days=1)
        relevant_names.add(dt.strftime("%Y-%m-%d"))

    by_day = defaultdict(list)
    paths = []
    for year, month in months:
        hourly_dir = bank_month_dir(mpfm_root, bank_code, year, month, "Hourly")
        if not hourly_dir.is_dir():
            continue
        for pdf_path in sorted(hourly_dir.rglob(f"{bank_code}_MPFM_Hourly-*.pdf")):
            if _filename_date_iso(pdf_path) not in relevant_names:
                continue
            paths.append(pdf_path)
    for pdf_name, record, error in parse_pdf_batch(paths, "hourly", workers):
        pdf_path = Path(pdf_name)
        if error:
            print(f"  [WARN] falha ao ler {pdf_path.name}: {error}")
            continue
        day = record.get("date_from")
        hour = record.get("hour")
        if not day or hour is None or day not in target_days:
            continue
        by_day[day].append(record)
    return by_day


# ═════════════════════════════════════════════════════════════════════════
# SEP: localizar os 3 TXT (óleo/gás/água) de um dia lendo o "Meter ID"
# ═════════════════════════════════════════════════════════════════════════

_METER_ID_RE = re.compile(r"Meter\s+ID\s+(\S+)", re.IGNORECASE)
_EXCLUDED_TXT_PREFIXES = (
    "alarmsandevents", "configuration", "run_daily", "parameters", "run_hourly", "_monthly", "run1_monthly",
    "run2_monthly", "run3_monthly", "security",
)


def _read_meter_id(txt_path: Path) -> str:
    try:
        with open(txt_path, encoding="utf-8", errors="replace") as f:
            head = f.read(400)
        m = _METER_ID_RE.search(head)
        return m.group(1).strip().upper() if m else ""
    except Exception:
        return ""


def find_sep_files_for_day(sep_root: Path, day_iso: str) -> dict:
    """Varre FC13/FC14/FC17 do dia (caminho direto, sem busca recursiva no
    ano inteiro) e identifica óleo/gás/água pelo Meter ID lido no cabeçalho
    de cada Run_24Hours*.txt."""
    found = {}
    for fc_folder in SEP_FC_FOLDERS:
        day_dir = sep_day_dir(sep_root, day_iso, fc_folder)
        if not day_dir.is_dir():
            continue
        for txt_path in day_dir.glob("Run_24Hours*.txt"):
            name_lower = txt_path.name.lower()
            if name_lower.startswith(_EXCLUDED_TXT_PREFIXES):
                continue
            meter_id = _read_meter_id(txt_path)
            phase = TARGET_PHASE_BY_METER.get(meter_id)
            if phase and phase not in found:
                found[phase] = txt_path
    return found


def load_sep_data_for_days(sep_root: Path, days_iso: list) -> dict:
    result = {}
    for day in days_iso:
        paths = find_sep_files_for_day(sep_root, day)
        missing = [k for k in ("oleo", "gas", "agua") if k not in paths]
        if missing:
            if paths:
                print(f"  [WARN] SEP incompleto em {day}: faltando {', '.join(missing)} — dia ignorado no merge SEP")
            continue
        try:
            result[day] = parse_sep_txt_set(paths["oleo"], paths["gas"], paths["agua"])
        except Exception as exc:
            print(f"  [WARN] falha ao ler TXT do SEP em {day}: {exc}")
    return result


def _pdf_text(path: Path, max_pages: int | None = None) -> str:
    reader = PdfReader(str(path))
    pages = reader.pages if max_pages is None else reader.pages[:max_pages]
    return "\n".join(page.extract_text() or "" for page in pages)


def _mpfm_support_root(mpfm_root: Path) -> Path:
    """Retorna a pasta que contém bancos MPFM, alarmes e eventos."""
    if (mpfm_root / ALARM_FOLDER_NAME).exists() or (mpfm_root / EVENT_FOLDER_NAME).exists():
        return mpfm_root
    return mpfm_root.parent


def alarm_root_from_mpfm(mpfm_root: Path) -> Path:
    return _mpfm_support_root(mpfm_root) / ALARM_FOLDER_NAME


def event_root_from_mpfm(mpfm_root: Path) -> Path:
    return _mpfm_support_root(mpfm_root) / EVENT_FOLDER_NAME


def _bank_from_text(value: str) -> str:
    m = re.search(r"Bank\s*0?(\d{1,2})", value or "", re.IGNORECASE)
    if not m:
        return ""
    return f"B{int(m.group(1)):02d}"


def _instrument_from_text(value: str) -> str:
    m = re.search(r"\b\d{2}FT\d{3,4}[A-Z]?\b", value or "", re.IGNORECASE)
    return m.group(0).upper() if m else ""


def _issue_flag(priority: str, description: str, state: str = "") -> str:
    text = f"{priority} {description} {state}".lower()
    flags = []
    if "critical" in text:
        flags.append("CRITICAL")
    for word, label in (
        ("communication", "COMM"), ("failed", "FAILED"), ("low flow", "LOW_FLOW"),
        ("low dp", "LOW_DP"), ("deviation", "DEVIATION"), ("pvt", "PVT"),
        ("parameter", "PARAMETER"), ("density", "DENSITY"), ("meter factor", "METER_FACTOR"),
    ):
        if word in text and label not in flags:
            flags.append(label)
    return ",".join(flags)


def _record_date_from_timestamp(timestamp: str, fallback_day: str) -> str:
    if not timestamp:
        return fallback_day
    for pattern in ("%Y-%m-%d %H:%M:%S", "%d/%m/%y %H:%M:%S", "%m/%d/%y %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(timestamp, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return fallback_day


def _alarm_file_for_day(alarm_root: Path, day_iso: str) -> Path:
    dt = datetime.strptime(day_iso, "%Y-%m-%d")
    return _month_dir(_year_dir(alarm_root, dt.year), dt.month) / f"{dt:%d-%m}_Alarmes FCS320.pdf"


def parse_alarm_pdf(path: Path, fallback_day: str) -> list[dict]:
    if not path.exists():
        return []
    try:
        text = re.sub(r"\s+", " ", _pdf_text(path))
    except Exception as exc:
        return [{
            "ProductionDate": fallback_day, "Timestamp": "", "RecordType": "ALARM", "SourceKind": "PDF",
            "Bank": "", "Priority": "", "Object": "", "Description": f"Falha ao ler PDF: {exc}",
            "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": "",
            "SourceFile": path.name, "IssueFlag": "READ_ERROR",
        }]
    pattern = re.compile(
        r"\b(Alarm|Advisory|Critical|General)\s+"
        r"(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+"
        r"([A-Za-z0-9_.-]+)\s+(?:[A-Za-z0-9_.-]+\s+)?"
        r"(.+?)\s+(On|Off)\s+([A-Za-z]+)\s+(-?\d+)\s+(-?\d+)",
        re.IGNORECASE,
    )
    rows = []
    for match in pattern.finditer(text):
        priority, timestamp, obj, description, state, detailed, source_id, signal = match.groups()
        rows.append({
            "ProductionDate": fallback_day,
            "Timestamp": timestamp,
            "RecordType": "ALARM",
            "SourceKind": "PDF",
            "Bank": _bank_from_text(obj),
            "Priority": priority.title(),
            "Object": obj,
            "Description": description.strip(),
            "State": state,
            "DetailedState": detailed,
            "SourceID": source_id,
            "SignalNumber": signal,
            "Instrumento": _instrument_from_text(obj),
            "SourceFile": path.name,
            "IssueFlag": _issue_flag(priority, description, detailed),
        })
    if not rows:
        rows.append({
            "ProductionDate": fallback_day, "Timestamp": "", "RecordType": "ALARM", "SourceKind": "PDF",
            "Bank": "", "Priority": "", "Object": "", "Description": "Nenhum alarme extraído do PDF",
            "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": "",
            "SourceFile": path.name, "IssueFlag": "EMPTY_OR_UNPARSED",
        })
    return rows


def _event_file_day(path: Path):
    name = path.name
    m = re.search(r"(\d{8})", name)
    if m:
        raw = m.group(1)
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:8]}"
    m = re.search(r"(\d{2})[.](\d{2})[.](\d{4})", name)
    if m:
        month, day, year = m.groups()
        return f"{year}-{month}-{day}"
    return None


def _event_paths_for_days(event_root: Path, days_iso: list) -> list[Path]:
    if not event_root.exists():
        return []
    wanted = set(days_iso)
    paths = []
    years = sorted({datetime.strptime(day, "%Y-%m-%d").year for day in days_iso})
    for year in years:
        year_dir = _year_dir(event_root, year)
        candidates = []
        if year_dir.exists():
            candidates.extend(year_dir.rglob("*.txt"))
            candidates.extend(year_dir.rglob("*.pdf"))
        candidates.extend(event_root.glob("*.pdf"))
        for path in candidates:
            day = _event_file_day(path)
            if day in wanted and path not in paths:
                paths.append(path)
    return paths


def parse_event_txt(path: Path, selected_days: set) -> list[dict]:
    rows = []
    text = path.read_text(encoding="utf-8", errors="replace")
    for line in text.splitlines():
        m = re.match(r"\s*(\d{2}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2})\s+(.+?)\s*$", line)
        if not m:
            continue
        timestamp, description = m.groups()
        prod_date = _record_date_from_timestamp(timestamp, "")
        if prod_date not in selected_days:
            continue
        rows.append({
            "ProductionDate": prod_date, "Timestamp": timestamp, "RecordType": "EVENT", "SourceKind": "TXT",
            "Bank": "", "Priority": "", "Object": "", "Description": description.strip(), "State": "",
            "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": _instrument_from_text(description),
            "SourceFile": path.name, "IssueFlag": _issue_flag("", description),
        })
    return rows


def parse_event_pdf(path: Path, selected_days: set) -> list[dict]:
    rows = []
    try:
        text = re.sub(r"\s+", " ", _pdf_text(path, max_pages=6))
    except Exception as exc:
        day = _event_file_day(path) or ""
        return [{
            "ProductionDate": day, "Timestamp": "", "RecordType": "EVENT", "SourceKind": "PDF",
            "Bank": "", "Priority": "", "Object": "", "Description": f"Falha ao ler PDF: {exc}",
            "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": _instrument_from_text(path.name),
            "SourceFile": path.name, "IssueFlag": "READ_ERROR",
        }]
    for match in re.finditer(r"(\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2})", text):
        timestamp = match.group(1)
        prod_date = _record_date_from_timestamp(timestamp, "")
        if prod_date not in selected_days:
            continue
        start = max(0, match.start() - 160)
        end = min(len(text), match.end() + 240)
        snippet = text[start:end].strip()
        rows.append({
            "ProductionDate": prod_date, "Timestamp": timestamp, "RecordType": "EVENT", "SourceKind": "PDF",
            "Bank": _bank_from_text(path.name), "Priority": "", "Object": "", "Description": snippet,
            "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": _instrument_from_text(path.name),
            "SourceFile": path.name, "IssueFlag": _issue_flag("", snippet),
        })
    return rows


def load_alarm_event_data(mpfm_root: Path, days_iso: list) -> pd.DataFrame:
    alarm_root = alarm_root_from_mpfm(mpfm_root)
    event_root = event_root_from_mpfm(mpfm_root)
    selected_days = set(days_iso)
    rows = []
    for day in days_iso:
        alarm_path = _alarm_file_for_day(alarm_root, day)
        if alarm_path.exists():
            rows.extend(parse_alarm_pdf(alarm_path, day))
        else:
            rows.append({
                "ProductionDate": day, "Timestamp": "", "RecordType": "ALARM", "SourceKind": "PDF",
                "Bank": "", "Priority": "", "Object": "", "Description": "Arquivo de alarmes não encontrado",
                "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": "",
                "SourceFile": str(alarm_path.name), "IssueFlag": "MISSING_FILE",
            })
    for path in _event_paths_for_days(event_root, days_iso):
        try:
            if path.suffix.lower() == ".txt":
                rows.extend(parse_event_txt(path, selected_days))
            elif path.suffix.lower() == ".pdf":
                rows.extend(parse_event_pdf(path, selected_days))
        except Exception as exc:
            rows.append({
                "ProductionDate": _event_file_day(path) or "", "Timestamp": "", "RecordType": "EVENT", "SourceKind": path.suffix.upper().strip("."),
                "Bank": "", "Priority": "", "Object": "", "Description": f"Falha ao ler evento: {exc}",
                "State": "", "DetailedState": "", "SourceID": "", "SignalNumber": "", "Instrumento": _instrument_from_text(path.name),
                "SourceFile": path.name, "IssueFlag": "READ_ERROR",
            })
    df = pd.DataFrame(rows, columns=ALARM_EVENT_COLUMNS)
    df = _normalize_alarm_event_columns(df)
    return df.where(pd.notna(df), "")


# ═════════════════════════════════════════════════════════════════════════
# PRÉ-VALIDAÇÃO E DASHBOARD
# ═════════════════════════════════════════════════════════════════════════

def _prompt_yes_no(question: str, default: bool = False) -> bool:
    suffix = "[S/n]" if default else "[s/N]"
    try:
        answer = input(f"{question} {suffix}: ").strip().lower()
    except EOFError:
        return default
    if not answer:
        return default
    return answer in ("s", "sim", "y", "yes")


def _clean_path_input(value: str) -> str:
    """Limpa aspas que costumam acompanhar caminhos copiados do Explorer."""
    return value.strip().strip('"').strip()


def _prompt_configured_path(label: str, default: str, required: bool = True) -> str:
    """Exibe um caminho padrão e permite confirmar ou substituí-lo.

    O Enter confirma o caminho sugerido. Se o caminho não existir, o usuário
    pode corrigi-lo imediatamente ou continuar para receber a validação final
    do processo.
    """
    suggested = _clean_path_input(default)
    while True:
        print(f"\n📁 {label}")
        if suggested:
            print(f"   Padrão: {suggested}")
            try:
                entered = input("   Pressione Enter para confirmar ou cole outro caminho: ")
            except EOFError:
                entered = ""
            value = _clean_path_input(entered) or suggested
        else:
            try:
                value = _clean_path_input(input("   Informe o caminho: "))
            except EOFError:
                value = ""

        if not value:
            if required:
                print("[WARN] Um caminho é necessário para esta etapa.")
                continue
            return value

        path = Path(value)
        if path.is_dir():
            print(f"[OK] Caminho confirmado: {path}")
            return value

        print(f"[WARN] Caminho ainda não encontrado: {path}")
        if _prompt_yes_no("Deseja editar este caminho agora?", default=True):
            suggested = value
            continue
        print("   O processo continuará e fará a validação final antes da leitura.")
        return value


def _prompt_operation_mode(default: str = "1") -> str:
    print("\n🚦 Escolha o caminho da automação")
    print("  1 - Automação completa: baixar PDF/TXT + PI + Base_Unica")
    print("  2 - Parcial: PI + Base_Unica")
    print("  3 - Parcial: baixar PDF/TXT + Base_Unica")
    print("  4 - Parcial: baixar PDF/TXT + PI")
    print("  5 - Apenas baixar PDF/TXT")
    print("  6 - Apenas PI")
    print("  7 - Apenas Base_Unica")
    while True:
        try:
            answer = input(f"Opção [{default}]: ").strip() or default
        except EOFError:
            answer = default
        if answer in OPERATION_MODES:
            return answer
        print("Opção inválida. Digite um número de 1 a 7.")


def _operation_flags(mode: str) -> dict:
    return OPERATION_MODES.get(str(mode), OPERATION_MODES["1"])


def count_hourly_pdf_files_for_day(mpfm_root: Path, bank_code: str, months: list, day_iso: str) -> int:
    relevant_names = {day_iso, (datetime.strptime(day_iso, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")}
    paths = set()
    for year, month in months:
        hourly_dir = bank_month_dir(mpfm_root, bank_code, year, month, "Hourly")
        if not hourly_dir.is_dir():
            continue
        for pdf_path in hourly_dir.rglob(f"{bank_code}_MPFM_Hourly-*.pdf"):
            if _filename_date_iso(pdf_path) in relevant_names:
                paths.add(pdf_path)
    return len(paths)


def build_preflight_report(mpfm_root: Path, sep_root: Path, months: list, selected_days: list, daily_by_bank: dict, aligned_bank: str) -> tuple[list, list]:
    rows = []
    issues = []
    sep_by_day = {day: find_sep_files_for_day(sep_root, day) for day in selected_days}
    alarm_root = alarm_root_from_mpfm(mpfm_root)
    event_root = event_root_from_mpfm(mpfm_root)
    event_paths_by_day = defaultdict(list)
    for path in _event_paths_for_days(event_root, selected_days):
        day = _event_file_day(path)
        if day:
            event_paths_by_day[day].append(path)
    for day in selected_days:
        sep_paths = sep_by_day.get(day, {})
        missing_sep = [phase for phase in ("oleo", "gas", "agua") if phase not in sep_paths]
        alarm_ok = _alarm_file_for_day(alarm_root, day).exists()
        event_count = len(event_paths_by_day.get(day, []))
        for bank_code in BANK_FOLDERS:
            daily_ok = day in daily_by_bank.get(bank_code, {})
            hourly_count = count_hourly_pdf_files_for_day(mpfm_root, bank_code, months, day)
            hourly_ok = hourly_count >= 24
            # O SEP é uma fonte independente. A comparação com qualquer MPFM
            # é escolhida no dashboard; não deve ficar artificialmente ligada
            # ao banco configurado como alinhamento padrão.
            sep_ok = not missing_sep
            row_issues = []
            if not daily_ok:
                row_issues.append("Daily PDF ausente")
            if not hourly_ok:
                row_issues.append(f"Hourly incompleto ({hourly_count}/24)")
            if not sep_ok:
                row_issues.append(f"SEP faltando {', '.join(missing_sep)}")
            if not alarm_ok:
                row_issues.append("arquivo de alarmes ausente")
            status = "OK" if not row_issues else "VERIFICAR"
            row = {
                "Dia": day,
                "Banco": bank_code,
                "Daily PDF": "OK" if daily_ok else "FALTA",
                "Hourly PDFs": f"{hourly_count}/24" if hourly_count < 24 else f"{hourly_count}",
                "SEP": "OK" if sep_ok else f"FALTA {', '.join(missing_sep)}",
                "Alarmes": "OK" if alarm_ok else "FALTA",
                "Eventos": str(event_count),
                "Status": status,
                "Observação": "; ".join(row_issues),
            }
            rows.append(row)
            if row_issues:
                issues.append(row)
    return rows, issues


def print_preflight_report(rows: list, issues: list):
    print("\n🔎 Pré-validação da janela selecionada")
    if not rows:
        print("  Nenhum item para validar.")
        return
    df = pd.DataFrame(rows)
    status_counts = df["Status"].value_counts().to_dict()
    print(f"  Bancos/dias avaliados: {len(df)} | OK: {status_counts.get('OK', 0)} | Verificar: {status_counts.get('VERIFICAR', 0)}")
    if issues:
        print("  Itens com pendência:")
        for row in issues[:30]:
            print(f"   - {row['Dia']} {row['Banco']}: {row['Observação']}")
        if len(issues) > 30:
            print(f"   ... mais {len(issues) - 30} item(ns)")


def _num(value):
    try:
        if value in ("", None):
            return np.nan
        return float(value)
    except Exception:
        return np.nan


def _row_has_production(row, threshold: float = 0.001) -> bool:
    """Indica se a linha representa operação com massa mensurável.

    Linhas zeradas são preservadas como evidência de disponibilidade/fechamento,
    mas não entram em desvios, comparações ou indicadores de estabilidade.
    """
    values = [_num(row.get("MPFM corr Total (t)", "")), _num(row.get("MPFM corr HC (t)", ""))]
    return any(not np.isnan(value) and abs(value) > threshold for value in values)


def _fmt(value, digits=2):
    value = _num(value)
    if np.isnan(value):
        return ""
    return f"{value:,.{digits}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _format_br_datetime(value) -> str:
    """Formata datas visíveis no padrão brasileiro sem alterar chaves internas."""
    if value in (None, ""):
        return ""
    if isinstance(value, pd.Timestamp):
        parsed = value
    elif isinstance(value, datetime):
        parsed = pd.Timestamp(value)
    else:
        text = str(value).strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}(?:[ T]\d{1,2}:?\d{0,2}(?::\d{2})?)?", text):
            return text
        parsed = pd.to_datetime(text, errors="coerce")
    if pd.isna(parsed):
        return str(value)
    text = str(value)
    has_time = bool(re.search(r"(?:[ T]\d{1,2}:\d{2})", text)) or (
        isinstance(value, (datetime, pd.Timestamp))
        and (parsed.hour or parsed.minute or parsed.second)
    )
    return parsed.strftime("%d/%m/%Y %H:%M" if has_time else "%d/%m/%Y")


def _format_html_value(value, column: str):
    if value in (None, ""):
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    column_text = str(column).lower()
    if isinstance(value, (datetime, pd.Timestamp)) or re.match(r"^\d{4}-\d{2}-\d{2}(?:[ T]|$)", str(value).strip()):
        return _format_br_datetime(value)
    if isinstance(value, (int, float, np.integer, np.floating)) and not isinstance(value, bool):
        digits = 0 if any(token in column_text for token in ("count", "qtd", "quant", "registros", "linhas", "hora", "ano", "dias consecutivos")) else 2
        if "%" in str(column) or "percent" in column_text or "desvio" in column_text:
            digits = 1
        return _fmt(value, digits)
    return value


def _html_table(rows: list, columns: list | None = None, max_rows: int = 200) -> str:
    if not rows:
        return "<p class='muted'>Sem registros.</p>"
    df = pd.DataFrame(rows)
    if columns:
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns]
    total_rows = len(df)
    truncation_note = ""
    if total_rows > max_rows:
        truncation_note = (
            f"<p class='muted table-note'>Mostrando {max_rows} de {total_rows} registros. "
            "Consulte a Base_Unica/Excel para a evidência completa.</p>"
        )
        df = df.head(max_rows)
    # O Excel mantém tipos e valores originais. Somente a camada visual do
    # HTML recebe data/hora e separadores numéricos brasileiros.
    for col in df.columns:
        df[col] = df[col].map(lambda value, column=col: _format_html_value(value, column))
    table_html = df.to_html(index=False, escape=True, classes="data-table")
    table_html = table_html.replace("<th>", '<th scope="col">')
    return truncation_note + table_html


def _daily_mpfm(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    out = _normalize_master_columns(df)
    for col in ("ProductionDate", "Granularity", "Origin", "SourceType", "Bank", "Tag", "Instrumento"):
        out[col] = out[col].astype(str)
    enabled = out.apply(lambda row: _mpfm_extraction_enabled(row.get("Tag"), row.get("Instrumento")), axis=1)
    return out[(out["Granularity"] == "Daily") & (out["Origin"] == "MPFM") & (out["SourceType"] == "PDF") & enabled].copy()


def _max_consecutive_hourly_delta(group: pd.DataFrame, column: str) -> float:
    """Retorna o maior salto absoluto entre horas consecutivas da série."""
    if column not in group.columns or "Hour" not in group.columns:
        return np.nan
    work = pd.DataFrame({
        "hour": pd.to_numeric(group["Hour"], errors="coerce"),
        "value": pd.to_numeric(group[column], errors="coerce"),
    }).dropna()
    if work.empty:
        return np.nan
    work = work.groupby("hour", as_index=False)["value"].mean().sort_values("hour")
    if len(work) < 2:
        return np.nan
    hour_gap = work["hour"].diff()
    delta = work["value"].diff().abs()
    valid = delta[hour_gap == 1]
    return float(valid.max()) if not valid.empty else np.nan


def dashboard_analysis(df_out: pd.DataFrame, context_df: pd.DataFrame, target_days: list, preflight_rows: list) -> dict:
    daily = _daily_mpfm(context_df)
    if not daily.empty:
        # Pontos não liberados para publicação não geram anomalias, spikes nem
        # linhas de auditoria no HTML.
        daily = daily[daily.apply(lambda row: _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    current_daily = daily[daily["ProductionDate"].isin(target_days)].copy()
    hourly = _normalize_master_columns(context_df)
    hourly = hourly[
        hourly["ProductionDate"].astype(str).isin({str(day) for day in target_days})
        & (hourly["Granularity"].astype(str) == "Hourly")
        & (hourly["Origin"].astype(str) == "MPFM")
        & (hourly["SourceType"].astype(str) == "PDF")
    ].copy()

    key_metrics = [
        "MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr HC (t)",
        "MPFM corr Água (t)", "MPFM corr Total (t)", "Pressão (barg)", "Temperatura (°C)",
    ]
    anomalies = []
    comparisons = []
    pressure_flags = []
    for _, row in current_daily.iterrows():
        day = str(row["ProductionDate"])
        bank = str(row["Bank"])
        tag = str(row["Tag"])
        if not _row_has_production(row):
            anomalies.append({
                "Dia": day, "Banco": bank, "Tag": tag,
                "Tipo": "EXCLUÍDO OPERACIONAL",
                "Campo": "MPFM corr Total (t)",
                "Valor": _fmt(row.get("MPFM corr Total (t)", "")),
                "Observação": "Linha zerada/sem produção; não participa dos desvios, spikes ou acumulados.",
            })
            continue
        day_dt = datetime.strptime(day, "%Y-%m-%d")
        prev_days = [(day_dt - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(1, 5)]
        prev = daily[(daily["Bank"] == bank) & (daily["Tag"] == tag) & (daily["ProductionDate"].isin(prev_days))]

        for metric in key_metrics:
            value = _num(row.get(metric, ""))
            if np.isnan(value):
                anomalies.append({"Dia": day, "Banco": bank, "Tag": tag, "Tipo": "ausente", "Campo": metric, "Valor": "", "Observação": "Valor não encontrado"})
                continue
            if value == 0 and metric not in ("MPFM corr Água (t)",):
                anomalies.append({"Dia": day, "Banco": bank, "Tag": tag, "Tipo": "zerado", "Campo": metric, "Valor": _fmt(value), "Observação": "Valor zerado; verificar condição operacional ou falha de leitura"})
            prev_values = pd.to_numeric(prev.get(metric, pd.Series(dtype=float)), errors="coerce").dropna()
            if len(prev_values) >= 2:
                mean_prev = float(prev_values.mean())
                std_prev = float(prev_values.std(ddof=0))
                delta_pct = np.nan if mean_prev == 0 else (value - mean_prev) / abs(mean_prev) * 100
                status = "OK"
                if not np.isnan(delta_pct) and abs(delta_pct) >= 30:
                    status = "SPIKE"
                    anomalies.append({"Dia": day, "Banco": bank, "Tag": tag, "Tipo": "spike", "Campo": metric, "Valor": _fmt(value), "Observação": f"Desvio de {_fmt(delta_pct, 1)}% contra média dos 4 dias anteriores"})
                elif std_prev > 0 and abs(value - mean_prev) > 3 * std_prev:
                    status = "SPIKE"
                    anomalies.append({"Dia": day, "Banco": bank, "Tag": tag, "Tipo": "spike", "Campo": metric, "Valor": _fmt(value), "Observação": "Fora de 3 desvios-padrão dos 4 dias anteriores"})
                comparisons.append({
                    "Dia": day, "Banco": bank, "Tag": tag, "Campo": metric,
                    "Valor dia": _fmt(value), "Média 4 dias anteriores": _fmt(mean_prev),
                    "Δ %": "" if np.isnan(delta_pct) else _fmt(delta_pct, 1), "Status": status,
                })

        pressure = _num(row.get("Pressão (barg)", ""))
        gas = _num(row.get("MPFM corr Gás (t)", ""))
        if not np.isnan(pressure) and pressure > 490:
            pressure_flags.append({
                "Dia": day, "Banco": bank, "Tag": tag, "Pressão (barg)": _fmt(pressure, 1),
                "Gás medido (t)": _fmt(gas),
                "Status": "VERIFICAR GÁS MEDIDO" if not np.isnan(gas) and gas > 0 else "OK - sem gás medido",
                "Observação": "Pressão acima de 490 barg; operação indicada acima do ponto de bolha.",
            })

    stability = []
    if not hourly.empty:
        for col in ("ProductionDate", "Bank", "Tag", "Hour"):
            hourly[col] = hourly[col].astype(str)
        for (day, bank, tag), group in hourly.groupby(["ProductionDate", "Bank", "Tag"], dropna=False):
            hc = pd.to_numeric(group["MPFM corr HC (t)"], errors="coerce").dropna()
            total = pd.to_numeric(group["MPFM corr Total (t)"], errors="coerce").dropna()
            pressure = pd.to_numeric(group["Pressão (barg)"], errors="coerce").dropna()

            if total.empty or not bool((total.abs() > 0.001).any()):
                stability.append({
                    "Dia": day, "Banco": bank, "Tag": tag, "Horas": len(group),
                    "CV HC (%)": "", "CV Total (%)": "", "Desv.Pressão (barg)": "",
                    "Máx |ΔP| (bar/h)": "", "Máx |ΔT| (°C/h)": "",
                    "Critério operacional": "Não aplicável — linha sem produção",
                    "Status": "EXCLUÍDO OPERACIONAL",
                    "Observação": "Linha horária zerada/sem produção; não participa da avaliação de estabilidade.",
                })
                continue

            def cv(series):
                if len(series) < 2 or float(series.mean()) == 0:
                    return np.nan
                return float(series.std(ddof=0) / abs(series.mean()) * 100)

            hc_cv = cv(hc)
            total_cv = cv(total)
            pressure_std = float(pressure.std(ddof=0)) if len(pressure) >= 2 else np.nan
            max_dp = _max_consecutive_hourly_delta(group, "Pressão (barg)")
            max_dt = _max_consecutive_hourly_delta(group, "Temperatura (°C)")
            status = "OK"
            notes = []
            if len(group) < 24:
                status = "VERIFICAR"
                notes.append(f"cobertura {len(group)}/24h")
            if not np.isnan(max_dp) and max_dp > 0.5:
                status = "VERIFICAR"
                notes.append("ΔP > 0,5 bar/h")
            if not np.isnan(max_dt) and max_dt > 1.0:
                status = "VERIFICAR"
                notes.append("ΔT > 1 °C/h")
            if not np.isnan(hc_cv) and hc_cv > 25:
                status = "VERIFICAR"
                notes.append("HC instável")
            if not np.isnan(total_cv) and total_cv > 25:
                status = "VERIFICAR"
                notes.append("Total instável")
            if not np.isnan(pressure_std) and pressure_std > 20:
                status = "VERIFICAR"
                notes.append("pressão oscilante")
            stability.append({
                "Dia": day, "Banco": bank, "Tag": tag, "Horas": len(group),
                "CV HC (%)": _fmt(hc_cv, 1), "CV Total (%)": _fmt(total_cv, 1),
                "Desv.Pressão (barg)": _fmt(pressure_std, 1),
                "Máx |ΔP| (bar/h)": _fmt(max_dp, 3),
                "Máx |ΔT| (°C/h)": _fmt(max_dt, 3),
                "Critério operacional": "ΔP ≤ 0,5 bar/h; ΔT ≤ 1 °C/h",
                "Status": status,
                "Observação": "; ".join(notes),
            })

    summary = []
    summary_source = context_df[context_df["ProductionDate"].astype(str).isin({str(day) for day in target_days})].copy()
    if not summary_source.empty:
        temp = summary_source
        for col in ("ProductionDate", "Bank", "Granularity"):
            temp[col] = temp[col].astype(str)
        grouped = temp.groupby(["ProductionDate", "Bank", "Granularity"], dropna=False).size().reset_index(name="Linhas")
        summary = grouped.rename(columns={"ProductionDate": "Dia", "Granularity": "Granularidade"}).to_dict("records")

    chart_items = []
    composition_rows = []
    hc_total_rows = []
    deviation_points = []
    control_hc = []
    control_total = []
    if not current_daily.empty:
        current_daily["_hc"] = pd.to_numeric(current_daily["MPFM corr HC (t)"], errors="coerce")
        chart = current_daily.groupby(["ProductionDate", "Bank"], dropna=False)["_hc"].sum().reset_index()
        chart_items = [(f"{r['ProductionDate']} {r['Bank']}", r["_hc"]) for _, r in chart.iterrows()]
        for _, row in current_daily.iterrows():
            label = f"{row['ProductionDate']} {row['Bank']}"
            composition_rows.append({
                "label": label,
                "Óleo": _num(row.get("MPFM corr Óleo (t)", "")),
                "Gás": _num(row.get("MPFM corr Gás (t)", "")),
                "Água": _num(row.get("MPFM corr Água (t)", "")),
            })
            hc_total_rows.append({
                "label": label,
                "HC": _num(row.get("MPFM corr HC (t)", "")),
                "Total": _num(row.get("MPFM corr Total (t)", "")),
            })

    sep_rows = _normalize_master_columns(context_df)
    sep_rows = sep_rows[sep_rows["ProductionDate"].astype(str).isin(target_days)].copy()
    if not sep_rows.empty:
        for _, row in sep_rows.iterrows():
            dhc = _num(row.get("Desvio HC (%)", ""))
            dtot = _num(row.get("Desvio Total (%)", ""))
            label = f"{row.get('ProductionDate', '')} {row.get('Bank', '')} H{row.get('Hour', '')}"
            if not np.isnan(dhc) and not np.isnan(dtot):
                deviation_points.append((dhc, dtot, label))
            if not np.isnan(dhc):
                control_hc.append((label, dhc))
            if not np.isnan(dtot):
                control_total.append((label, dtot))

    stability_chart = []
    for row in stability:
        stability_chart.append({
            "label": f"{row.get('Dia', '')} {row.get('Banco', '')}",
            "CV HC": _num(str(row.get("CV HC (%)", "")).replace(".", "").replace(",", ".")),
            "CV Total": _num(str(row.get("CV Total (%)", "")).replace(".", "").replace(",", ".")),
            "Desv.Pressão": _num(str(row.get("Desv.Pressão (barg)", "")).replace(".", "").replace(",", ".")),
        })

    return {
        "summary": summary,
        "anomalies": anomalies,
        "comparisons": comparisons,
        "pressure_flags": pressure_flags,
        "stability": stability,
        "chart_items": chart_items,
        "composition_rows": composition_rows,
        "hc_total_rows": hc_total_rows,
        "deviation_points": deviation_points,
        "control_hc": control_hc,
        "control_total": control_total,
        "stability_chart": stability_chart,
        "preflight": preflight_rows,
    }


def _dashboard_comparison_records(context_df: pd.DataFrame, target_days: list, aligned_bank: str = SEP_ALIGNED_BANK) -> list:
    """Monta somente os pares físicos Topside/Subsea configurados."""
    df = _normalize_master_columns(context_df)
    if df.empty:
        return []
    for col in ("ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Bank", "Tag"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    mpfm = df[
        df["ProductionDate"].isin({str(day) for day in target_days})
        & (df["Origin"] == "MPFM")
        & (df["SourceType"] == "PDF")
        & df["Granularity"].isin(["Daily", "Hourly"])
    ].copy()

    records = []
    for _, row in mpfm.iterrows():
        bank = str(row.get("Bank", ""))
        if not _row_has_production(row):
            continue
        if not _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")):
            continue
        granularity = str(row.get("Granularity", ""))
        hour = _canonical_master_key_value(row.get("Hour", ""), "Hour") if granularity == "Hourly" else ""
        base = {
            "date": str(row.get("ProductionDate", "")),
            "hour": hour,
            "granularity": granularity,
            "bank": bank,
            "tag": str(row.get("Tag", "")),
        }
        variables = {}
        for key, spec in DASHBOARD_VARIABLES.items():
            value = _num(row.get(spec["mpfm"], ""))
            if not np.isnan(value):
                variables[key] = value
        if not variables:
            continue
        tipo = str(row.get("Tipo", ""))
        tag = str(row.get("Tag", ""))
        pair_info = _comparison_info_for_row(row)
        if pair_info:
            records.append({**base, "pair": pair_info["pair"], "side": pair_info["side"], "role": pair_info["role"], "source": "MPFM", "variables": variables})
    return records


def _separator_frontend_records(context_df: pd.DataFrame, target_days: list) -> dict:
    """Entrega séries independentes ao HTML; não cria produto cartesiano MPFM x SEP."""
    df = _normalize_master_columns(context_df)
    if df.empty:
        return {"mpfm": [], "sep": []}
    for col in ("ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Bank", "Tag", "Tipo"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    days = {str(day) for day in target_days}
    metrics = {
        "hc": ("MPFM corr HC (t)", "SEP HC (t)"),
        "total": ("MPFM corr Total (t)", "SEP Total (t)"),
        "oil": ("MPFM corr Óleo (t)", "SEP Óleo Mass (t)"),
        "gas": ("MPFM corr Gás (t)", "SEP Gás Mass (t)"),
        "water": ("MPFM corr Água (t)", "SEP Água Mass (t)"),
    }
    mpfm_records = []
    mpfm = df[
        df["ProductionDate"].isin(days)
        & (df["Origin"] == "MPFM")
        & (df["SourceType"] == "PDF")
        & df["Granularity"].isin(["Daily", "Hourly"])
    ].copy()
    for _, row in mpfm.iterrows():
        if not _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")):
            continue
        values = {key: _num(row.get(cols[0], "")) for key, cols in metrics.items()}
        values = {key: value for key, value in values.items() if not np.isnan(value)}
        if not values:
            continue
        granularity = str(row.get("Granularity", ""))
        hour = _canonical_master_key_value(row.get("Hour", ""), "Hour") if granularity == "Hourly" else ""
        bank, tag, tipo = str(row.get("Bank", "")), str(row.get("Tag", "")), str(row.get("Tipo", ""))
        mpfm_records.append({
            "date": str(row.get("ProductionDate", "")), "hour": hour, "granularity": granularity,
            "point": f"{bank} — {tipo} — {tag}", "bank": bank, "tag": tag, "values": values,
        })
    sep_records = []
    sep = df[
        df["ProductionDate"].isin(days)
        & (df["Origin"] == "SEP")
        & df["Granularity"].isin(["Daily", "Hourly"])
    ].copy()
    for _, row in sep.iterrows():
        values = {key: _num(row.get(cols[1], "")) for key, cols in metrics.items()}
        values = {key: value for key, value in values.items() if not np.isnan(value)}
        if not values:
            continue
        granularity = str(row.get("Granularity", ""))
        hour = _canonical_master_key_value(row.get("Hour", ""), "Hour") if granularity == "Hourly" else ""
        sep_records.append({"date": str(row.get("ProductionDate", "")), "hour": hour, "granularity": granularity, "values": values})
    return {"mpfm": mpfm_records, "sep": sep_records}


def _interactive_separator_comparison_panel(records: dict, target_days: list) -> str:
    mpfm_records = records.get("mpfm", [])
    sep_records = records.get("sep", [])
    if not mpfm_records or not sep_records:
        return "<p class='muted'>Sem séries MPFM e SEP suficientes para a consulta livre.</p>"
    mpfm_json = json.dumps(mpfm_records, ensure_ascii=False)
    sep_json = json.dumps(sep_records, ensure_ascii=False)
    min_day = str(min(target_days)) if target_days else ""
    max_day = str(max(target_days)) if target_days else ""
    return f"""
<div class="info-box"><b>Consulta sob demanda:</b> o HTML mantém MPFM e Separador como séries independentes. O cruzamento só ocorre no navegador após a escolha do ponto, da variável, da granularidade e do período. A seleção do usuário não comprova alinhamento físico; por isso, os limites são apresentados como indicadores e não como veredito automático de conformidade.</div>
<div class="controls integrated-controls">
  <label>MPFM <select id="sepCmpPoint"></select></label>
  <label>Variável <select id="sepCmpMetric"><option value="hc">Massa HC (t)</option><option value="total">Massa Total (t)</option><option value="oil">Massa Óleo (t)</option><option value="gas">Massa Gás (t)</option><option value="water">Massa Água (t)</option></select></label>
  <label>Granularidade <select id="sepCmpGranularity"><option value="Daily">Diária</option><option value="Hourly">Horária</option></select></label>
  <label>Data inicial <select id="sepCmpFrom"></select></label>
  <label>Data final <select id="sepCmpTo"></select></label>
</div>
<button type="button" class="tab-button" id="sepCmpExport">Exportar comparação para Excel (.csv)</button>
<div id="sepCmpSummary" class="metric-grid"></div>
<div id="sepCmpNotice" class="muted"></div>
<div class="table-wrap"><table class="data-table" id="sepCmpTable"></table></div>
<script>
const SEP_CMP_MPFM = {mpfm_json};
const SEP_CMP_REF = {sep_json};
const sepCmpPoint=document.getElementById('sepCmpPoint'), sepCmpMetric=document.getElementById('sepCmpMetric'), sepCmpGranularity=document.getElementById('sepCmpGranularity');
const sepCmpFrom=document.getElementById('sepCmpFrom'), sepCmpTo=document.getElementById('sepCmpTo'), sepCmpSummary=document.getElementById('sepCmpSummary'), sepCmpNotice=document.getElementById('sepCmpNotice'), sepCmpTable=document.getElementById('sepCmpTable');
const sepCmpExport=document.getElementById('sepCmpExport');
const sepMetric={{hc:{{label:'Massa HC',limit:10}},total:{{label:'Massa Total',limit:7}},oil:{{label:'Massa Óleo',limit:null}},gas:{{label:'Massa Gás',limit:null}},water:{{label:'Massa Água',limit:null}}}};
const sepKey=r=>`${{r.date}}|${{r.granularity}}|${{r.granularity==='Hourly'?String(r.hour):''}}`;
const sepFmt=(v,d=2)=>Number.isFinite(v)?v.toLocaleString('pt-BR',{{minimumFractionDigits:d,maximumFractionDigits:d}}):'—';
const sepCmpDateBR=d=>{{const p=String(d||'').slice(0,10).split('-');return p.length===3?`${{p[2]}}/${{p[1]}}/${{p[0]}}`:String(d||'');}};
[...new Set(SEP_CMP_MPFM.map(r=>r.point))].sort().forEach(point=>sepCmpPoint.add(new Option(point,point)));
const sepCmpDays=[...new Set([...SEP_CMP_MPFM,...SEP_CMP_REF].map(r=>r.date))].filter(Boolean).sort();
sepCmpDays.forEach(d=>{{sepCmpFrom.add(new Option(sepCmpDateBR(d),d));sepCmpTo.add(new Option(sepCmpDateBR(d),d));}});
if(sepCmpDays.length){{sepCmpFrom.value='{html.escape(min_day)}';sepCmpTo.value='{html.escape(max_day)}';}}
let sepLastRows=[];
function sepRender() {{
  const metric=sepCmpMetric.value, spec=sepMetric[metric], gran=sepCmpGranularity.value, d0=sepCmpFrom.value||'0000-00-00', d1=sepCmpTo.value||'9999-99-99';
  const refs=new Map(SEP_CMP_REF.filter(r=>r.granularity===gran&&r.date>=d0&&r.date<=d1).map(r=>[sepKey(r),r]));
  const selected=SEP_CMP_MPFM.filter(r=>r.point===sepCmpPoint.value&&r.granularity===gran&&r.date>=d0&&r.date<=d1).sort((a,b)=>sepKey(a).localeCompare(sepKey(b)));
  const rows=selected.map(m=>{{
    const ref=refs.get(sepKey(m)), mv=Number(m.values?.[metric]), rv=Number(ref?.values?.[metric]);
    const valid=Number.isFinite(mv)&&Number.isFinite(rv)&&Math.abs(rv)>=0.1, dev=valid?(mv/rv-1)*100:null;
    let status='SEM REFERÊNCIA SEP';
    if(ref && Number.isFinite(rv) && Math.abs(rv)<0.1) status='REFERÊNCIA SEP < 0,1 t — percentual suprimido';
    else if(valid && spec.limit===null) status='COMPARAÇÃO DIAGNÓSTICA';
    else if(valid) status=Math.abs(dev)>spec.limit?'INDICADOR ACIMA DO LIMITE':'INDICADOR DENTRO DO LIMITE';
    return {{...m,ref,mv,rv,dev,status}};
  }});
  sepLastRows=rows;
  const validRows=rows.filter(r=>Number.isFinite(r.dev)), low=rows.filter(r=>r.ref&&Number.isFinite(r.rv)&&Math.abs(r.rv)<0.1), out=validRows.filter(r=>spec.limit!==null&&Math.abs(r.dev)>spec.limit);
  sepCmpSummary.innerHTML=`<div class="metric-card"><span>Pontos encontrados</span><b>${{rows.length}}</b><small>somente após os filtros</small></div><div class="metric-card good"><span>Desvios calculáveis</span><b>${{validRows.length}}</b><small>referência SEP ≥ 0,1 t</small></div><div class="metric-card warn"><span>Referência muito baixa</span><b>${{low.length}}</b><small>percentual suprimido</small></div><div class="metric-card ${{out.length?'warn':'good'}}"><span>Indicadores acima do limite</span><b>${{out.length}}</b><small>sem veredito regulatório</small></div>`;
  sepCmpNotice.textContent=rows.length?'Fórmula: (MPFM corrigido − Separador) / Separador × 100. A escolha do MPFM e a responsabilidade pelo alinhamento pertencem ao usuário.':'Sem dados para os filtros selecionados.';
  let body=''; rows.slice(0,500).forEach(r=>{{const moment=r.granularity==='Hourly'?`${{r.date}} H${{String(r.hour).padStart(2,'0')}}`:r.date; body+=`<tr><td>${{moment}}</td><td>${{sepFmt(r.mv)}}</td><td>${{sepFmt(r.rv)}}</td><td>${{Number.isFinite(r.dev)?sepFmt(r.dev)+'%':'—'}}</td><td>${{r.status}}</td></tr>`;}});
  sepCmpTable.innerHTML=`<thead><tr><th>Momento</th><th>MPFM ${{spec.label}}</th><th>SEP ${{spec.label}}</th><th>Desvio</th><th>Classificação informativa</th></tr></thead><tbody>${{body}}</tbody>`;
}}
function sepCsvCell(value) {{ const text=String(value??'').replaceAll('"','""'); return `"${{text}}"`; }}
function sepExportCsv() {{
  if(!sepLastRows.length) {{ alert('Sem dados para exportar nos filtros atuais.'); return; }}
  const metric=sepMetric[sepCmpMetric.value];
  const lines=[["Data","Hora","Granularidade","MPFM selecionado",`MPFM ${{metric.label}} (t)`,`SEP ${{metric.label}} (t)`,"Desvio (%)","Classificação informativa","Fórmula","Observação"]];
  sepLastRows.forEach(r=>lines.push([r.date,r.hour,r.granularity,r.point,Number.isFinite(r.mv)?r.mv:'',Number.isFinite(r.rv)?r.rv:'',Number.isFinite(r.dev)?r.dev:'',r.status,"(MPFM corrigido - SEP) / SEP * 100","Seleção do usuário; alinhamento físico não comprovado pela automação"]));
  const csv='\ufeff'+lines.map(row=>row.map(sepCsvCell).join(';')).join('\\r\\n');
  const url=URL.createObjectURL(new Blob([csv],{{type:'text/csv;charset=utf-8;'}})); const a=document.createElement('a'); a.href=url; a.download=`comparacao_mpfm_sep_${{sepCmpPoint.value.split(' — ')[0]||'selecionado'}}_${{sepCmpMetric.value}}.csv`; a.click(); URL.revokeObjectURL(url);
}}
[sepCmpPoint,sepCmpMetric,sepCmpGranularity,sepCmpFrom,sepCmpTo].forEach(el=>el.addEventListener('change',sepRender)); sepRender();
sepCmpExport.addEventListener('click',sepExportCsv);
</script>
"""


def _interactive_comparison_panel(records: list, target_days: list) -> str:
    if not records:
        return "<p class='muted'>Sem dados úteis para os pares configurados.</p>"
    records_json = json.dumps(records, ensure_ascii=False)
    variables_json = json.dumps({k: {"label": v["label"], "aggregation": v["aggregation"]} for k, v in DASHBOARD_VARIABLES.items()}, ensure_ascii=False)
    min_day = str(min(target_days)) if target_days else ""
    max_day = str(max(target_days)) if target_days else ""
    return f"""
<div class="controls">
    <label>Comparação <select id="cmpPair"></select></label>
    <label>Variável 1 <select id="cmpVar1"></select></label>
    <label>Variável 2 <select id="cmpVar2"><option value="">— nenhuma —</option></select></label>
    <label>Granularidade <select id="cmpGranularity"><option value="Hourly">Horária</option><option value="Daily">Diária</option></select></label>
    <label>Data inicial <select id="cmpDateFrom"></select></label>
    <label>Data final <select id="cmpDateTo"></select></label>
</div>
<div class="info-box">
    <b>O que está sendo comparado:</b> cada linha representa uma série operacional do par selecionado.
    Nos pares Topside/Subsea, compara-se o MPFM do poço/subsea contra o respectivo riser/topside.
</div>
<p class="muted">Gráfico em linhas. Selecione até duas variáveis. A legenda fica em área dedicada para não cobrir filtros, datas ou eixos.</p>
<div id="cmpEmpty" class="muted"></div>
<div id="cmpLegend" class="legend-row"></div>
<div class="chart-scroll"><svg id="cmpChart" class="chart wide-chart" role="img" aria-label="Comparativo operacional"></svg></div>
<div class="table-wrap"><table class="data-table" id="cmpTable"></table></div>
<script>
const CMP_RECORDS = {records_json};
const CMP_VARIABLES = {variables_json};
const pairSelect = document.getElementById('cmpPair');
const variableSelect1 = document.getElementById('cmpVar1');
const variableSelect2 = document.getElementById('cmpVar2');
const granularitySelect = document.getElementById('cmpGranularity');
const dateFromInput = document.getElementById('cmpDateFrom');
const dateToInput = document.getElementById('cmpDateTo');
const chartSvg = document.getElementById('cmpChart');
const legendEl = document.getElementById('cmpLegend');
const tableEl = document.getElementById('cmpTable');
const emptyEl = document.getElementById('cmpEmpty');
const palette = ['#007398', '#f59e0b', '#00a3ad', '#64748b'];

function uniq(values) {{ return [...new Set(values)].filter(Boolean).sort(); }}
function fmt(v, digits=1) {{ return Number.isFinite(v) ? v.toLocaleString('pt-BR', {{maximumFractionDigits: digits}}) : ''; }}
function dateBR(d) {{ const p=String(d||'').slice(0,10).split('-'); return p.length===3 ? `${{p[2]}}/${{p[1]}}/${{p[0]}}` : String(d||''); }}
function categoryOf(r) {{ return r.granularity === 'Hourly' ? `${{r.date}} H${{String(r.hour).padStart(2,'0')}}` : r.date; }}

function fillControls() {{
    uniq(CMP_RECORDS.map(r => r.pair)).forEach(pair => pairSelect.add(new Option(pair, pair)));
    const days = uniq(CMP_RECORDS.map(r => r.date)).sort();
    days.forEach(day => {{ dateFromInput.add(new Option(dateBR(day), day)); dateToInput.add(new Option(dateBR(day), day)); }});
    if (days.length) {{ dateFromInput.value = "{html.escape(min_day)}"; dateToInput.value = "{html.escape(max_day)}"; }}
    Object.entries(CMP_VARIABLES).forEach(([key, spec]) => {{ variableSelect1.add(new Option(spec.label, key)); variableSelect2.add(new Option(spec.label, key)); }});
    variableSelect1.value = 'mpfm_corr_total';
    variableSelect2.value = '';
}}

function selectedVariables() {{
    const vals = [variableSelect1.value, variableSelect2.value].filter((v, i, a) => v && a.indexOf(v) === i);
    if (!vals.length) {{ variableSelect1.value = 'mpfm_corr_total'; vals.push('mpfm_corr_total'); }}
    return vals;
}}

function aggregate(rows, variables) {{
    const points = [];
    for (const variable of variables) {{
    const spec = CMP_VARIABLES[variable];
    const byKey = new Map();
    for (const r of rows) {{
        const val = Number(r.variables?.[variable]);
        if (!Number.isFinite(val)) continue;
        const cat = categoryOf(r);
        const key = `${{cat}}|||${{r.side}}|||${{variable}}`;
        if (!byKey.has(key)) byKey.set(key, {{category: cat, side: r.side, role: r.role, source: r.source, variable, variableLabel: spec.label, date: r.date, values: [], bank: r.bank, tag: r.tag}});
        byKey.get(key).values.push(val);
    }}
    points.push(...[...byKey.values()].map(item => {{
        const value = spec.aggregation === 'avg' ? item.values.reduce((a,b)=>a+b,0)/item.values.length : item.values.reduce((a,b)=>a+b,0);
        return {{...item, value}};
    }}));
    }}
    return points;
}}

function drawChart(points, variables) {{
    chartSvg.innerHTML = '';
    legendEl.innerHTML = '';
    const categories = uniq(points.map(p => p.category));
    const seriesNames = uniq(points.map(p => `${{p.side}} — ${{p.variableLabel}}`));
    if (!categories.length || !seriesNames.length) {{ emptyEl.textContent = 'Sem dados para os filtros selecionados.'; return; }}
    emptyEl.textContent = '';
    const width = Math.max(1080, categories.length * 84);
    const height = 430, left = 92, top = 56, bottom = 100, right = 40;
    const plotW = width - left - right, plotH = height - top - bottom;
    chartSvg.setAttribute('viewBox', `0 0 ${{width}} ${{height}}`);
    chartSvg.style.minWidth = width + 'px';
    const ys = points.map(p => p.value).filter(Number.isFinite);
    const minY = Math.min(...ys, 0), maxY = Math.max(...ys, 1);
    const pad = Math.max((maxY - minY) * 0.12, 1);
    const yMin = minY - pad, yMax = maxY + pad;
    const ns = 'http://www.w3.org/2000/svg';
    function el(name, attrs, text) {{ const e = document.createElementNS(ns, name); Object.entries(attrs || {{}}).forEach(([k,v]) => e.setAttribute(k, v)); if (text !== undefined) e.textContent = text; chartSvg.appendChild(e); return e; }}
    el('text', {{x: 20, y: 28, class: 'chart-title'}}, `${{pairSelect.value}} — ${{variables.map(v => CMP_VARIABLES[v].label).join(' + ')}}`);
    for (let i=0;i<=4;i++) {{
        const y = top + plotH - (i/4)*plotH;
        const val = yMin + (yMax - yMin) * i/4;
        el('line', {{x1:left, y1:y, x2:width-right, y2:y, stroke:'#e2e8f0'}});
        el('text', {{x:12, y:y+4, class:'axis-label'}}, fmt(val));
    }}
    const xPos = cat => left + categories.indexOf(cat) / Math.max(categories.length - 1, 1) * plotW;
    const yPos = val => top + plotH - (val - yMin) / (yMax - yMin) * plotH;
    const bySeries = new Map();
    points.forEach(p => {{ const name = `${{p.side}} — ${{p.variableLabel}}`; if (!bySeries.has(name)) bySeries.set(name, []); bySeries.get(name).push(p); }});
    let sIdx = 0;
    bySeries.forEach((series, name) => {{
        series.sort((a,b) => categories.indexOf(a.category) - categories.indexOf(b.category));
        const color = palette[sIdx % palette.length];
        let d = '';
        series.forEach((p, idx) => {{ const x = xPos(p.category), y = yPos(p.value); d += `${{idx ? 'L' : 'M'}}${{x.toFixed(1)}} ${{y.toFixed(1)}} `; }});
        el('path', {{d, fill:'none', stroke:color, 'stroke-width':'3'}});
        series.forEach(p => {{ const c = el('circle', {{cx:xPos(p.category), cy:yPos(p.value), r:4, fill:color}}); c.appendChild(document.createElementNS(ns,'title')).textContent = `${{p.category}} | ${{name}}: ${{fmt(p.value,2)}}`; }});
        sIdx++;
    }});
    categories.forEach((cat, i) => {{
        const tx = xPos(cat);
        const label = cat.length > 14 ? cat.replace('2026-', '') : cat;
        el('text', {{x:tx, y:height-72, class:'axis-label', 'text-anchor':'end', transform:`rotate(-45 ${{tx}} ${{height-72}})`}}, label);
    }});
    const legendClasses = ['legend-swatch--blue', 'legend-swatch--amber', 'legend-swatch--teal', 'legend-swatch--ink'];
    legendEl.innerHTML = seriesNames.map((name, j) => `<span class="legend-item"><i class="legend-swatch ${{legendClasses[j % legendClasses.length]}}"></i>${{name}}</span>`).join('');
}}

function drawTable(points) {{
    const categories = uniq(points.map(p => p.category));
    const seriesNames = uniq(points.map(p => `${{p.side}} — ${{p.variableLabel}}`));
    let html = '<thead><tr><th>Momento</th>' + seriesNames.map(s => `<th>${{s}}</th>`).join('') + '<th>Desvio oficial (%)</th><th>Referência</th></tr></thead><tbody>';
    const byCatSeries = new Map(points.map(p => [`${{p.category}}|||${{p.side}} — ${{p.variableLabel}}`, p.value]));
    const byCatPoints = new Map();
    points.forEach(p => {{ if (!byCatPoints.has(p.category)) byCatPoints.set(p.category, []); byCatPoints.get(p.category).push(p); }});
    categories.slice(0, 300).forEach(cat => {{
        const vals = seriesNames.map(s => byCatSeries.get(`${{cat}}|||${{s}}`));
        const catPoints = byCatPoints.get(cat) || [];
        let mpfm = catPoints.find(p => p.source === 'MPFM' && p.role === 'Subsea') || catPoints.find(p => p.source === 'MPFM' && p.role === 'MPFM') || catPoints.find(p => p.source === 'MPFM');
        let ref = catPoints.find(p => p.source === 'SEP') || catPoints.find(p => p.role === 'Topside');
        let dev = '', refLabel = '';
        if (mpfm && ref && Number.isFinite(mpfm.value) && Number.isFinite(ref.value) && Math.abs(ref.value) >= 0.1) {{
            dev = fmt((mpfm.value / ref.value - 1) * 100, 2) + '%';
            refLabel = `${{mpfm.side}} ÷ ${{ref.side}}`;
        }} else if (mpfm && ref && Number.isFinite(ref.value) && Math.abs(ref.value) < 0.1) {{
            refLabel = 'Referência < 0,1 — desvio percentual suprimido';
        }}
        html += `<tr><td>${{cat}}</td>${{vals.map(v => `<td>${{fmt(v,2)}}</td>`).join('')}}<td>${{dev}}</td><td>${{refLabel}}</td></tr>`;
    }});
    html += '</tbody>';
    tableEl.innerHTML = html;
}}

function renderComparison() {{
    const pair = pairSelect.value, variables = selectedVariables(), gran = granularitySelect.value;
    const d0 = dateFromInput.value || '0000-00-00', d1 = dateToInput.value || '9999-99-99';
    const rows = CMP_RECORDS.filter(r => r.pair === pair && r.granularity === gran && r.date >= d0 && r.date <= d1);
    const points = aggregate(rows, variables);
    drawChart(points, variables);
    drawTable(points);
}}
fillControls();
[pairSelect, variableSelect1, variableSelect2, granularitySelect, dateFromInput, dateToInput].forEach(el => el.addEventListener('change', renderComparison));
renderComparison();
</script>
"""


def _availability_visual(preflight_rows: list) -> str:
    if not preflight_rows:
        return "<p class='muted'>Sem pré-validação disponível.</p>"
    df = pd.DataFrame(preflight_rows)
    if "Status" not in df.columns:
        return _html_table(preflight_rows)
    counts = df["Status"].astype(str).value_counts().to_dict()
    total = max(sum(counts.values()), 1)
    ok = counts.get("OK", 0)
    not_ok = total - ok
    parts = [
        "<div class='info-box'><b>Como interpretar:</b> cada item avaliado representa a disponibilidade esperada de PDF MPFM, dados hourly/daily, SEP e alarmes/eventos na janela. Status OK indica que a automação encontrou o insumo; status diferente de OK indica pendência operacional antes de confiar em 100% da janela.</div>",
        "<div class='metric-grid'>",
        f"<div class='metric-card good'><span>Itens OK</span><b>{ok}</b><small>{ok / total * 100:.1f}% da janela</small></div>",
        f"<div class='metric-card warn'><span>Pendências</span><b>{not_ok}</b><small>{not_ok / total * 100:.1f}% da janela</small></div>",
        f"<div class='metric-card'><span>Total avaliado</span><b>{total}</b><small>bancos, dias e fontes</small></div>",
        "</div><div class='stacked-bar'>",
    ]
    offset = 0.0
    for status, count in sorted(counts.items()):
        pct = count / total * 100
        color = "#22c55e" if status == "OK" else "#f59e0b"
        parts.append(f"<rect x='{offset:.3f}' y='0' width='{pct:.3f}' height='24' fill='{color}'><title>{html.escape(status)}: {count}</title></rect>")
        offset += pct
    parts.append("</div><div class='legend-row'>")
    for status, count in sorted(counts.items()):
        color = "#22c55e" if status == "OK" else "#f59e0b"
        parts.append(f"<span class='legend-item'><i class='legend-swatch {'legend-swatch--green' if status == 'OK' else 'legend-swatch--amber'}'></i>{html.escape(status)}: {count}</span>")
    parts.append("</div>")
    return "".join(parts)


def _summary_visual(summary_rows: list) -> str:
    if not summary_rows:
        return "<p class='muted'>Sem linhas extraídas.</p>"
    df = pd.DataFrame(summary_rows)
    if "Granularidade" not in df.columns or "Linhas" not in df.columns:
        return ""
    grouped = df.groupby("Granularidade", dropna=False)["Linhas"].sum().reset_index()
    total = max(float(grouped["Linhas"].sum()), 1.0)
    bank_grouped = df.groupby("Banco", dropna=False)["Linhas"].sum().reset_index().sort_values("Linhas", ascending=False).head(8) if "Banco" in df.columns else pd.DataFrame()
    parts = [
        "<div class='info-box'><b>Como interpretar:</b> este resumo mostra quanto dado entrou na Base_Unica. Use a distribuição por granularidade para conferir se a janela tem Daily/Hourly esperados e a distribuição por banco para perceber lacunas ou concentração de dados.</div>",
        "<div class='metric-grid'>",
        f"<div class='metric-card'><span>Total de linhas</span><b>{int(total)}</b><small>extraídas para a janela</small></div>",
    ]
    for _, row in grouped.iterrows():
        label = str(row["Granularidade"])
        count = float(row["Linhas"])
        pct = count / total * 100
        parts.append(f"<div class='metric-card accent'><span>{html.escape(label)}</span><b>{int(count)}</b><small>{pct:.1f}% do total</small></div>")
    parts.append("</div><div class='mini-bars'>")
    for _, row in bank_grouped.iterrows():
        label = str(row["Banco"] or "Sem banco")
        count = float(row["Linhas"])
        pct = count / total * 100
        parts.append(f"<div class='mini-row'><b>{html.escape(label)}</b><span>{int(count)} linhas</span><svg class='mini-track-svg' viewBox='0 0 100 18' preserveAspectRatio='none' role='img' aria-label='{html.escape(label)}'><rect x='0' y='0' width='{pct:.1f}' height='18' fill='#007398'></rect></svg></div>")
    parts.append("<p class='muted'>Itens não extraídos aparecem na seção de disponibilidade da janela, com status diferente de OK.</p></div>")
    return "".join(parts)


def _coverage_calendar_panel(context_df: pd.DataFrame, target_days: list) -> str:
    """Calendário de cobertura por ponto MPFM, fase e granularidade."""
    df = _normalize_master_columns(context_df)
    if df.empty or not target_days:
        return "<p class='muted'>Sem dados para montar o calendário de cobertura.</p>"
    days = sorted({str(day) for day in target_days})
    mpfm = df[
        (df["Origin"].astype(str) == "MPFM")
        & (df["Granularity"].astype(str).isin(["Daily", "Hourly"]))
        & (df["ProductionDate"].astype(str).isin(days))
    ].copy()
    if mpfm.empty:
        return "<p class='muted'>Nenhum MPFM foi encontrado na janela.</p>"
    for col in ("ProductionDate", "Granularity", "Bank", "Tipo", "Tag", "Hour"):
        mpfm[col] = mpfm[col].where(pd.notna(mpfm[col]), "").astype(str)
    mpfm = mpfm[mpfm.apply(lambda row: _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    if mpfm.empty:
        return "<p class='muted'>Nenhum MPFM liberado para publicação foi encontrado na janela.</p>"
    points = sorted({(row["Bank"], row["Tipo"], row["Tag"]) for _, row in mpfm.iterrows()}, key=lambda item: item)
    phase_columns = {"oleo": "MPFM corr Óleo (t)", "gas": "MPFM corr Gás (t)", "agua": "MPFM corr Água (t)", "total": "MPFM corr Total (t)"}
    group_columns = ["ProductionDate", "Granularity", "Bank", "Tipo", "Tag"]
    numeric_columns = list(phase_columns.values())
    for column in numeric_columns:
        mpfm[column] = pd.to_numeric(mpfm.get(column, pd.Series(dtype=float)), errors="coerce")
    grouped = mpfm.groupby(group_columns, dropna=False).agg(
        Actual=("Hour", "nunique"),
        **{key: (column, "sum") for key, column in phase_columns.items()},
    ).reset_index()
    grouped_lookup = {
        (str(row["ProductionDate"]), str(row["Granularity"]), str(row["Bank"]), str(row["Tipo"]), str(row["Tag"])): row
        for _, row in grouped.iterrows()
    }
    records = []
    for granularity in ("Daily", "Hourly"):
        for day in days:
            for bank, tipo, tag in points:
                grouped_row = grouped_lookup.get((day, granularity, bank, tipo, tag))
                expected = 1 if granularity == "Daily" else 24
                if granularity == "Daily":
                    actual = 1 if grouped_row is not None else 0
                else:
                    actual = int(grouped_row["Actual"]) if grouped_row is not None else 0
                total_sum = float(grouped_row["total"]) if grouped_row is not None and pd.notna(grouped_row["total"]) else 0.0
                if actual == 0:
                    state = "AUSENTE"
                elif granularity == "Hourly" and actual < expected:
                    state = "PARCIAL"
                elif abs(float(total_sum)) <= 0.001:
                    state = "FECHADO"
                else:
                    state = "OK"
                values = {}
                for key, column in phase_columns.items():
                    value = float(grouped_row[key]) if grouped_row is not None and actual and pd.notna(grouped_row[key]) else np.nan
                    values[key] = {"value": None if np.isnan(value) else value, "state": "AUSENTE" if actual == 0 else ("ZERO" if abs(value) <= 0.001 else "OK")}
                records.append({"day": day, "granularity": granularity, "bank": bank, "tipo": tipo, "tag": tag, "expected": expected, "actual": actual, "state": state, "values": values})
    data_json = json.dumps(records, ensure_ascii=False)
    points_json = json.dumps([{"bank": b, "tipo": t, "tag": g} for b, t, g in points], ensure_ascii=False)
    days_json = json.dumps(days, ensure_ascii=False)
    return f"""
<div class="controls">
  <label>Granularidade <select id="calGranularity"><option value="Daily">Diária — 1 ponto/dia</option><option value="Hourly">Horária — 24 pontos/dia</option></select></label>
  <label>Variável <select id="calVariable"><option value="oleo">Óleo</option><option value="gas">Gás</option><option value="agua">Água</option><option value="total">Total</option></select></label>
  <label>Filtro MPFM <select id="calPoint"><option value="ALL">Todos os pontos</option></select></label>
</div>
<div class="info-box"><b>Como ler:</b> o calendário mostra o que era esperado por ponto. <b>OK</b> = dado extraído; <b>PARCIAL</b> = Hourly abaixo de 24 pontos; <b>AUSENTE</b> = não localizado; <b>FECHADO</b> = ponto presente, mas produção total zerada. O zero de uma fase pode ser fisicamente válido e aparece como <b>ZERO</b>.</div>
<div id="calLegend" class="legend-row"><span class="legend-item"><i class="cal-ok"></i>OK</span><span class="legend-item"><i class="cal-partial"></i>PARCIAL</span><span class="legend-item"><i class="cal-absent"></i>AUSENTE</span><span class="legend-item"><i class="cal-closed"></i>FECHADO</span></div>
<div class="table-wrap calendar-wrap"><table class="data-table calendar-table" id="coverageCalendar"></table></div>
<script>
const CALENDAR_DATA = {data_json};
const CALENDAR_POINTS = {points_json};
const CALENDAR_DAYS = {days_json};
const calGranularity = document.getElementById('calGranularity');
const calVariable = document.getElementById('calVariable');
const calPoint = document.getElementById('calPoint');
const calTable = document.getElementById('coverageCalendar');
function calFmt(value) {{ return Number.isFinite(Number(value)) ? Number(value).toLocaleString('pt-BR', {{maximumFractionDigits: 2}}) : '—'; }}
function calClass(state) {{ return {{OK:'cal-ok', PARCIAL:'cal-partial', AUSENTE:'cal-absent', FECHADO:'cal-closed'}}[state] || 'cal-absent'; }}
function fillCalPoints() {{ CALENDAR_POINTS.forEach(p => calPoint.add(new Option(`${{p.bank}} — ${{p.tipo}} — ${{p.tag}}`, `${{p.bank}}|||${{p.tipo}}|||${{p.tag}}`))); }}
function renderCalendar() {{
  const gran = calGranularity.value, variable = calVariable.value, selected = calPoint.value;
  const rows = CALENDAR_DATA.filter(r => r.granularity === gran && (selected === 'ALL' || `${{r.bank}}|||${{r.tipo}}|||${{r.tag}}` === selected));
  const byPoint = new Map();
  rows.forEach(r => {{ const key = `${{r.bank}}|||${{r.tipo}}|||${{r.tag}}`; if (!byPoint.has(key)) byPoint.set(key, {{...r, cells: new Map()}}); byPoint.get(key).cells.set(r.day, r); }});
  let out = '<thead><tr><th>MPFM / ponto</th>' + CALENDAR_DAYS.map(d => `<th>${{d.slice(5)}}</th>`).join('') + '</tr></thead><tbody>';
  byPoint.forEach(p => {{ out += `<tr><th>${{p.bank}}<br><small>${{p.tipo}} — ${{p.tag}}</small></th>`; CALENDAR_DAYS.forEach(day => {{ const r=p.cells.get(day); const v=r?.values?.[variable] || {{value:null,state:'AUSENTE'}}; out += `<td><div class="calendar-cell ${{calClass(r?.state)}}"><b>${{calFmt(v.value)}}</b><small>${{r ? `${{r.actual}}/${{r.expected}} ${{gran === 'Hourly' ? 'h' : 'registro'}}` : '0/1 registro'}}</small><em>${{r?.state || 'AUSENTE'}} · ${{v.state}}</em></div></td>`; }}); out += '</tr>'; }});
  calTable.innerHTML = out + '</tbody>';
}}
fillCalPoints(); [calGranularity, calVariable, calPoint].forEach(el => el.addEventListener('change', renderCalendar)); renderCalendar();
</script>
"""


def _alarm_category_visual(alarm_event_rows: list) -> str:
    """Pareto executivo de alarmes por categoria/criticidade."""
    if not alarm_event_rows:
        return "<p class='muted'>Sem alarmes/eventos para classificar.</p>"
    df = pd.DataFrame(alarm_event_rows)
    if df.empty:
        return "<p class='muted'>Sem alarmes/eventos para classificar.</p>"
    counts = defaultdict(lambda: {"Total": 0, "Critical": 0, "General": 0})
    for _, row in df.iterrows():
        priority = str(row.get("Priority", "General") or "General")
        raw = str(row.get("IssueFlag", "") or "").strip()
        categories = [item.strip().upper() for item in raw.split(",") if item.strip()] or ["SEM CLASSIFICAÇÃO"]
        for category in categories:
            counts[category]["Total"] += 1
            counts[category][priority] = counts[category].get(priority, 0) + 1
    ordered = sorted(counts.items(), key=lambda item: (-item[1]["Total"], item[0]))
    total = max(sum(value["Total"] for _, value in ordered), 1)
    width, height, left, right, top, bottom = max(820, len(ordered) * 92), 380, 64, 64, 42, 100
    plot_w, plot_h = width - left - right, height - top - bottom
    max_count = max((value["Total"] for _, value in ordered), default=1)
    cumulative = 0
    parts = ["<div class='info-box'><b>Pareto de alarmes por categoria:</b> categorias em ordem decrescente; a linha mostra o percentual acumulado. A categoria é derivada de <i>IssueFlag</i>; flags múltiplas contam em cada categoria aplicável.</div>", f"<div class='chart-scroll'><svg class='chart wide-chart' viewBox='0 0 {width} {height}' width='{width}' role='img' aria-label='Pareto de alarmes por categoria'>"]
    parts.append(f"<text x='18' y='25' class='chart-title'>Pareto de alarmes/eventos — categorias e acumulado</text>")
    for i in range(5):
        y = top + plot_h - i / 4 * plot_h
        parts.append(f"<line x1='{left}' y1='{y:.1f}' x2='{width-right}' y2='{y:.1f}' stroke='#e2e8f0'></line><text x='10' y='{y+4:.1f}' class='axis-label'>{int(max_count*i/4)}</text>")
    points = []
    for idx, (category, value) in enumerate(ordered):
        x = left + (idx + .5) * plot_w / len(ordered)
        bar_w = plot_w / len(ordered) * .62
        bar_h = value["Total"] / max_count * plot_h
        y = top + plot_h - bar_h
        critical = value.get("Critical", 0)
        parts.append(f"<rect x='{x-bar_w/2:.1f}' y='{y:.1f}' width='{bar_w:.1f}' height='{bar_h:.1f}' fill='{ '#dc2626' if critical else '#007398' }'><title>{html.escape(category)}: {value['Total']} ocorrências — {critical} Critical / {value.get('General', 0)} General</title></rect>")
        parts.append(f"<text x='{x:.1f}' y='{y-6:.1f}' class='axis-label' text-anchor='middle'>{value['Total']}</text><text x='{x:.1f}' y='{height-54}' class='axis-label' text-anchor='end' transform='rotate(-45 {x:.1f} {height-54})'>{html.escape(category[:18])}</text>")
        cumulative += value["Total"]
        points.append((x, top + plot_h - cumulative / total * plot_h, cumulative / total * 100))
    path = ' '.join(f"{'M' if i == 0 else 'L'}{x:.1f} {y:.1f}" for i, (x, y, _) in enumerate(points))
    parts.append(f"<path d='{path}' fill='none' stroke='#a15c00' stroke-width='3'></path>")
    for x, y, pct in points:
        parts.append(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='4' fill='#a15c00'><title>{pct:.1f}% acumulado</title></circle>")
    parts.append(f"<text x='{width-right+8}' y='{top+4}' class='axis-label'>100%</text></svg></div><div class='legend-row'><span class='legend-item'><i class='legend-swatch legend-swatch--blue'></i>Barras: ocorrências</span><span class='legend-item'><i class='legend-swatch legend-swatch--amber'></i>Linha: acumulado (%)</span><span class='legend-item'><i class='legend-swatch legend-swatch--red'></i>Contém Critical</span></div>")
    return "".join(parts)


def _alarm_history_visual(alarm_event_rows: list) -> str:
    if not alarm_event_rows:
        return "<p class='muted'>Sem alarmes/eventos extraídos para a janela.</p>"
    df = pd.DataFrame(alarm_event_rows)
    if "ProductionDate" not in df.columns or "RecordType" not in df.columns:
        return ""
    df["ProductionDate"] = df["ProductionDate"].astype(str)
    df["RecordType"] = df["RecordType"].astype(str).str.upper()
    grouped = df.groupby(["ProductionDate", "RecordType"], dropna=False).size().reset_index(name="Quantidade")
    days = sorted(grouped["ProductionDate"].unique())
    types = [t for t in ["ALARM", "EVENT"] if t in set(grouped["RecordType"])]
    if not days or not types:
        return ""
    values = {f"{str(r['ProductionDate'])}|{str(r['RecordType']).upper()}": int(r["Quantidade"]) for _, r in grouped.iterrows()}
    chart_data = json.dumps({"days": days, "types": types, "values": values}, ensure_ascii=False).replace("</", "<\\/")
    controls = "<div class='controls alarm-history-controls'><label>Exibir no gráfico<select id='alarmHistoryType'><option value='ALL'>Alarmes e eventos</option><option value='ALARM'>Somente alarmes</option><option value='EVENT'>Somente eventos</option></select></label></div>"
    script = """
<script>
(() => {
  const data = __ALARM_HISTORY_DATA__;
  const select = document.getElementById('alarmHistoryType');
  const host = document.getElementById('alarmHistoryChart');
  const colors = {ALARM:'#dc2626', EVENT:'#007398'};
  function render() {
    const selected = select.value;
    const types = selected === 'ALL' ? data.types : data.types.filter(t => t === selected);
    const width = Math.max(720, data.days.length * 70 + 120), height = 300, left = 58, top = 34, bottom = 70;
    const plotH = height - top - bottom, maxTotal = Math.max(1, ...data.days.map(day => types.reduce((sum,t) => sum + (data.values[day+'|'+t] || 0), 0)));
    const barW = Math.min(42, Math.max(18, (width-left-40)/Math.max(data.days.length,1)*.58));
    const ns='http://www.w3.org/2000/svg', svg=document.createElementNS(ns,'svg');
    svg.setAttribute('class','chart wide-chart'); svg.setAttribute('viewBox',`0 0 ${width} ${height}`); svg.setAttribute('width',width); svg.setAttribute('role','img'); svg.setAttribute('aria-label','Histórico filtrado de alarmes e eventos');
    const add=(name,attrs,text)=>{const e=document.createElementNS(ns,name);Object.entries(attrs||{}).forEach(([k,v])=>e.setAttribute(k,v));if(text!==undefined)e.textContent=text;svg.appendChild(e);return e;};
    add('text',{x:18,y:24,class:'chart-title'}, selected==='ALL'?'Histórico diário de alarmes/eventos':'Histórico diário — '+(selected==='ALARM'?'alarmes':'eventos'));
    for(let i=0;i<5;i++){const y=top+plotH-i/4*plotH,val=maxTotal*i/4;add('line',{x1:left,y1:y,x2:width-30,y2:y,stroke:'#e2e8f0'});add('text',{x:10,y:y+4,class:'axis-label'},Math.round(val));}
    data.days.forEach((day,idx)=>{const x=left+idx*((width-left-50)/Math.max(data.days.length,1))+16;let cursor=top+plotH;types.forEach(t=>{const count=data.values[day+'|'+t]||0,h=count/maxTotal*plotH;cursor-=h;const r=add('rect',{x:x.toFixed(1),y:cursor.toFixed(1),width:barW.toFixed(1),height:h.toFixed(1),fill:colors[t]||'#64748b'});const title=document.createElementNS(ns,'title');title.textContent=day+' '+t+': '+count;r.appendChild(title);});add('text',{x:x+barW/2,y:height-38,class:'axis-label','text-anchor':'end',transform:`rotate(-45 ${x+barW/2} ${height-38})`},day.length>=10?day.slice(5):day);});
    host.replaceChildren(svg);
  }
  select.addEventListener('change',render); render();
})();
</script>""".replace("__ALARM_HISTORY_DATA__", chart_data)
    legend = "<div class='legend-row'><span class='legend-item'><i class='legend-swatch legend-swatch--red'></i>ALARM</span><span class='legend-item'><i class='legend-swatch legend-swatch--blue'></i>EVENT</span></div>"
    return "<div class='info-box'><b>Como interpretar:</b> barras maiores indicam dias com maior atividade. Use o filtro para analisar somente alarmes ou somente eventos; os detalhes permanecem na tabela abaixo.</div>" + controls + "<div class='chart-scroll' id='alarmHistoryChart'></div>" + legend + script


def _methodology_panel() -> str:
    return """
<div class="formula-grid">
  <div class="formula-card"><b>Desvio HC — memorial/RANP 44</b><span>d_HC (%) = [(Mass_HC_MPFM_Corrected / Mass_HC_Ref) − 1] × 100</span><small>Limite crítico: ±10,0%. No cálculo automático oficial, a referência é o riser/topside do par físico. MPFM×SEP é consulta selecionada pelo usuário no frontend. Percentual suprimido se |referência| &lt; 0,1 t.</small></div>
  <div class="formula-card"><b>Desvio Total — memorial/RANP 44</b><span>d_Total (%) = [(Mass_Total_MPFM_Corrected / Mass_Total_Ref) − 1] × 100</span><small>Limite crítico: ±7,0%. Persistência por 10 dias válidos consecutivos é tratada como gatilho de investigação/restabelecimento, conforme a base aprovada aplicável.</small></div>
  <div class="formula-card"><b>Desvios por fase — diagnóstico</b><span>d_fase (%) = [((Mass_fase_MPFM_Uncorrected × K_fase) / Mass_fase_Ref) − 1] × 100</span><small>Na Base_Unica, Mass_fase_MPFM_Uncorrected × K_fase é representado por MPFM corr Óleo/Gás/Água. Fases suportam diagnóstico PVT/K-factor e não disparam desenquadramento automático.</small></div>
  <div class="formula-card"><b>Fatores K oficiais</b><span>Ko/Kg/Kw = Mass_fase_Ref_Back-flashed / Mass_fase_MPFM_Uncorrected</span><small>Regras especiais do memorial: acima do ponto de bolha Kg = 1,0000 e Ko = Khc; se óleo/água ≥ 25, Kw = 1,0000.</small></div>
  <div class="formula-card"><b>Spike / ponto suspeito</b><span>Spike quando valor &gt; média móvel + 3σ ou valor &lt; média móvel − 3σ</span><small>É um alerta estatístico, não prova de falha. Investigue contexto operacional e alarmes.</small></div>
  <div class="formula-card"><b>Estabilidade</b><span>Média móvel e σ calculados na janela: 24 pontos horários ou 7 pontos diários</span><small>Série estável tende a permanecer dentro dos limites ±3σ e sem saltos persistentes.</small></div>
</div>
"""


def _official_deviation_pct(mpfm_value, reference_value, min_reference: float = 0.1):
    """(MPFM - referência) / referência × 100, com denominador confiável."""
    mpfm = _num(mpfm_value)
    ref = _num(reference_value)
    if np.isnan(mpfm) or np.isnan(ref) or abs(ref) < abs(float(min_reference)):
        return np.nan
    return (mpfm / ref - 1.0) * 100.0


def _official_deviation_rows(
    context_df: pd.DataFrame,
    target_days: list,
    aligned_bank: str = SEP_ALIGNED_BANK,
) -> list:
    """Calcula somente os desvios dos pares físicos Topside/Subsea."""
    df = _normalize_master_columns(context_df)
    if df.empty:
        return []
    for col in ("ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Bank", "Tag"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    df = df[
        df["ProductionDate"].isin({str(day) for day in target_days})
        & (df["Origin"] == "MPFM")
        & (df["SourceType"] == "PDF")
        & (df["Granularity"].isin(["Daily", "Hourly"]))
    ].copy()
    rows = []
    key_cols = ["ProductionDate", "Hour", "Granularity"]
    # Métricas cujas colunas realmente existem nesta janela de dados.
    pair_specs = [spec for spec in PAIR_DEVIATION_SPECS if spec[2] in df.columns]
    pair_columns = [spec[2] for spec in pair_specs]
    # Guarda de plausibilidade: algumas leituras horárias de PDF chegam com
    # erro de extração (ex.: OCR/tabela) e resultam em massas fisicamente
    # impossíveis (ordens de grandeza acima do restante da série do mesmo
    # banco). Comparar tais leituras contra a referência gera %desvio
    # absurdos (ex.: >1.000.000%) que não refletem um problema real de
    # medição. Além disso, quando a referência está muito próxima de zero
    # (transiente de partida/parada de poço), a razão explode mesmo com as
    # duas leituras fisicamente plausíveis isoladamente — o %desvio deixa de
    # ser um indicador útil nesse instante. Usamos a mediana de cada banco
    # como linha de base robusta e sinalizamos como suspeita qualquer leitura
    # muito acima OU muito abaixo dela, mantendo a linha na planilha (para
    # auditoria) mas fora da contagem de gatilhos de desenquadramento.
    PLAUSIBILITY_FACTOR = 15.0
    LOW_REFERENCE_FACTOR = 0.1
    for pair in COMPARISON_PAIRS:
        subsea = df[_comparison_side_mask(df, pair, "subsea")].copy()
        topside = df[_comparison_side_mask(df, pair, "topside")].copy()
        if subsea.empty or topside.empty:
            continue
        # Agrega somente as massas. Isso evita duplicação caso uma fonte tenha
        # mais de uma linha/tags para o mesmo momento operacional.
        sub_group = subsea.groupby(key_cols, dropna=False).agg(
            **{column: (column, lambda s: pd.to_numeric(s, errors="coerce").sum(min_count=1)) for column in pair_columns}
        ).reset_index()
        top_group = topside.groupby(key_cols, dropna=False).agg(
            **{column: (column, lambda s: pd.to_numeric(s, errors="coerce").sum(min_count=1)) for column in pair_columns}
        ).reset_index()
        merged = sub_group.merge(top_group, on=key_cols, how="inner", suffixes=("_Subsea", "_Topside"))
        # Daily e Hourly têm ordens de grandeza naturalmente diferentes.
        # Misturar as duas medianas marcava os totais diários como spikes por
        # compará-los contra uma linha de base horária.
        sub_medians = {}
        top_medians = {}
        for granularity in ("Daily", "Hourly"):
            sub_values = pd.to_numeric(
                subsea.loc[subsea["Granularity"] == granularity, "MPFM corr Total (t)"], errors="coerce"
            )
            top_values = pd.to_numeric(
                topside.loc[topside["Granularity"] == granularity, "MPFM corr Total (t)"], errors="coerce"
            )
            sub_values = sub_values[sub_values.abs() > 0.001]
            top_values = top_values[top_values.abs() > 0.001]
            sub_medians[granularity] = float(sub_values.median()) if not sub_values.empty else None
            top_medians[granularity] = float(top_values.median()) if not top_values.empty else None
        for _, joined in merged.iterrows():
            granularity = str(joined.get("Granularity", ""))
            sub_median = sub_medians.get(granularity)
            top_median = top_medians.get(granularity)
            subsea_total = _num(joined.get("MPFM corr Total (t)_Subsea", ""))
            topside_total = _num(joined.get("MPFM corr Total (t)_Topside", ""))
            if np.isnan(subsea_total) or np.isnan(topside_total) or abs(subsea_total) <= 0.001 or abs(topside_total) <= 0.001:
                continue
            row_suspect_high = bool(
                (sub_median and abs(subsea_total) > PLAUSIBILITY_FACTOR * abs(sub_median))
                or (top_median and abs(topside_total) > PLAUSIBILITY_FACTOR * abs(top_median))
            )
            row_suspect_low = bool(
                (sub_median and abs(subsea_total) < LOW_REFERENCE_FACTOR * abs(sub_median))
                or (top_median and abs(topside_total) < LOW_REFERENCE_FACTOR * abs(top_median))
            )
            for metric, label, column, limit in pair_specs:
                sub_value = joined.get(f"{column}_Subsea", np.nan)
                top_value = joined.get(f"{column}_Topside", np.nan)
                dev = _official_deviation_pct(sub_value, top_value, min_reference=0.1)
                if np.isnan(dev):
                    continue
                if row_suspect_high:
                    status = "DADO SUSPEITO (revisar PDF fonte)"
                elif row_suspect_low:
                    status = "REFERÊNCIA PRÓXIMA DE ZERO (transiente)"
                elif limit is None:
                    status = "DIAGNÓSTICO"
                else:
                    status = "CONFORME" if abs(dev) <= limit else "FORA DO LIMITE"
                rows.append({
                    "Dia": joined.get("ProductionDate", ""),
                    "Hora": joined.get("Hour", ""),
                    "Granularidade": joined.get("Granularity", ""),
                    "Banco": pair["pair"],
                    "Tag": f"{pair['subsea_label']} × {pair['topside_label']}",
                    "Métrica": f"{label} — Subsea × Topside",
                    "MPFM corrigido (t)": _fmt(sub_value, 2),
                    "Referência (t)": _fmt(top_value, 2),
                    "Desvio oficial (%)": _fmt(dev, 2),
                    "Limite": "—" if limit is None else f"±{_fmt(limit, 1)}%",
                    "Classe": "DIAGNÓSTICO" if limit is None else "CRÍTICO RANP 44",
                    "Referência da comparação": pair["topside_label"],
                    "Status": status,
                    "DesvioNum": float(dev),
                    "MPFMNum": _num(sub_value),
                    "ReferenciaNum": _num(top_value),
                    "MetricaChave": metric,
                    "LimiteNum": limit,
                })

    # Implementa os três gatilhos temporais. Somente desvios críticos diários
    # consecutivos acumulam dias; horas e linhas diagnósticas não acumulam.
    offending_days = defaultdict(set)
    for item in rows:
        if item.get("Status") == "FORA DO LIMITE" and item.get("Granularidade") == "Daily":
            offending_days[(item.get("Banco", ""), item.get("Tag", ""), item.get("Métrica", ""))].add(str(item.get("Dia", ""))[:10])

    def consecutive_ending(day_text: str, days: set[str]) -> int:
        try:
            current = datetime.strptime(day_text[:10], "%Y-%m-%d")
        except ValueError:
            return 0
        count = 0
        while current.strftime("%Y-%m-%d") in days:
            count += 1
            current -= timedelta(days=1)
        return count

    for item in rows:
        key = (item.get("Banco", ""), item.get("Tag", ""), item.get("Métrica", ""))
        count = consecutive_ending(str(item.get("Dia", "")), offending_days.get(key, set())) if item.get("Granularidade") == "Daily" else 0
        item["Dias consecutivos fora"] = count
        item["Gatilho 3 dias"] = "ATIVO" if count >= 3 else "—"
        item["Gatilho 6 dias"] = "ATIVO" if count >= 6 else "—"
        item["Gatilho 10 dias"] = "ATIVO" if count >= 10 else "—"
    return rows


def _interactive_cep_panel(rows: list) -> str:
    """Mostra a comparação Topside/Subsea em gráfico CEP selecionável.

    O usuário escolhe até duas métricas e ambas são plotadas com suas próprias
    linhas de limite (±10% para HC, ±7% para Massa Total). As métricas de fase
    em volume padrão são diagnósticas e entram sem linha de limite.
    """
    records = []
    metric_labels = {}
    source_days = {str(row.get("Dia", ""))[:10] for row in rows or [] if str(row.get("Dia", "")).strip()}
    preferred_granularity = "Hourly" if len(source_days) == 1 else "Daily"
    for row in rows or []:
        metric_text = str(row.get("Métrica", ""))
        if "Subsea × Topside" not in metric_text or str(row.get("Granularidade", "")) != preferred_granularity:
            continue
        metric = str(row.get("MetricaChave", "")).strip()
        if not metric:
            continue
        deviation = row.get("DesvioNum")
        try:
            deviation = float(deviation)
        except (TypeError, ValueError):
            continue
        limit = row.get("LimiteNum")
        limit = None if limit is None else float(limit)
        metric_labels.setdefault(metric, metric_text.split(" — ")[0])
        records.append({
            "pair": str(row.get("Banco", "")),
            "pairLabel": str(row.get("Tag", "")),
            "date": str(row.get("Dia", ""))[:10],
            "month": str(row.get("Dia", ""))[:7],
            "metric": metric,
            "deviation": deviation,
            "limit": limit,
            "status": str(row.get("Status", "")),
            "consecutive": int(row.get("Dias consecutivos fora", 0) or 0),
            "trigger3": str(row.get("Gatilho 3 dias", "—")),
            "trigger6": str(row.get("Gatilho 6 dias", "—")),
            "trigger10": str(row.get("Gatilho 10 dias", "—")),
        })
    if not records:
        return "<p class='muted'>Sem dados diários válidos para comparação Topside × Subsea.</p>"
    data_json = json.dumps(records, ensure_ascii=False)
    labels_json = json.dumps(metric_labels, ensure_ascii=False)
    return """
<div class="info-box"><b>Visão mensal padrão:</b> este gráfico usa todos os dias disponíveis na Base Única para o mês escolhido, mesmo quando o HTML foi gerado a partir de uma janela menor. Selecione o par e até duas métricas. A linha central é zero; as linhas tracejadas são ±10% para HC e ±7% para Massa Total. As fases são diagnósticas.</div>
<div class="controls">
  <label>Mês de referência <select id="cepMonth"></select></label>
  <label>Par Subsea × Topside <select id="cepPair"></select></label>
  <label>Métrica 1 <select id="cepVar1"></select></label>
  <label>Métrica 2 <select id="cepVar2"><option value="">— nenhuma —</option></select></label>
</div>
<div id="cepCards" class="metric-grid"></div>
<div id="cepLegend" class="legend-row"></div>
<div id="cepEmpty" class="muted"></div>
<div class="chart-scroll"><svg id="cepChart" class="chart wide-chart" role="img" aria-label="Gráfico CEP Topside e Subsea"></svg></div>
<div class="table-wrap"><table class="data-table" id="cepTable"></table></div>
<script>
const CEP_RECORDS = __CEP_DATA__;
const CEP_LABELS = __CEP_LABELS__;
const cepMonth = document.getElementById('cepMonth'), cepPair = document.getElementById('cepPair'), cepVar1 = document.getElementById('cepVar1'), cepVar2 = document.getElementById('cepVar2');
const cepCards = document.getElementById('cepCards'), cepLegend = document.getElementById('cepLegend'), cepEmpty = document.getElementById('cepEmpty'), cepChart = document.getElementById('cepChart'), cepTable = document.getElementById('cepTable');
const cepColors = ['#007398', '#a15c00'];
const cepLimitColors = {10:'#dc2626',7:'#9a4a1f'};
const cepFmt = (v, d=1) => Number.isFinite(v) ? v.toLocaleString('pt-BR', {maximumFractionDigits:d}) + '%' : '';
const cepEsc = v => String(v ?? '').replace(/[&<>"']/g, m => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
const cepUniq = values => [...new Set(values)].filter(Boolean);
function cepLimitOf(metric) { const found = CEP_RECORDS.find(r => r.metric === metric && r.limit !== null && r.limit !== undefined); return found ? found.limit : null; }
function cepFill() {
  cepUniq(CEP_RECORDS.map(r => r.month)).sort().forEach(month => cepMonth.add(new Option(month, month)));
  if (cepMonth.options.length) cepMonth.value = cepMonth.options[cepMonth.options.length - 1].value;
  cepUniq(CEP_RECORDS.map(r => r.pair)).forEach(pair => {
    const sample = CEP_RECORDS.find(r => r.pair === pair);
    cepPair.add(new Option(`${pair} | ${sample?.pairLabel || ''}`, pair));
  });
  cepUniq(CEP_RECORDS.map(r => r.metric)).forEach(metric => {
    const limit = cepLimitOf(metric);
    const text = `${CEP_LABELS[metric] || metric}${limit === null ? ' — diagnóstico' : ` — limite ±${limit}%`}`;
    cepVar1.add(new Option(text, metric)); cepVar2.add(new Option(text, metric));
  });
  cepVar1.value = cepVar1.options.length ? cepVar1.options[0].value : '';
  if ([...cepVar1.options].some(o => o.value === 'HC')) cepVar1.value = 'HC';
  cepVar2.value = [...cepVar2.options].some(o => o.value === 'Total') ? 'Total' : '';
}
function cepSelected() { return [cepVar1.value, cepVar2.value].filter((v,i,a) => v && a.indexOf(v) === i); }
function cepRender() {
  const month = cepMonth.value, pair = cepPair.value, metrics = cepSelected();
  cepChart.innerHTML = ''; cepTable.innerHTML = ''; cepCards.innerHTML = ''; cepLegend.innerHTML = '';
  const series = metrics.map((metric, i) => ({
    metric, color: cepColors[i % cepColors.length], limit: cepLimitOf(metric),
    label: CEP_LABELS[metric] || metric,
    rows: CEP_RECORDS.filter(r => r.month === month && r.pair === pair && r.metric === metric).sort((a,b) => a.date.localeCompare(b.date)),
  })).filter(s => s.rows.length);
  if (!series.length) { cepEmpty.textContent = 'Sem dados para esta combinação.'; return; }
  cepEmpty.textContent = '';
  cepCards.innerHTML = series.map(s => {
    const out = s.limit === null ? 0 : s.rows.filter(r => Math.abs(r.deviation) > s.limit).length;
    const maxConsecutive = Math.max(...s.rows.map(r => r.consecutive || 0), 0);
    const active3 = s.rows.filter(r => (r.consecutive||0) >= 3).length;
    const active6 = s.rows.filter(r => (r.consecutive||0) >= 6).length;
    const active10 = s.rows.filter(r => (r.consecutive||0) >= 10).length;
    const cls = s.limit === null ? '' : (out ? 'warn' : 'good');
    return `<div class="metric-card ${cls}"><span>${cepEsc(s.label)}</span><div class="kv-list"><div class="kv"><span>Fora do limite</span><b>${s.limit === null ? '—' : out}</b></div><div class="kv"><span>Máx. consecutivo</span><b>${maxConsecutive} d</b></div><div class="kv"><span>Gatilhos 3/6/10 d</span><b>${active3}/${active6}/${active10}</b></div></div><small>${s.limit === null ? 'diagnóstico — sem limite RANP' : `limite ±${s.limit}%`}</small></div>`;
  }).join('');
  const cats = [...new Set(series.flatMap(s => s.rows.map(r => r.date)))].sort();
  const limits = series.map(s => s.limit).filter(v => v !== null && v !== undefined);
  const width = Math.max(980, cats.length * 72), height = 460, left = 84, top = 82, right = 78, bottom = 86, plotW = width-left-right, plotH = height-top-bottom;
  cepChart.setAttribute('viewBox', `0 0 ${width} ${height}`); cepChart.style.minWidth = width + 'px';
  const guard = limits.length ? Math.max(...limits) + 2 : 2;
  const values = series.flatMap(s => s.rows.map(r => r.deviation)).concat(limits, limits.map(v => -v), [0]);
  const yMin = Math.min(...values, -guard), yMax = Math.max(...values, guard);
  const x = i => left + (cats.length > 1 ? i / (cats.length-1) * plotW : plotW/2), y = v => top + plotH - (v-yMin)/(yMax-yMin)*plotH;
  const ns = 'http://www.w3.org/2000/svg';
  function el(name, attrs, text) { const e=document.createElementNS(ns,name); Object.entries(attrs||{}).forEach(([k,v])=>e.setAttribute(k,v)); if(text!==undefined)e.textContent=text; cepChart.appendChild(e); return e; }
  el('text', {x:20,y:28,class:'chart-title'}, `${pair} — ${month} — ${series.map(s=>s.label).join('  ×  ')}`);
  const legendItems = [
    ...series.map(s => ({label:s.label, color:s.color, dash:'', kind:'line'})),
    ...[...new Set(limits)].sort((a,b)=>b-a).map(limit => ({label:`Limite ±${limit}%`, color:cepLimitColors[Number(limit)] || '#dc2626', dash:'7 5', kind:'line'})),
    {label:'Linha zero', color:'#141f20', dash:'', kind:'line'}
  ];
  let lx = 20, ly = 52;
  legendItems.forEach(item => {
    const labelWidth = Math.min(210, 38 + item.label.length * 7);
    if (lx + labelWidth > width - 20) { lx = 20; ly += 18; }
    el('line', {x1:lx, y1:ly-4, x2:lx+24, y2:ly-4, stroke:item.color, 'stroke-width': item.label === 'Linha zero' ? '2' : '3', 'stroke-dasharray':item.dash});
    el('text', {x:lx+30, y:ly, class:'axis-label'}, item.label);
    lx += labelWidth;
  });
  for (let i=0;i<=4;i++) { const value=yMin+(yMax-yMin)*i/4, yy=y(value); el('line',{x1:left,y1:yy,x2:width-right,y2:yy,stroke:'#e1e7e9'}); el('text',{x:10,y:yy+4,class:'axis-label'},cepFmt(value,1)); }
  const yy0 = y(0); el('line',{x1:left,y1:yy0,x2:width-right,y2:yy0,stroke:'#141f20','stroke-width':'2'}); el('text',{x:width-right+8,y:yy0+4,class:'axis-label'},'0%');
  series.forEach(s => { if (s.limit === null) return;
    [s.limit, -s.limit].forEach(value => { const yy=y(value), limitColor=cepLimitColors[Number(s.limit)] || '#dc2626'; el('line',{x1:left,y1:yy,x2:width-right,y2:yy,stroke:limitColor,'stroke-width':'1.8','stroke-dasharray':'7 5'}); el('text',{x:width-right+8,y:yy+4,class:'axis-label'},`${value>0?'+':''}${value}%`); }); });
  series.forEach(s => {
    const byDate = new Map(s.rows.map(r => [r.date, r]));
    const pts = cats.map((cat,i) => ({i, r: byDate.get(cat)})).filter(p => p.r);
    const path = pts.map((p,k) => `${k?'L':'M'}${x(p.i).toFixed(1)} ${y(p.r.deviation).toFixed(1)}`).join(' ');
    el('path',{d:path,fill:'none',stroke:s.color,'stroke-width':'3'});
    pts.forEach(p => { const bad = s.limit !== null && Math.abs(p.r.deviation) > s.limit;
      const c = el('circle',{cx:x(p.i),cy:y(p.r.deviation),r:bad?6:4,fill:s.color,stroke:bad?'#a12020':'none','stroke-width':bad?'2':'0'});
      c.appendChild(document.createElementNS(ns,'title')).textContent=`${s.label} | ${p.r.date} | ${cepFmt(p.r.deviation,2)} | ${s.limit===null?'DIAGNÓSTICO':(bad?'FORA DO LIMITE':'CONFORME')}`; });
  });
  cats.forEach((cat,i)=>el('text',{x:x(i),y:height-bottom+24,class:'axis-label','text-anchor':'middle'},cat));
  cepLegend.innerHTML = series.map(s=>`<span class="legend-item"><i class="legend-swatch" style="background:${s.color}"></i>${cepEsc(s.label)}${s.limit===null?' (diagnóstico)':` (±${s.limit}%)`}</span>`).join('');
  const tableRows = series.flatMap(s => s.rows.map(r => ({s, r}))).sort((a,b) => a.r.date.localeCompare(b.r.date) || a.s.label.localeCompare(b.s.label));
  cepTable.innerHTML = '<thead><tr><th>Data</th><th>Métrica</th><th>Desvio</th><th>Limite</th><th>Status</th><th>Dias consecutivos</th><th>3 dias</th><th>6 dias</th><th>10 dias</th></tr></thead><tbody>' + tableRows.map(({s,r})=>`<tr><td data-label="Data">${cepEsc(r.date)}</td><td data-label="Métrica">${cepEsc(s.label)}</td><td data-label="Desvio">${cepFmt(r.deviation,2)}</td><td data-label="Limite">${s.limit===null?'—':'±'+s.limit+'%'}</td><td data-label="Status">${cepEsc(r.status)}</td><td data-label="Dias consecutivos">${r.consecutive}</td><td data-label="3 dias">${cepEsc(r.trigger3)}</td><td data-label="6 dias">${cepEsc(r.trigger6)}</td><td data-label="10 dias">${cepEsc(r.trigger10)}</td></tr>`).join('') + '</tbody>';
}
cepFill(); [cepMonth,cepPair,cepVar1,cepVar2].forEach(e=>e.addEventListener('change',cepRender)); cepRender();
</script>
""".replace("__CEP_DATA__", data_json).replace("__CEP_LABELS__", labels_json)


def _dashboard_numeric(value):
    try:
        number = float(value)
        return number if np.isfinite(number) else None
    except (TypeError, ValueError):
        return None


def _monthly_cep_days(context_df: pd.DataFrame, selected_days: list) -> list[str]:
    """Dias da Base Única pertencentes ao(s) mês(es) da janela publicada."""
    selected_months = {str(day)[:7] for day in selected_days if len(str(day)) >= 7}
    if context_df is None or context_df.empty or not selected_months:
        return sorted({str(day)[:10] for day in selected_days if str(day).strip()})
    available = context_df.get("ProductionDate", pd.Series(dtype=str)).astype(str).str[:10]
    return sorted({day for day in available if day[:7] in selected_months and re.fullmatch(r"\d{4}-\d{2}-\d{2}", day)})


def _dashboard_preferred_granularity(target_days: list, context_df: pd.DataFrame | None = None) -> str:
    """Prefere horário em um dia, com fallback para Daily quando necessário."""
    preferred = "Hourly" if len({str(day)[:10] for day in target_days}) == 1 else "Daily"
    if context_df is None or context_df.empty or preferred != "Hourly":
        return preferred
    work = _normalize_master_columns(context_df)
    days = {str(day)[:10] for day in target_days}
    day_mask = work["ProductionDate"].astype(str).str[:10].isin(days) & work["Origin"].astype(str).eq("MPFM")
    granularities = set(work.loc[day_mask, "Granularity"].astype(str))
    return "Hourly" if "Hourly" in granularities else "Daily"


def _dashboard_mpfm_records(
    context_df: pd.DataFrame,
    target_days: list,
    comparativo_df: pd.DataFrame | None = None,
) -> list:
    """Extrato compacto para a visão integrada; não inclui o SEP.

    As massas vêm da Base Única. Variáveis de processo/modo, quando
    disponíveis, vêm da COMPARATIVO_TOTAL e são ligadas somente pela chave
    explícita data+banco+TAG. A ausência continua como ausência: nenhum zero
    é criado para preencher GVF, WLR, GOR ou modo de cálculo.
    """
    df = _normalize_master_columns(context_df).copy()
    all_rows = df.copy()
    required = ["ProductionDate", "Hour", "Granularity", "Origin", "Bank", "Tipo", "Tag", "Entity", "Instrumento"]
    for col in required:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    days = {str(day)[:10] for day in target_days}
    df = df[
        df["ProductionDate"].str[:10].isin(days)
        & df["Granularity"].isin(["Daily", "Hourly"])
        & df["Origin"].eq("MPFM")
    ].copy()
    if not df.empty:
        df = df[df.apply(lambda row: _mpfm_extraction_enabled(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    if not df.empty:
        df = df[df.apply(lambda row: _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    if df.empty:
        return []
    recon_lookup = {}
    recon = all_rows[all_rows["Origin"].astype(str).eq("RECON")].copy()
    for _, recon_row in recon.iterrows():
        key = (
            str(recon_row.get("ProductionDate", ""))[:10],
            str(recon_row.get("Bank", "")),
            _point_key(recon_row.get("Tag", ""), recon_row.get("Instrumento", "")),
        )
        recon_lookup[key] = {
            "reconCoverage": str(recon_row.get("Recon Cobertura", "") or ""),
            "reconHours": str(recon_row.get("Recon Horas", "") or ""),
            "reconStatusHC": str(recon_row.get("Status HC", "") or ""),
            "reconStatusTotal": str(recon_row.get("Status Água", "") or ""),
        }

    process_lookup = {}
    if comparativo_df is not None and not comparativo_df.empty:
        process = comparativo_df.copy().where(pd.notna(comparativo_df), "")
        for _, process_row in process.iterrows():
            key = (
                str(process_row.get("Data", ""))[:10],
                str(process_row.get("Banco", "")),
                _point_key(process_row.get("TAG", ""), process_row.get("TAG", "")),
            )
            process_lookup[key] = {
                "gvf": _dashboard_numeric(process_row.get("GVF (%)")),
                "wlr": _dashboard_numeric(process_row.get("WLR (%)")),
                "gor": _dashboard_numeric(process_row.get("GOR")),
                "velocity": _dashboard_numeric(process_row.get("Velocidade Escoamento (m/s)")),
                "dpIn": _dashboard_numeric(process_row.get("ΔP - Inlet (mbar)")),
                "dpOut": _dashboard_numeric(process_row.get("ΔP - Outlet (mbar)")),
                "continuousPhase": str(process_row.get("Continuous Phase", "") or ""),
                "calculationMode": str(process_row.get("Calculation Mode", "") or ""),
                "meterStatus1": str(process_row.get("Meter Status 1", "") or ""),
                "meterStatus2": str(process_row.get("Meter Status 2", "") or ""),
                "flowWarning": str(process_row.get("Flow Calculation Warn.", "") or ""),
            }
    fields = {
        "oil": "MPFM corr Óleo (t)", "gas": "MPFM corr Gás (t)", "water": "MPFM corr Água (t)",
        "hc": "MPFM corr HC (t)", "total": "MPFM corr Total (t)",
        "oilVol": "PVT @20 vol Óleo (m³)", "gasVol": "PVT @20 vol Gás (Sm³)",
        "waterVol": "PVT @20 vol Água (m³)", "pressure": "Pressão (barg)",
        "temperature": "Temperatura (°C)", "gvf": "GVF (%)", "wlr": "WLR (%)",
    }
    records = []
    for _, row in df.iterrows():
        display_tag = str(row.get("Tag", ""))
        instrument = str(row.get("Instrumento", ""))
        bank = str(row.get("Bank", ""))
        record = {
            "date": str(row.get("ProductionDate", ""))[:10], "bank": bank,
            "point": _point_key(display_tag, instrument),
            "pointName": _point_name(display_tag, instrument),
            "pointLabel": _point_display_label(display_tag, bank, instrument),
            "hour": _canonical_master_key_value(row.get("Hour", ""), "Hour") if str(row.get("Granularity", "")) == "Hourly" else "",
            "granularity": str(row.get("Granularity", "")),
            "type": str(row.get("Tipo", "")), "tag": display_tag,
            "entity": str(row.get("Entity", "")),
        }
        record.update({key: _dashboard_numeric(row.get(column)) for key, column in fields.items()})
        lookup_key = (record["date"], bank, record["point"])
        record.update(process_lookup.get(lookup_key, {}))
        record.update(recon_lookup.get(lookup_key, {}))
        records.append(record)
    return records


def _dashboard_separator_records(context_df: pd.DataFrame, target_days: list) -> list:
    """Extrato próprio do Separador, mantido fora do conjunto MPFM."""
    df = _normalize_master_columns(context_df).copy()
    required = ["ProductionDate", "Granularity", "Origin", "Hour", "Tag", "Entity", "SEP TAG", "SEP Medidor", "SEP Local", "SEP Status"]
    for col in required:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    days = {str(day)[:10] for day in target_days}
    df = df[df["ProductionDate"].str[:10].isin(days) & df["Origin"].eq("SEP") & df["Granularity"].isin(["Daily", "Hourly"])].copy()
    if df.empty:
        return []
    fields = {
        "oil": "SEP Óleo Mass (t)", "gas": "SEP Gás Mass (t)", "water": "SEP Água Mass (t)",
        "hc": "SEP HC (t)", "total": "SEP Total (t)", "oilVol": "SEP Óleo NSV (sm³)",
        "gasVol": "SEP Gás St. Vol. (m³)", "waterVol": "SEP Água NSV (sm³)",
        "pressure": "SEP Pressão Méd. (barg)", "temperature": "SEP Temperatura Méd. (°C)",
        "oilBsw": "SEP Óleo BSW (%)", "waterBsw": "SEP Água BSW (%)",
    }
    records = []
    for _, row in df.iterrows():
        record = {
            "date": str(row.get("ProductionDate", ""))[:10], "hour": str(row.get("Hour", "")),
            "granularity": str(row.get("Granularity", "")), "tag": str(row.get("SEP TAG") or row.get("Tag", "")),
            "meter": str(row.get("SEP Medidor", "")), "local": str(row.get("SEP Local", "")),
            "status": str(row.get("SEP Status", "")),
        }
        record.update({key: _dashboard_numeric(row.get(column)) for key, column in fields.items()})
        records.append(record)
    return records


def _leadership_dashboard_panel(
    context_df: pd.DataFrame,
    target_days: list,
    comparativo_df: pd.DataFrame | None = None,
    official_rows: list | None = None,
) -> str:
    records = _dashboard_mpfm_records(context_df, target_days, comparativo_df)
    if not records:
        return "<p class='muted'>Sem dados Daily/MPFM para a visão de liderança.</p>"
    separator_records = _dashboard_separator_records(context_df, target_days)
    data_json = json.dumps(records, ensure_ascii=False, allow_nan=False)
    sep_json = json.dumps(separator_records, ensure_ascii=False, allow_nan=False)
    days_json = json.dumps(sorted({r["date"] for r in records}), ensure_ascii=False)
    official_json = json.dumps([
        {
            "date": str(row.get("Dia", ""))[:10],
            "hour": str(row.get("Hora", "")),
            "granularity": str(row.get("Granularidade", "")),
            "pair": str(row.get("Banco", "")),
            "metric": str(row.get("MetricaChave", "")),
            "deviation": _dashboard_numeric(row.get("DesvioNum")),
            "limit": _dashboard_numeric(row.get("LimiteNum")),
            "status": str(row.get("Status", "")),
            "consecutive": int(row.get("Dias consecutivos fora", 0) or 0),
        }
        for row in (official_rows or [])
        if row.get("MetricaChave") in {"HC", "Total"}
    ], ensure_ascii=False, allow_nan=False)
    return """
<div class="integrated-note"><b>Visão para decisão</b><span>Um dia usa dados horários; janelas maiores usam a consolidação diária. Campos ausentes aparecem como “Sem dado”.</span></div>
<div class="controls integrated-controls">
  <label>Data inicial <select id="leadFrom"></select></label>
  <label>Data final <select id="leadTo"></select></label>
  <label>Granularidade <select id="leadGranularity"><option value="Daily">Diária</option><option value="Hourly">Horária</option></select></label>
  <label>Tipo <select id="leadType"><option value="ALL">Todos</option></select></label>
  <label>Ponto de medição <select id="leadPoint"><option value="ALL">Todos</option></select></label>
  <label>Par físico <select id="leadPair"><option value="ALL">Todos os pares</option></select></label>
  <label>Variável 1 <select id="leadMetric1"></select></label>
  <label>Variável 2 <select id="leadMetric2"><option value="">— nenhuma —</option></select></label>
  <label>Busca na tabela <input id="leadSearch" type="search" placeholder="Ponto, TAG, banco ou local"></label>
</div>
<div id="leadOverview" class="integrated-overview"></div>
<div id="leadPairSummary" class="comparison-strip"></div>
<div id="leadCards" class="metric-grid"></div>
<div class="chart-grid"><div><h3>Evolução diária por ponto de medição</h3><div id="leadLegend" class="legend-row"></div><div class="chart-scroll"><svg id="leadTrend" class="chart wide-chart" role="img" aria-label="Evolução por ponto de medição"></svg></div></div><div><h3>Ranking de pontos no último dia filtrado</h3><div id="leadRankControls" class="rank-filter-row" aria-label="Variável do ranking"></div><div class="chart-scroll"><svg id="leadRank" class="chart wide-chart" role="img" aria-label="Ranking dos pontos"></svg></div></div></div>
<div id="leadFlow" class="measurement-flow" aria-label="Cadeia resumida da medição"></div>
<h3>Medições por ponto</h3><div class="table-wrap"><table class="data-table" id="leadTable"></table></div>
<script>
const LEAD_RECORDS = __LEAD_DATA__;
const LEAD_SEP_RECORDS = __LEAD_SEP_DATA__;
const LEAD_DAYS = __LEAD_DAYS__;
const LEAD_OFFICIAL = __LEAD_OFFICIAL__;
const leadFrom = document.getElementById('leadFrom'), leadTo = document.getElementById('leadTo'), leadGranularity = document.getElementById('leadGranularity'), leadType = document.getElementById('leadType'), leadPoint = document.getElementById('leadPoint'), leadPair = document.getElementById('leadPair'), leadMetric1 = document.getElementById('leadMetric1'), leadMetric2 = document.getElementById('leadMetric2'), leadSearch = document.getElementById('leadSearch');
const leadOverview = document.getElementById('leadOverview'), leadFlow = document.getElementById('leadFlow'), leadPairSummary = document.getElementById('leadPairSummary'), leadCards = document.getElementById('leadCards'), leadLegend = document.getElementById('leadLegend'), leadTrend = document.getElementById('leadTrend'), leadRank = document.getElementById('leadRank'), leadTable = document.getElementById('leadTable');
const leadLabels = {total:'Massa total corrigida (t)',hc:'HC corrigida (t)',oil:'Óleo corrigido (t)',gas:'Gás corrigido (t)',water:'Água corrigida (t)',oilVol:'Volume óleo padrão (m³)',gasVol:'Volume gás padrão (Sm³)',waterVol:'Volume água padrão (m³)'};
const leadPalette = ['#007398','#a15c00','#206f77','#8a3d68','#4a5f2a','#1d5d8f','#9a4a1f','#5b4b8a','#2f6d4f','#7a2f3f','#3f6f8a','#6b6b1f'];
const leadFmt = (v,d=1) => v !== null && v !== '' && Number.isFinite(Number(v)) ? Number(v).toLocaleString('pt-BR',{minimumFractionDigits:d,maximumFractionDigits:d}) : 'Sem dado';
const leadEsc = v => String(v ?? '').replace(/[&<>"']/g, m => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
const leadDateBR=d=>{const p=String(d||'').slice(0,10).split('-');return p.length===3?`${p[2]}/${p[1]}/${p[0]}`:String(d||'');};
function leadFill(){
  LEAD_DAYS.forEach(d=>{ leadFrom.add(new Option(leadDateBR(d),d)); leadTo.add(new Option(leadDateBR(d),d)); });
  leadFrom.value=LEAD_DAYS[0]; leadTo.value=LEAD_DAYS[LEAD_DAYS.length-1];
  [...new Set(LEAD_RECORDS.map(r=>r.type).filter(Boolean))].sort().forEach(v=>leadType.add(new Option(v,v)));
  const points = new Map(); LEAD_RECORDS.forEach(r=>{ if(r.point && !points.has(r.point)) points.set(r.point, r.pointLabel || r.point); });
  [...points.entries()].sort((a,b)=>a[1].localeCompare(b[1],'pt-BR')).forEach(([value,label])=>leadPoint.add(new Option(label,value)));
  [...new Set(LEAD_OFFICIAL.map(r=>r.pair).filter(Boolean))].sort().forEach(v=>leadPair.add(new Option(v,v)));
  Object.entries(leadLabels).forEach(([k,label])=>{ leadMetric1.add(new Option(label,k)); leadMetric2.add(new Option(label,k)); });
  leadMetric1.value='total'; leadMetric2.value='';
}
function leadFiltered(){ const a=leadFrom.value,b=leadTo.value,g=leadGranularity.value,type=leadType.value,point=leadPoint.value,q=leadSearch.value.toLowerCase().trim(); return LEAD_RECORDS.filter(r=>r.date>=a&&r.date<=b&&r.granularity===g&&(type==='ALL'||r.type===type)&&(point==='ALL'||r.point===point)&&(!q||[r.pointName,r.bank,r.tag,r.entity,r.type].join(' ').toLowerCase().includes(q))); }
function leadSepFiltered(){ const a=leadFrom.value,b=leadTo.value; const rows=LEAD_SEP_RECORDS.filter(r=>r.date>=a&&r.date<=b); const daily=rows.filter(r=>r.granularity==='Daily'); return daily.length?daily:rows; }
const leadValues=(rs,k)=>rs.map(r=>r[k]).filter(v=>v!==null&&v!==''&&Number.isFinite(Number(v))).map(Number);
const leadSum=(rs,k)=>{const v=leadValues(rs,k);return v.length?v.reduce((a,x)=>a+x,0):null};
const leadAvg=(rs,k)=>{const v=leadValues(rs,k);return v.length?v.reduce((a,x)=>a+x,0)/v.length:null};
const leadGor=(rs)=>{const gas=leadSum(rs,'gasVol'),oil=leadSum(rs,'oilVol');return gas!==null&&oil!==null&&Math.abs(oil)>0.001?gas/oil:null};
function leadKvCard(title,rows,cls,note){ const body=rows.map(([label,value,unit,digits])=>`<div class="kv"><span>${leadEsc(label)}</span><b>${leadFmt(value,digits)} ${leadEsc(unit)}</b></div>`).join(''); return `<div class="metric-card ${cls||''}"><span>${leadEsc(title)}</span><div class="kv-list">${body}</div>${note?`<small>${leadEsc(note)}</small>`:''}</div>`; }
function leadGroupCards(rs){
  return `<div class="metric-grid">`
    + leadKvCard('Corrigidos',[['HC corrigido',leadSum(rs,'hc'),'t',1],['Total corrigido',leadSum(rs,'total'),'t',1]],'accent','HC = óleo + gás · Total = óleo + gás + água')
    + leadKvCard('Massas por fase',[['Óleo',leadSum(rs,'oil'),'t',1],['Gás',leadSum(rs,'gas'),'t',1],['Água',leadSum(rs,'water'),'t',2]],'','massas corrigidas')
    + leadKvCard('Volumes por fase',[['Óleo',leadSum(rs,'oilVol'),'m³',1],['Gás',leadSum(rs,'gasVol'),'Sm³',0],['Água',leadSum(rs,'waterVol'),'m³',2]],'','20 °C | 1 atm')
    + `</div><details class="metric-detail"><summary>Validar registros usados nestes consolidados</summary><div class="table-wrap"><table class="data-table"><thead><tr><th>Data</th><th>MPFM</th><th>Tipo</th><th>Óleo (t)</th><th>Gás (t)</th><th>Água (t)</th><th>HC (t)</th><th>Total (t)</th><th>GOR (Sm³/m³)</th></tr></thead><tbody>${rs.map(r=>`<tr><td>${leadEsc(r.date)}</td><td>${leadEsc(r.pointName||r.tag)}</td><td>${leadEsc(r.type||'Separador')}</td><td>${leadFmt(r.oil,3)}</td><td>${leadFmt(r.gas,3)}</td><td>${leadFmt(r.water,3)}</td><td>${leadFmt(r.hc,3)}</td><td>${leadFmt(r.total,3)}</td><td>${leadFmt((Number(r.oilVol)>0?Number(r.gasVol)/Number(r.oilVol):null),3)}</td></tr>`).join('')}</tbody></table></div></details>`;
}
function leadTextValue(rows,key){ const values=[...new Set(rows.map(r=>String(r[key]??'').trim()).filter(Boolean))]; return values.length?values.slice(0,2).join(' · '):'Sem dado'; }
function leadRenderOverview(rows){
  const group=(title,kind)=>{
    const active=rows.filter(r=>r.type===kind&&Number(r.total)>0.001);
    const coverage=active.filter(r=>String(r.reconCoverage||'').startsWith('OK')).length;
    const points=new Set(active.map(r=>r.point).filter(Boolean)).size;
    const statusRows=active.filter(r=>[r.meterStatus1,r.meterStatus2,r.flowWarning].some(v=>String(v||'').trim())).length;
    const cards=[
      ['Pressão média',leadFmt(leadAvg(active,'pressure'),1),'barg','processo'],
      ['Temperatura média',leadFmt(leadAvg(active,'temperature'),1),'°C','processo'],
      ['GVF médio',leadFmt(leadAvg(active,'gvf'),1),'%','PI/Comparativo'],
      ['WLR médio',leadFmt(leadAvg(active,'wlr'),2),'%','PI/Comparativo'],
      ['GOR médio',leadFmt(leadGor(active),3),'Sm³/m³','calculado: gás padrão ÷ óleo padrão'],
      ['Reconciliação',active.length?`${coverage}/${active.length}`:'Sem dado','','pontos com cobertura OK'],
    ];
    return `<section class="integrated-overview-group"><div class="overview-heading"><div><span class="eyebrow-dark">Resumo estatístico da janela</span><h3>${leadEsc(title)}</h3></div><div class="overview-meta">${points} medidor(es) com produção · ${active.length} registros · ${statusRows} com status PI informado</div></div><div class="overview-kpis">${cards.map(([label,value,unit,note])=>`<article class="overview-kpi"><span>${leadEsc(label)}</span><b>${leadEsc(value)}${value==='Sem dado'?'':` <small>${leadEsc(unit)}</small>`}</b><em>${leadEsc(note)}</em></article>`).join('')}</div><div class="mode-line"><span><b>Fase contínua:</b> ${leadEsc(leadTextValue(active,'continuousPhase'))}</span><span><b>Modo de cálculo:</b> ${leadEsc(leadTextValue(active,'calculationMode'))}</span></div></section>`;
  };
  leadOverview.innerHTML=group('MPFM Topside','Topside')+group('MPFM Subsea','Subsea')+`<div class="integrated-note">Separador de Testes: fonte independente. Não entra na contagem, médias, reconciliação ou KPIs acima.</div>`;
  leadOverview.querySelectorAll('.integrated-overview-group').forEach(card=>{const kpi=[...card.querySelectorAll('.overview-kpi')].find(x=>x.querySelector('span')?.textContent==='Pressão média');const value=Number((kpi?.querySelector('b')?.textContent||'').replace(/\\./g,'').replace(',','.').match(/[0-9.]+/)?.[0]);if(Number.isFinite(value)&&value>490){kpi.classList.add('warn');kpi.querySelector('em').textContent='ATENÇÃO · acima de 490 barg / ponto de bolha';}}});
}
function leadRenderFlow(rows){
  const byType=type=>rows.filter(r=>r.type===type), sub=byType('Subsea'), top=byType('Topside'), sep=leadSepFiltered();
  const node=(kind,title,body,note)=>`<article class="flow-node flow-${kind}"><span>${leadEsc(title)}</span><b>${leadEsc(body)}</b><small>${leadEsc(note)}</small></article>`;
  leadFlow.innerHTML=node('subsea','MPFM Subsea',`${leadFmt(leadSum(sub,'total'),1)} t`,'massa total corrigida no filtro')+`<div class="flow-link" aria-hidden="true"></div>`+node('topside','MPFM Topside',`${leadFmt(leadSum(top,'total'),1)} t`,'referência dos pares físicos')+`<div class="flow-link" aria-hidden="true"></div>`+node('balance','Balanço e CEP',`${LEAD_OFFICIAL.filter(r=>r.date>=leadFrom.value&&r.date<=leadTo.value).length} resultados`,'HC ±10% · Total ±7%')+node('separator','Separador de Testes',`${leadFmt(leadSum(sep,'total'),1)} t`,'fonte independente · consulta manual');
}
function leadRenderPairSummary(){
  const a=leadFrom.value,b=leadTo.value,pair=leadPair.value;
  const preferred=leadGranularity.value;
  let rows=LEAD_OFFICIAL.filter(r=>r.date>=a&&r.date<=b&&r.granularity===preferred&&(pair==='ALL'||r.pair===pair));
  const pairs=[...new Set(rows.map(r=>r.pair))];
  if(!rows.length){leadPairSummary.innerHTML='<p class="muted">Sem comparação física válida para o filtro atual.</p>';return;}
  const isSuppressed=r=>r&&/SUSPEITO|PRÓXIMA DE ZERO/.test(r.status);
  const shown=r=>!r?'Sem dado':(isSuppressed(r)?'Suprimido':leadFmt(r.deviation,2)+'%');
  const note=r=>isSuppressed(r)?r.status:(r?'limite ±'+r.limit+'%':'sem resultado');
  const statusText=r=>!r?'Sem dado':(isSuppressed(r)?r.status:(r.limit!==null&&Math.abs(r.deviation)>r.limit?'NÃO CONFORME':'CONFORME'));
  const metricAt=(items,m)=>items.find(r=>r.metric===m);
  const cell=r=>!r?'<b>Sem dado</b><br><small>sem resultado</small>':(isSuppressed(r)?`<b>Não classificável</b><br><small>${leadEsc(r.status)}</small>`:`<b>${leadFmt(r.deviation,2)}%</b><br><small>${statusText(r)}</small>`);
  const detailFor=p=>{const pr=rows.filter(r=>r.pair===p);const moments=[...new Set(pr.map(r=>`${r.date}|${r.granularity==='Hourly'?String(r.hour).padStart(2,'0')+':00':'Diário'}`))].sort();const trs=moments.map(k=>{const [date,moment]=k.split('|');const items=pr.filter(r=>r.date===date&&(r.granularity==='Hourly'?String(r.hour).padStart(2,'0')+':00':'Diário')===moment);const hc=metricAt(items,'HC'),tot=metricAt(items,'Total');return `<tr><td>${leadEsc(date)}</td><td>${leadEsc(moment)}</td><td>${cell(hc)}</td><td>${hc?'±'+leadEsc(hc.limit)+'%':'—'}</td><td>${cell(tot)}</td><td>${tot?'±'+leadEsc(tot.limit)+'%':'—'}</td></tr>`}).join('');return `<div class="pair-detail" hidden data-detail-pair="${leadEsc(p)}"><h4>${leadEsc(p)} · valores da janela filtrada</h4><div class="table-wrap"><table class="data-table"><thead><tr><th>Data</th><th>Momento</th><th>HC</th><th>Limite HC</th><th>Total</th><th>Limite Total</th></tr></thead><tbody>${trs}</tbody></table></div></div>`};
  const cards=pairs.map(p=>{const pr=rows.filter(r=>r.pair===p);const latest=pr.reduce((m,r)=>`${r.date}|${String(r.hour).padStart(2,'0')}`>`${m.date}|${String(m.hour).padStart(2,'0')}`?r:m,pr[0]);const atMoment=pr.filter(r=>r.date===latest.date&&r.hour===latest.hour);const hc=metricAt(atMoment,'HC'),tot=metricAt(atMoment,'Total');const suspect=[hc,tot].some(isSuppressed);const bad=!suspect&&[hc,tot].some(r=>r&&r.limit!==null&&Math.abs(r.deviation)>r.limit);return `<article class="pair-card ${suspect?'pair-warn':(bad?'pair-bad':'pair-ok')}" tabindex="0" role="button" aria-expanded="false" title="Clique para ver os valores da janela filtrada" data-pair="${leadEsc(p)}"><span>${leadEsc(p)}</span><div><b>HC ${shown(hc)}</b><small>${leadEsc(note(hc))}</small></div><div><b>Total ${shown(tot)}</b><small>${leadEsc(note(tot))}</small></div><em>${leadEsc(latest.date)}${latest.granularity==='Hourly'?' · '+String(latest.hour).padStart(2,'0')+':00':''}</em></article>`}).join('');
  leadPairSummary.innerHTML=`<div class="comparison-strip-title"><div><span class="eyebrow-dark">Comparações físicas</span><h3>Subsea × Topside</h3></div><small>desvio = (Subsea corrigido − Topside) / Topside × 100</small></div><div class="pair-card-grid">${cards}${pairs.map(detailFor).join('')}</div>`;
  leadPairSummary.querySelectorAll('.pair-card').forEach(card=>{
    const toggle=()=>{
      const pair=card.dataset.pair||'';
      const detail=[...leadPairSummary.querySelectorAll('.pair-detail')].find(el=>el.dataset.detailPair===pair);
      if(!detail)return;
      const opening=detail.hidden;
      leadPairSummary.querySelectorAll('.pair-detail').forEach(el=>el.hidden=true);
      leadPairSummary.querySelectorAll('.pair-card').forEach(el=>el.setAttribute('aria-expanded','false'));
      if(opening){
        detail.hidden=false;
        card.setAttribute('aria-expanded','true');
      }
    };
    card.addEventListener('click',toggle);
    card.addEventListener('keydown',e=>{
      if(e.key==='Enter'||e.key===' '){
        e.preventDefault();
        toggle();
      }
    });
  });
}
function leadRenderCards(rows){
  const sepRows=leadSepFiltered();
  const groups=['Topside','Subsea'];
  const extra=[...new Set(rows.map(r=>r.type).filter(Boolean))].filter(t=>!groups.includes(t));
  const blocks=[...groups,...extra].map(type=>{
    const rs=rows.filter(r=>r.type===type);
    const count=new Set(rs.map(r=>r.point).filter(Boolean)).size;
    return `<section class="dashboard-card-container"><h3>MPFM ${leadEsc(type)}</h3><p class="muted">${count} ponto(s) medido(s) no filtro atual</p>${leadGroupCards(rs)}</section>`;
  });
  const sepCount=new Set(sepRows.map(r=>r.tag).filter(Boolean)).size;
  blocks.push(`<section class="dashboard-card-container"><h3>Separador de Testes</h3><p class="muted">${sepCount} medidor(es) no período — não somado aos MPFMs</p>${leadGroupCards(sepRows)}</section>`);
  leadCards.innerHTML=`<div class="integrated-note"><b>Topside e Subsea são apresentados separadamente.</b> O total não soma grupos físicos distintos. O Separador de Testes é referência independente e nunca entra na soma dos MPFMs.</div>`+blocks.join('');
}
function leadSvgEl(svg,name,attrs,text){ const e=document.createElementNS('http://www.w3.org/2000/svg',name); Object.entries(attrs||{}).forEach(([k,v])=>e.setAttribute(k,v)); if(text!==undefined)e.textContent=text; svg.appendChild(e); return e; }
function leadCategory(r){ return r.granularity==='Hourly' ? `${r.date} H${String(r.hour).padStart(2,'0')}` : r.date; }
function leadRenderLine(rows){
  leadTrend.innerHTML=''; leadLegend.innerHTML='';
  if(!rows.length) return;
  const keys=[leadMetric1.value,leadMetric2.value].filter((v,i,a)=>v&&a.indexOf(v)===i);
  if(!keys.length) return;
  const cats=[...new Set(rows.map(leadCategory))].sort();
  const pointNames=new Map(); rows.forEach(r=>{ if(!pointNames.has(r.point)) pointNames.set(r.point, r.pointName||r.tag||r.point); });
  const points=[...pointNames.keys()];
  const series=[];
  keys.forEach((key,ki)=>points.forEach((point,pi)=>{
    const values=cats.map(cat=>rows.filter(r=>r.point===point&&leadCategory(r)===cat).reduce((a,r)=>a+(Number(r[key])||0),0));
    if(values.some(v=>v!==0)) series.push({key,point,axis:ki,color:leadPalette[pi%leadPalette.length],label:pointNames.get(point),values});
  }));
  if(!series.length) return;
  const width=Math.max(920,cats.length*78),height=380,left=86,right=keys.length>1?96:40,top=44,bottom=64,plotW=width-left-right,plotH=height-top-bottom;
  leadTrend.setAttribute('viewBox',`0 0 ${width} ${height}`); leadTrend.style.minWidth=width+'px';
  const scales=keys.map((key,ki)=>{ const vals=series.filter(s=>s.axis===ki).flatMap(s=>s.values); const max=Math.max(...vals,0), min=Math.min(...vals,0); return {min, max: max===min?max+1:max}; });
  const x=i=>left+(cats.length>1?i/(cats.length-1)*plotW:plotW/2);
  const y=(v,ki)=>{ const s=scales[ki]; return top+plotH-(v-s.min)/(s.max-s.min)*plotH; };
  for(let i=0;i<=4;i++){ const yy=top+plotH-plotH*i/4; leadSvgEl(leadTrend,'line',{x1:left,y1:yy,x2:width-right,y2:yy,stroke:'#e1e7e9'});
    leadSvgEl(leadTrend,'text',{x:10,y:yy+4,class:'axis-label'},leadFmt(scales[0].min+(scales[0].max-scales[0].min)*i/4,1));
    if(keys.length>1) leadSvgEl(leadTrend,'text',{x:width-right+8,y:yy+4,class:'axis-label'},leadFmt(scales[1].min+(scales[1].max-scales[1].min)*i/4,1)); }
  leadSvgEl(leadTrend,'text',{x:20,y:28,class:'chart-title'},keys.map(k=>leadLabels[k]).join('  ×  '));
  series.forEach(s=>{ const d=s.values.map((v,i)=>`${i?'L':'M'}${x(i).toFixed(1)} ${y(v,s.axis).toFixed(1)}`).join(' ');
    const attrs={d,fill:'none',stroke:s.color,'stroke-width':'2.5','data-point':s.point}; if(s.axis===1) attrs['stroke-dasharray']='7 5';
    leadSvgEl(leadTrend,'path',attrs);
    s.values.forEach((v,i)=>{ const c=leadSvgEl(leadTrend,'circle',{cx:x(i),cy:y(v,s.axis),r:3.5,fill:s.color,'data-point':s.point}); c.appendChild(document.createElementNS('http://www.w3.org/2000/svg','title')).textContent=`${s.label} | ${cats[i]} | ${leadLabels[s.key]}: ${leadFmt(v,2)}`; }); });
  cats.forEach((cat,i)=>leadSvgEl(leadTrend,'text',{x:x(i),y:height-bottom+26,class:'axis-label','text-anchor':'middle'},cat));
  const legend=[...pointNames.entries()].map(([point,label])=>{ const color=leadPalette[points.indexOf(point)%leadPalette.length]; return `<span class="legend-item legend-toggle" role="button" tabindex="0" aria-pressed="true" data-point="${leadEsc(point)}"><i class="legend-swatch" style="background:${color}"></i>${leadEsc(label)}</span>`; });
  if(keys.length>1) legend.push(`<span class="legend-item"><i class="legend-swatch legend-swatch--ink"></i>traço contínuo: ${leadEsc(leadLabels[keys[0]])} (eixo esq.) · tracejado: ${leadEsc(leadLabels[keys[1]])} (eixo dir.)</span>`);
  leadLegend.innerHTML=legend.join('');
  leadLegend.querySelectorAll('.legend-toggle').forEach(item=>{ const toggle=()=>{ const point=item.dataset.point; const hidden=item.classList.toggle('legend-off'); item.setAttribute('aria-pressed',String(!hidden)); leadTrend.querySelectorAll('[data-point]').forEach(el=>{if(el.dataset.point===point)el.style.display=hidden?'none':'';}); }; item.addEventListener('click',toggle); item.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key===' '){e.preventDefault();toggle();}}}); });
}
function refreshChartZoom(){ document.querySelectorAll('.chart-scroll[data-zoom-ready="1"]').forEach(wrap=>{ const svg=wrap.querySelector('svg'); const input=wrap.previousElementSibling?.querySelector('input[type="number"]'); if(!svg||!input)return; const scale=Math.min(3,Math.max(.5,Number(wrap.dataset.zoomScale||1))); svg.style.transform='none'; svg.style.zoom=''; svg.style.width='100%'; svg.style.minWidth='0'; svg.style.height='auto'; svg.style.display='block'; if(!wrap.dataset.zoomBaseHeight)wrap.dataset.zoomBaseHeight=String(Math.ceil(svg.getBoundingClientRect().height)); const baseH=Number(wrap.dataset.zoomBaseHeight)||326; const scaledH=Math.max(1,Math.ceil(baseH*scale)); svg.style.transformOrigin='left top'; svg.style.transform=`scale(${scale})`; wrap.style.height=`${scaledH}px`; wrap.style.minHeight=`${scaledH}px`; wrap.style.overflowY='hidden'; input.value=Math.round(scale*100); }); }
function installChartZoom(){ document.querySelectorAll('.chart-scroll').forEach(wrap=>{ if(wrap.dataset.zoomReady){return;} wrap.dataset.zoomReady='1'; wrap.dataset.zoomScale='1'; const bar=document.createElement('div'); bar.className='chart-toolbar'; const out=document.createElement('button'); out.type='button'; out.textContent='−'; out.title='Reduzir zoom'; const reset=document.createElement('input'); reset.type='number'; reset.value='100'; reset.min='50'; reset.max='300'; reset.step='25'; reset.title='Percentual de zoom'; reset.setAttribute('aria-label','Percentual de zoom'); const inc=document.createElement('button'); inc.type='button'; inc.textContent='+'; inc.title='Aumentar zoom'; const apply=(value)=>{wrap.dataset.zoomScale=String(Math.min(3,Math.max(.5,Number(value)||1)));refreshChartZoom();}; out.onclick=()=>apply(Number(wrap.dataset.zoomScale||1)-.25); reset.onchange=()=>apply(Number(reset.value)/100); reset.oninput=()=>{if(reset.value!=='')apply(Number(reset.value)/100);}; inc.onclick=()=>apply(Number(wrap.dataset.zoomScale||1)+.25); bar.append(out,reset,inc); wrap.parentElement.insertBefore(bar,wrap); }); refreshChartZoom(); if(!window.__chartZoomResponsive){ window.__chartZoomResponsive=true; window.addEventListener('resize',()=>{ document.querySelectorAll('.chart-scroll[data-zoom-ready="1"]').forEach(w=>{w.dataset.zoomBaseHeight='';}); requestAnimationFrame(refreshChartZoom); },{passive:true}); } }
function leadRenderRank(rows){ const controls=document.getElementById('leadRankControls'); if(controls){const keys=['total','hc','oil','gas','water','pressure','temperature','gvf','wlr','gor'].filter(k=>leadLabels[k]); controls.innerHTML=keys.map(k=>`<button type="button" class="${leadMetric1.value===k?'active':''}" data-rank-key="${k}">${leadEsc(leadLabels[k])}</button>`).join(''); controls.querySelectorAll('button').forEach(b=>b.addEventListener('click',()=>{leadMetric1.value=b.dataset.rankKey;leadRender();}));} leadRank.innerHTML=''; if(!rows.length)return; const key=leadMetric1.value,last=rows.reduce((a,r)=>r.date>a?r.date:a,'');const groups={};rows.filter(r=>r.date===last).forEach(r=>{const k=r.pointName||r.tag;groups[k]=(groups[k]||0)+(Number(r[key])||0)});const items=Object.entries(groups).sort((a,b)=>b[1]-a[1]).slice(0,10);if(!items.length)return;const width=760,height=Math.max(300,items.length*32+70),left=200,right=90,top=30,bottom=24,plotW=width-left-right,max=Math.max(...items.map(x=>x[1]),1);leadRank.setAttribute('viewBox',`0 0 ${width} ${height}`);leadRank.style.minWidth=width+'px';leadSvgEl(leadRank,'text',{x:20,y:20,class:'chart-title'},`${leadLabels[key]} — ${last}`);items.forEach(([label,value],i)=>{const yy=top+i*32;leadSvgEl(leadRank,'text',{x:left-10,y:yy+16,class:'axis-label','text-anchor':'end'},label);leadSvgEl(leadRank,'rect',{x:left,y:yy+4,width:Math.max(2,value/max*plotW),height:18,fill:'#007398',rx:3});leadSvgEl(leadRank,'text',{x:left+Math.max(2,value/max*plotW)+8,y:yy+18,class:'axis-label'},leadFmt(value,1));}); }
function leadRenderTable(rows){ const cols=['Data','Hora','Ponto','Tipo','TAG','Banco','Pressão (barg)','Temperatura (°C)','Óleo corr. (t)','Gás corr. (t)','Água corr. (t)','HC corr. (t)','Total corr. (t)','Óleo padrão (m³)','Gás padrão (Sm³)','Água padrão (m³)','GVF (%)','WLR (%)','GOR','Fase contínua','Modo de cálculo','Recon cobertura']; const cells=r=>[r.date,r.granularity==='Hourly'?String(r.hour).padStart(2,'0')+':00':'—',r.pointName||r.tag,r.type,r.tag,r.bank,leadFmt(r.pressure,2),leadFmt(r.temperature,2),leadFmt(r.oil,3),leadFmt(r.gas,3),leadFmt(r.water,3),leadFmt(r.hc,3),leadFmt(r.total,3),leadFmt(r.oilVol,2),leadFmt(r.gasVol,0),leadFmt(r.waterVol,2),leadFmt(r.gvf,1),leadFmt(r.wlr,2),leadFmt(r.gor,3),r.continuousPhase||'Sem dado',r.calculationMode||'Sem dado',r.reconCoverage||'Sem dado']; leadTable.innerHTML='<thead><tr>'+cols.map(c=>`<th>${leadEsc(c)}</th>`).join('')+'</tr></thead><tbody>'+rows.slice(0,500).map(r=>'<tr>'+cells(r).map((v,i)=>`<td data-label="${leadEsc(cols[i])}">${leadEsc(v)}</td>`).join('')+'</tr>').join('')+'</tbody>'; }
const LEAD_FILTER_KEY='mpfm-dashboard-integrated-filters-v1';
function leadRestore(){try{const state=JSON.parse(localStorage.getItem(LEAD_FILTER_KEY)||'{}');[[leadFrom,'from'],[leadTo,'to'],[leadType,'type'],[leadPoint,'point'],[leadPair,'pair'],[leadMetric1,'metric1'],[leadMetric2,'metric2']].forEach(([el,key])=>{if(state[key]&&[...el.options].some(o=>o.value===state[key]))el.value=state[key]});}catch(e){}}}
function leadPersist(){try{localStorage.setItem(LEAD_FILTER_KEY,JSON.stringify({from:leadFrom.value,to:leadTo.value,type:leadType.value,point:leadPoint.value,pair:leadPair.value,metric1:leadMetric1.value,metric2:leadMetric2.value}));}catch(e){}}}
function leadRender(){ const rows=leadFiltered(); leadPersist(); leadRenderOverview(rows); leadRenderFlow(rows); leadRenderPairSummary(); leadRenderCards(rows); leadRenderLine(rows); leadRenderRank(rows); leadRenderTable(rows); if(typeof refreshChartZoom==='function')refreshChartZoom(); }
leadFill();leadRestore();[leadFrom,leadTo,leadGranularity,leadType,leadPoint,leadPair,leadMetric1,leadMetric2,leadSearch].forEach(e=>e.addEventListener('input',leadRender));leadRender();installChartZoom();
</script>
""".replace("__LEAD_DATA__", data_json).replace("__LEAD_SEP_DATA__", sep_json).replace("__LEAD_DAYS__", days_json).replace("__LEAD_OFFICIAL__", official_json).replace("{{", "{").replace("}}", "}")


def _separator_dashboard_panel(context_df: pd.DataFrame, target_days: list) -> str:
    records = _dashboard_separator_records(context_df, target_days)
    if not records:
        return "<p class='muted'>Nenhuma linha própria do Separador foi encontrada na janela.</p>"
    data_json = json.dumps(records, ensure_ascii=False, allow_nan=False)
    days_json = json.dumps(sorted({r["date"] for r in records}), ensure_ascii=False)
    return """
<div class="info-box"><b>Aba independente do Separador:</b> os valores abaixo são dados próprios do Separador de Testes. Eles não são somados aos MPFMs e podem ser usados como referência para qualquer MPFM selecionado na aba Comparações.</div>
<div class="controls">
  <label>Data inicial <select id="sepFrom"></select></label>
  <label>Data final <select id="sepTo"></select></label>
  <label>Granularidade <select id="sepGran"><option value="Daily">Diária</option><option value="Hourly">Horária</option></select></label>
  <label>Métrica principal <select id="sepMetric"><option value="total">Massa total (t)</option><option value="hc">Massa HC (t)</option><option value="oil">Massa óleo (t)</option><option value="gas">Massa gás (t)</option><option value="water">Massa água (t)</option><option value="oilVol">Óleo NSV (sm³)</option><option value="gasVol">Gás padrão (m³)</option><option value="waterVol">Água NSV (sm³)</option><option value="pressure">Pressão (barg)</option><option value="temperature">Temperatura (°C)</option></select></label>
  <label>Busca <input id="sepSearch" type="search" placeholder="TAG, medidor ou status"></label>
</div>
<div class="info-box subtle"><b>Variáveis no gráfico:</b> clique nos botões para ocultar/aparecer séries. Com mais de uma variável ativa, o gráfico usa escala normalizada para comparar tendência; o valor real fica no tooltip.</div>
<div id="sepMetricToggles" class="legend-row" aria-label="Variáveis do gráfico do separador"></div>
<div id="sepCards" class="metric-grid"></div>
<div class="chart-grid"><div><h3>Evolução do Separador</h3><div class="chart-scroll"><svg id="sepTrend" class="chart wide-chart" role="img" aria-label="Evolução do Separador"></svg></div></div><div><h3>Composição de massa</h3><div class="chart-scroll"><svg id="sepMix" class="chart wide-chart" role="img" aria-label="Composição de massa do Separador"></svg></div></div></div>
<h3>Dados próprios do Separador</h3><div class="table-wrap"><table class="data-table" id="sepTable"></table></div>
<script>
const SEP_RECORDS = __SEP_DATA__, SEP_DAYS = __SEP_DAYS__;
const sepFrom=document.getElementById('sepFrom'),sepTo=document.getElementById('sepTo'),sepGran=document.getElementById('sepGran'),sepMetricSelect=document.getElementById('sepMetric'),sepSearch=document.getElementById('sepSearch'),sepMetricToggles=document.getElementById('sepMetricToggles'),sepCards=document.getElementById('sepCards'),sepTrend=document.getElementById('sepTrend'),sepMix=document.getElementById('sepMix'),sepTable=document.getElementById('sepTable');
const sepMetricDefs=[
  {key:'total',label:'Total',unit:'t',phase:'Síntese',color:'#0b3a4a',digits:1},
  {key:'hc',label:'HC',unit:'t',phase:'Síntese',color:'#007398',digits:1},
  {key:'oil',label:'Óleo massa',unit:'t',phase:'Óleo',color:'#206f77',digits:2},
  {key:'gas',label:'Gás massa',unit:'t',phase:'Gás',color:'#a15c00',digits:2},
  {key:'water',label:'Água massa',unit:'t',phase:'Água',color:'#1d5d8f',digits:2},
  {key:'oilVol',label:'Óleo NSV',unit:'sm³',phase:'Óleo',color:'#2dd4bf',digits:2},
  {key:'gasVol',label:'Gás padrão',unit:'m³',phase:'Gás',color:'#f97316',digits:1},
  {key:'waterVol',label:'Água NSV',unit:'sm³',phase:'Água',color:'#60a5fa',digits:2},
  {key:'pressure',label:'Pressão',unit:'barg',phase:'Processo',color:'#7c3aed',digits:2},
  {key:'temperature',label:'Temperatura',unit:'°C',phase:'Processo',color:'#dc2626',digits:2}
];
const sepLabels=Object.fromEntries(sepMetricDefs.map(m=>[m.key,`${m.label}${m.unit?` (${m.unit})`:''}`]));
let sepActiveMetrics=new Set(['total']);
const sepDashboardFmt=(v,d=1)=>Number.isFinite(Number(v))?Number(v).toLocaleString('pt-BR',{maximumFractionDigits:d}):'—'; const sepEsc=v=>String(v??'').replace(/[&<>"']/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
const sepDateBR=d=>{const p=String(d||'').slice(0,10).split('-');return p.length===3?`${p[2]}/${p[1]}/${p[0]}`:String(d||'');};
function sepFill(){SEP_DAYS.forEach(d=>{sepFrom.add(new Option(sepDateBR(d),d));sepTo.add(new Option(sepDateBR(d),d));});sepFrom.value=SEP_DAYS[0];sepTo.value=SEP_DAYS[SEP_DAYS.length-1];}
function sepRenderMetricToggles(){sepMetricToggles.innerHTML=sepMetricDefs.map(m=>`<span class="legend-item legend-toggle ${sepActiveMetrics.has(m.key)?'':'legend-off'}" role="button" tabindex="0" aria-pressed="${sepActiveMetrics.has(m.key)}" data-metric="${m.key}" title="${sepEsc(m.phase)}"><i class="legend-swatch" style="background:${m.color}"></i>${sepEsc(m.phase)} · ${sepEsc(m.label)}</span>`).join('');sepMetricToggles.querySelectorAll('[data-metric]').forEach(el=>{el.addEventListener('click',()=>sepToggleMetric(el.dataset.metric));el.addEventListener('keydown',ev=>{if(ev.key==='Enter'||ev.key===' '){ev.preventDefault();sepToggleMetric(el.dataset.metric);}}}});});}
function sepToggleMetric(key){if(sepActiveMetrics.has(key)&&sepActiveMetrics.size>1)sepActiveMetrics.delete(key);else sepActiveMetrics.add(key);sepMetricSelect.value=key;sepDashRender();}
function sepFiltered(){const q=sepSearch.value.toLowerCase().trim();return SEP_RECORDS.filter(r=>r.date>=sepFrom.value&&r.date<=sepTo.value&&r.granularity===sepGran.value&&(!q||[r.tag,r.meter,r.local,r.status].join(' ').toLowerCase().includes(q)));}
function sepSvgLine(rows){sepTrend.innerHTML='';if(!rows.length)return;const metrics=sepMetricDefs.filter(m=>sepActiveMetrics.has(m.key));const by={};rows.forEach(r=>{const k=r.granularity==='Hourly'?`${r.date} ${String(r.hour).padStart(2,'0')}`:r.date;if(!by[k])by[k]={};metrics.forEach(m=>{by[k][m.key]=(by[k][m.key]||0)+(Number(r[m.key])||0);});});const pts=Object.entries(by).sort((a,b)=>a[0].localeCompare(b[0]));const multi=metrics.length>1,width=Math.max(920,pts.length*68),height=390,left=78,right=42,top=48,bottom=76,pw=width-left-right,ph=height-top-bottom,ns='http://www.w3.org/2000/svg';sepTrend.setAttribute('viewBox',`0 0 ${width} ${height}`);sepTrend.style.minWidth=width+'px';const el=(n,a,t)=>{const e=document.createElementNS(ns,n);Object.entries(a||{}).forEach(([k,v])=>e.setAttribute(k,v));if(t!==undefined)e.textContent=t;sepTrend.appendChild(e);return e;};const x=i=>left+i/Math.max(pts.length-1,1)*pw;const series=metrics.map(m=>{const vals=pts.map(p=>Number(p[1][m.key])||0),max=Math.max(...vals.map(v=>Math.abs(v)),0)||1;return {...m,vals,max,plotVals:multi?vals.map(v=>v/max*100):vals};});const allPlot=series.flatMap(s=>s.plotVals),min=Math.min(...allPlot,0),max=Math.max(...allPlot,0),span=max-min||1,y=v=>top+ph-(v-min)/span*ph;for(let i=0;i<=4;i++){const v=min+span*i/4,yy=y(v);el('line',{x1:left,y1:yy,x2:width-right,y2:yy,stroke:'#d8e5e8'});el('text',{x:8,y:yy+4,class:'axis-label'},multi?`${sepDashboardFmt(v,0)}%`:sepDashboardFmt(v,1));}el('text',{x:16,y:24,class:'chart-title'},multi?'Separador — tendências por variável (escala normalizada)':'Separador — '+sepLabels[metrics[0].key]);series.forEach(s=>{el('path',{d:s.plotVals.map((v,i)=>`${i?'L':'M'}${x(i).toFixed(1)} ${y(v).toFixed(1)}`).join(' '),fill:'none',stroke:s.color,'stroke-width':3});s.plotVals.forEach((v,i)=>{const c=el('circle',{cx:x(i),cy:y(v),r:3.8,fill:s.color});c.appendChild(document.createElementNS(ns,'title')).textContent=`${s.label} | ${pts[i][0]} | ${sepDashboardFmt(s.vals[i],s.digits)} ${s.unit}${multi?` | normalizado: ${sepDashboardFmt(v,0)}%`:''}`;});});pts.forEach((p,i)=>{if(i%Math.max(1,Math.floor(pts.length/10))===0)el('text',{x:x(i)-18,y:height-bottom+28,class:'axis-label'},p[0].slice(5));});}
function sepRenderMix(rows){sepMix.innerHTML='';const sum=k=>rows.reduce((a,r)=>a+(Number(r[k])||0),0),vals=[['Óleo',sum('oil'),'#206f77'],['Gás',sum('gas'),'#a15c00'],['Água',sum('water'),'#1d5d8f']],total=vals.reduce((a,x)=>a+x[1],0)||1,width=760,height=230,left=120,right=40,barW=width-left-right,ns='http://www.w3.org/2000/svg';sepMix.setAttribute('viewBox',`0 0 ${width} ${height}`);const el=(n,a,t)=>{const e=document.createElementNS(ns,n);Object.entries(a||{}).forEach(([k,v])=>e.setAttribute(k,v));if(t!==undefined)e.textContent=t;sepMix.appendChild(e);return e;};el('text',{x:12,y:24,class:'chart-title'},'Massa por fase — Separador');vals.forEach((v,i)=>{const yy=55+i*48,pct=v[1]/total;el('text',{x:8,y:yy+16,class:'axis-label'},v[0]);el('rect',{x:left,y:yy,width:barW,height:22,fill:'#e6eef0',rx:3});el('rect',{x:left,y:yy,width:barW*pct,height:22,fill:v[2],rx:3});el('text',{x:left+barW*pct+8,y:yy+16,class:'axis-label'},`${sepDashboardFmt(v[1],2)} t — ${sepDashboardFmt(pct*100,1)}%`);});}
function sepRenderTable(rows){const cols=['Data','Hora','Gran.','TAG','Medidor','Status','Pressão (barg)','Temperatura (°C)','Óleo (t)','Gás (t)','Água (t)','HC (t)','Total (t)','Óleo NSV','Gás padrão','Água NSV'];sepTable.innerHTML='<thead><tr>'+cols.map(c=>`<th>${c}</th>`).join('')+'</tr></thead><tbody>'+rows.slice(0,500).map(r=>`<tr><td>${sepEsc(r.date)}</td><td>${sepEsc(r.hour)}</td><td>${sepEsc(r.granularity)}</td><td>${sepEsc(r.tag)}</td><td>${sepEsc(r.meter)}</td><td>${sepEsc(r.status)}</td><td>${sepDashboardFmt(r.pressure,2)}</td><td>${sepDashboardFmt(r.temperature,2)}</td><td>${sepDashboardFmt(r.oil,3)}</td><td>${sepDashboardFmt(r.gas,3)}</td><td>${sepDashboardFmt(r.water,3)}</td><td>${sepDashboardFmt(r.hc,3)}</td><td>${sepDashboardFmt(r.total,3)}</td><td>${sepDashboardFmt(r.oilVol,2)}</td><td>${sepDashboardFmt(r.gasVol,1)}</td><td>${sepDashboardFmt(r.waterVol,2)}</td></tr>`).join('')+'</tbody>';}
function sepDashRender(){const rows=sepFiltered(),sum=k=>rows.reduce((a,r)=>a+(Number(r[k])||0),0);sepCards.innerHTML=`<div class="metric-card accent"><span>HC do Separador</span><b>${sepDashboardFmt(sum('hc'),1)} t</b><small>óleo + gás</small></div><div class="metric-card accent"><span>Total do Separador</span><b>${sepDashboardFmt(sum('total'),1)} t</b><small>óleo + gás + água</small></div><div class="metric-card"><span>Óleo</span><b>${sepDashboardFmt(sum('oil'),1)} t</b><small>massa de fase</small></div><div class="metric-card"><span>Gás</span><b>${sepDashboardFmt(sum('gas'),1)} t</b><small>massa de fase</small></div><div class="metric-card"><span>Água</span><b>${sepDashboardFmt(sum('water'),1)} t</b><small>massa de fase</small></div><div class="metric-card"><span>Registros</span><b>${rows.length}</b><small>no filtro atual</small></div>`;sepRenderMetricToggles();sepSvgLine(rows);sepRenderMix(rows);sepRenderTable(rows);}
sepFill();[sepFrom,sepTo,sepGran,sepSearch].forEach(e=>e.addEventListener('input',sepDashRender));sepMetricSelect.addEventListener('change',()=>{sepActiveMetrics=new Set([sepMetricSelect.value]);sepDashRender();});sepDashRender();
</script>
""".replace("__SEP_DATA__", data_json).replace("__SEP_DAYS__", days_json).replace("{{", "{").replace("}}", "}")


def _boundary_conditions_panel(comparativo_df: pd.DataFrame | None, target_days: list) -> str:
    """Resume condições de contorno vindas do PI ou da entrada manual do Excel."""
    if comparativo_df is None or comparativo_df.empty:
        return "<p class='muted'>Nenhuma condição de contorno disponível. Os campos serão preenchidos pelo PI ou pela aba COMPARATIVO_MANUAL do Excel.</p>"
    df = comparativo_df.copy().where(pd.notna(comparativo_df), "")
    if "Data" in df.columns and target_days:
        df = df[df["Data"].astype(str).isin({str(d) for d in target_days})].copy()
    if "TAG" in df.columns and not df.empty:
        df = df[df["TAG"].map(lambda tag: _dashboard_point_visible(tag))].copy()
    if df.empty:
        return "<p class='muted'>Não há condições de contorno para a janela publicada.</p>"
    df = df.sort_values(by=[c for c in ["Data", "Banco", "TAG"] if c in df.columns], key=lambda c: c.astype(str))
    source = df.get("Origem", pd.Series([""] * len(df), index=df.index)).astype(str)
    df["Fonte condições"] = np.where(source.str.contains("PI", case=False, na=False), "PI Vision", "Excel / manual")
    columns = [
        "Data", "Banco", "TAG", "Fonte condições", "Pressão MPFM (barg)", "Temperatura MPFM (°C)",
        "Velocidade Escoamento (m/s)", "GVF (%)", "ΔP - Inlet (mbar)", "ΔP - Outlet (mbar)", "WVF (%)", "WLR (%)", "GOR",
        "Water Conductivity (mS/cm)", "Water Conductivity Input (mS/cm)", "Meter Status 1", "Meter Status 2", "Flow Calculation Warn.",
        "Continuous Phase", "Calculation Mode", "Observações",
    ]
    return "<h3>Condições operacionais e status por medidor</h3><div class='info-box'><b>Granularidade:</b> cada linha representa um registro diário por Data, Banco e TAG. As condições podem ser originadas no PI Vision ou preenchidas na aba COMPARATIVO_MANUAL; não representam uma hora específica e campos vazios não devem ser interpretados como zero.</div><div class='table-wrap'>" + _html_table(df.to_dict("records"), columns=columns, max_rows=180) + "</div>"


def _latest_target_day(target_days: list) -> str:
    return max([str(day) for day in target_days], default="")


def _coverage_hours(value) -> float:
    text = str(value or "")
    match = re.search(r"\((\d+)/24h\)", text)
    if match:
        return float(match.group(1))
    if text.upper().startswith("OK"):
        return 24.0
    return np.nan


def _deviation_cards_html(official_rows: list, day: str) -> str:
    """Cards de %desvio do dia por par Subsea × Topside.

    Bloco A: um card por par com os desvios críticos RANP 44 (HC ±10% e Massa
    Total ±7%), com semáforo. Bloco B: card único com os desvios por fase em
    volume padrão (20 °C / 1 atm), que são diagnósticos e não têm limite.
    """
    daily_rows = [
        row for row in (official_rows or [])
        if str(row.get("Dia", ""))[:10] == day
        and str(row.get("Granularidade", "")) == "Daily"
        and "Subsea × Topside" in str(row.get("Métrica", ""))
    ]
    if not daily_rows:
        return "<p class='muted'>Sem par Subsea × Topside com desvio calculável no dia.</p>"
    by_pair: dict = {}
    for row in daily_rows:
        by_pair.setdefault(str(row.get("Banco", "")), {})[str(row.get("MetricaChave", ""))] = row

    def cell(row) -> tuple:
        if row is None:
            return "—", ""
        dev = row.get("DesvioNum")
        limit = row.get("LimiteNum")
        status = str(row.get("Status", ""))
        text = f"{_fmt(dev, 2)}%"
        if limit is None:
            return text, ""
        if status in ("DADO SUSPEITO (revisar PDF fonte)", "REFERÊNCIA PRÓXIMA DE ZERO (transiente)"):
            return text, "warn"
        return text, ("good" if abs(float(dev)) <= float(limit) else "warn")

    critical, phases = [], []
    for pair_name, metrics in by_pair.items():
        lines, worst = [], "good"
        for key, limit_text in (("HC", "±10%"), ("Total", "±7%")):
            text, state = cell(metrics.get(key))
            if state == "warn":
                worst = "warn"
            label = "Desvio HC" if key == "HC" else "Desvio Total"
            lines.append(
                f"<div class='kv kv--{state or 'none'}'><span>{html.escape(label)} · limite {limit_text}</span>"
                f"<b>{html.escape(text)}</b></div>"
            )
        sample = next(iter(metrics.values()))
        critical.append(
            f"<div class='metric-card {worst}'><span>{html.escape(pair_name)}</span>"
            f"<div class='kv-list'>{''.join(lines)}</div>"
            f"<small>referência: {html.escape(str(sample.get('Referência da comparação', '')))} | {html.escape(day)}</small></div>"
        )
        phase_lines = []
        for key, label in (("Óleo vol", "Óleo"), ("Gás vol", "Gás"), ("Água vol", "Água")):
            text, _ = cell(metrics.get(key))
            phase_lines.append(f"<div class='kv'><span>{label}</span><b>{html.escape(text)}</b></div>")
        phases.append(
            f"<div class='metric-card'><span>{html.escape(pair_name)}</span>"
            f"<div class='kv-list'>{''.join(phase_lines)}</div>"
            f"<small>volume padrão 20 °C | 1 atm — diagnóstico, sem limite RANP</small></div>"
        )
    return (
        "<h3>Desvios críticos RANP 44 — Subsea × Topside</h3>"
        f"<div class='metric-grid'>{''.join(critical)}</div>"
        "<h3>Desvios por fase em volume padrão</h3>"
        f"<div class='metric-grid'>{''.join(phases)}</div>"
    )


def _executive_daily_panel(context_df: pd.DataFrame, target_days: list, official_rows: list | None = None) -> str:
    day = _latest_target_day(target_days)
    if not day:
        return "<p class='muted'>Sem dia processado para resumo executivo.</p>"
    df = _normalize_master_columns(context_df)
    for col in ("ProductionDate", "Granularity", "Origin", "Bank", "Tipo", "Tag", "Entity"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    daily = _daily_mpfm(df)
    daily = daily[daily["ProductionDate"] == day].copy()
    if not daily.empty:
        daily = daily[daily.apply(lambda row: _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    recon = df[(df["ProductionDate"] == day) & (df["Granularity"] == "Daily") & (df["Origin"] == "RECON")].copy()
    if daily.empty:
        return "<p class='muted'>Sem linhas Daily/MPFM para resumo executivo.</p>"
    for col in ["MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr Água (t)", "MPFM corr HC (t)", "MPFM corr Total (t)"]:
        daily[col] = pd.to_numeric(daily[col], errors="coerce").fillna(0)
    active = daily[daily["MPFM corr Total (t)"].abs() > 0.001].copy()
    totals = {col: float(active[col].sum()) for col in ["MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr Água (t)", "MPFM corr HC (t)", "MPFM corr Total (t)"]}
    # As massas e volumes absolutos por fase são apresentados nos cards da aba
    # de liderança; aqui o foco é o %desvio que governa a decisão (RANP 44).
    prev_day = (datetime.strptime(day, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    prev = df[(df["ProductionDate"] == prev_day) & (df["Granularity"] == "Daily") & (df["Origin"] == "MPFM")].copy()
    prev_total = pd.to_numeric(prev.get("MPFM corr Total (t)", pd.Series(dtype=float)), errors="coerce").fillna(0).sum() if not prev.empty else np.nan
    trend = np.nan if np.isnan(prev_total) or prev_total == 0 else (totals["MPFM corr Total (t)"] / prev_total - 1) * 100
    if not recon.empty:
        coverage_values = recon["Recon Cobertura"].map(_coverage_hours)
        coverage_pct = float(coverage_values.fillna(0).sum() / (24 * len(recon)) * 100) if len(recon) else np.nan
    else:
        coverage_pct = np.nan
    active_count = len(active)
    total_count = len(daily)
    status_cols = ["Status Gás", "Status Óleo", "Status HC", "Status Água"]
    flags = 0
    if not recon.empty and all(col in recon.columns for col in status_cols):
        flags = int(recon[status_cols].fillna("OK").ne("OK").any(axis=1).sum())
    deviation_flags = sum(1 for item in (official_rows or []) if item.get("metric") in {"HC", "Total"} and item.get("limit") is not None and item.get("deviation") is not None and abs(float(item["deviation"])) > float(item["limit"]) and not re.search(r"SUSPEITO|PRÓXIMA DE ZERO", str(item.get("status", "")), re.I))
    flags += deviation_flags
    status = "NORMAL" if flags == 0 else ("ATENÇÃO" if flags <= 3 else "CRÍTICO")
    status_class = "good" if flags == 0 else "warn"
    trend_text = "—" if np.isnan(trend) else (("+" if trend >= 0 else "") + _fmt(trend, 1) + "%")
    parts = [
        f"<div class='info-box'><b>Mensagem para gestão:</b> em {html.escape(day)}, a produção total corrigida foi de <b>{_fmt(totals['MPFM corr Total (t)'], 1)} t</b>, composta por <b>{_fmt(totals['MPFM corr Óleo (t)'], 1)} t de óleo</b>, <b>{_fmt(totals['MPFM corr Gás (t)'], 1)} t de gás</b> e <b>{_fmt(totals['MPFM corr Água (t)'], 2)} t de água</b>. Status executivo: <b>{status}</b>.</div>",
        "<div class='metric-grid'>",
        f"<div class='metric-card {status_class}'><span>Status executivo</span><b>{status}</b><small>{flags} item(ns) de reconciliação/desvio requer(em) atenção</small></div>",
        f"<div class='metric-card accent'><span>Produção do dia</span><div class='kv-list'><div class='kv'><span>HC corrigido</span><b>{_fmt(totals['MPFM corr HC (t)'], 1)} t</b></div><div class='kv'><span>Total corrigido</span><b>{_fmt(totals['MPFM corr Total (t)'], 1)} t</b></div></div><small>D-1 (total): {trend_text}</small></div>",
        "</div>",
        _deviation_cards_html(official_rows or [], day),
    ]
    rows = []
    for _, row in active.sort_values(["Tipo", "Bank", "Entity"]).iterrows():
        rows.append({
            "Local": row.get("Tipo", ""), "Banco": row.get("Bank", ""), "TAG": row.get("Tag", ""),
            "Volume óleo 20 °C / 1 atm (m³)": _fmt(row.get("PVT @20 vol Óleo (m³)", ""), 1),
            "Volume gás 20 °C / 1 atm (Sm³)": _fmt(row.get("PVT @20 vol Gás (Sm³)", ""), 0),
            "Volume água 20 °C / 1 atm (m³)": _fmt(row.get("PVT @20 vol Água (m³)", ""), 3),
            "Massa HC corrigida (t)": _fmt(row.get("MPFM corr HC (t)", ""), 2),
            "Massa total corrigida (t)": _fmt(row.get("MPFM corr Total (t)", ""), 2),
        })
    parts.append("<h3>Dados para reporte XML 042 por medidor</h3><div class='info-box'>Volumes padrão a 20 °C e 1 atm — campos de conferência para o XML. As massas HC e Total permanecem como validação.</div><div class='table-wrap'>")
    parts.append(_html_table(rows, max_rows=80))
    parts.append("</div>")
    return "".join(parts)


def _hourly_profile_svg(context_df: pd.DataFrame, target_days: list) -> str:
    day = _latest_target_day(target_days)
    if not day:
        return "<p class='muted'>Sem dia processado para perfil horário.</p>"
    df = _normalize_master_columns(context_df)
    for col in ("ProductionDate", "Granularity", "Origin", "Hour"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    hourly = df[(df["ProductionDate"] == day) & (df["Granularity"] == "Hourly") & (df["Origin"] == "MPFM")].copy()
    if not hourly.empty and "Instrumento" in hourly.columns:
        hourly = hourly[hourly.apply(lambda row: _mpfm_extraction_enabled(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    if not hourly.empty:
        hourly = hourly[hourly.apply(lambda row: _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")), axis=1)].copy()
    if hourly.empty:
        return "<p class='muted'>Sem linhas Hourly/MPFM para o dia mais recente.</p>"
    hourly["HourNum"] = pd.to_numeric(hourly["Hour"], errors="coerce")
    for col in ["MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr Água (t)"]:
        hourly[col] = pd.to_numeric(hourly[col], errors="coerce").fillna(0)
    grouped = hourly.groupby("HourNum", dropna=True)[["MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr Água (t)"]].sum().reset_index().sort_values("HourNum")
    values = []
    for _, row in grouped.iterrows():
        values.append((int(row["HourNum"]), float(row["MPFM corr Gás (t)"] + row["MPFM corr Óleo (t)"] + row["MPFM corr Água (t)"])))
    if not values:
        return "<p class='muted'>Sem valores horários numéricos.</p>"
    point_rows = []
    grouped_points = hourly.groupby(["Bank", "Tag", "HourNum"], dropna=True)[["MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr Água (t)", "MPFM corr HC (t)", "MPFM corr Total (t)"]].sum().reset_index()
    for _, row in grouped_points.iterrows():
        point_rows.append({"point": f"{row['Bank']} — {row['Tag']}", "hour": int(row["HourNum"]), "oil": float(row["MPFM corr Óleo (t)"]), "gas": float(row["MPFM corr Gás (t)"]), "water": float(row["MPFM corr Água (t)"]), "hc": float(row["MPFM corr HC (t)"]), "total": float(row["MPFM corr Total (t)"])})
    point_json = json.dumps(point_rows, ensure_ascii=False, allow_nan=False)
    width, height, pad = 980, 260, 46
    max_v = max(v for _, v in values) or 1.0
    points = []
    for idx, (hour, value) in enumerate(values):
        x = pad + (width - 2 * pad) * (idx / max(len(values) - 1, 1))
        y = height - pad - (height - 2 * pad) * (value / max_v)
        points.append((x, y, hour, value))
    svg_points = " ".join(f"{x:.1f},{y:.1f}" for x, y, _, _ in points)
    points_options = "".join(f"<option>{html.escape(item)}</option>" for item in sorted({r['point'] for r in point_rows}))
    parts = [f"<div class='info-box'><b>Perfil horário de 24 horas:</b> selecione MPFM e fase para evitar misturar Topside e Subsea. Dia {html.escape(day)}.</div><div class='controls'><label>MPFM <select id='hourlyPoint'>{points_options}</select></label><label>Fase <select id='hourlyPhase'><option value='total'>Total</option><option value='hc'>HC</option><option value='oil'>Óleo</option><option value='gas'>Gás</option><option value='water'>Água</option></select></label></div>"]
    parts.append(f"<div class='chart-scroll'><svg id='hourlyProfileChart' viewBox='0 0 {width} {height}' class='chart wide-chart' width='{width}' role='img' aria-label='Perfil horário consolidado'>")
    parts.append(f"<text x='18' y='26' class='chart-title'>Massa total corrigida por hora (t)</text>")
    parts.append(f"<line x1='{pad}' y1='{height-pad}' x2='{width-pad}' y2='{height-pad}' stroke='#aab5c4'></line><line x1='{pad}' y1='{pad}' x2='{pad}' y2='{height-pad}' stroke='#aab5c4'></line>")
    parts.append(f"<polyline fill='none' stroke='#007398' stroke-width='4' points='{svg_points}'></polyline>")
    for x, y, hour, value in points:
        parts.append(f"<circle cx='{x:.1f}' cy='{y:.1f}' r='3.5' fill='#007398'><title>{hour:02d}:00 — {_fmt(value, 1)} t</title></circle>")
    parts.append(f"</svg></div><script>const HP_DATA={point_json};const hpPoint=document.getElementById('hourlyPoint'),hpPhase=document.getElementById('hourlyPhase'),hpSvg=document.getElementById('hourlyProfileChart');function hpRender(){{const rows=HP_DATA.filter(r=>r.point===hpPoint.value),key=hpPhase.value;const by=Object.fromEntries(rows.map(r=>[r.hour,r[key]]));const vals=Array.from({{length:24}},(_,i)=>Number(by[i]??by[i+1]??0));const max=Math.max(...vals,1),W={width},H={height},P={pad},pts=vals.map((v,i)=>`${{P+(W-2*P)*i/23}},${{H-P-(H-2*P)*v/max}}`).join(' ');const old=hpSvg.querySelector('polyline');if(old)old.setAttribute('points',pts);hpSvg.querySelector('.chart-title').textContent=hpPoint.value+' — '+key+' — 24 horas';}}[hpPoint,hpPhase].forEach(e=>e.addEventListener('change',hpRender));hpRender();</script>")
    return "".join(parts)


def _comparison_delta_visual(comparisons: list) -> str:
    rows = []
    for row in comparisons:
        delta = _num(str(row.get("Δ %", "")).replace(".", "").replace(",", "."))
        if np.isnan(delta):
            continue
        arrow = "▲" if delta > 0 else "▼" if delta < 0 else "→"
        klass = "up" if delta > 0 else "down" if delta < 0 else "flat"
        rows.append({
            "Dia": row.get("Dia", ""), "Banco": row.get("Banco", ""), "Tag": row.get("Tag", ""),
            "Campo": row.get("Campo", ""), "Variação": f"<span class='delta {klass}'>{arrow} {_fmt(delta, 1)}%</span>",
            "Status": row.get("Status", ""),
        })
    if not rows:
        return "<p class='muted'>Sem comparação suficiente contra os 4 dias anteriores.</p>"
    return pd.DataFrame(rows[:120]).to_html(index=False, escape=False, classes="data-table")


def _dashboard_timeseries_records(context_df: pd.DataFrame, target_days: list) -> list:
    df = _normalize_master_columns(context_df)
    if df.empty:
        return []
    for col in ("ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Bank", "Tag"):
        df[col] = df[col].where(pd.notna(df[col]), "").astype(str)
    df = df[
        df["ProductionDate"].isin({str(day) for day in target_days})
        & (df["Origin"] == "MPFM")
        & (df["SourceType"] == "PDF")
        & (df["Bank"].isin(COMPARISON_BANKS))
        & (df["Granularity"].isin(["Daily", "Hourly"]))
    ].copy()
    records = []
    for _, row in df.iterrows():
        if not _row_has_production(row):
            continue
        if not _dashboard_point_visible(row.get("Tag"), row.get("Instrumento")):
            continue
        values = {}
        for key, spec in DASHBOARD_VARIABLES.items():
            value = _num(row.get(spec["mpfm"], ""))
            if not np.isnan(value):
                values[key] = value
        if values:
            bank = str(row.get("Bank", ""))
            info = _comparison_info_for_row(row)
            records.append({
                "date": str(row.get("ProductionDate", "")),
                "hour": str(row.get("Hour", "")) if str(row.get("Granularity", "")) == "Hourly" else "",
                "granularity": str(row.get("Granularity", "")),
                "bank": bank,
                "label": info.get("side", bank),
                "values": values,
            })
    return records


def _interactive_stability_panel(records: list, target_days: list) -> str:
    if not records:
        return "<p class='muted'>Sem série horária/diária para estabilidade.</p>"
    records_json = json.dumps(records, ensure_ascii=False)
    variables_json = json.dumps({k: {"label": v["label"], "aggregation": v["aggregation"]} for k, v in DASHBOARD_VARIABLES.items()}, ensure_ascii=False)
    min_day = str(min(target_days)) if target_days else ""
    max_day = str(max(target_days)) if target_days else ""
    return f"""
<div class="controls">
  <label>MPFM <select id="stabBank"></select></label>
  <label>Variável <select id="stabVariable"></select></label>
  <label>Granularidade <select id="stabGranularity"><option value="Hourly">Horária (janela 24 pontos)</option><option value="Daily">Diária (janela 7 pontos)</option></select></label>
  <label>Data inicial <select id="stabDateFrom"></select></label>
  <label>Data final <select id="stabDateTo"></select></label>
</div>
<p class="muted">A linha azul é a medição. A linha verde é a média móvel anterior. Linhas vermelhas mostram limites estatísticos ±3σ calculados sem incluir o ponto atual.</p>
<div class="legend-row"><span class="legend-item"><i class="legend-swatch legend-swatch--blue"></i>Medição</span><span class="legend-item"><i class="legend-swatch legend-swatch--green"></i>Média móvel anterior</span><span class="legend-item"><i class="legend-swatch legend-swatch--red"></i>Limites ±3σ</span><span class="legend-item"><i class="legend-swatch legend-swatch--dot legend-swatch--red"></i>Ponto suspeito</span></div>
<div id="stabEmpty" class="muted"></div>
<div class="chart-scroll"><svg id="stabChart" class="chart wide-chart" role="img" aria-label="Estabilidade e spikes"></svg></div>
<script>
const STAB_RECORDS = {records_json};
const STAB_VARIABLES = {variables_json};
const stabBank = document.getElementById('stabBank'), stabVariable = document.getElementById('stabVariable'), stabGranularity = document.getElementById('stabGranularity');
const stabDateFrom = document.getElementById('stabDateFrom'), stabDateTo = document.getElementById('stabDateTo'), stabChart = document.getElementById('stabChart'), stabEmpty = document.getElementById('stabEmpty');
function sUniq(values) {{ return [...new Set(values)].filter(Boolean).sort(); }}
function sFmt(v) {{ return Number.isFinite(v) ? v.toLocaleString('pt-BR', {{maximumFractionDigits: 2}}) : ''; }}
function sDateBR(d) {{ const p=String(d||'').slice(0,10).split('-'); return p.length===3 ? `${{p[2]}}/${{p[1]}}/${{p[0]}}` : String(d||''); }}
function sCat(r) {{ return r.granularity === 'Hourly' ? `${{r.date}} H${{String(r.hour).padStart(2,'0')}}` : r.date; }}
function sFill() {{
  sUniq(STAB_RECORDS.map(r => `${{r.bank}} — ${{r.label}}`)).forEach(x => stabBank.add(new Option(x, x.split(' — ')[0])));
  Object.entries(STAB_VARIABLES).forEach(([k,v]) => stabVariable.add(new Option(v.label, k)));
  sUniq(STAB_RECORDS.map(r => r.date)).forEach(d => {{ stabDateFrom.add(new Option(sDateBR(d), d)); stabDateTo.add(new Option(sDateBR(d), d)); }});
  if (stabDateFrom.options.length) {{ stabDateFrom.value='{html.escape(min_day)}'; stabDateTo.value='{html.escape(max_day)}'; }}
  stabVariable.value='mpfm_corr_total';
}}
function rolling(vals, n) {{ return vals.map((value, i) => {{ const prior = vals.slice(Math.max(0, i-n), i).filter(Number.isFinite); const w = prior.length ? prior : [value]; const m = w.reduce((a,b)=>a+b,0)/Math.max(w.length,1); const sd = Math.sqrt(w.reduce((a,b)=>a+(b-m)*(b-m),0)/Math.max(w.length,1)); return {{mean:m, ucl:m+3*sd, lcl:m-3*sd}}; }}); }}
function sRender() {{
  stabChart.innerHTML='';
  const rows = STAB_RECORDS.filter(r => r.bank===stabBank.value && r.granularity===stabGranularity.value && r.date >= (stabDateFrom.value||'0000-00-00') && r.date <= (stabDateTo.value||'9999-99-99'))
    .map(r => ({{cat:sCat(r), value:Number(r.values?.[stabVariable.value])}})).filter(r => Number.isFinite(r.value));
  if (!rows.length) {{ stabEmpty.textContent='Sem dados para os filtros selecionados.'; return; }}
  stabEmpty.textContent='';
  const n = stabGranularity.value === 'Hourly' ? 24 : 7;
  const bands = rolling(rows.map(r=>r.value), n);
  const width=Math.max(1080, rows.length*58), height=430, left=88, top=44, bottom=110, right=40, plotW=width-left-right, plotH=height-top-bottom;
  stabChart.setAttribute('viewBox', `0 0 ${{width}} ${{height}}`); stabChart.style.minWidth=width+'px';
  const allY = rows.flatMap((r,i)=>[r.value,bands[i].mean,bands[i].ucl,bands[i].lcl]).filter(Number.isFinite); const minY=Math.min(...allY), maxY=Math.max(...allY); const pad=Math.max((maxY-minY)*.12,1); const yMin=minY-pad, yMax=maxY+pad;
  const ns='http://www.w3.org/2000/svg'; function el(n,a,t){{const e=document.createElementNS(ns,n); Object.entries(a||{{}}).forEach(([k,v])=>e.setAttribute(k,v)); if(t!==undefined)e.textContent=t; stabChart.appendChild(e); return e;}}
  const x=i=>left+i/Math.max(rows.length-1,1)*plotW, y=v=>top+plotH-(v-yMin)/(yMax-yMin)*plotH;
  el('text',{{x:20,y:28,class:'chart-title'}},`${{stabBank.options[stabBank.selectedIndex]?.text || stabBank.value}} — ${{STAB_VARIABLES[stabVariable.value].label}}`);
  for(let i=0;i<=4;i++){{const yy=top+plotH-i/4*plotH; el('line',{{x1:left,y1:yy,x2:width-right,y2:yy,stroke:'#e2e8f0'}}); el('text',{{x:12,y:yy+4,class:'axis-label'}},sFmt(yMin+(yMax-yMin)*i/4));}}
  function path(vals){{return vals.map((v,i)=>`${{i?'L':'M'}}${{x(i).toFixed(1)}} ${{y(v).toFixed(1)}}`).join(' ');}}
  el('path',{{d:path(rows.map(r=>r.value)),fill:'none',stroke:'#007398','stroke-width':'3'}}); el('path',{{d:path(bands.map(b=>b.mean)),fill:'none',stroke:'#16a34a','stroke-width':'2','stroke-dasharray':'6 4'}}); el('path',{{d:path(bands.map(b=>b.ucl)),fill:'none',stroke:'#dc2626','stroke-width':'2','stroke-dasharray':'6 4'}}); el('path',{{d:path(bands.map(b=>b.lcl)),fill:'none',stroke:'#dc2626','stroke-width':'2','stroke-dasharray':'6 4'}});
  rows.forEach((r,i)=>{{const spike=r.value>bands[i].ucl||r.value<bands[i].lcl; const c=el('circle',{{cx:x(i),cy:y(r.value),r:spike?5:3.5,fill:spike?'#dc2626':'#007398'}}); c.appendChild(document.createElementNS(ns,'title')).textContent=`${{r.cat}}: ${{sFmt(r.value)}}${{spike?' | SUSPEITO':''}}`; const label=r.cat.replace('2026-',''); el('text',{{x:x(i),y:height-64,class:'axis-label','text-anchor':'end',transform:`rotate(-45 ${{x(i)}} ${{height-64}})`}},label);}});
}}
sFill(); [stabBank,stabVariable,stabGranularity,stabDateFrom,stabDateTo].forEach(e=>e.addEventListener('change',sRender)); sRender();
</script>
"""


def _merge_dashboard_alarm_events(
    master_path: Path | None,
    current_events: pd.DataFrame | None,
    selected_days: list,
) -> pd.DataFrame:
    """Usa o histórico master e a execução atual para a janela publicada."""
    pieces = []
    master_events = read_master_alarm_events(master_path)
    if not master_events.empty:
        pieces.append(master_events)
    if current_events is not None and not current_events.empty:
        pieces.append(_normalize_alarm_event_columns(current_events))
    if not pieces:
        return pd.DataFrame(columns=ALARM_EVENT_COLUMNS)
    events = pd.concat(pieces, ignore_index=True)
    events = events.where(pd.notna(events), "")
    events = _normalize_alarm_event_columns(events)
    if selected_days:
        events = events[events["ProductionDate"].astype(str).isin({str(day) for day in selected_days})].copy()
    dedup_cols = ["ProductionDate", "Timestamp", "RecordType", "SourceKind", "Object", "Description", "SourceFile"]
    events.drop_duplicates(subset=dedup_cols, keep="last", inplace=True)
    return events


def _alarm_explanation(issue_flag: str, record_type: str = "ALARM") -> str:
    explanations = {
        "CRITICAL": "Prioridade crítica: avaliar impacto operacional e ação imediata.",
        "COMM": "Comunicação: verificar rede, controlador, instrumento e atualização da variável.",
        "FAILED": "Estado Failed: a função ou cálculo não está aprovado pelo sistema.",
        "PVT": "PVT: verificar entradas, validade do modelo, composição e cálculo termodinâmico.",
        "MISSING_FILE": "Arquivo esperado não foi localizado na pasta de origem.",
        "READ_ERROR": "Arquivo localizado, mas não pôde ser interpretado pela rotina.",
        "EVENT": "Evento operacional: cruzar horário com modo, alinhamento, produção e alarmes.",
    }
    tokens = [token.strip().upper() for token in str(issue_flag or "").split(",") if token.strip()]
    notes = [explanations[token] for token in tokens if token in explanations]
    if str(record_type).upper() == "EVENT" and not any("Evento" in note for note in notes):
        notes.insert(0, explanations["EVENT"])
    return " ".join(dict.fromkeys(notes)) or "Revisar descrição, estado, horário e impacto no ponto de medição."


def _pi_quality_panel(master_path: Path | None, target_days: list, pi_df: pd.DataFrame | None = None) -> str:
    """Apresenta o que o PI entregou, o que faltou e a rastreabilidade da coleta."""
    pi = pi_df if pi_df is not None else read_master_pi_extract(master_path)
    if pi.empty:
        return "<div class='info-box'><b>PI sem evidência:</b> a coleta não entregou linhas para a janela. Verifique login, URL, aba do Edge, período e o diagnóstico salvo pelo coletor.</div>"
    day_col = "PI Dia Coleta" if "PI Dia Coleta" in pi.columns else "PI Inicio"
    pi = pi[pi[day_col].map(_pi_day_iso).isin({str(day) for day in target_days})].copy()
    if pi.empty:
        missing = ", ".join(str(day) for day in target_days) or "janela selecionada"
        return f"<div class='info-box'><b>PI sem evidência na janela publicada:</b> nenhum registro foi localizado para {html.escape(missing)}.</div>"
    status_col = "PI Status Coleta" if "PI Status Coleta" in pi.columns else None
    status_counts = pi[status_col].astype(str).value_counts().to_dict() if status_col else {"OK": len(pi)}
    failed_mask = pi[status_col].astype(str).str.upper().isin({"OK", "SUCESSO", "APROVADO"}) == False if status_col else pd.Series(False, index=pi.index)
    error_mask = failed_mask.copy()
    if "PI Erro" in pi.columns:
        error_mask |= pi["PI Erro"].astype(str).str.strip().ne("")
    good_quality = pi["Qualidade"].astype(str).str.upper().isin({"GOOD", "OK", "VALID", "VALIDO", "VÁLIDO"}).sum() if "Qualidade" in pi.columns else 0
    days_captured = sorted({str(value) for value in pi[day_col].map(_pi_day_iso) if str(value)})
    missing_days = [str(day) for day in target_days if str(day) not in set(days_captured)]
    meters = sorted({str(value) for value in pi.get("Medidor", pd.Series(dtype=str)).astype(str) if str(value).strip() and str(value).lower() != "nan"})
    variables = sorted({str(value) for value in pi.get("Variavel", pd.Series(dtype=str)).astype(str) if str(value).strip() and str(value).lower() != "nan"})
    urls = sorted({str(value) for value in pi.get("PI URL Aplicada", pd.Series(dtype=str)).astype(str) if str(value).startswith("http")})
    status_rows = [{"Status PI": status, "Linhas": count} for status, count in status_counts.items()]
    missing_rows = [{"Dia esperado": day, "Status": "NÃO RECEBIDO"} for day in missing_days]
    error_rows = pi.loc[error_mask].copy()
    error_columns = [c for c in ["PI Dia Coleta", "PI Status Coleta", "PI Erro", "PI Diagnóstico", "PI Arquivo Origem", "PI URL Aplicada"] if c in error_rows.columns]
    klass = "warn" if int(error_mask.sum()) or missing_days else "good"
    source_rows = [{"Item": "Medidores", "Valor": ", ".join(meters) or "Não informado"}, {"Item": "Variáveis recebidas", "Valor": ", ".join(variables) or "Não informado"}, {"Item": "Dias recebidos", "Valor": ", ".join(days_captured) or "Nenhum"}, {"Item": "URL principal aplicada", "Valor": urls[0] if urls else "Não informado"}]
    return (
        "<div class='metric-grid'>"
        f"<div class='metric-card {klass}'><span>Status da coleta PI</span><b>{'ATENÇÃO' if klass == 'warn' else 'OK'}</b><small>{int(error_mask.sum())} linha(s) com falha/erro; {len(missing_days)} dia(s) sem recebimento</small></div>"
        f"<div class='metric-card accent'><span>Linhas recebidas</span><b>{len(pi)}</b><small>evidência PI na janela</small></div>"
        f"<div class='metric-card'><span>Qualidade GOOD/OK</span><b>{int(good_quality)}</b><small>linhas com qualidade aprovada</small></div>"
        f"<div class='metric-card'><span>Medidores identificados</span><b>{len(meters)}</b><small>{html.escape(', '.join(meters) or 'não informado')}</small></div>"
        "</div>"
        "<h3>Resumo da extração</h3><div class='table-wrap'>" + _html_table(status_rows, max_rows=30) + "</div>"
        "<h3>Rastreabilidade da evidência</h3><div class='table-wrap'>" + _html_table(source_rows, max_rows=20) + "</div>"
        + ("<h3>Dias esperados sem recebimento</h3><div class='table-wrap'>" + _html_table(missing_rows, max_rows=80) + "</div>" if missing_rows else "")
        + ("<h3>Falhas e diagnósticos do PI</h3><div class='table-wrap'>" + _html_table(error_rows.to_dict("records"), columns=error_columns, max_rows=100) + "</div>" if not error_rows.empty else "<div class='info-box'>Nenhuma falha de coleta foi identificada nesta janela.</div>")
    )


def _xml042_panel(master_path: Path | None = None, target_days: list | None = None, context_df: pd.DataFrame | None = None) -> str:
    """Mostra o pacote standalone de XML 042 (ANP); geração continua manual/sob demanda, fora desta automação."""
    package_dir = Path(__file__).resolve().parent / XML042_AUTOMATION_DIR
    config_path = package_dir / XML042_CONFIG_FILE
    script_path = package_dir / XML042_SCRIPT_FILE
    available = script_path.exists()
    catalog_rows = []
    pending_rows = []
    cnpj8 = ""
    if config_path.exists():
        try:
            config = json.loads(config_path.read_text("utf-8"))
        except Exception:
            config = {}
        cnpj8 = str(config.get("cnpj8", "") or "")
        for item in config.get("catalog", []) or []:
            if not item.get("active", True) or not item.get("enabled_042", True):
                if item.get("active", True) and not item.get("enabled_042", True):
                    pending_rows.append({
                        "Poço operacional": item.get("well_operator_name", ""),
                        "Poço ANP": item.get("well_anp_name", ""),
                        "TAG subsea": item.get("subsea_tag", ""),
                        "Status": item.get("status", "Pendente de autorização ANP"),
                    })
                continue
            catalog_rows.append({
                "Poço operacional": item.get("well_operator_name", ""),
                "Poço ANP": item.get("well_anp_name", ""),
                "Código cadastro": item.get("cod_cadastro_poco", ""),
                "TAG subsea": item.get("subsea_tag", ""),
                "Banco": item.get("bank", ""),
            })
    preview_rows = []
    preview_xml = {}
    preview_filenames = {}
    preview_keys = set()
    preview_rejected = []
    if available and master_path and master_path.exists() and target_days:
        try:
            spec = importlib.util.spec_from_file_location("xml042_rules_preview", script_path)
            xml_rules = importlib.util.module_from_spec(spec)
            sys.modules[spec.name] = xml_rules
            spec.loader.exec_module(xml_rules)
            base_df, _ = xml_rules.load_base_unica(master_path)
            config_full = xml_rules.load_config(config_path)
            date_values = [str(day)[:10] for day in target_days]
            date_from, date_to = min(date_values), max(date_values)
            filtered = xml_rules.filter_daily_mpfm_subsea(base_df, date_from, date_to, config_full)
            candidates, preview_rejected = xml_rules.candidates_from_base(filtered, config_full.get("catalog", []), config_full)
            for candidate in candidates:
                key = f"{candidate.production_day}|{candidate.catalog.get('cod_cadastro_poco','')}"
                xml_text = xml_rules.build_xml042_text(candidate)
                content_hash = hashlib.sha256(xml_text.encode("iso-8859-1", errors="xmlcharrefreplace")).hexdigest()
                if content_hash in preview_keys:
                    preview_rejected.append({"day": candidate.production_day, "code": candidate.catalog.get("cod_cadastro_poco", ""), "reason": "conteúdo idêntico já representado; duplicidade bloqueada"})
                    continue
                preview_keys.add(content_hash)
                preview_xml[key] = xml_text
                preview_filenames[key] = xml_rules.build_anp_filename(cnpj8, package_dir / "xml042_gerados", set(preview_filenames.values()))
                preview_rows.append({"Data": candidate.production_day, "Poço": candidate.catalog.get("well_operator_name", ""), "TAG": candidate.subsea_tag, "Código": candidate.catalog.get("cod_cadastro_poco", ""), "Óleo 20 °C (m³)": candidate.oil_sm3, "Gás 20 °C (Sm³)": candidate.gas_sm3, "Água 20 °C (m³)": candidate.water_sm3})
        except Exception as exc:
            preview_rejected.append({"reason": f"Falha ao montar prévia: {exc}"})
    status_klass = "good" if available else "warn"
    status_text = "Disponível" if available else "Pacote não encontrado"
    info = (
        "<div class='info-box'><b>Geração sob demanda:</b> o XML 042 (ANP) é produzido por uma automação "
        "separada e standalone, executada manualmente pelo usuário quando necessário — ela NÃO roda "
        "automaticamente junto com esta automação Base_Unica. A entrada é o próprio "
        "BASE_UNICA_TOTAL.xlsx já gerado.</div>"
        f"<div class='metric-grid'><div class='metric-card {status_klass}'><span>Pacote XML 042</span>"
        f"<b>{html.escape(status_text)}</b><small>{html.escape(str(package_dir))}</small></div>"
        f"<div class='metric-card'><span>CNPJ raiz</span><b>{html.escape(cnpj8 or '—')}</b>"
        "<small>usado no nome do arquivo ANP</small></div>"
        f"<div class='metric-card'><span>Poços habilitados</span><b>{len(catalog_rows)}</b>"
        "<small>autorizados/cadastrados para XML042</small></div>"
        f"<div class='metric-card warn'><span>Poços pendentes</span><b>{len(pending_rows)}</b>"
        "<small>aguardando autorização ANP</small></div></div>"
    )
    if available:
        example_cmd = f'python "{script_path}" --base-unica "BASE_UNICA_TOTAL.xlsx" --date-from 01/08/2026 --date-to 03/08/2026'
        how_to_body = (
            f"1. Abra a pasta <code>{html.escape(XML042_AUTOMATION_DIR)}</code>.<br>"
            f"2. Dê duplo clique em <code>{html.escape(XML042_BAT_FILE)}</code> e informe o Excel Base_Unica e a janela de "
            "datas quando solicitado; ou execute por linha de comando:<br>"
            f"<code>{html.escape(example_cmd)}</code>"
        )
    else:
        how_to_body = f"Pasta <code>{html.escape(XML042_AUTOMATION_DIR)}</code> não encontrada ao lado do script principal."
    how_to = "<h3>Como gerar (sob demanda)</h3><div class='info-box'>" + how_to_body + "</div>"
    catalog_table = (
        "<h3>Poços habilitados para XML 042</h3><div class='table-wrap'>" + _html_table(catalog_rows, max_rows=50) + "</div>"
        if catalog_rows else "<div class='info-box'>Nenhum poço ativo cadastrado no catálogo XML042.</div>"
    )
    pending_table = (
        "<h3>Poços novos — pendentes de autorização ANP</h3><div class='table-wrap'>" + _html_table(pending_rows, max_rows=50) + "</div>"
        if pending_rows else ""
    )
    preview_json = json.dumps(preview_xml, ensure_ascii=False).replace("</", "<\\/")
    preview_labels = {f"{row.get('Data')}|{row.get('Código')}": f"{row.get('Data')} — {row.get('Poço')}" for row in preview_rows}
    preview_panel = ""
    if available:
        preview_panel = (
            "<h3>Prévia do XML 042 — volumes padrão Subsea</h3>"
            "<div class='info-box'><b>Fonte:</b> somente linhas Daily, MPFM, Subsea e volumes PVT a 20 °C / 1 atm. Esta prévia aplica as mesmas regras do pacote XML042 standalone; linhas rejeitadas não são disponibilizadas para download.</div>"
            "<div class='controls'><label>XML para visualizar/download<select id='xml042Select'>" + "".join(f"<option value='{html.escape(key)}'>{html.escape(preview_labels.get(key, key))}</option>" for key in preview_xml) + "</select></label><button type='button' class='tab-button' id='xml042Download'>Baixar XML selecionado</button></div>"
            "<div class='table-wrap'>" + _html_table(preview_rows, max_rows=100) + "</div>"
            "<div class='info-box'><b>Legenda — código do poço:</b> " + " | ".join(f"{html.escape(str(row.get('Poço','')))} — {html.escape(str(row.get('Código','')))}" for row in preview_rows) + "</div>"
            "<pre id='xml042Preview' style='max-height:420px;overflow:auto;background:#0b1f33;color:#dbeafe;padding:14px;border-radius:8px;white-space:pre-wrap'></pre>"
            f"<script>\n(() => {{ const data={preview_json}; const names={json.dumps(preview_filenames, ensure_ascii=False)}; const sel=document.getElementById('xml042Select'), pre=document.getElementById('xml042Preview'), btn=document.getElementById('xml042Download'); function render(){{pre.textContent=data[sel.value]||'Nenhum XML elegível na janela.';}} sel.addEventListener('change',render); btn.addEventListener('click',()=>{{const blob=new Blob([data[sel.value]||''],{{type:'application/xml;charset=utf-8'}});const a=document.createElement('a');a.href=URL.createObjectURL(blob);a.download=names[sel.value]||'042_{html.escape(cnpj8)}_preview.xml';a.click();setTimeout(()=>URL.revokeObjectURL(a.href),1000);}}); render(); }})();</script>"
            + ("<div class='table-wrap'><h4>Linhas rejeitadas pelas regras XML042</h4>" + _html_table(preview_rejected, max_rows=100) + "</div>" if preview_rejected else "")
        )
    return info + how_to + preview_panel + catalog_table + pending_table


def _load_dashboard_source_frames(workbook_path: Path | None) -> dict[str, pd.DataFrame]:
    """Lê abas sequencialmente e audita todas sem manter tudo simultaneamente em RAM."""
    if not workbook_path or not workbook_path.exists():
        return {}
    core_sheets = {
        MASTER_SHEET_NAME, "BASE_UNICA_STANDALONE",
        MPFM_MEASUREMENTS_SHEET_NAME, SEP_OIL_SHEET_NAME, SEP_GAS_SHEET_NAME,
        SEP_WATER_SHEET_NAME, RECONCILIATION_SHEET_NAME,
        PI_MASTER_SHEET_NAME, PI_SHEET_NAME,
        ALARM_EVENT_MASTER_SHEET_NAME, ALARM_EVENT_SHEET_NAME,
        COMPARATIVO_TOTAL_SHEET_NAME,
    }
    result = {}
    with pd.ExcelFile(workbook_path, engine="openpyxl") as workbook:
        for name in workbook.sheet_names:
            frame = pd.read_excel(workbook, sheet_name=name, dtype=object).where(lambda value: pd.notna(value), "")
            fingerprint = _source_frame_fingerprint(frame)
            rows = int(len(frame))
            columns = int(len(frame.columns))
            if name in core_sheets:
                stored = frame
            else:
                stored = pd.DataFrame(columns=frame.columns)
            stored.attrs["source_rows"] = rows
            stored.attrs["source_columns"] = columns
            stored.attrs["source_fingerprint"] = fingerprint
            result[name] = stored
            if stored is not frame:
                del frame
            gc.collect()
    return result


def _source_frame_fingerprint(frame: pd.DataFrame) -> str:
    """Impressão digital determinística de colunas, ordem e valores da aba."""
    cached = frame.attrs.get("source_fingerprint")
    if cached:
        return str(cached)
    digest = hashlib.sha256()
    digest.update(json.dumps([str(column) for column in frame.columns], ensure_ascii=False).encode("utf-8"))
    if not frame.empty:
        normalized = frame.where(pd.notna(frame), "").astype(str)
        digest.update(pd.util.hash_pandas_object(normalized, index=True).values.tobytes())
    return digest.hexdigest()


def _source_frame_rows(frame: pd.DataFrame) -> int:
    return int(frame.attrs.get("source_rows", len(frame)))


def _source_frame_columns(frame: pd.DataFrame) -> int:
    return int(frame.attrs.get("source_columns", len(frame.columns)))


def _dashboard_source_manifest(source_frames: dict[str, pd.DataFrame], workbook_path: Path | None) -> dict:
    return {
        "workbook": str(workbook_path.resolve()) if workbook_path else "",
        "locale": "pt-BR",
        "dateTimeFormat": "dd/mm/yyyy hh:mm",
        "decimalSeparator": ",",
        "sheets": [
            {
                "name": name,
                "rows": _source_frame_rows(frame),
                "columns": _source_frame_columns(frame),
                "fingerprint": _source_frame_fingerprint(frame),
            }
            for name, frame in source_frames.items()
        ],
    }


def _sheet_purpose(sheet_name: str) -> tuple[str, str]:
    descriptions = {
        MASTER_SHEET_NAME: ("Base operacional", "Executivo, Separador, Cobertura e Detalhes"),
        "BASE_UNICA_STANDALONE": ("Base operacional da janela", "Executivo, Separador, Cobertura e Detalhes"),
        MPFM_MEASUREMENTS_SHEET_NAME: ("Medições MPFM normalizadas", "Visão integrada, CEP, XML042 e detalhes"),
        SEP_OIL_SHEET_NAME: ("C.V. — fase óleo", "Consolidação dinâmica do Separador"),
        SEP_GAS_SHEET_NAME: ("C.V. — fase gás", "Consolidação dinâmica do Separador"),
        SEP_WATER_SHEET_NAME: ("C.V. — fase água", "Consolidação dinâmica do Separador"),
        RECONCILIATION_SHEET_NAME: ("Reconciliação normalizada", "CEP, desvios e status executivo"),
        METER_CATALOG_SHEET_NAME: ("Cadastro de medidores", "Rastreabilidade e mapeamento técnico"),
        IMPORT_SOURCES_SHEET_NAME: ("Fontes importadas", "Rastreabilidade e validação de conteúdo"),
        IMPORT_LOG_SHEET_NAME: ("Log de importação", "Rastreabilidade e validação de conteúdo"),
        PI_MASTER_SHEET_NAME: ("Histórico PI Vision", "PI Vision e condições de contorno"),
        PI_SHEET_NAME: ("PI Vision da janela", "PI Vision e condições de contorno"),
        "PI_EXTRACT_RESUMO": ("Resumo de rastreabilidade PI", "PI Vision"),
        ALARM_EVENT_MASTER_SHEET_NAME: ("Alarmes e eventos FCS320", "Alarmes/eventos e Pareto executivo"),
        ALARM_EVENT_SHEET_NAME: ("Alarmes e eventos FCS320 da janela", "Alarmes/eventos e Pareto executivo"),
        COMPARATIVO_TOTAL_SHEET_NAME: ("Base + PI consolidado", "Comparações e condições de contorno"),
        COMPARATIVO_PARES_SHEET_NAME: ("Topside × Subsea", "CEP e limites RANP 44"),
        MANUAL_SHEET_NAME: ("Entrada complementar manual", "Condições de contorno, quando preenchida"),
    }
    return descriptions.get(sheet_name, ("Aba complementar do arquivo-fonte", "Rastreabilidade e validação de conteúdo"))


def _dashboard_sources_panel(source_frames: dict[str, pd.DataFrame], workbook_path: Path | None) -> str:
    """Inventário visual de todas as abas reais e de como entram no dashboard."""
    rows = []
    available = 0
    total_records = 0
    for sheet_name, frame in source_frames.items():
        purpose, destination = _sheet_purpose(sheet_name)
        count = _source_frame_rows(frame) if frame is not None else 0
        loaded = frame is not None
        available += int(loaded)
        total_records += count
        rows.append({
            "Aba / fonte": sheet_name,
            "Status": "INGESTÃO ATIVA" if loaded else "NÃO DISPONÍVEL",
            "Registros": count,
            "Colunas": _source_frame_columns(frame) if loaded else 0,
            "Impressão digital": _source_frame_fingerprint(frame)[:16] if loaded else "",
            "Conteúdo": purpose,
            "Uso no dashboard": destination,
        })
    master_label = html.escape(str(workbook_path)) if workbook_path else "Fonte em memória — arquivo não associado"
    record_count = f"{total_records:,}".replace(",", ".")
    cards = (
        "<div class='metric-grid source-kpis'>"
        f"<div class='metric-card good'><span>Abas carregadas</span><b>{available}/{len(source_frames)}</b><small>todas as abas do arquivo-fonte</small></div>"
        f"<div class='metric-card accent'><span>Registros indexados</span><b>{record_count}</b><small>somatório das fontes carregadas</small></div>"
        f"<div class='metric-card'><span>Arquivo-fonte</span><b>{'TOTAL' if workbook_path and workbook_path.name == 'BASE_UNICA_TOTAL.xlsx' else 'STANDALONE'}</b><small>origem desta publicação</small></div>"
        "</div>"
    )
    return (
        f"<div class='info-box'><b>Rastreabilidade:</b> {master_label}. Esta visão mostra quais abas estão efetivamente "
        "disponíveis para a publicação e onde cada conjunto é utilizado. Aba sem dados não é tratada como zero.</div>"
        + cards
        + "<div class='table-wrap source-table'>" + _html_table(rows, max_rows=30) + "</div>"
    )


def _data_lineage_panel(context_df: pd.DataFrame, target_days: list, source_manifest: dict) -> str:
    """Explica, sem criar variáveis, como as evidências chegam às saídas."""
    work = _normalize_master_columns(context_df)
    days = {str(day)[:10] for day in target_days}
    if not work.empty:
        work = work[work["ProductionDate"].astype(str).str[:10].isin(days)].copy()
    origins = work.get("Origin", pd.Series(dtype=str)).astype(str).value_counts().to_dict()
    sheets = {item.get("name", ""): int(item.get("rows", 0) or 0) for item in source_manifest.get("sheets", [])}
    stages = [
        ("01", "Aquisição", f"{origins.get('MPFM', 0)} linhas MPFM", "PDF Daily/Hourly preservado por data, banco, TAG e instrumento"),
        ("02", "Separador", f"{origins.get('SEP', 0)} linhas SEP", "TXT de óleo, gás e água como fonte independente"),
        ("03", "Contexto PI", f"{sheets.get(PI_MASTER_SHEET_NAME, sheets.get(PI_SHEET_NAME, 0))} linhas PI", "variáveis de processo, status e modos quando disponíveis"),
        ("04", "Base Única", f"{len(work)} linhas na janela", "camada consolidada, deduplicada e rastreável"),
        ("05", "Cálculos", f"{sheets.get(COMPARATIVO_PARES_SHEET_NAME, 0)} comparações", "massas corrigidas, reconciliação e pares físicos Subsea × Topside"),
        ("06", "Publicação", "HTML + Excel + XML sob demanda", "painel navegável, evidência tabular e XML042 com trava de unicidade"),
    ]
    stage_html = "".join(
        f"<article class='lineage-stage'><span>{number}</span><div><h3>{html.escape(title)}</h3><b>{html.escape(value)}</b><small>{html.escape(note)}</small></div></article>"
        for number, title, value, note in stages
    )
    rows = [
        {"Dado": "Massas MPFM corrigidas", "Origem": "PDF MPFM", "Transformação": "massa bruta × K de fase", "Uso": "HC, Total, pares físicos e CEP"},
        {"Dado": "Volumes padrão @20 °C/1 atm", "Origem": "PDF/PVT", "Transformação": "rota de condição padrão preservada", "Uso": "tabelas, diagnóstico e XML042"},
        {"Dado": "Óleo, gás e água do SEP", "Origem": "TXT Separador", "Transformação": "extração por fase sem alinhamento automático", "Uso": "aba SEP e comparação escolhida pelo usuário"},
        {"Dado": "GVF, WLR, GOR e modos", "Origem": "PI Vision / COMPARATIVO_TOTAL", "Transformação": "chave data+banco+TAG", "Uso": "contexto operacional da Visão Integrada"},
        {"Dado": "Desvio HC e Total", "Origem": "Subsea + Topside físico", "Transformação": "(Subsea corrigido − Topside) / Topside × 100", "Uso": "limites ±10% e ±7%; gatilho temporal"},
        {"Dado": "XML042", "Origem": "BASE_UNICA_TOTAL.xlsx", "Transformação": "uma emissão por data+poço no registro SQLite", "Uso": "geração ANP sob demanda"},
    ]
    return (
        "<div class='info-box'><b>Cadeia auditável:</b> cada bloco abaixo indica a origem, o tratamento e o destino. O Separador não é vinculado automaticamente a nenhum MPFM; Topside × Subsea usa apenas os pares físicos configurados.</div>"
        f"<div class='lineage-grid'>{stage_html}</div>"
        "<h3>Mapa de rastreabilidade por grupo de dados</h3>"
        f"<div class='table-wrap'>{_html_table(rows, max_rows=30)}</div>"
    )


def _acceptance_panel(source_manifest: dict, findings: list[dict] | None = None) -> str:
    sheets = source_manifest.get("sheets", [])
    rules = [
        ("Arquivo Excel identificado", bool(source_manifest.get("workbook"))),
        ("Todas as abas carregadas", bool(sheets)),
        ("Estrutura registrada", all(item.get("columns", 0) > 0 for item in sheets)),
        ("Conteúdo com impressão digital", all(len(item.get("fingerprint", "")) == 64 for item in sheets)),
        ("Data/hora em padrão brasileiro", source_manifest.get("dateTimeFormat") == "dd/mm/yyyy hh:mm"),
        ("Vírgula como separador decimal", source_manifest.get("decimalSeparator") == ","),
    ]
    rows = [{"Critério de aceitação": label, "Resultado": "APROVADO" if passed else "REPROVADO"} for label, passed in rules]
    overall = all(passed for _, passed in rules)
    findings = findings or []
    technical_overall = all(passed for _, passed in rules)
    analytical_ok = not findings
    overall_label = "APROVADA" if technical_overall and analytical_ok else ("APROVADA COM RESSALVAS" if technical_overall else "REPROVADA")
    overall_class = "good" if technical_overall and analytical_ok else "warn"
    finding_rows = findings or [{"Origem": "Todas as abas", "Situação": "Nenhuma pendência analítica identificada", "Impacto": "Janela sem ressalvas registradas"}]
    sheet_rows = [{"Aba do Excel": item.get("name", ""), "Validação": "APROVADA" if item.get("columns", 0) and len(item.get("fingerprint", "")) == 64 else "REPROVADA", "Registros": item.get("rows", 0), "Colunas": item.get("columns", 0)} for item in sheets]
    return (
        f"<div class='metric-grid'><div class='metric-card {'good' if overall else 'warn'}'><span>Validação final</span>"
        f"<b class='{overall_class}'>{overall_label}</b><small>{sum(passed for _, passed in rules)}/{len(rules)} critérios técnicos; {len(findings)} ressalva(s) analítica(s)</small></div>"
        f"<div class='metric-card accent'><span>Abas verificadas</span><b>{len(sheets)}</b><small>linhas, colunas e conteúdo confrontados</small></div></div>"
        "<div class='info-box'><b>Como interpretar:</b> APROVADA significa que a entrega técnica Excel → HTML foi conferida. A situação só fica plenamente aprovada quando também não existem ressalvas operacionais nas abas PI, alarmes/eventos, detalhes, cobertura ou comparações.</div>"
        "<h3>Critérios técnicos de publicação</h3><div class='table-wrap'>" + _html_table(rows, max_rows=30) + "</div>"
        "<h3>Abas do Excel verificadas</h3><div class='info-box'>A contagem abaixo se refere às abas existentes no arquivo Excel-fonte, não a telas do HTML.</div><div class='table-wrap'>" + _html_table(sheet_rows, max_rows=50) + "</div>"
        "<h3>Ressalvas consolidadas das abas</h3><div class='table-wrap'>" + _html_table(finding_rows, max_rows=30) + "</div>"
    )


def validate_dashboard_delivery(dashboard_path: Path, source_frames: dict[str, pd.DataFrame], workbook_path: Path | None) -> dict:
    """Última camada: confronta o manifesto gravado no HTML com o Excel carregado."""
    text = dashboard_path.read_text(encoding="utf-8")
    match = re.search(r'<script id="source-manifest" type="application/json">(.*?)</script>', text, re.DOTALL)
    expected = _dashboard_source_manifest(source_frames, workbook_path)
    actual = json.loads(html.unescape(match.group(1))) if match else {}
    expected_by_name = {item["name"]: item for item in expected["sheets"]}
    actual_by_name = {item["name"]: item for item in actual.get("sheets", [])}
    rules = {
        "html_rico_gerado": "Dashboard rico não foi concluído" not in text and 'id="leadCards"' in text and "leadRender();" in text,
        "layout_responsivo_sem_lacuna": "grid-auto-rows: 1fr" not in text,
        "manifesto_embutido": bool(match),
        "arquivo_fonte_correto": actual.get("workbook") == expected.get("workbook"),
        "todas_abas_representadas": set(actual_by_name) == set(expected_by_name),
        "linhas_colunas_conferem": all(actual_by_name.get(name, {}).get("rows") == item["rows"] and actual_by_name.get(name, {}).get("columns") == item["columns"] for name, item in expected_by_name.items()),
        "conteudo_conferido_por_hash": all(actual_by_name.get(name, {}).get("fingerprint") == item["fingerprint"] for name, item in expected_by_name.items()),
        "localidade_pt_br": 'data-locale="pt-BR"' in text and actual.get("decimalSeparator") == ",",
        "formato_data_hora_br": actual.get("dateTimeFormat") == "dd/mm/yyyy hh:mm",
    }
    result = {
        "status": "APROVADO" if all(rules.values()) else "REPROVADO",
        "validatedAt": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "dashboard": str(dashboard_path.resolve()),
        "workbook": expected.get("workbook"),
        "rules": rules,
        "sheets": expected["sheets"],
    }
    report_path = dashboard_path.with_name(f"{dashboard_path.stem}_VALIDACAO.json")
    report_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    result["reportPath"] = str(report_path)
    return result


def write_dashboard(
    dashboard_path: Path,
    df_out: pd.DataFrame,
    target_days: list,
    selected_days: list,
    preflight_rows: list,
    master_path: Path | None = None,
    df_alarm_events: pd.DataFrame | None = None,
    aligned_bank: str = SEP_ALIGNED_BANK,
    preloaded_master_df: pd.DataFrame | None = None,
    preloaded_pi_df: pd.DataFrame | None = None,
    preloaded_comparativo_df: pd.DataFrame | None = None,
    preloaded_alarm_events: pd.DataFrame | None = None,
    source_workbook_path: Path | None = None,
    preloaded_source_frames: dict[str, pd.DataFrame] | None = None,
) -> Path:
    dashboard_started = time.perf_counter()
    context_parts = []
    if preloaded_master_df is not None:
        # Quem chamou já tem a Base_Unica total em memória (acabou de
        # atualizá-la/lê-la) — evita reabrir o arquivo inteiro do disco de
        # novo, o que fica cada vez mais caro conforme o histórico cresce.
        context_parts.append(preloaded_master_df)
    elif master_path and master_path.exists():
        # Em bases migradas, a visão do HTML é montada das tabelas
        # normalizadas. Bases antigas continuam usando o fallback legado.
        context_parts.append(read_dashboard_context(master_path))
    if preloaded_master_df is not df_out:
        context_parts.append(df_out)
    context_df = pd.concat(context_parts, ignore_index=True) if context_parts else df_out
    context_df = _normalize_master_columns(context_df)
    for key in MASTER_DEDUP_KEYS:
        context_df[key] = context_df[key].map(lambda value, key=key: _canonical_master_key_value(value, key))
    context_df.drop_duplicates(subset=MASTER_DEDUP_KEYS, keep="last", inplace=True)

    analysis_started = time.perf_counter()
    analysis = dashboard_analysis(df_out, context_df, target_days, preflight_rows)
    print(f"[TIME] dashboard_analysis: {time.perf_counter() - analysis_started:.1f}s", flush=True)
    panel_started = time.perf_counter()
    official_rows = _official_deviation_rows(context_df, target_days, aligned_bank)
    monthly_cep_days = _monthly_cep_days(context_df, selected_days)
    monthly_official_rows = _official_deviation_rows(context_df, monthly_cep_days, aligned_bank)
    comparison_panel = _interactive_comparison_panel(_dashboard_comparison_records(context_df, target_days, aligned_bank), target_days)
    separator_comparison_panel = _interactive_separator_comparison_panel(_separator_frontend_records(context_df, target_days), target_days)
    stability_panel = _interactive_stability_panel(_dashboard_timeseries_records(context_df, target_days), target_days)
    executive_panel = _executive_daily_panel(context_df, target_days, official_rows)
    overview_comparativo_df = (
        preloaded_comparativo_df
        if preloaded_comparativo_df is not None
        else _read_master_sheet(master_path, COMPARATIVO_TOTAL_SHEET_NAME)
    )
    leadership_panel = _leadership_dashboard_panel(
        context_df,
        target_days,
        overview_comparativo_df,
        official_rows,
    )
    calendar_panel = _coverage_calendar_panel(context_df, target_days)
    separator_panel = _separator_dashboard_panel(context_df, target_days)
    hourly_profile_panel = _hourly_profile_svg(context_df, target_days)
    print(f"[TIME] comparison/stability/executive/leadership/coverage/separator/profile: {time.perf_counter() - panel_started:.1f}s", flush=True)
    panel_started = time.perf_counter()
    cep_panel = _interactive_cep_panel(monthly_official_rows)
    xml042_panel = _xml042_panel(master_path=master_path, target_days=target_days, context_df=context_df)
    print(f"[TIME] official/CEP/XML042: {time.perf_counter() - panel_started:.1f}s", flush=True)
    source_frames = preloaded_source_frames if preloaded_source_frames is not None else _load_dashboard_source_frames(source_workbook_path)
    pi_source = source_frames.get(PI_MASTER_SHEET_NAME, source_frames.get(PI_SHEET_NAME, pd.DataFrame()))
    alarm_source = source_frames.get(ALARM_EVENT_MASTER_SHEET_NAME, source_frames.get(ALARM_EVENT_SHEET_NAME, pd.DataFrame()))
    pi_master_df = preloaded_pi_df if preloaded_pi_df is not None else (pi_source if not pi_source.empty else read_master_pi_extract(master_path))
    comparativo_df = overview_comparativo_df
    if comparativo_df is None or comparativo_df.empty:
        comparativo_df = source_frames.get(COMPARATIVO_TOTAL_SHEET_NAME, _read_master_sheet(master_path, COMPARATIVO_TOTAL_SHEET_NAME))
    if comparativo_df.empty:
        comparativo_df = build_comparativo_total_df(context_df, pi_master_df)
    boundary_panel = _boundary_conditions_panel(comparativo_df, target_days)
    if not source_frames:
        source_frames = {MASTER_SHEET_NAME if master_path else "BASE_UNICA_STANDALONE": context_df}
    source_manifest = _dashboard_source_manifest(source_frames, source_workbook_path)
    sources_panel = _dashboard_sources_panel(source_frames, source_workbook_path)
    lineage_panel = _data_lineage_panel(context_df, target_days, source_manifest)
    alarm_event_rows = []
    alarm_event_summary = []
    alarm_flag_rows = []
    ae = preloaded_alarm_events if preloaded_alarm_events is not None else (alarm_source if not alarm_source.empty else _merge_dashboard_alarm_events(master_path, df_alarm_events, selected_days))
    if not ae.empty:
        ae = ae.where(pd.notna(ae), "")
        alarm_event_rows = ae.to_dict("records")
        grouped_ae = ae.groupby(["ProductionDate", "RecordType", "Priority", "IssueFlag"], dropna=False).size().reset_index(name="Quantidade")
        alarm_event_summary = grouped_ae.rename(columns={"ProductionDate": "Dia", "RecordType": "Tipo", "IssueFlag": "Flag"}).to_dict("records")
        flagged = ae[ae["IssueFlag"].astype(str).str.len() > 0].copy()
        flagged["Interpretação"] = flagged.apply(lambda row: _alarm_explanation(row.get("IssueFlag", ""), row.get("RecordType", "")), axis=1)
        alarm_flag_rows = flagged.sort_values(by=["ProductionDate", "Timestamp"], key=lambda col: col.astype(str), ascending=[True, False]).head(200).to_dict("records")
    acceptance_findings = []
    if analysis.get("pressure_flags"):
        acceptance_findings.append({"Origem": "Detalhes", "Situação": f"{len(analysis['pressure_flags'])} ponto(s) acima de 490 barg", "Impacto": "Requer verificar gás medido e condição acima do ponto de bolha"})
    if analysis.get("anomalies"):
        acceptance_findings.append({"Origem": "Detalhes/Cobertura", "Situação": f"{len(analysis['anomalies'])} anomalia(s), zerado(s) ou ausente(s)", "Impacto": "Não significa falha automática; exige avaliação operacional"})
    if not pi_master_df.empty:
        pi_day_col = "PI Dia Coleta" if "PI Dia Coleta" in pi_master_df.columns else "PI Inicio"
        pi_window = pi_master_df[pi_master_df[pi_day_col].map(_pi_day_iso).isin({str(day) for day in target_days})]
        if "PI Status Coleta" in pi_window.columns:
            pi_bad = int((~pi_window["PI Status Coleta"].astype(str).str.upper().isin({"OK", "SUCESSO", "APROVADO"})).sum())
            if pi_bad:
                acceptance_findings.append({"Origem": "PI Vision", "Situação": f"{pi_bad} linha(s) com falha/erro de coleta", "Impacto": "Contexto PI não está 100% aprovado"})
    alarm_issue_count = int(ae["IssueFlag"].astype(str).str.contains("MISSING_FILE|READ_ERROR|EMPTY_OR_UNPARSED", regex=True, na=False).sum()) if not ae.empty else 0
    if alarm_issue_count:
        acceptance_findings.append({"Origem": "Alarmes/eventos", "Situação": f"{alarm_issue_count} registro(s) com problema de arquivo/leitura", "Impacto": "Cobertura de alarmes/eventos requer conferência"})
    acceptance_panel = _acceptance_panel(source_manifest, acceptance_findings)
    dashboard_path.parent.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    try:
        dashboard_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        dashboard_ip = "não identificado"
    target_day_set = {str(day) for day in target_days}
    context_dates = context_df.get("ProductionDate", pd.Series(dtype=str)).astype(str).str[:10]
    context_window_df = context_df.loc[context_dates.isin(target_day_set)]
    if preflight_rows:
        preflight_ok_count = sum(1 for row in preflight_rows if str(row.get("Status", "")).upper() == "OK")
        preflight_not_ok_count = max(0, len(preflight_rows) - preflight_ok_count)
    else:
        days_with_data = set(context_window_df.get("ProductionDate", pd.Series(dtype=str)).astype(str).str[:10])
        preflight_ok_count = len(days_with_data & target_day_set)
        preflight_not_ok_count = max(0, len(target_day_set - days_with_data))
    header_record_count = len(context_window_df)
    header_summary = (
        f"Gerado em {generated_at} | IP: {dashboard_ip} | "
        f"Dias na janela: {len(selected_days)} | Dias processados: {len(target_days)} | "
        f"Dados OK: {preflight_ok_count} | Dados não OK: {preflight_not_ok_count} | "
        f"Registros analisados: {header_record_count} | SEP: fonte independente"
    )
    logo_data_uri = _dashboard_logo_data_uri()
    logo_markup = (
        f'<img src="{logo_data_uri}" alt="Logo do sistema multifásico">'
        if logo_data_uri else ""
    )
    issue_count = sum(1 for row in preflight_rows if row.get("Status") != "OK")
    configured_point_rows = []
    for instrument, metadata in MPFM_INSTRUMENT_METADATA.items():
        point_data = context_df[
            (context_df.get("Instrumento", pd.Series(dtype=str)).astype(str).map(_normalize_instrument) == instrument)
            & (context_df.get("ProductionDate", pd.Series(dtype=str)).astype(str).str[:10].isin({str(day) for day in target_days}))
        ]
        configured_point_rows.append({
            "Banco": metadata.get("bank", ""),
            "Local": metadata.get("entity", ""),
            "TAG": metadata.get("tag", ""),
            "Tipo": metadata.get("tipo", ""),
            "Instrumento": instrument,
            "Cobertura na janela": "Com dados" if not point_data.empty else "Sem dados — aguardando Daily/PI",
        })
    configured_points_panel = (
        "<div class='info-box'><b>Cadastro físico dos pontos:</b> instrumentos distintos podem compartilhar o mesmo banco e o mesmo arquivo PDF sem criar bancos artificiais.</div>"
        "<div class='table-wrap'>" + _html_table(configured_point_rows, max_rows=20) + "</div>"
        if configured_point_rows else ""
    )
    source_manifest_json = json.dumps(source_manifest, ensure_ascii=False).replace("</", "<\\/")
    html_doc = fr"""<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Relatório de Monitoramento Multifásico</title>
  <style>
        :root {{
          --eds-bg-canvas:#eaf8fa; --eds-bg-surface:#f6ffff; --eds-bg-emphasis:#206f77; --eds-bg-emphasis-hover:#205c62;
          --eds-border-subtle:#bbdbdf; --eds-border-medium:#7cbac1; --eds-border-strong:#21767e;
          --eds-text-strong:#141f20; --eds-text-subtle:#1f6369; --eds-text-on-emphasis:#ffffff;
          --eds-info:#1d5d8f; --eds-info-bg:#e4f1f8; --eds-success:#256b43; --eds-success-bg:#e6f4eb;
          --eds-warning:#a15c00; --eds-warning-bg:#fff1d6; --eds-danger:#b3261e; --eds-danger-bg:#fde8e7;
          --eq-blue:#0b3a4a; --eq-teal:#206f77; --eq-cyan:#7cbac1; --ink:var(--eds-text-strong); --muted:#4f6266; --line:var(--eds-border-subtle); --panel:var(--eds-bg-surface); --soft:var(--eds-bg-canvas);
        }}
        * {{ box-sizing: border-box; }}
        body {{ font-family: Segoe UI, Arial, sans-serif; margin: 0; background: var(--eds-bg-canvas); color: var(--ink); }}
        header {{ background: var(--eq-blue); color: var(--eds-text-on-emphasis); padding: 24px 34px; border-bottom: 4px solid var(--eq-cyan); }}
        .brand-row {{ display: flex; gap: 14px; align-items: center; }}
        .brand-mark {{ width: 44px; height: 44px; display: grid; place-items: center; border: 1px solid rgba(255,255,255,.38); border-radius: 4px; background: var(--eq-teal); }}
        .brand-mark img {{ width: 42px; height: 42px; object-fit: contain; display: block; }}
        .eyebrow {{ font-size: 11px; letter-spacing: .12em; text-transform: uppercase; opacity: .78; font-weight: 700; }}
        header h1 {{ letter-spacing: -.02em; margin: 3px 0 7px; }}
        header .meta {{ color: #d8eff1; font-size: 12px; line-height: 1.45; }}
        main {{ padding: 24px 32px; max-width: 1680px; margin: 0 auto; }}
        h1, h2 {{ margin: 0 0 12px; }}
        h2 {{ color: var(--eq-blue); font-size: 22px; border-left: 4px solid var(--eq-teal); padding-left: 10px; }}
        h3 {{ color: var(--eq-blue); margin-top: 22px; }}
        section {{ background: var(--panel); border-radius: 4px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 8px rgba(20,31,32,.10); border: 1px solid var(--eds-border-subtle); }}
    .cards {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; }}
        .card {{ background: var(--eds-bg-surface); border-left: 4px solid var(--eq-teal); border-radius: 4px; padding: 14px; box-shadow: inset 0 1px 0 rgba(255,255,255,.7); }}
    .card b {{ display: block; font-size: 24px; margin-top: 4px; }}
        .muted {{ color: var(--muted); }}
    .data-table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    .data-table th, .data-table td {{ border-bottom: 1px solid #e5e7eb; padding: 7px 8px; text-align: left; vertical-align: top; }}
    .data-table th {{ background: #dceff1; position: sticky; top: 0; color: var(--eq-blue); }}
    .table-note {{ margin: 8px 0; padding: 7px 10px; background: #fff7ed; border-left: 4px solid #f59e0b; border-radius: 7px; }}
    .table-wrap {{ max-height: 420px; overflow: auto; border: 1px solid var(--eds-border-subtle); border-radius: 4px; }}
    .chart {{ width: 100%; height: auto; background: #fbfdff; border: 1px solid #e5e7eb; border-radius: 10px; }}
    .wide-chart {{ width: auto; max-width: none; }}
    .chart-scroll > svg.wide-chart:not([id]) {{ width: 100%; max-width: 100%; min-width: 0; display: block; }}
    .chart-scroll {{ overflow-x: auto; overflow-y: hidden; border: 1px solid #e5e7eb; border-radius: 10px; background: #fbfdff; }}
    .controls {{ display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 12px; align-items: end; }}
    .controls label {{ display: flex; flex-direction: column; gap: 4px; font-size: 12px; color: #475569; font-weight: 600; }}
    .controls select, .controls input {{ min-width: 160px; padding: 8px 10px; border: 1px solid #cbd5e1; border-radius: 8px; background: white; color: #0f172a; }}
    .controls select[multiple] {{ min-width: 260px; }}
    .info-box {{ background: var(--eds-info-bg); border-left: 4px solid var(--eds-info); border-radius: 4px; padding: 12px 14px; margin: 10px 0 14px; color: #244858; line-height: 1.45; }}
    .formula-grid, .metric-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(230px, 1fr)); gap: 12px; margin: 12px 0; }}
    #leadCards {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
    .formula-card, .metric-card {{ background: var(--eds-bg-surface); border: 1px solid var(--line); border-radius: 4px; padding: 14px; }}
    .formula-card b, .metric-card span {{ display: block; color: var(--eq-blue); margin-bottom: 7px; font-weight: 700; }}
    .formula-card span {{ display: block; font-family: Consolas, monospace; background: #f8fafc; border-radius: 8px; padding: 8px; color: #0f172a; }}
    .formula-card small, .metric-card small {{ color: var(--muted); display: block; margin-top: 7px; }}
    .metric-card b {{ display: block; font-size: 26px; color: var(--ink); }}
    .metric-card.good {{ border-left: 5px solid var(--eds-success); }}
    .metric-card.warn {{ border-left: 5px solid var(--eds-warning); }}
    .metric-card.accent {{ border-left: 5px solid var(--eq-teal); }}
    .dashboard-card-container {{ background: var(--eds-bg-surface); border: 1px solid var(--line); border-radius: 8px; padding: 12px; margin: 12px 0; }}
    .dashboard-card-container > h3 {{ margin: 0 0 8px; color: var(--eq-blue); }}
    .integrated-note {{ grid-column: 1 / -1; margin: 0 0 10px; padding: 8px 12px; color: #244858; background: #eef8fa; border-left: 4px solid var(--eq-teal); border-radius: 6px; font-size: 12px; }}
    .pair-card {{ cursor: pointer; }} .pair-card:focus {{ outline: 3px solid rgba(0,125,138,.25); outline-offset: 2px; }}
    .pair-detail {{ grid-column: 1 / -1; padding: 10px; border: 1px solid var(--line); border-radius: 8px; background: #f8fafc; }} .pair-detail[hidden] {{ display:none; }} .pair-detail h4 {{ margin:0 0 8px; color:var(--eq-blue); }}
    .rank-filter-row {{ display:flex; flex-wrap:wrap; gap:6px; margin:6px 0 8px; }}
    .rank-filter-row button {{ border:1px solid #cbd5e1; border-radius:999px; background:#fff; color:#334155; padding:5px 10px; cursor:pointer; font-size:11px; font-weight:700; }}
    .rank-filter-row button.active, .rank-filter-row button:hover {{ background:#007d8a; border-color:#007d8a; color:#fff; }}
    .stacked-bar {{ display: flex; height: 26px; border-radius: 999px; overflow: hidden; background: #e2e8f0; margin: 10px 0; box-shadow: inset 0 1px 3px rgba(15,23,42,.15); }}
    .mini-bars {{ display: grid; gap: 10px; }}
    .mini-row {{ display: grid; grid-template-columns: 160px 110px 1fr; gap: 10px; align-items: center; }}
    .mini-track {{ height: 18px; background: #e2e8f0; border-radius: 999px; overflow: hidden; }}
    .mini-track div {{ height: 100%; border-radius: 999px; }}
    .legend-row {{ display: flex; flex-wrap: wrap; gap: 10px 18px; align-items: center; padding: 8px 0 10px; }}
    .legend-item {{ display: inline-flex; align-items: center; gap: 7px; font-size: 12px; color: #334155; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 999px; padding: 5px 9px; }}
    .legend-item.legend-toggle {{ cursor: pointer; user-select: none; }}
    .legend-item.legend-toggle.legend-off {{ opacity: .38; text-decoration: line-through; }}
    .chart-toolbar {{ display: flex; justify-content: flex-end; gap: 6px; margin: 6px 0; }}
    .chart-toolbar button, .chart-toolbar input {{ min-width: 32px; padding: 4px 9px; border: 1px solid #cbd5e1; border-radius: 6px; background: #fff; color: #0f172a; cursor: pointer; font-weight: 700; box-sizing: border-box; }}
    .chart-toolbar input {{ width: 58px; text-align: center; cursor: text; }}
    .chart-toolbar button:hover {{ background: #e0f2f1; border-color: #007d8a; }}
    .legend-item i {{ display: inline-block; width: 22px; height: 4px; border-radius: 999px; }}
    .calendar-wrap {{ max-height: 620px; }}
    .calendar-table {{ min-width: 980px; }}
    .calendar-table th:first-child {{ min-width: 190px; position: sticky; left: 0; z-index: 2; }}
    .calendar-table td, .calendar-table th {{ text-align: center; min-width: 118px; }}
    .calendar-cell {{ min-height: 54px; border-radius: 8px; padding: 6px; border-left: 4px solid #94a3b8; background: #f8fafc; }}
    .calendar-cell b, .calendar-cell small, .calendar-cell em {{ display: block; }}
    .calendar-cell b {{ font-size: 15px; }}
    .calendar-cell small, .calendar-cell em {{ font-size: 10px; font-style: normal; color: #475569; }}
    .calendar-cell.cal-ok {{ border-left-color: #16a34a; background: #f0fdf4; }}
    .calendar-cell.cal-partial {{ border-left-color: #f59e0b; background: #fffbeb; }}
    .calendar-cell.cal-absent {{ border-left-color: #dc2626; background: #fef2f2; }}
    .calendar-cell.cal-closed {{ border-left-color: #64748b; background: #f1f5f9; }}
    .legend-item i.cal-ok, .legend-item i.cal-partial, .legend-item i.cal-absent, .legend-item i.cal-closed {{ height: 10px; width: 10px; border-radius: 3px; }}
    .legend-item i.cal-ok {{ background: #16a34a; }} .legend-item i.cal-partial {{ background: #f59e0b; }} .legend-item i.cal-absent {{ background: #dc2626; }} .legend-item i.cal-closed {{ background: #64748b; }}
    .alarm-category-list {{ display: grid; gap: 10px; margin: 12px 0; }}
    .alarm-category-row {{ display: grid; grid-template-columns: minmax(220px, 0.8fr) minmax(260px, 2fr); gap: 12px; align-items: center; }}
    .alarm-category-row small {{ display: block; color: var(--muted); margin-top: 3px; }}
    .alarm-category-track {{ height: 18px; background: #e2e8f0; border-radius: 999px; overflow: hidden; }}
    .alarm-category-track span {{ display: block; height: 100%; border-radius: 999px; }}
    .tab-nav {{ display: flex; flex-wrap: wrap; gap: 6px; padding: 8px; margin: 0 0 18px; background: var(--eds-bg-surface); border: 1px solid var(--line); border-radius: 4px; position: sticky; top: 8px; z-index: 10; box-shadow: 0 2px 8px rgba(20,31,32,.10); }}
    .tab-button {{ border: 1px solid var(--eds-border-medium); background: var(--eds-bg-surface); color: var(--eq-blue); padding: 9px 13px; border-radius: 4px; font-weight: 700; cursor: pointer; }}
    .tab-button:hover {{ background: #dceff1; border-color: var(--eq-teal); }}
    .tab-button.active {{ background: var(--eq-blue); color: #ffffff; border-color: var(--eq-blue); }}
    .tab-panel[hidden] {{ display: none !important; }}
    @media print {{ .tab-nav {{ display: none; }} .tab-panel[hidden] {{ display: block !important; }} }}
    @media (max-width: 900px) {{ #leadCards {{ grid-template-columns: 1fr; }} }}
    .delta {{ display: inline-block; min-width: 80px; padding: 3px 8px; border-radius: 999px; font-weight: 700; }}
    .delta.up {{ color: var(--eds-success); background: var(--eds-success-bg); }}
    .delta.down {{ color: var(--eds-danger); background: var(--eds-danger-bg); }}
    .delta.flat {{ color: #334155; background: #e2e8f0; }}
    .chart-title {{ font-weight: 700; fill: #0f172a; }}
    .axis-label, .value-label {{ font-size: 12px; fill: #334155; }}
    .bar {{ fill: #007398; }}
    .chart-grid {{ display: grid; grid-template-columns: 1fr; gap: 18px; align-items: start; }}
    .chart-grid:has(#leadTrend), .chart-grid:has(#leadRank), .chart-grid:has(#sepTrend), .chart-grid:has(#sepMix) {{ grid-template-columns: 1fr; }}
    .chart-grid > div {{ min-width: 0; }}
    /* Tema visual baseado no modelo BASE_UNICA_CORPORATE_DASHBOARD_20260806-1. */
    :root {{
      --navy:#071e41; --navy2:#0a2d5d; --teal:#007d8a; --cyan:#11a8b7; --ink:#102238; --muted:#627287; --line:#d9e1e8; --bg:#f5f8fb; --card:#ffffff;
      --green:#1b8e47; --amber:#e98a00; --red:#d73535; --blue:#1565c0; --shadow:0 8px 22px rgba(16,34,56,.07);
      --eds-bg-canvas:var(--bg); --eds-bg-surface:var(--card); --eds-border-subtle:var(--line); --eds-border-medium:#cbd6e1;
      --eds-success:var(--green); --eds-success-bg:#e0f4e5; --eds-warning:var(--amber); --eds-warning-bg:#fff0d0; --eds-danger:var(--red); --eds-danger-bg:#ffe0e0;
      --eq-blue:var(--navy); --eq-teal:var(--teal); --eq-cyan:var(--cyan); --panel:var(--card); --soft:var(--bg);
    }}
    html {{ scroll-behavior: smooth; }}
    body {{ margin: 0; font-family: Inter, "Segoe UI", Arial, sans-serif; background: var(--bg); color: var(--ink); font-size: 14px; }}
    button, select, input {{ font: inherit; }}
    header {{ position:fixed; inset:0 0 auto 0; z-index:30; background:linear-gradient(110deg,#062b59 0%,#0b477d 55%,#12639a 100%); color:#fff; height:auto; min-height:96px; padding:12px 28px; display:flex; align-items:center; justify-content:space-between; border-bottom:3px solid #d7263d; box-shadow:0 3px 12px rgba(16,34,56,.16); transition:min-height .2s ease,padding .2s ease; }}
    header.header--compact {{ min-height:48px !important; padding-top:6px !important; padding-bottom:6px !important; }}
    header.header--compact .meta, header.header--compact .eyebrow {{ display:none !important; }}
    header.header--compact h1 {{ font-size:15px; margin:0; }}
    header h1 {{ font-size:21px; font-weight:780; letter-spacing:-.03em; margin:2px 0 4px; }}
    header .meta {{ color:#dbeafe; font-size:11px; line-height:1.4; }}
    .eyebrow {{ color:#bfdbfe; }}
    .brand-row {{ width:100%; min-width:0; flex-wrap:wrap; align-items:center; }}
    .brand-mark {{ width:clamp(38px,5vw,56px); height:clamp(38px,5vw,56px); flex:0 0 clamp(38px,5vw,56px); border-color:#d8e2ec; border-radius:8px; background:#f4f8fc; }}
    .brand-mark img {{ width:calc(100% - 2px); height:calc(100% - 2px); max-width:100%; object-fit:contain; }}
    .header-status {{ display:flex; align-items:center; gap:10px; padding-left:22px; color:#e0f2fe; font-size:11px; white-space:nowrap; }}
    .header-status:before {{ content:""; width:9px; height:9px; border-radius:50%; background:var(--green); box-shadow:0 0 0 4px #e4f5ea; }}
    main {{ max-width:none; margin:0 0 0 214px; padding:calc(var(--header-height,96px) + 20px) 24px 38px; }}
    @media(max-width:1100px) {{ main {{ padding-top:116px !important; }} }}
    h2 {{ color: var(--navy); font-size: 19px; letter-spacing: -.02em; border-left: 4px solid var(--teal); padding-left: 10px; margin-top: 18px; }}
    h3 {{ color: var(--navy); font-size: 15px; letter-spacing: -.01em; margin: 22px 0 10px; }}
    section {{ background: var(--card); border: 1px solid var(--line); border-radius: 10px; box-shadow: var(--shadow); padding: 16px; margin-bottom: 14px; }}
    section.tab-panel {{ background:transparent; border:0; box-shadow:none; padding:0; margin-bottom:0; }}
    .tab-nav {{ position:fixed; left:0; top:var(--header-height,96px); bottom:auto; width:214px; z-index:20; height:calc(100vh - var(--header-height,96px)); max-height:calc(100vh - var(--header-height,96px)); background:#fff; border:0; border-right:1px solid var(--line); display:flex; flex-direction:column; align-items:stretch; gap:5px; padding:20px 12px; margin:0; overflow-y:auto; box-shadow:none; border-radius:0; }}
    .tab-button {{ min-height:42px; height:auto; border:0; background:transparent; color:#53667a; font-weight:680; border-radius:7px; padding:10px 12px; cursor:pointer; white-space:normal; text-align:left; }}
    .tab-button:hover {{ color:var(--navy); background:#f0f5fa; }}
    .tab-button.active {{ color:#fff; background:var(--blue); box-shadow:0 5px 12px rgba(21,101,192,.18); }}
    .filter-panel {{ background: #fff; border: 1px solid var(--line); border-radius: 10px; padding: 0 14px; margin: 0 0 12px; box-shadow: 0 4px 14px rgba(16,34,56,.04); }}
    .filter-panel summary {{ cursor: pointer; color: var(--navy); font-weight: 750; padding: 11px 2px; list-style: none; }}
    .filter-panel summary::-webkit-details-marker {{ display: none; }}
    .filter-panel summary:before {{ content: "＋"; display: inline-block; width: 20px; color: var(--teal); font-size: 16px; }}
    .filter-panel[open] summary:before {{ content: "−"; }}
    .controls {{ background: transparent; border: 0; border-radius: 0; padding: 2px 0 14px; margin-bottom: 0; box-shadow: none; }}
    .controls label {{ color: var(--muted); font-weight: 750; text-transform: uppercase; letter-spacing: .04em; }}
    .controls select, .controls input {{ height: 36px; border: 1px solid #cbd6e1; border-radius: 7px; color: var(--ink); }}
    .info-box {{ background: #f4f8fc; border-left: 4px solid var(--blue); border-radius: 7px; color: #4e6276; }}
    .table-wrap {{ border: 1px solid var(--line); border-radius: 8px; background: #fff; }}
    .data-table {{ font-size: 12px; border-spacing: 0; overflow: hidden; }}
    .data-table th, .data-table td {{ border-bottom: 1px solid #edf1f4; padding: 10px 12px; white-space: nowrap; text-align: center; vertical-align: middle; }}
    .data-table th {{ background: #f5f8fb; color: #53667a; font-size: 10px; text-transform: uppercase; letter-spacing: .04em; position: sticky; top: 0; z-index: 1; }}
    .data-table tbody tr:nth-child(even) {{ background: #fbfdff; }}
    .data-table tbody tr:hover {{ background: #eaf4fb; }}
    .data-table tbody tr:last-child td {{ border-bottom: 0; }}
    .metric-card, .formula-card, .card {{ background: var(--card); border: 1px solid var(--line); border-radius: 8px; box-shadow: none; }}
    .metric-card {{ min-height: 106px; position: relative; overflow: hidden; border-left: 1px solid var(--line); }}
    .metric-card:before {{ content: ""; position: absolute; left: 0; top: 0; bottom: 0; width: 4px; background: var(--teal); }}
    .metric-card span {{ color: var(--muted); font-size: 12px; }}
    .metric-card b {{ font-size: 27px; font-weight: 800; letter-spacing: -.04em; color: var(--ink); }}
    .metric-card.good:before {{ background: var(--green); }}
    .metric-card.warn:before {{ background: var(--amber); }}
    .metric-card.accent:before {{ background: var(--teal); }}
    .metric-card span, .metric-card b, .metric-card small {{ text-align: center; }}
    .chart, .chart-scroll {{ background: #fff; border: 1px solid var(--line); border-radius: 8px; }}
    .chart-grid {{ gap: 14px; }}
        header {{ overflow:hidden; }}
        .brand-row {{ min-width: 0; position: relative; z-index: 1; flex-wrap: wrap; align-items: flex-start; }}
        .brand-row > div:last-child {{ min-width: 0; }}
        .brand-row > div:nth-child(2) {{ min-width: 0; flex: 1 1 520px; }}
        header .meta {{ overflow-wrap: anywhere; word-break: break-word; white-space: normal; line-height: 1.55; }}
        .header-status {{ margin-left: auto; white-space: normal; text-align: right; max-width: 100%; }}
        .tab-button {{ position: relative; transition: color .18s ease, border-color .18s ease; }}
        .tab-button:focus-visible, select:focus-visible, input:focus-visible {{ outline: 3px solid rgba(21,101,192,.28); outline-offset: 2px; }}
        .mobile-tab-select {{ display: none; width: 100%; height: 42px; border: 1px solid var(--line); border-radius: 8px; background:#fff; color:var(--navy); font-weight:750; padding:0 12px; }}
        .tab-panel > h2 {{ margin: 20px 0 12px; padding: 12px 14px; border: 1px solid var(--line); border-left: 5px solid var(--teal); border-radius: 9px; background: linear-gradient(90deg,#fff,#f7fbfd); box-shadow: 0 3px 12px rgba(16,34,56,.04); }}
        .metric-grid {{ align-items: stretch; }}
        .metric-card {{ min-width: 0; min-height: 126px; height: 100%; display:flex; flex-direction:column; justify-content:center; gap:4px; padding:17px 15px 15px; transition: transform .18s ease, box-shadow .18s ease; }}
        .metric-card:hover {{ transform: translateY(-2px); box-shadow: 0 9px 22px rgba(16,34,56,.09); }}
        .metric-card:after {{ content:""; position:absolute; right:-24px; top:-24px; width:76px; height:76px; border-radius:50%; background:rgba(0,125,138,.06); }}
        .metric-card span {{ min-height: 30px; display:flex; align-items:flex-end; justify-content:center; text-transform:uppercase; letter-spacing:.045em; font-size:10px; }}
        .metric-card b {{ line-height:1.1; overflow-wrap:anywhere; }}
        .metric-card small {{ min-height: 30px; line-height:1.35; }}
        .dashboard-card-container {{ padding: 16px; border-radius: 10px; box-shadow: 0 5px 16px rgba(16,34,56,.05); }}
        .controls {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; align-items:end; }}
        .controls label, .controls select, .controls input {{ min-width:0; width:100%; max-width:100%; }}
        .controls select[multiple] {{ height:auto; min-height:120px; }}
        .chart-grid > div {{ background:#fff; border:1px solid var(--line); border-radius:10px; padding:14px; box-shadow:0 4px 14px rgba(16,34,56,.04); overflow:hidden; }}
        .chart-scroll {{ max-width:100%; position:relative;  }}
        .chart-scroll:after {{ content:"Deslize para explorar →"; position:sticky; left:12px; bottom:7px; display:inline-block; padding:3px 8px; border-radius:999px; background:rgba(7,30,65,.78); color:#fff; font-size:10px; pointer-events:none; }}
        .table-wrap {{ max-width:100%; position:relative; overflow:auto;  }}
        .table-toolbar {{ display:flex; align-items:center; justify-content:space-between; gap:10px; padding:9px 10px; background:#f7fafc; border-bottom:1px solid var(--line); position:sticky; left:0; z-index:3; }}
        .table-search {{ min-width:220px; max-width:380px; width:40%; height:34px; border:1px solid #cbd6e1; border-radius:7px; padding:0 10px; background:#fff; }}
        .table-count {{ color:var(--muted); font-size:11px; white-space:nowrap; }}
        .data-table th.sortable {{ cursor:pointer; -webkit-user-select:none; user-select:none; }}
        .data-table th.sortable:after {{ content:" ↕"; color:#94a3b8; }}
        .data-table td {{ max-width:360px; overflow:hidden; text-overflow:ellipsis; }}
        .data-table td:hover {{ white-space:normal; overflow:visible; }}
        .status-chip {{ display:inline-flex; align-items:center; justify-content:center; padding:4px 8px; border-radius:999px; font-size:10px; font-weight:800; letter-spacing:.025em; }}
        .status-ok {{ color:#116832; background:#e0f4e5; }} .status-warn {{ color:#8a5200; background:#fff0d0; }} .status-bad {{ color:#a12020; background:#ffe3e3; }} .status-info {{ color:#1556a0; background:#e5f0fb; }}
        .cell-bar {{ background-image:linear-gradient(90deg,rgba(17,168,183,.14) var(--cell-pct),transparent var(--cell-pct)); }}
        #cepTable .cell-bar {{ background-image:none; }}
        .source-table .data-table td:first-child {{ font-weight:800; color:var(--navy); }}
        .mini-row {{ grid-template-columns:minmax(130px,180px) minmax(80px,110px) minmax(160px,1fr); }}
        /* Tabelas corporativas: cabeçalho fixo ao rolar, zebra e hover. O
           border-collapse precisa ser "separate" para o thead sticky não
           perder as bordas durante a rolagem. */
        .table-wrap {{ max-height:70vh; }}
        .data-table {{ border-collapse:separate; border-spacing:0; }}
        .data-table thead th {{ position:sticky; top:0; z-index:2; background:#0f3b40; color:#fff; font-size:10px; font-weight:800; text-transform:none; letter-spacing:.02em; border-bottom:0; box-shadow:inset 0 -3px 0 var(--teal); }}
        .volume-source-table thead th.volume-col {{ background:#075e68; color:#fff; box-shadow:inset 0 -3px 0 #7ee2df; }}
        .volume-source-table tbody td.volume-col {{ background:#e8f8f7; font-weight:700; }}
        .data-table thead th.sortable:after {{ color:rgba(255,255,255,.7); }}
        .data-table tbody td {{ border-bottom:1px solid #e8eef2; }}
        .data-table tbody tr:nth-child(even) {{ background:#f4f8fa; }}
        .data-table tbody tr:hover {{ background:#e2f1f5; }}
        #leadTable tbody td:nth-child(n+7) {{ text-align:right; font-variant-numeric:tabular-nums; }}
        /* Cards multivalor: uma linha por medida dentro do mesmo card. */
        .metric-card .kv-list {{ display:flex; flex-direction:column; gap:2px; width:100%; }}
        .metric-card .kv {{ display:flex; align-items:baseline; justify-content:space-between; gap:12px; padding:5px 0; border-bottom:1px dashed #e3eaee; }}
        .metric-card .kv:last-child {{ border-bottom:0; }}
        .metric-card .kv span {{ display:inline; min-height:0; margin:0; color:var(--muted); font-size:11px; font-weight:600; text-align:left; text-transform:none; letter-spacing:0; }}
        .metric-card .kv b {{ display:inline; font-size:17px; font-weight:800; letter-spacing:-.02em; color:var(--ink); text-align:right; white-space:nowrap; }}
        .metric-card .kv--good b {{ color:#116832; }}
        .metric-card .kv--warn b {{ color:#a12020; }}
        .metric-card .kv--good span:after {{ content:" ● conforme"; color:#116832; font-weight:800; }}
        .metric-card .kv--warn span:after {{ content:" ● fora do limite"; color:#a12020; font-weight:800; }}
        .integrated-overview {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:18px; margin:12px 0; box-shadow:0 5px 16px rgba(16,34,56,.05); }}
        .integrated-controls {{ display:grid; grid-template-columns:repeat(7,minmax(130px,1fr)); gap:10px; align-items:end; margin:0 0 12px; padding:14px; background:#fff; border:1px solid var(--line); border-radius:10px; box-shadow:0 4px 14px rgba(16,34,56,.04); }}
        .integrated-controls label {{ color:#5b6d80; font-size:9px; font-weight:850; letter-spacing:.05em; text-transform:uppercase; }}
        .integrated-controls select,.integrated-controls input {{ min-width:0; width:100%; height:38px; padding:7px 9px; border:1px solid #cfd9e4; border-radius:6px; background:#fff; color:var(--navy); }}
        .integrated-note {{ display:flex; align-items:center; gap:12px; margin:-2px 0 10px; color:#6b7b8c; font-size:11px; }}
        .integrated-note b {{ color:var(--blue); font-size:10px; letter-spacing:.06em; text-transform:uppercase; }}
        .overview-heading,.comparison-strip-title {{ display:flex; align-items:flex-end; justify-content:space-between; gap:16px; margin-bottom:14px; }}
        .overview-heading h3,.comparison-strip-title h3 {{ margin:3px 0 0; font-size:18px; }}
        .eyebrow-dark {{ color:var(--teal); font-size:10px; font-weight:850; letter-spacing:.11em; text-transform:uppercase; }}
        .overview-meta {{ color:var(--muted); font-size:12px; text-align:right; }}
        .overview-kpis {{ display:grid; grid-template-columns:repeat(6,minmax(125px,1fr)); gap:10px; }}
        .overview-kpi {{ min-width:0; padding:12px; border:1px solid #e0e7ee; border-radius:9px; background:#f9fbfd; }}
        .overview-kpi > span {{ display:block; min-height:28px; color:var(--muted); font-size:10px; font-weight:800; letter-spacing:.045em; text-transform:uppercase; }}
        .overview-kpi > b {{ display:block; color:var(--navy); font-size:22px; letter-spacing:-.03em; overflow-wrap:anywhere; }}
        .overview-kpi > b small {{ font-size:11px; letter-spacing:0; }}
        .overview-kpi > em {{ display:block; margin-top:5px; color:var(--muted); font-size:9px; font-style:normal; }}
        .overview-kpi.warn {{ border:2px solid #dc2626; background:#fff1f2; box-shadow:0 0 0 3px rgba(220,38,38,.12); }} .overview-kpi.warn > span,.overview-kpi.warn > b,.overview-kpi.warn > em {{ color:#991b1b; }}
        .metric-detail {{ grid-column:1/-1; margin-top:10px; border-top:1px solid var(--line); padding-top:8px; }} .metric-detail summary {{ cursor:pointer; color:var(--teal); font-weight:800; font-size:12px; }}
        .mode-line {{ display:flex; flex-wrap:wrap; gap:10px 24px; margin-top:12px; padding-top:11px; border-top:1px solid #e6ecf1; color:#43566a; font-size:12px; }}
        .measurement-flow {{ display:grid; grid-template-columns:minmax(160px,1fr) 28px minmax(160px,1fr) 28px minmax(180px,1.1fr) minmax(180px,1fr); gap:10px; align-items:stretch; margin:14px 0; }}
        .flow-node {{ padding:15px; background:#fff; border:1px solid var(--line); border-top:4px solid var(--teal); border-radius:10px; box-shadow:0 4px 12px rgba(16,34,56,.04); }}
        .flow-node > span,.pair-card > span {{ display:block; color:var(--muted); font-size:10px; font-weight:850; letter-spacing:.055em; text-transform:uppercase; }}
        .flow-node > b {{ display:block; margin:8px 0 4px; color:var(--navy); font-size:18px; }}
        .flow-node > small {{ color:var(--muted); line-height:1.35; }}
        #leadFlow {{ display:none; }}
        .flow-topside {{ border-top-color:var(--blue); }} .flow-balance {{ border-top-color:var(--green); }} .flow-separator {{ border-top-color:var(--amber); }}
        .flow-link {{ align-self:center; height:2px; background:#a8bac9; }}
        .comparison-strip {{ background:#fff; border:1px solid var(--line); border-radius:12px; padding:16px; margin:14px 0; }}
        .comparison-strip-title > small {{ max-width:520px; color:var(--muted); text-align:right; }}
        .pair-card-grid {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(230px,1fr)); gap:10px; }}
        .pair-card {{ padding:13px; border:1px solid var(--line); border-left:5px solid var(--green); border-radius:9px; background:#fbfdfc; }}
        .pair-card.pair-bad {{ border-left-color:var(--red); background:#fffafa; }} .pair-card.pair-bad:after {{ content:'NÃO CONFORME'; display:block; margin-top:8px; color:#b91c1c; font-size:10px; font-weight:850; letter-spacing:.08em; }}
        .pair-card.pair-warn {{ border-left-color:var(--amber); background:#fffaf0; }} .pair-card.pair-warn:after {{ content:'NÃO CLASSIFICÁVEL'; display:block; margin-top:8px; color:#a15c00; font-size:10px; font-weight:850; letter-spacing:.08em; }}
        .pair-card > div {{ display:flex; align-items:baseline; justify-content:space-between; gap:10px; padding:8px 0 4px; border-bottom:1px dashed #e4eaef; }}
        .pair-card > div b {{ font-size:15px; }} .pair-card > div small,.pair-card > em {{ color:var(--muted); font-size:10px; font-style:normal; }}
        .pair-card > em {{ display:block; margin-top:8px; }}
        .lineage-grid {{ display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:10px; margin:14px 0 22px; }}
        .lineage-stage {{ position:relative; display:grid; grid-template-rows:42px 1fr; gap:10px; align-items:start; min-height:190px; padding:15px; background:#fff; border:1px solid var(--line); border-top:4px solid var(--blue); border-radius:10px; box-shadow:0 4px 12px rgba(16,34,56,.04); }}
        .lineage-stage:not(:last-child):after {{ content:""; position:absolute; right:-11px; top:50%; width:11px; height:2px; background:#a9bac9; z-index:2; }}
        .lineage-stage > span {{ display:grid; place-items:center; width:40px; height:40px; border-radius:8px; background:var(--navy); color:#fff; font-weight:850; }}
        .lineage-stage h3 {{ margin:0 0 9px; }} .lineage-stage b {{ display:block; margin-bottom:7px; color:var(--teal); }} .lineage-stage small {{ color:var(--muted); line-height:1.45; }}
        @media(max-width:1400px) {{ .lineage-grid {{ grid-template-columns:repeat(3,minmax(0,1fr)); }} .lineage-stage:not(:last-child):after {{ display:none; }} }}
        @media(max-width:1100px) {{ main {{ padding:92px 16px 30px; }} .metric-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .chart-grid {{ grid-template-columns:1fr; }} .alarm-category-row {{ grid-template-columns:1fr; }} .overview-kpis {{ grid-template-columns:repeat(3,minmax(0,1fr)); }} .integrated-controls {{ grid-template-columns:repeat(3,minmax(0,1fr)); }} .measurement-flow {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .flow-link {{ display:none; }} .lineage-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} }}
        @media(max-width:720px) {{
            header {{ position:fixed; height:auto; min-height:72px; padding:12px 14px; }} header h1 {{ font-size:18px; }} header .meta {{ display:block; font-size:10px; line-height:1.45; }} .header-status {{ margin-left:0; text-align:left; width:100%; font-size:10px; }} .brand-row > div:nth-child(2) {{ flex-basis:calc(100% - 52px); }} .brand-mark {{ width:38px; height:38px; flex:0 0 38px; }} .brand-mark img {{ width:36px; height:36px; }} main {{ padding-top:calc(var(--header-height,160px) + 10px) !important; }}
            .tab-nav {{ position:sticky; left:auto; top:0; bottom:auto; width:auto; height:auto; padding:9px 12px; border-right:0; border-bottom:1px solid var(--line); }} .tab-nav:before,.tab-button {{ display:none; }} .mobile-tab-select {{ display:block; }}
            main {{ margin:0; padding:12px 10px 28px; }} .metric-grid,.formula-grid {{ grid-template-columns:1fr; gap:9px; }} .metric-card {{ min-height:104px; }} .metric-card span,.metric-card small {{ min-height:0; }}
            .tab-panel > h2 {{ font-size:17px; margin:14px 0 9px; padding:10px 11px; }}
            .controls {{ grid-template-columns:1fr; }} .filter-panel {{ padding:0 11px; }}
            .table-toolbar {{ align-items:stretch; flex-direction:column; }} .table-search {{ width:100%; max-width:none; min-width:0; }}
            .table-wrap:not(.calendar-wrap) .data-table.enhanced-table thead {{ display:none; }}
            .table-wrap:not(.calendar-wrap) .data-table.enhanced-table,.table-wrap:not(.calendar-wrap) .data-table.enhanced-table tbody,.table-wrap:not(.calendar-wrap) .data-table.enhanced-table tr,.table-wrap:not(.calendar-wrap) .data-table.enhanced-table td {{ display:block; width:100%; }}
            .table-wrap:not(.calendar-wrap) .data-table.enhanced-table tr {{ margin:9px; width:calc(100% - 18px); border:1px solid var(--line); border-radius:9px; background:#fff; overflow:hidden; }}
            .table-wrap:not(.calendar-wrap) .data-table.enhanced-table td {{ display:grid; grid-template-columns:minmax(105px,38%) 1fr; gap:9px; text-align:right; white-space:normal; max-width:none; padding:8px 10px; }}
            .table-wrap:not(.calendar-wrap) .data-table.enhanced-table td:before {{ content:attr(data-label); text-align:left; color:#64748b; font-size:9px; font-weight:800; text-transform:uppercase; letter-spacing:.04em; }}
            .calendar-table {{ min-width:760px; }} .calendar-table td,.calendar-table th {{ min-width:102px; }}
            .mini-row {{ grid-template-columns:1fr; gap:4px; }} .chart-scroll:after {{ display:none; }}
            .overview-heading,.comparison-strip-title {{ align-items:flex-start; flex-direction:column; }} .overview-meta,.comparison-strip-title > small {{ text-align:left; }}
            .overview-kpis,.integrated-controls,.measurement-flow,.lineage-grid {{ grid-template-columns:1fr; }} .mode-line {{ flex-direction:column; gap:8px; }}
        }}
        @media print {{ .table-toolbar,.mobile-tab-select {{ display:none !important; }} .data-table td {{ max-width:none; white-space:normal; }} }}
  </style>
</head>
<body data-locale="pt-BR" data-date-time-format="dd/mm/yyyy hh:mm" data-decimal-separator=",">
    <script id="source-manifest" type="application/json">{source_manifest_json}</script>
  <header>
    <div class="brand-row">
    <div class="brand-mark" aria-label="Logo do sistema multifásico">{logo_markup}</div>
    <div><div class="eyebrow">Metering Management | Bacalhau</div><h1>Relatório de Monitoramento Multifásico</h1><div class="meta">{html.escape(header_summary)}<br>Janela: {html.escape(', '.join(map(str, selected_days)))}<br>Processados: {html.escape(', '.join(map(str, target_days)))}</div></div>
    </div>
    <div class="header-status">Base Única carregada</div>
  </header>
  <main>
    <nav class="tab-nav" id="dashboardTabs" role="tablist" aria-label="Seções do dashboard">
      <button class="tab-button" data-tab="executivo" role="tab">Visão integrada</button>
      <button class="tab-button" data-tab="comparacoes" role="tab">Comparações</button>
      <button class="tab-button" data-tab="separador" role="tab">Separador</button>
      <button class="tab-button" data-tab="cobertura" role="tab">Cobertura</button>
      <button class="tab-button" data-tab="pi" role="tab">PI Vision</button>
      <button class="tab-button" data-tab="alarmes" role="tab">Alarmes/eventos</button>
      <button class="tab-button" data-tab="detalhes" role="tab">Detalhes</button>
      <button class="tab-button" data-tab="auditoria" role="tab">Auditoria e rastreabilidade</button>
            <button class="tab-button" data-tab="validacao" role="tab">Validação final</button>
      <button class="tab-button" data-tab="xml042" role="tab">XML 042 (ANP)</button>
    </nav>
    <section class="tab-panel" data-tab-group="executivo"><h2>Visão Integrada do MPFM</h2>{leadership_panel}<h2>Resumo do último dia processado</h2>{executive_panel}{_alarm_category_visual(alarm_event_rows)}</section>
    <section class="tab-panel" data-tab-group="cobertura"><h2>Cobertura e recebimento dos dados</h2><div class="metric-grid"><div class="metric-card accent"><span>Linhas recebidas</span><b>{len(df_out)}</b><small>PDF/TXT na janela</small></div><div class="metric-card"><span>Dias processados</span><b>{len(target_days)}</b><small>janela publicada</small></div><div class="metric-card {'warn' if issue_count else 'good'}"><span>Pendências de origem</span><b>{issue_count}</b><small>pré-validação de arquivos</small></div><div class="metric-card"><span>Alertas analíticos</span><b>{len(analysis['anomalies']) + len(analysis['pressure_flags'])}</b><small>triagem para investigação</small></div></div><h3>Calendário de extração e cobertura</h3>{calendar_panel}<h3>Disponibilidade esperada × recebida</h3>{_availability_visual(analysis['preflight'])}<div class="table-wrap">{_html_table(analysis['preflight'])}</div><h3>Contagem por fonte, ponto e granularidade</h3>{_summary_visual(analysis['summary'])}<div class="table-wrap">{_html_table(analysis['summary'])}</div><h3>Cadastro de pontos adicionais</h3>{configured_points_panel}</section>
    <section class="tab-panel" data-tab-group="pi"><h2>Extração PI Vision</h2>{_pi_quality_panel(master_path, target_days, pi_master_df)}<div class="info-box"><b>Telas configuradas:</b><br>Metering Monitor — https://pivision.equinor.com/PIVision/#/Displays/54854/Metering-Monitor<br>Metering Daily Control — https://pivision.equinor.com/PIVision/#/Displays/56466/Metering-Daily-Control</div></section>
    <section class="tab-panel" data-tab-group="alarmes"><h2>Resumo de alarmes e eventos FCS320</h2>{_alarm_history_visual(alarm_event_rows)}<div class="table-wrap">{_html_table(alarm_event_summary)}</div></section>
    <section class="tab-panel" data-tab-group="alarmes"><h2>Alarmes/eventos que devem ser vistos</h2><div class="info-box"><b>Como interpretar:</b> a coluna “Interpretação” traduz a categoria técnica para orientar a primeira verificação. Ela não substitui a análise do horário, estado do alarme, condição de processo e impacto no balanço.</div><div class="table-wrap">{_html_table(alarm_flag_rows)}</div></section>
    <section class="tab-panel" data-tab-group="comparacoes"><h2>Condições de contorno da medição</h2>{boundary_panel}</section>
    <section class="tab-panel" data-tab-group="comparacoes"><h2>Comparação CEP — Topside × Subsea</h2>{cep_panel}</section>
    <section class="tab-panel" data-tab-group="comparacoes"><h2>Exploração detalhada Topside × Subsea</h2>{comparison_panel}</section>
    <section class="tab-panel" data-tab-group="separador"><h2>Dados próprios do Separador de Testes</h2>{separator_panel}<h2>Consulta MPFM × Separador sob demanda</h2>{separator_comparison_panel}</section>
    <section class="tab-panel" data-tab-group="detalhes"><h2>Perfil horário consolidado</h2>{hourly_profile_panel}<h2>Estabilidade e spikes por MPFM</h2>{stability_panel}<h2>Dados a verificar: spikes, zerados e ausentes</h2><div class="table-wrap">{_html_table(analysis['anomalies'])}</div><h2>Comparação contra os 4 dias anteriores</h2><div class="table-wrap">{_comparison_delta_visual(analysis['comparisons'])}</div><h2>Metodologia e fórmulas</h2>{_methodology_panel()}<h2>Pressão acima de 490 barg e gás medido</h2><div class="table-wrap">{_html_table(analysis['pressure_flags'])}</div></section>
    <section class="tab-panel" data-tab-group="auditoria"><h2>Fontes ingeridas e alinhamento com a Base Única Total</h2>{sources_panel}<h2>Cadeia do dado e rastreabilidade</h2>{lineage_panel}</section>
    <section class="tab-panel" data-tab-group="validacao"><h2>Critérios de aceitação e conferência Excel → HTML</h2>{acceptance_panel}</section>
    <section class="tab-panel" data-tab-group="xml042"><h2>XML 042 (ANP) — geração sob demanda</h2>{xml042_panel}</section>
  </main>
  <script>
    (() => {{
      const header = document.querySelector('header');
      const syncLayout = () => {{
        if (!header) return;
        const height = Math.ceil(header.getBoundingClientRect().height);
        document.documentElement.style.setProperty('--header-height', `${{height}}px`);
      }};
      syncLayout();
      window.addEventListener('resize', syncLayout, {{ passive: true }});
      if (window.ResizeObserver && header) new ResizeObserver(syncLayout).observe(header);
      window.addEventListener('scroll', () => {{
        const compact = window.scrollY > 80;
        if (header.classList.contains('header--compact') !== compact) {{
          header.classList.toggle('header--compact', compact);
          requestAnimationFrame(syncLayout);
        }}
      }}, {{ passive: true }});
    }})();
    document.querySelectorAll('.controls').forEach((controls, index) => {{
        if (controls.classList.contains('integrated-controls')) return;
        const details = document.createElement('details');
        details.className = 'filter-panel';
        const summary = document.createElement('summary');
        summary.textContent = index === 0 ? 'Filtros e seleção de período' : 'Filtros desta seção';
        controls.parentNode.insertBefore(details, controls);
        details.appendChild(summary);
        details.appendChild(controls);
    }});
  (() => {{
    const buttons = [...document.querySelectorAll('#dashboardTabs .tab-button')];
    const panels = [...document.querySelectorAll('main .tab-panel')];
        const nav = document.getElementById('dashboardTabs');
        const mobileSelect = document.createElement('select');
        mobileSelect.className = 'mobile-tab-select';
        mobileSelect.setAttribute('aria-label', 'Selecionar seção do dashboard');
        buttons.forEach(button => {{
            button.id = `dashboard-tab-${{button.dataset.tab}}`;
            if (!mobileSelect.querySelector(`option[value="${{button.dataset.tab}}"]`)) mobileSelect.add(new Option(button.textContent, button.dataset.tab));
        }});
        panels.forEach((panel, index) => {{ panel.id = `dashboard-panel-${{panel.dataset.tabGroup}}-${{index}}`; panel.setAttribute('role','tabpanel'); }});
        nav.appendChild(mobileSelect);
    function activate(group) {{
      buttons.forEach(button => {{ const active = button.dataset.tab === group; button.classList.toggle('active', active); button.setAttribute('aria-selected', active ? 'true' : 'false'); }});
      panels.forEach(panel => {{ panel.hidden = panel.dataset.tabGroup !== group; }});
            mobileSelect.value = group;
      requestAnimationFrame(() => {{
        document.querySelectorAll('.chart-scroll[data-zoom-ready="1"]').forEach(wrap => {{
          if (wrap.getBoundingClientRect().width > 0) wrap.dataset.zoomBaseHeight = '';
        }});
        if (typeof refreshChartZoom === 'function') refreshChartZoom();
      }});
      window.scrollTo({{top: 0, behavior: 'smooth'}});
    }}
    buttons.forEach(button => button.addEventListener('click', () => activate(button.dataset.tab)));
        mobileSelect.addEventListener('change', () => activate(mobileSelect.value));
    activate('executivo');
  }})();
    (() => {{
        const normalize = value => String(value ?? '').trim().toLocaleLowerCase('pt-BR');
        document.querySelectorAll('.table-wrap .data-table').forEach(table => {{
            if (table.classList.contains('calendar-table')) return;
            table.classList.add('enhanced-table');
            const headers = [...table.querySelectorAll('thead th')];
            const body = table.tBodies[0];
            if (!body || !headers.length) return;
            [...body.rows].forEach(row => [...row.cells].forEach((cell, index) => {{
                cell.dataset.label = headers[index]?.textContent?.trim() || `Campo ${{index + 1}}`;
                const text = cell.textContent.trim();
                const upper = text.toUpperCase();
                if (/^(OK|CONFORME|NORMAL|ATIVA|INGESTÃO ATIVA)$/.test(upper)) cell.innerHTML = `<span class="status-chip status-ok">${{text}}</span>`;
                else if (/(ATENÇÃO|PARCIAL|PENDENTE|REVISAR|SUSPEITO)/.test(upper)) cell.innerHTML = `<span class="status-chip status-warn">${{text}}</span>`;
                else if (/(FALHA|CRÍTICO|FORA DO LIMITE|AUSENTE|NÃO DISPONÍVEL)/.test(upper)) cell.innerHTML = `<span class="status-chip status-bad">${{text}}</span>`;
                else if (/(HOURLY|DAILY|MPFM|SEP|PI VISION)/.test(upper) && text.length < 36) cell.innerHTML = `<span class="status-chip status-info">${{text}}</span>`;
                const pct = Number(text.replace('%','').replace(/\./g,'').replace(',','.'));
                if (text.includes('%') && Number.isFinite(pct)) {{ cell.classList.add('cell-bar'); cell.style.setProperty('--cell-pct', `${{Math.max(0,Math.min(100,Math.abs(pct)))}}%`); }}
            }}));
            const wrap = table.closest('.table-wrap');
            const toolbar = document.createElement('div'); toolbar.className = 'table-toolbar';
            const search = document.createElement('input'); search.className = 'table-search'; search.type = 'search'; search.placeholder = 'Pesquisar nesta tabela…'; search.setAttribute('aria-label','Pesquisar nesta tabela');
            const count = document.createElement('span'); count.className = 'table-count';
            toolbar.append(search, count); wrap.insertBefore(toolbar, wrap.firstChild);
            const refreshCount = () => {{ const visible = [...body.rows].filter(row => !row.hidden).length; count.textContent = `${{visible}} de ${{body.rows.length}} registros`; }};
            search.addEventListener('input', () => {{ const query = normalize(search.value); [...body.rows].forEach(row => row.hidden = query && !normalize(row.textContent).includes(query)); refreshCount(); }});
            headers.forEach((header, column) => {{ header.classList.add('sortable'); header.tabIndex = 0; let ascending = true; const sort = () => {{ const rows = [...body.rows]; rows.sort((a,b) => normalize(a.cells[column]?.textContent).localeCompare(normalize(b.cells[column]?.textContent), 'pt-BR', {{numeric:true}}) * (ascending ? 1 : -1)); rows.forEach(row => body.appendChild(row)); ascending = !ascending; }}; header.addEventListener('click', sort); header.addEventListener('keydown', event => {{ if (event.key === 'Enter' || event.key === ' ') {{ event.preventDefault(); sort(); }} }}); }});
            refreshCount();
        }});
    }})();
    (() => {{
        const localizeText = root => {{
            const walker = document.createTreeWalker(root, NodeFilter.SHOW_TEXT);
            const nodes = [];
            while (walker.nextNode()) nodes.push(walker.currentNode);
            nodes.forEach(node => {{
                if (['SCRIPT','STYLE','CODE'].includes(node.parentElement?.tagName)) return;
                node.nodeValue = node.nodeValue
                    .replace(/(\d{{4}})-(\d{{2}})-(\d{{2}})[ T]H?(\d{{1,2}})(?::(\d{{2}}))?/g, (_,y,m,d,h,min) => `${{d}}/${{m}}/${{y}} ${{String(h).padStart(2,'0')}}:${{min || '00'}}`)
                    .replace(/(\d{{4}})-(\d{{2}})-(\d{{2}})/g, '$3/$2/$1');
            }});
        }};
        localizeText(document.body);
        new MutationObserver(mutations => mutations.forEach(mutation => mutation.addedNodes.forEach(node => localizeText(node)))).observe(document.body, {{childList:true, subtree:true}});
    }})();
  </script>
  <script>document.querySelectorAll('table.data-table').forEach(table=>{{const headers=[...table.querySelectorAll('thead th')];const volumeIdx=headers.map((h,i)=>/volume (óleo|gás|água) 20/i.test(h.textContent)?i:-1).filter(i=>i>=0);if(volumeIdx.length===3){{table.classList.add('volume-source-table');headers.forEach((h,i)=>{{if(volumeIdx.includes(i))h.classList.add('volume-col')}});table.querySelectorAll('tbody tr').forEach(row=>row.querySelectorAll('td').forEach((cell,i)=>{{if(volumeIdx.includes(i))cell.classList.add('volume-col')}}));}}}});</script>
  <script>(()=>{{document.querySelectorAll('table.volume-source-table').forEach(table=>table.querySelectorAll('tbody tr').forEach(row=>{{if(/topside/i.test(row.textContent||''))row.style.setProperty('display','none','important');}}));}})();</script>
  <script>if (typeof installChartZoom === 'function') installChartZoom();</script>
</body>
</html>"""
    dashboard_path.write_text(html_doc, encoding="utf-8")
    print(f"[TIME] write_dashboard total: {time.perf_counter() - dashboard_started:.1f}s", flush=True)
    return dashboard_path


def write_fallback_dashboard(
    dashboard_path: Path,
    exc: Exception,
    df_out: pd.DataFrame,
    target_days: list,
    selected_days: list,
) -> Path:
    """Gera um HTML mínimo quando algum painel rico falha.

    A automação não deve terminar sem dashboard: este arquivo mantém a
    rastreabilidade do erro e uma amostra dos dados para decisão/diagnóstico.
    """
    dashboard_path.parent.mkdir(parents=True, exist_ok=True)
    sample = pd.DataFrame()
    if df_out is not None and not df_out.empty:
        visible_columns = [col for col in [
            "ProductionDate", "Granularity", "Origin", "Bank", "Tipo", "Tag",
            "MPFM corr HC (t)", "MPFM corr Total (t)", "SEP HC (t)", "SEP Total (t)",
            "Desvio HC (%)", "Desvio Total (%)",
        ] if col in df_out.columns]
        sample = df_out[visible_columns].head(200).copy() if visible_columns else df_out.head(200).copy()
    sample_html = _html_table(sample.to_dict("records"), max_rows=200) if not sample.empty else "<p>Sem amostra disponível.</p>"
    logo_data_uri = _dashboard_logo_data_uri()
    logo_markup = f'<img src="{logo_data_uri}" alt="Logo do sistema multifásico" class="inline-logo">' if logo_data_uri else ""
    html_doc = f"""<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Relatório de Monitoramento Multifásico - Diagnóstico</title>
  <style>
    body {{ font-family: Segoe UI, Arial, sans-serif; margin: 0; background: #f5f8fb; color: #102238; }}
    header {{ background: #071e41; color: white; padding: 22px 30px; }}
    header img {{ background: rgba(255,255,255,.1); border-radius: 8px; }}
    main {{ padding: 24px 30px; }}
    .box {{ background: white; border: 1px solid #d9e1e8; border-radius: 8px; padding: 16px; margin-bottom: 16px; }}
    .warn {{ border-left: 5px solid #e98a00; }}
    .data-table {{ border-collapse: collapse; width: 100%; font-size: 12px; }}
    .data-table th, .data-table td {{ border-bottom: 1px solid #edf1f4; padding: 8px 9px; text-align: left; white-space: nowrap; }}
    .data-table th {{ background: #f5f8fb; color: #53667a; }}
    .table-wrap {{ max-height: 500px; overflow: auto; border: 1px solid #d9e1e8; border-radius: 8px; }}
    code {{ background: #f1f5f9; padding: 2px 5px; border-radius: 4px; }}
  </style>
</head>
<body>
    <header>{logo_markup}<h1 class="inline-title">Relatório de Monitoramento Multifásico</h1><p>Dashboard de contingência gerado em {html.escape(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))}</p></header>
  <main>
    <div class="box warn"><b>Dashboard rico não foi concluído.</b><p>A automação gravou este HTML mínimo para preservar a saída final. Erro: <code>{html.escape(str(exc))}</code></p></div>
    <div class="box"><b>Janela solicitada:</b> {html.escape(", ".join(str(day) for day in selected_days) or "não informada")}<br><b>Dias processados:</b> {html.escape(", ".join(str(day) for day in target_days) or "não informado")}</div>
    <div class="box"><h2>Amostra da Base Única</h2><div class="table-wrap">{sample_html}</div></div>
  </main>
</body>
</html>"""
    dashboard_path.write_text(html_doc, encoding="utf-8")
    return dashboard_path


def publish_dashboard(
    dashboard_path: Path,
    df_out: pd.DataFrame,
    target_days: list,
    selected_days: list,
    preflight_rows: list,
    master_path: Path | None = None,
    df_alarm_events: pd.DataFrame | None = None,
    aligned_bank: str = SEP_ALIGNED_BANK,
    preloaded_master_df: pd.DataFrame | None = None,
    preloaded_pi_df: pd.DataFrame | None = None,
    preloaded_comparativo_df: pd.DataFrame | None = None,
    preloaded_alarm_events: pd.DataFrame | None = None,
    source_workbook_path: Path | None = None,
    preloaded_source_frames: dict[str, pd.DataFrame] | None = None,
) -> Path:
    """Publica um dashboard rico e preserva um HTML de diagnóstico em caso de falha."""
    try:
        # As rotas normais passam datas ISO como texto; republicações diretas
        # da Base Única também podem fornecer datetime/date. Normalizar aqui
        # evita que filtros como pressão, cobertura e séries diárias comparem
        # strings com objetos date e acabem aparentando ausência de dados.
        def _dashboard_iso_day(value):
            if isinstance(value, (datetime, pd.Timestamp)):
                return value.strftime("%Y-%m-%d")
            return str(value)[:10]
        target_days = [_dashboard_iso_day(value) for value in target_days]
        selected_days = [_dashboard_iso_day(value) for value in selected_days]
        source_frames = preloaded_source_frames
        if source_frames is None:
            source_frames = _load_dashboard_source_frames(source_workbook_path)
        if not source_frames:
            fallback_sheet = MASTER_SHEET_NAME if source_workbook_path and "TOTAL" in source_workbook_path.stem.upper() else "BASE_UNICA_STANDALONE"
            source_frames = {fallback_sheet: df_out.copy()}
        write_dashboard(
            dashboard_path,
            df_out,
            target_days,
            selected_days,
            preflight_rows,
            master_path,
            df_alarm_events,
            aligned_bank=aligned_bank,
            preloaded_master_df=preloaded_master_df,
            preloaded_pi_df=preloaded_pi_df,
            preloaded_comparativo_df=preloaded_comparativo_df,
            preloaded_alarm_events=preloaded_alarm_events,
            source_workbook_path=source_workbook_path,
            preloaded_source_frames=source_frames,
        )
        validation = validate_dashboard_delivery(dashboard_path, source_frames, source_workbook_path)
        print(f"[INFO] Dashboard gerado: {dashboard_path}")
        print(f"[VALIDAÇÃO] {validation['status']}: {validation['reportPath']}")
        if validation["status"] != "APROVADO":
            failed = ", ".join(name for name, passed in validation["rules"].items() if not passed)
            print(f"[WARN] HTML não aprovado para entrega. Critérios reprovados: {failed}")
    except Exception as exc:
        print(f"[WARN] Falha ao gerar dashboard rico ({dashboard_path}): {exc}")
        traceback.print_exc()
        try:
            write_fallback_dashboard(dashboard_path, exc, df_out, target_days, selected_days)
            print(f"[INFO] HTML de diagnóstico gerado: {dashboard_path}")
        except Exception as fallback_exc:
            print(f"[WARN] Também não foi possível gerar o HTML de diagnóstico ({dashboard_path}): {fallback_exc}")
            traceback.print_exc()
    return dashboard_path


# ═════════════════════════════════════════════════════════════════════════
# CLI / prompts interativos
# ═════════════════════════════════════════════════════════════════════════

def _resolve_config():
    parser = argparse.ArgumentParser(description="Gerador Base_Unica standalone (MPFM + SEP)")
    parser.add_argument("--operation-mode", choices=sorted(OPERATION_MODES), default="", help="Modo operacional: 1=todas, 2=PI+Base, 3=Email+Base, 4=Email+PI, 5=Email, 6=PI, 7=Base")
    parser.add_argument("--mpfm-root", default="", help="Pasta raiz dos PDFs MPFM (3.1.x por banco)")
    parser.add_argument("--sep-root", default="", help="Pasta raiz dos Daily Reports (FC13/FC14/FC17)")
    parser.add_argument("--output", default=OUTPUT_PATH, help="Caminho do .xlsx de saída")
    parser.add_argument("--master-output", default=MASTER_OUTPUT_PATH, help="Caminho do .xlsx incremental total (padrão: BASE_UNICA_TOTAL.xlsx ao lado do script)")
    parser.add_argument("--no-master", action="store_true", help="Não atualizar a Base_Unica total incremental nesta execução")
    parser.add_argument("--force-refresh", action="store_true", help="Reprocessa todos os dias selecionados, mesmo se já existirem na Base_Unica total")
    parser.add_argument("--ask-period", action="store_true", help="Pergunta interativamente se o usuário deseja informar uma janela de datas")
    parser.add_argument("--approve-missing", action="store_true", help="Continua a execução mesmo se a pré-validação encontrar dados faltantes")
    parser.add_argument("--preflight-only", action="store_true", help="Somente valida a disponibilidade da janela e encerra sem gerar Excel")
    parser.add_argument("--no-dashboard", action="store_true", help="Não gera dashboard HTML ao final da análise")
    parser.add_argument("--dashboard-output", default="", help="Caminho do dashboard HTML (padrão: ao lado do Excel de saída)")
    parser.add_argument("--standalone-dashboard-output", default="", help="Caminho do HTML da janela processada (STANDALONE)")
    parser.add_argument("--total-dashboard-output", default="", help="Caminho do HTML completo baseado na Base_Unica total")
    parser.add_argument("--no-open-dashboard", action="store_true", help="Gera o dashboard, mas não tenta abrir no navegador")
    parser.add_argument("--no-pi", action="store_true", help="Não executar a coleta PI Vision antes dos PDFs/TXTs")
    parser.add_argument("--pi-root", default=PI_COLLECTOR_ROOT, help="Pasta do coletor PI Vision")
    parser.add_argument("--pi-config", default=PI_COLLECTOR_CONFIG, help="Arquivo config do coletor PI Vision")
    parser.add_argument("--pi-output", default=PI_EXTRACT_OUTPUT, help="Arquivo normalizado gerado pelo coletor PI Vision")
    parser.add_argument("--pi-period-output", default=PI_PERIOD_OUTPUT, help="JSON do período aplicado pelo coletor PI Vision")
    parser.add_argument("--continue-without-pi", action="store_true", help="Continua o processamento se a coleta PI falhar")
    parser.add_argument("--pi-retries", type=int, default=PI_RETRIES, help="Tentativas adicionais para cada dia PI quando a captura vier incompleta")
    parser.add_argument("--pi-from-excel", default="", help="Lê aba PI_EXTRACT de um Excel gerado anteriormente, evitando nova coleta (ex: BASE_UNICA_STANDALONE_20260816_221423.xlsx)")
    parser.add_argument("--continue-without-email", action="store_true", help="Continua o processamento se a automação de e-mail falhar")
    parser.add_argument("--days", type=int, default=DAYS_COUNT, help="Quantidade de dias mais recentes a exportar")
    parser.add_argument("--aligned-bank", default=SEP_ALIGNED_BANK, help=argparse.SUPPRESS)
    parser.add_argument("--months-lookback", type=int, default=MONTHS_LOOKBACK, help="Máximo de meses pesquisados")
    parser.add_argument("--workers", type=int, default=PDF_WORKERS, help="Processos paralelos para leitura de PDF")
    parser.add_argument("--date-from", default="", help="Início do intervalo (DD/MM/AAAA ou AAAA-MM-DD)")
    parser.add_argument("--date-to", default="", help="Fim do intervalo (DD/MM/AAAA ou AAAA-MM-DD)")
    args = parser.parse_args()

    operation_mode = args.operation_mode.strip()
    if not operation_mode:
        operation_mode = _prompt_operation_mode(default="1")
    operation = _operation_flags(operation_mode)
    run_email = bool(operation["email"])
    run_pi = bool(operation["pi"]) and not args.no_pi
    run_base = bool(operation["base"])
    print(f"\n[OK] Caminho selecionado: {operation_mode} - {operation['label']}")

    mpfm_root = _clean_path_input(args.mpfm_root)
    if not mpfm_root and (run_email or run_base):
        mpfm_root = _prompt_configured_path(
            "Pasta raiz dos PDFs MPFM (3.1.x por banco)", MPFM_ROOT
        )
    sep_root = _clean_path_input(args.sep_root)
    if not sep_root and run_base:
        sep_root = _prompt_configured_path(
            "Pasta raiz dos Daily Reports / TXT do Separador (FC13/FC14/FC17)", SEP_ROOT
        )

    output_path = args.output.strip()
    if not output_path:
        output_path = str(EXCEL_OUTPUT_DIR / f"BASE_UNICA_STANDALONE_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

    update_master = bool(UPDATE_MASTER) and not args.no_master and run_base
    master_output_path = args.master_output.strip()
    if update_master and not master_output_path:
        master_output_path = str(EXCEL_OUTPUT_DIR / "BASE_UNICA_TOTAL.xlsx")

    date_from_arg = args.date_from.strip()
    date_to_arg = args.date_to.strip()
    if args.ask_period and not date_from_arg and not date_to_arg:
        must_ask_pi_window = run_pi and not run_base
        if must_ask_pi_window or _prompt_yes_no("Deseja informar uma janela específica de análise?", default=False):
            if must_ask_pi_window:
                print("Para executar PI sem Base_Unica, informe a janela que será coletada no PI.")
            date_from_arg = input("Data inicial (DD/MM/AAAA ou AAAA-MM-DD): ").strip()
            date_to_arg = input("Data final (DD/MM/AAAA ou AAAA-MM-DD): ").strip()

    try:
        date_from = _parse_date_argument(date_from_arg, "--date-from")
        date_to = _parse_date_argument(date_to_arg, "--date-to")
    except ValueError as exc:
        parser.error(str(exc))
    if date_from and not date_to:
        date_to = datetime.now() - timedelta(days=1)
    if date_to and not date_from:
        parser.error("--date-to exige --date-from.")
    if date_from and date_to and date_from > date_to:
        parser.error("--date-from não pode ser posterior a --date-to.")
    if run_pi and not run_base and not (date_from and date_to):
        parser.error("Modos com PI sem Base_Unica exigem uma janela: informe --date-from/--date-to ou use --ask-period.")

    return (
        operation_mode, run_email, run_pi, run_base,
        Path(mpfm_root) if mpfm_root else Path(), Path(sep_root) if sep_root else Path(), Path(output_path), args.days,
        args.aligned_bank.strip().upper(), args.months_lookback, args.workers, date_from, date_to,
        update_master, Path(master_output_path) if master_output_path else None, args.force_refresh,
        args.approve_missing, args.preflight_only, not args.no_dashboard,
        Path(args.dashboard_output.strip()) if args.dashboard_output.strip() else None,
        not args.no_open_dashboard,
        Path(args.standalone_dashboard_output.strip()) if args.standalone_dashboard_output.strip() else None,
        Path(args.total_dashboard_output.strip()) if args.total_dashboard_output.strip() else None,
        Path(args.pi_root.strip()), args.pi_config.strip(),
        Path(args.pi_output.strip()), Path(args.pi_period_output.strip()), args.continue_without_pi,
        max(0, args.pi_retries), args.continue_without_email, args.pi_from_excel.strip(),
    )


# ═════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════

def main() -> int | None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(line_buffering=True)

    (
        operation_mode, run_email, run_pi, run_base,
        mpfm_root, sep_root, output_path, days_count, aligned_bank, months_lookback, workers,
        date_from, date_to, update_master, master_output_path, force_refresh,
        approve_missing, preflight_only, generate_dashboard, dashboard_output_path, open_dashboard,
        standalone_dashboard_output_path, total_dashboard_output_path,
        pi_root, pi_config, pi_output_path, pi_period_output_path, continue_without_pi,
        pi_retries, continue_without_email, pi_from_excel_path,
    ) = _resolve_config()

    standalone_dashboard_path = (
        standalone_dashboard_output_path
        or dashboard_output_path
        or HTML_OUTPUT_DIR / f"{output_path.stem}_STANDALONE_DASHBOARD.html"
    )
    total_dashboard_path = (
        total_dashboard_output_path
        or (
            HTML_OUTPUT_DIR / f"{master_output_path.stem}_RELATORIO_COMPLETO.html"
            if master_output_path
            else HTML_OUTPUT_DIR / f"{output_path.stem}_BASE_UNICA_TOTAL_RELATORIO_COMPLETO.html"
        )
    )

    if (run_email or run_base) and not mpfm_root.is_dir():
        print(f"[ERROR] Pasta MPFM_ROOT não encontrada: {mpfm_root}")
        sys.exit(1)
    if run_base and not sep_root.is_dir():
        print(f"[ERROR] Pasta SEP_ROOT não encontrada: {sep_root}")
        sys.exit(1)
    if days_count < 1:
        print("[ERROR] --days deve ser maior que zero.")
        sys.exit(1)
    if months_lookback < 1:
        print("[ERROR] --months-lookback deve ser maior que zero.")
        sys.exit(1)
    if workers < 1:
        print("[ERROR] --workers deve ser maior que zero.")
        sys.exit(1)

    if run_email:
        try:
            run_email_download_automation(mpfm_root, sep_root, continue_on_warning=continue_without_email)
        except Exception as exc:
            print(f"[WARN] Falha na automação de baixar/organizar PDF/TXT: {exc}")
            if not continue_without_email and (run_pi or run_base):
                if not _prompt_yes_no("Deseja continuar para as próximas etapas mesmo sem concluir baixar/organizar PDF/TXT?", default=False):
                    print("Execução cancelada pelo usuário após falha na automação de e-mail.")
                    return 1
            elif not (run_pi or run_base):
                print("Execução encerrada: modo selecionado era apenas baixar/organizar PDF/TXT.")
                return 1

    if not run_base:
        if run_pi:
            try:
                pi_days = _iso_days_between(date_from, date_to)
                df_pi_only = run_pi_collection_for_days(
                    pi_root, pi_config, pi_output_path, pi_period_output_path, output_path, pi_days, retries=pi_retries
                )
                print(f"\n[OK] Automação PI concluída: {len(df_pi_only)} linha(s) gravadas em {output_path} / aba {PI_SHEET_NAME}")
            except Exception as exc:
                print(f"[WARN] Falha na coleta PI Vision: {exc}")
                if not continue_without_pi:
                    return 1
        elif run_email:
            print("\n[OK] Automação de baixar/organizar PDF/TXT concluída. Modo selecionado não executa PI nem Base_Unica.")
        return

    df_pi_extract = pd.DataFrame()
    pi_attempted = False
    if pi_from_excel_path and Path(pi_from_excel_path).exists():
        try:
            df_pi_extract = pd.read_excel(pi_from_excel_path, sheet_name=PI_SHEET_NAME)
            print(f"[OK] PI_EXTRACT carregada de Excel existente ({len(df_pi_extract)} linha(s)): {pi_from_excel_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao ler PI_EXTRACT de {pi_from_excel_path}: {exc}")
    if run_pi and date_from and date_to and not preflight_only:
        pi_attempted = True
        try:
            pi_days = _iso_days_between(date_from, date_to)
            df_pi_extract = run_pi_collection_for_days(
                pi_root, pi_config, pi_output_path, pi_period_output_path, output_path, pi_days, retries=pi_retries
            )
            print(f"[OK] Aba {PI_SHEET_NAME} salva antes da descoberta/leitura dos PDFs/TXTs: {output_path}")
        except Exception as exc:
            print(f"[WARN] Falha na coleta PI Vision: {exc}")
            print("   Verifique Edge/CDP, login PI Vision, display Daily Control V5 e, se aplicável, o fallback V4.9.")
            if not continue_without_pi:
                if not _prompt_yes_no("Deseja continuar a leitura dos PDFs/TXTs mesmo sem dados PI nesta execução?", default=False):
                    print("Execução cancelada pelo usuário antes da leitura dos PDFs/TXTs.")
                    return
            else:
                print("   --continue-without-pi informado. Continuando sem aba PI nesta execução.")

    today = datetime.now()
    if date_from and date_to:
        months = _months_in_range(date_from, date_to + timedelta(days=1))
        requested_from = date_from.strftime("%Y-%m-%d")
        requested_to = date_to.strftime("%Y-%m-%d")
        print(f"📆 Intervalo solicitado: {requested_from} a {requested_to}")
    else:
        months = _months_to_scan(today, back=months_lookback - 1)
        requested_from = requested_to = ""

    print("📅 Descobrindo dias disponíveis (Daily PDF) por banco...")
    daily_by_bank = {}
    for bank_code in BANK_FOLDERS:
        daily_by_bank[bank_code] = discover_daily_records(
            mpfm_root, bank_code, months,
            keep=None if requested_from else days_count + 5,
            date_from=requested_from or None,
            date_to=requested_to or None,
            workers=workers,
        )
        print(f"  {bank_code}: {len(daily_by_bank[bank_code])} dia(s) encontrados")

    all_days = sorted({day for recs in daily_by_bank.values() for day in recs}, reverse=True)
    if not all_days:
        print("[ERROR] Nenhum dia com PDF Daily encontrado em nenhum banco. Abortando.")
        return 1

    selected_days = sorted(all_days if requested_from else all_days[:days_count])
    selection_label = "Dias selecionados" if requested_from else f"Últimos {len(selected_days)} dia(s) selecionados"
    print(f"\n[OK] {selection_label}: {', '.join(selected_days)}")
    period_selected = bool(requested_from)
    if period_selected:
        print("[INFO] Período informado pelo usuário: os dias da janela serão reprocessados e sobrescritos no consolidado.")

    most_recent = max(all_days)
    expected_most_recent = (date_to or (today - timedelta(days=1))).strftime("%Y-%m-%d")
    if most_recent < expected_most_recent:
        print(
            f"[WARN] ATENÇÃO: o dia mais recente disponível ({most_recent}) é anterior a ontem "
            f"({expected_most_recent}). Os dados podem não estar em dia."
        )

    for bank_code in BANK_FOLDERS:
        missing = [d for d in selected_days if d not in daily_by_bank[bank_code]]
        if missing:
            print(f"[WARN] {bank_code}: sem Daily PDF para {', '.join(missing)}")

    preflight_rows, preflight_issues = build_preflight_report(
        mpfm_root, sep_root, months, selected_days, daily_by_bank, aligned_bank
    )
    print_preflight_report(preflight_rows, preflight_issues)
    if preflight_only:
        print("\n[OK] Pré-validação concluída. Execução encerrada por --preflight-only.")
        return
    if preflight_issues and not approve_missing:
        print("\n[WARN] A janela selecionada possui dados faltantes/incompletos.")
        if not _prompt_yes_no("Deseja continuar mesmo assim?", default=False):
            print("Execução cancelada pelo usuário antes do processamento pesado.")
            return
    elif preflight_issues and approve_missing:
        print("\n[WARN] Dados faltantes encontrados, mas --approve-missing foi informado. Continuando execução.")

    skipped_days = []
    if update_master and master_output_path and master_output_path.exists() and not force_refresh and not period_selected:
        loaded_days = loaded_days_in_master(master_output_path, selected_days, daily_by_bank)
        skipped_days = [day for day in selected_days if day in loaded_days]
        if skipped_days:
            print(
                "\n[SKIP] Modo incremental: pulando dia(s) já carregado(s) na Base_Unica total: "
                f"{', '.join(skipped_days)}"
            )
            print("   Use --force-refresh se precisar reprocessar correções retroativas.")
    elif force_refresh:
        print("\n[INFO] --force-refresh informado: todos os dias selecionados serão reprocessados.")

    target_days = [day for day in selected_days if day not in skipped_days]
    if not target_days:
        print("\n[OK] Todos os dias selecionados já estão carregados na Base_Unica total. Nada novo para processar.")
        if run_pi and not pi_attempted:
            try:
                print("[INFO] A Base_Unica já estava carregada, mas a rota PI ainda será executada para atualizar a evidência operacional.")
                df_pi_extract = run_pi_collection_for_days(
                    pi_root, pi_config, pi_output_path, pi_period_output_path, output_path, selected_days, retries=pi_retries
                )
                print(f"[OK] Aba {PI_SHEET_NAME} salva: {output_path}")
            except Exception as exc:
                print(f"[WARN] Falha na coleta PI Vision: {exc}")
                print("   A falha será registrada quando possível e o dashboard será republicado com a evidência disponível.")
                if not continue_without_pi:
                    if not _prompt_yes_no("Deseja republicar o dashboard mesmo sem dados PI novos nesta execução?", default=True):
                        print("Execução cancelada pelo usuário após falha na rota PI.")
                        return
        if update_master and master_output_path and not df_pi_extract.empty:
            try:
                pi_stats = update_master_pi_extract(master_output_path, df_pi_extract)
                print(
                    "[OK] Dados PI consolidados: "
                    f"{pi_stats['increment_rows']} novas linhas; "
                    f"{pi_stats['total_rows']} totais na aba {PI_MASTER_SHEET_NAME}"
                )
            except PermissionError:
                print(f"[WARN] Não foi possível atualizar dados PI porque o arquivo está aberto: {master_output_path}")
            except Exception as exc:
                print(f"[WARN] Falha ao atualizar dados PI no consolidado ({master_output_path}): {exc}")
        if generate_dashboard and update_master and master_output_path and master_output_path.exists():
            print("[INFO] Republicando o dashboard a partir da Base_Unica total, sem reler PDFs/TXTs...")
            try:
                master_df = read_master_base_unica(master_output_path)
                master_window = master_df[master_df["ProductionDate"].astype(str).isin({str(day) for day in selected_days})].copy()
            except Exception as exc:
                print(f"[WARN] Falha ao ler a Base_Unica total para republicar o dashboard ({master_output_path}): {exc}")
                return
            try:
                comp_stats = refresh_master_comparativo_total(master_output_path, df_base=master_df)
                print(f"[OK] Aba {COMPARATIVO_TOTAL_SHEET_NAME} atualizada: {comp_stats['rows']} linha(s)")
            except Exception as exc:
                print(f"[WARN] Não foi possível atualizar {COMPARATIVO_TOTAL_SHEET_NAME}: {exc}")
            try:
                pair_stats = refresh_master_comparativo_pares(master_output_path, aligned_bank, df_base=master_df)
                print(f"[OK] Aba {COMPARATIVO_PARES_SHEET_NAME} atualizada: {pair_stats['rows']} linha(s)")
            except Exception as exc:
                print(f"[WARN] Não foi possível atualizar {COMPARATIVO_PARES_SHEET_NAME}: {exc}")
            try:
                _remove_separator_comparison_sheets(master_output_path)
                print("[OK] Comparações MPFM × SEP permanecem somente no HTML sob demanda")
            except Exception as exc:
                print(f"[WARN] Não foi possível remover abas comparativas SEP antigas: {exc}")
            try:
                format_workbook(master_output_path)
            except Exception as exc:
                print(f"[WARN] Não foi possível aplicar formatação rápida ao consolidado: {exc}")
            try:
                master_alarm_events = read_master_alarm_events(master_output_path)
            except Exception as exc:
                print(f"[WARN] Falha ao ler alarmes/eventos consolidados para o dashboard ({master_output_path}): {exc}")
                master_alarm_events = pd.DataFrame()
            if generate_dashboard:
                publish_dashboard(
                    standalone_dashboard_path,
                    master_window,
                    selected_days,
                    selected_days,
                    preflight_rows,
                    None,
                    master_alarm_events,
                    aligned_bank=aligned_bank,
                    source_workbook_path=output_path,
                )
                publish_dashboard(
                    total_dashboard_path,
                    master_window,
                    selected_days,
                    selected_days,
                    preflight_rows,
                    master_output_path,
                    master_alarm_events,
                    aligned_bank=aligned_bank,
                    preloaded_master_df=master_df,
                    source_workbook_path=master_output_path,
                )
            if open_dashboard:
                try:
                    webbrowser.open(total_dashboard_path.resolve().as_uri())
                except Exception as exc:
                    print(f"[WARN] Dashboards gerados, mas não foi possível abrir o relatório completo: {exc}")
        return
    target_days_set = set(target_days)
    if skipped_days:
        print(f"[INFO] Dia(s) que serão processados nesta execução: {', '.join(target_days)}")

    if run_pi and not pi_attempted:
        try:
            df_pi_extract = run_pi_collection_for_days(
                pi_root, pi_config, pi_output_path, pi_period_output_path, output_path, target_days, retries=pi_retries
            )
            print(f"[OK] Aba {PI_SHEET_NAME} salva antes da leitura dos PDFs/TXTs: {output_path}")
        except Exception as exc:
            print(f"[WARN] Falha na coleta PI Vision: {exc}")
            print("   Verifique Edge/CDP, login PI Vision, display Daily Control V5 e, se aplicável, o fallback V4.9.")
            if not continue_without_pi:
                if not _prompt_yes_no("Deseja continuar a leitura dos PDFs/TXTs mesmo sem dados PI nesta execução?", default=False):
                    print("Execução cancelada pelo usuário antes da leitura dos PDFs/TXTs.")
                    return
            else:
                print("   --continue-without-pi informado. Continuando sem aba PI nesta execução.")
    elif not run_pi:
        print("\n[SKIP] Coleta PI Vision desativada por --no-pi.")

    print("\n📄 Lendo relatórios Hourly (apenas dos dias selecionados)...")
    hourly_by_bank_day = {}
    for bank_code in BANK_FOLDERS:
        hourly_by_bank_day[bank_code] = discover_hourly_records_for_days(mpfm_root, bank_code, months, target_days_set, workers)
        total_hours = sum(len(v) for v in hourly_by_bank_day[bank_code].values())
        print(f"  {bank_code}: {total_hours} registro(s) horário(s)")

    print("\n🧪 Localizando TXT do Separador de Testes (FC13/FC14/FC17)...")
    sep_data_by_day = load_sep_data_for_days(sep_root, target_days)
    for day in target_days:
        if day not in sep_data_by_day:
            print(f"  [WARN] SEP não disponível para {day} — a Base Única ficará sem linha própria do separador nesse dia")

    all_rows = []
    for day in target_days:
        sep_day_data = sep_data_by_day.get(day)
        for bank_code in BANK_FOLDERS:
            daily_entry = daily_by_bank[bank_code].get(day)
            hourly_recs = hourly_by_bank_day[bank_code].get(day, [])
            # O PDF B05 contém PE_4/18FT1506 e PE_EO105/18FT1706; ambas as
            # seções são emitidas uma única vez no processamento nativo B05.
            strict = False

            if hourly_recs:
                df_hourly = build_hourly_df_with_sep(hourly_recs, bank_code, None, strict=strict)
                all_rows.extend(hourly_df_to_rows(df_hourly, sep_merged=False))

            if daily_entry:
                _, daily_record = daily_entry
                df_daily = build_daily_df(daily_record, bank_code, strict=strict)
                all_rows.extend(daily_df_to_rows(df_daily, None))

                df_recon = build_recon_df(daily_record, hourly_recs, bank_code, strict=strict)
                all_rows.extend(recon_df_to_rows(df_recon))

        all_rows.extend(sep_data_to_rows(day, sep_day_data))

        print(f"  📦 {day}: processado")

    print("\n🚨 Lendo alarmes e eventos FCS320...")
    df_alarm_events = load_alarm_event_data(mpfm_root, target_days)
    alarm_count = int((df_alarm_events["RecordType"].astype(str) == "ALARM").sum()) if not df_alarm_events.empty else 0
    event_count = int((df_alarm_events["RecordType"].astype(str) == "EVENT").sum()) if not df_alarm_events.empty else 0
    print(f"  Alarmes/eventos extraídos: {len(df_alarm_events)} linha(s) ({alarm_count} alarmes, {event_count} eventos)")

    if not all_rows:
        print("[ERROR] Nenhuma linha gerada. Abortando sem escrever arquivo.")
        return 1

    df_out = pd.DataFrame(all_rows, columns=BASE_UNICA_COLUMNS)
    dedup_keys = df_out[MASTER_DEDUP_KEYS].copy()
    for key in MASTER_DEDUP_KEYS:
        dedup_keys[key] = dedup_keys[key].map(lambda value, key=key: _canonical_master_key_value(value, key))
    before_dedup_rows = len(df_out)
    df_out = df_out.loc[~dedup_keys.duplicated(subset=MASTER_DEDUP_KEYS, keep="last")].copy()
    removed_duplicate_rows = before_dedup_rows - len(df_out)
    if removed_duplicate_rows:
        print(f"🧹 Removidas {removed_duplicate_rows} linha(s) duplicada(s) antes de gravar a Base_Unica.")
    df_out.sort_values(
        by=["ProductionDate", "Bank", "Granularity", "Hour", "Tag"],
        inplace=True,
        na_position="last",
        key=lambda col: col.astype(str) if col.name != "Hour" else col,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _remove_separator_comparison_sheets(output_path)
    writer_mode = "a" if output_path.exists() else "w"
    writer_kwargs = {"engine": "openpyxl", "mode": writer_mode}
    if writer_mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"
    df_comparativo_total = build_comparativo_total_df(df_out, df_pi_extract)
    df_comparativo_pares = build_comparativo_pares_df(df_out, target_days, aligned_bank)
    with pd.ExcelWriter(output_path, **writer_kwargs) as writer:
        df_out.to_excel(writer, sheet_name="BASE_UNICA_STANDALONE", index=False)
        df_alarm_events.to_excel(writer, sheet_name=ALARM_EVENT_SHEET_NAME, index=False)
        df_comparativo_total.to_excel(writer, sheet_name=COMPARATIVO_TOTAL_SHEET_NAME, index=False)
        df_comparativo_pares.to_excel(writer, sheet_name=COMPARATIVO_PARES_SHEET_NAME, index=False)
    add_comparativo_sheet(output_path)
    format_workbook(output_path)
    print(f"\n[OK] Concluído: {len(df_out)} linha(s) gravadas em {output_path}")
    print(f"[OK] Aba {COMPARATIVO_TOTAL_SHEET_NAME} gerada: {len(df_comparativo_total)} linha(s)")
    print(f"[OK] Aba {COMPARATIVO_PARES_SHEET_NAME} gerada: {len(df_comparativo_pares)} linha(s)")
    print("[OK] Dados próprios do SEP mantidos na Base Única; comparação disponível somente no HTML")

    master_combined_df = None
    if update_master and master_output_path:
        _t_master = time.time()
        try:
            master_replace_days = target_days if (period_selected or force_refresh) else None
            stats = update_master_base_unica(master_output_path, df_out, replace_days=master_replace_days)
            master_combined_df = stats.get("df")
            print(
                "[OK] Base_Unica total atualizada: "
                f"{master_output_path} "
                f"({stats['previous_rows']} anteriores + {stats['increment_rows']} novas "
                f"=> {stats['total_rows']} totais; "
                f"{stats['replaced_or_duplicate_rows']} substituídas/duplicadas; "
                f"{stats['replaced_rows']} removidas por reprocessamento da janela) "
                f"[{time.time() - _t_master:.1f}s]"
            )
        except PermissionError:
            print(f"[WARN] Não foi possível atualizar a Base_Unica total porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao atualizar a Base_Unica total ({master_output_path}): {exc}")

        try:
            alarm_stats = update_master_alarm_events(master_output_path, df_alarm_events, replace_days=target_days)
            print(
                "[OK] Alarmes/eventos consolidados: "
                f"{alarm_stats['increment_rows']} novas linhas; "
                f"{alarm_stats['replaced_rows']} linha(s) substituída(s); "
                f"{alarm_stats['total_rows']} totais"
            )
        except PermissionError:
            print(f"[WARN] Não foi possível atualizar alarmes/eventos porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao atualizar alarmes/eventos no consolidado ({master_output_path}): {exc}")

        if not df_pi_extract.empty:
            try:
                pi_stats = update_master_pi_extract(master_output_path, df_pi_extract)
                print(
                    "[OK] Dados PI consolidados: "
                    f"{pi_stats['increment_rows']} novas linhas; "
                    f"{pi_stats['total_rows']} totais na aba {PI_MASTER_SHEET_NAME}"
                )
            except PermissionError:
                print(f"[WARN] Não foi possível atualizar dados PI porque o arquivo está aberto: {master_output_path}")
            except Exception as exc:
                print(f"[WARN] Falha ao atualizar dados PI no consolidado ({master_output_path}): {exc}")

        try:
            _t = time.time()
            comp_stats = refresh_master_comparativo_total(master_output_path, df_base=master_combined_df)
            print(f"[OK] Aba {COMPARATIVO_TOTAL_SHEET_NAME} consolidada: {comp_stats['rows']} linha(s) [{time.time() - _t:.1f}s]")
        except PermissionError:
            print(f"[WARN] Não foi possível atualizar {COMPARATIVO_TOTAL_SHEET_NAME} porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao atualizar {COMPARATIVO_TOTAL_SHEET_NAME} no consolidado ({master_output_path}): {exc}")

        try:
            _t = time.time()
            pair_stats = refresh_master_comparativo_pares(master_output_path, aligned_bank, df_base=master_combined_df)
            print(f"[OK] Aba {COMPARATIVO_PARES_SHEET_NAME} consolidada: {pair_stats['rows']} linha(s) [{time.time() - _t:.1f}s]")
        except PermissionError:
            print(f"[WARN] Não foi possível atualizar {COMPARATIVO_PARES_SHEET_NAME} porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao atualizar {COMPARATIVO_PARES_SHEET_NAME} no consolidado ({master_output_path}): {exc}")

        try:
            _remove_separator_comparison_sheets(master_output_path)
            print("[OK] Abas comparativas SEP antigas removidas do consolidado")
        except PermissionError:
            print(f"[WARN] Não foi possível remover abas comparativas SEP porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha ao remover abas comparativas SEP antigas ({master_output_path}): {exc}")

        try:
            _t = time.time()
            format_workbook(master_output_path)
            print(f"[OK] Formatação corporativa aplicada ao consolidado Excel [{time.time() - _t:.1f}s]")
        except PermissionError:
            print(f"[WARN] Não foi possível formatar o consolidado porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"[WARN] Falha na formatação do consolidado Excel: {exc}")

    if generate_dashboard:
        monthly_master_path = master_output_path if master_output_path and master_output_path.exists() else None
        if monthly_master_path and master_combined_df is None:
            try:
                master_combined_df = read_master_base_unica(monthly_master_path)
            except Exception as exc:
                print(f"[WARN] Base Única mensal não pôde ser lida para o CEP do HTML: {exc}")
        publish_dashboard(
            standalone_dashboard_path,
            df_out,
            target_days,
            selected_days,
            preflight_rows,
            monthly_master_path,
            df_alarm_events,
            aligned_bank=aligned_bank,
            preloaded_master_df=master_combined_df,
            source_workbook_path=monthly_master_path or output_path,
        )
        total_df = None
        total_alarm_events = None
        if update_master and master_output_path and master_output_path.exists():
            try:
                # Reaproveita o DataFrame combinado já atualizado em memória
                # nesta mesma execução em vez de reler o arquivo do disco.
                total_df = master_combined_df if master_combined_df is not None else read_master_base_unica(master_output_path)
                total_alarm_events = read_master_alarm_events(master_output_path)
            except Exception as exc:
                # Não deixa uma falha de leitura do consolidado (ex.: arquivo
                # ainda sendo sincronizado pelo OneDrive) abortar o script sem
                # gerar o relatório completo — cai para os dados da janela.
                print(f"[WARN] Falha ao ler a Base_Unica total para o relatório completo ({master_output_path}): {exc}")
                print("   O relatório completo será gerado com os dados da janela desta execução.")
                total_df = None
                total_alarm_events = None
        if total_df is None:
            if update_master and master_output_path:
                print("[WARN] Base_Unica total não disponível; relatório completo será gerado com os dados da janela.")
            _t_total_html = time.time()
            publish_dashboard(
                total_dashboard_path,
                df_out,
                target_days,
                selected_days,
                preflight_rows,
                None,
                df_alarm_events,
                aligned_bank=aligned_bank,
                source_workbook_path=master_output_path if master_output_path and master_output_path.exists() else output_path,
            )
            print(f"[TIME] Relatório HTML total publicado em {time.time() - _t_total_html:.1f}s: {total_dashboard_path}")
        else:
            _t_total_html = time.time()
            publish_dashboard(
                total_dashboard_path,
                total_df,
                selected_days,
                selected_days,
                preflight_rows,
                master_output_path,
                total_alarm_events,
                aligned_bank=aligned_bank,
                preloaded_master_df=total_df,
                source_workbook_path=master_output_path,
            )
            print(f"[TIME] Relatório HTML total publicado em {time.time() - _t_total_html:.1f}s: {total_dashboard_path}")
        if open_dashboard:
            try:
                webbrowser.open(total_dashboard_path.resolve().as_uri())
            except Exception as exc:
                print(f"[WARN] Dashboards gerados, mas não foi possível abrir o relatório completo no navegador: {exc}")

    # ── Persistência PI Vision no banco local ────────────────────────────────
    if not df_pi_extract.empty:
        try:
            import sqlite3 as _sqlite3
            _db_path = Path(__file__).parents[1] / "data" / "mpfm_local.db"
            if _db_path.exists():
                import sys as _sys
                _proj_root = str(Path(__file__).parents[1])
                if _proj_root not in _sys.path:
                    _sys.path.insert(0, _proj_root)
                from services.importing.pi_vision_import_service import import_pi_excel as _import_pi_excel
                _conn = _sqlite3.connect(_db_path)
                try:
                    _pi_res = _import_pi_excel(_conn, excel_path=pi_output_path)
                    print(
                        f"[PI→DB] {_pi_res.get('inserted', 0)} leituras inseridas, "
                        f"{_pi_res.get('skipped', 0)} ignoradas "
                        f"({_pi_res.get('elapsed_s', 0):.1f}s) — {pi_output_path}"
                    )
                finally:
                    _conn.close()
            else:
                print(f"[PI→DB] Banco não encontrado em {_db_path} — dados PI não persistidos.")
        except Exception as _exc:
            print(f"[PI→DB] Falha ao persistir leituras PI no banco (não crítico): {_exc}")


if __name__ == "__main__":
    raise SystemExit(main())
