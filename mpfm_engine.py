#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MPFM Generic Reconciliation Engine – Bacalhau FPSO
Suporta qualquer unidade (B03/B05/B08/B10/...), North/South, Topside/Subsea.
Uso: configurar a lista UNITS no bloco __main__ e executar.
"""

import re, glob, os
import numpy as np
import pandas as pd
from datetime import date as _date, datetime as _datetime
from pypdf import PdfReader
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Formatação BR ─────────────────────────────────────────────────────────────
_DATE_PAT     = re.compile(r'^\d{4}-\d{2}-\d{2}$')
_DATETIME_PAT = re.compile(r'^\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}')
_DATE_COLS    = {'Dia', 'Dia ref.'}
_DTIME_COLS   = {'DT Início', 'DT Fim'}

def _to_date(v):
    """'YYYY-MM-DD' → datetime.date  (para exibir dd-mm-yyyy no Excel)."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    s = str(v).strip()
    if _DATE_PAT.match(s):
        y, m, d = s.split('-')
        return _date(int(y), int(m), int(d))
    return v

def _to_datetime(v):
    """'YYYY-MM-DD HH:MM' → datetime.datetime."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    s = str(v).strip()
    if _DATETIME_PAT.match(s):
        try:
            return _datetime.strptime(s[:16], '%Y-%m-%d %H:%M')
        except ValueError:
            return v
    return v

def _prep_dates(df):
    """Converte colunas de data/hora para objetos Python antes de escrever no Excel."""
    df = df.copy()
    for col in df.columns:
        if col in _DATE_COLS:
            df[col] = df[col].apply(_to_date)
        elif col in _DTIME_COLS:
            df[col] = df[col].apply(_to_datetime)
    return df

# Números gravados como float — formato controlado via number_format no openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def to_float(x):
    x = str(x).strip().replace('\u00A0', '').replace(' ', '')
    if x in ('-', '', 'None', 'nan', 'NaN'):
        return np.nan
    try:
        return float(x.replace(',', '.'))
    except Exception:
        return np.nan

def flat_text(pdf_path):
    reader = PdfReader(pdf_path)
    raw = '\n'.join(page.extract_text() or '' for page in reader.pages)
    return re.sub(r'\s+', ' ', raw)

NUM = r'([\-\d]+(?:\.\d+)?)'
C5  = ['gas', 'oil', 'hc', 'water', 'total']
C4  = ['gas', 'oil', 'hc', 'water']

def nan_dict(keys):
    return {k: np.nan for k in keys}

# ─────────────────────────────────────────────────────────────────────────────
# Metric row parsers
# ─────────────────────────────────────────────────────────────────────────────

def parse_row(label_pat, text, ncols=5):
    cols = (C5 if ncols == 5 else C4)[:ncols]
    m = re.search(
        label_pat + r'[^\[]*\[(?:t|Sm[³3]?)\]\s*' + r'\s+'.join([NUM] * ncols),
        text, re.IGNORECASE
    )
    if m:
        return {c: to_float(m.group(i + 1)) for i, c in enumerate(cols)}
    return nan_dict(cols)

def parse_fwa(text):
    """
    Extrai Pressão, Temperatura e Densidades da seção Flow Weighted Averages.
    Colunas: Gas | Oil | Meter | Water
    Pressão e Temperatura → coluna Meter (3ª)
    Density → Gas / Oil / Water (coluna Meter é '-' para densidade)
    Suporta multiline (subsea PDFs) e inline (daily PDFs).
    """
    def _find_meter_value(label, unit_pat):
        """Captura valor da 3ª coluna (Meter) — pula Gas e Oil que são '-'."""
        # Formato multiline
        m = re.search(
            label + r'\s*\n?\s*' + unit_pat +
            r'\s*\n(?:-[^\n]*\n){0,2}\s*(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return to_float(m.group(1))
        # Formato inline: "Pressure [barg] - - 538.91"
        m = re.search(
            label + r'\s*' + unit_pat + r'(?:\s*-\s*){1,3}(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return to_float(m.group(1))
        return np.nan

    def _extract_density(text):
        """
        Extrai Gas, Oil, Water de Density [kg/m³].
        Para em 'Production Total' ou próximo TAG para evitar capturar
        totais da p.2 (flat_text colapsa newlines em espaços).
        Aceita max 4 tokens diretos após a unidade.
        """
        m = re.search(
            r'Density\s*\[kg/m\S*\]\s*(.*?)'
            r'(?=Production\s+Total|PE_|Riser\s+P|\Z)',
            text, re.DOTALL | re.IGNORECASE
        )
        if not m:
            return np.nan, np.nan, np.nan
        tokens = []
        for part in re.split(r'\s+', m.group(1).strip()):
            if re.fullmatch(r'-|\d+(?:\.\d+)?', part):
                tokens.append(part)
                if len(tokens) == 4:
                    break
            elif part:   # palavra não-numérica → para
                break
        def t2f(t): return np.nan if t == '-' else to_float(t)
        gas   = t2f(tokens[0]) if len(tokens) > 0 else np.nan
        oil   = t2f(tokens[1]) if len(tokens) > 1 else np.nan
        water = t2f(tokens[3]) if len(tokens) > 3 else np.nan
        return gas, oil, water

    pres = _find_meter_value(r'Pressure',    r'\[barg\]')
    temp = _find_meter_value(r'Temperature', r'\[.{1,3}C\]')
    dens_gas, dens_oil, dens_water = _extract_density(text)

    return {
        'pressure':    pres,
        'temperature': temp,
        'dens_gas':    dens_gas,
        'dens_oil':    dens_oil,
        'dens_water':  dens_water,
    }

def parse_metrics(block, ncols=5):
    corr = parse_row(r'MPFM\s+corrected\s+mass', block, ncols)
    # TAG com produção zero não tem FWA fisicamente definida
    all_zero = all(
        (v == 0 or (isinstance(v, float) and (v == 0.0 or np.isnan(v))))
        for v in corr.values()
    )
    fwa = {k: np.nan for k in ('pressure','temperature','dens_gas','dens_oil','dens_water')} \
          if all_zero else parse_fwa(block)
    return {
        'mpfm_uncorr': parse_row(r'MPFM\s+uncorrected\s+mass',                   block, ncols),
        'mpfm_corr':   corr,
        'pvt_mass':    parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*(?!@)',          block, ncols),
        'pvt_vol':     parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*(?!@)',   block, ncols),
        'pvt20_mass':  parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*@20',            block, ncols),
        'pvt20_vol':   parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*@20',     block, ncols),
        'fwa':         fwa,
    }

# ─────────────────────────────────────────────────────────────────────────────
# PDF parser — universal (Daily & Hourly, any unit)
# ─────────────────────────────────────────────────────────────────────────────

# Matches: "Riser P5", "Riser P1-", "PE_4", "PE_EO4", "PI_2", "PW-104DA" etc.
TAG_RE = re.compile(
    r'((?:Riser\s+P\d+|P[EI]_[A-Z0-9]+-?|PW[-_][A-Z0-9]+-?))\s*-\s*(\d{2}FT\d+)\s+'
    r'Production\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)
AREA_TOTAL_RE = re.compile(
    r'(North|South)\s*-\s*(Topside|Subsea)\s+MPFM\s+Production\s+Total\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)

def parse_pdf(pdf_path, report_type='daily'):
    text = flat_text(pdf_path)

    # Detect FPSO side and unit type from title line
    if 'FPSO South' in text[:300]:   fpso_side = 'South'
    elif 'FPSO West' in text[:300]:    fpso_side = 'West'
    elif 'FPSO North' in text[:300]:   fpso_side = 'North'
    else:                               fpso_side = 'South'
    unit_type  = 'Subsea' if 'Subsea' in text[:300] else 'Topside'

    if report_type == 'daily':
        dm = re.search(r'Daily Report from\s+([\d.]+)\s+[\d:]+\s+to\s+([\d.]+)', text)
        date_from = dm.group(1).replace('.', '-') if dm else None
        date_to   = dm.group(2).replace('.', '-') if dm else None
        hour = None; dt_from = dt_to = None
    else:
        hm = re.search(r'Hourly Report from\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)\s+to\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)', text)
        dt_from   = f"{hm.group(1).replace('.', '-')} {hm.group(2)}" if hm else None
        dt_to     = f"{hm.group(3).replace('.', '-')} {hm.group(4)}" if hm else None
        date_from = hm.group(1).replace('.', '-') if hm else None
        date_to   = None
        # A hora operacional do PDF hourly é a hora inicial da janela.
        hour      = int(hm.group(2).split(':')[0]) if hm else None

    # Individual tag blocks
    headers = list(TAG_RE.finditer(text))
    tags = {}
    for i, h in enumerate(headers):
        tag = h.group(1).strip().replace(' ', '_').rstrip('-')
        instr = h.group(2)
        block = text[h.start(): headers[i+1].start() if i+1 < len(headers) else None]
        tags[tag] = {'instrument': instr, 'metrics': parse_metrics(block, 5)}

    # Area total
    am = AREA_TOTAL_RE.search(text)
    area_total = {}
    if am:
        block = text[am.start():]
        area_total = {
            'area':      am.group(1),
            'unit_type': am.group(2),
            'metrics':   parse_metrics(block, 4),
        }

    return {
        'pdf_path':  pdf_path,
        'date_from': date_from,
        'date_to':   date_to,
        'dt_from':   dt_from,
        'dt_to':     dt_to,
        'hour':      hour,
        'fpso_side': fpso_side,
        'unit_type': unit_type,
        'tags':      tags,
        'area_total': area_total,
    }

# ─────────────────────────────────────────────────────────────────────────────
# Column definitions
# ─────────────────────────────────────────────────────────────────────────────

HOURLY_COLS = [
    'Dia ref.','Hora','Bank','Loop','Tipo','TAG','Instrumento','DT Início','DT Fim',
    'MPFM uncorr Gás (t)','MPFM uncorr Óleo (t)','MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)','MPFM uncorr Total (t)',
    'MPFM corr Gás (t)','MPFM corr Óleo (t)','MPFM corr HC (t)',
    'MPFM corr Água (t)','MPFM corr Total (t)',
    'PVT mass Gás (t)','PVT mass Óleo (t)','PVT mass Água (t)',
    'PVT vol Gás (Sm³)','PVT vol Óleo (m³)','PVT vol Água (m³)',
    'PVT @20 mass Gás (t)','PVT @20 mass Óleo (t)','PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)','PVT @20 vol Óleo (m³)','PVT @20 vol Água (m³)',
    'Pressão (barg)','Temperatura (°C)',
    'Dens. Gás (kg/m³)','Dens. Óleo (kg/m³)','Dens. Água (kg/m³)','Fonte',
]

DAILY_COLS = [
    'Bank','Loop','Tipo','TAG','Instrumento','Dia',
    'MPFM uncorr Gás (t)','MPFM uncorr Óleo (t)','MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)','MPFM uncorr Total (t)',
    'MPFM corr Gás (t)','MPFM corr Óleo (t)','MPFM corr HC (t)',
    'MPFM corr Água (t)','MPFM corr Total (t)',
    'PVT mass Gás (t)','PVT mass Óleo (t)','PVT mass Água (t)',
    'PVT vol Gás (Sm³)','PVT vol Óleo (m³)','PVT vol Água (m³)',
    'PVT @20 mass Gás (t)','PVT @20 mass Óleo (t)','PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)','PVT @20 vol Óleo (m³)','PVT @20 vol Água (m³)',
    'Pressão (barg)','Temperatura (°C)',
    'Dens. Gás (kg/m³)','Dens. Óleo (kg/m³)','Dens. Água (kg/m³)','Fonte (Daily)',
]

def _mrow(m):
    u=m['mpfm_uncorr']; c=m['mpfm_corr']
    pm=m['pvt_mass'];   pv=m['pvt_vol']
    p2=m['pvt20_mass']; v2=m['pvt20_vol']
    f=m['fwa']
    return [
        u.get('gas'),u.get('oil'),u.get('hc'),u.get('water'),u.get('total'),
        c.get('gas'),c.get('oil'),c.get('hc'),c.get('water'),c.get('total'),
        pm.get('gas'),pm.get('oil'),pm.get('water'),
        pv.get('gas'),pv.get('oil'),pv.get('water'),
        p2.get('gas'),p2.get('oil'),p2.get('water'),
        v2.get('gas'),v2.get('oil'),v2.get('water'),
        f['pressure'],f['temperature'],
        f['dens_gas'],f['dens_oil'],f['dens_water'],
    ]

# ─────────────────────────────────────────────────────────────────────────────
# DataFrame builders
# ─────────────────────────────────────────────────────────────────────────────

def build_hourly_df(hourly_records, unit_code):
    rows = []
    for rec in sorted(hourly_records, key=lambda r: r['hour'] or 0):
        for tag, td in rec['tags'].items():
            base = [rec['date_from'], rec['hour'],
                    unit_code, rec['fpso_side'], rec['unit_type'],
                    tag, td['instrument'], rec['dt_from'], rec['dt_to']]
            rows.append(base + _mrow(td['metrics']) + [os.path.basename(rec['pdf_path'])])
    return pd.DataFrame(rows, columns=HOURLY_COLS)

def build_daily_df(daily, unit_code):
    rows = []
    for tag, td in daily['tags'].items():
        base = [unit_code, daily['fpso_side'], daily['unit_type'],
                tag, td['instrument'], daily['date_from']]
        rows.append(base + _mrow(td['metrics']) + [os.path.basename(daily['pdf_path'])])
    return pd.DataFrame(rows, columns=DAILY_COLS)

def build_recon_df(daily, hourly_records, unit_code, abs_tol=0.5):
    hours_found = sorted(set(r['hour'] for r in hourly_records if r['hour'] is not None))
    hours_str   = ','.join(f'{h:02d}' for h in hours_found)
    n           = len(hours_found)
    coverage    = 'OK (24/24h)' if n == 24 else (f'PARCIAL ({n}/24h)' if n > 0 else 'SEM HORÁRIOS')

    rows = []
    for tag, td in daily['tags'].items():
        d_c  = td['metrics']['mpfm_corr']
        d_pv = td['metrics']['pvt_vol']

        s = {k: 0.0 for k in ['gas','oil','hc','water','total',
                                'pv_gas','pv_oil','pv_water']}
        for rec in hourly_records:
            if tag not in rec['tags']:
                continue
            hm = rec['tags'][tag]['metrics']
            def v(x): return x if isinstance(x, float) and not np.isnan(x) else 0.0
            for col in C5:
                s[col]     += v(hm['mpfm_corr'].get(col, 0.0))
            s['pv_gas']   += v(hm['pvt_vol'].get('gas', 0.0))
            s['pv_oil']   += v(hm['pvt_vol'].get('oil', 0.0))
            s['pv_water'] += v(hm['pvt_vol'].get('water', 0.0))

        def delta(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)): return np.nan
            return round(sv - dv, 4)

        def status(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)): return '-'
            if n == 0: return 'SEM DADOS'
            d = abs(sv - dv)
            return 'OK' if d <= max(abs_tol, 0.0005 * abs(dv)) else 'VERIFICAR'

        row = {
            'Bank':       unit_code,
            'Loop':       daily['fpso_side'],
            'Tipo':       daily['unit_type'],
            'TAG':        tag,
            'Instrumento':td['instrument'],
            'Dia':        daily['date_from'],
            'Cobertura':  coverage,
            'Horas':      hours_str if n > 0 else '-',
            # MPFM corrected mass
            'Daily Gás (t)':         d_c.get('gas'),
            'Soma h. Gás (t)':       round(s['gas'],   4),
            'Δ Gás (t)':             delta(d_c.get('gas'),   s['gas']),
            'Status Gás':            status(d_c.get('gas'),  s['gas']),
            'Daily Óleo (t)':        d_c.get('oil'),
            'Soma h. Óleo (t)':      round(s['oil'],   4),
            'Δ Óleo (t)':            delta(d_c.get('oil'),   s['oil']),
            'Status Óleo':           status(d_c.get('oil'),  s['oil']),
            'Daily HC (t)':          d_c.get('hc'),
            'Soma h. HC (t)':        round(s['hc'],    4),
            'Δ HC (t)':              delta(d_c.get('hc'),    s['hc']),
            'Status HC':             status(d_c.get('hc'),   s['hc']),
            'Daily Água (t)':        d_c.get('water'),
            'Soma h. Água (t)':      round(s['water'], 4),
            'Δ Água (t)':            delta(d_c.get('water'), s['water']),
            'Status Água':           status(d_c.get('water'),s['water']),
            # PVT vol
            'Daily PVT Gás (Sm³)':   d_pv.get('gas'),
            'Soma h. PVT Gás (Sm³)': round(s['pv_gas'],   1),
            'Δ PVT Gás':             delta(d_pv.get('gas'),   s['pv_gas']),
            'Daily PVT Óleo (m³)':   d_pv.get('oil'),
            'Soma h. PVT Óleo (m³)': round(s['pv_oil'],   4),
            'Δ PVT Óleo':            delta(d_pv.get('oil'),   s['pv_oil']),
        }
        rows.append(row)
    return pd.DataFrame(rows)

def build_area_totals_row(daily, unit_code):
    at = daily.get('area_total', {})
    if not at:
        return None
    c  = at['metrics']['mpfm_corr']
    u  = at['metrics']['mpfm_uncorr']
    pv = at['metrics']['pvt_vol']
    pm = at['metrics']['pvt_mass']
    return {
        'Bank':     unit_code,
        'Loop':     at['area'],
        'Tipo':     at.get('unit_type', '-'),
        'Dia':      daily['date_from'],
        'Uncorr Gás (t)':   u.get('gas'),
        'Uncorr Óleo (t)':  u.get('oil'),
        'Uncorr HC (t)':    u.get('hc'),
        'Uncorr Água (t)':  u.get('water'),
        'Corr Gás (t)':     c.get('gas'),
        'Corr Óleo (t)':    c.get('oil'),
        'Corr HC (t)':      c.get('hc'),
        'Corr Água (t)':    c.get('water'),
        'PVT mass Gás (t)':  pm.get('gas'),
        'PVT mass Óleo (t)': pm.get('oil'),
        'PVT mass Água (t)': pm.get('water'),
        'PVT vol Gás (Sm³)': pv.get('gas'),
        'PVT vol Óleo (m³)': pv.get('oil'),
    }

# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────

CLR = {'hdr':'4472C4','white':'FFFFFF','alt':'EEF2FF',
       'ok':'C6EFCE','warn':'FFC7CE','partial':'FFEB9C',
       'nodata':'DDDDDD','sep':'2F5496'}

def _hdr(cell, bg='hdr'):
    cell.font      = Font(bold=True, color=CLR['white'], size=10)
    cell.fill      = PatternFill('solid', fgColor=CLR[bg])
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_excel(output_path, sheets_dict):
    # Prepara datas antes de escrever (números mantidos como float)
    sheets_prep = {name: _prep_dates(df) for name, df in sheets_dict.items()}

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for name, df in sheets_prep.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])

        wb  = writer.book
        thn = Side(style='thin', color='CCCCCC')
        brd = Border(left=thn, right=thn, top=thn, bottom=thn)

        for name, df in sheets_prep.items():
            ws = wb[name[:31]]
            for cell in ws[1]:
                _hdr(cell)
                cell.border = brd
            ws.row_dimensions[1].height = 34
            ws.freeze_panes = 'A2'

            # Mapeia colunas de data/hora pelo cabeçalho da linha 1
            date_cols  = set()
            dtime_cols = set()
            for cell in ws[1]:
                col_name = str(cell.value or '')
                if col_name in _DATE_COLS:
                    date_cols.add(cell.column)
                elif col_name in _DTIME_COLS:
                    dtime_cols.add(cell.column)

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                alt_fill = PatternFill('solid', fgColor=CLR['alt']) if row_idx % 2 == 0 else None
                for cell in row:
                    cell.border = brd
                    v = cell.value

                    # ── Formato data dd-mm-yyyy ───────────────────────────────
                    if cell.column in date_cols:
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # ── Formato datetime dd/mm/yyyy hh:mm ────────────────────
                    elif cell.column in dtime_cols:
                        cell.number_format = 'DD/MM/YYYY HH:MM'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # ── Float: número real com 4 casas decimais ──────────
                    elif isinstance(v, float) and v == v:  # not NaN
                        cell.number_format = '#,##0.0000'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif isinstance(v, int):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                    if alt_fill and not isinstance(v, str):
                        cell.fill = alt_fill
                    if v == 'VERIFICAR':
                        cell.fill = PatternFill('solid', fgColor=CLR['warn'])
                        cell.font = Font(bold=True, color='9C0006')
                    elif v == 'OK (24/24h)':
                        cell.fill = PatternFill('solid', fgColor=CLR['ok'])
                        cell.font = Font(bold=True, color='276221')
                    elif isinstance(v, str) and 'PARCIAL' in v:
                        cell.fill = PatternFill('solid', fgColor=CLR['partial'])
                        cell.font = Font(bold=True, color='9C6500')
                    elif v in ('SEM DADOS', 'SEM HORÁRIOS'):
                        cell.fill = PatternFill('solid', fgColor=CLR['nodata'])

            # Auto width — usa DD-MM-YYYY como referência para colunas de data
            for ci in range(1, ws.max_column + 1):
                ltr = get_column_letter(ci)
                if ci in date_cols:
                    ws.column_dimensions[ltr].width = 14
                elif ci in dtime_cols:
                    ws.column_dimensions[ltr].width = 18
                else:
                    vals = [str(ws.cell(r, ci).value or '') for r in range(1, min(ws.max_row+1, 300))]
                    ws.column_dimensions[ltr].width = min(max(len(t) for t in vals) + 3, 42)

    print(f'OK  {output_path}')

# ─────────────────────────────────────────────────────────────────────────────
# Separador de Testes – parser dos TXTs de CV (OLEO / GAS / AGUA)
# ─────────────────────────────────────────────────────────────────────────────

SEP_COLS_HOURLY = [
    'SEP Temperatura Méd. (°C)',
    'SEP Pressão Méd. (barg)',
    'SEP Óleo Vol. Bruto (m³) CV',
    'SEP Óleo (t) CV',
    'SEP Gás (t) CV',
    'SEP Água (t) CV',
    'SEP HC (t)',
    'SEP Total (t)',
    'Desvio HC (%)',
    'Desvio Total (%)',
]

def _parse_hourly_txt_oleo(path):
    """
    Lê Run_24Hours*_OLEO.txt (ou _AGUA.txt com mesmo formato).
    Retorna dict  {hour_int: {temp, pressure_barg, gv_m3, mass_t}, 'DAY': {...}}
    hour_int: 1..24  (24 = última hora, corresponde ao PDF com hour=0 do dia seguinte)
    """
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAY':
                # DAY  pressure  temperature  SD  MD  IV  GV  GSV  Mass  NSV  BSW  CPL  CTL
                if len(parts) >= 9:
                    result['DAY'] = {
                        'temp':          float(parts[2]),
                        'pressure_barg': float(parts[1]) / 100.0,
                        'gv_m3':         float(parts[6]),
                        'mass_t':        float(parts[8]),
                    }
            elif key.isdigit():
                h = int(key)
                if len(parts) >= 9:
                    result[h] = {
                        'temp':          float(parts[2]),
                        'pressure_barg': float(parts[1]) / 100.0,
                        'gv_m3':         float(parts[6]),
                        'mass_t':        float(parts[8]),
                    }
    return result


_SEP_TOKEN_SPLIT_DECIMALS = (5, 4, 3, 6, 2)


def _recover_concatenated_float_token(token):
    raw = str(token or '').strip()
    if raw.count('.') < 2:
        return None
    if not re.fullmatch(r'[+-]?\d+(?:\.\d+)+', raw):
        return None
    first_dot = raw.find('.')
    for frac_len in _SEP_TOKEN_SPLIT_DECIMALS:
        split_at = first_dot + 1 + frac_len
        if split_at <= first_dot + 1 or split_at >= len(raw):
            continue
        left = raw[:split_at]
        right = raw[split_at:]
        if not right or not right[0].isdigit() or right.count('.') != 1:
            continue
        if _is_number(left) and _is_number(right):
            return left, right
    return None


def _parse_sep_float_token(raw_value, *, file_path, row_key, field_name, line, trace_hook=None):
    token = str(raw_value or '').strip()
    try:
        return float(token)
    except ValueError:
        recovered = _recover_concatenated_float_token(token)
        if not recovered:
            raise
        primary_token, overflow_token = recovered
        value = float(primary_token)
        if trace_hook:
            trace_hook({
                'code': 'sep_parser_recovered_token',
                'file_path': str(file_path or ''),
                'row_key': str(row_key),
                'field_name': str(field_name),
                'raw_token': token,
                'recovered_token': primary_token,
                'overflow_token': overflow_token,
                'line': str(line or '').strip(),
            })
        return value

def _parse_hourly_txt_oleo(path, trace_hook=None):
    """
    Lê Run_24Hours*_OLEO.txt (ou _AGUA.txt com mesmo formato).
    Retorna dict  {hour_int: {temp, pressure_barg, gv_m3, mass_t}, 'DAY': {...}}
    hour_int: 1..24  (24 = última hora, corresponde ao PDF com hour=0 do dia seguinte)
    """
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAY':
                # DAY  pressure  temperature  SD  MD  IV  GV  GSV  Mass  NSV  BSW  CPL  CTL
                if len(parts) >= 9:
                    result['DAY'] = {
                        'temp': _parse_sep_float_token(parts[2], file_path=path, row_key=key, field_name='temp', line=line, trace_hook=trace_hook),
                        'pressure_barg': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_raw', line=line, trace_hook=trace_hook) / 100.0,
                        'gv_m3': _parse_sep_float_token(parts[6], file_path=path, row_key=key, field_name='gv_m3', line=line, trace_hook=trace_hook),
                        'mass_t': _parse_sep_float_token(parts[8], file_path=path, row_key=key, field_name='mass_t', line=line, trace_hook=trace_hook),
                    }
            elif key.isdigit():
                h = int(key)
                if len(parts) >= 9:
                    result[h] = {
                        'temp': _parse_sep_float_token(parts[2], file_path=path, row_key=key, field_name='temp', line=line, trace_hook=trace_hook),
                        'pressure_barg': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_raw', line=line, trace_hook=trace_hook) / 100.0,
                        'gv_m3': _parse_sep_float_token(parts[6], file_path=path, row_key=key, field_name='gv_m3', line=line, trace_hook=trace_hook),
                        'mass_t': _parse_sep_float_token(parts[8], file_path=path, row_key=key, field_name='mass_t', line=line, trace_hook=trace_hook),
                    }
    return result


def _parse_hourly_txt_gas(path, trace_hook=None):
    """
    Lê Run_24Hours*_GAS.txt.
    Retorna dict {hour_int: {mass_t}, 'DAY': {mass_t}}
    ATENÇÃO: o campo Mass no arquivo de gás está em kg (apesar do header dizer 't').
             Divisão por 1000 é feita aqui para converter para toneladas.
    """
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAILY':
                # DAILY  [empty cols]  GrVol  StVol  Mass  Energy  [DiffPress]  Flowtime
                nums = [p for p in parts[1:] if _is_number(p)]
                if len(nums) >= 3:
                    result['DAY'] = {
                        'mass_t': _parse_sep_float_token(nums[2], file_path=path, row_key=key, field_name='mass_raw', line=line, trace_hook=trace_hook) / 1000.0
                    }
            elif key.isdigit():
                h = int(key)
                # Hour  Pressure  Temperature  SD  DT  GrVol  StVol  Mass  Energy  DiffPress  Flowtime
                if len(parts) >= 8:
                    result[h] = {
                        'mass_t': _parse_sep_float_token(parts[7], file_path=path, row_key=key, field_name='mass_raw', line=line, trace_hook=trace_hook) / 1000.0
                    }
    return result

def _is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def parse_sep_txt_set(oleo_path, gas_path, agua_path, density_sim=None, trace_hook=None):
    """
    Lê os 3 arquivos TXT do separador de testes e monta um dict unificado:
      {hour: {temp, pressure_barg, oil_m3, oil_t, gas_t, water_t, hc_t, total_t}, 'DAY': {...}}

    NOTA: usa mass_t diretamente do campo Mass do TXT (já em toneladas pelo medidor).
    Parâmetro density_sim mantido apenas por compatibilidade — não é mais utilizado.
    """
    oleo_data = _parse_hourly_txt_oleo(oleo_path, trace_hook=trace_hook)
    gas_data  = _parse_hourly_txt_gas(gas_path, trace_hook=trace_hook)
    agua_data = _parse_hourly_txt_oleo(agua_path, trace_hook=trace_hook)   # mesmo formato

    combined = {}
    all_keys = sorted(set(list(oleo_data.keys()) + list(gas_data.keys()) + list(agua_data.keys())),
                      key=lambda x: 999 if x == 'DAY' else x)

    for k in all_keys:
        o = oleo_data.get(k, {})
        g = gas_data.get(k, {})
        w = agua_data.get(k, {})

        oil_m3 = o.get('gv_m3',   np.nan)
        oil_t  = o.get('mass_t',  np.nan)   # direto do TXT, sem conversão por densidade
        gas_t  = g.get('mass_t',  np.nan)
        wat_t  = w.get('mass_t',  np.nan)

        def _s(a, b):
            if np.isnan(a) or np.isnan(b): return np.nan
            return a + b

        hc_t    = _s(oil_t, gas_t)
        total_t = _s(hc_t,  wat_t)

        combined[k] = {
            'temp':          o.get('temp',          np.nan),
            'pressure_barg': o.get('pressure_barg', np.nan),
            'oil_m3':        oil_m3,
            'oil_t':         oil_t,
            'gas_t':         gas_t,
            'water_t':       wat_t,
            'hc_t':          hc_t,
            'total_t':       total_t,
        }
    return combined

def _desvio(mpfm_val, sep_val):
    """Calcula desvio percentual entre MPFM e Separador."""
    try:
        if mpfm_val is None or sep_val is None: return np.nan
        if np.isnan(float(mpfm_val)) or np.isnan(float(sep_val)): return np.nan
        if float(sep_val) == 0: return np.nan
        return round((float(mpfm_val) - float(sep_val)) / float(sep_val) * 100, 2)
    except Exception:
        return np.nan

def _sep_row(sep_hour, mpfm_hc_t, mpfm_total_t):
    """
    Monta os valores das colunas SEP_COLS_HOURLY para uma hora.
    sep_hour: dict do parse_sep_txt_set para essa hora (ou None).
    Retorna lista de 10 valores (SEP_COLS_HOURLY).
    """
    if not sep_hour:
        return [np.nan] * 10
    oil_t = sep_hour.get('oil_t',   np.nan)
    gas_t = sep_hour.get('gas_t',   np.nan)
    wat_t = sep_hour.get('water_t', np.nan)
    hc_t  = sep_hour.get('hc_t',   np.nan)
    tot_t = sep_hour.get('total_t', np.nan)
    def _r(v): return round(v, 4) if (v is not None and not (isinstance(v, float) and np.isnan(v))) else np.nan
    return [
        sep_hour.get('temp',          np.nan),
        sep_hour.get('pressure_barg', np.nan),
        sep_hour.get('oil_m3',        np.nan),
        _r(oil_t),
        _r(gas_t),
        _r(wat_t),
        _r(hc_t),
        _r(tot_t),
        _desvio(mpfm_hc_t, hc_t),
        _desvio(mpfm_total_t, tot_t),
    ]

def build_hourly_df_with_sep(hourly_records, unit_code, sep_data=None):
    """
    Extensão de build_hourly_df que acrescenta colunas do Separador de Testes
    quando sep_data (resultado de parse_sep_txt_set) é fornecido.
    Mantém TODOS os campos originais inalterados — apenas adiciona colunas ao final.
    """
    cols = HOURLY_COLS + (SEP_COLS_HOURLY if sep_data else [])
    rows = []
    for rec in sorted(hourly_records, key=lambda r: r['hour'] or 0):
        for tag, td in rec['tags'].items():
            base = [rec['date_from'], rec['hour'],
                    unit_code, rec['fpso_side'], rec['unit_type'],
                    tag, td['instrument'], rec['dt_from'], rec['dt_to']]
            mpfm_vals = _mrow(td['metrics'])
            row = base + mpfm_vals + [os.path.basename(rec['pdf_path'])]

            if sep_data:
                # TXT hora 24 = PDF hour 0 (última hora do dia)
                pdf_hour = rec['hour']
                txt_key  = 24 if pdf_hour == 0 else pdf_hour
                sh       = sep_data.get(txt_key)
                # MPFM corrected HC e Total para cálculo dos desvios
                c = td['metrics']['mpfm_corr']
                mpfm_hc  = c.get('hc')
                mpfm_tot = c.get('total')
                row += _sep_row(sh, mpfm_hc, mpfm_tot)

            rows.append(row)
    return pd.DataFrame(rows, columns=cols)

# ─────────────────────────────────────────────────────────────────────────────
# Public runner — chamado externamente
# ─────────────────────────────────────────────────────────────────────────────

def run(units_cfg, output_path):
    """
    units_cfg: lista de dicts com chaves:
      unit_code    str   ex: 'B10'
      daily_pdf    str   caminho do Daily PDF
      hourly_glob  str   glob para os hourly PDFs (pode ser '' ou None)
      sep_oleo_txt str   (opcional) Run_24Hours*_OLEO.txt
      sep_gas_txt  str   (opcional) Run_24Hours*_GAS.txt
      sep_agua_txt str   (opcional) Run_24Hours*_AGUA.txt
      sep_density  float (opcional) densidade simulação kg/m³ (default 790.78)
    output_path: caminho do Excel de saída
    """
    sheets    = {}
    area_rows = []

    for cfg in units_cfg:
        code   = cfg['unit_code']
        daily  = parse_pdf(cfg['daily_pdf'], 'daily')
        hourly_pdfs = sorted(glob.glob(cfg.get('hourly_glob') or '')) if cfg.get('hourly_glob') else []
        hourly = [parse_pdf(p, 'hourly') for p in hourly_pdfs]

        # ── Separador de Testes (opcional) ──────────────────────────────────
        sep_data = None
        if cfg.get('sep_oleo_txt') and cfg.get('sep_gas_txt') and cfg.get('sep_agua_txt'):
            density  = cfg.get('sep_density', 790.78)
            sep_data = parse_sep_txt_set(
                cfg['sep_oleo_txt'], cfg['sep_gas_txt'], cfg['sep_agua_txt'],
                density_sim=density
            )
            sep_hrs = [k for k in sep_data if k != 'DAY']
            print(f'\n📦 {code}  ({daily["fpso_side"]} – {daily["unit_type"]})')
            print(f'   Daily   : {os.path.basename(cfg["daily_pdf"])}')
            print(f'   Hourly  : {len(hourly)} arquivo(s)')
            print(f'   TAGs    : {list(daily["tags"].keys())}')
            print(f'   SEP TXT : {len(sep_hrs)} horas | density={density} kg/m³')
        else:
            print(f'\n📦 {code}  ({daily["fpso_side"]} – {daily["unit_type"]})')
            print(f'   Daily  : {os.path.basename(cfg["daily_pdf"])}')
            print(f'   Hourly : {len(hourly)} arquivo(s)')
            print(f'   TAGs   : {list(daily["tags"].keys())}')

        if hourly:
            hrs = sorted(r["hour"] for r in hourly if r["hour"] is not None)
            print(f'   Horas  : {hrs}')

        day_tag = (daily['date_from'] or '00-00-00')[8:10] + '_' + (daily['date_from'] or '00-00-00')[5:7]

        # HOURLY — usa versão com sep se disponível
        if hourly:
            sheets[f'HOURLY_{code}_{day_tag}'] = build_hourly_df_with_sep(hourly, code, sep_data)
        else:
            sheets[f'HOURLY_{code}_{day_tag}'] = pd.DataFrame(
                columns=HOURLY_COLS + (SEP_COLS_HOURLY if sep_data else [])
            )

        sheets[f'DAILY_{code}_{day_tag}'] = build_daily_df(daily, code)
        sheets[f'RECON_{code}_{day_tag}'] = build_recon_df(daily, hourly, code)

        at_row = build_area_totals_row(daily, code)
        if at_row:
            area_rows.append(at_row)

    if area_rows:
        sheets['AREA_TOTALS'] = pd.DataFrame(area_rows)

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    write_excel(output_path, sheets)

    print(f'\n📊 Abas geradas:')
    for n, df in sheets.items():
        print(f'   {n}: {len(df)} linhas × {len(df.columns)} cols')

    return sheets

# ─────────────────────────────────────────────────────────────────────────────
# __main__  — execução direta
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Monthly base manager — processa config JSON e atualiza Excel incremental
# ─────────────────────────────────────────────────────────────────────────────

import json
from datetime import datetime
from openpyxl import load_workbook as _load_wb

def load_config(config_path):
    with open(config_path, encoding="utf-8") as f:
        return json.load(f)

def save_config(cfg_dict, config_path):
    os.makedirs(os.path.dirname(config_path) or ".", exist_ok=True)
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(cfg_dict, f, indent=2, ensure_ascii=False, default=str)

def run_incremental(config_path, base_output_dir="/mnt/user-data/outputs"):
    cfg             = load_config(config_path)
    month_label     = cfg.get("month_label", "UNKNOWN")
    output_filename = cfg.get("output_xlsx", f"MPFM_{month_label}.xlsx")
    output_path     = os.path.join(base_output_dir, output_filename)
    processed_jobs  = set(cfg.get("processed_jobs", []))
    jobs            = cfg.get("jobs", [])

    print(f"\nMes: {month_label}  |  Output: {output_path}")
    print(f"OK Ja processados: {sorted(processed_jobs) or 'nenhum'}")

    new_sheets    = {}
    new_area_rows = []

    for job in jobs:
        daily_path = job["daily_pdf"]
        if "*" in daily_path:
            matches = sorted(glob.glob(daily_path))
            daily_pdfs = matches if matches else []
        else:
            daily_pdfs = [daily_path] if os.path.exists(daily_path) else []

        if not daily_pdfs:
            print(f"   WARN  Daily PDF nao encontrado: {daily_path}")
            continue

        for daily_pdf in daily_pdfs:
            rec     = parse_pdf(daily_pdf, "daily")
            day_tag = (rec["date_from"] or "00-00-00")[8:10] + "_" + (rec["date_from"] or "00-00-00")[5:7]
            job_id  = f'{job["unit_code"]}_{day_tag}'

            if job_id in processed_jobs:
                print(f"   ⏭️  {job_id} já processado, pulando.")
                continue

            # Build unit config for this day
            unit_cfg = dict(job)
            unit_cfg["daily_pdf"] = daily_pdf
            for key in ("sep_oleo_txt", "sep_gas_txt", "sep_agua_txt"):
                val = job.get(key, "")
                if val and "*" in val:
                    m = sorted(glob.glob(val))
                    unit_cfg[key] = m[0] if m else ""

            code   = unit_cfg["unit_code"]
            daily  = parse_pdf(daily_pdf, "daily")
            hourly_pdfs = sorted(glob.glob(unit_cfg.get("hourly_glob") or "")) if unit_cfg.get("hourly_glob") else []
            hourly = [parse_pdf(p, "hourly") for p in hourly_pdfs]

            sep_data = None
            if (unit_cfg.get("sep_oleo_txt") and unit_cfg.get("sep_gas_txt") and
                    unit_cfg.get("sep_agua_txt") and
                    os.path.exists(unit_cfg["sep_oleo_txt"]) and
                    os.path.exists(unit_cfg["sep_gas_txt"]) and
                    os.path.exists(unit_cfg["sep_agua_txt"])):
                sep_data = parse_sep_txt_set(
                    unit_cfg["sep_oleo_txt"], unit_cfg["sep_gas_txt"], unit_cfg["sep_agua_txt"],
                    density_sim=unit_cfg.get("sep_density", 790.78)
                )

            print(f"\n   PACK {code} - {day_tag}  ({daily['fpso_side']} - {daily['unit_type']})")
            print(f"      Hourly: {len(hourly)}h  |  SEP: {'OK' if sep_data else '-'}")

            if hourly:
                new_sheets[f"HOURLY_{code}_{day_tag}"] = build_hourly_df_with_sep(hourly, code, sep_data)
            else:
                new_sheets[f"HOURLY_{code}_{day_tag}"] = pd.DataFrame(
                    columns=HOURLY_COLS + (SEP_COLS_HOURLY if sep_data else [])
                )
            new_sheets[f"DAILY_{code}_{day_tag}"]  = build_daily_df(daily, code)
            new_sheets[f"RECON_{code}_{day_tag}"]  = build_recon_df(daily, hourly, code)

            at = build_area_totals_row(daily, code)
            if at:
                at["job_id"] = job_id
                new_area_rows.append(at)

            processed_jobs.add(job_id)

    if not new_sheets:
        print("\nOK Nenhum job novo.")
        return cfg

    os.makedirs(base_output_dir, exist_ok=True)
    _merge_excel(output_path, new_sheets, new_area_rows)

    cfg["processed_jobs"] = sorted(processed_jobs)
    cfg["last_run"]       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_config(cfg, config_path)

    print(f"\nOK Excel atualizado: {output_path}")
    print(f"   Novas abas: {list(new_sheets.keys())}")
    return cfg

def _apply_sheet_styles(wb, sheet_names):
    thn = Side(style="thin", color="CCCCCC")
    brd = Border(left=thn, right=thn, top=thn, bottom=thn)
    for name in sheet_names:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for cell in ws[1]:
            cell.font  = Font(bold=True, color=CLR["white"], size=10)
            cell.fill  = PatternFill("solid", fgColor=CLR["hdr"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
        ws.row_dimensions[1].height = 34
        ws.freeze_panes = "A2"
        for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
            alt = PatternFill("solid", fgColor=CLR["alt"]) if ri % 2 == 0 else None
            for cell in row:
                cell.border = brd
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(cell.value, (int, float)) else "left",
                    vertical="center"
                )
                if alt and not isinstance(cell.value, str):
                    cell.fill = alt
                v = cell.value
                if v == "VERIFICAR":
                    cell.fill = PatternFill("solid", fgColor=CLR["warn"])
                    cell.font = Font(bold=True, color="9C0006")
                elif v == "OK (24/24h)":
                    cell.fill = PatternFill("solid", fgColor=CLR["ok"])
                    cell.font = Font(bold=True, color="276221")
                elif isinstance(v, str) and "PARCIAL" in v:
                    cell.fill = PatternFill("solid", fgColor=CLR["partial"])
                    cell.font = Font(bold=True, color="9C6500")
        for ci in range(1, ws.max_column + 1):
            ltr  = get_column_letter(ci)
            vals = [str(ws.cell(r, ci).value or "") for r in range(1, min(ws.max_row + 1, 300))]
            ws.column_dimensions[ltr].width = min(max(len(t) for t in vals) + 3, 42)

def _merge_excel(output_path, new_sheets, new_area_rows):
    if os.path.exists(output_path):
        from openpyxl import load_workbook as lw
        dest_wb = None
        src_wb = None
        tmp = output_path + ".tmp.xlsx"
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
            dest_wb = lw(output_path)
            # Write new sheets to temp, style them, then copy into dest
            with pd.ExcelWriter(tmp, engine="openpyxl") as writer:
                for name, df in new_sheets.items():
                    df.to_excel(writer, index=False, sheet_name=name[:31])
            src_wb = lw(tmp)
            _apply_sheet_styles(src_wb, list(new_sheets.keys()))
            src_wb.save(tmp)
            src_wb.close()
            src_wb = lw(tmp)
            for sname in src_wb.sheetnames:
                if sname in dest_wb.sheetnames:
                    del dest_wb[sname]
                src_ws = src_wb[sname]
                dest_ws = dest_wb.create_sheet(sname)
                # Copia valor E estilos para manter formatação azul em todas as abas
                for row in src_ws.iter_rows():
                    for cell in row:
                        nc = dest_ws.cell(cell.row, cell.column, cell.value)
                        if cell.has_style:
                            nc.font      = cell.font.copy()
                            nc.fill      = cell.fill.copy()
                            nc.border    = cell.border.copy()
                            nc.alignment = cell.alignment.copy()
                # Copia altura das linhas
                for rn, rd in src_ws.row_dimensions.items():
                    dest_ws.row_dimensions[rn].height = rd.height
                # Copia largura das colunas
                for col_letter, cd in src_ws.column_dimensions.items():
                    dest_ws.column_dimensions[col_letter].width = cd.width
                # Freeze panes
                if src_ws.freeze_panes:
                    dest_ws.freeze_panes = src_ws.freeze_panes
            _update_area_totals_wb(dest_wb, new_area_rows)
            dest_wb.save(output_path)
        finally:
            if src_wb is not None:
                try:
                    src_wb.close()
                except Exception:
                    pass
            if dest_wb is not None:
                try:
                    dest_wb.close()
                except Exception:
                    pass
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except Exception:
                    pass
    else:
        clean = [{k: v for k, v in r.items() if k != "job_id"} for r in new_area_rows]
        all_s = dict(new_sheets)
        if clean:
            all_s["AREA_TOTALS"] = pd.DataFrame(clean)
        write_excel(output_path, all_s)

def _update_area_totals_wb(wb, new_area_rows):
    if not new_area_rows:
        return
    clean  = [{k: v for k, v in r.items() if k != "job_id"} for r in new_area_rows]
    new_df = pd.DataFrame(clean)
    if "AREA_TOTALS" in wb.sheetnames:
        ws   = wb["AREA_TOTALS"]
        cols = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            row = {cols[c]: ws.cell(r, c + 1).value for c in range(len(cols))}
            if any(v is not None for v in row.values()):
                rows.append(row)
        exist = pd.DataFrame(rows)
        combined = pd.concat([exist, new_df], ignore_index=True).drop_duplicates(
            subset=["Bank", "Dia"], keep="last"
        )
        del wb["AREA_TOTALS"]
    else:
        combined = new_df
    ws = wb.create_sheet("AREA_TOTALS")
    for ci, col in enumerate(combined.columns, 1):
        ws.cell(1, ci, col)
    for data_ri, (_, row) in enumerate(combined.iterrows(), start=2):
        for ci, v in enumerate(row, 1):
            ws.cell(data_ri, ci, v)
    _apply_sheet_styles(wb, ["AREA_TOTALS"])



def build_status_sheet(state: dict) -> "pd.DataFrame":
    """
    Gera aba STATUS_MES mostrando o que foi recebido e o que falta por dia.
    Requerido por dia: 24h Hourly + 1 Daily + SEP (3 TXTs)
    """
    import pandas as pd, calendar
    yr  = int(state.get('yr',  0))
    mo  = int(state.get('mo',  0))
    if not yr or not mo:
        return pd.DataFrame()

    n_days = calendar.monthrange(yr, mo)[1]
    processed = state.get('processed', [])  # lista de keys tipo 'B10_04_03'

    # Decompõe keys processados
    daily_done  = set()
    hourly_done = {}  # day_tag → set of hours
    sep_done    = set()

    for key in processed:
        parts = key.split('_')
        if len(parts) < 3: continue
        unit = parts[0]
        day_tag = '_'.join(parts[1:])   # '04_03'
        day_done_key = day_tag
        daily_done.add(day_done_key)

    # Reconstrói horas a partir do estado (guardamos no processed somente a key do dia)
    # Para horas precisamos de info adicional — usamos processed_hours do state
    processed_hours = state.get('processed_hours', {})  # day_tag → [lista horas]

    rows = []
    for d in range(1, n_days + 1):
        day_tag = f'{d:02d}_{mo:02d}'
        h_list  = processed_hours.get(day_tag, [])
        n_h     = len(h_list)
        has_day = day_tag in daily_done
        has_sep = day_tag in state.get('sep_days', [])

        h_ok    = n_h == 24
        all_ok  = h_ok and has_day

        missing_h = sorted(set(range(24)) - set(h_list)) if not h_ok else []

        rows.append({
            'Dia':          f'{d:02d}/{mo:02d}/{yr}',
            'Hourly (h)':   n_h,
            'Daily':        'OK' if has_day  else '-',
            'SEP':          'OK' if has_sep  else '-',
            'Status':       'COMPLETO' if all_ok else ('PARCIAL' if (n_h > 0 or has_day) else 'PENDENTE'),
            'Horas faltando': ', '.join(str(h) for h in missing_h) if missing_h else '',
        })

    return pd.DataFrame(rows)


def _cli():
    import sys
    if len(sys.argv) >= 3 and sys.argv[1] == '--config':
        config_path = sys.argv[2]
        base_dir    = sys.argv[3] if len(sys.argv) >= 4 else '/mnt/user-data/outputs'
        print(f'\n🚀 Modo incremental — config: {config_path}')
        cfg = run_incremental(config_path, base_dir)
        print(f'\n📋 Processados: {cfg["processed_jobs"]}')
    else:
        print('Uso: python3 mpfm_engine.py --config /path/config.json [output_dir]')

if __name__ == '__main__':
    import sys
    if '--config' in sys.argv:
        _cli()
    else:
        # Original __main__ block
        BASE_24  = '/home/claude/zip_contents/24-01'
        BASE_B10 = '/home/claude/zip_b10'
        BASE_SEP = '/home/claude/mpfm_dados/SEP_TESTE_CV'
        B08_DAILY = '/mnt/user-data/uploads/B08_MPFM_Daily-20260223-000000_0000.pdf'

        UNITS = [
            {'unit_code':'B03','daily_pdf':f'{BASE_24}/B03_MPFM_Daily-20260125-000000+0000.pdf','hourly_glob':f'{BASE_24}/B03_MPFM_Hourly-20260124-*.pdf'},
            {'unit_code':'B05','daily_pdf':f'{BASE_24}/B05_MPFM_Daily-20260125-000000+0000.pdf','hourly_glob':f'{BASE_24}/B05_MPFM_Hourly-20260124-*.pdf'},
            {'unit_code':'B08','daily_pdf':B08_DAILY,'hourly_glob':None},
            {'unit_code':'B10','daily_pdf':f'{BASE_B10}/MPFM_Daily_FCS/B10_MPFM_Daily-20260302-000000+0000.pdf','hourly_glob':f'{BASE_B10}/MPFM_Hourly_FCS/B10_MPFM_Hourly-20260301-*.pdf','sep_oleo_txt':f'{BASE_SEP}/Run_24Hours1-1.20260302000000_OLEO.txt','sep_gas_txt':f'{BASE_SEP}/Run_24Hours2-1.20260302000000_GAS.txt','sep_agua_txt':f'{BASE_SEP}/Run_24Hours1-1.20260302000000_AGUA.txt','sep_density':790.78},
        ]
        OUTPUT = '/mnt/user-data/outputs/MPFM_ALL_UNITS_RECONCILIATION.xlsx'
        sheets = run(UNITS, OUTPUT)
def _cli():
    import sys
    if len(sys.argv) >= 3 and sys.argv[1] == '--config':
        config_path = sys.argv[2]
        base_dir    = sys.argv[3] if len(sys.argv) >= 4 else '/mnt/user-data/outputs'
        print(f'\n🚀 Modo incremental — config: {config_path}')
        cfg = run_incremental(config_path, base_dir)
        print(f'\n📋 Processados: {cfg["processed_jobs"]}')
    else:
        print('Uso: python3 mpfm_engine.py --config /path/config.json [output_dir]')

if __name__ == '__main__':
    import sys
    if '--config' in sys.argv:
        _cli()
    else:
        # Original __main__ block
        BASE_24  = '/home/claude/zip_contents/24-01'
        BASE_B10 = '/home/claude/zip_b10'
        BASE_SEP = '/home/claude/mpfm_dados/SEP_TESTE_CV'
        B08_DAILY = '/mnt/user-data/uploads/B08_MPFM_Daily-20260223-000000_0000.pdf'

        UNITS = [
            {'unit_code':'B03','daily_pdf':f'{BASE_24}/B03_MPFM_Daily-20260125-000000+0000.pdf','hourly_glob':f'{BASE_24}/B03_MPFM_Hourly-20260124-*.pdf'},
            {'unit_code':'B05','daily_pdf':f'{BASE_24}/B05_MPFM_Daily-20260125-000000+0000.pdf','hourly_glob':f'{BASE_24}/B05_MPFM_Hourly-20260124-*.pdf'},
            {'unit_code':'B08','daily_pdf':B08_DAILY,'hourly_glob':None},
            {'unit_code':'B10','daily_pdf':f'{BASE_B10}/MPFM_Daily_FCS/B10_MPFM_Daily-20260302-000000+0000.pdf','hourly_glob':f'{BASE_B10}/MPFM_Hourly_FCS/B10_MPFM_Hourly-20260301-*.pdf','sep_oleo_txt':f'{BASE_SEP}/Run_24Hours1-1.20260302000000_OLEO.txt','sep_gas_txt':f'{BASE_SEP}/Run_24Hours2-1.20260302000000_GAS.txt','sep_agua_txt':f'{BASE_SEP}/Run_24Hours1-1.20260302000000_AGUA.txt','sep_density':790.78},
        ]
        OUTPUT = '/mnt/user-data/outputs/MPFM_ALL_UNITS_RECONCILIATION.xlsx'
        sheets = run(UNITS, OUTPUT)
