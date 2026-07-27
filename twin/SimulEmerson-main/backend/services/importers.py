"""Twin MPFM — Importadores de dados externos (planilhas MPFM mensais .xlsx).

Refator (Iteração 4): função monolítica de CC=19 → helpers menores e testáveis.
Sem mudança de comportamento.
"""
from __future__ import annotations
from io import BytesIO
from typing import Any, Dict, List, Tuple


def _find_header_row(rows: List[Tuple[Any, ...]]) -> int:
    """Procura linha de cabeçalho nas primeiras 20 linhas."""
    for idx, row in enumerate(rows[:20]):
        joined = "|".join(str(c or "") for c in row).lower()
        if "productiondate" in joined or ("entity" in joined and "tag" in joined):
            return idx
    return 0


def _normalize_headers(row: Tuple[Any, ...]) -> List[str]:
    """Garante header não-vazio para cada coluna."""
    return [
        str(c).strip() if c is not None else f"col_{i}"
        for i, c in enumerate(row)
    ]


def _row_is_empty(row: Tuple[Any, ...]) -> bool:
    return not any(cell is not None and str(cell).strip() != "" for cell in row)


def _cell_to_serializable(value: Any) -> Any:
    """Converte datetime/date em ISO string; demais valores ficam como estão."""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _parse_row(row: Tuple[Any, ...], headers: List[str]) -> Dict[str, Any]:
    """Mapeia uma linha em dict {header: cell}."""
    return {
        (headers[i] if i < len(headers) and headers[i] else f"col_{i}"):
            _cell_to_serializable(row[i]) if i < len(row) else None
        for i in range(len(headers))
    }


def import_mpfm_xlsx(content: bytes, max_rows: int = 5000) -> List[Dict[str, Any]]:
    """Lê planilha MPFM .xlsx/.xlsm e devolve lista de registros normalizados."""
    from openpyxl import load_workbook  # local import para acelerar boot
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = _find_header_row(rows)
    headers = _normalize_headers(rows[header_idx])

    records: List[Dict[str, Any]] = []
    for row in rows[header_idx + 1: header_idx + 1 + max_rows]:
        if _row_is_empty(row):
            continue
        records.append(_parse_row(row, headers))
    return records
