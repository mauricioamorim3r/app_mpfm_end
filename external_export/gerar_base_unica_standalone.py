#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
Gerador Base_Unica — versão 100% AUTOCONTIDA (sem depender do repositório
app_mpfm_end, do banco de dados ou do server.py).

Este único arquivo contém tudo que precisa para rodar em qualquer máquina:
  - parsing dos PDFs MPFM (Daily/Hourly)         -> copiado de mpfm_engine.py
  - parsing dos TXT do Separador de Testes (SEP) -> copiado de mpfm_engine.py
  - localização dos TXT do SEP por Meter ID      -> lógica própria, simples
  - montagem das linhas no formato Base_Unica    -> lógica própria
    - atualização incremental de uma Base_Unica total consolidada

DEPENDÊNCIAS PYTHON (instale com pip se a máquina não tiver):
    pip install pandas numpy pypdf openpyxl

COMO CONFIGURAR OS CAMINHOS
----------------------------
Existem 3 formas (use a que preferir):

1) Editar as constantes na seção "CONFIG" logo abaixo (mais simples).
2) Passar por linha de comando, ex.:
     python gerar_base_unica_standalone.py --mpfm-root "D:\Relatorios\MPFM" ^
         --sep-root "D:\Relatorios\SEP" --output "D:\saida.xlsx" --days 5 ^
         --aligned-bank B10 --master-output "D:\BASE_UNICA_TOTAL.xlsx"
3) Se não configurar nem passar argumentos, o script pergunta os caminhos
   interativamente (input) na hora de rodar.

BASE_UNICA_TOTAL INCREMENTAL
----------------------------
Por padrão, depois de gerar o Excel individual da análise, o script também
atualiza `BASE_UNICA_TOTAL.xlsx` ao lado deste arquivo. A integração é
incremental: histórico + nova análise são combinados e linhas repetidas são
substituídas pela versão mais recente usando a chave técnica:
ProductionDate, Hour, Granularity, Origin, SourceType, Bank, Entity, Tag,
Instrumento. Use `--no-master` para desativar ou `--master-output` para escolher
outro caminho.

ESTRUTURA DE PASTAS ESPERADA
-----------------------------
MPFM_ROOT/
  <pasta do banco B03>/<ano>/<NN. NomeMes>/Daily/B03_MPFM_Daily-YYYYMMDD-......pdf
  <pasta do banco B03>/<ano>/<NN. NomeMes>/Hourly/B03_MPFM_Hourly-YYYYMMDD-HHMMSS...pdf
  ... (um subdiretório por banco, ver BANK_FOLDERS abaixo)

SEP_ROOT/
  <ano>/<NN. NomeMes>/FPSO-Bacalhau_Daily reports_YYYY-MM-DD/01 - CV_Reports/FC13/Run_24Hours*.txt
  ... (idem para FC14 e FC17)

Se a estrutura de pastas da sua instalação for diferente, ajuste as funções
`bank_month_dir()` e `sep_day_dir()` na seção "CAMINHOS DE PASTA" abaixo.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from pypdf import PdfReader

# ═════════════════════════════════════════════════════════════════════════
# CONFIG — ajuste aqui (ou use os argumentos de linha de comando / prompts)
# ═════════════════════════════════════════════════════════════════════════

# Deixe em branco ("") para o script perguntar interativamente ao rodar.
MPFM_ROOT = r""
SEP_ROOT = r""

# Nome da subpasta de cada banco dentro de MPFM_ROOT.
BANK_FOLDERS = {
    "B03": "3.1.1_13-FT-0367 Riser P5 - Topside B03",
    "B08": "3.1.2_13-FT-0167 Riser P2 - Topside B08",
    "B13": "3.1.3_13-FT-0317 Riser P4 - Topside B13",
    "B05": "3.1.4_18-FT-1506 PE 4 e PE_EO105 - Subsea B05",
    "B10": "3.1.5_18-FT-0506 PE 2 - Subsea B10",
    "B15": "3.1.6_18-FT-1106 PW_104DA - Subsea B15",
}

DAYS_COUNT = 5              # quantos dias (mais recentes disponíveis) exportar
SEP_ALIGNED_BANK = "B10"    # banco ao qual o SEP deve ser mesclado na mesma linha
OUTPUT_PATH = r""           # deixe em branco para salvar ao lado deste script
MASTER_OUTPUT_PATH = r""    # deixe em branco para salvar BASE_UNICA_TOTAL.xlsx ao lado deste script
UPDATE_MASTER = True         # atualiza arquivo incremental total após cada execução
MONTHS_LOOKBACK = 12         # meses máximos pesquisados para encontrar os dias disponíveis
PDF_WORKERS = 4              # processos paralelos para PDFs em exportações históricas

MONTH_PT = {
    1: "01. Janeiro", 2: "02. Fevereiro", 3: "03. Março", 4: "04. Abril",
    5: "05. Maio", 6: "06. Junho", 7: "07. Julho", 8: "08. Agosto",
    9: "09. Setembro", 10: "10. Outubro", 11: "11. Novembro", 12: "12. Dezembro",
}

# Meter ID -> fluido, usado para identificar os TXT do Separador de Testes
TARGET_PHASE_BY_METER = {
    "20FT0247": "oleo",
    "20FT0251": "agua",
    "20FT0244": "gas",
}
SEP_FC_FOLDERS = ["FC13", "FC14", "FC17"]


# ═════════════════════════════════════════════════════════════════════════
# MPFM ENGINE — copiado/adaptado de mpfm_engine.py (parsing de PDF e TXT)
# ═════════════════════════════════════════════════════════════════════════

def _to_float(x):
    x = str(x).strip().replace('\u00A0', '').replace(' ', '')
    if x in ('-', '', 'None', 'nan', 'NaN'):
        return np.nan
    try:
        return float(x.replace(',', '.'))
    except Exception:
        return np.nan


def _flat_text(pdf_path):
    reader = PdfReader(pdf_path)
    raw = '\n'.join(page.extract_text() or '' for page in reader.pages)
    return re.sub(r'\s+', ' ', raw)


NUM = r'([\-\d]+(?:\.\d+)?)'
C5 = ['gas', 'oil', 'hc', 'water', 'total']
C4 = ['gas', 'oil', 'hc', 'water']


def _nan_dict(keys):
    return {k: np.nan for k in keys}


def _parse_row(label_pat, text, ncols=5):
    cols = (C5 if ncols == 5 else C4)[:ncols]
    m = re.search(
        label_pat + r'[^\[]*\[(?:t|Sm[³3]?)\]\s*' + r'\s+'.join([NUM] * ncols),
        text, re.IGNORECASE
    )
    if m:
        return {c: _to_float(m.group(i + 1)) for i, c in enumerate(cols)}
    return _nan_dict(cols)


def _parse_fwa(text):
    """Extrai Pressão, Temperatura e Densidades (Flow Weighted Averages)."""
    def _find_meter_value(label, unit_pat):
        m = re.search(
            label + r'\s*\n?\s*' + unit_pat +
            r'\s*\n(?:-[^\n]*\n){0,2}\s*(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
        m = re.search(
            label + r'\s*' + unit_pat + r'(?:\s*-\s*){1,3}(\d[\d.]*)',
            text, re.IGNORECASE)
        if m:
            return _to_float(m.group(1))
        return np.nan

    def _extract_density(text):
        m = re.search(
            r'Density\s*\[kg/m\S*\]\s*(.*?)'
            r'(?=Production\s+Total|PE_|Riser\s+P|\Z)',
            text, re.DOTALL | re.IGNORECASE
        )
        if not m:
            return np.nan, np.nan, np.nan
        tokens = []
        for part in re.split(r'\s+', m.group(1).strip()):
            if re.fullmatch(r'-|\d+(?:\.\d+)?', part):
                tokens.append(part)
                if len(tokens) == 4:
                    break
            elif part:
                break

        def t2f(t):
            return np.nan if t == '-' else _to_float(t)
        gas = t2f(tokens[0]) if len(tokens) > 0 else np.nan
        oil = t2f(tokens[1]) if len(tokens) > 1 else np.nan
        water = t2f(tokens[3]) if len(tokens) > 3 else np.nan
        return gas, oil, water

    pres = _find_meter_value(r'Pressure', r'\[barg\]')
    temp = _find_meter_value(r'Temperature', r'\[.{1,3}C\]')
    dens_gas, dens_oil, dens_water = _extract_density(text)
    return {
        'pressure': pres, 'temperature': temp,
        'dens_gas': dens_gas, 'dens_oil': dens_oil, 'dens_water': dens_water,
    }


def _parse_metrics(block, ncols=5):
    corr = _parse_row(r'MPFM\s+corrected\s+mass', block, ncols)
    all_zero = all(
        (v == 0 or (isinstance(v, float) and (v == 0.0 or np.isnan(v))))
        for v in corr.values()
    )
    fwa = {k: np.nan for k in ('pressure', 'temperature', 'dens_gas', 'dens_oil', 'dens_water')} \
        if all_zero else _parse_fwa(block)
    return {
        'mpfm_uncorr': _parse_row(r'MPFM\s+uncorrected\s+mass', block, ncols),
        'mpfm_corr': corr,
        'pvt_mass': _parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*(?!@)', block, ncols),
        'pvt_vol': _parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*(?!@)', block, ncols),
        'pvt20_mass': _parse_row(r'PVT\s+ref(?:erence)?\s+mass\s*@20', block, ncols),
        'pvt20_vol': _parse_row(r'PVT\s+ref(?:erence)?\s+vol(?:ume)?\s*@20', block, ncols),
        'fwa': fwa,
    }


TAG_RE = re.compile(
    r'((?:Riser\s+P\d+|P[EI]_[A-Z0-9]+-?|PW[-_][A-Z0-9]+-?))\s*-\s*(\d{2}FT\d+)\s+'
    r'Production\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)
AREA_TOTAL_RE = re.compile(
    r'(North|South)\s*-\s*(Topside|Subsea)\s+MPFM\s+Production\s+Total\s+Previous\s+(Day|Hour)',
    re.IGNORECASE
)


def parse_pdf(pdf_path, report_type='daily'):
    text = _flat_text(pdf_path)

    if 'FPSO South' in text[:300]:
        fpso_side = 'South'
    elif 'FPSO West' in text[:300]:
        fpso_side = 'West'
    elif 'FPSO North' in text[:300]:
        fpso_side = 'North'
    else:
        fpso_side = 'South'
    unit_type = 'Subsea' if 'Subsea' in text[:300] else 'Topside'

    if report_type == 'daily':
        dm = re.search(r'Daily Report from\s+([\d.]+)\s+[\d:]+\s+to\s+([\d.]+)', text)
        date_from = dm.group(1).replace('.', '-') if dm else None
        date_to = dm.group(2).replace('.', '-') if dm else None
        hour = None
        dt_from = dt_to = None
    else:
        hm = re.search(r'Hourly Report from\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)\s+to\s+([\d.]+)\s+(\d{1,2}:\d{2}(?::\d{2})?)', text)
        dt_from = f"{hm.group(1).replace('.', '-')} {hm.group(2)}" if hm else None
        dt_to = f"{hm.group(3).replace('.', '-')} {hm.group(4)}" if hm else None
        date_from = hm.group(1).replace('.', '-') if hm else None
        date_to = None
        hour = int(hm.group(2).split(':')[0]) if hm else None

    headers = list(TAG_RE.finditer(text))
    tags = {}
    for i, h in enumerate(headers):
        tag = h.group(1).strip().replace(' ', '_').rstrip('-')
        instr = h.group(2)
        block = text[h.start(): headers[i + 1].start() if i + 1 < len(headers) else None]
        tags[tag] = {'instrument': instr, 'metrics': _parse_metrics(block, 5)}

    am = AREA_TOTAL_RE.search(text)
    area_total = {}
    if am:
        block = text[am.start():]
        area_total = {'area': am.group(1), 'unit_type': am.group(2), 'metrics': _parse_metrics(block, 4)}

    return {
        'pdf_path': pdf_path, 'date_from': date_from, 'date_to': date_to,
        'dt_from': dt_from, 'dt_to': dt_to, 'hour': hour,
        'fpso_side': fpso_side, 'unit_type': unit_type,
        'tags': tags, 'area_total': area_total,
    }


def _parse_pdf_worker(task):
    pdf_path, report_type = task
    try:
        return pdf_path, parse_pdf(pdf_path, report_type), ""
    except Exception as exc:
        return pdf_path, None, str(exc)


def parse_pdf_batch(paths, report_type: str, workers: int):
    tasks = [(str(path), report_type) for path in paths]
    if workers == 1 or len(tasks) < 2:
        return [_parse_pdf_worker(task) for task in tasks]
    with ProcessPoolExecutor(max_workers=workers) as executor:
        return list(executor.map(_parse_pdf_worker, tasks, chunksize=8))


HOURLY_COLS = [
    'Dia ref.', 'Hora', 'Bank', 'Loop', 'Tipo', 'TAG', 'Instrumento', 'DT Início', 'DT Fim',
    'MPFM uncorr Gás (t)', 'MPFM uncorr Óleo (t)', 'MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)', 'MPFM uncorr Total (t)',
    'MPFM corr Gás (t)', 'MPFM corr Óleo (t)', 'MPFM corr HC (t)',
    'MPFM corr Água (t)', 'MPFM corr Total (t)',
    'PVT mass Gás (t)', 'PVT mass Óleo (t)', 'PVT mass Água (t)',
    'PVT vol Gás (Sm³)', 'PVT vol Óleo (m³)', 'PVT vol Água (m³)',
    'PVT @20 mass Gás (t)', 'PVT @20 mass Óleo (t)', 'PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)', 'PVT @20 vol Óleo (m³)', 'PVT @20 vol Água (m³)',
    'Pressão (barg)', 'Temperatura (°C)',
    'Dens. Gás (kg/m³)', 'Dens. Óleo (kg/m³)', 'Dens. Água (kg/m³)', 'Fonte',
]

DAILY_COLS = [
    'Bank', 'Loop', 'Tipo', 'TAG', 'Instrumento', 'Dia',
    'MPFM uncorr Gás (t)', 'MPFM uncorr Óleo (t)', 'MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)', 'MPFM uncorr Total (t)',
    'MPFM corr Gás (t)', 'MPFM corr Óleo (t)', 'MPFM corr HC (t)',
    'MPFM corr Água (t)', 'MPFM corr Total (t)',
    'PVT mass Gás (t)', 'PVT mass Óleo (t)', 'PVT mass Água (t)',
    'PVT vol Gás (Sm³)', 'PVT vol Óleo (m³)', 'PVT vol Água (m³)',
    'PVT @20 mass Gás (t)', 'PVT @20 mass Óleo (t)', 'PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)', 'PVT @20 vol Óleo (m³)', 'PVT @20 vol Água (m³)',
    'Pressão (barg)', 'Temperatura (°C)',
    'Dens. Gás (kg/m³)', 'Dens. Óleo (kg/m³)', 'Dens. Água (kg/m³)', 'Fonte (Daily)',
]

SEP_COLS_HOURLY = [
    'SEP Temperatura Méd. (°C)', 'SEP Pressão Méd. (barg)', 'SEP Óleo Vol. Bruto (m³) CV',
    'SEP Óleo (t) CV', 'SEP Gás (t) CV', 'SEP Água (t) CV', 'SEP HC (t)', 'SEP Total (t)',
    'Desvio HC (%)', 'Desvio Total (%)',
]


def _mrow(m):
    u = m['mpfm_uncorr']
    c = m['mpfm_corr']
    pm = m['pvt_mass']
    pv = m['pvt_vol']
    p2 = m['pvt20_mass']
    v2 = m['pvt20_vol']
    f = m['fwa']
    return [
        u.get('gas'), u.get('oil'), u.get('hc'), u.get('water'), u.get('total'),
        c.get('gas'), c.get('oil'), c.get('hc'), c.get('water'), c.get('total'),
        pm.get('gas'), pm.get('oil'), pm.get('water'),
        pv.get('gas'), pv.get('oil'), pv.get('water'),
        p2.get('gas'), p2.get('oil'), p2.get('water'),
        v2.get('gas'), v2.get('oil'), v2.get('water'),
        f['pressure'], f['temperature'],
        f['dens_gas'], f['dens_oil'], f['dens_water'],
    ]


def build_daily_df(daily, unit_code):
    import os
    rows = []
    for tag, td in daily['tags'].items():
        base = [unit_code, daily['fpso_side'], daily['unit_type'], tag, td['instrument'], daily['date_from']]
        rows.append(base + _mrow(td['metrics']) + [os.path.basename(daily['pdf_path'])])
    return pd.DataFrame(rows, columns=DAILY_COLS)


def build_recon_df(daily, hourly_records, unit_code, abs_tol=0.5):
    hours_found = sorted(set(r['hour'] for r in hourly_records if r['hour'] is not None))
    hours_str = ','.join(f'{h:02d}' for h in hours_found)
    n = len(hours_found)
    coverage = 'OK (24/24h)' if n == 24 else (f'PARCIAL ({n}/24h)' if n > 0 else 'SEM HORÁRIOS')

    rows = []
    for tag, td in daily['tags'].items():
        d_c = td['metrics']['mpfm_corr']
        d_pv = td['metrics']['pvt_vol']
        s = {k: 0.0 for k in ['gas', 'oil', 'hc', 'water', 'total', 'pv_gas', 'pv_oil', 'pv_water']}
        for rec in hourly_records:
            if tag not in rec['tags']:
                continue
            hm = rec['tags'][tag]['metrics']

            def v(x):
                return x if isinstance(x, float) and not np.isnan(x) else 0.0
            for col in C5:
                s[col] += v(hm['mpfm_corr'].get(col, 0.0))
            s['pv_gas'] += v(hm['pvt_vol'].get('gas', 0.0))
            s['pv_oil'] += v(hm['pvt_vol'].get('oil', 0.0))
            s['pv_water'] += v(hm['pvt_vol'].get('water', 0.0))

        def delta(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)):
                return np.nan
            return round(sv - dv, 4)

        def status(dv, sv):
            if dv is None or (isinstance(dv, float) and np.isnan(dv)):
                return '-'
            if n == 0:
                return 'SEM DADOS'
            d = abs(sv - dv)
            return 'OK' if d <= max(abs_tol, 0.0005 * abs(dv)) else 'VERIFICAR'

        row = {
            'Bank': unit_code, 'Loop': daily['fpso_side'], 'Tipo': daily['unit_type'],
            'TAG': tag, 'Instrumento': td['instrument'], 'Dia': daily['date_from'],
            'Cobertura': coverage, 'Horas': hours_str if n > 0 else '-',
            'Daily Gás (t)': d_c.get('gas'), 'Soma h. Gás (t)': round(s['gas'], 4),
            'Δ Gás (t)': delta(d_c.get('gas'), s['gas']), 'Status Gás': status(d_c.get('gas'), s['gas']),
            'Daily Óleo (t)': d_c.get('oil'), 'Soma h. Óleo (t)': round(s['oil'], 4),
            'Δ Óleo (t)': delta(d_c.get('oil'), s['oil']), 'Status Óleo': status(d_c.get('oil'), s['oil']),
            'Daily HC (t)': d_c.get('hc'), 'Soma h. HC (t)': round(s['hc'], 4),
            'Δ HC (t)': delta(d_c.get('hc'), s['hc']), 'Status HC': status(d_c.get('hc'), s['hc']),
            'Daily Água (t)': d_c.get('water'), 'Soma h. Água (t)': round(s['water'], 4),
            'Δ Água (t)': delta(d_c.get('water'), s['water']), 'Status Água': status(d_c.get('water'), s['water']),
            'Daily PVT Gás (Sm³)': d_pv.get('gas'), 'Soma h. PVT Gás (Sm³)': round(s['pv_gas'], 1),
            'Δ PVT Gás': delta(d_pv.get('gas'), s['pv_gas']),
            'Daily PVT Óleo (m³)': d_pv.get('oil'), 'Soma h. PVT Óleo (m³)': round(s['pv_oil'], 4),
            'Δ PVT Óleo': delta(d_pv.get('oil'), s['pv_oil']),
        }
        rows.append(row)
    return pd.DataFrame(rows)


def _desvio(mpfm_val, sep_val):
    try:
        if mpfm_val is None or sep_val is None:
            return np.nan
        if np.isnan(float(mpfm_val)) or np.isnan(float(sep_val)):
            return np.nan
        if float(sep_val) == 0:
            return np.nan
        return round((float(mpfm_val) - float(sep_val)) / float(sep_val) * 100, 2)
    except Exception:
        return np.nan


def _sep_row(sep_hour, mpfm_hc_t, mpfm_total_t):
    if not sep_hour:
        return [np.nan] * 10
    oil_t = sep_hour.get('oil_t', np.nan)
    gas_t = sep_hour.get('gas_t', np.nan)
    wat_t = sep_hour.get('water_t', np.nan)
    hc_t = sep_hour.get('hc_t', np.nan)
    tot_t = sep_hour.get('total_t', np.nan)

    def _r(v):
        return round(v, 4) if (v is not None and not (isinstance(v, float) and np.isnan(v))) else np.nan
    return [
        sep_hour.get('temp', np.nan), sep_hour.get('pressure_barg', np.nan), sep_hour.get('oil_m3', np.nan),
        _r(oil_t), _r(gas_t), _r(wat_t), _r(hc_t), _r(tot_t),
        _desvio(mpfm_hc_t, hc_t), _desvio(mpfm_total_t, tot_t),
    ]


def build_hourly_df_with_sep(hourly_records, unit_code, sep_data=None):
    import os
    cols = HOURLY_COLS + (SEP_COLS_HOURLY if sep_data else [])
    rows = []
    for rec in sorted(hourly_records, key=lambda r: r['hour'] or 0):
        for tag, td in rec['tags'].items():
            base = [rec['date_from'], rec['hour'], unit_code, rec['fpso_side'], rec['unit_type'],
                    tag, td['instrument'], rec['dt_from'], rec['dt_to']]
            mpfm_vals = _mrow(td['metrics'])
            row = base + mpfm_vals + [os.path.basename(rec['pdf_path'])]
            if sep_data:
                pdf_hour = rec['hour']
                txt_key = 24 if pdf_hour == 0 else pdf_hour
                sh = sep_data.get(txt_key)
                c = td['metrics']['mpfm_corr']
                mpfm_hc = c.get('hc')
                mpfm_tot = c.get('total')
                row += _sep_row(sh, mpfm_hc, mpfm_tot)
            rows.append(row)
    return pd.DataFrame(rows, columns=cols)


_SEP_TOKEN_SPLIT_DECIMALS = (5, 4, 3, 6, 2)


def _is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def _recover_concatenated_float_token(token):
    raw = str(token or '').strip()
    if raw.count('.') < 2:
        return None
    if not re.fullmatch(r'[+-]?\d+(?:\.\d+)+', raw):
        return None
    first_dot = raw.find('.')
    for frac_len in _SEP_TOKEN_SPLIT_DECIMALS:
        split_at = first_dot + 1 + frac_len
        if split_at <= first_dot + 1 or split_at >= len(raw):
            continue
        left = raw[:split_at]
        right = raw[split_at:]
        if not right or not right[0].isdigit() or right.count('.') != 1:
            continue
        if _is_number(left) and _is_number(right):
            return left, right
    return None


def _parse_sep_float_token(raw_value, *, file_path, row_key, field_name, line):
    token = str(raw_value or '').strip()
    try:
        return float(token)
    except ValueError:
        recovered = _recover_concatenated_float_token(token)
        if not recovered:
            raise
        primary_token, _overflow_token = recovered
        return float(primary_token)


def _parse_hourly_txt_oleo(path):
    """Lê Run_24Hours*_OLEO.txt (ou _AGUA.txt, mesmo formato).
    Retorna {hour_int: {temp, pressure_barg, gv_m3, mass_t}, 'DAY': {...}}.
    hour_int: 1..24 (24 = última hora, equivale ao PDF hour=0 do dia seguinte)."""
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAY':
                if len(parts) >= 9:
                    result['DAY'] = {
                        'temp': _parse_sep_float_token(parts[2], file_path=path, row_key=key, field_name='temp', line=line),
                        'pressure_barg': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_raw', line=line) / 100.0,
                        'gv_m3': _parse_sep_float_token(parts[6], file_path=path, row_key=key, field_name='gv_m3', line=line),
                        'mass_t': _parse_sep_float_token(parts[8], file_path=path, row_key=key, field_name='mass_t', line=line),
                    }
            elif key.isdigit():
                h = int(key)
                if len(parts) >= 9:
                    result[h] = {
                        'temp': _parse_sep_float_token(parts[2], file_path=path, row_key=key, field_name='temp', line=line),
                        'pressure_barg': _parse_sep_float_token(parts[1], file_path=path, row_key=key, field_name='pressure_raw', line=line) / 100.0,
                        'gv_m3': _parse_sep_float_token(parts[6], file_path=path, row_key=key, field_name='gv_m3', line=line),
                        'mass_t': _parse_sep_float_token(parts[8], file_path=path, row_key=key, field_name='mass_t', line=line),
                    }
    return result


def _parse_hourly_txt_gas(path):
    """Lê Run_24Hours*_GAS.txt. Retorna {hour_int: {mass_t}, 'DAY': {mass_t}}.
    O campo Mass no arquivo de gás está em kg -> convertido aqui para toneladas."""
    result = {}
    with open(path, encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.rstrip('\r\n')
            parts = line.split()
            if not parts:
                continue
            key = parts[0].upper()
            if key == 'DAILY':
                nums = [p for p in parts[1:] if _is_number(p)]
                if len(nums) >= 3:
                    result['DAY'] = {
                        'mass_t': _parse_sep_float_token(nums[2], file_path=path, row_key=key, field_name='mass_raw', line=line) / 1000.0
                    }
            elif key.isdigit():
                h = int(key)
                if len(parts) >= 8:
                    result[h] = {
                        'mass_t': _parse_sep_float_token(parts[7], file_path=path, row_key=key, field_name='mass_raw', line=line) / 1000.0
                    }
    return result


def parse_sep_txt_set(oleo_path, gas_path, agua_path):
    """Lê os 3 TXT do separador de testes e monta um dict unificado:
    {hour: {temp, pressure_barg, oil_m3, oil_t, gas_t, water_t, hc_t, total_t}, 'DAY': {...}}."""
    oleo_data = _parse_hourly_txt_oleo(oleo_path)
    gas_data = _parse_hourly_txt_gas(gas_path)
    agua_data = _parse_hourly_txt_oleo(agua_path)

    combined = {}
    all_keys = sorted(set(list(oleo_data.keys()) + list(gas_data.keys()) + list(agua_data.keys())),
                       key=lambda x: 999 if x == 'DAY' else x)
    for k in all_keys:
        o = oleo_data.get(k, {})
        g = gas_data.get(k, {})
        w = agua_data.get(k, {})
        oil_m3 = o.get('gv_m3', np.nan)
        oil_t = o.get('mass_t', np.nan)
        gas_t = g.get('mass_t', np.nan)
        wat_t = w.get('mass_t', np.nan)

        def _s(a, b):
            if np.isnan(a) or np.isnan(b):
                return np.nan
            return a + b

        hc_t = _s(oil_t, gas_t)
        total_t = _s(hc_t, wat_t)
        combined[k] = {
            'temp': o.get('temp', np.nan), 'pressure_barg': o.get('pressure_barg', np.nan),
            'oil_m3': oil_m3, 'oil_t': oil_t, 'gas_t': gas_t, 'water_t': wat_t,
            'hc_t': hc_t, 'total_t': total_t,
        }
    return combined


def _mpfm_hour_from_sep_hour(sep_hour_ref):
    if sep_hour_ref is None:
        return None
    try:
        hour = int(sep_hour_ref)
    except Exception:
        return None
    return 0 if hour == 24 else hour


def _sep_desvio_pct(mpfm_value, sep_value):
    try:
        if mpfm_value in (None, "") or sep_value in (None, ""):
            return ""
        mpfm_f = float(mpfm_value)
        sep_f = float(sep_value)
        if sep_f == 0:
            return ""
        return round((mpfm_f - sep_f) / sep_f * 100, 2)
    except Exception:
        return ""


# ═════════════════════════════════════════════════════════════════════════
# BASE_UNICA — colunas e mapeamentos (copiado de monthly_workbook_service.py)
# ═════════════════════════════════════════════════════════════════════════

BASE_UNICA_COLUMNS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType", "Area", "System", "Bank", "Loop", "Tipo", "Entity", "Tag", "Instrumento", "PI Tag",
    "MPFM uncorr Gás (t)", "MPFM uncorr Óleo (t)", "MPFM uncorr HC (t)", "MPFM uncorr Água (t)", "MPFM uncorr Total (t)",
    "MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr HC (t)", "MPFM corr Água (t)", "MPFM corr Total (t)",
    "PVT mass Gás (t)", "PVT mass Óleo (t)", "PVT mass HC (t)", "PVT mass Água (t)", "PVT mass Total (t)",
    "PVT vol Gás (Sm³)", "PVT vol Óleo (m³)", "PVT vol HC (m³)", "PVT vol Água (m³)", "PVT vol Total (m³)",
    "PVT @20 mass Gás (t)", "PVT @20 mass Óleo (t)", "PVT @20 mass HC (t)", "PVT @20 mass Água (t)", "PVT @20 mass Total (t)",
    "PVT @20 vol Gás (Sm³)", "PVT @20 vol Óleo (m³)", "PVT @20 vol HC (m³)", "PVT @20 vol Água (m³)", "PVT @20 vol Total (m³)",
    "Pressão (barg)", "Temperatura (°C)", "Dens. Gás (kg/m³)", "Dens. Óleo (kg/m³)", "Dens. Água (kg/m³)",
    "SEP TAG", "SEP Medidor", "SEP Local", "SEP Status", "Bancos alinhados",
    "SEP Óleo Vol. Bruto (m³) CV", "SEP Óleo (t) CV", "SEP Gás (t) CV", "SEP Água (t) CV", "SEP HC (t)", "SEP Total (t)", "SEP Temperatura Méd. (°C)", "SEP Pressão Méd. (barg)",
    "Desvio HC (%)", "Desvio Total (%)",
    "Recon Cobertura", "Recon Horas", "Recon Daily Gás (t)", "Recon Daily Óleo (t)", "Recon Daily HC (t)", "Recon Daily Água (t)",
    "Recon Soma h. Gás (t)", "Recon Soma h. Óleo (t)", "Recon Soma h. HC (t)", "Recon Soma h. Água (t)",
    "Recon Δ Gás (t)", "Recon Δ Óleo (t)", "Recon Δ HC (t)", "Recon Δ Água (t)",
    "Status Gás", "Status Óleo", "Status HC", "Status Água", "Fonte", "SourceFile", "IsOfficial",
]

MANUAL_SHEET_NAME = "COMPARATIVO_MANUAL"

COMPARATIVO_COLUMNS = [
    "Data", "Hora", "Banco", "TAG", "Origem", "Pressão MPFM (barg)", "Temperatura MPFM (°C)",
    "Densidade Óleo (kg/m³)", "Densidade Gás (kg/m³)", "Densidade Água (kg/m³)",
    "Massa HC Não Corrigida (t)", "Massa Total Não Corrigida (t)", "Massa Óleo Não Corrigida (t)",
    "Massa Gás Não Corrigida (t)", "Massa Água Não Corrigida (t)",
    "Massa HC Corrigida (t)", "Massa Total Corrigida (t)", "Massa Óleo Corrigida (t)",
    "Massa Gás Corrigida (t)", "Massa Água Corrigida (t)",
    "Massa HC Padrão (t)", "Massa Total Padrão (t)", "Massa Óleo Padrão (t)",
    "Massa Gás Padrão (t)", "Massa Água Padrão (t)",
    "Volume Óleo STD 20°C, 1 atm (m³)", "Volume Gás STD 20°C, 1 atm (Sm³)",
    "Volume Água STD 20°C, 1 atm (m³)",
    "Velocidade Escoamento (m/s)", "GVF (%)", "ΔP - Inlet (mbar)", "VVF (%)", "WLR (%)",
    "Choke O?", "Pressão MPFM acima do ponto de bolha?", "Continuous Phase", "Calculation Mode", "Observações",
]

COMPARATIVO_NOTES = [
    ("Metodologia do desvio", ""),
    ("PE-02 × Riser P2", "% Desvio HC = ((Massa HC Corrigida PE-02 subsea / Massa HC Corrigida Riser P2 topside) - 1) × 100"),
    ("PE-02 × Riser P2", "% Desvio Total = ((Massa Total Corrigida PE-02 subsea / Massa Total Corrigida Riser P2 topside) - 1) × 100"),
    ("PE-04 × Riser P5", "Aplicar a mesma metodologia de desvio HC e Total."),
    ("PW-104 × Riser P4", "Aplicar a mesma metodologia de desvio HC e Total."),
    ("Referência", "Riser (topside). Limites de aceitação: ±10% (HC) e ±7% (Total)."),
    ("Separador de Testes", "Não possui trilha de correção independente; os valores de massa são repetidos nas colunas Corrigida e Padrão."),
]

RECON_MAP = {
    "Cobertura": "Recon Cobertura", "Horas": "Recon Horas",
    "Daily Gás (t)": "Recon Daily Gás (t)", "Daily Óleo (t)": "Recon Daily Óleo (t)",
    "Daily HC (t)": "Recon Daily HC (t)", "Daily Água (t)": "Recon Daily Água (t)",
    "Soma h. Gás (t)": "Recon Soma h. Gás (t)", "Soma h. Óleo (t)": "Recon Soma h. Óleo (t)",
    "Soma h. HC (t)": "Recon Soma h. HC (t)", "Soma h. Água (t)": "Recon Soma h. Água (t)",
    "Δ Gás (t)": "Recon Δ Gás (t)", "Δ Óleo (t)": "Recon Δ Óleo (t)",
    "Δ HC (t)": "Recon Δ HC (t)", "Δ Água (t)": "Recon Δ Água (t)",
    "Status Gás": "Status Gás", "Status Óleo": "Status Óleo",
    "Status HC": "Status HC", "Status Água": "Status Água",
}

SEP_COLS_PASSTHROUGH = [
    "SEP Temperatura Méd. (°C)", "SEP Pressão Méd. (barg)", "SEP Óleo Vol. Bruto (m³) CV",
    "SEP Óleo (t) CV", "SEP Gás (t) CV", "SEP Água (t) CV", "SEP HC (t)", "SEP Total (t)",
    "Desvio HC (%)", "Desvio Total (%)",
]

_MPFM_METRIC_COLS = [
    "MPFM uncorr Gás (t)", "MPFM uncorr Óleo (t)", "MPFM uncorr HC (t)", "MPFM uncorr Água (t)", "MPFM uncorr Total (t)",
    "MPFM corr Gás (t)", "MPFM corr Óleo (t)", "MPFM corr HC (t)", "MPFM corr Água (t)", "MPFM corr Total (t)",
    "PVT mass Gás (t)", "PVT mass Óleo (t)", "PVT mass Água (t)",
    "PVT vol Gás (Sm³)", "PVT vol Óleo (m³)", "PVT vol Água (m³)",
    "PVT @20 mass Gás (t)", "PVT @20 mass Óleo (t)", "PVT @20 mass Água (t)",
    "PVT @20 vol Gás (Sm³)", "PVT @20 vol Óleo (m³)", "PVT @20 vol Água (m³)",
    "Pressão (barg)", "Temperatura (°C)", "Dens. Gás (kg/m³)", "Dens. Óleo (kg/m³)", "Dens. Água (kg/m³)",
]


def _new_row() -> dict:
    return {col: "" for col in BASE_UNICA_COLUMNS}


def add_comparativo_sheet(workbook_path: Path, manual_rows: int = 200, replace_existing: bool = True):
    """Cria a aba fixa de preenchimento manual e preserva a aba Base_Unica."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(workbook_path)
    if MANUAL_SHEET_NAME in workbook.sheetnames:
        if not replace_existing:
            workbook.close()
            return
        del workbook[MANUAL_SHEET_NAME]
    sheet = workbook.create_sheet(MANUAL_SHEET_NAME)
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = True

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COMPARATIVO_COLUMNS))
    title = sheet.cell(row=1, column=1, value="COMPARATIVO MANUAL — PREENCHIMENTO PELO USUÁRIO")
    title.font = Font(bold=True, color="FFFFFF", size=12)
    title.fill = PatternFill("solid", fgColor="1F4E78")
    title.alignment = Alignment(horizontal="center")
    sheet.cell(row=2, column=1, value="Campos mantidos em branco para preenchimento manual; fórmulas poderão ser incluídas posteriormente.")
    sheet.cell(row=2, column=1).font = Font(italic=True, color="666666")

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for column_index, column_name in enumerate(COMPARATIVO_COLUMNS, start=1):
        cell = sheet.cell(row=3, column=column_index, value=column_name)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column_index)].width = max(14, min(len(column_name) + 3, 35))

    for row_index in range(4, 4 + manual_rows):
        for column_index in range(1, len(COMPARATIVO_COLUMNS) + 1):
            sheet.cell(row=row_index, column=column_index).alignment = Alignment(vertical="center")

    notes_row = 6 + manual_rows
    sheet.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=len(COMPARATIVO_COLUMNS))
    notes_title = sheet.cell(row=notes_row, column=1, value="METODOLOGIA E REFERÊNCIAS")
    notes_title.font = Font(bold=True, color="FFFFFF")
    notes_title.fill = PatternFill("solid", fgColor="1F4E78")
    notes_title.alignment = Alignment(horizontal="center")

    for row_offset, (topic, description) in enumerate(COMPARATIVO_NOTES, start=1):
        row_index = notes_row + row_offset
        sheet.cell(row=row_index, column=1, value=topic).font = Font(bold=True)
        sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=len(COMPARATIVO_COLUMNS))
        note_cell = sheet.cell(row=row_index, column=2, value=description)
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(workbook_path)
    workbook.close()


def hourly_df_to_rows(df, sep_merged: bool) -> list:
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia ref."], "Hour": r["Hora"], "Granularity": "Hourly",
            "Origin": "MPFM", "SourceType": "PDF", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "MPFM", "SourceFile": r["Fonte"], "IsOfficial": 1,
        })
        for col in _MPFM_METRIC_COLS:
            row[col] = r[col]
        if sep_merged:
            for col in SEP_COLS_PASSTHROUGH:
                if col in df.columns:
                    row[col] = r[col]
            row["SEP TAG"] = "SEP_Dados"
            row["SEP Medidor"] = "20FT0244/20FT0247/20FT0251"
            row["SEP Local"] = "Separador de Testes"
            row["SEP Status"] = "Alinhado"
            row["Bancos alinhados"] = row["Bank"]
        rows.append(row)
    return rows


def daily_df_to_rows(df, sep_day: dict | None) -> list:
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia"], "Hour": "", "Granularity": "Daily",
            "Origin": "MPFM", "SourceType": "PDF", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "MPFM", "SourceFile": r["Fonte (Daily)"], "IsOfficial": 1,
        })
        for col in _MPFM_METRIC_COLS:
            row[col] = r[col]
        if sep_day:
            mpfm_hc = row["MPFM corr HC (t)"]
            mpfm_tot = row["MPFM corr Total (t)"]
            row["SEP Temperatura Méd. (°C)"] = sep_day.get("temp", "")
            row["SEP Pressão Méd. (barg)"] = sep_day.get("pressure_barg", "")
            row["SEP Óleo Vol. Bruto (m³) CV"] = sep_day.get("oil_m3", "")
            row["SEP Óleo (t) CV"] = sep_day.get("oil_t", "")
            row["SEP Gás (t) CV"] = sep_day.get("gas_t", "")
            row["SEP Água (t) CV"] = sep_day.get("water_t", "")
            row["SEP HC (t)"] = sep_day.get("hc_t", "")
            row["SEP Total (t)"] = sep_day.get("total_t", "")
            row["Desvio HC (%)"] = _sep_desvio_pct(mpfm_hc, sep_day.get("hc_t"))
            row["Desvio Total (%)"] = _sep_desvio_pct(mpfm_tot, sep_day.get("total_t"))
            row["SEP TAG"] = "SEP_Dados"
            row["SEP Medidor"] = "20FT0244/20FT0247/20FT0251"
            row["SEP Local"] = "Separador de Testes"
            row["SEP Status"] = "Alinhado"
            row["Bancos alinhados"] = row["Bank"]
        rows.append(row)
    return rows


def recon_df_to_rows(df) -> list:
    rows = []
    for _, r in df.iterrows():
        row = _new_row()
        row.update({
            "ProductionDate": r["Dia"], "Hour": "", "Granularity": "Daily",
            "Origin": "RECON", "SourceType": "CALC", "Bank": r["Bank"], "Loop": r["Loop"], "Tipo": r["Tipo"],
            "Entity": r["TAG"], "Tag": r["TAG"], "Instrumento": r["Instrumento"],
            "Fonte": "Reconciliação", "SourceFile": "", "IsOfficial": 1,
        })
        for metric, column in RECON_MAP.items():
            if metric in df.columns:
                row[column] = r[metric]
        rows.append(row)
    return rows


MASTER_SHEET_NAME = "BASE_UNICA_TOTAL"
MASTER_DEDUP_KEYS = [
    "ProductionDate", "Hour", "Granularity", "Origin", "SourceType",
    "Bank", "Entity", "Tag", "Instrumento",
]


def _normalize_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Garante a ordem e o conjunto esperado de colunas da Base_Unica."""
    if df is None or df.empty:
        return pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    out = df.copy()
    for col in BASE_UNICA_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    return out[BASE_UNICA_COLUMNS]


def update_master_base_unica(master_path: Path, df_new: pd.DataFrame) -> dict:
    """Integra incrementalmente a execução atual em um arquivo Base_Unica total.

    Regra: atualiza somente a aba BASE_UNICA_TOTAL. A aba COMPARATIVO_MANUAL,
    usada para preenchimento do usuário, é criada na primeira execução e depois
    preservada sem sobrescrever valores digitados. O histórico automatizado é
    combinado com a nova execução e deduplicado por chave técnica, mantendo a
    versão mais recente.
    """
    master_path.parent.mkdir(parents=True, exist_ok=True)
    df_current = pd.DataFrame(columns=BASE_UNICA_COLUMNS)
    previous_rows = 0

    if master_path.exists():
        try:
            df_current = pd.read_excel(master_path, sheet_name=MASTER_SHEET_NAME, dtype=object)
        except ValueError:
            from openpyxl import load_workbook

            workbook = load_workbook(master_path, read_only=True, data_only=True)
            try:
                candidate_sheets = [name for name in workbook.sheetnames if name != MANUAL_SHEET_NAME]
            finally:
                workbook.close()
            if not candidate_sheets:
                df_current = pd.DataFrame(columns=BASE_UNICA_COLUMNS)
            else:
                df_current = pd.read_excel(master_path, sheet_name=candidate_sheets[0], dtype=object)
        df_current = _normalize_master_columns(df_current)
        df_current = df_current.where(pd.notna(df_current), "")
        previous_rows = len(df_current)

    df_increment = _normalize_master_columns(df_new)
    df_increment = df_increment.where(pd.notna(df_increment), "")
    combined = pd.concat([df_current, df_increment], ignore_index=True)

    for key in MASTER_DEDUP_KEYS:
        combined[key] = combined[key].where(pd.notna(combined[key]), "").astype(str)

    combined.drop_duplicates(subset=MASTER_DEDUP_KEYS, keep="last", inplace=True)
    combined.sort_values(
        by=["ProductionDate", "Bank", "Granularity", "Hour", "Tag", "Origin"],
        inplace=True,
        na_position="last",
        key=lambda col: col.astype(str) if col.name != "Hour" else col,
    )
    combined = combined[BASE_UNICA_COLUMNS]

    if master_path.exists():
        with pd.ExcelWriter(master_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            combined.to_excel(writer, sheet_name=MASTER_SHEET_NAME, index=False)
    else:
        combined.to_excel(master_path, sheet_name=MASTER_SHEET_NAME, index=False)

    add_comparativo_sheet(master_path, replace_existing=False)
    return {
        "previous_rows": previous_rows,
        "increment_rows": len(df_increment),
        "total_rows": len(combined),
        "replaced_or_duplicate_rows": previous_rows + len(df_increment) - len(combined),
    }


# ═════════════════════════════════════════════════════════════════════════
# CAMINHOS DE PASTA — ajuste aqui se a estrutura da sua instalação for diferente
# ═════════════════════════════════════════════════════════════════════════

_FILENAME_DATE_RE = re.compile(r"-(\d{8})-")


def _filename_date_iso(path: Path):
    m = _FILENAME_DATE_RE.search(path.name)
    if not m:
        return None
    raw = m.group(1)
    return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"


def bank_month_dir(mpfm_root: Path, bank_code: str, year: int, month: int, sub: str) -> Path:
    return mpfm_root / BANK_FOLDERS[bank_code] / str(year) / MONTH_PT[month] / sub


def sep_day_dir(sep_root: Path, day_iso: str, fc_folder: str) -> Path:
    dt = datetime.strptime(day_iso, "%Y-%m-%d")
    return (
        sep_root / str(dt.year) / MONTH_PT[dt.month]
        / f"FPSO-Bacalhau_Daily reports_{day_iso}" / "01 - CV_Reports" / fc_folder
    )


def _months_to_scan(reference: datetime, back: int = 1):
    months = []
    y, m = reference.year, reference.month
    for _ in range(back + 1):
        months.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return months


def _months_in_range(date_from: datetime, date_to: datetime):
    months = []
    year, month = date_from.year, date_from.month
    while (year, month) <= (date_to.year, date_to.month):
        months.append((year, month))
        month += 1
        if month == 13:
            year += 1
            month = 1
    return months


def _parse_date_argument(value: str, option_name: str):
    if not value:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue
    raise ValueError(f"{option_name} deve usar DD/MM/AAAA ou AAAA-MM-DD.")


# ═════════════════════════════════════════════════════════════════════════
# DESCOBERTA — dias disponíveis (Daily PDF) e leitura de Hourly PDF
# ═════════════════════════════════════════════════════════════════════════

def _candidate_daily_paths(
    mpfm_root: Path, bank_code: str, months: list, keep: int | None = None,
    date_from: str | None = None, date_to: str | None = None,
) -> list:
    paths = []
    last_report_name_day = ""
    if date_to:
        last_report_name_day = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    for year, month in months:
        daily_dir = bank_month_dir(mpfm_root, bank_code, year, month, "Daily")
        if daily_dir.is_dir():
            for path in daily_dir.glob(f"{bank_code}_MPFM_Daily-*.pdf"):
                file_day = _filename_date_iso(path)
                if date_from and file_day and file_day < date_from:
                    continue
                if last_report_name_day and file_day and file_day > last_report_name_day:
                    continue
                paths.append(path)
    paths.sort(key=lambda p: _filename_date_iso(p) or "", reverse=True)
    return paths[:keep] if keep else paths


def discover_daily_records(
    mpfm_root: Path, bank_code: str, months: list, keep: int | None = None,
    date_from: str | None = None, date_to: str | None = None, workers: int = 1,
) -> dict:
    out = {}
    paths = _candidate_daily_paths(mpfm_root, bank_code, months, keep, date_from, date_to)
    for pdf_name, record, error in parse_pdf_batch(paths, "daily", workers):
        pdf_path = Path(pdf_name)
        if error:
            print(f"  ⚠️  falha ao ler {pdf_path.name}: {error}")
            continue
        day = record.get("date_from")
        if not day:
            continue
        if date_from and day < date_from:
            continue
        if date_to and day > date_to:
            continue
        out.setdefault(day, (pdf_path, record))
    return out


def discover_hourly_records_for_days(mpfm_root: Path, bank_code: str, months: list, target_days: set, workers: int = 1) -> dict:
    relevant_names = set(target_days)
    for d in target_days:
        dt = datetime.strptime(d, "%Y-%m-%d") + timedelta(days=1)
        relevant_names.add(dt.strftime("%Y-%m-%d"))

    by_day = defaultdict(list)
    paths = []
    for year, month in months:
        hourly_dir = bank_month_dir(mpfm_root, bank_code, year, month, "Hourly")
        if not hourly_dir.is_dir():
            continue
        for pdf_path in sorted(hourly_dir.glob(f"{bank_code}_MPFM_Hourly-*.pdf")):
            if _filename_date_iso(pdf_path) not in relevant_names:
                continue
            paths.append(pdf_path)
    for pdf_name, record, error in parse_pdf_batch(paths, "hourly", workers):
        pdf_path = Path(pdf_name)
        if error:
            print(f"  ⚠️  falha ao ler {pdf_path.name}: {error}")
            continue
        day = record.get("date_from")
        hour = record.get("hour")
        if not day or hour is None or day not in target_days:
            continue
        by_day[day].append(record)
    return by_day


# ═════════════════════════════════════════════════════════════════════════
# SEP: localizar os 3 TXT (óleo/gás/água) de um dia lendo o "Meter ID"
# ═════════════════════════════════════════════════════════════════════════

_METER_ID_RE = re.compile(r"Meter\s+ID\s+(\S+)", re.IGNORECASE)
_EXCLUDED_TXT_PREFIXES = (
    "alarmsandevents", "configuration", "run_daily", "parameters", "run_hourly", "_monthly", "run1_monthly",
    "run2_monthly", "run3_monthly", "security",
)


def _read_meter_id(txt_path: Path) -> str:
    try:
        with open(txt_path, encoding="utf-8", errors="replace") as f:
            head = f.read(400)
        m = _METER_ID_RE.search(head)
        return m.group(1).strip().upper() if m else ""
    except Exception:
        return ""


def find_sep_files_for_day(sep_root: Path, day_iso: str) -> dict:
    """Varre FC13/FC14/FC17 do dia (caminho direto, sem busca recursiva no
    ano inteiro) e identifica óleo/gás/água pelo Meter ID lido no cabeçalho
    de cada Run_24Hours*.txt."""
    found = {}
    for fc_folder in SEP_FC_FOLDERS:
        day_dir = sep_day_dir(sep_root, day_iso, fc_folder)
        if not day_dir.is_dir():
            continue
        for txt_path in day_dir.glob("Run_24Hours*.txt"):
            name_lower = txt_path.name.lower()
            if name_lower.startswith(_EXCLUDED_TXT_PREFIXES):
                continue
            meter_id = _read_meter_id(txt_path)
            phase = TARGET_PHASE_BY_METER.get(meter_id)
            if phase and phase not in found:
                found[phase] = txt_path
    return found


def load_sep_data_for_days(sep_root: Path, days_iso: list) -> dict:
    result = {}
    for day in days_iso:
        paths = find_sep_files_for_day(sep_root, day)
        missing = [k for k in ("oleo", "gas", "agua") if k not in paths]
        if missing:
            if paths:
                print(f"  ⚠️  SEP incompleto em {day}: faltando {', '.join(missing)} — dia ignorado no merge SEP")
            continue
        try:
            result[day] = parse_sep_txt_set(paths["oleo"], paths["gas"], paths["agua"])
        except Exception as exc:
            print(f"  ⚠️  falha ao ler TXT do SEP em {day}: {exc}")
    return result


# ═════════════════════════════════════════════════════════════════════════
# CLI / prompts interativos
# ═════════════════════════════════════════════════════════════════════════

def _resolve_config():
    parser = argparse.ArgumentParser(description="Gerador Base_Unica standalone (MPFM + SEP)")
    parser.add_argument("--mpfm-root", default=MPFM_ROOT, help="Pasta raiz dos PDFs MPFM (3.1.x por banco)")
    parser.add_argument("--sep-root", default=SEP_ROOT, help="Pasta raiz dos Daily Reports (FC13/FC14/FC17)")
    parser.add_argument("--output", default=OUTPUT_PATH, help="Caminho do .xlsx de saída")
    parser.add_argument("--master-output", default=MASTER_OUTPUT_PATH, help="Caminho do .xlsx incremental total (padrão: BASE_UNICA_TOTAL.xlsx ao lado do script)")
    parser.add_argument("--no-master", action="store_true", help="Não atualizar a Base_Unica total incremental nesta execução")
    parser.add_argument("--days", type=int, default=DAYS_COUNT, help="Quantidade de dias mais recentes a exportar")
    parser.add_argument("--aligned-bank", default=SEP_ALIGNED_BANK, help="Banco ao qual o SEP será mesclado")
    parser.add_argument("--months-lookback", type=int, default=MONTHS_LOOKBACK, help="Máximo de meses pesquisados")
    parser.add_argument("--workers", type=int, default=PDF_WORKERS, help="Processos paralelos para leitura de PDF")
    parser.add_argument("--date-from", default="", help="Início do intervalo (DD/MM/AAAA ou AAAA-MM-DD)")
    parser.add_argument("--date-to", default="", help="Fim do intervalo (DD/MM/AAAA ou AAAA-MM-DD)")
    args = parser.parse_args()

    mpfm_root = args.mpfm_root.strip()
    if not mpfm_root:
        mpfm_root = input("Caminho da pasta raiz dos PDFs MPFM (3.1.x por banco): ").strip()
    sep_root = args.sep_root.strip()
    if not sep_root:
        sep_root = input("Caminho da pasta raiz dos Daily Reports (FC13/FC14/FC17): ").strip()

    output_path = args.output.strip()
    if not output_path:
        output_path = str(Path(__file__).resolve().parent / f"BASE_UNICA_STANDALONE_{datetime.now():%Y%m%d_%H%M%S}.xlsx")

    update_master = bool(UPDATE_MASTER) and not args.no_master
    master_output_path = args.master_output.strip()
    if update_master and not master_output_path:
        master_output_path = str(Path(__file__).resolve().parent / "BASE_UNICA_TOTAL.xlsx")

    try:
        date_from = _parse_date_argument(args.date_from.strip(), "--date-from")
        date_to = _parse_date_argument(args.date_to.strip(), "--date-to")
    except ValueError as exc:
        parser.error(str(exc))
    if date_from and not date_to:
        date_to = datetime.now() - timedelta(days=1)
    if date_to and not date_from:
        parser.error("--date-to exige --date-from.")
    if date_from and date_to and date_from > date_to:
        parser.error("--date-from não pode ser posterior a --date-to.")

    return (
        Path(mpfm_root), Path(sep_root), Path(output_path), args.days,
        args.aligned_bank.strip().upper(), args.months_lookback, args.workers, date_from, date_to,
        update_master, Path(master_output_path) if master_output_path else None,
    )


# ═════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════

def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(line_buffering=True)

    mpfm_root, sep_root, output_path, days_count, aligned_bank, months_lookback, workers, date_from, date_to, update_master, master_output_path = _resolve_config()

    if not mpfm_root.is_dir():
        print(f"❌ Pasta MPFM_ROOT não encontrada: {mpfm_root}")
        sys.exit(1)
    if not sep_root.is_dir():
        print(f"❌ Pasta SEP_ROOT não encontrada: {sep_root}")
        sys.exit(1)
    if days_count < 1:
        print("❌ --days deve ser maior que zero.")
        sys.exit(1)
    if months_lookback < 1:
        print("❌ --months-lookback deve ser maior que zero.")
        sys.exit(1)
    if workers < 1:
        print("❌ --workers deve ser maior que zero.")
        sys.exit(1)
    if aligned_bank not in BANK_FOLDERS:
        print(f"❌ Banco SEP inválido: {aligned_bank}. Opções: {', '.join(BANK_FOLDERS)}")
        sys.exit(1)

    today = datetime.now()
    if date_from and date_to:
        months = _months_in_range(date_from, date_to + timedelta(days=1))
        requested_from = date_from.strftime("%Y-%m-%d")
        requested_to = date_to.strftime("%Y-%m-%d")
        print(f"📆 Intervalo solicitado: {requested_from} a {requested_to}")
    else:
        months = _months_to_scan(today, back=months_lookback - 1)
        requested_from = requested_to = ""

    print("📅 Descobrindo dias disponíveis (Daily PDF) por banco...")
    daily_by_bank = {}
    for bank_code in BANK_FOLDERS:
        daily_by_bank[bank_code] = discover_daily_records(
            mpfm_root, bank_code, months,
            keep=None if requested_from else days_count + 5,
            date_from=requested_from or None,
            date_to=requested_to or None,
            workers=workers,
        )
        print(f"  {bank_code}: {len(daily_by_bank[bank_code])} dia(s) encontrados")

    all_days = sorted({day for recs in daily_by_bank.values() for day in recs}, reverse=True)
    if not all_days:
        print("❌ Nenhum dia com PDF Daily encontrado em nenhum banco. Abortando.")
        return

    target_days = sorted(all_days if requested_from else all_days[:days_count])
    target_days_set = set(target_days)
    selection_label = "Dias selecionados" if requested_from else f"Últimos {len(target_days)} dia(s) selecionados"
    print(f"\n✅ {selection_label}: {', '.join(target_days)}")

    most_recent = max(all_days)
    expected_most_recent = (date_to or (today - timedelta(days=1))).strftime("%Y-%m-%d")
    if most_recent < expected_most_recent:
        print(
            f"⚠️  ATENÇÃO: o dia mais recente disponível ({most_recent}) é anterior a ontem "
            f"({expected_most_recent}). Os dados podem não estar em dia."
        )

    for bank_code in BANK_FOLDERS:
        missing = [d for d in target_days if d not in daily_by_bank[bank_code]]
        if missing:
            print(f"⚠️  {bank_code}: sem Daily PDF para {', '.join(missing)}")

    print("\n📄 Lendo relatórios Hourly (apenas dos dias selecionados)...")
    hourly_by_bank_day = {}
    for bank_code in BANK_FOLDERS:
        hourly_by_bank_day[bank_code] = discover_hourly_records_for_days(mpfm_root, bank_code, months, target_days_set, workers)
        total_hours = sum(len(v) for v in hourly_by_bank_day[bank_code].values())
        print(f"  {bank_code}: {total_hours} registro(s) horário(s)")

    print("\n🧪 Localizando TXT do Separador de Testes (FC13/FC14/FC17)...")
    sep_data_by_day = load_sep_data_for_days(sep_root, target_days)
    for day in target_days:
        if day not in sep_data_by_day:
            print(f"  ⚠️  SEP não disponível para {day} — banco {aligned_bank} ficará sem colunas SEP nesse dia")

    all_rows = []
    for day in target_days:
        for bank_code in BANK_FOLDERS:
            daily_entry = daily_by_bank[bank_code].get(day)
            hourly_recs = hourly_by_bank_day[bank_code].get(day, [])
            is_aligned = bank_code == aligned_bank
            sep_data = sep_data_by_day.get(day) if is_aligned else None

            if hourly_recs:
                df_hourly = build_hourly_df_with_sep(hourly_recs, bank_code, sep_data)
                all_rows.extend(hourly_df_to_rows(df_hourly, sep_merged=bool(sep_data)))

            if daily_entry:
                _, daily_record = daily_entry
                df_daily = build_daily_df(daily_record, bank_code)
                sep_day = sep_data.get("DAY") if sep_data else None
                all_rows.extend(daily_df_to_rows(df_daily, sep_day))

                df_recon = build_recon_df(daily_record, hourly_recs, bank_code)
                all_rows.extend(recon_df_to_rows(df_recon))

        print(f"  📦 {day}: processado")

    if not all_rows:
        print("❌ Nenhuma linha gerada. Abortando sem escrever arquivo.")
        return

    df_out = pd.DataFrame(all_rows, columns=BASE_UNICA_COLUMNS)
    df_out.sort_values(
        by=["ProductionDate", "Bank", "Granularity", "Hour", "Tag"],
        inplace=True,
        na_position="last",
        key=lambda col: col.astype(str) if col.name != "Hour" else col,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_out.to_excel(output_path, sheet_name="BASE_UNICA_STANDALONE", index=False)
    add_comparativo_sheet(output_path)
    print(f"\n✅ Concluído: {len(df_out)} linha(s) gravadas em {output_path}")

    if update_master and master_output_path:
        try:
            stats = update_master_base_unica(master_output_path, df_out)
            print(
                "✅ Base_Unica total atualizada: "
                f"{master_output_path} "
                f"({stats['previous_rows']} anteriores + {stats['increment_rows']} novas "
                f"=> {stats['total_rows']} totais; "
                f"{stats['replaced_or_duplicate_rows']} substituídas/duplicadas)"
            )
        except PermissionError:
            print(f"⚠️  Não foi possível atualizar a Base_Unica total porque o arquivo está aberto: {master_output_path}")
        except Exception as exc:
            print(f"⚠️  Falha ao atualizar a Base_Unica total ({master_output_path}): {exc}")


if __name__ == "__main__":
    main()
