import openpyxl

path = "data/outputs/Relatorio_Corporativo_PE04_PE02_RiserP2_SEP_2026-02-03_a_2026-07-30_20260801_232054.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets:", wb.sheetnames)

ws = wb["Comparativo_PE02xRiserP2_D"]
print("\nComparativo diario - header row 2:")
print([c.value for c in ws[2]])
print("\nAmostra de linhas com desvio calculado:")
count = 0
for row in ws.iter_rows(min_row=3, values_only=True):
    # last two cols are % Desvio HC / % Desvio Total
    if row[-1] is not None or row[-2] is not None:
        print(row[0], "HC:", row[-2], "Total:", row[-1])
        count += 1
    if count >= 8:
        break

ws2 = wb["SepTeste_Diario"]
print("\nSepTeste_Diario header:")
print([c.value for c in ws2[2]])
count = 0
for row in ws2.iter_rows(min_row=3, values_only=True):
    if row[1] is not None:
        print(row[:6])
        count += 1
    if count >= 5:
        break

ws3 = wb["PE-04_Diario"]
print("\nPE-04_Diario amostra:")
for row in ws3.iter_rows(min_row=3, max_row=6, values_only=True):
    print(row)
