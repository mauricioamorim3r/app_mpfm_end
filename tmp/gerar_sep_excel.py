"""Gera Excel SEP diretamente (sem HTTP), copiando a lógica do endpoint."""
import sys, os
sys.path.insert(0, os.path.abspath("."))

from collections import defaultdict
from routes.export_routes import register_export_routes  # noqa - só para confirmar importação
import sqlite3
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from services.excel_template_service import SEP_EXPORT_TEMPLATE
from routes.date_utils import normalize_date_range

DATE_FROM = "2026-06-01"
DATE_TO   = "2026-08-13"
OUT = "data/outputs/SEP_JUN_AGO_2026.xlsx"

print(f"Gerando SEP: {DATE_FROM} -> {DATE_TO}")
print(f"Template: {SEP_EXPORT_TEMPLATE}")

def copy_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)

conn = sqlite3.connect("data/mpfm_local.db")
conn.row_factory = sqlite3.Row
cur = conn.cursor()

detail_kind_by_fluid = {
    "oleo": "sep_oleo_detail",
    "gas": "sep_gas_detail",
    "agua": "sep_agua_detail",
}
fluid_by_detail_kind = {v: k for k, v in detail_kind_by_fluid.items()}

detail_sql = (
    "SELECT day_ref, hour_ref, bank, tag, instrument, row_kind, metric_name, metric_value "
    "FROM measurements_active "
    "WHERE row_kind IN ('sep_oleo_detail','sep_gas_detail','sep_agua_detail') "
    "AND COALESCE(is_official,1)=1 AND day_ref BETWEEN ? AND ? "
    "ORDER BY day_ref, row_kind, COALESCE(hour_ref,-1), tag, metric_name"
)

detail_sheet_config = {
    "oleo": [("Hour","Hora"),("Pressure_kPa","Pressure (kpa)"),("Pressure_barg","Pressure (barg)"),
             ("Temperature_degC","Temperature (deg c)"),("SD_kg_sm3","SD (kg/sm³)"),
             ("MD_kg_m3","MD (kg/m³)"),("IV_m3","IV (m³)"),("GV_m3","GV (m³)"),
             ("GSV_sm3","GSV (sm³)"),("Mass_ton","Mass (t)"),("NSV_sm3","NSV (sm³)"),
             ("BSW_pct","BSW (%)"),("CPL","CPL"),("CTL","CTL")],
    "gas":  [("Hour","Hora"),("Pressure_kPa_g","Pressure (kpa_g)"),
             ("Temperature_degC","Temperature (deg c)"),("SD_kg_sm3","SD (kg/sm³)"),
             ("DT_kg_m3","DT (kg/m³)"),("GrVol_m3","Gr. vol. (m³)"),("StVol_m3","St. vol. (m³)"),
             ("Mass_t","Mass (t)"),("Energy_GJ","Energy (gj)"),
             ("DiffPress_kPa","Diff. press. (kpa)"),("Flowtime_min","Flowtime (min)")],
    "agua": [("Hour","Hora"),("Pressure_kPa","Pressure (kpa)"),
             ("Temperature_degC","Temperature (deg c)"),("SD_kg_sm3","SD (kg/sm³)"),
             ("MD_kg_m3","MD (kg/m³)"),("IV_m3","IV (m³)"),("GV_m3","GV (m³)"),
             ("GSV_sm3","GSV (sm³)"),("Mass_ton","Mass (t)"),("NSV_sm3","NSV (sm³)"),
             ("BSW_pct","BSW (%)"),("CPL","CPL"),("CTL","CTL")],
}

print("Consultando banco...", flush=True)
detail_store = {}
for fluid in detail_kind_by_fluid:
    detail_store[fluid] = {"pivot": defaultdict(dict), "meta": {}, "grouped": defaultdict(list), "day_rows": {}}

rows = cur.execute(detail_sql, [DATE_FROM, DATE_TO]).fetchall()
print(f"  {len(rows):,} linhas lidas", flush=True)

for row in rows:
    fluid = fluid_by_detail_kind[row["row_kind"]]
    key = (row["day_ref"], row["hour_ref"], row["tag"], row["instrument"])
    detail_store[fluid]["pivot"][key][row["metric_name"]] = row["metric_value"]
    detail_store[fluid]["meta"][key] = {
        "day_ref": row["day_ref"], "hour_ref": row["hour_ref"],
        "bank": row["bank"], "tag": row["tag"], "instrument": row["instrument"],
    }

conn.close()

for fluid, payload in detail_store.items():
    for key in sorted(payload["pivot"].keys(), key=lambda item: (item[0] or "", -1 if item[1] is None else int(item[1]), item[2] or "", item[3] or "")):
        meta = payload["meta"][key]
        payload["grouped"][meta["day_ref"]].append((key, meta))
        if meta["hour_ref"] is None:
            payload["day_rows"][meta["day_ref"]] = (key, meta)

print("Abrindo template...", flush=True)
wb = openpyxl.load_workbook(str(SEP_EXPORT_TEMPLATE)) if SEP_EXPORT_TEMPLATE.exists() else openpyxl.Workbook()

def add_fluid_sheet(sheet_name: str, fluid: str):
    config = detail_sheet_config[fluid]
    metric_keys = [k for k, _ in config if k != "Hour"]
    metric_headers = {k: h for k, h in config}
    payload = detail_store[fluid]
    all_days = sorted(set(k[0] for k in payload["pivot"].keys()))
    if not all_days:
        return

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    ws.delete_rows(2, ws.max_row)

    ws.cell(1, 1).value = "Data"
    ws.cell(1, 2).value = "Tag"
    ws.cell(1, 3).value = "Instrumento"
    ws.cell(1, 4).value = "Hora"
    for ci, key in enumerate(metric_keys, 5):
        ws.cell(1, ci).value = metric_headers.get(key, key)

    row_idx = 2
    for day in all_days:
        day_entries = payload["grouped"].get(day, [])
        for key, meta in day_entries:
            metrics = payload["pivot"][key]
            hr = meta["hour_ref"]
            ws.cell(row_idx, 1).value = day
            ws.cell(row_idx, 2).value = meta["tag"]
            ws.cell(row_idx, 3).value = meta["instrument"]
            ws.cell(row_idx, 4).value = "Diário" if hr is None else int(hr)
            for ci, mkey in enumerate(metric_keys, 5):
                ws.cell(row_idx, ci).value = metrics.get(mkey)
            row_idx += 1
    print(f"  {sheet_name}: {row_idx-2} linhas", flush=True)

print("Gerando abas...", flush=True)
add_fluid_sheet("separador oleo", "oleo")
add_fluid_sheet("separador gas", "gas")
add_fluid_sheet("separador agua", "agua")

print(f"Salvando {OUT}...", flush=True)
wb.save(OUT)
print(f"Concluido: {os.path.getsize(OUT)/1024:.1f} KB")
