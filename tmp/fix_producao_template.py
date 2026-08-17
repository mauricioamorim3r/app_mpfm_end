"""Limpa o template dados_producao.xlsx mantendo só as linhas de estilo (rows 1-6)."""
import sys, shutil, os
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from services.excel_template_service import PRODUCTION_EXPORT_TEMPLATE

# Linhas a manter: 1-2 (header/subtitle), 3 (vazia), 4 (colunas), 5-6 (estilo even/odd)
KEEP = 6

bak = Path(str(PRODUCTION_EXPORT_TEMPLATE).replace(".xlsx", "_BACKUP.xlsx"))
shutil.copy2(PRODUCTION_EXPORT_TEMPLATE, bak)
print(f"Backup: {bak.name} ({round(os.path.getsize(bak)/1024,1)} KB)")

wb = openpyxl.load_workbook(str(PRODUCTION_EXPORT_TEMPLATE))

for name in wb.sheetnames:
    ws = wb[name]
    original = ws.max_row
    if original <= KEEP:
        print(f"  [{name}] OK ({original} linhas)")
        continue
    # Remove merged cells nas linhas a deletar
    for m in list(ws.merged_cells.ranges):
        if m.min_row > KEEP:
            ws.unmerge_cells(str(m))
    # Deleta linhas extras
    ws.delete_rows(KEEP + 1, original - KEEP)
    print(f"  [{name}] {original} -> {ws.max_row} linhas (removidas {original - ws.max_row})")

wb.save(str(PRODUCTION_EXPORT_TEMPLATE))
print(f"Salvo: {PRODUCTION_EXPORT_TEMPLATE.name} ({round(os.path.getsize(PRODUCTION_EXPORT_TEMPLATE)/1024,1)} KB)")
