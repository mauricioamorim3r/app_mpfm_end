"""
Relatorio corporativo consolidado - PE-04, PE-02 x Riser P2 e Separador de Testes.

Periodo: 2026-02-03 a 2026-07-30
Fonte: measurements_curated (SQLite, is_official=1)

Abas geradas:
  - Capa                              (metodologia e legenda)
  - PE-04_Diario / PE-04_Horario      (subsea, sem par topside solicitado)
  - PE-02_Diario / PE-02_Horario      (subsea)
  - RiserP2_Diario / RiserP2_Horario  (topside)
  - Comparativo_PE02_RiserP2_Diario / _Horario  (desvio HC e Total, ref. Riser)
  - SepTeste_Diario / SepTeste_Horario

Colunas (nesta ordem, para cada aba de dados):
  Massa HC/Total/Oleo/Gas/Agua CORRIGIDA (t)
  Massa HC/Total/Oleo/Gas/Agua PADRAO (t)             [condicao padrao / PVT @20]
  Volume Oleo/Gas/Agua STD 20C 1atm (m3 / Sm3)
"""
from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableStyleInfo

DB_PATH = "data/mpfm_local.db"
DATE_FROM = "2026-02-03"
DATE_TO = "2026-07-30"
HOURS = list(range(1, 25))
LIMITE_HC_PCT = 10.0
LIMITE_TOTAL_PCT = 7.0
OUT_PATH = Path("data/outputs") / (
    f"Relatorio_Corporativo_PE04_PE02_RiserP2_SEP_{DATE_FROM}_a_{DATE_TO}_"
    f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
)

MASS_CORR_METRICS = {
    "hc": "MPFM corr HC (t)",
    "total": "MPFM corr Total (t)",
    "oleo": "MPFM corr Óleo (t)",
    "gas": "MPFM corr Gás (t)",
    "agua": "MPFM corr Água (t)",
}
MASS_STD_METRICS = {
    "oleo": "PVT @20 mass Óleo (t)",
    "gas": "PVT @20 mass Gás (t)",
    "agua": "PVT @20 mass Água (t)",
}
VOL_STD_METRICS = {
    "oleo": "PVT @20 vol Óleo (m³)",
    "gas": "PVT @20 vol Gás (Sm³)",
    "agua": "PVT @20 vol Água (m³)",
}

FINAL_COLUMNS = [
    "Massa HC Corrigida (t)",
    "Massa Total Corrigida (t)",
    "Massa Óleo Corrigida (t)",
    "Massa Gás Corrigida (t)",
    "Massa Água Corrigida (t)",
    "Massa HC Padrão (t)",
    "Massa Total Padrão (t)",
    "Massa Óleo Padrão (t)",
    "Massa Gás Padrão (t)",
    "Massa Água Padrão (t)",
    "Volume Óleo STD 20°C/1atm (m³)",
    "Volume Gás STD 20°C/1atm (Sm³)",
    "Volume Água STD 20°C/1atm (m³)",
]

MPFM_ENTITIES = {
    "PE-04": {"bank": "B05", "tag": "PE_4"},
    "PE-02": {"bank": "B10", "tag": "PE_2"},
    "RiserP2": {"bank": "B08", "tag": "Riser_P2"},
}


def date_list(d0: str, d1: str) -> list[str]:
    start = datetime.strptime(d0, "%Y-%m-%d")
    end = datetime.strptime(d1, "%Y-%m-%d")
    out = []
    cur = start
    while cur <= end:
        out.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return out


DAYS = date_list(DATE_FROM, DATE_TO)


def fetch_metric_map(conn, row_kind, bank, tag, metric_names, hour_filter=None):
    """Returns dict {(day_ref[, hour_ref]): {metric_name: value}}."""
    ph = ",".join("?" * len(metric_names))
    sql = (
        "SELECT day_ref, hour_ref, metric_name, metric_value "
        "FROM measurements_curated "
        "WHERE row_kind=? AND bank=? AND tag=? AND day_ref BETWEEN ? AND ? "
        f"AND metric_name IN ({ph}) AND COALESCE(is_official,1)=1"
    )
    params = [row_kind, bank, tag, DATE_FROM, DATE_TO, *metric_names]
    if hour_filter == "null":
        sql += " AND hour_ref IS NULL"
    elif hour_filter == "notnull":
        sql += " AND hour_ref IS NOT NULL"
    out = {}
    for day_ref, hour_ref, metric_name, value in conn.execute(sql, params):
        key = (day_ref, hour_ref) if hour_filter == "notnull" else day_ref
        out.setdefault(key, {})[metric_name] = value
    return out


def build_mpfm_rows(conn, bank, tag, hourly: bool) -> pd.DataFrame:
    row_kind = "hourly" if hourly else "daily"
    hour_filter = "notnull" if hourly else "null"
    corr = fetch_metric_map(conn, row_kind, bank, tag, list(MASS_CORR_METRICS.values()), hour_filter)
    std_mass = fetch_metric_map(conn, row_kind, bank, tag, list(MASS_STD_METRICS.values()), hour_filter)
    std_vol = fetch_metric_map(conn, row_kind, bank, tag, list(VOL_STD_METRICS.values()), hour_filter)

    records = []
    keys = [(d, h) for d in DAYS for h in HOURS] if hourly else DAYS
    for key in keys:
        day_ref, hour_ref = key if hourly else (key, None)
        c = corr.get(key, {})
        sm = std_mass.get(key, {})
        sv = std_vol.get(key, {})
        oleo_std = sm.get(MASS_STD_METRICS["oleo"])
        gas_std = sm.get(MASS_STD_METRICS["gas"])
        agua_std = sm.get(MASS_STD_METRICS["agua"])
        hc_std = (oleo_std or 0) + (gas_std or 0) if (oleo_std is not None or gas_std is not None) else None
        total_std = (hc_std or 0) + (agua_std or 0) if (hc_std is not None or agua_std is not None) else None
        row = {
            "Data": day_ref,
            "Massa HC Corrigida (t)": c.get(MASS_CORR_METRICS["hc"]),
            "Massa Total Corrigida (t)": c.get(MASS_CORR_METRICS["total"]),
            "Massa Óleo Corrigida (t)": c.get(MASS_CORR_METRICS["oleo"]),
            "Massa Gás Corrigida (t)": c.get(MASS_CORR_METRICS["gas"]),
            "Massa Água Corrigida (t)": c.get(MASS_CORR_METRICS["agua"]),
            "Massa HC Padrão (t)": hc_std,
            "Massa Total Padrão (t)": total_std,
            "Massa Óleo Padrão (t)": oleo_std,
            "Massa Gás Padrão (t)": gas_std,
            "Massa Água Padrão (t)": agua_std,
            "Volume Óleo STD 20°C/1atm (m³)": sv.get(VOL_STD_METRICS["oleo"]),
            "Volume Gás STD 20°C/1atm (Sm³)": sv.get(VOL_STD_METRICS["gas"]),
            "Volume Água STD 20°C/1atm (m³)": sv.get(VOL_STD_METRICS["agua"]),
        }
        if hourly:
            row["Hora"] = hour_ref
        records.append(row)

    cols = (["Data", "Hora"] if hourly else ["Data"]) + FINAL_COLUMNS
    df = pd.DataFrame.from_records(records, columns=cols)
    return df


def fetch_sep_mass(conn, hour_filter) -> dict:
    metrics = ["oil_t", "gas_t", "water_t", "hc_t", "total_t"]
    ph = ",".join("?" * len(metrics))
    sql = (
        "SELECT day_ref, hour_ref, metric_name, metric_value FROM measurements_curated "
        "WHERE row_kind='sep' AND bank='SEP' AND day_ref BETWEEN ? AND ? "
        f"AND metric_name IN ({ph}) AND COALESCE(is_official,1)=1"
    )
    params = [DATE_FROM, DATE_TO, *metrics]
    sql += " AND hour_ref IS NULL" if hour_filter == "null" else " AND hour_ref IS NOT NULL"
    out = {}
    for day_ref, hour_ref, metric_name, value in conn.execute(sql, params):
        key = (day_ref, hour_ref) if hour_filter == "notnull" else day_ref
        out.setdefault(key, {})[metric_name] = value
    return out


def fetch_sep_volume(conn, row_kind, tag, metric_name, hour_filter) -> dict:
    sql = (
        "SELECT day_ref, hour_ref, metric_value FROM measurements_curated "
        "WHERE row_kind=? AND bank='SEP' AND tag=? AND day_ref BETWEEN ? AND ? "
        "AND metric_name=? AND COALESCE(is_official,1)=1"
    )
    params = [row_kind, tag, DATE_FROM, DATE_TO, metric_name]
    sql += " AND hour_ref IS NULL" if hour_filter == "null" else " AND hour_ref IS NOT NULL"
    out = {}
    for day_ref, hour_ref, value in conn.execute(sql, params):
        key = (day_ref, hour_ref) if hour_filter == "notnull" else day_ref
        out[key] = value
    return out


def build_sep_rows(conn, hourly: bool) -> pd.DataFrame:
    hour_filter = "notnull" if hourly else "null"
    mass = fetch_sep_mass(conn, hour_filter)
    vol_oleo = fetch_sep_volume(conn, "sep_oleo_detail", "20FT0247", "GSV_sm3", hour_filter)
    vol_gas = fetch_sep_volume(conn, "sep_gas_detail", "20FT0244", "StVol_m3", hour_filter)
    vol_agua = fetch_sep_volume(conn, "sep_agua_detail", "20FT0251", "GSV_sm3", hour_filter)

    records = []
    keys = [(d, h) for d in DAYS for h in HOURS] if hourly else DAYS
    for key in keys:
        day_ref, hour_ref = key if hourly else (key, None)
        m = mass.get(key, {})
        oleo_m = m.get("oil_t")
        gas_m = m.get("gas_t")
        agua_m = m.get("water_t")
        hc_m = m.get("hc_t")
        total_m = m.get("total_t")
        row = {
            "Data": day_ref,
            # Separador de testes: massa unica (corrigida = padrao de referencia da apropriacao)
            "Massa HC Corrigida (t)": hc_m,
            "Massa Total Corrigida (t)": total_m,
            "Massa Óleo Corrigida (t)": oleo_m,
            "Massa Gás Corrigida (t)": gas_m,
            "Massa Água Corrigida (t)": agua_m,
            "Massa HC Padrão (t)": hc_m,
            "Massa Total Padrão (t)": total_m,
            "Massa Óleo Padrão (t)": oleo_m,
            "Massa Gás Padrão (t)": gas_m,
            "Massa Água Padrão (t)": agua_m,
            "Volume Óleo STD 20°C/1atm (m³)": vol_oleo.get(key),
            "Volume Gás STD 20°C/1atm (Sm³)": vol_gas.get(key),
            "Volume Água STD 20°C/1atm (m³)": vol_agua.get(key),
        }
        if hourly:
            row["Hora"] = hour_ref
        records.append(row)

    cols = (["Data", "Hora"] if hourly else ["Data"]) + FINAL_COLUMNS
    return pd.DataFrame.from_records(records, columns=cols)


def build_comparison_rows(subsea_df: pd.DataFrame, topside_df: pd.DataFrame, hourly: bool) -> pd.DataFrame:
    key_cols = ["Data", "Hora"] if hourly else ["Data"]
    s = subsea_df.set_index(key_cols)
    t = topside_df.set_index(key_cols)
    records = []
    for key in s.index:
        srow = s.loc[key]
        trow = t.loc[key] if key in t.index else None
        hc_subsea = srow["Massa HC Corrigida (t)"]
        total_subsea = srow["Massa Total Corrigida (t)"]
        hc_riser = trow["Massa HC Corrigida (t)"] if trow is not None else None
        total_riser = trow["Massa Total Corrigida (t)"] if trow is not None else None

        desvio_hc = (
            round(((hc_subsea / hc_riser) - 1) * 100, 2)
            if hc_subsea is not None and pd.notna(hc_subsea) and hc_riser not in (None, 0) and pd.notna(hc_riser)
            else None
        )
        desvio_total = (
            round(((total_subsea / total_riser) - 1) * 100, 2)
            if total_subsea is not None and pd.notna(total_subsea) and total_riser not in (None, 0) and pd.notna(total_riser)
            else None
        )

        rec = {}
        if hourly:
            rec["Data"], rec["Hora"] = key
        else:
            rec["Data"] = key
        for col in FINAL_COLUMNS:
            rec[f"PE-02 (Subsea) · {col}"] = srow[col]
        for col in FINAL_COLUMNS:
            rec[f"Riser P2 (Topside) · {col}"] = trow[col] if trow is not None else None
        rec["% Desvio HC"] = desvio_hc
        rec["% Desvio Total"] = desvio_total
        records.append(rec)

    key_hdr = ["Data", "Hora"] if hourly else ["Data"]
    cols = key_hdr + [f"PE-02 (Subsea) · {c}" for c in FINAL_COLUMNS] + \
        [f"Riser P2 (Topside) · {c}" for c in FINAL_COLUMNS] + ["% Desvio HC", "% Desvio Total"]
    return pd.DataFrame.from_records(records, columns=cols)


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBTITLE_FONT = Font(italic=True, size=9, color="555555")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _mass_or_vol_col(col_name: str) -> bool:
    return col_name.startswith("Massa") or col_name.startswith("Volume") or "· Massa" in col_name or "· Volume" in col_name


def write_sheet(wb: Workbook, title: str, subtitle: str, df: pd.DataFrame, percent_cols=None):
    ws = wb.create_sheet(title=title[:31])
    percent_cols = percent_cols or []

    ws.cell(1, 1, subtitle)
    ws.cell(1, 1).font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 1))

    header_row = 2
    for c_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(header_row, c_idx, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for r_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for c_idx, (col, value) in enumerate(zip(df.columns, row), start=1):
            cell = ws.cell(r_idx, c_idx, value if pd.notna(value) else None)
            cell.border = BORDER
            if col == "Data":
                cell.number_format = "yyyy-mm-dd"
            elif col == "Hora":
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center")
            elif col in percent_cols:
                cell.number_format = '0.00"%"'
            elif _mass_or_vol_col(col):
                cell.number_format = "#,##0.000"

    last_row = header_row + len(df)
    last_col = len(df.columns)
    if last_row > header_row:
        table_ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"
        ws.auto_filter.ref = table_ref
        safe_name = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in title)[:28]
        table = Table(displayName=safe_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    # Conditional formatting for deviation columns
    for pct_col in percent_cols:
        if pct_col not in df.columns:
            continue
        col_idx = list(df.columns).index(pct_col) + 1
        col_letter = get_column_letter(col_idx)
        limit = LIMITE_HC_PCT if "HC" in pct_col else LIMITE_TOTAL_PCT
        rng = f"{col_letter}{header_row + 1}:{col_letter}{max(last_row, header_row + 1)}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f"ABS({col_letter}{header_row + 1})>{limit}"], fill=PatternFill("solid", fgColor="F8CBAD")),
        )
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f"AND(ABS({col_letter}{header_row + 1})<={limit},{col_letter}{header_row + 1}<>\"\")"],
                        fill=PatternFill("solid", fgColor="C6E0B4")),
        )

    ws.freeze_panes = ws.cell(header_row + 1, 2 if "Hora" in df.columns else 2)
    for c_idx, col in enumerate(df.columns, start=1):
        width = 14
        if col in ("Data", "Hora"):
            width = 10
        elif "·" in col:
            width = 30
        elif len(col) > 20:
            width = 22
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.sheet_view.showGridLines = False
    return ws


def write_cover(wb: Workbook):
    ws = wb.create_sheet(title="Capa", index=0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    lines = [
        ("Relatório Corporativo — Medição MPFM e Separador de Testes", TITLE_FONT),
        ("", None),
        (f"Período: {DATE_FROM} a {DATE_TO}", Font(bold=True, size=11)),
        (f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", Font(size=10, color="555555")),
        ("", None),
        ("Escopo", Font(bold=True, size=12, color="1F4E78")),
        ("• PE-04 (subsea) — banco B05 / tag PE_4", None),
        ("• PE-02 (subsea) — banco B10 / tag PE_2", None),
        ("• Riser P2 (topside) — banco B08 / tag Riser_P2", None),
        ("• Separador de Testes — bank SEP (20FT0247 óleo, 20FT0244 gás, 20FT0251 água)", None),
        ("", None),
        ("Estrutura de colunas", Font(bold=True, size=12, color="1F4E78")),
        ("Massa HC / Total / Óleo / Gás / Água CORRIGIDA (t) — condição de linha (medição MPFM corrigida / massa do Separador).", None),
        ("Massa HC / Total / Óleo / Gás / Água PADRÃO (t) — condição padrão (PVT @20°C). HC = Óleo + Gás; Total = HC + Água.", None),
        ("Volume Óleo / Gás / Água STD 20°C, 1 atm — volume em condição padrão (PVT @20 / GSV / St. vol. do Separador).", None),
        ("", None),
        ("Metodologia do desvio (aba Comparativo PE-02 × Riser P2)", Font(bold=True, size=12, color="1F4E78")),
        ("% Desvio HC   = ((Massa HC Corrigida PE-02 subsea / Massa HC Corrigida Riser P2 topside) − 1) × 100", None),
        ("% Desvio Total = ((Massa Total Corrigida PE-02 subsea / Massa Total Corrigida Riser P2 topside) − 1) × 100", None),
        ("Referência: Riser P2 (topside). Limites de aceitação: ± 10% (HC) e ± 7% (Total).", None),
        ("Verde = dentro do limite · Vermelho = fora do limite.", None),
        ("", None),
        ("Observação — Separador de Testes", Font(bold=True, size=12, color="1F4E78")),
        ("O separador de testes não possui trilha de correção independente (linha vs. padrão); os valores", None),
        ("de massa reportados representam a referência de apropriação e são repetidos nas colunas", None),
        ("CORRIGIDA e PADRÃO para manter a mesma estrutura de colunas das demais abas.", None),
        ("", None),
        ("Abas do relatório", Font(bold=True, size=12, color="1F4E78")),
        ("PE-04_Diario / PE-04_Horario", None),
        ("PE-02_Diario / PE-02_Horario", None),
        ("RiserP2_Diario / RiserP2_Horario", None),
        ("Comparativo_PE02xRiserP2_Diario / _Horario", None),
        ("SepTeste_Diario / SepTeste_Horario", None),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(i, 1, text)
        cell.font = font or Font(size=10)
    return ws


def main():
    conn = sqlite3.connect(DB_PATH)

    pe04_daily = build_mpfm_rows(conn, MPFM_ENTITIES["PE-04"]["bank"], MPFM_ENTITIES["PE-04"]["tag"], hourly=False)
    pe04_hourly = build_mpfm_rows(conn, MPFM_ENTITIES["PE-04"]["bank"], MPFM_ENTITIES["PE-04"]["tag"], hourly=True)

    pe02_daily = build_mpfm_rows(conn, MPFM_ENTITIES["PE-02"]["bank"], MPFM_ENTITIES["PE-02"]["tag"], hourly=False)
    pe02_hourly = build_mpfm_rows(conn, MPFM_ENTITIES["PE-02"]["bank"], MPFM_ENTITIES["PE-02"]["tag"], hourly=True)

    riserp2_daily = build_mpfm_rows(conn, MPFM_ENTITIES["RiserP2"]["bank"], MPFM_ENTITIES["RiserP2"]["tag"], hourly=False)
    riserp2_hourly = build_mpfm_rows(conn, MPFM_ENTITIES["RiserP2"]["bank"], MPFM_ENTITIES["RiserP2"]["tag"], hourly=True)

    comp_daily = build_comparison_rows(pe02_daily, riserp2_daily, hourly=False)
    comp_hourly = build_comparison_rows(pe02_hourly, riserp2_hourly, hourly=True)

    sep_daily = build_sep_rows(conn, hourly=False)
    sep_hourly = build_sep_rows(conn, hourly=True)

    conn.close()

    wb = Workbook()
    wb.remove(wb.active)
    write_cover(wb)

    write_sheet(wb, "PE-04_Diario", "PE-04 (Subsea) — dados diários — banco B05 / tag PE_4", pe04_daily)
    write_sheet(wb, "PE-04_Horario", "PE-04 (Subsea) — dados horários — banco B05 / tag PE_4", pe04_hourly)

    write_sheet(wb, "PE-02_Diario", "PE-02 (Subsea) — dados diários — banco B10 / tag PE_2", pe02_daily)
    write_sheet(wb, "PE-02_Horario", "PE-02 (Subsea) — dados horários — banco B10 / tag PE_2", pe02_hourly)

    write_sheet(wb, "RiserP2_Diario", "Riser P2 (Topside) — dados diários — banco B08 / tag Riser_P2", riserp2_daily)
    write_sheet(wb, "RiserP2_Horario", "Riser P2 (Topside) — dados horários — banco B08 / tag Riser_P2", riserp2_hourly)

    write_sheet(
        wb, "Comparativo_PE02xRiserP2_D",
        "Comparativo diário PE-02 (Subsea) × Riser P2 (Topside) — % Desvio = ((Subsea/Topside)-1)×100, ref. Riser",
        comp_daily, percent_cols=["% Desvio HC", "% Desvio Total"],
    )
    write_sheet(
        wb, "Comparativo_PE02xRiserP2_H",
        "Comparativo horário PE-02 (Subsea) × Riser P2 (Topside) — % Desvio = ((Subsea/Topside)-1)×100, ref. Riser",
        comp_hourly, percent_cols=["% Desvio HC", "% Desvio Total"],
    )

    write_sheet(wb, "SepTeste_Diario", "Separador de Testes — dados diários — bank SEP (20FT0247/20FT0244/20FT0251)", sep_daily)
    write_sheet(wb, "SepTeste_Horario", "Separador de Testes — dados horários — bank SEP (20FT0247/20FT0244/20FT0251)", sep_hourly)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print("Arquivo gerado:", OUT_PATH)
    print("Linhas por aba:")
    for name, df in [
        ("PE-04_Diario", pe04_daily), ("PE-04_Horario", pe04_hourly),
        ("PE-02_Diario", pe02_daily), ("PE-02_Horario", pe02_hourly),
        ("RiserP2_Diario", riserp2_daily), ("RiserP2_Horario", riserp2_hourly),
        ("Comparativo_Diario", comp_daily), ("Comparativo_Horario", comp_hourly),
        ("SepTeste_Diario", sep_daily), ("SepTeste_Horario", sep_hourly),
    ]:
        preenchidas = df["Massa HC Corrigida (t)"].notna().sum() if "Massa HC Corrigida (t)" in df.columns else "-"
        print(f"  {name}: {len(df)} linhas ({preenchidas} com Massa HC preenchida)")


if __name__ == "__main__":
    main()
