"""
Exporta dados brutos (MPFM subsea/topside + Separador) para Excel,
com indicacao explicita de linhas ausentes (dia ou dia+hora sem nenhum dado).

Escopo:
  - Periodo: 2026-02-01 a 2026-07-31 (ontem, considerando "hoje" = 2026-08-01)
  - Pontos MPFM: PE-04, PE-02, PW-104DA (subsea) e Riser P2, Riser P4, Riser P5 (topside)
  - Todas as variaveis (metric_name) ja existentes em measurements_curated para
    row_kind = daily, hourly e recon
  - Dados do Separador (bank='SEP'): consolidado (row_kind='sep') e detalhes
    horarios de oleo/gas/agua (sep_oleo_detail, sep_gas_detail, sep_agua_detail)

Cada aba de dados recebe a grade completa de datas (e horas, quando aplicavel)
do periodo solicitado. Onde nao ha nenhum registro no banco, a linha e mantida
com as colunas de metrica em branco e a coluna "Status_Dados" marcada como
"SEM DADOS" (linha destacada em vermelho) ou "PARCIAL" (algumas metricas
ausentes, destacada em amarelo).

Saida: data/outputs/Export_MPFM_SEP_2026-02-01_a_2026-07-31_<timestamp>.xlsx
"""
import itertools
import sqlite3
import sys
from datetime import datetime

import pandas as pd

DB_PATH = "data/mpfm_local.db"
DATE_FROM = "2026-02-01"
DATE_TO = "2026-07-31"
ALL_DATES = pd.date_range(DATE_FROM, DATE_TO).strftime("%Y-%m-%d").tolist()

MPFM_POINTS = [
    ("B05", "PE_4", "PE-04 (Subsea)"),
    ("B10", "PE_2", "PE-02 (Subsea)"),
    ("B15", "PW-104DA", "PW-104DA (Subsea)"),
    ("B08", "Riser_P2", "Riser P2 (Topside)"),
    ("B13", "Riser_P4", "Riser P4 (Topside)"),
    ("B03", "Riser_P5", "Riser P5 (Topside)"),
]

SEP_POINTS = [
    ("SEP", "Consolidado 24h (SEP)"),
    ("20FT0247", "Oleo (20FT0247)"),
    ("20FT0244", "Gas (20FT0244)"),
    ("20FT0251", "Agua (20FT0251)"),
]
SEP_TAG_LABELS = dict(SEP_POINTS)


def load_rows(conn, row_kind, bank_tag_pairs=None, bank=None):
    q = """
        SELECT bank, tag, day_ref, hour_ref, metric_name, metric_value, metric_unit
        FROM measurements_curated
        WHERE row_kind = ?
          AND day_ref BETWEEN ? AND ?
    """
    params = [row_kind, DATE_FROM, DATE_TO]
    if bank_tag_pairs:
        placeholders = ",".join(["(?,?)"] * len(bank_tag_pairs))
        q += f" AND (bank,tag) IN ({placeholders})"
        for b, t in bank_tag_pairs:
            params.extend([b, t])
    if bank:
        q += " AND bank = ?"
        params.append(bank)
    return pd.read_sql_query(q, conn, params=params)


def distinct_hours(conn, row_kind):
    rows = conn.execute(
        "SELECT DISTINCT hour_ref FROM measurements_curated WHERE row_kind=? AND hour_ref IS NOT NULL ORDER BY hour_ref",
        (row_kind,),
    ).fetchall()
    return [r[0] for r in rows]


def pivot_wide(df, index_cols):
    if df.empty:
        return pd.DataFrame(columns=index_cols)
    wide = df.pivot_table(
        index=index_cols, columns="metric_name", values="metric_value", aggfunc="first"
    )
    wide = wide.reset_index()
    wide.columns.name = None
    return wide


def build_full_grid(ids, id_cols, hours=None):
    """ids: list of tuples matching id_cols (sem data/hora). Retorna DataFrame com a grade completa."""
    if hours:
        combos = list(itertools.product(ids, ALL_DATES, hours))
        rows = [(*i, d, h) for i, d, h in combos]
        cols = id_cols + ["day_ref", "hour_ref"]
    else:
        combos = list(itertools.product(ids, ALL_DATES))
        rows = [(*i, d) for i, d in combos]
        cols = id_cols + ["day_ref"]
    return pd.DataFrame(rows, columns=cols)


def merge_with_status(grid, wide, id_cols, label_map, label_col_name):
    merge_cols = id_cols + ["day_ref"] + (["hour_ref"] if "hour_ref" in grid.columns else [])
    merged = grid.merge(wide, on=merge_cols, how="left")
    metric_cols = [c for c in merged.columns if c not in merge_cols]

    def status(row):
        vals = row[metric_cols]
        n_present = vals.notna().sum()
        if n_present == 0:
            return "SEM DADOS"
        if n_present < len(metric_cols):
            return "PARCIAL"
        return "OK"

    merged["Status_Dados"] = merged.apply(status, axis=1) if metric_cols else "SEM DADOS"

    if label_map is not None:
        if len(id_cols) == 2:
            keys = list(zip(merged[id_cols[0]], merged[id_cols[1]]))
            merged.insert(0, label_col_name, [label_map.get(k, f"{k[0]}/{k[1]}") for k in keys])
        else:
            merged.insert(0, label_col_name, merged[id_cols[0]].map(label_map).fillna(merged[id_cols[0]]))

    sort_cols = id_cols + ["day_ref"] + (["hour_ref"] if "hour_ref" in merged.columns else [])
    merged = merged.sort_values(by=sort_cols).reset_index(drop=True)

    status_col = merged.pop("Status_Dados")
    insert_at = len(id_cols) + (2 if label_map is not None else 1)
    merged.insert(insert_at, "Status_Dados", status_col)

    rename = {"day_ref": "Data", "hour_ref": "Hora", "bank": "Bank", "tag": "Tag"}
    merged = merged.rename(columns=rename)
    return merged


def style_status_column(ws, header_row=1):
    from openpyxl.styles import Font, PatternFill

    header = [c.value for c in ws[header_row]]
    if "Status_Dados" not in header:
        return
    col_idx = header.index("Status_Dados") + 1
    red_fill = PatternFill("solid", fgColor="F8CBAD")
    red_font = Font(color="9C0006", bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    yellow_font = Font(color="9C6500", bold=True)
    for row_cells in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        cell = row_cells[col_idx - 1]
        if cell.value == "SEM DADOS":
            for c in row_cells:
                c.fill = red_fill
            cell.font = red_font
        elif cell.value == "PARCIAL":
            for c in row_cells:
                c.fill = yellow_fill
            cell.font = yellow_font


def main():
    conn = sqlite3.connect(DB_PATH)
    label_map_mpfm = {(b, t): lbl for b, t, lbl in MPFM_POINTS}
    bank_tag_pairs = [(b, t) for b, t, _ in MPFM_POINTS]

    print("Lendo MPFM diario...")
    df_daily = load_rows(conn, "daily", bank_tag_pairs=bank_tag_pairs)
    wide_daily = pivot_wide(df_daily, ["bank", "tag", "day_ref"])
    grid_daily = build_full_grid(bank_tag_pairs, ["bank", "tag"])
    sheet_daily = merge_with_status(grid_daily, wide_daily, ["bank", "tag"], label_map_mpfm, "Ponto")

    print("Lendo MPFM horario...")
    hourly_hours = distinct_hours(conn, "hourly") or list(range(24))
    df_hourly = load_rows(conn, "hourly", bank_tag_pairs=bank_tag_pairs)
    wide_hourly = pivot_wide(df_hourly, ["bank", "tag", "day_ref", "hour_ref"])
    grid_hourly = build_full_grid(bank_tag_pairs, ["bank", "tag"], hours=hourly_hours)
    sheet_hourly = merge_with_status(grid_hourly, wide_hourly, ["bank", "tag"], label_map_mpfm, "Ponto")

    print("Lendo MPFM reconciliacao...")
    df_recon = load_rows(conn, "recon", bank_tag_pairs=bank_tag_pairs)
    wide_recon = pivot_wide(df_recon, ["bank", "tag", "day_ref"])
    grid_recon = build_full_grid(bank_tag_pairs, ["bank", "tag"])
    sheet_recon = merge_with_status(grid_recon, wide_recon, ["bank", "tag"], label_map_mpfm, "Ponto")

    sep_tags = [t for t, _ in SEP_POINTS]

    print("Lendo Separador consolidado...")
    sep_hours = distinct_hours(conn, "sep") or list(range(24))
    df_sep = load_rows(conn, "sep", bank="SEP")
    wide_sep = pivot_wide(df_sep, ["tag", "day_ref", "hour_ref"])
    grid_sep = build_full_grid([(t,) for t in sep_tags], ["tag"], hours=sep_hours)
    sheet_sep = merge_with_status(grid_sep, wide_sep, ["tag"], SEP_TAG_LABELS, "Ponto")

    print("Lendo Separador detalhe oleo...")
    oleo_hours = distinct_hours(conn, "sep_oleo_detail") or list(range(24))
    df_sep_oleo = load_rows(conn, "sep_oleo_detail", bank="SEP")
    wide_sep_oleo = pivot_wide(df_sep_oleo, ["tag", "day_ref", "hour_ref"])
    grid_sep_oleo = build_full_grid([("20FT0247",)], ["tag"], hours=oleo_hours)
    sheet_sep_oleo = merge_with_status(grid_sep_oleo, wide_sep_oleo, ["tag"], SEP_TAG_LABELS, "Ponto")

    print("Lendo Separador detalhe gas...")
    gas_hours = distinct_hours(conn, "sep_gas_detail") or list(range(24))
    df_sep_gas = load_rows(conn, "sep_gas_detail", bank="SEP")
    wide_sep_gas = pivot_wide(df_sep_gas, ["tag", "day_ref", "hour_ref"])
    grid_sep_gas = build_full_grid([("20FT0244",)], ["tag"], hours=gas_hours)
    sheet_sep_gas = merge_with_status(grid_sep_gas, wide_sep_gas, ["tag"], SEP_TAG_LABELS, "Ponto")

    print("Lendo Separador detalhe agua...")
    agua_hours = distinct_hours(conn, "sep_agua_detail") or list(range(24))
    df_sep_agua = load_rows(conn, "sep_agua_detail", bank="SEP")
    wide_sep_agua = pivot_wide(df_sep_agua, ["tag", "day_ref", "hour_ref"])
    grid_sep_agua = build_full_grid([("20FT0251",)], ["tag"], hours=agua_hours)
    sheet_sep_agua = merge_with_status(grid_sep_agua, wide_sep_agua, ["tag"], SEP_TAG_LABELS, "Ponto")

    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"data/outputs/Export_MPFM_SEP_2026-02-01_a_2026-07-31_{ts}.xlsx"

    sheets = [
        ("MPFM_Diario", sheet_daily),
        ("MPFM_Horario", sheet_hourly),
        ("MPFM_Reconciliacao", sheet_recon),
        ("SEP_Consolidado", sheet_sep),
        ("SEP_Detalhe_Oleo", sheet_sep_oleo),
        ("SEP_Detalhe_Gas", sheet_sep_gas),
        ("SEP_Detalhe_Agua", sheet_sep_agua),
    ]

    legenda = pd.DataFrame(
        {
            "Status_Dados": ["OK", "PARCIAL", "SEM DADOS"],
            "Significado": [
                "Todas as variaveis presentes para o dia/hora",
                "Algumas variaveis ausentes para o dia/hora (celulas em branco)",
                "Nenhum dado encontrado no banco para o dia/hora (linha destacada em vermelho)",
            ],
        }
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        legenda.to_excel(writer, sheet_name="Legenda", index=False)
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for name, _df in sheets:
            style_status_column(writer.sheets[name[:31]])

    print()
    print("=== RESUMO ===")
    for name, df in sheets:
        n_sem = (df["Status_Dados"] == "SEM DADOS").sum() if "Status_Dados" in df.columns else 0
        n_par = (df["Status_Dados"] == "PARCIAL").sum() if "Status_Dados" in df.columns else 0
        print(f"{name}: {len(df)} linhas, {len(df.columns)} colunas | SEM DADOS={n_sem} PARCIAL={n_par}")
    print(f"Arquivo gerado: {out_path}")


if __name__ == "__main__":
    sys.exit(main())
