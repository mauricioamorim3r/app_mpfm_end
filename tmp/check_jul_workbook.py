from openpyxl import load_workbook
import time

t0 = time.time()
wb = load_workbook(r"data/outputs/MPFM_JUL_2026.xlsx", read_only=True)
print(f"Carregado em {time.time()-t0:.1f}s (read_only)")
print(f"Total de abas: {len(wb.sheetnames)}")
from collections import Counter
prefixes = Counter(name.split('_')[0] for name in wb.sheetnames)
print(prefixes)
# amostra de nomes
print(wb.sheetnames[:15])
wb.close()
