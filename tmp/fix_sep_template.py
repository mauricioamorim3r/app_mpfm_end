"""Limpa o template SEP, removendo dados antigos e preservando só as linhas de estilo."""
import sys, shutil
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from services.excel_template_service import SEP_EXPORT_TEMPLATE

KEEP_ROWS = {
    "separador oleo":  27,  # row1=header, row2=day-sep, row3=DAY, rows4-27=horas
    "separador gas":   27,
    "separador agua":  27,
    "Separador_Totais": 2,  # row1=header, row2=summary style
}

# Backup
bak = Path(str(SEP_EXPORT_TEMPLATE).replace(".xlsx", "_BACKUP.xlsx"))
shutil.copy2(SEP_EXPORT_TEMPLATE, bak)
print(f"Backup salvo: {bak.name}")

wb = openpyxl.load_workbook(str(SEP_EXPORT_TEMPLATE))

for sheet_name, keep_rows in KEEP_ROWS.items():
    if sheet_name not in wb.sheetnames:
        print(f"  [{sheet_name}] NAO encontrada, pulando")
        continue
    ws = wb[sheet_name]
    original = ws.max_row
    if original <= keep_rows:
        print(f"  [{sheet_name}] ja OK ({original} linhas), sem alteracao")
        continue

    # Remove merged cells que estejam nas linhas a deletar
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row > keep_rows:
            ws.unmerge_cells(str(merged))

    # Limpa os valores das linhas extras (nao delete — mantém estrutura)
    for row_idx in range(keep_rows + 1, original + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None

    # Remove linhas de dados extras explicitamente
    rows_to_delete = original - keep_rows
    if rows_to_delete > 0:
        ws.delete_rows(keep_rows + 1, rows_to_delete)

    print(f"  [{sheet_name}] {original} -> {ws.max_row} linhas (removidas {original - ws.max_row})")

wb.save(str(SEP_EXPORT_TEMPLATE))
print(f"\nTemplate salvo: {SEP_EXPORT_TEMPLATE.name}")

# Confirma tamanho final
import os
print(f"Tamanho antes: {os.path.getsize(bak)/1024:.1f} KB")
print(f"Tamanho depois: {os.path.getsize(SEP_EXPORT_TEMPLATE)/1024:.1f} KB")
