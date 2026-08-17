"""
Exporta dados brutos (MPFM subsea/topside + Separador) para Excel,
com indicacao explicita de linhas ausentes (dia ou dia+hora sem nenhum dado).

Escopo:
  - Periodo: 2026-02-01 a 2026-07-31 (ontem, considerando "hoje" = 2026-08-01)
  - Pontos MPFM: PE-04, PE-02, PW-104DA (subsea) e Riser P2, Riser P4, Riser P5 (topside)
  - Todas as variaveis (metric_name) ja existentes em measurements_curated para
    row_kind = daily, hourly e recon
  - Dados do Separador (bank='SEP', tag='SEP' para o consolidado; 20FT0247/0244/0251
    para os detalhes de oleo/gas/agua), cada um com linha "diaria" (hour_ref NULL)
    e linhas horarias (hour_ref 1..24)

Cada aba recebe a grade completa de datas (e horas, quando aplicavel) do periodo
solicitado. Onde nao ha nenhum registro no banco, a linha e mantida com as
colunas de metrica em branco e a coluna "Status_Dados":
  OK        -> todas as variaveis presentes
  PARCIAL   -> algumas variaveis ausentes (linha em amarelo)
  SEM DADOS -> nenhuma variavel presente (linha em vermelho)

Saida: data/outputs/Export_MPFM_SEP_2026-02-01_a_2026-07-31_<timestamp>.xlsx
"""
import itertools
import sqlite3
import sys
from datetime import datetime

import pandas as pd

DB_PATH = "data/mpfm_local.db"
DATE_FROM = "2026-02-01"
DATE_TO = "2026-07-31"
ALL_DATES = pd.date_range(DATE_FROM, DATE_TO).strftime("%Y-%m-%d").tolist()

MPFM_POINTS = [
    ("B05", "PE_4", "PE-04 (Subsea)"),
    ("B10", "PE_2", "PE-02 (Subsea)"),
    ("B15", "PW-104DA", "PW-104DA (Subsea)"),
    ("B08", "Riser_P2", "Riser P2 (Topside)"),
    ("B13", "Riser_P4", "Riser P4 (Topside)"),
    ("B03", "Riser_P5", "Riser P5 (Topside)"),
]

SEP_LABELS = {
    "SEP": "Consolidado 24h (SEP)",
    "20FT0247": "Oleo (20FT0247)",
    "20FT0244": "Gas (20FT0244)",
    "20FT0251": "Agua (20FT0251)",
}


def load_rows(conn, row_kind, tag, hour_is_null):
    q = """
        SELECT bank, tag, day_ref, hour_ref, metric_name, metric_value
        FROM measurements_curated
        WHERE row_kind = ? AND tag = ? AND day_ref BETWEEN ? AND ?
          AND hour_ref IS {} NULL
    """.format("" if hour_is_null else "NOT")
    return pd.read_sql_query(q, conn, params=[row_kind, tag, DATE_FROM, DATE_TO])


def load_rows_mpfm(conn, row_kind, bank_tag_pairs):
    placeholders = ",".join(["(?,?)"] * len(bank_tag_pairs))
    q = f"""
        SELECT bank, tag, day_ref, hour_ref, metric_name, metric_value
        FROM measurements_curated
        WHERE row_kind = ? AND day_ref BETWEEN ? AND ?
          AND (bank,tag) IN ({placeholders})
    """
    params = [row_kind, DATE_FROM, DATE_TO]
    for b, t in bank_tag_pairs:
        params.extend([b, t])
    return pd.read_sql_query(q, conn, params=params)


def distinct_hours(conn, row_kind, tag=None):
    q = "SELECT DISTINCT hour_ref FROM measurements_curated WHERE row_kind=? AND hour_ref IS NOT NULL"
    params = [row_kind]
    if tag:
        q += " AND tag=?"
        params.append(tag)
    q += " ORDER BY hour_ref"
    return [r[0] for r in conn.execute(q, params).fetchall()]


def pivot_wide(df, index_cols):
    if df.empty:
        return pd.DataFrame(columns=index_cols)
    wide = df.pivot_table(index=index_cols, columns="metric_name", values="metric_value", aggfunc="first")
    wide = wide.reset_index()
    wide.columns.name = None
    return wide


def build_full_grid(ids, id_cols, hours=None):
    if hours:
        rows = [(*i, d, h) for i, d, h in itertools.product(ids, ALL_DATES, hours)]
        cols = id_cols + ["day_ref", "hour_ref"]
    else:
        rows = [(*i, d) for i, d in itertools.product(ids, ALL_DATES)]
        cols = id_cols + ["day_ref"]
    return pd.DataFrame(rows, columns=cols)


def merge_with_status(grid, wide, id_cols, label_map, label_col_name):
    merge_cols = id_cols + ["day_ref"] + (["hour_ref"] if "hour_ref" in grid.columns else [])
    merged = grid.merge(wide, on=merge_cols, how="left")
    metric_cols = [c for c in merged.columns if c not in merge_cols]

    if metric_cols:
        present_count = merged[metric_cols].notna().sum(axis=1)
        total = len(metric_cols)
        status = pd.Series("SEM DADOS", index=merged.index)
        status[present_count == total] = "OK"
        status[(present_count > 0) & (present_count < total)] = "PARCIAL"
    else:
        status = pd.Series("SEM DADOS", index=merged.index)
    merged["Status_Dados"] = status

    if label_map is not None:
        if len(id_cols) == 2:
            keys = list(zip(merged[id_cols[0]], merged[id_cols[1]]))
            merged.insert(0, label_col_name, [label_map.get(k, f"{k[0]}/{k[1]}") for k in keys])
        else:
            merged.insert(0, label_col_name, merged[id_cols[0]].map(label_map).fillna(merged[id_cols[0]]))

    sort_cols = id_cols + ["day_ref"] + (["hour_ref"] if "hour_ref" in merged.columns else [])
    merged = merged.sort_values(by=sort_cols).reset_index(drop=True)

    status_col = merged.pop("Status_Dados")
    insert_at = len(id_cols) + (2 if label_map is not None else 1)
    merged.insert(insert_at, "Status_Dados", status_col)

    rename = {"day_ref": "Data", "hour_ref": "Hora", "bank": "Bank", "tag": "Tag"}
    merged = merged.rename(columns=rename)
    return merged


def style_status_column(ws, header_row=1):
    from openpyxl.styles import Font, PatternFill

    header = [c.value for c in ws[header_row]]
    if "Status_Dados" not in header:
        return
    col_idx = header.index("Status_Dados") + 1
    red_fill = PatternFill("solid", fgColor="F8CBAD")
    red_font = Font(color="9C0006", bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    yellow_font = Font(color="9C6500", bold=True)
    for row_cells in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        cell = row_cells[col_idx - 1]
        if cell.value == "SEM DADOS":
            for c in row_cells:
                c.fill = red_fill
            cell.font = red_font
        elif cell.value == "PARCIAL":
            for c in row_cells:
                c.fill = yellow_fill
            cell.font = yellow_font


def build_sep_sheets(conn, row_kind, tag):
    label_map = SEP_LABELS
    df_daily = load_rows(conn, row_kind, tag, hour_is_null=True)
    wide_daily = pivot_wide(df_daily, ["tag", "day_ref"])
    grid_daily = build_full_grid([(tag,)], ["tag"])
    sheet_daily = merge_with_status(grid_daily, wide_daily, ["tag"], label_map, "Ponto")

    hours = distinct_hours(conn, row_kind, tag) or list(range(1, 25))
    df_hourly = load_rows(conn, row_kind, tag, hour_is_null=False)
    wide_hourly = pivot_wide(df_hourly, ["tag", "day_ref", "hour_ref"])
    grid_hourly = build_full_grid([(tag,)], ["tag"], hours=hours)
    sheet_hourly = merge_with_status(grid_hourly, wide_hourly, ["tag"], label_map, "Ponto")
    return sheet_daily, sheet_hourly


def main():
    conn = sqlite3.connect(DB_PATH)
    label_map_mpfm = {(b, t): lbl for b, t, lbl in MPFM_POINTS}
    bank_tag_pairs = [(b, t) for b, t, _ in MPFM_POINTS]

    print("Lendo MPFM diario...")
    df_daily = load_rows_mpfm(conn, "daily", bank_tag_pairs)
    wide_daily = pivot_wide(df_daily, ["bank", "tag", "day_ref"])
    grid_daily = build_full_grid(bank_tag_pairs, ["bank", "tag"])
    sheet_daily = merge_with_status(grid_daily, wide_daily, ["bank", "tag"], label_map_mpfm, "Ponto")

    print("Lendo MPFM horario...")
    hourly_hours = distinct_hours(conn, "hourly") or list(range(24))
    df_hourly = load_rows_mpfm(conn, "hourly", bank_tag_pairs)
    wide_hourly = pivot_wide(df_hourly, ["bank", "tag", "day_ref", "hour_ref"])
    grid_hourly = build_full_grid(bank_tag_pairs, ["bank", "tag"], hours=hourly_hours)
    sheet_hourly = merge_with_status(grid_hourly, wide_hourly, ["bank", "tag"], label_map_mpfm, "Ponto")

    print("Lendo MPFM reconciliacao...")
    df_recon = load_rows_mpfm(conn, "recon", bank_tag_pairs)
    wide_recon = pivot_wide(df_recon, ["bank", "tag", "day_ref"])
    grid_recon = build_full_grid(bank_tag_pairs, ["bank", "tag"])
    sheet_recon = merge_with_status(grid_recon, wide_recon, ["bank", "tag"], label_map_mpfm, "Ponto")

    print("Lendo Separador consolidado...")
    sep_diario, sep_horario = build_sep_sheets(conn, "sep", "SEP")

    print("Lendo Separador detalhe oleo...")
    oleo_diario, oleo_horario = build_sep_sheets(conn, "sep_oleo_detail", "20FT0247")

    print("Lendo Separador detalhe gas...")
    gas_diario, gas_horario = build_sep_sheets(conn, "sep_gas_detail", "20FT0244")

    print("Lendo Separador detalhe agua...")
    agua_diario, agua_horario = build_sep_sheets(conn, "sep_agua_detail", "20FT0251")

    conn.close()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = f"data/outputs/Export_MPFM_SEP_2026-02-01_a_2026-07-31_{ts}.xlsx"

    sheets = [
        ("MPFM_Diario", sheet_daily),
        ("MPFM_Horario", sheet_hourly),
        ("MPFM_Reconciliacao", sheet_recon),
        ("SEP_Diario", sep_diario),
        ("SEP_Horario", sep_horario),
        ("SEP_Oleo_Diario", oleo_diario),
        ("SEP_Oleo_Horario", oleo_horario),
        ("SEP_Gas_Diario", gas_diario),
        ("SEP_Gas_Horario", gas_horario),
        ("SEP_Agua_Diario", agua_diario),
        ("SEP_Agua_Horario", agua_horario),
    ]

    legenda = pd.DataFrame(
        {
            "Status_Dados": ["OK", "PARCIAL", "SEM DADOS"],
            "Significado": [
                "Todas as variaveis presentes para o dia/hora",
                "Algumas variaveis ausentes para o dia/hora (celulas em branco) - linha em amarelo",
                "Nenhum dado encontrado no banco para o dia/hora - linha em vermelho",
            ],
        }
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        legenda.to_excel(writer, sheet_name="Legenda", index=False)
        for name, df in sheets:
            df.to_excel(writer, sheet_name=name[:31], index=False)
        for name, _df in sheets:
            style_status_column(writer.sheets[name[:31]])

    print()
    print("=== RESUMO ===")
    for name, df in sheets:
        n_sem = (df["Status_Dados"] == "SEM DADOS").sum() if "Status_Dados" in df.columns else 0
        n_par = (df["Status_Dados"] == "PARCIAL").sum() if "Status_Dados" in df.columns else 0
        print(f"{name}: {len(df)} linhas, {len(df.columns)} colunas | SEM DADOS={n_sem} PARCIAL={n_par}")
    print(f"Arquivo gerado: {out_path}")


if __name__ == "__main__":
    sys.exit(main())
