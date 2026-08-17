from openpyxl import load_workbook

try:
    wb = load_workbook("data/outputs/MPFM_JUL_2026.xlsx")
    print("OK, sheets:", wb.sheetnames)
    wb.close()
except Exception as exc:
    print("ERRO ao abrir:", type(exc).__name__, exc)
