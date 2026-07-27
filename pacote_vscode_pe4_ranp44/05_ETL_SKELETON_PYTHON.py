"""Skeleton ETL PE-4 para Relatório Semestral MPFM RANP44.
Adaptar conectores `load_from_app_db` e `download_drive_files` para a aplicação local.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


@dataclass
class JobConfig:
    well: str = "PE_4"
    days: int = 180
    end_date: date | None = None
    template_path: Path = Path("Template_Relatorio_Desempenho_Semestral_MPFM_RANP44.xlsx")
    output_path: Path = Path("Relatorio_Desempenho_Semestral_MPFM_PE4.xlsx")
    raw_drive_dir: Path = Path("data/raw_drive")


def window_dates(end_date: date, days: int = 180) -> list[date]:
    start = end_date - timedelta(days=days - 1)
    return [start + timedelta(days=i) for i in range(days)]


def load_from_app_db(cfg: JobConfig) -> pd.DataFrame:
    """Implementar com o repositório real da aplicação.
    Deve retornar uma linha por dia ou dados horários que serão agregados.
    """
    return pd.DataFrame()


def load_drive_monthly_excels(cfg: JobConfig) -> pd.DataFrame:
    frames = []
    for path in cfg.raw_drive_dir.glob("MPFM_*.xlsx"):
        df = pd.read_excel(path)
        df["__source_file"] = path.name
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def normalize_drive_mpfm(df: pd.DataFrame, cfg: JobConfig) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    # Filtro estrito PE_4. Ajustar nomes conforme dataframe real.
    mask = False
    for col in ["Entity", "Well", "Poço", "Poco"]:
        if col in df.columns:
            mask = mask | (df[col].astype(str).str.upper() == cfg.well.upper())
    out = df.loc[mask].copy()
    if out.empty:
        return out
    out["ProductionDate"] = pd.to_datetime(out["ProductionDate"]).dt.date
    out["Massa_oleo_MPFM_t"] = out.get("MPFM corr Óleo (t)")
    out["Massa_gas_MPFM_t"] = out.get("MPFM corr Gás (t)")
    out["Massa_agua_MPFM_t"] = out.get("MPFM corr Água (t)")
    out["Fonte_arquivo"] = out.get("__source_file")
    out["Fonte_mestre"] = "drive_mpfm_monthly_excel"
    out["Status_dados"] = "PARCIAL_DRIVE_NORMALIZADO"
    return out


def consolidate_daily(frames: Iterable[pd.DataFrame], cfg: JobConfig) -> pd.DataFrame:
    dates = pd.DataFrame({"ProductionDate": window_dates(cfg.end_date or date.today(), cfg.days)})
    data = pd.concat([f for f in frames if f is not None and not f.empty], ignore_index=True) if frames else pd.DataFrame()
    if data.empty:
        dates["Usar?"] = "NÃO"
        dates["Status_dados"] = "BLOQUEIO_SEM_DADOS"
        return dates

    # Prioridade: app_db > drive_mpfm > xml042. Ajustar conforme origem real.
    # Aqui mantemos agregação conservadora por data.
    agg = data.groupby("ProductionDate", dropna=False).agg({
        "Massa_oleo_MPFM_t": "sum",
        "Massa_gas_MPFM_t": "sum",
        "Massa_agua_MPFM_t": "sum",
        "Fonte_arquivo": lambda x: "; ".join(sorted(set(map(str, x.dropna())))),
        "Status_dados": lambda x: "; ".join(sorted(set(map(str, x.dropna())))),
    }).reset_index()
    out = dates.merge(agg, on="ProductionDate", how="left")
    out["Usar?"] = out["Massa_oleo_MPFM_t"].notna().map({True: "SIM", False: "NÃO"})
    out["Status_dados"] = out["Status_dados"].fillna("BLOQUEIO_DIA_SEM_DADOS")
    return out


def write_template(cfg: JobConfig, daily: pd.DataFrame) -> None:
    wb = load_workbook(cfg.template_path)
    ws = wb["05_Historico_Diario_180d"]
    start_row = 6
    col_map = {
        "ProductionDate": 1,
        "Usar?": 5,
        "Massa_oleo_MPFM_t": 18,
        "Massa_gas_MPFM_t": 19,
        "Massa_agua_MPFM_t": 20,
        "Fonte_arquivo": 42,
        "Status_dados": 8,
    }
    for idx, row in daily.head(180).iterrows():
        excel_row = start_row + idx
        for field, col in col_map.items():
            if field in daily.columns:
                ws.cell(excel_row, col).value = row[field]
    wb.save(cfg.output_path)


def main() -> None:
    cfg = JobConfig(end_date=date.today())
    drive_raw = load_drive_monthly_excels(cfg)
    drive_norm = normalize_drive_mpfm(drive_raw, cfg)
    app_df = load_from_app_db(cfg)
    daily = consolidate_daily([app_df, drive_norm], cfg)
    write_template(cfg, daily)


if __name__ == "__main__":
    main()
