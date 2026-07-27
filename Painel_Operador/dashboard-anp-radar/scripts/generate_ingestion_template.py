from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parents[1]
OUTPUT = APP_DIR / "templates" / "Radar_ANP_Template_Ingestao.xlsx"


SHEETS = {
    "controle_ingestao_pi": [
        "lote_id",
        "origem_pi",
        "periodo_inicio",
        "periodo_fim",
        "data_exportacao",
        "responsavel",
        "status_revisao",
        "arquivo_origem",
        "observacao",
    ],
    "pi_series_export": [
        "Fonte de dados",
        "Tempo",
        "Valor",
        "unidade",
        "qualidade",
        "tag_ponto",
        "equipamento",
        "dominio_sugerido",
        "tipo_sinal_sugerido",
        "observacao",
    ],
    "pi_catalogo_sinais": [
        "Fonte de dados",
        "apelido_app",
        "dominio_sugerido",
        "tipo_sinal_sugerido",
        "tag_ponto",
        "equipamento",
        "fluido",
        "unidade",
        "prioridade_app",
        "usar_no_dashboard",
        "observacao",
    ],
    "pi_mapeamento_app": [
        "area_dashboard",
        "dominio_modelo",
        "tipo_sinal",
        "regra_uso",
        "unidade_saida",
        "comparacao",
        "observacao",
    ],
    "fontes_dados": [
        "source_id",
        "grupo",
        "descricao",
        "caminho",
        "tipo",
        "recursivo",
        "responsavel",
        "periodicidade_atualizacao",
        "observacao",
    ],
    "pontos_medicao": [
        "tag_ponto",
        "cod_instalacao",
        "instalacao",
        "fluido",
        "tipo_medicao_principal",
        "tipo_medicao_secundaria",
        "tipo_medidor",
        "computador_vazao",
        "numero_serie_medidor",
        "status_ativo",
        "min_operacao",
        "max_operacao",
        "incerteza_maxima",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "medicao_diaria": [
        "data_operacional",
        "tag_ponto",
        "familia_xml",
        "camada",
        "volume_bruto",
        "volume_corrigido",
        "volume_liquido",
        "totalizador_inicial",
        "totalizador_final",
        "pressao",
        "temperatura",
        "diferencial_pressao",
        "duracao_fluxo_min",
        "unidade",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "alarmes_eventos": [
        "data_hora_evento",
        "data_operacional",
        "origem",
        "tag_ponto",
        "equipamento",
        "tipo_evento",
        "parametro",
        "valor_anterior",
        "valor_novo",
        "unidade",
        "usuario",
        "motivo",
        "evidencia_esperada",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "analises_fisico_quimicas": [
        "data_amostra",
        "data_resultado",
        "ponto_ou_corrente",
        "fluido",
        "boletim",
        "tipo_analise",
        "api",
        "densidade",
        "bsw",
        "cromatografia_json",
        "pvt_versao",
        "metodo",
        "validade_inicio",
        "validade_fim",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "certificados_calibracao": [
        "instrumento_tag",
        "tag_ponto",
        "numero_serie",
        "tipo_instrumento",
        "laboratorio",
        "certificado",
        "data_calibracao",
        "data_validade",
        "faixa_calibrada_min",
        "faixa_calibrada_max",
        "unidade",
        "erro",
        "incerteza",
        "criterio_aceitacao",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "incerteza_medicao": [
        "tag_ponto",
        "data_referencia",
        "versao_memoria",
        "incerteza_calculada",
        "incerteza_maxima_permitida",
        "componentes_json",
        "metodo_calculo",
        "responsavel",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "planos_coleta": [
        "ponto_ou_corrente",
        "fluido",
        "tipo_analise",
        "periodicidade",
        "janela_inicio",
        "janela_fim",
        "proxima_execucao",
        "responsavel",
        "base_normativa",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "pam_limites": [
        "tag_ponto",
        "equipamento",
        "parametro",
        "pam_min",
        "pam_max",
        "limite_alarme_min",
        "limite_alarme_max",
        "faixa_medicao_min",
        "faixa_medicao_max",
        "faixa_calibrada_min",
        "faixa_calibrada_max",
        "unidade",
        "revisao_documento",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "obrigacoes_regulatorias": [
        "obrigacao_id",
        "obrigacao",
        "base_normativa",
        "periodicidade",
        "prazo",
        "evidencia_esperada",
        "aplicavel_a",
        "regra_conformidade",
        "source_file",
        "evidence_ref",
        "review_status",
    ],
    "regras_validacao": [
        "rule_id",
        "descricao",
        "entrada_necessaria",
        "calculo_ou_comparacao",
        "tolerancia",
        "severidade_quando_falha",
        "evidencia_obrigatoria",
        "status_regra",
    ],
    "evidencias": [
        "evidence_id",
        "source_file",
        "tipo_arquivo",
        "pagina",
        "aba",
        "linha",
        "timestamp",
        "hash_arquivo",
        "descricao",
        "extraido_por_ia",
        "review_status",
    ],
}


EXAMPLES = {
    "controle_ingestao_pi": [
        "PI-2026-06-18-001",
        "PI System / MODELOS",
        "2026-05-01",
        "2026-06-30",
        "2026-06-18 08:00:00",
        "Medição",
        "pendente",
        "MODELOS\\Bacalhau_Fiscal_metering_oil.csv",
        "Exportar CSV do PI ou colar as series na aba pi_series_export.",
    ],
    "pi_series_export": [
        "\\Bacalhau\\Fiscal Metering\\Oil\\FT-001|GSV",
        "2026-06-01 00:00:00",
        "1234,56",
        "m3",
        "Good",
        "FT-001",
        "Fiscal Metering Oil",
        "fiscal_metering",
        "oil",
        "Colunas Fonte de dados, Tempo e Valor são obrigatórias para o ETL.",
    ],
    "pi_catalogo_sinais": [
        "\\Bacalhau\\Fiscal Metering\\Oil\\FT-001|GSV",
        "Óleo fiscal FT-001",
        "fiscal_metering",
        "oil",
        "FT-001",
        "Fiscal Metering Oil",
        "Óleo",
        "m3",
        "alta",
        "sim",
        "Usado na comparação fiscal x multifásico.",
    ],
    "pi_mapeamento_app": [
        "comparacao_fiscal_mpfm",
        "fiscal_metering",
        "oil",
        "media_diaria_ponderada_por_amostras",
        "m3/d",
        "fiscal_x_mpfm",
        "O app usa agregados diários no painel de inteligência de medição.",
    ],
    "fontes_dados": [
        "anpPanel",
        "Exports Painel ANP",
        "Planilhas exportadas do Painel do Operador",
        r"C:\Dados\ANP\Painel",
        "folder",
        "sim",
        "Medição",
        "diaria",
        "varrer subpastas",
    ],
    "alarmes_eventos": [
        "2026-06-02 00:10:00",
        "2026-06-01",
        "PMAE 004",
        "43FT0102",
        "21JN111",
        "alteracao_parametro",
        "densidade",
        "0.8566",
        "0.8537",
        "-",
        "operador",
        "atualizacao por boletim",
        "lab_report_densidade",
        "004_....zip",
        "DADOS_BASICOS[43FT0102]",
        "pendente",
    ],
}


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="15202B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 2, 14), 42)


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)
        if sheet_name in EXAMPLES:
            ws.append(EXAMPLES[sheet_name])
        style_sheet(ws)

    readme = wb.create_sheet("README")
    readme.append(["campo", "descrição"])
    readme.append(["objetivo", "Template geral de contingência para ingestão auditável do Radar ANP."])
    readme.append(["fluxo_pi", "PI/MODELOS -> pi_series_export ou CSV em MODELOS -> build_dashboard_data.py -> dashboard-data.json e radar-anp.sqlite -> aplicação."])
    readme.append(["colunas_pi_obrigatorias", "Fonte de dados, Tempo e Valor. As demais colunas ajudam auditoria, mapeamento e revisão."])
    readme.append(["regra", "Todo valor manual deve apontar para source_file e evidence_ref."])
    readme.append(["status", "Use pendente, validado, rejeitado ou substituido."])
    style_sheet(readme)
    wb.move_sheet(readme, offset=-(len(wb.sheetnames) - 1))

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"wrote {OUTPUT}")


if __name__ == "__main__":
    main()
