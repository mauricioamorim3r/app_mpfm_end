from __future__ import annotations

from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parents[1]
TEMPLATE = APP_DIR / "templates" / "Radar_ANP_Template_Ingestao.xlsx"
REQUIRED_SHEETS = {
    "README",
    "controle_ingestao_pi",
    "pi_series_export",
    "pi_catalogo_sinais",
    "pi_mapeamento_app",
}


workbook = load_workbook(TEMPLATE, read_only=True, data_only=True)
try:
    missing = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing:
        raise SystemExit(f"missing sheets: {', '.join(missing)}")
    pi_sheet = workbook["pi_series_export"]
    header = [cell.value for cell in next(pi_sheet.iter_rows(max_row=1))]
    required_columns = ["Fonte de dados", "Tempo", "Valor"]
    missing_columns = [column for column in required_columns if column not in header]
    if missing_columns:
        raise SystemExit(f"missing PI columns: {', '.join(missing_columns)}")
    example = [cell.value for cell in next(pi_sheet.iter_rows(min_row=2, max_row=2))]
    example_by_column = dict(zip(header, example))
    for column in required_columns:
        if not example_by_column.get(column):
            raise SystemExit(f"missing PI example value: {column}")
    datetime.fromisoformat(str(example_by_column["Tempo"]))
    float(str(example_by_column["Valor"]).replace(".", "").replace(",", "."))
    print("template ok")
    print("sheets=" + ",".join(workbook.sheetnames[:5]))
finally:
    workbook.close()
