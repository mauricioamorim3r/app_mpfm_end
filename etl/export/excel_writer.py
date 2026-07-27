"""Escrita de dados no template Excel RANP 44."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


COLUMN_MAP = {
    # Conforme 02_CONFIG_RANP44_PE4.yaml e 03_MAPEAMENTO_CAMPOS
    "A": "ProductionDate",
    "B": "Mes_auto",
    "C": "Semana_ISO_auto",
    "D": "Tipo_dia",
    "E": "Usar?",
    "F": "MPFM_Tag",
    "G": "Ref_ID",
    "H": "Status_dados",
    "I": "GVF_avg_pct",
    "J": "GVF_min_pct",
    "K": "GVF_max_pct",
    "L": "Salinidade_avg",
    "M": "BSW_avg_pct",
    "N": "Pressao_min_bar",
    "O": "Pressao_max_bar",
    "P": "Temperatura_min_C",
    "Q": "Temperatura_max_C",
    "R": "Massa_oleo_MPFM_t",
    "S": "Massa_gas_MPFM_t",
    "T": "Massa_agua_MPFM_t",
    "U": "Massa_oleo_REF_t",
    "V": "Massa_gas_REF_t",
    "W": "Massa_agua_REF_t",
    "X": "Volume_oleo_MPFM_Sm3",
    "Y": "Volume_gas_MPFM_kSm3",
    "Z": "Volume_agua_MPFM_m3",
    "AP": "Fonte_arquivo",
    "AQ": "XML042_ref",
    "AR": "Qualidade",
}


def write_to_template(
    df: pd.DataFrame,
    template_path: Path,
    output_path: Path,
    well: str = "PE_4",
    master_sheet: str = "05_Historico_Diario_180d",
    first_data_row: int = 6,
) -> None:
    """Escreve dados consolidados no template Excel.
    
    Args:
        df: DataFrame com dados diários (180 linhas)
        template_path: Caminho do template Excel
        output_path: Caminho do arquivo de saída
        well: Nome do poço
        master_sheet: Nome da aba principal
        first_data_row: Primeira linha de dados (após cabeçalho)
    """
    print(f"\n📝 Escrevendo dados no template...")
    print(f"   Template: {template_path.name}")
    print(f"   Saída: {output_path.name}")
    print(f"   Aba: {master_sheet}")
    print(f"   Linhas: {len(df)}")
    
    # Carregar template
    try:
        wb = load_workbook(template_path)
    except PermissionError:
        raise PermissionError(
            f"Arquivo de template está aberto: {template_path}\n"
            "Feche o Excel e tente novamente."
        )
    
    if master_sheet not in wb.sheetnames:
        raise ValueError(
            f"Aba '{master_sheet}' não encontrada no template.\n"
            f"Abas disponíveis: {wb.sheetnames}"
        )
    
    ws = wb[master_sheet]
    
    # Adicionar colunas derivadas
    df = enrich_dataframe(df)
    
    # Escrever dados linha por linha
    for i, row_data in df.iterrows():
        excel_row = first_data_row + i
        
        for col_letter, field_name in COLUMN_MAP.items():
            if field_name not in df.columns:
                continue
            
            value = row_data[field_name]
            cell = ws[f"{col_letter}{excel_row}"]
            
            # Tratar valores nulos
            if pd.isna(value):
                cell.value = None
                continue
            
            # Formatar datas
            if field_name == "ProductionDate" and hasattr(value, "strftime"):
                cell.value = value
                cell.number_format = "DD/MM/YYYY"
            # Formatar números
            elif isinstance(value, (int, float)):
                cell.value = float(value)
                # Massas/volumes com 3 decimais
                if "Massa" in field_name or "Volume" in field_name:
                    cell.number_format = "#,##0.000"
                # Porcentagens com 2 decimais
                elif "pct" in field_name or "avg" in field_name:
                    cell.number_format = "0.00"
                # Pressão/temperatura com 2 decimais
                elif "Pressao" in field_name or "Temperatura" in field_name:
                    cell.number_format = "0.00"
            # Texto
            else:
                cell.value = str(value)
            
            # Destacar linhas com bloqueio
            if field_name == "Qualidade" and value == "BLOQUEIO":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True)
            elif field_name == "Qualidade" and value == "PARCIAL":
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif field_name == "Qualidade" and value == "OK":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Metadados no rodapé
    footer_row = first_data_row + len(df) + 2
    ws[f"A{footer_row}"] = f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws[f"A{footer_row}"].font = Font(italic=True, size=9)
    
    ws[f"A{footer_row + 1}"] = f"Poço: {well} | Período: {len(df)} dias"
    ws[f"A{footer_row + 1}"].font = Font(italic=True, size=9)
    
    # Salvar
    wb.save(output_path)
    print(f"   ✅ Arquivo salvo: {output_path}")
    print(f"   📄 Tamanho: {output_path.stat().st_size / 1024:.1f} KB")


def enrich_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Adiciona colunas derivadas (Mês auto, Semana ISO, Tipo dia)."""
    df = df.copy()
    
    # Mês automático
    if "ProductionDate" in df.columns:
        df["Mes_auto"] = pd.to_datetime(df["ProductionDate"]).dt.strftime("%Y-%m")
        df["Semana_ISO_auto"] = pd.to_datetime(df["ProductionDate"]).dt.isocalendar().week
        
        # Tipo de dia (útil/fim de semana)
        df["Tipo_dia"] = pd.to_datetime(df["ProductionDate"]).dt.dayofweek.apply(
            lambda x: "Útil" if x < 5 else "Fim de semana"
        )
    
    # Ref_ID (simplificado - usar separador/tag se disponível)
    if "MPFM_Tag" in df.columns:
        df["Ref_ID"] = df["MPFM_Tag"]
    else:
        df["Ref_ID"] = "PE_4_MPFM"
    
    # Campos de condição de contorno (min/max)
    # Se só temos média, usar como min/max também
    if "Pressao_bar" in df.columns:
        if "Pressao_min_bar" not in df.columns:
            df["Pressao_min_bar"] = df["Pressao_bar"]
        if "Pressao_max_bar" not in df.columns:
            df["Pressao_max_bar"] = df["Pressao_bar"]
    
    if "Temperatura_C" in df.columns:
        if "Temperatura_min_C" not in df.columns:
            df["Temperatura_min_C"] = df["Temperatura_C"]
        if "Temperatura_max_C" not in df.columns:
            df["Temperatura_max_C"] = df["Temperatura_C"]
    
    return df


def create_summary_sheet(
    wb,
    validation_result,
    audit_log: dict,
    sheet_name: str = "00_Resumo_Validacao",
) -> None:
    """Cria aba de resumo com validação e auditoria (opcional).
    
    Args:
        wb: Workbook openpyxl
        validation_result: ValidationResult
        audit_log: Dicionário de auditoria
        sheet_name: Nome da aba de resumo
    """
    # Criar nova aba (se não existe)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name, 0)  # Inserir como primeira aba
    
    # Limpar aba
    ws.delete_rows(1, ws.max_row)
    
    # Cabeçalho
    ws["A1"] = "RESUMO DE VALIDAÇÃO RANP 44"
    ws["A1"].font = Font(size=14, bold=True)
    
    row = 3
    
    # Status geral
    ws[f"A{row}"] = "Status:"
    ws[f"B{row}"] = "APROVADO ✅" if validation_result.is_valid else "BLOQUEADO ❌"
    ws[f"B{row}"].font = Font(
        bold=True,
        color="006100" if validation_result.is_valid else "9C0006"
    )
    row += 2
    
    # Bloqueios
    if validation_result.blocking_errors:
        ws[f"A{row}"] = "Bloqueios Críticos:"
        ws[f"A{row}"].font = Font(bold=True, color="9C0006")
        row += 1
        for error in validation_result.blocking_errors:
            ws[f"B{row}"] = f"• {error}"
            row += 1
        row += 1
    
    # Alertas
    if validation_result.warnings:
        ws[f"A{row}"] = "Alertas:"
        ws[f"A{row}"].font = Font(bold=True, color="FF6600")
        row += 1
        for warning in validation_result.warnings:
            ws[f"B{row}"] = f"• {warning}"
            row += 1
        row += 1
    
    # Métricas
    ws[f"A{row}"] = "Métricas:"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1
    
    for key, value in audit_log.items():
        if key in ["blocking_errors", "warnings", "status_breakdown", "arquivos_lista"]:
            continue  # Pular listas complexas
        ws[f"B{row}"] = f"{key}:"
        ws[f"C{row}"] = str(value)
        row += 1
    
    # Ajustar largura de colunas
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
