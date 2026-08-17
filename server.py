#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MPFM Manager Local v4
- Processamento local em Python
- Excel mensal incremental
- SQLite local para histórico/monitoramento
- Classificação de TXT por conteúdo
"""

import os, sys, re, json, shutil, tempfile, glob, pathlib, sqlite3, math, zipfile, hashlib, threading, warnings, time, secrets, base64, asyncio, logging
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Optional

from app_config import (
    APP_TITLE,
    APP_VERSION,
    AUTH_ENABLED,
    AUTH_PASSWORD,
    AUTH_USERNAME,
    BACKUP_DB_RETENTION,
    BACKUP_ZIP_RETENTION,
    BBL_PER_M3,
    DB_PATH,
    DEFAULT_DENSITY,
    DEFAULT_HOST,
    DEFAULT_PORT,
    GAS_SM3_PER_BOE,
    MONTH_PT,
    OUTPUT_DIR,
    PUBLIC_BASE_URL,
    DAILY_BACKUP_ENABLED,
    STARTUP_BACKUP_ENABLED,
    STATIC_DIR,
    UPLOAD_DIR,
    WORK_DIR,
    ensure_workdirs,
)
from repositories.cards import CardsRepository
from repositories.importing import ImportRepository
from repositories.sgmfm import SgmfmRepository
from repositories.sep import SepRepository
from routes.admin_routes import register_admin_routes
from db_schema import apply_schema, run_migrations, rebuild_active_view
from routes.cards_routes import register_cards_routes
from routes.export_routes import register_export_routes
from routes.monthly_reports_routes import register_monthly_reports_routes
from routes.methodology_flow_routes import register_methodology_flow_routes
from routes.mpfm_adjustment_routes import register_mpfm_adjustment_routes
from routes.ops_routes import register_ops_routes
from services.ops.monitoring_service import invalidate_months_cache
from cache_manager import invalidate_cache
from services.measurement_dimensions import ensure_measurement_dimensions, refresh_measurement_dimensions
from routes.painel_operador_routes import register_painel_operador_routes
from routes.recon_routes import register_recon_routes
from routes.sep_routes import register_sep_routes
from routes.sgmfm_routes import register_sgmfm_routes
from routes.system_routes import register_system_routes
from routes.xml042_routes import register_xml042_routes
from routes.ai_routes import router as ai_router
from routes.ai_agent_routes import router as ai_agent_router
from services.cards import build_daily_cards
from services.importing import SEP_SOURCE_UNIT_CODE, SEP_UNIT_BY_METER, classify_input, inspect_txt_content
from services.importing import prepare_ingestion_batches
from services.importing import load_import_state, save_import_state
from services.importing import process_monthly_mpfm_inputs
from services.importing import process_monthly_sep_inputs
from services.importing import (
    build_monthly_base_unica as build_monthly_base_unica_service,
    cleanup_workbook as cleanup_workbook_service,
    excel_name as excel_name_service,
)
from services.importing.base_unica_import_service import (
    apply_base_unica_import as apply_base_unica_import_service,
    preview_base_unica_import as preview_base_unica_import_service,
)
from services.painel_operador import PainelOperadorStagingService
from services.sep import (
    sep_detail_headers,
    sep_detail_kind,
    store_sep_fluid_detail,
    store_sep_measurements,
    upsert_sep_detail_row,
)
from services.excel_template_service import (
    MONTHLY_WORKBOOK_TEMPLATE,
    clear_value_region,
    reset_sheet_from_template,
    seed_row_from_template,
)
from services.sgmfm import (
    build_prefill_payload as build_sgmfm_prefill_payload,
    build_record_summary as build_sgmfm_record_summary,
    build_schema_payload as build_sgmfm_schema_payload,
    generate_record_code as generate_sgmfm_record_code,
    render_record_html as render_sgmfm_record_html,
)

try:
    from repositories.alarme import AlarmRepository
    _ALARM_REPOSITORY_IMPORT_ERROR = None
except Exception as exc:
    AlarmRepository = None
    _ALARM_REPOSITORY_IMPORT_ERROR = exc

try:
    from routes.alarme_routes import register_alarme_routes
    _ALARM_ROUTES_IMPORT_ERROR = None
except Exception as exc:
    register_alarme_routes = None
    _ALARM_ROUTES_IMPORT_ERROR = exc

# ── Engine ────────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))
import importlib.util

def load_engine():
    for candidate in [
        Path(__file__).parent / "mpfm_engine.py",
        Path(__file__).parent.parent / "mpfm_engine.py",
    ]:
        if candidate.exists():
            spec = importlib.util.spec_from_file_location("mpfm_engine", str(candidate))
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    raise FileNotFoundError("mpfm_engine.py não encontrado.")

engine = load_engine()

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import Response, FileResponse, HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from starlette.types import ASGIApp, Scope, Receive, Send
import uvicorn

ensure_workdirs()

class BasicAuthMiddleware:
    def __init__(self, app: ASGIApp):
        self.app = app

    async def __call__(self, scope: Scope, receive: Receive, send: Send):
        if scope["type"] != "http":
            await self.app(scope, receive, send)
            return

        state = scope.setdefault("state", {})
        state["mpfm_authenticated"] = False

        path = scope.get("path", "")
        method = scope.get("method", "")
        if not AUTH_ENABLED or path == "/api/health" or method == "OPTIONS":
            try:
                await self.app(scope, receive, send)
            except asyncio.CancelledError:
                pass
            return

        auth_header = ""
        for name, value in scope.get("headers", []):
            if name.lower() == b"authorization":
                auth_header = value.decode("latin1")
                break

        if auth_header.startswith("Basic "):
            try:
                decoded = base64.b64decode(auth_header[6:]).decode("utf-8")
                username, _, password = decoded.partition(":")
                ok = secrets.compare_digest(username, AUTH_USERNAME) & \
                     secrets.compare_digest(password, AUTH_PASSWORD)
                if ok:
                    state["mpfm_authenticated"] = True
                    try:
                        await self.app(scope, receive, send)
                    except asyncio.CancelledError:
                        pass
                    return
            except (ValueError, UnicodeDecodeError):
                pass

        response = Response(
            content="Acesso não autorizado. Informe usuário e senha.",
            status_code=401,
            headers={"WWW-Authenticate": 'Basic realm="MPFM Manager"'},
            media_type="text/plain; charset=utf-8",
        )
        await response(scope, receive, send)


def _require_auth_config():
    """Raise early if authentication is enabled but credentials are missing."""
    if not AUTH_ENABLED:
        return
    if not AUTH_USERNAME or not AUTH_PASSWORD:
        raise RuntimeError(
            "Autenticação habilitada (MPFM_AUTH_ENABLED=true) mas MPFM_AUTH_USER "
            "e/ou MPFM_AUTH_PASS não estão configurados."
        )


# Validate auth configuration at import time so the server fails fast.
_require_auth_config()


app = FastAPI(title=APP_TITLE, version=APP_VERSION)

# CORS: restrict to configured origins instead of allowing every domain.
_allowed_origins = os.getenv("MPFM_ALLOWED_ORIGINS", "http://localhost:8765,http://127.0.0.1:8765").split(",")
_allowed_origins = [origin.strip() for origin in _allowed_origins if origin.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "Accept", "X-Requested-With"],
    allow_credentials=True,
)
app.add_middleware(GZipMiddleware, minimum_size=1024, compresslevel=6)
app.add_middleware(BasicAuthMiddleware)


@app.middleware("http")
async def add_security_headers(request, call_next):
    """Attach essential HTTP security headers to every response."""
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

warnings.filterwarnings(
    "ignore",
    message="Conditional Formatting extension is not supported and will be removed",
    category=UserWarning,
)


# ── SQLite ────────────────────────────────────────────────────────────────────

BACKUP_DIR = WORK_DIR / "backups"

def db_conn():
    # Se houver uma conexão compartilhada no escopo da requisição, reutiliza.
    from services.db_scope import get_scoped_db_conn
    scoped = get_scoped_db_conn()
    if scoped is not None:
        return scoped()
    db_target = str(DB_PATH)
    conn = sqlite3.connect(db_target, uri=db_target.startswith("file:"), timeout=30.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 30000")
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA cache_size = -64000")
    except Exception as _pragma_err:
        print(f"[WARNING] SQLite PRAGMA failed: {_pragma_err}")
    return conn


def _backup_db_now(label: str = "auto") -> None:
    """Hot-copy the live DB via SQLite backup API (WAL-safe). Keeps last 14 copies."""
    if not DB_PATH.exists():
        return
    try:
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = BACKUP_DIR / f"mpfm_{label}_{ts}.db"
        src = sqlite3.connect(str(DB_PATH))
        dst = sqlite3.connect(str(dest))
        src.backup(dst)
        dst.close()
        src.close()
        _prune_backups(keep=BACKUP_DB_RETENTION)
    except Exception as e:
        print(f"[BACKUP] Falha ao criar backup: {e}")


def _prune_backups(keep: int = 14) -> None:
    """Remove oldest backup files beyond `keep` most recent."""
    files = sorted(BACKUP_DIR.glob("mpfm_*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in files[keep:]:
        try:
            old.unlink()
        except Exception:
            pass


def _daily_backup_loop() -> None:
    """Background thread: backup once per day, then sleep 24 h."""
    import time as _time
    _time.sleep(3600)  # first run after 1 h, not immediately at startup
    while True:
        try:
            _backup_db_now("daily")
            _build_backup_zip()
            _prune_backup_zips(keep=BACKUP_ZIP_RETENTION)
        except Exception as e:
            print(f"[BACKUP] Erro no backup diário: {e}")
        _time.sleep(86400)


def _prune_backup_zips(keep: int = 7) -> None:
    zips = sorted(OUTPUT_DIR.glob("MPFM_backup_*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in zips[keep:]:
        try:
            old.unlink()
        except Exception:
            pass


def init_db():
    conn = db_conn()
    cur = conn.cursor()
    apply_schema(cur)
    try:
        run_migrations(cur)
    except Exception as _mig_err:
        print(f"[ERROR] Falha em migracao do banco: {_mig_err}")
    conn.commit()
    from app_config import get_inactive_tag_associados
    rebuild_active_view(conn, get_inactive_tag_associados())
    ensure_measurement_dimensions(conn)
    conn.close()


# A cópia integral no bootstrap é opt-in: mesmo em thread ela compete por disco
# com as primeiras consultas do dashboard.
if STARTUP_BACKUP_ENABLED:
    threading.Thread(target=_backup_db_now, args=("startup",), daemon=True, name="startup-backup").start()
init_db()

if DAILY_BACKUP_ENABLED:
    threading.Thread(target=_daily_backup_loop, daemon=True, name="daily-backup").start()


def start_run(source_type: str, source_ref: str, density: float, files_count: int) -> int:
    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO processing_runs(started_at, source_type, source_ref, density, files_count) VALUES(?,?,?,?,?)",
        (datetime.now().isoformat(timespec='seconds'), source_type, source_ref, density, files_count),
    )
    run_id = cur.lastrowid
    try:
        conn.commit()
    finally:
        conn.close()
    return run_id


def finish_run(run_id: int, status: str, notes: Optional[dict] = None):
    conn = db_conn()
    try:
        conn.execute(
            "UPDATE processing_runs SET finished_at=?, status=?, notes_json=? WHERE id=?",
            (datetime.now().isoformat(timespec='seconds'), status, json.dumps(notes or {}, ensure_ascii=False), run_id),
        )
        conn.commit()
    finally:
        conn.close()



def _day_tag_to_iso(year: str, month: str, day_tag: str) -> str:
    try:
        parts = str(day_tag or '').split('_')
        dd = int(parts[0])
        mm = int(parts[1]) if len(parts) > 1 and str(parts[1]).isdigit() else int(month)
        yyyy = int(parts[2]) if len(parts) > 2 and str(parts[2]).isdigit() else int(year)
        return f"{yyyy:04d}-{mm:02d}-{dd:02d}"
    except Exception:
        return ''


def _get_sep_alignment(bank: str, production_date: str):
    return SepRepository(db_conn).get_sep_alignment(bank, production_date)


def _upsert_card_override(payload: dict):
    if not (payload.get('production_date') and payload.get('bank') and payload.get('card_type')):
        raise HTTPException(400, 'production_date, bank e card_type são obrigatórios')
    return CardsRepository(db_conn).upsert_card_override(payload)


def _daily_metric_groups(date_from: str, date_to: str, bank: str = ''):
    return build_daily_cards(CardsRepository(db_conn), date_from, date_to, bank)


def _has_sep_alignment(bank: str, production_date: str) -> bool:
    return _get_sep_alignment(bank, production_date) is not None


def _load_sep_data_by_day(production_date: str) -> dict:
    if not production_date:
        return {}
    conn = db_conn()
    rows = conn.execute(
        "SELECT hour_ref, metric_name, metric_value FROM measurements_curated WHERE row_kind='sep' AND bank='SEP' AND COALESCE(is_official,1)=1 AND day_ref=? ORDER BY COALESCE(hour_ref,999), metric_name",
        (production_date,)
    ).fetchall()
    conn.close()
    out = {}
    for hour_ref, metric_name, metric_value in rows:
        key = 'DAY' if hour_ref is None else int(hour_ref)
        out.setdefault(key, {})[metric_name] = metric_value
    return out


def _load_sep_data_by_range(date_from: str, date_to: str) -> dict[str, dict]:
    """Carrega dados SEP para vários dias em uma única query (evita N+1)."""
    if not date_from or not date_to:
        return {}
    conn = db_conn()
    rows = conn.execute(
        """
        SELECT day_ref, hour_ref, metric_name, metric_value
        FROM measurements_curated
        WHERE row_kind='sep' AND bank='SEP' AND COALESCE(is_official,1)=1
          AND day_ref>=? AND day_ref<=
        ? ORDER BY day_ref, COALESCE(hour_ref,999), metric_name
        """,
        (date_from, date_to),
    ).fetchall()
    conn.close()
    out: dict[str, dict] = {}
    for day_ref, hour_ref, metric_name, metric_value in rows:
        day_data = out.setdefault(day_ref, {})
        key = 'DAY' if hour_ref is None else int(hour_ref)
        day_data.setdefault(key, {})[metric_name] = metric_value
    return out


def _recompute_alignment_resolution(production_date: str, bank: str):
    return SepRepository(db_conn).recompute_alignment_resolution(production_date, bank)


def _recompute_card_resolution(production_date: str, bank: str, card_type: str, tag: str = '', instrument: str = ''):
    return CardsRepository(db_conn).recompute_card_resolution(production_date, bank, card_type, tag, instrument)


def _rebuild_sep_summary_for_day(production_date: str):
    conn = db_conn()
    row = conn.execute(
        """
        SELECT meter_id
        FROM sep_source_files
        WHERE production_date=? AND fluid_kind='sep_oleo' AND is_active=1 AND is_official=1
        ORDER BY id DESC
        LIMIT 1
        """,
        (production_date,),
    ).fetchone()
    conn.close()
    meter_id = (row["meter_id"] if row else "") or ""
    unit_code = SEP_UNIT_BY_METER.get(meter_id, SEP_SOURCE_UNIT_CODE)
    return SepRepository(db_conn).rebuild_sep_summary_from_detail(
        production_date,
        unit_code=unit_code,
    )


def db_upsert_sep_alignment(production_date: str, bank: str, mpfm_tag: str = '', sep_meter_id: str = '', sep_tag: str = 'SEP', notes: str = '', force_new: bool = True):
    now = datetime.now().isoformat(timespec='seconds')
    conn = db_conn(); conn.row_factory = sqlite3.Row; cur = conn.cursor()
    existing_official = cur.execute(
        "SELECT * FROM sep_alignments WHERE production_date=? AND bank=? AND is_active=1 AND COALESCE(is_official,1)=1 ORDER BY id DESC LIMIT 1",
        (production_date, bank)
    ).fetchone()
    if existing_official and not force_new and str(existing_official['sep_meter_id'] or '') == str(sep_meter_id or '') and str(existing_official['sep_tag'] or 'SEP') == str(sep_tag or 'SEP') and str(existing_official['mpfm_tag'] or '') == str(mpfm_tag or ''):
        cur.execute("UPDATE sep_alignments SET notes=?, updated_at=? WHERE id=?", (notes or '', now, existing_official['id']))
        new_id = existing_official['id']
    else:
        is_official = 0 if existing_official else 1
        status = 'pending' if existing_official else 'official'
        cur.execute(
            "INSERT INTO sep_alignments(production_date, bank, mpfm_tag, sep_meter_id, sep_tag, notes, is_active, is_official, resolution_status, created_at, updated_at) VALUES(?,?,?,?,?,?,1,?,?,?,?)",
            (production_date, bank, mpfm_tag or '', sep_meter_id or '', sep_tag or 'SEP', notes or '', is_official, status, now, now)
        )
        new_id = cur.lastrowid
    try:
        conn.commit()
    finally:
        conn.close()
    try:
        if production_date and len(production_date)>=7:
            yr, mo = production_date[:4], production_date[5:7]
            schedule_monthly_base_unica(OUTPUT_DIR / excel_name(yr, mo), yr, mo)
    except Exception:
        pass
    return new_id


def db_delete_sep_alignment(alignment_id: int):
    now = datetime.now().isoformat(timespec='seconds')
    conn = db_conn()
    row = conn.execute("SELECT production_date, bank FROM sep_alignments WHERE id=?", (alignment_id,)).fetchone()
    conn.execute("UPDATE sep_alignments SET is_active=0, is_official=0, resolution_status='deleted', updated_at=? WHERE id=?", (now, alignment_id))
    try:
        conn.commit()
    finally:
        conn.close()
    try:
        if row:
            _recompute_alignment_resolution(row['production_date'], row['bank'] if 'bank' in row.keys() else '')
    except Exception:
        pass
    try:
        production_date = row['production_date'] if row else ''
        if production_date and len(production_date)>=7:
            yr, mo = production_date[:4], production_date[5:7]
            schedule_monthly_base_unica(OUTPUT_DIR / excel_name(yr, mo), yr, mo)
    except Exception:
        pass

def db_log_file(
    run_id: int,
    filename: str,
    ext: str,
    file_type: str,
    unit_code: str,
    meter_id: str,
    location: str,
    content_date: str,
    report_start: str,
    report_end: str,
    excel_month: str,
    identity_key: str = '',
    time_source: str = '',
    file_hash: str = '',
    processed_ok: bool = True,
    message: str = '',
):
    ImportRepository(db_conn, _file_sha1, _infer_metric_unit).log_file(
        run_id,
        filename,
        ext,
        file_type,
        unit_code,
        meter_id,
        location,
        content_date,
        report_start,
        report_end,
        excel_month,
        identity_key,
        time_source,
        file_hash,
        processed_ok,
        message,
    )


def db_log_raw_file(run_id: int, source_path: Path, source_type: str, meta: dict) -> int | None:
    return ImportRepository(db_conn, _file_sha1, _infer_metric_unit).log_raw_file(run_id, source_path, source_type, meta)


def db_log_parsing_event(run_id: int, source_file_raw_id: int | None, parser_name: str, parser_stage: str, status: str, details: dict | None = None):
    ImportRepository(db_conn, _file_sha1, _infer_metric_unit).log_parsing_event(
        run_id, source_file_raw_id, parser_name, parser_stage, status, details
    )


def db_find_import_by_identity(identity_key: str) -> dict | None:
    return ImportRepository(db_conn, _file_sha1, _infer_metric_unit).find_latest_import_by_identity(identity_key)


def db_add_issue(run_id: int, excel_file: str, issue_type: str, severity: str, ref_key: str, day_ref: str, details: str):
    ImportRepository(db_conn, _file_sha1, _infer_metric_unit).add_issue(
        run_id, excel_file, issue_type, severity, ref_key, _normalize_day_ref(day_ref), details
    )


def _infer_metric_unit(col: str) -> str:
    m = re.search(r'\(([^)]+)\)', col or '')
    return m.group(1) if m else ''


def _normalize_day_ref(v) -> str:
    s = str(v or '').strip()
    if not s:
        return ''
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s[:10]):
        return s[:10]
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', s[:10])
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = re.match(r'^(\d{2})-(\d{2})-(\d{4})$', s[:10])
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return s[:10]

def _txt_report_kind(name: str) -> str:
    n = (name or '').lower()
    if '24hours' in n:
        return '24hours'
    if 'daily' in n:
        return 'daily'
    return 'other'


def _file_sha1(path: str) -> str:
    h = hashlib.sha1()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b''):
            h.update(chunk)
    return h.hexdigest()


def _recompute_sep_source_resolution(production_date: str, fluid_kind: str, meter_id: str):
    return ImportRepository(db_conn, _file_sha1, _infer_metric_unit).recompute_sep_source_resolution(
        production_date,
        fluid_kind,
        meter_id,
    )


def db_register_sep_source_file(
    file_path: str,
    fluid_kind: str,
    meter_id: str,
    location: str,
    production_date: str,
    report_start: str = '',
    report_end: str = '',
    identity_key: str = '',
    time_source: str = 'content',
):
    return ImportRepository(db_conn, _file_sha1, _infer_metric_unit).register_sep_source_file(
        file_path,
        fluid_kind,
        meter_id,
        location,
        production_date,
        report_start,
        report_end,
        identity_key,
        time_source,
    )


def db_store_sheet_rows(run_id: int, excel_file: str, sheet_name: str, rows):
    ImportRepository(db_conn, _file_sha1, _infer_metric_unit).store_sheet_rows(run_id, excel_file, sheet_name, rows)


def _sanitize_files_imported_history(target_month: str = ''):
    return ImportRepository(db_conn, _file_sha1, _infer_metric_unit).sanitize_files_imported_history(target_month)


def _normalize_admin_date_input(value: str) -> str:
    raw = str(value or '').strip()
    if not raw:
        return ''
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return raw


def _db_health_snapshot():
    try:
        conn = db_conn()
        cur = conn.cursor()
        payload = {
            'ok': True,
            'db_path': str(DB_PATH),
            'processing_runs': cur.execute("SELECT COUNT(*) FROM processing_runs").fetchone()[0],
            'source_files_raw': cur.execute("SELECT COUNT(*) FROM source_files_raw").fetchone()[0],
            'files_imported': cur.execute("SELECT COUNT(*) FROM files_imported").fetchone()[0],
            'measurements_curated': cur.execute("SELECT COUNT(*) FROM measurements_curated").fetchone()[0],
            'validation_issues': cur.execute("SELECT COUNT(*) FROM validation_issues").fetchone()[0],
            'latest_day_ref': cur.execute("SELECT MAX(day_ref) FROM measurements_curated").fetchone()[0] or '',
        }
        conn.close()
        return payload
    except Exception as exc:
        return {
            'ok': False,
            'db_path': str(DB_PATH),
            'error': str(exc),
        }


def _remove_path_with_retries(path: Path, *, recursive: bool = False, attempts: int = 5, delay: float = 0.2) -> bool:
    if not path.exists():
        return False
    for attempt in range(attempts):
        try:
            if recursive:
                shutil.rmtree(path)
            else:
                path.unlink()
            return True
        except FileNotFoundError:
            return False
        except PermissionError:
            if attempt == attempts - 1:
                raise
            time.sleep(delay)
    return False


def _truncate_local_database():
    conn = db_conn()
    try:
        cur = conn.cursor()
        cur.execute("PRAGMA foreign_keys=OFF")
        tables = [
            row[0]
            for row in cur.execute(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table' AND name NOT LIKE 'sqlite_%'
                """
            ).fetchall()
        ]
        for table_name in tables:
            cur.execute(f'DELETE FROM "{table_name}"')
        try:
            cur.execute("DELETE FROM sqlite_sequence")
        except sqlite3.OperationalError:
            pass
        try:
            cur.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        except sqlite3.OperationalError:
            pass
        conn.commit()
        return {
            'mode': 'truncate_fallback',
            'tables_cleared': len(tables),
        }
    finally:
        conn.close()


def _reset_local_data(keep_backup_zip: bool = True, hard_restart: bool = False):
    global _monthly_reset_in_progress, _monthly_refresh_generation
    removed = {
        'db_files': 0,
        'state_files': 0,
        'output_excels': 0,
        'backup_zips': 0,
        'upload_items': 0,
    }
    reset_mode = 'safe_restart' if hard_restart else 'truncate_reset'
    with _monthly_refresh_lock:
        _monthly_reset_in_progress = True
        _monthly_refresh_generation += 1
        _monthly_refresh_guard.clear()
        _monthly_refresh_threads.clear()
        fallback = _truncate_local_database()
        removed['db_files'] = 1 if fallback.get('tables_cleared') else 0

    try:
        if hard_restart:
            for path in (Path(f'{DB_PATH}-wal'), Path(f'{DB_PATH}-shm')):
                try:
                    if _remove_path_with_retries(path):
                        removed['db_files'] += 1
                except Exception:
                    pass
        else:
            for suffix in ('-wal', '-shm'):
                try:
                    if _remove_path_with_retries(Path(f'{DB_PATH}{suffix}')):
                        removed['db_files'] += 1
                except Exception:
                    pass

        for path in WORK_DIR.glob('state_*.json'):
            if _remove_path_with_retries(path):
                removed['state_files'] += 1

        if _monthly_refresh_tmp_dir.exists():
            for path in _monthly_refresh_tmp_dir.iterdir():
                if _remove_path_with_retries(path, recursive=path.is_dir()):
                    removed['state_files'] += 1

        for path in OUTPUT_DIR.glob('*.xlsx'):
            if _remove_path_with_retries(path):
                removed['output_excels'] += 1

        if not keep_backup_zip:
            for path in OUTPUT_DIR.glob('MPFM_backup_*.zip'):
                if _remove_path_with_retries(path):
                    removed['backup_zips'] += 1

        if UPLOAD_DIR.exists():
            for path in UPLOAD_DIR.iterdir():
                if _remove_path_with_retries(path, recursive=path.is_dir()):
                    removed['upload_items'] += 1

        WORK_DIR.mkdir(parents=True, exist_ok=True)
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        init_db()

        payload = {
            'ok': True,
            'mode': reset_mode,
            'removed': removed,
            'health': _db_health_snapshot(),
        }
        return payload
    finally:
        with _monthly_refresh_lock:
            _monthly_reset_in_progress = False


def _delete_all_data_for_day(target_date: str):
    target_date = _normalize_admin_date_input(target_date)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(target_date or "").strip()):
        return {"ok": False, "error": "Dia inválido. Use o formato YYYY-MM-DD."}

    month_key = target_date[:7]
    yr = target_date[:4]
    mo = target_date[5:7]
    day_tag = f"{target_date[8:10]}_{target_date[5:7]}"
    workbook_path = OUTPUT_DIR / excel_name(yr, mo)
    deleted_xml_files = []

    conn = db_conn()
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    imported_rows = cur.execute(
        """
        SELECT id, run_id, filename, COALESCE(file_hash,'') AS file_hash
        FROM files_imported
        WHERE content_date=?
        """,
        (target_date,),
    ).fetchall()
    imported_ids = [int(row["id"]) for row in imported_rows]
    run_ids = sorted({int(row["run_id"]) for row in imported_rows if row["run_id"] is not None})
    filenames = sorted({str(row["filename"] or "") for row in imported_rows if str(row["filename"] or "").strip()})

    sep_source_rows = cur.execute(
        """
        SELECT id, source_file, COALESCE(source_hash,'') AS source_hash
        FROM sep_source_files
        WHERE production_date=?
        """,
        (target_date,),
    ).fetchall()
    sep_source_ids = [int(row["id"]) for row in sep_source_rows]
    sep_filenames = sorted({str(row["source_file"] or "") for row in sep_source_rows if str(row["source_file"] or "").strip()})

    source_file_ids = set()
    if filenames and run_ids:
        placeholders_names = ",".join("?" * len(filenames))
        placeholders_runs = ",".join("?" * len(run_ids))
        sql = (
            f"SELECT id FROM source_files_raw WHERE filename IN ({placeholders_names}) "
            f"AND run_id IN ({placeholders_runs})"
        )
        params = filenames + run_ids
        source_file_ids.update(
            int(row["id"])
            for row in cur.execute(sql, params).fetchall()
        )
    source_file_ids.update(
        int(row["id"])
        for row in cur.execute(
            "SELECT id FROM source_files_raw WHERE content_date=?",
            (target_date,),
        ).fetchall()
    )

    xml_rows = cur.execute(
        "SELECT id, file_path FROM xml042_documents WHERE production_day=?",
        (target_date,),
    ).fetchall()
    xml_ids = [int(row["id"]) for row in xml_rows]

    daily_card_rows = cur.execute(
        "SELECT id FROM daily_cards WHERE production_date=?",
        (target_date,),
    ).fetchall()
    daily_card_ids = [int(row["id"]) for row in daily_card_rows]

    deleted = {
        "files_imported": 0,
        "source_files_raw": 0,
        "parsing_events_raw": 0,
        "measurements_raw": 0,
        "measurements_curated": 0,
        "validation_issues": 0,
        "sep_source_files": 0,
        "sep_alignments": 0,
        "daily_cards": 0,
        "daily_card_edits": 0,
        "mpfm_monitoring_daily": 0,
        "recon_runs": 0,
        "tpoc_daily_potential_curated": 0,
        "xml042_documents": 0,
        "processing_runs_pruned": 0,
    }

    try:
        cur.execute("BEGIN")

        deleted["measurements_curated"] = cur.execute(
            "DELETE FROM measurements_curated WHERE day_ref=? OR source_record_id IN (SELECT id FROM sep_source_files WHERE production_date=?)",
            (target_date, target_date),
        ).rowcount or 0
        deleted["validation_issues"] = cur.execute(
            "DELETE FROM validation_issues WHERE day_ref=?",
            (target_date,),
        ).rowcount or 0
        deleted["recon_runs"] = cur.execute(
            "DELETE FROM recon_runs WHERE day_ref=?",
            (target_date,),
        ).rowcount or 0
        deleted["mpfm_monitoring_daily"] = cur.execute(
            "DELETE FROM mpfm_monitoring_daily WHERE production_date=?",
            (target_date,),
        ).rowcount or 0
        deleted["tpoc_daily_potential_curated"] = cur.execute(
            "DELETE FROM tpoc_daily_potential_curated WHERE production_day=?",
            (target_date,),
        ).rowcount or 0

        if daily_card_ids:
            placeholders = ",".join("?" * len(daily_card_ids))
            deleted["daily_card_edits"] = cur.execute(
                f"DELETE FROM daily_card_edits WHERE daily_card_id IN ({placeholders})",
                daily_card_ids,
            ).rowcount or 0
        deleted["daily_cards"] = cur.execute(
            "DELETE FROM daily_cards WHERE production_date=?",
            (target_date,),
        ).rowcount or 0
        deleted["sep_alignments"] = cur.execute(
            "DELETE FROM sep_alignments WHERE production_date=?",
            (target_date,),
        ).rowcount or 0

        if source_file_ids:
            placeholders = ",".join("?" * len(source_file_ids))
            deleted["parsing_events_raw"] = cur.execute(
                f"DELETE FROM parsing_events_raw WHERE source_file_raw_id IN ({placeholders})",
                list(source_file_ids),
            ).rowcount or 0
            if sep_source_ids:
                sep_placeholders = ",".join("?" * len(sep_source_ids))
                deleted["measurements_raw"] = cur.execute(
                    f"DELETE FROM measurements_raw WHERE source_file_raw_id IN ({placeholders}) OR source_record_id IN ({sep_placeholders}) OR content_date=?",
                    list(source_file_ids) + sep_source_ids + [target_date],
                ).rowcount or 0
            else:
                deleted["measurements_raw"] = cur.execute(
                    f"DELETE FROM measurements_raw WHERE source_file_raw_id IN ({placeholders}) OR content_date=?",
                    list(source_file_ids) + [target_date],
                ).rowcount or 0
        elif sep_source_ids:
            sep_placeholders = ",".join("?" * len(sep_source_ids))
            deleted["measurements_raw"] = cur.execute(
                f"DELETE FROM measurements_raw WHERE source_record_id IN ({sep_placeholders}) OR content_date=?",
                sep_source_ids + [target_date],
            ).rowcount or 0

        if imported_ids:
            placeholders = ",".join("?" * len(imported_ids))
            deleted["files_imported"] = cur.execute(
                f"DELETE FROM files_imported WHERE id IN ({placeholders})",
                imported_ids,
            ).rowcount or 0
        else:
            deleted["files_imported"] = cur.execute(
                "DELETE FROM files_imported WHERE content_date=?",
                (target_date,),
            ).rowcount or 0

        deleted["sep_source_files"] = cur.execute(
            "DELETE FROM sep_source_files WHERE production_date=?",
            (target_date,),
        ).rowcount or 0

        if source_file_ids:
            placeholders = ",".join("?" * len(source_file_ids))
            deleted["source_files_raw"] = cur.execute(
                f"DELETE FROM source_files_raw WHERE id IN ({placeholders})",
                list(source_file_ids),
            ).rowcount or 0
        else:
            deleted["source_files_raw"] = cur.execute(
                "DELETE FROM source_files_raw WHERE content_date=?",
                (target_date,),
            ).rowcount or 0

        if xml_ids:
            placeholders = ",".join("?" * len(xml_ids))
            deleted["xml042_documents"] = cur.execute(
                f"DELETE FROM xml042_documents WHERE id IN ({placeholders})",
                xml_ids,
            ).rowcount or 0

        for run_id in run_ids:
            has_remaining = (
                cur.execute("SELECT 1 FROM files_imported WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM source_files_raw WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM measurements_curated WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM measurements_raw WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM validation_issues WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
                or cur.execute("SELECT 1 FROM parsing_events_raw WHERE run_id=? LIMIT 1", (run_id,)).fetchone()
            )
            if has_remaining:
                continue
            deleted["processing_runs_pruned"] += cur.execute(
                "DELETE FROM processing_runs WHERE id=?",
                (run_id,),
            ).rowcount or 0

        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        raise
    finally:
        conn.close()

    for row in xml_rows:
        file_path = str(row["file_path"] or "").strip()
        if not file_path:
            continue
        try:
            xml_path = Path(file_path)
            if xml_path.exists():
                xml_path.unlink()
                deleted_xml_files.append(xml_path.name)
        except Exception:
            continue

    state = load_state(yr, mo)
    state["processed"] = [key for key in (state.get("processed") or []) if not str(key).endswith(f"_{day_tag}")]
    for key in list((state.get("processed_hours_by_key") or {}).keys()):
        if str(key).endswith(f"_{day_tag}"):
            state["processed_hours_by_key"].pop(key, None)
    state.setdefault("processed_hours", {}).pop(day_tag, None)
    state.setdefault("sep_by_day", {}).pop(target_date, None)
    state["sep_days"] = [day for day in (state.get("sep_days") or []) if str(day) != target_date]
    save_state(state)

    validation_snapshot = _rebuild_validation_snapshot_for_month(month_key)
    queued = schedule_monthly_base_unica(workbook_path, yr, mo)

    return {
        "ok": True,
        "target_date": target_date,
        "month": month_key,
        "deleted": deleted,
        "deleted_xml_files": deleted_xml_files,
        "state": {
            "processed_day_tag_removed": day_tag,
            "state_file": f"state_{yr}_{mo}.json",
            "workbook_name": workbook_path.name,
        },
        "validation_snapshot": validation_snapshot.get("recomputed", {}),
        "queued": queued,
        "workbook_exists": workbook_path.exists(),
    }


# ── Helpers ───────────────────────────────────────────────────────────────────

def excel_name(yr, mo):
    return excel_name_service(MONTH_PT, yr, mo)


def _ser(sep):
    def c(v):
        return None if (v is None or (isinstance(v, float) and math.isnan(v))) else v
    return {str(k): {kk: c(vv) for kk, vv in d.items()} for k, d in sep.items()}


def _des(raw):
    if not raw:
        return None
    return {"DAY" if k == "DAY" else int(k): d for k, d in raw.items()}



def build_monthly_base_unica(workbook_path: Path, yr: str, mo: str):
    build_monthly_base_unica_service(
        db_conn,
        workbook_path,
        yr,
        mo,
        write_cards_to_workbook_fn=_write_cards_to_workbook,
        serialize_sep_row_fn=_ser,
        month_pt=MONTH_PT,
        engine=engine,
        load_state_fn=load_state,
    )


_monthly_refresh_guard = set()
_monthly_refresh_lock = threading.Lock()
_monthly_refresh_threads = {}
_monthly_reset_in_progress = False
_monthly_refresh_generation = 0
_monthly_refresh_tmp_dir = WORK_DIR / "_monthly_refresh"


def is_monthly_workbook_rebuilding(workbook_path: Path) -> bool:
    key = str(Path(workbook_path).resolve())
    with _monthly_refresh_lock:
        return key in _monthly_refresh_guard


def schedule_monthly_base_unica(workbook_path: Path, yr: str, mo: str):
    global _monthly_refresh_generation
    key = str(Path(workbook_path).resolve())
    with _monthly_refresh_lock:
        if _monthly_reset_in_progress:
            return False
        if key in _monthly_refresh_guard:
            return False
        _monthly_refresh_guard.add(key)
        generation = _monthly_refresh_generation

    def _runner():
        temp_path = None
        try:
            _monthly_refresh_tmp_dir.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".xlsx",
                dir=str(_monthly_refresh_tmp_dir),
                prefix=f"monthly_{yr}_{mo}_",
            ) as tmp_file:
                temp_path = Path(tmp_file.name)
            build_monthly_base_unica(temp_path, yr, mo)
            with _monthly_refresh_lock:
                can_publish = (
                    generation == _monthly_refresh_generation and
                    not _monthly_reset_in_progress
                )
            if can_publish and temp_path.exists():
                temp_path.replace(workbook_path)
        except Exception as exc:
            print(f"WARN monthly workbook refresh failed for {workbook_path}: {exc}")
        finally:
            if temp_path and temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            with _monthly_refresh_lock:
                _monthly_refresh_guard.discard(key)
                _monthly_refresh_threads.pop(key, None)

    worker = threading.Thread(target=_runner, name=f"monthly-refresh-{yr}-{mo}", daemon=True)
    with _monthly_refresh_lock:
        _monthly_refresh_threads[key] = worker
    worker.start()
    return True


def load_state(yr, mo):
    return load_import_state(WORK_DIR, yr, mo)


def save_state(s):
    save_import_state(WORK_DIR, s)


def dataframe_to_records(df):
    if df is None or df.empty:
        return []
    out = []
    for row in df.to_dict(orient='records'):
        clean = {}
        for k, v in row.items():
            if hasattr(v, 'isoformat'):
                clean[k] = v.isoformat()
            elif isinstance(v, float) and math.isnan(v):
                clean[k] = None
            else:
                clean[k] = v
        out.append(clean)
    return out


# ── Core ──────────────────────────────────────────────────────────────────────

# ── Cadastro-driven validation ───────────────────────────────────────────────

def _build_cadastro_index():
    """
    Builds lookup structures from cadastro.json:
      - expected_tags[bank_code] = set of expected TAG/sistema strings
      - tag_to_instrument[bank_code][tag] = instrument_tag (ex: 18FT0506)
      - bank_loop[bank_code] = loop string
      - sep_banks: compatibility set for separator-linked flows (the separator itself is a
        standalone source and no longer belongs to a fixed MPFM bank)
    """
    cad = _load_cadastro()
    expected_tags   = {}
    tag_to_instrument = {}

    for entry in cad.get('banks_subsea', []) + cad.get('banks_topside', []):
        if not entry.get('ativo', True):
            continue
        bc  = entry.get('bank_code', '').strip()
        tag = (entry.get('sistema') or '').strip()
        ins = (entry.get('tag_associado') or '').strip()
        if not bc:
            continue
        normalized_tag = _normalize_tag_name(tag)
        if not normalized_tag:
            continue
        expected_tags.setdefault(bc, set()).add(normalized_tag)
        tag_to_instrument.setdefault(bc, {})[normalized_tag] = ins

    bank_loop = cad.get('banco_loop', {})
    sep_banks = {SEP_SOURCE_UNIT_CODE}

    return {
        'expected_tags':    expected_tags,
        'tag_to_instrument': tag_to_instrument,
        'bank_loop':         bank_loop,
        'sep_banks':         sep_banks,
    }


def _validate_with_cadastro(run_id: int, excel_file: str, unit_code: str,
                             found_tags: list, day_ref: str):
    """
    Compares tags found in the PDF against the cadastro expectations.
    Issues logged to validation_issues:
      - unknown_tag   : tag in PDF not in cadastro for that bank (warn)
      - missing_tag   : tag expected in cadastro but absent from PDF (warn)
      - no_sep_data   : bank has sep_meter but no TXT data for the day (warn)
    """
    idx = _build_cadastro_index()
    expected = idx['expected_tags'].get(unit_code, set())

    found_set    = {_normalize_tag_name(tag) for tag in found_tags if _normalize_tag_name(tag)}
    unknown_tags = found_set - expected
    missing_tags = expected - found_set

    for tag in sorted(unknown_tags):
        if tag:   # ignore blank tags
            db_add_issue(run_id, excel_file, 'unknown_tag', 'warn',
                         f'{unit_code}/{tag}', day_ref,
                         f'TAG "{tag}" encontrado no PDF mas não está no cadastro do {unit_code}.')

    for tag in sorted(missing_tags):
        if tag:
            db_add_issue(run_id, excel_file, 'missing_tag', 'warn',
                         f'{unit_code}/{tag}', day_ref,
                         f'TAG "{tag}" esperado no cadastro do {unit_code} mas ausente no PDF do dia.')

    return {'unknown': list(unknown_tags), 'missing': list(missing_tags)}


def _rebuild_validation_snapshot_for_month(target_month: str):
    if not re.match(r'^\d{4}-\d{2}$', str(target_month or '')):
        return {'ok': False, 'month': str(target_month or ''), 'recomputed': {}}

    import calendar

    yr, mo = int(target_month[:4]), int(target_month[5:7])
    date_from = f'{target_month}-01'
    date_to = f'{target_month}-{calendar.monthrange(yr, mo)[1]:02d}'
    cadastro = _build_cadastro_index()
    now = datetime.now().isoformat(timespec='seconds')

    conn = db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        DELETE FROM validation_issues
        WHERE issue_type IN ('missing_hours','missing_tag','unknown_tag','sep_missing_files','sep_duplicate_candidate','recon_partial','recon_verify')
          AND (
            substr(day_ref,1,7)=?
            OR substr(created_at,1,7)=?
            OR (
              length(COALESCE(day_ref,''))=5
              AND substr(day_ref,3,1)='_'
              AND substr(day_ref,4,2)=?
              AND substr(created_at,1,4)=?
            )
          )
        """,
        (target_month, target_month, target_month[5:7], target_month[:4]),
    )

    daily_pairs = cur.execute(
        """
        SELECT DISTINCT day_ref, bank
        FROM measurements_curated
        WHERE row_kind='daily' AND day_ref BETWEEN ? AND ?
        ORDER BY day_ref, bank
        """,
        (date_from, date_to),
    ).fetchall()
    daily_tags_rows = cur.execute(
        """
        SELECT DISTINCT day_ref, bank, tag
        FROM measurements_curated
        WHERE row_kind='daily' AND day_ref BETWEEN ? AND ? AND COALESCE(tag,'')<>''
        """,
        (date_from, date_to),
    ).fetchall()
    hourly_rows = cur.execute(
        """
        SELECT DISTINCT day_ref, bank, hour_ref
        FROM measurements_curated
        WHERE row_kind='hourly' AND day_ref BETWEEN ? AND ? AND hour_ref IS NOT NULL
        """,
        (date_from, date_to),
    ).fetchall()
    sep_counts = cur.execute(
        """
        SELECT production_date, COUNT(DISTINCT fluid_kind) AS n_fluids
        FROM sep_source_files
        WHERE is_active=1 AND is_official=1 AND production_date BETWEEN ? AND ?
        GROUP BY production_date
        """,
        (date_from, date_to),
    ).fetchall()
    sep_duplicates = cur.execute(
        """
        SELECT production_date, fluid_kind, meter_id, COUNT(*) AS n_rows
        FROM sep_source_files
        WHERE is_active=1 AND production_date BETWEEN ? AND ?
        GROUP BY production_date, fluid_kind, meter_id
        HAVING COUNT(*) > 1
        """,
        (date_from, date_to),
    ).fetchall()

    found_tags_by_day_bank = defaultdict(set)
    for row in daily_tags_rows:
        normalized = _normalize_tag_name(row['tag'])
        if normalized:
            found_tags_by_day_bank[(row['day_ref'], row['bank'])].add(normalized)

    hours_by_day_bank = defaultdict(set)
    for row in hourly_rows:
        hours_by_day_bank[(row['day_ref'], row['bank'])].add(int(row['hour_ref']))

    counts = {
        'missing_hours': 0,
        'missing_tag': 0,
        'unknown_tag': 0,
        'sep_missing_files': 0,
        'sep_duplicate_candidate': 0,
    }

    for row in daily_pairs:
        day_ref = row['day_ref']
        day_ref_iso = _normalize_day_ref(day_ref)
        bank = row['bank']
        expected = cadastro['expected_tags'].get(bank, set())
        found = found_tags_by_day_bank.get((day_ref, bank), set())

        for tag in sorted(found - expected):
            cur.execute(
                """
                INSERT INTO validation_issues(run_id, excel_file, issue_type, severity, ref_key, day_ref, details, created_at)
                VALUES(NULL,'','unknown_tag','warn',?,?,?,?)
                """,
                (f'{bank}/{tag}', day_ref_iso, f'TAG "{tag}" encontrado nos dados do dia mas não está no cadastro do {bank}.', now),
            )
            counts['unknown_tag'] += 1

        for tag in sorted(expected - found):
            cur.execute(
                """
                INSERT INTO validation_issues(run_id, excel_file, issue_type, severity, ref_key, day_ref, details, created_at)
                VALUES(NULL,'','missing_tag','warn',?,?,?,?)
                """,
                (f'{bank}/{tag}', day_ref_iso, f'TAG "{tag}" esperado no cadastro do {bank} mas ausente nos dados do dia.', now),
            )
            counts['missing_tag'] += 1

        rec_hours = hours_by_day_bank.get((day_ref, bank), set())
        missing_hours = sorted(set(range(24)) - rec_hours)
        if missing_hours:
            details = 'Horas faltando: ' + ', '.join(f'{hour:02d}' for hour in missing_hours)
            cur.execute(
                """
                INSERT INTO validation_issues(run_id, excel_file, issue_type, severity, ref_key, day_ref, details, created_at)
                VALUES(NULL,'','missing_hours','warn',?,?,?,?)
                """,
                (f'{bank}_{day_ref}', day_ref_iso, details, now),
            )
            counts['missing_hours'] += 1

    for row in sep_counts:
        if int(row['n_fluids'] or 0) >= 3:
            continue
        cur.execute(
            """
            INSERT INTO validation_issues(run_id, excel_file, issue_type, severity, ref_key, day_ref, details, created_at)
            VALUES(NULL,'','sep_missing_files','warn',?,?,?,?)
            """,
            ('B10', row['production_date'], f"Separador com apenas {int(row['n_fluids'] or 0)}/3 fluido(s) oficiais no dia.", now),
        )
        counts['sep_missing_files'] += 1

    for row in sep_duplicates:
        cur.execute(
            """
            INSERT INTO validation_issues(run_id, excel_file, issue_type, severity, ref_key, day_ref, details, created_at)
            VALUES(NULL,'','sep_duplicate_candidate','info',?,?,?,?)
            """,
            (
                f"B10/{row['fluid_kind']}",
                row['production_date'],
                f"{int(row['n_rows'] or 0)} arquivos candidatos para {row['fluid_kind']}. Um fica oficial e os demais pendentes para rastreabilidade.",
                now,
            ),
        )
        counts['sep_duplicate_candidate'] += 1

    conn.commit()
    conn.close()
    return {'ok': True, 'month': target_month, 'recomputed': counts}

def _db_store_sep_measurements(run_id: int, excel_file: str, unit: str, sep_data: dict, yr: str, mo: str, actual_day: str = None, source_file: str = "", source_record_id: int = None, is_official: bool = True):
    store_sep_measurements(db_conn, run_id, excel_file, unit, sep_data, yr, mo, actual_day, source_file, source_record_id, is_official)


def _db_store_sep_fluid_detail(run_id: int, excel_file: str, fluid_kind: str, file_path: str, actual_day: str = None, source_record_id: int = None, is_official: bool = True):
    store_sep_fluid_detail(db_conn, inspect_txt_content, run_id, excel_file, fluid_kind, file_path, actual_day, source_record_id, is_official)


def _sep_detail_headers(fluid: str):
    return sep_detail_headers(fluid)


def _sep_detail_kind(fluid: str) -> str:
    return sep_detail_kind(fluid)


def _upsert_sep_detail_row(fluid: str, day_ref: str, hour_ref, tag: str, instrument: str = '', values: dict | None = None, source_file: str = '', source_record_id: int | None = None, is_official: bool = True):
    upsert_sep_detail_row(db_conn, fluid, day_ref, hour_ref, tag, instrument, values, source_file, source_record_id, is_official)


_RECON_DAILY_METRIC_MAP = {
    "MPFM corr Gás (t)": ("mpfm_corr", "gas"),
    "MPFM corr Óleo (t)": ("mpfm_corr", "oil"),
    "MPFM corr HC (t)": ("mpfm_corr", "hc"),
    "MPFM corr Água (t)": ("mpfm_corr", "water"),
    "MPFM corr Total (t)": ("mpfm_corr", "total"),
    "PVT vol Gás (Sm³)": ("pvt_vol", "gas"),
    "PVT vol Óleo (m³)": ("pvt_vol", "oil"),
    "PVT vol Água (m³)": ("pvt_vol", "water"),
}

_RECON_HOURLY_METRIC_MAP = {
    "MPFM corr Gás (t)": ("mpfm_corr", "gas"),
    "MPFM corr Óleo (t)": ("mpfm_corr", "oil"),
    "MPFM corr HC (t)": ("mpfm_corr", "hc"),
    "MPFM corr Água (t)": ("mpfm_corr", "water"),
    "MPFM corr Total (t)": ("mpfm_corr", "total"),
    "PVT vol Gás (Sm³)": ("pvt_vol", "gas"),
    "PVT vol Óleo (m³)": ("pvt_vol", "oil"),
    "PVT vol Água (m³)": ("pvt_vol", "water"),
}


def _clear_mpfm_state_for_daily(state: dict, key: str):
    processed = list(state.get("processed", []))
    state["processed"] = [item for item in processed if item != key]


def _clear_mpfm_state_for_hours(state: dict, key: str, hours_to_remove: list[int]):
    existing = list(state.setdefault("processed_hours_by_key", {}).get(key, []))
    state["processed_hours_by_key"][key] = sorted({int(hour) for hour in existing if int(hour) not in set(hours_to_remove)})
    day_tag = "_".join(key.split("_")[1:])
    day_existing = list(state.setdefault("processed_hours", {}).get(day_tag, []))
    state["processed_hours"][day_tag] = sorted({int(hour) for hour in day_existing if int(hour) not in set(hours_to_remove)})


def _purge_mpfm_rows(bank: str, day_ref: str, row_kind: str, hour_refs: list[int] | None = None):
    conn = db_conn()
    cur = conn.cursor()
    if row_kind == 'hourly' and hour_refs:
        placeholders = ",".join("?" * len(hour_refs))
        cur.execute(
            f"DELETE FROM measurements_curated WHERE bank=? AND day_ref=? AND row_kind='hourly' AND hour_ref IN ({placeholders})",
            [bank, day_ref] + list(hour_refs),
        )
    else:
        cur.execute(
            "DELETE FROM measurements_curated WHERE bank=? AND day_ref=? AND row_kind=?",
            (bank, day_ref, row_kind),
        )
    conn.commit()
    conn.close()


def _rebuild_recon_from_curated(bank: str, production_date: str, run_id: int, excel_file: str):
    conn = db_conn()
    cur = conn.cursor()
    daily_rows = cur.execute(
        """
        SELECT bank, loop, tipo, tag, instrument, metric_name, metric_value
        FROM measurements_curated
        WHERE row_kind='daily' AND bank=? AND day_ref=?
        ORDER BY tag, metric_name
        """,
        (bank, production_date),
    ).fetchall()
    hourly_rows = cur.execute(
        """
        SELECT hour_ref, tag, instrument, metric_name, metric_value
        FROM measurements_curated
        WHERE row_kind='hourly' AND bank=? AND day_ref=?
        ORDER BY hour_ref, tag, metric_name
        """,
        (bank, production_date),
    ).fetchall()
    cur.execute(
        "DELETE FROM measurements_curated WHERE row_kind='recon' AND bank=? AND day_ref=?",
        (bank, production_date),
    )
    conn.commit()
    conn.close()

    if not daily_rows:
        return False

    daily = {
        'date_from': production_date,
        'fpso_side': daily_rows[0]['loop'] or '',
        'unit_type': daily_rows[0]['tipo'] or '',
        'tags': {},
    }
    for row in daily_rows:
        tag = row['tag'] or ''
        entry = daily['tags'].setdefault(
            tag,
            {
                'instrument': row['instrument'] or '',
                'metrics': {
                    'mpfm_corr': {'gas': None, 'oil': None, 'hc': None, 'water': None, 'total': None},
                    'pvt_vol': {'gas': None, 'oil': None, 'water': None},
                },
            },
        )
        metric_map = _RECON_DAILY_METRIC_MAP.get(row['metric_name'] or '')
        if metric_map:
            group_name, metric_name = metric_map
            entry['metrics'][group_name][metric_name] = float(row['metric_value'])

    hourly_records_map = {}
    for row in hourly_rows:
        hour_ref = row['hour_ref']
        if hour_ref is None:
            continue
        record = hourly_records_map.setdefault(
            int(hour_ref),
            {
                'hour': int(hour_ref),
                'date_from': production_date,
                'tags': {},
            },
        )
        tag_entry = record['tags'].setdefault(
            row['tag'] or '',
            {
                'instrument': row['instrument'] or '',
                'metrics': {
                    'mpfm_corr': {'gas': 0.0, 'oil': 0.0, 'hc': 0.0, 'water': 0.0, 'total': 0.0},
                    'pvt_vol': {'gas': 0.0, 'oil': 0.0, 'water': 0.0},
                },
            },
        )
        metric_map = _RECON_HOURLY_METRIC_MAP.get(row['metric_name'] or '')
        if metric_map:
            group_name, metric_name = metric_map
            tag_entry['metrics'][group_name][metric_name] = float(row['metric_value'])

    hourly_records = [hourly_records_map[key] for key in sorted(hourly_records_map)]
    df_recon = engine.build_recon_df(daily, hourly_records, bank)
    rows = dataframe_to_records(df_recon)
    if not rows:
        return False
    db_store_sheet_rows(run_id, excel_file, f"RECON_{bank}_{production_date[8:10]}_{production_date[5:7]}", rows)
    return True


def _apply_mpfm_overwrite_purges(data: dict, state: dict, run_id: int, excel_file: str, logger):
    recon_targets = set()
    for key, daily_item in list(data.get('daily', {}).items()):
        daily_rec, unit_code = daily_item
        if not daily_rec or not daily_rec.get('_overwrite_existing'):
            continue
        production_date = daily_rec.get('date_from') or ''
        if not production_date:
            continue
        _purge_mpfm_rows(unit_code, production_date, 'daily')
        _purge_mpfm_rows(unit_code, production_date, 'recon')
        _clear_mpfm_state_for_daily(state, key)
        recon_targets.add((unit_code, production_date))
        db_add_issue(run_id, excel_file, 'measurement_overwritten', 'info', key, production_date, 'PDF daily mais recente sobrescreveu medicao da mesma janela.')
        logger(f'  ♻️  {unit_code} {production_date}: daily anterior removido para sobrescrita')

    for key, hourly_recs in list(data.get('hourly', {}).items()):
        overwritten_hours = sorted({int(rec.get('hour')) for rec in hourly_recs if rec.get('_overwrite_existing') and rec.get('hour') is not None})
        if not overwritten_hours:
            continue
        unit_code = key.split('_')[0]
        if not hourly_recs:
            continue
        production_date = hourly_recs[0].get('date_from') or ''
        if not production_date:
            continue
        _purge_mpfm_rows(unit_code, production_date, 'hourly', overwritten_hours)
        _purge_mpfm_rows(unit_code, production_date, 'recon')
        _clear_mpfm_state_for_hours(state, key, overwritten_hours)
        recon_targets.add((unit_code, production_date))
        hours_label = ", ".join(f"{hour:02d}" for hour in overwritten_hours)
        db_add_issue(run_id, excel_file, 'measurement_overwritten', 'info', key, production_date, f'PDF hourly mais recente sobrescreveu hora(s): {hours_label}.')
        logger(f'  ♻️  {unit_code} {production_date}: hourly {hours_label} removidas para sobrescrita')

    return recon_targets


def process_file_list(paths_names, density=DEFAULT_DENSITY, source_type='upload', source_ref='', force_overwrite=False):
    """
    FLUXO:
    - PDFs: parseia conteúdo e usa date_from interno para mês/dia.
    - TXTs: classifica por conteúdo (Meter ID / Location / layout), não só por nome.
    - Salva histórico em SQLite.
    """
    import pandas as pd
    import time as _time
    run_id = start_run(source_type, source_ref, density, len(paths_names))
    print(f"[diag] run {run_id}: iniciando parse de {len(paths_names)} arquivo(s)", flush=True)
    _t_parse0 = _time.monotonic()
    import_repo = ImportRepository(db_conn, _file_sha1, _infer_metric_unit)
    with import_repo.batch_writes():
        prepared = prepare_ingestion_batches(
            paths_names,
            run_id,
            source_type,
            engine.parse_pdf,
            _build_cadastro_index,
            import_repo.log_raw_file,
            import_repo.log_file,
            import_repo.find_latest_import_by_identity,
            import_repo.find_latest_import_by_hash,
            import_repo.log_parsing_event,
            import_repo.add_issue,
            force_overwrite=force_overwrite,
        )
    log = list(prepared['log'])
    by_month = prepared['by_month']
    months_found = prepared['months_found']
    print(f"[diag] run {run_id}: parse concluido em {_time.monotonic()-_t_parse0:.2f}s ({len(months_found)} mes(es))", flush=True)

    def _postcheck_run_payload():
        conn = db_conn()
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        problems = []
        checked = 0
        files_ok = 0
        file_errors = [
            dict(row)
            for row in cur.execute(
                """
                SELECT filename, file_type, content_date, report_start, report_end, COALESCE(message,'') AS message
                FROM files_imported
                WHERE run_id=? AND COALESCE(processed_ok,1)=0
                ORDER BY id
                """,
                (run_id,),
            ).fetchall()
        ]

        for record, unit_code, report_type, name, content_date in prepared['parsed_pdfs']:
            checked += 1
            expected_tags = {
                _normalize_tag_name(tag)
                for tag in (record.get('tags') or {}).keys()
                if _normalize_tag_name(tag)
            }
            if report_type == 'daily':
                rows = cur.execute(
                    """
                    SELECT DISTINCT tag
                    FROM measurements_curated
                    WHERE run_id=? AND bank=? AND row_kind='daily' AND day_ref=?
                    """,
                    (run_id, unit_code, content_date),
                ).fetchall()
            else:
                hour_ref = record.get('hour')
                rows = cur.execute(
                    """
                    SELECT DISTINCT tag
                    FROM measurements_curated
                    WHERE run_id=? AND bank=? AND row_kind='hourly' AND day_ref=? AND hour_ref=?
                    """,
                    (run_id, unit_code, content_date, int(hour_ref) if hour_ref is not None else None),
                ).fetchall()
            actual_tags = {_normalize_tag_name(row['tag']) for row in rows if _normalize_tag_name(row['tag'])}
            if not actual_tags:
                problems.append({
                    'filename': name,
                    'file_type': report_type,
                    'content_date': content_date,
                    'detail': 'Arquivo lido, mas não gerou medições na aplicação.',
                })
                continue
            missing_tags = sorted(tag for tag in expected_tags if tag not in actual_tags)
            if missing_tags:
                problems.append({
                    'filename': name,
                    'file_type': report_type,
                    'content_date': content_date,
                    'detail': f'TAGs esperadas ausentes após a carga: {", ".join(missing_tags)}',
                })
                continue
            files_ok += 1

        for _, name, file_type, unit, meter_id, location, content_date, _file_hash, info in prepared['txt_files']:
            checked += 1
            identity_key = info.get('identity_key') or ''
            sep_row = None
            if identity_key:
                sep_row = cur.execute(
                    """
                    SELECT id
                    FROM sep_source_files
                    WHERE identity_key=?
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (identity_key,),
                ).fetchone()
            if not sep_row:
                sep_row = cur.execute(
                    """
                    SELECT id
                    FROM sep_source_files
                    WHERE source_file=? AND production_date=? AND fluid_kind=?
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (name, content_date, file_type),
                ).fetchone()
            if not sep_row:
                problems.append({
                    'filename': name,
                    'file_type': file_type,
                    'content_date': content_date,
                    'detail': 'TXT aceito, mas não gerou registro de origem do separador.',
                })
                continue
            row_count = cur.execute(
                "SELECT COUNT(*) FROM measurements_curated WHERE source_record_id=?",
                (int(sep_row['id']),),
            ).fetchone()[0] or 0
            if row_count <= 0:
                problems.append({
                    'filename': name,
                    'file_type': file_type,
                    'content_date': content_date,
                    'detail': 'TXT registrado, mas não gerou dados do separador na aplicação.',
                })
                continue
            files_ok += 1

        conn.close()
        return {
            'checked_files': checked,
            'validated_files': files_ok,
            'problem_files': problems,
            'failed_files': file_errors,
        }

    if not prepared['parsed_pdfs'] and not prepared['txt_files']:
        import_check = _postcheck_run_payload()
        if import_check['failed_files']:
            log.append('\n⚠️  Nenhum arquivo válido seguiu para a carga.')
            for item in import_check['failed_files']:
                log.append(f"  - {item['filename']}: {item.get('message') or 'Falha ao processar arquivo.'}")
            finish_run(run_id, 'warn', {'log': log, 'import_check': import_check})
            return {'log': log, 'excels': [], 'run_id': run_id, 'import_check': import_check, 'status': 'warn'}
        finish_run(run_id, 'empty', {'log': log, 'import_check': import_check})
        return {'log': log, 'excels': [], 'run_id': run_id, 'import_check': import_check, 'status': 'empty'}
    log.append(f"\n📅 Meses: {', '.join(MONTH_PT.get(mo,mo)+'/'+yr for yr,mo in months_found)}")

    excels = []

    # PASSO 3 — gerar Excel por mês
    for (yr, mo) in months_found:
        data = by_month[(yr, mo)]
        label = f"{MONTH_PT.get(mo, mo)}/{yr}"
        fname = excel_name(yr, mo)
        outxls = OUTPUT_DIR / fname
        state = load_state(yr, mo)

        log.append(f'\n━━━ {label} ━━━')
        _t_month0 = _time.monotonic()
        print(f"[diag] run {run_id}: iniciando mes {label}", flush=True)

        recon_targets = _apply_mpfm_overwrite_purges(data, state, run_id, fname, log.append)
        print(f"[diag] run {run_id}: purge de overwrite concluido em {_time.monotonic()-_t_month0:.2f}s", flush=True)

        _t_sep0 = _time.monotonic()
        process_monthly_sep_inputs(
            data['txts'],
            run_id=run_id,
            excel_file=fname,
            year=yr,
            month=mo,
            density=density,
            state=state,
            engine=engine,
            inspect_txt_content_fn=inspect_txt_content,
            register_sep_source_file_fn=db_register_sep_source_file,
            store_sep_fluid_detail_fn=_db_store_sep_fluid_detail,
            store_sep_measurements_fn=_db_store_sep_measurements,
            add_issue_fn=db_add_issue,
            ser_fn=_ser,
            logger=log.append,
        )
        print(f"[diag] run {run_id}: SEP mensal concluido em {_time.monotonic()-_t_sep0:.2f}s", flush=True)

        new_sheets = {}
        _t_mpfm0 = _time.monotonic()
        month_result = process_monthly_mpfm_inputs(
            data,
            run_id=run_id,
            excel_file=fname,
            year=yr,
            month=mo,
            state=state,
            engine=engine,
            dataframe_to_records_fn=dataframe_to_records,
            day_tag_to_iso_fn=_day_tag_to_iso,
            get_sep_alignment_fn=_get_sep_alignment,
            deserialize_sep_fn=_des,
            load_sep_data_by_day_fn=_load_sep_data_by_day,
            has_sep_alignment_fn=_has_sep_alignment,
            validate_with_cadastro_fn=_validate_with_cadastro,
            add_issue_fn=db_add_issue,
            logger=log.append,
        )
        area_rows = month_result['area_rows']
        sheet_records_for_db = month_result['sheet_records_for_db']
        state = month_result['state']
        print(f"[diag] run {run_id}: MPFM mensal concluido em {_time.monotonic()-_t_mpfm0:.2f}s ({len(sheet_records_for_db)} sheet(s) para gravar)", flush=True)

        # STATUS_MES
        try:
            status_df = engine.build_status_sheet(state)
            if not status_df.empty:
                new_sheets['STATUS_MES'] = status_df
        except Exception as e:
            log.append(f'  ⚠️  Status sheet: {e}')

        if new_sheets:
            try:
                _t_merge0 = _time.monotonic()
                engine._merge_excel(str(outxls), new_sheets, area_rows)
                print(f"[diag] run {run_id}: _merge_excel concluido em {_time.monotonic()-_t_merge0:.2f}s", flush=True)
                log.append(f'  ✅ {fname}  ({len(new_sheets)} abas + BASE_UNICA_MES)')
                excels.append(fname)
                _t_dbwrite0 = _time.monotonic()
                with import_repo.batch_writes():
                    for sheet_name, rows in sheet_records_for_db:
                        import_repo.store_sheet_rows(run_id, fname, sheet_name, rows)
                print(f"[diag] run {run_id}: gravacao em measurements_curated concluida em {_time.monotonic()-_t_dbwrite0:.2f}s", flush=True)
                touched_keys = sorted(set(list(data.get('daily', {}).keys()) + list(data.get('hourly', {}).keys())))
                for key in touched_keys:
                    unit_code = key.split('_')[0]
                    production_date = _day_tag_to_iso(yr, mo, "_".join(key.split("_")[1:]))
                    if production_date:
                        recon_targets.add((unit_code, production_date))
                # Reconstrução de reconciliação e base única são executadas em background
                # para não bloquear a resposta do upload.
                def _background_recon_and_base():
                    for unit_code, production_date in sorted(recon_targets):
                        try:
                            _rebuild_recon_from_curated(unit_code, production_date, run_id, fname)
                        except Exception as exc:
                            try:
                                db_add_issue(run_id, fname, 'recon_rebuild_error', 'warn', f'{unit_code}/{production_date}', production_date, str(exc))
                            except Exception:
                                pass
                    try:
                        schedule_monthly_base_unica(outxls, yr, mo)
                    except Exception as exc:
                        try:
                            db_add_issue(run_id, fname, 'base_unica_schedule_error', 'warn', fname, '', str(exc))
                        except Exception:
                            pass

                threading.Thread(target=_background_recon_and_base, daemon=True).start()
                log.append('  ℹ️  Reconciliação/base única em atualização assíncrona')
            except Exception as e:
                db_add_issue(run_id, fname, 'excel_save_error', 'error', fname, '', str(e))
                log.append(f'  ❌ Erro ao salvar Excel: {e}')
        else:
            if outxls.exists():
                try:
                    status_df = engine.build_status_sheet(state)
                    if not status_df.empty:
                        engine._merge_excel(str(outxls), {'STATUS_MES': status_df}, [])
                except Exception:
                    pass
                _cleanup_workbook(outxls)
                log.append('  ℹ️  Sem abas novas — status atualizado')
                excels.append(fname)

        state['last_run'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_state(state)
        print(f"[diag] run {run_id}: mes {label} finalizado em {_time.monotonic()-_t_month0:.2f}s", flush=True)

    _t_postcheck0 = _time.monotonic()
    import_check = _postcheck_run_payload()
    print(f"[diag] run {run_id}: postcheck concluido em {_time.monotonic()-_t_postcheck0:.2f}s ({import_check['checked_files']} arquivo(s) checados)", flush=True)
    if import_check['problem_files'] or import_check['failed_files']:
        log.append('\n⚠️  Pós-checagem da carga:')
        for item in import_check['problem_files']:
            db_add_issue(run_id, '', 'postcheck_missing_output', 'warn', item['filename'], item['content_date'], item['detail'])
            log.append(f"  - {item['filename']}: {item['detail']}")
        for item in import_check['failed_files']:
            detail = item.get('message') or 'Falha ao processar arquivo.'
            log.append(f"  - {item['filename']}: {detail}")
    else:
        log.append(f"\n✅ Pós-checagem: {import_check['validated_files']} arquivo(s) validados após a carga.")

    run_status = 'warn' if import_check['problem_files'] or import_check['failed_files'] else 'ok'
    finish_run(run_id, run_status, {'excels': excels, 'log_lines': len(log), 'import_check': import_check})
    dimension_conn = db_conn()
    try:
        refresh_measurement_dimensions(dimension_conn, run_id=run_id)
    finally:
        dimension_conn.close()
    # Invalida o cache de meses para que o painel reflita os novos dados imediatamente
    invalidate_months_cache()
    invalidate_cache("mpfm_metadata")
    invalidate_cache("ops_months")
    invalidate_cache("ops_month_summary")
    return {'log': log, 'excels': list(dict.fromkeys(excels)), 'run_id': run_id, 'import_check': import_check, 'status': run_status}


# ── API Routes ────────────────────────────────────────────────────────────────
register_system_routes(
    app,
    {
        'db_conn': db_conn,
        'process_file_list': process_file_list,
        'default_density': DEFAULT_DENSITY,
        'upload_dir': UPLOAD_DIR,
        'output_dir': OUTPUT_DIR,
        'static_dir': STATIC_DIR,
        'work_dir': WORK_DIR,
        'app_title': APP_TITLE,
        'app_version': APP_VERSION,
        'public_base_url': PUBLIC_BASE_URL,
        'db_path': DB_PATH,
        'is_monthly_workbook_rebuilding': is_monthly_workbook_rebuilding,
    },
)

app.include_router(ai_router)
app.include_router(ai_agent_router)

CADASTRO_PATH = WORK_DIR / 'cadastro.json'

# Cache baseado em mtime — só relê o arquivo quando ele for modificado no disco
_cadastro_cache: dict = {"data": None, "mtime": -1.0}

def _load_cadastro():
    src = pathlib.Path(__file__).parent / 'cadastro.json'
    path = CADASTRO_PATH if CADASTRO_PATH.exists() else (src if src.exists() else None)
    if path is None:
        return {}
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return {}
    if _cadastro_cache["data"] is not None and mtime == _cadastro_cache["mtime"]:
        return _cadastro_cache["data"]
    data = json.loads(path.read_text('utf-8'))
    if path is src and not CADASTRO_PATH.exists():
        CADASTRO_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), 'utf-8')
    _cadastro_cache["data"] = data
    _cadastro_cache["mtime"] = mtime
    return data

def _safe_pct(delta, base):
    try:
        if base in (None, 0):
            return None
        return abs(float(delta)) / abs(float(base)) * 100.0
    except Exception:
        return None


def _normalize_tag_name(tag: str) -> str:
    value = str(tag or '').strip()
    if not value:
        return ''
    value = re.sub(r'[-_\s]+', '', value)
    return value.upper()


def _recon_deviation_metrics(day: str):
    mo = day[5:7] if day else ''
    yr = day[:4] if day else ''
    if not mo or not yr:
        return {'max_dev_hc_pct': None, 'max_dev_total_pct': None}
    fpath = OUTPUT_DIR / excel_name(yr, mo)
    if not fpath.exists():
        return {'max_dev_hc_pct': None, 'max_dev_total_pct': None}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fpath, read_only=True, data_only=True)
        max_hc = None
        max_total = None
        for sname in wb.sheetnames:
            if not sname.startswith('RECON_'):
                continue
            ws = wb[sname]
            hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
            need = ['Dia', 'Daily HC (t)', 'Δ HC (t)', 'Daily Água (t)', 'Δ Água (t)']
            if not all(n in hdr for n in need):
                continue
            idx = {n: hdr.index(n) for n in need}
            for row in ws.iter_rows(min_row=2, values_only=True):
                dval = row[idx['Dia']]
                dstr = str(dval.date()) if hasattr(dval, 'date') else str(dval)[:10]
                if dstr != day:
                    continue
                daily_hc = row[idx['Daily HC (t)']] or 0
                delta_hc = row[idx['Δ HC (t)']] or 0
                daily_w = row[idx['Daily Água (t)']] or 0
                delta_w = row[idx['Δ Água (t)']] or 0
                phc = _safe_pct(delta_hc, daily_hc)
                ptot = _safe_pct((delta_hc or 0) + (delta_w or 0), (daily_hc or 0) + (daily_w or 0))
                if phc is not None:
                    max_hc = phc if max_hc is None else max(max_hc, phc)
                if ptot is not None:
                    max_total = ptot if max_total is None else max(max_total, ptot)
        wb.close()
        return {
            'max_dev_hc_pct': round(max_hc, 2) if max_hc is not None else None,
            'max_dev_total_pct': round(max_total, 2) if max_total is not None else None,
        }
    except Exception:
        return {'max_dev_hc_pct': None, 'max_dev_total_pct': None}


def _month_calendar_status(day: str):
    if not day:
        return []
    import calendar
    yr, mo = int(day[:4]), int(day[5:7])
    last_day = calendar.monthrange(yr, mo)[1]
    conn = db_conn(); cur = conn.cursor()
    out = []
    for d in range(1, last_day + 1):
        day_ref = f"{yr:04d}-{mo:02d}-{d:02d}"
        banks = cur.execute("SELECT COUNT(DISTINCT bank) FROM measurements_curated WHERE day_ref=? AND row_kind='daily' AND bank<>'' AND bank<>'SEP'", (day_ref,)).fetchone()[0]
        daily_rows = cur.execute("SELECT COUNT(*) FROM measurements_curated WHERE day_ref=? AND row_kind='daily' AND bank<>''", (day_ref,)).fetchone()[0]
        hours_received = cur.execute("SELECT COUNT(DISTINCT printf('%s-%02d', bank, hour_ref)) FROM measurements_curated WHERE day_ref=? AND row_kind='hourly' AND bank<>'' AND bank<>'SEP' AND hour_ref IS NOT NULL", (day_ref,)).fetchone()[0]
        expected = banks * 24
        alerts = cur.execute("SELECT COUNT(*) FROM validation_issues WHERE day_ref=?", (day_ref,)).fetchone()[0]
        if daily_rows == 0 and hours_received == 0 and alerts == 0:
            status = 'empty'
        elif alerts > 0:
            status = 'attention'
        elif expected > 0 and hours_received >= expected:
            status = 'ok'
        else:
            status = 'partial'
        out.append({'day': day_ref, 'n': d, 'status': status, 'hours': hours_received, 'expected': expected, 'alerts': alerts})
    conn.close()
    return out


def _build_backup_zip():
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    zpath = OUTPUT_DIR / f'MPFM_backup_{ts}.zip'
    with zipfile.ZipFile(zpath, 'w', zipfile.ZIP_DEFLATED) as zf:
        for path in [DB_PATH, CADASTRO_PATH]:
            if path.exists():
                zf.write(path, arcname=path.name)
        for path in WORK_DIR.glob('state_*.json'):
            zf.write(path, arcname=path.name)
        for path in OUTPUT_DIR.glob('*.xlsx'):
            zf.write(path, arcname=f'outputs/{path.name}')
    return zpath


def _clear_local_data(keep_backup_zip: bool = True):
    return _reset_local_data(keep_backup_zip=keep_backup_zip)


def _restart_local_data(keep_backup_zip: bool = True):
    return _reset_local_data(keep_backup_zip=keep_backup_zip, hard_restart=True)


# ── Separator CRUD ──────────────────────────────────────────────────────────────


def _safe_card_sheet_name(base: str, used: set[str]) -> str:
    import re
    s = re.sub(r'[\[\]\*\?\/\:]', '_', base or 'CARD')
    s = s.strip() or 'CARD'
    s = s[:31]
    original = s
    i = 1
    while s in used:
        suffix = f"_{i}"
        s = (original[:31-len(suffix)] + suffix)
        i += 1
    used.add(s)
    return s


def _write_cards_to_workbook(wb, date_from: str, date_to: str, bank: str = '', template_wb=None):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    cards = _daily_metric_groups(date_from, date_to, bank)
    owned_template = template_wb is None and MONTHLY_WORKBOOK_TEMPLATE.exists()
    if owned_template:
        template_wb = load_workbook(MONTHLY_WORKBOOK_TEMPLATE)
    if template_wb and 'CARDS_RESUMO' in template_wb.sheetnames:
        ws, template_ws = reset_sheet_from_template(wb, template_wb, 'CARDS_RESUMO')
        start_col = 3
    else:
        if 'CARDS_RESUMO' in wb.sheetnames:
            del wb['CARDS_RESUMO']
        ws = wb.create_sheet('CARDS_RESUMO')
        template_ws = None
        start_col = 1

    clear_value_region(ws, 4, start_col)
    ws.cell(1, start_col, 'CARDS DIÁRIOS DE MEDIÇÃO')
    ws.cell(2, start_col, f'Período: {date_from} até {date_to}' + (f' · Banco: {bank}' if bank else ' · Todos os bancos'))
    headers = ['Data','Banco','Tipo','Título','Tag','Instrumento','Origem','Oil Sm³','Gas MSm³','Water Sm³','Oil t','Gas t','Water t','HC %','Total %','Observações']
    for i, h in enumerate(headers, start_col):
        ws.cell(4, i, h)
    row = 5
    def num(v):
        try:
            if v in (None,''): return None
            return float(v)
        except Exception:
            return None
    for card in cards:
        if template_ws and row > template_ws.max_row:
            style_row = 5 if (row - 5) % 2 == 0 else min(6, template_ws.max_row)
            seed_row_from_template(ws, template_ws, row, style_row, start_col, start_col + len(headers) - 1)
        vols = card.get('volumes', {}) or {}
        masses = card.get('masses', {}) or {}
        bal = card.get('balance', {}) or {}
        vals = [
            card.get('production_date',''), card.get('bank',''), card.get('card_type',''), card.get('title',''), card.get('tag',''), card.get('instrument',''), card.get('source',''),
            num(vols.get('oil_sm3')), num(vols.get('gas_msm3')), num(vols.get('water_sm3')),
            num(masses.get('oil_t')), num(masses.get('gas_t')), num(masses.get('water_t')),
            (num(bal.get('hc_pct'))/100 if num(bal.get('hc_pct')) is not None else None),
            (num(bal.get('total_pct'))/100 if num(bal.get('total_pct')) is not None else None),
            card.get('observations','') or ''
        ]
        for i, v in enumerate(vals, start_col):
            c = ws.cell(row, i, v)
            if i in (start_col + 13, start_col + 14) and isinstance(v,(int,float)):
                c.number_format = '0.0%'
            elif start_col + 7 <= i <= start_col + 12 and isinstance(v,(int,float)):
                c.number_format = '#,##0.0000'
        row += 1
    if headers:
        last_col = get_column_letter(start_col + len(headers) - 1)
        ws.auto_filter.ref = f"{get_column_letter(start_col)}4:{last_col}{max(row - 1, 4)}"
    if owned_template and template_wb:
        template_wb.close()
    return len(cards)


# ─── User column preferences ──────────────────────────────────────────────────
PREFS_PATH = WORK_DIR / 'user_prefs.json'

DEFAULT_MPFM_COLUMNS = [
    'day_ref','hour_ref','bank','tag',
    'metric_name','metric_value','metric_unit','source_file'
]

ALL_METRIC_NAMES = [
    'MPFM uncorr Gás (t)','MPFM uncorr Óleo (t)','MPFM uncorr HC (t)',
    'MPFM uncorr Água (t)','MPFM uncorr Total (t)',
    'MPFM corr Gás (t)','MPFM corr Óleo (t)','MPFM corr HC (t)',
    'MPFM corr Água (t)','MPFM corr Total (t)',
    'PVT mass Gás (t)','PVT mass Óleo (t)','PVT mass Água (t)',
    'PVT vol Gás (Sm³)','PVT vol Óleo (m³)','PVT vol Água (m³)',
    'PVT @20 mass Gás (t)','PVT @20 mass Óleo (t)','PVT @20 mass Água (t)',
    'PVT @20 vol Gás (Sm³)','PVT @20 vol Óleo (m³)','PVT @20 vol Água (m³)',
    'Pressão (barg)','Temperatura (°C)',
    'Dens. Gás (kg/m³)','Dens. Óleo (kg/m³)','Dens. Água (kg/m³)',
    'Daily Gás (t)','Daily Óleo (t)','Daily HC (t)','Daily Água (t)',
    'Soma h. Gás (t)','Soma h. Óleo (t)','Soma h. HC (t)','Soma h. Água (t)',
    'Δ Gás (t)','Δ Óleo (t)','Δ HC (t)','Δ Água (t)',
    'Horas','Cobertura',
]

def _load_prefs():
    if PREFS_PATH.exists():
        try: return json.loads(PREFS_PATH.read_text('utf-8'))
        except: pass
    return {'selected_metrics': [
        'MPFM corr Óleo (t)','MPFM corr Gás (t)','MPFM corr HC (t)',
        'MPFM corr Total (t)','MPFM corr Água (t)',
        'Pressão (barg)','Temperatura (°C)',
    ]}

register_admin_routes(
    app,
    {
        'load_cadastro': _load_cadastro,
        'cadastro_path': CADASTRO_PATH,
        'output_dir': OUTPUT_DIR,
        'load_prefs': _load_prefs,
        'prefs_path': PREFS_PATH,
        'all_metric_names': ALL_METRIC_NAMES,
        'db_conn': db_conn,
    },
)

register_sgmfm_routes(
    app,
    {
        'repo': SgmfmRepository(db_conn),
        'db_conn': db_conn,
        'load_cadastro': _load_cadastro,
        'build_schema_payload': build_sgmfm_schema_payload,
        'build_record_summary': build_sgmfm_record_summary,
        'build_prefill_payload': build_sgmfm_prefill_payload,
        'render_record_html': render_sgmfm_record_html,
        'generate_record_code': generate_sgmfm_record_code,
        'normalize_tag_name': _normalize_tag_name,
    },
)

def _bbl_from_m3(v):
    try:
        return round(float(v or 0) * BBL_PER_M3, 3)
    except Exception:
        return 0.0


def _boe_from(oil_m3=None, gas_sm3=None):
    try:
        oil_bbl = float(oil_m3 or 0) * BBL_PER_M3
        gas_boe = float(gas_sm3 or 0) / GAS_SM3_PER_BOE
        return round(oil_bbl + gas_boe, 3)
    except Exception:
        return 0.0


def _parse_date(d):
    if isinstance(d, datetime):
        return d
    value = str(d or "").strip()
    if not value:
        raise ValueError("empty date")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(f"time data {value!r} does not match supported formats")


def _add_period(base: str, periodicity: str, periodicity_days: int = 0) -> str:
    from calendar import monthrange
    if not base:
        return ''
    dt = _parse_date(base)
    p = (periodicity or 'custom').lower()
    if p == 'daily':
        nd = dt.replace() + __import__('datetime').timedelta(days=1)
    elif p == 'weekly':
        nd = dt.replace() + __import__('datetime').timedelta(days=7)
    elif p == 'monthly':
        y, m = dt.year, dt.month + 1
        if m > 12: y, m = y + 1, 1
        day = min(dt.day, monthrange(y, m)[1])
        nd = dt.replace(year=y, month=m, day=day)
    elif p == 'quarterly':
        y, m = dt.year, dt.month + 3
        while m > 12: y, m = y + 1, m - 12
        day = min(dt.day, monthrange(y, m)[1])
        nd = dt.replace(year=y, month=m, day=day)
    elif p == 'semiannual':
        y, m = dt.year, dt.month + 6
        while m > 12: y, m = y + 1, m - 12
        day = min(dt.day, monthrange(y, m)[1])
        nd = dt.replace(year=y, month=m, day=day)
    elif p == 'annual':
        y, m = dt.year + 1, dt.month
        day = min(dt.day, monthrange(y, m)[1])
        nd = dt.replace(year=y, month=m, day=day)
    else:
        nd = dt + __import__('datetime').timedelta(days=int(periodicity_days or 0))
    return nd.strftime('%Y-%m-%d')


def _deadline_payload(r):
    def get(key, default=''):
        try:
            return r[key]
        except (KeyError, IndexError):
            return default

    periodicity_map = {
        'custom': 'Definido pelo usuário',
        'daily': 'Diário',
        'weekly': 'Semanal',
        'monthly': 'Mensal',
        'quarterly': 'Trimestral',
        'semiannual': 'Semestral',
        'annual': 'Anual',
    }
    today = datetime.now().strftime('%Y-%m-%d')
    due = get('due_date') or (_add_period(get('start_date'), get('periodicity'), get('periodicity_days')) if get('start_date') else '')
    completion_date = get('completion_date')
    source_status = get('source_status')
    days_remaining = None
    status = 'Sem data'
    if due:
        days_remaining = (_parse_date(due) - _parse_date(today)).days
        status = 'Atrasado' if days_remaining < 0 else 'Hoje' if days_remaining == 0 else 'Em dia'
    is_closed = bool(completion_date) or any(term in str(source_status or '').upper() for term in ('CONCLU', 'ENVIADO', 'EVIDÊNCIA'))
    if is_closed and due:
        if completion_date:
            delta = (_parse_date(completion_date) - _parse_date(due)).days
            status = 'Concluído com atraso' if delta > 0 else 'Concluído'
        else:
            status = 'Concluído'
    return {
        'id': get('id'), 'subject': get('subject'), 'category': get('category'), 'start_date': get('start_date'), 'due_date': due,
        'periodicity': get('periodicity'), 'periodicity_days': get('periodicity_days'), 'notes': get('notes'), 'icon': get('icon') or '⏳',
        'source_ref': get('source_ref'), 'source_file': get('source_file'), 'norm_ref': get('norm_ref'),
        'evidence_required': get('evidence_required'), 'responsible_area': get('responsible_area'),
        'trigger_event': get('trigger_event'), 'risk_level': get('risk_level'), 'recommended_action': get('recommended_action'),
        'completion_date': completion_date, 'source_status': source_status,
        'periodicity_label': periodicity_map.get(get('periodicity'), get('periodicity')),
        'days_remaining': days_remaining, 'status': status, 'is_closed': is_closed
    }


def _cleanup_workbook(workbook_path: Path):
    cleanup_workbook_service(workbook_path)


register_ops_routes(
    app,
    {
        'db_conn': db_conn,
        'recon_deviation_metrics': _recon_deviation_metrics,
        'month_calendar_status': _month_calendar_status,
        'bbl_from_m3': _bbl_from_m3,
        'boe_from': _boe_from,
        'build_backup_zip': _build_backup_zip,
        'build_monthly_base_unica': build_monthly_base_unica,
        'clear_local_data': _clear_local_data,
        'restart_local_data': _restart_local_data,
        'cleanup_workbook': _cleanup_workbook,
        'db_path': DB_PATH,
        'init_db': init_db,
        'recompute_alignment_resolution': _recompute_alignment_resolution,
        'recompute_card_resolution': _recompute_card_resolution,
        'recompute_sep_source_resolution': _recompute_sep_source_resolution,
        'rebuild_sep_summary_for_day': _rebuild_sep_summary_for_day,
        'rebuild_validation_snapshot_for_month': _rebuild_validation_snapshot_for_month,
        'sanitize_files_imported_history': _sanitize_files_imported_history,
        'delete_all_data_for_day': _delete_all_data_for_day,
        'preview_base_unica_import': preview_base_unica_import_service,
        'apply_base_unica_import': apply_base_unica_import_service,
        'schedule_monthly_base_unica': schedule_monthly_base_unica,
        'work_dir': WORK_DIR,
        'output_dir': OUTPUT_DIR,
        'month_pt': MONTH_PT,
        'excel_name': excel_name,
        'load_prefs': _load_prefs,
        'load_cadastro': _load_cadastro,
        'normalize_tag_name': _normalize_tag_name,
        'load_state': load_state,
        'save_state': save_state,
        'load_sep_data_by_day': _load_sep_data_by_day,
        'load_sep_data_by_range': _load_sep_data_by_range,
        'serialize_sep_row': _ser,
    },
)

register_sep_routes(
    app,
    {
        'db_conn': db_conn,
        'db_upsert_sep_alignment': db_upsert_sep_alignment,
        'db_delete_sep_alignment': db_delete_sep_alignment,
        'recompute_alignment_resolution': _recompute_alignment_resolution,
        'recompute_sep_source_resolution': _recompute_sep_source_resolution,
        'output_dir': OUTPUT_DIR,
        'excel_name': excel_name,
        'build_monthly_base_unica': build_monthly_base_unica,
        'schedule_monthly_base_unica': schedule_monthly_base_unica,
        'cleanup_workbook': _cleanup_workbook,
        'sep_detail_headers': _sep_detail_headers,
        'sep_detail_kind': _sep_detail_kind,
        'upsert_sep_detail_row': _upsert_sep_detail_row,
    },
)

register_cards_routes(
    app,
    {
        'upsert_card_override': _upsert_card_override,
        'recompute_card_resolution': _recompute_card_resolution,
        'db_conn': db_conn,
        'deadline_payload': _deadline_payload,
        'sqlite_row_factory': sqlite3.Row,
    },
)


register_recon_routes(
    app,
    {
        'db_conn': db_conn,
        'has_sep_alignment': _has_sep_alignment,
        'get_sep_alignment': _get_sep_alignment,
    },
)

register_methodology_flow_routes(
    app,
    {
        'db_conn': db_conn,
    },
)

register_export_routes(
    app,
    {
        'db_conn': db_conn,
        'daily_metric_groups': _daily_metric_groups,
        'load_prefs': _load_prefs,
        'sep_detail_headers': _sep_detail_headers,
        'sep_detail_kind': _sep_detail_kind,
        'write_cards_to_workbook': _write_cards_to_workbook,
        'is_monthly_workbook_rebuilding': is_monthly_workbook_rebuilding,
    },
)

register_mpfm_adjustment_routes(
    app,
    {
        'db_conn': db_conn,
        'invalidate_cache': invalidate_cache,
    },
)

register_xml042_routes(
    app,
    {
        'db_conn': db_conn,
        'output_dir': OUTPUT_DIR,
        'load_cadastro': _load_cadastro,
        'normalize_tag_name': _normalize_tag_name,
    },
)

register_monthly_reports_routes(
    app,
    {
        'db_conn': db_conn,
        'normalize_tag_name': _normalize_tag_name,
        'month_pt': MONTH_PT,
    },
)

from repositories.xml042 import Xml042Repository

register_painel_operador_routes(
    app,
    {
        'service': PainelOperadorStagingService(Path(__file__).resolve().parent / 'Painel_Operador'),
        'db_conn': db_conn,
        'xml042_repo': Xml042Repository(db_conn, _normalize_tag_name),
    },
)

if AlarmRepository is not None and register_alarme_routes is not None:
    register_alarme_routes(
        app,
        {
            'repo': AlarmRepository(db_conn),
        },
    )
else:
    alarm_bootstrap_error = _ALARM_REPOSITORY_IMPORT_ERROR or _ALARM_ROUTES_IMPORT_ERROR
    print(f"[WARNING] Rotas de alarmes desabilitadas: {alarm_bootstrap_error}")


if __name__ == '__main__':
    print('\n' + '='*52)
    print(f'  {APP_TITLE.upper()} v{APP_VERSION} — Servidor Local')
    print('='*52)
    print(f'  URL:    {PUBLIC_BASE_URL}')
    print(f'  Dados:  {WORK_DIR.resolve()}')
    print(f'  SQLite: {DB_PATH.resolve()}')
    print('='*52 + '\n')
    uvicorn.run('server:app', host=DEFAULT_HOST, port=DEFAULT_PORT, reload=False, use_colors=False)
